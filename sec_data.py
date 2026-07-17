import os
import time
import httpx
import pandas as pd
import numpy as np
import re
import argparse
import warnings
import atexit
import copy
import hashlib
import json
import pickle
import sys
import logging
import shutil
from collections import defaultdict
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor
from functools import wraps, lru_cache
from io import StringIO, BytesIO
import html as html_lib
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from edgar import Company, set_identity

# Precompiled once at import; used by the HTML table-cleaning helpers below to
# strip hidden (display:none) nodes.  Identical to the inline re.compile(...) it
# replaces, but not rebuilt on each call.
_DISPLAY_NONE_RE = re.compile(r'display:\s*none', re.I)

# BUILD: annual-only foreign private issuer path
ANNUAL_ONLY_BUILD_ID = "annual-only-working-v8"
GLOBAL_CALC_PARENT = {}
# Inversion engine state: where each filing ACTUALLY presents each face
# concept (accumulated across filings), and concepts whose static-map
# placement has been overridden by the company's own linkbase.
_FACE_PRESENTED = {}
# Per-concept face-statement presentation positions captured from XBRL
# presentation trees during a live run.  Used by the final row sorter so all
# three core statements can follow the company's own filing order first.
_FACE_PRESENTATION_POS = {}
_RESOLVE_OVERRIDDEN = set()


class _FastTupleRow:
    """Minimal Series-like view over an ``itertuples`` row.

    ``DataFrame.iterrows()`` constructs a pandas Series for every fact, which
    is expensive on large XBRL fact sets.  This adapter preserves the exact
    ``row["column"]`` / ``row.get(...)`` access pattern used by the extractor
    while reading directly from tuple storage.
    """
    __slots__ = ("_values", "_positions")

    def __init__(self, values, positions):
        self._values = values
        self._positions = positions

    def reset(self, values):
        self._values = values
        return self

    def __getitem__(self, key):
        return self._values[self._positions[key]]

    def get(self, key, default=None):
        pos = self._positions.get(key)
        return default if pos is None else self._values[pos]


def _calc_section_of(concept, max_hops=5):
    """Walk the company's calculation linkbase upward to a recognized
    statement root and return the implied category, or None."""
    # The root sets are unions of module-level constants that are never mutated
    # at runtime, so they are built once and reused.  They are only used for
    # membership tests below, so the cached sets are identical in effect to
    # rebuilding them on every call.
    _roots = _calc_section_of._roots
    if _roots is None:
        _IS_ROOTS = (_OPEX_ROLLUP_PARENTS | _REVENUE_ROLLUP_PARENTS |
                     {'NetIncomeLoss', 'OperatingIncomeLoss', 'GrossProfit',
                      'IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest',
                      'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'})
        _CF_ROOTS = _CF_OPERATING_PARENTS | _CF_INVESTING_PARENTS | _CF_FINANCING_PARENTS
        _BS_ROOTS = {'Assets', 'Liabilities', 'LiabilitiesAndStockholdersEquity',
                     'StockholdersEquity',
                     'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest',
                     'AssetsCurrent', 'LiabilitiesCurrent'}
        _roots = _calc_section_of._roots = (_IS_ROOTS, _CF_ROOTS, _BS_ROOTS)
    _IS_ROOTS, _CF_ROOTS, _BS_ROOTS = _roots
    frontier = {concept}
    for _ in range(max_hops):
        nxt = set()
        for c in frontier:
            if c in _CF_ROOTS:
                return '3_Cash_Flow'
            if c in _IS_ROOTS:
                return '1_Income_Statement'
            if c in _BS_ROOTS:
                return '2_Balance_Sheet'
            for p, _w in GLOBAL_CALC_PARENT.get(c, ()):
                nxt.add(p)
        if not nxt:
            return None
        frontier = nxt
    return None


# Cache slot for the immutable statement-root sets, populated on first call.
_calc_section_of._roots = None

# Roll-up roots used to classify lines via the calculation linkbase.
_OPEX_ROLLUP_PARENTS = {
    'OperatingExpenses', 'CostsAndExpenses', 'OperatingCostsAndExpenses',
    'BenefitsLossesAndExpenses', 'NoninterestExpense',
}
_REVENUE_ROLLUP_PARENTS = {
    'Revenues', 'RevenueFromContractWithCustomerExcludingAssessedTax',
    'RevenueFromContractWithCustomerIncludingAssessedTax',
    'RevenuesNetOfInterestExpense', 'NoninterestIncome',
    'InterestAndDividendIncomeOperating',
}
_CF_OPERATING_PARENTS = {
    'NetCashProvidedByUsedInOperatingActivities',
    'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations',
}
_CF_INVESTING_PARENTS = {
    'NetCashProvidedByUsedInInvestingActivities',
    'NetCashProvidedByUsedInInvestingActivitiesContinuingOperations',
}
_CF_FINANCING_PARENTS = {
    'NetCashProvidedByUsedInFinancingActivities',
    'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations',
}
_CF_NONOPERATING_PARENTS = {
    'NetCashProvidedByUsedInInvestingActivities',
    'NetCashProvidedByUsedInInvestingActivitiesContinuingOperations',
    'NetCashProvidedByUsedInFinancingActivities',
    'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations',
}


_CF_BRIDGE_SPEC = {}
_IBM_STYLE_STATE = {'active': False}  # interest folded into the operating-expense block (no separate operating-income line)
_BRIDGE_USED_LABELS = set()


def _spec_from_contribs(contribs, getrow):
    """Derive per-label weights from realized contributions. Returns
    (spec, complete): complete is False when a materially nonzero
    contribution has no recoverable row weight, in which case the
    residual for that section cannot be faithfully recomputed later."""
    spec, complete = {}, True
    for lbl, c in contribs.items():
        r = pd.to_numeric(getrow(lbl), errors='coerce').fillna(0)
        m = (r != 0)
        if m.any():
            ratio = (c[m] / r[m]).median()
            if pd.notna(ratio) and ratio != 0:
                spec[lbl] = float(ratio)
                continue
        if c.abs().max() > 1e6:
            complete = False
    return spec, complete


def _recompute_cf_residuals(df):
    """Bridge residuals are computed inside calculate_kpis, but later
    repair passes (annual scope fixes, accounting engine, industry KPIs)
    can update component rows afterwards -- leaving stale plugs (GOOGL's
    Wiz acquisition value arrived after the bridge ran, leaving a +$31.6B
    ghost in 'Other Investing Adjustments (Net)'). Recompute the three CF
    residual rows from FINAL displayed values using the recorded spec."""
    if not _CF_BRIDGE_SPEC or df is None or df.empty:
        return df

    _num_row_cache = {}

    def _num_row(idx):
        cached = _num_row_cache.get(idx)
        if cached is not None:
            return cached
        s = pd.to_numeric(df.loc[idx], errors='coerce')
        _num_row_cache[idx] = s
        return s

    changed = []
    for section, total_lbl, res_lbl in (
            ('op', 'Operating Cash Flow', 'Other Operating Adjustments (Net)'),
            ('inv', 'Investing Cash Flow', 'Other Investing Adjustments (Net)'),
            ('fin', 'Financing Cash Flow', 'Other Financing Adjustments (Net)')):
        entry = _CF_BRIDGE_SPEC.get(section)
        if not entry or not entry.get('complete'):
            continue
        spec = entry.get('spec') or {}
        tidx, ridx = ('3_Cash_Flow', total_lbl), ('3_Cash_Flow', res_lbl)
        if not spec or tidx not in df.index or ridx not in df.index:
            continue
        total = _num_row(tidx)
        _missing = [lbl for lbl in spec if ('3_Cash_Flow', lbl) not in df.index]
        if _missing:
            print(f"  [Bridge] NOTE: '{res_lbl}' NOT recomputed -- bridge component "
                  f"row(s) vanished after KPIs: {', '.join(_missing[:6])}"
                  + (" ..." if len(_missing) > 6 else "")
                  + " (keeping KPI-time residual; investigate what dropped them)")
            continue
        ssum = pd.Series(0.0, index=df.columns)
        for lbl, w in spec.items():
            li = ('3_Cash_Flow', lbl)
            if li in df.index:
                ssum = ssum + _num_row(li).fillna(0) * w
        res = (total.fillna(0) - ssum).where(total.notna())
        old = _num_row(ridx)
        if ((old.fillna(0) - res.fillna(0)).abs() > 1e6).any():
            changed.append(res_lbl)
        df.loc[ridx, :] = res.values
        _num_row_cache[ridx] = pd.to_numeric(df.loc[ridx], errors='coerce')
    if changed:
        print(f"  [Bridge] Recomputed {', '.join(changed)} after repair passes "
              f"(stale residuals refreshed).")
    return df


_CF_SUPPLEMENTAL_MARKERS = ('paid', 'obtained in exchange', 'incurred',
                            'period increase', 'noncash', 'non-cash')


def _classify_cf_label_fallback(label, section):
    """
    Semantic fallback when the calculation linkbase has no arc for a learned
    CF concept (common for custom tags). Returns a weight or None. Cash-paid
    and noncash supplemental disclosures are never bridge components.
    """
    ll = str(label).lower()
    if any(m in ll for m in _CF_SUPPLEMENTAL_MARKERS) and 'repaid' not in ll:
        return None
    if section == 'fin':
        if any(k in ll for k in ('repayment', 'repurchase', 'buyback',
                                 'tax receivable agreement', 'tax withholding',
                                 'redemption', 'debt restructuring cost',
                                 'dividends')):
            return -1.0
        if any(k in ll for k in ('proceeds from issuance', 'treasury stock',
                                 'proceeds from stock', 'borrowings under',
                                 'collateral held under securities lending',
                                 'excess tax benefit')):
            return 1.0
    elif section == 'inv':
        is_invest = any(k in ll for k in ('investment', 'securities', 'acquisition',
                                          'business combination', 'intangible',
                                          'credit card receivable', 'loans receivable'))
        if is_invest and any(k in ll for k in ('purchase', 'payments to acquire',
                                               'net of cash acquired', 'loans originated')):
            return -1.0
        if is_invest and any(k in ll for k in ('sale', 'maturit', 'distribution',
                                               'proceeds', 'redemption')):
            return 1.0
        if 'capital expenditure' in ll:
            return -1.0
    elif section == 'op':
        if any(k in ll for k in ('collateral', 'securities lending', 'restricted cash')):
            return None
        if ('increase decrease in' in ll or ll.startswith('change in')
                or 'changes in' in ll or ll.startswith('increase (decrease)')
                or ll.startswith('decrease (increase)')):
            # liabilities first: 'securities loaned' is a liability even
            # though 'securities' alone reads as an asset
            if any(k in ll for k in ('payable', 'loaned', 'borrowed', 'accrued',
                                     'deposits received', 'owed to')):
                return 1.0
            is_asset = any(k in ll for k in ('receivable', 'asset', 'inventor',
                                             'prepaid', 'deposit', 'securities',
                                             'financial instruments', 'segregated',
                                             'owned', 'restricted'))
            return -1.0 if is_asset else 1.0
        if any(k in ll for k in ('amortization', 'depreciation', 'impairment',
                                 'provision', 'write-off', 'write off')):
            return 1.0
    return None


def _bridge_sign_self_check(total_n, sum_s, contribs, section_name):
    """
    A component summed with the wrong polarity leaves a residual of exactly
    -2x its value in every period. Test flipping each component's sign; adopt
    a flip only when it collapses the median residual by >60%. Catches both
    bad fallback heuristics and miswired calculation weights, self-healing
    per company instead of relying on label keywords.
    """
    mask = total_n.notna()
    if not mask.any() or not contribs:
        return sum_s
    flips = 0
    while flips < 3:
        resid = total_n.fillna(0) - sum_s
        best_lbl, best_score = None, 0.0
        for lbl, contrib in contribs.items():
            # quarters where this component is material
            material = mask & (contrib.abs() > np.maximum(0.05 * total_n.abs().fillna(0), 1e7))
            n_mat = int(material.sum())
            if n_mat < 3:
                continue
            # wrong-polarity signature: residual == -2x the contribution
            # wherever the component is material. Magnitude-weighted so the
            # quarters that matter dominate the verdict and baseline noise
            # in small quarters cannot veto an obvious correction.
            hit = (resid[material] + 2 * contrib[material]).abs() < 0.25 * (2 * contrib[material].abs())
            w_mat = float(contrib[material].abs().sum())
            w_hit = float(contrib[material].abs()[hit].sum())
            if w_mat > 0 and (w_hit / w_mat) >= 0.8:
                if w_mat > best_score:
                    best_lbl, best_score = lbl, w_mat
        if best_lbl is None:
            break
        sum_s = sum_s - 2 * contribs[best_lbl]
        contribs[best_lbl] = -contribs[best_lbl]
        print(f"  [Bridge] Sign-corrected '{best_lbl}' in the {section_name} bridge "
              f"(residual matched -2x its value in material quarters).")
        flips += 1

    # Stage 2 -- hill climb for multiple simultaneous polarity errors
    # (broker reconciliations): when several components are wrong at once,
    # no single flip matches the -2x signature. Greedily adopt the flip
    # that most reduces total |residual|, and revert everything unless the
    # final residual is at least 60% smaller than where stage 2 started.
    resid = total_n.fillna(0) - sum_s
    start_abs = float(resid[mask].abs().sum())
    tot_abs = float(total_n[mask].abs().sum())
    if start_abs > 0.10 * max(tot_abs, 1e9):
        cur_sum = sum_s.copy()
        cur_abs = start_abs
        flipped = []
        for _ in range(5):
            best_lbl2, best_abs2 = None, cur_abs
            cur_resid = total_n.fillna(0) - cur_sum
            for lbl, contrib in contribs.items():
                test = float((cur_resid + 2 * contrib)[mask].abs().sum())
                if test < 0.80 * cur_abs and test < best_abs2:
                    best_lbl2, best_abs2 = lbl, test
            if best_lbl2 is None:
                break
            cur_sum = cur_sum - 2 * contribs[best_lbl2]
            contribs[best_lbl2] = -contribs[best_lbl2]
            flipped.append(best_lbl2)
            cur_abs = best_abs2
        if flipped and cur_abs < 0.40 * start_abs:
            for lbl in flipped:
                print(f"  [Bridge] Sign-corrected '{lbl}' in the {section_name} bridge "
                      f"(multi-component polarity solve).")
            return cur_sum
        # revert contribs mutations from the failed climb
        for lbl in flipped:
            contribs[lbl] = -contribs[lbl]
    return sum_s


def _classify_calc_lineage(concepts, counted_concepts, targets, blockers, max_hops=4):
    """
    Walk the SEC calculation linkbase upward from any of `concepts`,
    carrying the product of arc weights.  Returns:
      ('target', weight)  -- first mapped ancestor is in `targets`
      ('counted', w) / ('blocked', w) -- hit an already-counted concept or a
                                          blocker root first (do not add)
      (None, 0.0)         -- no classification possible
    Used to wire dynamically learned face lines into the bridge math so
    'Other Adjustments' plugs stay true residuals.
    """
    frontier = {(c, 1.0) for c in (concepts or []) if c}
    seen = set()
    for _ in range(max_hops):
        nxt = set()
        for c, w in frontier:
            if c in seen:
                continue
            seen.add(c)
            for p, pw in GLOBAL_CALC_PARENT.get(c, ()):
                try:
                    eff = w * (float(pw) if pw is not None else 1.0)
                except (TypeError, ValueError):
                    eff = w
                if p in targets:
                    return ('target', eff)
                if p in blockers:
                    return ('blocked', eff)
                if p in counted_concepts:
                    return ('counted', eff)
                nxt.add((p, eff))
        frontier = nxt
        if not frontier:
            break
    return (None, 0.0)

import threading
import concurrent.futures
from collections import deque
import itertools


class PipelineProgress:
    """One end-to-end progress bar for the complete extraction pipeline.

    The old implementation measured only completed SEC filing futures, so it
    could report 100% while statement assembly, reconciliation, KPI creation,
    cleanup, and output writing were still running.  This tracker reserves
    progress ranges for those downstream stages and can also be updated from
    the isolated foreign-filer pipeline.
    """

    def __init__(self, enabled=True, description="SEC extraction"):
        self.enabled = bool(enabled)
        self.current = 0.0
        self._bar = None
        self._last_stage = ""
        self._ticker, initial_stage = self._split_description(description)
        self._stage_plain = self._stage_text(initial_stage)
        self._last_desc = ""
        self._stream = sys.stdout
        self._color = _stream_supports_color(self._stream)
        self._spinner_frames = self._spinner_frames_for_stream(self._stream)
        self._spinner_index = 0
        self._outline_width = _terminal_ui_width(self._stream)
        self._box_chars = _box_chars_for_stream(self._stream)
        self._attached_outline = False
        self._bottom_bar = None
        self._tqdm_cls = None
        self._result_footer = None
        self._stats_done = 0
        self._stats_total = 0
        self._stats_latest = self._stage_plain
        self._last_stats_text = ""
        self._warn_count = 0
        self._retry_count = 0
        self._fail_count = 0
        self._last_warn_text = ""
        self._spinner_stop = threading.Event()
        self._display_lock = threading.RLock()
        self._spinner_thread = None
        self._pulse_stop = threading.Event()
        self._pulse_thread = None
        if not self.enabled:
            return
        try:
            from tqdm.auto import tqdm
            self._tqdm_cls = tqdm
            left_border = (
                _ansi(self._box_chars["v"], "border")
                if self._color else self._box_chars["v"]
            )
            right_border = left_border
            bar_kwargs = dict(
                total=100.0,
                desc=self._decorated_stage(),
                file=self._stream,
                ascii=self._bar_charset_for_stream(self._stream),
                ncols=self._outline_width,
                dynamic_ncols=False,
                position=0,
                mininterval=0.5 if _PROGRESS_LIGHT_ENABLED else 0.15,
                smoothing=0.05,
                bar_format=(
                    f"{left_border} "
                    "{desc} {bar} {percentage:3.0f}% | {elapsed}"
                    f" {right_border}"
                ),
            )
            if self._color:
                bar_kwargs["colour"] = "white"
            try:
                self._bar = tqdm(**bar_kwargs)
            except TypeError:
                bar_kwargs.pop("colour", None)
                self._bar = tqdm(**bar_kwargs)
            self._last_desc = self._decorated_stage()
            if not _PROGRESS_LIGHT_ENABLED:
                self._spinner_thread = threading.Thread(
                    target=self._spin,
                    name="PipelineProgressSpinner",
                    daemon=True,
                )
                self._spinner_thread.start()
        except Exception:
            # Progress is a UI aid only; extraction must still work when tqdm
            # is unavailable or the output stream cannot render a dynamic bar.
            self.enabled = False
            self._bar = None

    @staticmethod
    def _spinner_frames_for_stream(stream):
        frames = (
            "\u280b", "\u2819", "\u2839", "\u2838", "\u283c",
            "\u2834", "\u2826", "\u2827", "\u2807", "\u280f",
        )
        encoding = getattr(stream, "encoding", None) or "utf-8"
        try:
            "".join(frames).encode(encoding)
            return frames
        except Exception:
            return ("|", "/", "-", "\\")

    @staticmethod
    def _bar_charset_for_stream(stream):
        charset = "\u00b7\u2501"
        encoding = getattr(stream, "encoding", None) or "utf-8"
        try:
            charset.encode(encoding)
            return charset
        except Exception:
            return True

    @staticmethod
    def _split_description(description):
        text = re.sub(r"\s+", " ", str(description or "")).strip()
        if ":" in text:
            prefix, rest = text.split(":", 1)
            prefix = prefix.strip()
            if prefix and len(prefix) <= 12:
                return prefix.upper(), rest.strip() or "starting"
        return "", text or "starting"

    @staticmethod
    def _stage_text(stage):
        text = re.sub(r"\s+", " ", str(stage or "")).strip()
        return text[:39] + "..." if len(text) > 42 else text

    @staticmethod
    def _phase_text(value):
        try:
            pct = float(value)
        except Exception:
            pct = 0.0
        if pct < 8.0:
            return "Phase 1/4 Setup"
        if pct < 70.0:
            return "Phase 2/4 SEC retrieval"
        if pct < 98.0:
            return "Phase 3/4 Build/reconcile"
        return "Phase 4/4 Write"

    def _decorated_stage(self, value=None):
        prefix = f"{self._ticker} | " if self._ticker else ""
        phase = self._phase_text(self.current if value is None else value)
        spinner = self._spinner_frames[self._spinner_index]
        body = f"{prefix}{phase} | {self._stage_plain}"
        text = f"{spinner} {body}"
        if len(text) > 60:
            body = body[:max(0, 56)] + "..."
        if not self._color:
            return f"{spinner} {body}"
        return f"{_ansi(spinner, 'orange')}{_ansi(' ' + body, 'bright')}"

    def attach_outline(self):
        """Attach the live progress row to the run-card outline."""
        with self._display_lock:
            self._attached_outline = self._bar is not None
            self._last_warn_text = self.warn_text()
            self._last_stats_text = self.stats_text()
            if (
                not self._attached_outline
                or self._bottom_bar is not None
                or self._tqdm_cls is None
            ):
                return
            bottom = _format_box_bottom(self._stream, self._outline_width)
            self._bottom_bar = self._tqdm_cls(
                total=1,
                desc=bottom,
                file=self._stream,
                ncols=self._outline_width,
                dynamic_ncols=False,
                position=1,
                leave=True,
                bar_format="{desc}",
            )
            self._bar.refresh()

    def set_result_footer(self, text):
        self._result_footer = str(text or "").strip() or None

    def sync_terminal_width(self):
        with self._display_lock:
            if self._attached_outline:
                return self._outline_width
            self._outline_width = _terminal_ui_width(self._stream)
            if self._bar is not None:
                self._bar.ncols = self._outline_width
            return self._outline_width

    def set_stats(self, done=None, total=None, latest=None):
        if done is not None:
            try:
                self._stats_done = max(0, int(done))
            except Exception:
                pass
        if total is not None:
            try:
                self._stats_total = max(0, int(total))
            except Exception:
                pass
        if latest:
            self._stats_latest = _clean_log_text(latest)
        return self.stats_text()

    def note_log(self, *args, **kwargs):
        sep = kwargs.pop("sep", " ")
        kwargs.pop("end", "\n")
        kwargs.pop("file", None)
        kwargs.pop("flush", False)
        text = sep.join(str(a) for a in args)
        text = _clean_log_text(text)
        if not text:
            return
        with self._display_lock:
            self._stats_latest = text
            self._update_warn_counts_from_log(text)
            self._rewrite_warn_row_if_changed()
            self._rewrite_stats_row_if_changed()

    def warn_text(self):
        return _format_warn_text(
            self._warn_count,
            self._retry_count,
            self._fail_count,
        )

    def stats_text(self):
        return _format_stats_text(
            self._stats_done,
            self._stats_total,
            self._stats_latest,
        )

    def _spin(self):
        while not self._spinner_stop.wait(0.12):
            with self._display_lock:
                if self._bar is None:
                    return
                self._spinner_index = (
                    self._spinner_index + 1
                ) % len(self._spinner_frames)
                desc = self._decorated_stage()
                self._bar.set_description_str(desc, refresh=True)
                self._last_desc = desc

    def start_pulse(self, end, stage=None, expected_seconds=45.0):
        """Smoothly advance inside a long synchronous stage.

        Heavy pandas/accounting passes do not report internal progress.  This
        keeps the UI honest enough for humans: it moves quickly at first, then
        slows and leaves a small reserve until the stage explicitly completes.
        """
        if self._bar is None:
            return
        self.stop_pulse()
        self.set(self.current, stage)
        try:
            end = max(self.current, min(100.0, float(end)))
            expected_seconds = max(1.0, float(expected_seconds))
        except Exception:
            return
        if end <= self.current:
            return
        self._pulse_stop = threading.Event()

        def _pulse():
            start = self.current
            span = end - start
            started = time.monotonic()
            while not self._pulse_stop.wait(0.45):
                elapsed = time.monotonic() - started
                # Asymptotic curve: never consumes the final 4% of this stage.
                frac = min(0.96, elapsed / (expected_seconds + elapsed))
                target = start + span * frac
                if target > self.current:
                    self.set(target)

        self._pulse_thread = threading.Thread(
            target=_pulse,
            name="PipelineProgressPulse",
            daemon=True,
        )
        self._pulse_thread.start()

    def stop_pulse(self):
        pulse = self._pulse_thread
        if pulse is None:
            return
        self._pulse_stop.set()
        if pulse is not threading.current_thread():
            pulse.join(timeout=0.8)
        self._pulse_thread = None

    def set(self, value, stage=None):
        """Advance monotonically to an absolute percentage."""
        if self._bar is None:
            return
        with self._display_lock:
            target = max(self.current, min(100.0, float(value)))
            if stage:
                stage_text = self._stage_text(stage)
                if stage_text != self._last_stage:
                    self._stage_plain = stage_text
                    self._last_stage = stage_text
                self._update_stats_from_stage(stage_text)
            desc = self._decorated_stage(target)
            if desc != self._last_desc:
                self._bar.set_description_str(desc, refresh=False)
                self._last_desc = desc
            delta = target - self.current
            if delta > 0:
                self._bar.update(delta)
                self.current = target
            else:
                self._bar.refresh()
            self._rewrite_stats_row_if_changed()

    def write(self, *args, **kwargs):
        text = kwargs.pop("sep", " ").join(str(a) for a in args)
        end = kwargs.pop("end", "\n")
        file = kwargs.pop("file", None)
        flush = kwargs.pop("flush", False)
        if kwargs:
            text = f"{text} {' '.join(f'{k}={v}' for k, v in kwargs.items())}"
        with self._display_lock:
            if self._bar is not None:
                self._bar.write(text, file=(file or self._stream), end=end)
            else:
                print(text, end=end, file=file, flush=flush)

    def finish(self, stage="Complete", footer=None):
        self.stop_pulse()
        self._spinner_stop.set()
        footer = footer or self._result_footer
        with self._display_lock:
            if self._bar is not None:
                self.set(100.0, stage)
                if footer and self._rewrite_status_row(footer):
                    footer = None
                self._bar.close()
                self._bar = None
            self._write_outline_bottom(footer)

    def close(self):
        """Close without falsely forcing the bar to 100% after an error."""
        self.stop_pulse()
        self._spinner_stop.set()
        with self._display_lock:
            if self._bar is not None:
                self._bar.close()
                self._bar = None
            self._write_outline_bottom()

    def _update_stats_from_stage(self, stage_text):
        text = str(stage_text or "").strip()
        match = re.search(r"(\d+)\s*/\s*(\d+)", text)
        if match:
            self._stats_done = int(match.group(1))
            self._stats_total = int(match.group(2))
        else:
            queued = re.search(r"Queued\s+(\d+)\s+SEC\s+filings", text, re.I)
            if queued:
                self._stats_done = 0
                self._stats_total = int(queued.group(1))

    def _update_warn_counts_from_log(self, text):
        lower = str(text or "").lower()
        if not lower:
            return
        if "retry" in lower or "rate-limited" in lower or "throttled" in lower:
            self._retry_count += 1
        if (
            "warning" in lower
            or "[warn]" in lower
            or " warn]" in lower
            or "throttled" in lower
            or "rate-limited" in lower
        ):
            self._warn_count += 1
        if (
            " failed" in lower
            or "[error]" in lower
            or " error]" in lower
            or "permanently failed" in lower
            or "could not recover" in lower
        ):
            self._fail_count += 1

    def _rewrite_card_row(self, label, value, rows_above_progress):
        if not self._attached_outline or not _stream_supports_cursor_control(self._stream):
            return False
        line = _format_card_row(
            label,
            value,
            stream=self._stream,
            width=self._outline_width,
        )
        self._stream.write(f"\r\033[{rows_above_progress}A{line}\033[{rows_above_progress}B\r")
        self._stream.flush()
        if self._bar is not None:
            self._bar.refresh()
        if self._bottom_bar is not None:
            self._bottom_bar.refresh()
        return True

    def _rewrite_status_row(self, footer):
        # Status is above Warn, Stats, and the progress separator.
        return self._rewrite_card_row(
            "Status",
            _saved_footer_text(footer, self._stream),
            4,
        )

    def _rewrite_warn_row_if_changed(self):
        text = self.warn_text()
        if text == self._last_warn_text:
            return
        if self._rewrite_card_row("Warn", text, 3):
            self._last_warn_text = text

    def _rewrite_stats_row_if_changed(self):
        text = self.stats_text()
        if text == self._last_stats_text:
            return
        if self._rewrite_card_row("Stats", text, 2):
            self._last_stats_text = text

    def _write_outline_bottom(self, footer=None):
        if not self._attached_outline:
            return
        footer = str(footer or "").strip()
        if self._bottom_bar is not None:
            if footer:
                self._bottom_bar.set_description_str(
                    _format_card_row("Status",
                                     _saved_footer_text(footer, self._stream),
                                     stream=self._stream,
                                     width=self._outline_width),
                    refresh=True,
                )
            else:
                self._bottom_bar.refresh()
            self._bottom_bar.close()
            self._bottom_bar = None
            if footer:
                self._stream.write(_format_box_bottom(self._stream, self._outline_width) + "\n")
                self._stream.flush()
            self._attached_outline = False
            return
        # tqdm can leave the cursor on the rendered progress row in some
        # terminals. Start a fresh line so the closing border is never eaten.
        if footer:
            self._stream.write(
                "\n"
                + _format_card_row("Status",
                                   _saved_footer_text(footer, self._stream),
                                   stream=self._stream,
                                   width=self._outline_width)
                + "\n"
                + _format_box_bottom(self._stream, self._outline_width)
                + "\n"
            )
        else:
            self._stream.write("\n" + _format_box_bottom(self._stream, self._outline_width) + "\n")
        self._stream.flush()
        self._attached_outline = False


def _stream_supports_text(stream, text):
    encoding = getattr(stream, "encoding", None) or "utf-8"
    try:
        str(text).encode(encoding)
        return True
    except Exception:
        return False


def _terminal_ui_width(stream=None, fallback=100):
    stream = stream or sys.stdout
    try:
        is_tty = stream.isatty()
    except Exception:
        is_tty = False
    if not is_tty:
        return int(fallback)
    try:
        columns = shutil.get_terminal_size((int(fallback), 24)).columns
    except Exception:
        columns = int(fallback)
    # Leave a tiny guard band so Windows terminals do not wrap the right border
    # when the cursor lands on the final column.
    target = max(48, int(columns) - 2)
    return target


_ANSI_STYLES = {
    "reset": "0",
    "dim": "2",
    "border": "90",
    "bright": "97",
    "title": "1;97",
    "orange": "38;5;215",
    "blue": "38;5;75",
    "cyan": "38;5;80",
    "purple": "38;5;141",
    "green": "38;5;114",
    "yellow": "38;5;221",
    "red": "38;5;203",
}


def _stream_supports_color(stream):
    if os.environ.get("NO_COLOR"):
        return False
    if os.environ.get("CLICOLOR_FORCE") not in (None, "", "0"):
        return True
    try:
        if not stream.isatty():
            return False
    except Exception:
        return False
    term = os.environ.get("TERM", "")
    if term.lower() == "dumb":
        return False
    if os.name != "nt":
        return bool(term)
    return bool(
        os.environ.get("WT_SESSION")
        or os.environ.get("ANSICON")
        or os.environ.get("ConEmuANSI", "").upper() == "ON"
        or os.environ.get("TERM_PROGRAM")
        or "xterm" in term.lower()
    )


def _stream_supports_cursor_control(stream):
    if os.environ.get("NO_COLOR"):
        return False
    try:
        if not stream.isatty():
            return False
    except Exception:
        return False
    term = os.environ.get("TERM", "")
    if term.lower() == "dumb":
        return False
    if os.name != "nt":
        return bool(term)
    return bool(
        os.environ.get("WT_SESSION")
        or os.environ.get("ANSICON")
        or os.environ.get("ConEmuANSI", "").upper() == "ON"
        or os.environ.get("TERM_PROGRAM")
        or "xterm" in term.lower()
    )


def _ansi(text, style, stream=None):
    if stream is not None and not _stream_supports_color(stream):
        return str(text)
    code = _ANSI_STYLES.get(style)
    if not code:
        return str(text)
    return f"\033[{code}m{text}\033[0m"


def _maybe_ansi(text, style, stream):
    return _ansi(text, style) if _stream_supports_color(stream) else str(text)


def _box_chars_for_stream(stream):
    if _stream_supports_text(stream, "\u256d\u2500\u256e\u2502\u251c\u2524\u2570\u256f"):
        return {
            "tl": "\u256d", "tr": "\u256e",
            "bl": "\u2570", "br": "\u256f",
            "h": "\u2500", "v": "\u2502",
            "lt": "\u251c", "rt": "\u2524",
        }
    return {
        "tl": "+", "tr": "+",
        "bl": "+", "br": "+",
        "h": "-", "v": "|",
        "lt": "+", "rt": "+",
    }


def _format_box_bottom(stream=None, width=100):
    stream = stream or sys.stdout
    chars = _box_chars_for_stream(stream or sys.stdout)
    width = max(48, int(width))
    line = chars["bl"] + chars["h"] * (width - 2) + chars["br"]
    return _maybe_ansi(line, "border", stream)


def _saved_footer_text(text, stream=None):
    text = str(text or "").strip()
    if not text:
        return ""
    if text.lower().startswith(("saved ", "saved to ", "no data")):
        return text
    return "Saved to " + text


def _format_box_row(text, stream=None, width=100, value_style=None):
    stream = stream or sys.stdout
    chars = _box_chars_for_stream(stream)
    width = max(48, int(width))
    inner = width - 4
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(text) > inner:
        text = text[:max(0, inner - 3)] + "..."
    padding = " " * max(0, inner - len(text))
    if not _stream_supports_color(stream):
        return f"{chars['v']} {text}{padding} {chars['v']}"
    value = _ansi(text, value_style) if value_style else text
    return (
        _ansi(chars["v"], "border")
        + " "
        + value
        + padding
        + " "
        + _ansi(chars["v"], "border")
    )


def _format_card_row(label, value, route=None, stream=None, width=100):
    stream = stream or sys.stdout
    chars = _box_chars_for_stream(stream)
    width = max(48, int(width))
    inner = width - 4
    label = str(label)
    value = str(value)
    content = f"{label:<7} {value}"
    if len(content) > inner:
        value = value[:max(0, inner - 11)] + "..."
        content = f"{label:<7} {value}"
    padding = " " * max(0, inner - len(content))
    if not _stream_supports_color(stream):
        return f"{chars['v']} {content:<{inner}} {chars['v']}"
    return (
        _ansi(chars["v"], "border")
        + " "
        + _ansi(f"{label:<7}", "dim")
        + " "
        + _color_card_value(label, value, route, stream)
        + padding
        + " "
        + _ansi(chars["v"], "border")
    )


def _format_stats_text(done, total, latest):
    try:
        done = max(0, int(done))
    except Exception:
        done = 0
    try:
        total = max(0, int(total))
    except Exception:
        total = 0
    latest = re.sub(r"\s+", " ", str(latest or "Starting")).strip()
    if total > 0:
        return f"{done}/{total} Filings | {latest}"
    return f"0/0 Filings | {latest}"


def _format_warn_text(warnings, retries, fails):
    try:
        warnings = max(0, int(warnings))
    except Exception:
        warnings = 0
    try:
        retries = max(0, int(retries))
    except Exception:
        retries = 0
    try:
        fails = max(0, int(fails))
    except Exception:
        fails = 0
    return f"{warnings} warnings | {retries} retries | {fails} fails"


def _format_cik_display(cik):
    text = re.sub(r"\s+", "", str(cik or ""))
    return text.zfill(10) if text.isdigit() else (text or "Unknown")


def _clean_log_text(text):
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    return text


def _format_progress_separator(stream=None, width=100):
    stream = stream or sys.stdout
    chars = _box_chars_for_stream(stream)
    width = max(48, int(width))
    title = " Progress "
    fill = chars["h"] * max(0, width - len(title) - 3)
    if not _stream_supports_color(stream):
        return chars["lt"] + chars["h"] + title + fill + chars["rt"]
    return (
        _ansi(chars["lt"] + chars["h"], "border")
        + _ansi(title, "title")
        + _ansi(fill + chars["rt"], "border")
    )


def _route_display_name(route):
    return {
        "US_NATIVE": "Native 10-K / 10-Q",
        "US_NATIVE_ANNUAL": "Native 10-K annual only",
        "FOREIGN_20F": "Foreign 20-F annual only",
        "FOREIGN_40F": "Foreign 40-F annual only",
        "UNSUPPORTED": "Unsupported filing route",
    }.get(str(route or "").strip(), str(route or "Unknown"))


def _mode_display_name(route=None, use_arelle=None, save_xlsx=False, mode=None):
    if mode:
        return str(mode)
    output = "XLSX" if save_xlsx else "CSV"
    route_key = str(route or "").strip()
    if route_key in ("FOREIGN_20F", "FOREIGN_40F"):
        if use_arelle is None:
            return f"annual | foreign FY | {output}"
        arelle_state = "on" if use_arelle else "off"
        return f"annual | foreign XBRL + Arelle {arelle_state} | {output}"
    period_mode = "annual" if route_key == "US_NATIVE_ANNUAL" else "quarterly"
    if use_arelle is None:
        return f"{period_mode} | {output}"
    arelle_state = "on" if use_arelle else "off"
    return f"{period_mode} | Arelle {arelle_state} | {output}"


def _cache_display_name():
    local_on = os.environ.get("SEC_LOCAL_STORAGE", "").strip().lower() in (
        "1", "true", "yes", "on"
    )
    parts = [f"edgar local {'on' if local_on else 'off'}"]
    native_cache = globals().get("_NATIVE_EXTRACTION_CACHE_ENABLED", None)
    if native_cache is not None:
        parts.append(f"native extract {'on' if native_cache else 'off'}")
    final_cache = globals().get("_FINAL_PIVOT_CACHE_ENABLED", None)
    if final_cache is not None:
        parts.append(f"final pivot {'on' if final_cache else 'off'}")
    html_cache = globals().get("_PERSISTENT_HTML_CACHE_ENABLED", None)
    if html_cache is not None:
        parts.append(f"HTML parse {'on' if html_cache else 'off'}")
    fx_cache = globals().get("_FX_PERSISTENT_CACHE_ENABLED", None)
    if fx_cache is not None:
        parts.append(f"20-F {'on' if fx_cache else 'off'}")
    return "; ".join(parts)


def _route_style(route):
    route = str(route or "").strip()
    if route == "US_NATIVE":
        return "blue"
    if route == "US_NATIVE_ANNUAL":
        return "green"
    if route == "FOREIGN_20F":
        return "cyan"
    if route == "FOREIGN_40F":
        return "purple"
    return "yellow"


def _color_card_value(label, value, route, stream):
    text = str(value)
    if not _stream_supports_color(stream):
        return text
    if label == "Ticker":
        return _ansi(text, "title")
    if label == "Route":
        return _ansi(text, _route_style(route))
    if label == "Mode":
        return _ansi(text, "bright")
    if label == "Cache":
        style = "green" if " off" not in text.lower() else "yellow"
        return _ansi(text, style)
    if label == "Workers":
        return _ansi(text, "bright")
    if label == "Status":
        lower = text.lower()
        if "saved" in lower:
            return _ansi(text, "green")
        if "no data" in lower:
            return _ansi(text, "yellow")
        return _ansi(text, "orange")
    if label == "Warn":
        lower = text.lower()
        fail_match = re.search(r"(\d+)\s+fails?", lower)
        warn_match = re.search(r"(\d+)\s+warnings?", lower)
        retry_match = re.search(r"(\d+)\s+retries?", lower)
        fails = int(fail_match.group(1)) if fail_match else 0
        warnings = int(warn_match.group(1)) if warn_match else 0
        retries = int(retry_match.group(1)) if retry_match else 0
        if fails:
            return _ansi(text, "red")
        if warnings or retries:
            return _ansi(text, "yellow")
        return _ansi(text, "dim")
    if label == "Stats":
        return _ansi(text, "bright")
    return text


def _format_run_card(ticker, limit, route, workers, cache, stream=None, width=100,
                     attach_progress=False, status="Running", warn=None, stats=None,
                     company_name=None, cik=None, mode=None):
    stream = stream or sys.stdout
    width = max(48, int(width))
    rows = [
        ("Ticker", str(ticker).upper()),
        ("Company", str(company_name or "Unknown")),
        ("CIK", _format_cik_display(cik)),
        ("Limit", (f"{limit} annual filings" if str(route or "").strip() in ("FOREIGN_20F", "FOREIGN_40F", "US_NATIVE_ANNUAL") else f"{limit} filings")),
        ("Route", _route_display_name(route)),
        ("Mode", _mode_display_name(route=route, mode=mode)),
        ("Workers", str(workers)),
        ("Cache", str(cache)),
        ("Status", str(status)),
        ("Warn", str(warn or _format_warn_text(0, 0, 0))),
        ("Stats", str(stats or _format_stats_text(0, 0, "Starting"))),
    ]

    chars = _box_chars_for_stream(stream)
    title = " SEC Financials "
    top_fill = chars["h"] * max(0, width - len(title) - 3)
    if _stream_supports_color(stream):
        top = (
            _ansi(chars["tl"] + chars["h"], "border")
            + _ansi(title, "title")
            + _ansi(top_fill + chars["tr"], "border")
        )
    else:
        top = chars["tl"] + chars["h"] + title + top_fill + chars["tr"]
    bottom = _format_box_bottom(stream, width)
    left, right = chars["v"], chars["v"]

    inner = width - 4
    lines = [top]
    for label, value in rows:
        lines.append(_format_card_row(label, value, route, stream, width))
    lines.append(_format_progress_separator(stream, width) if attach_progress else bottom)
    return "\n".join(lines)


class _QuietConsoleStream:
    """Sink for third-party console writes while the progress bar owns output."""

    def __init__(self, real_stream):
        self._real_stream = real_stream
        self.encoding = getattr(real_stream, "encoding", "utf-8")
        self.errors = getattr(real_stream, "errors", "replace")

    def write(self, data):
        return len(data or "")

    def writelines(self, lines):
        return None

    def flush(self):
        return None

    def isatty(self):
        return False


class _ExternalConsoleSilencer:
    """Temporarily suppress library stdout/stderr/log warnings in progress mode.

    Some edgartools parsing paths write directly to stdout/stderr or logging,
    bypassing our normal print suppression and tearing through tqdm's one-line
    progress bar. This guard is intentionally UI-only: it does not catch
    exceptions or alter return values from the wrapped SEC calls.
    """

    _lock = threading.RLock()
    _depth = 0
    _saved_stdout = None
    _saved_stderr = None
    _saved_logging_disable = None

    def __init__(self, enabled=True):
        self.enabled = bool(enabled)

    def __enter__(self):
        if not self.enabled:
            return self
        with self._lock:
            if self.__class__._depth == 0:
                self.__class__._saved_stdout = sys.stdout
                self.__class__._saved_stderr = sys.stderr
                self.__class__._saved_logging_disable = logging.root.manager.disable
                sys.stdout = _QuietConsoleStream(sys.stdout)
                sys.stderr = _QuietConsoleStream(sys.stderr)
                logging.disable(logging.WARNING)
            self.__class__._depth += 1
        return self

    def __exit__(self, exc_type, exc, tb):
        if not self.enabled:
            return False
        with self._lock:
            self.__class__._depth = max(0, self.__class__._depth - 1)
            if self.__class__._depth == 0:
                if self.__class__._saved_stdout is not None:
                    sys.stdout = self.__class__._saved_stdout
                if self.__class__._saved_stderr is not None:
                    sys.stderr = self.__class__._saved_stderr
                if self.__class__._saved_logging_disable is not None:
                    logging.disable(self.__class__._saved_logging_disable)
                self.__class__._saved_stdout = None
                self.__class__._saved_stderr = None
                self.__class__._saved_logging_disable = None
        return False


class SECRateLimiter:
    """A thread-safe sliding-window rate limiter.

    Uses a monotonic clock and never sleeps while holding the mutex. That
    preserves the exact request ceiling while allowing other worker threads to
    make progress instead of queueing behind a sleeping thread.
    """
    def __init__(self, max_calls: int, period: float = 1.0):
        self.max_calls = max_calls
        self.period = period
        self.timestamps = deque()
        self.lock = threading.Lock()

    def wait(self):
        while True:
            with self.lock:
                now = time.monotonic()
                while self.timestamps and now - self.timestamps[0] >= self.period:
                    self.timestamps.popleft()

                if len(self.timestamps) < self.max_calls:
                    self.timestamps.append(now)
                    return

                sleep_time = max(0.0, self.period - (now - self.timestamps[0]))

            # Sleep outside the lock. Every waking thread re-checks the
            # window, so the maximum request rate remains unchanged.
            if sleep_time > 0:
                time.sleep(sleep_time)

sec_limiter = SECRateLimiter(max_calls=9, period=1.0)


# Reuse TCP/TLS connections for direct HTTP calls outside edgartools.
_SHARED_HTTP_CLIENT = httpx.Client(follow_redirects=True)
atexit.register(_SHARED_HTTP_CLIENT.close)

# Suppress warnings from pandas read_html
warnings.filterwarnings('ignore', category=FutureWarning)


# Optional lightweight progress mode for profiling/repeated development runs.
# Disabled by default so the normal console UI remains unchanged.
_PROGRESS_LIGHT_ENABLED = os.environ.get("SEC_PROGRESS_LIGHT", "").strip().lower() in (
    "1", "true", "yes", "on"
)

# ---------------------------------------------------------------------------
# Optional profiling instrumentation (disabled by default)
# ---------------------------------------------------------------------------
_PROFILE_ENABLED = os.environ.get("SEC_PROFILE", "").strip().lower() in (
    "1", "true", "yes", "on"
)
_PROFILE_TIMINGS: dict[str, float] = {}
_PROFILE_COUNTS: dict[str, int] = {}
_PROFILE_COUNTERS: dict[str, int] = {}

_DEBUG_OUTPUT_ENABLED = os.environ.get("SEC_DEBUG", "").strip().lower() in (
    "1", "true", "yes", "on"
)


def _debug_print(*args, **kwargs):
    """Debug-only console output. Keeps normal CSV/XLSX outputs unchanged."""
    if _DEBUG_OUTPUT_ENABLED:
        print(*args, **kwargs)


def _profile_count(name: str, amount: int = 1):
    """Increment a profile-only counter.  No-op unless SEC_PROFILE=1."""
    if not _PROFILE_ENABLED:
        return
    try:
        amount = int(amount)
    except Exception:
        amount = 1
    _PROFILE_COUNTERS[name] = _PROFILE_COUNTERS.get(name, 0) + amount


class _ProfileTimer:
    """Tiny no-op-by-default timer for high-level runtime profiling."""

    def __init__(self, name: str):
        self.name = name
        self.start = None

    def __enter__(self):
        if _PROFILE_ENABLED:
            self.start = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        if _PROFILE_ENABLED and self.start is not None:
            elapsed = time.perf_counter() - self.start
            _PROFILE_TIMINGS[self.name] = _PROFILE_TIMINGS.get(self.name, 0.0) + elapsed
            _PROFILE_COUNTS[self.name] = _PROFILE_COUNTS.get(self.name, 0) + 1
        return False




def _profile_call(name, func, *args, **kwargs):
    """Profile a single call when SEC_PROFILE is enabled; otherwise direct."""
    with _ProfileTimer(name):
        return func(*args, **kwargs)

def _print_profile_report():
    if not _PROFILE_ENABLED:
        return
    if _PROFILE_TIMINGS:
        print("\n[Profile] Stage timing summary")
        for name, seconds in sorted(_PROFILE_TIMINGS.items(), key=lambda kv: kv[1], reverse=True):
            count = _PROFILE_COUNTS.get(name, 0)
            avg = seconds / count if count else seconds
            print(f"  {name:<45} {seconds:8.2f}s  calls={count:<5} avg={avg:.4f}s")
    if _PROFILE_COUNTERS:
        print("\n[Profile] Counter summary")
        for name, value in sorted(_PROFILE_COUNTERS.items(), key=lambda kv: kv[0]):
            print(f"  {name:<45} {value}")


atexit.register(_print_profile_report)

# ---------------------------------------------------------------------------
# First-run SEC identity setup
# ---------------------------------------------------------------------------
# The SEC requires automated clients to identify themselves with a real name
# and contact email.  Keep that contact identity local instead of hardcoding a
# maintainer's personal details into an open-source script.
_SEC_IDENTITY_INITIALIZED = False
_SEC_IDENTITY_VALUE = None
_SEC_IDENTITY_FILE_NAME = "sec_identity.json"
_SEC_IDENTITY_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _sec_identity_cache_path():
    """Return the identity file inside the script's shared cache directory."""
    cache_root = os.path.abspath(os.environ.get("SEC_CACHE_DIR", ".cache"))
    return os.path.join(cache_root, _SEC_IDENTITY_FILE_NAME)


def _normalize_sec_identity(name, email):
    """Validate and format the SEC contact identity used by edgartools."""
    name = re.sub(r"\s+", " ", str(name or "")).strip()
    email = re.sub(r"\s+", "", str(email or "")).strip()
    if len(name) < 2:
        raise ValueError("Enter your real name or organization name.")
    if not _SEC_IDENTITY_EMAIL_RE.match(email):
        raise ValueError("Enter a valid contact email address.")
    return name, email, f"{name} ({email})"


def _load_cached_sec_identity():
    path = _sec_identity_cache_path()
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if not isinstance(payload, dict):
            return None
        name, email, identity = _normalize_sec_identity(
            payload.get("name"), payload.get("email")
        )
        return name, email, identity
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return None


def _save_cached_sec_identity(name, email):
    """Atomically save the non-secret SEC contact identity under .cache/."""
    path = _sec_identity_cache_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = f"{path}.{os.getpid()}.tmp"
    payload = {
        "version": 1,
        "name": name,
        "email": email,
    }
    try:
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2)
            fh.write("\n")
        try:
            os.chmod(tmp_path, 0o600)
        except OSError:
            pass
        os.replace(tmp_path, path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
    return path


def _prompt_for_sec_identity():
    if not sys.stdin or not sys.stdin.isatty():
        raise RuntimeError(
            "No SEC identity is configured and this session is non-interactive. "
            "Set SEC_IDENTITY='Your Name (you@example.com)' or run the script "
            "once in an interactive terminal."
        )

    print("\n" + "=" * 64)
    print("SEC identity setup")
    print("=" * 64)
    print("The SEC requires automated downloads to include a real contact")
    print("name and email. This is not an account or password.")
    print("It will be saved locally and requested only once.\n")

    while True:
        try:
            name = input("Name or organization: ").strip()
            email = input("Contact email: ").strip()
            name, email, identity = _normalize_sec_identity(name, email)
        except (EOFError, KeyboardInterrupt):
            raise RuntimeError("SEC identity setup was cancelled.") from None
        except ValueError as exc:
            print(f"Invalid identity: {exc}\n")
            continue

        path = _save_cached_sec_identity(name, email)
        print(f"\nSEC identity saved to {path}\n")
        return name, email, identity


def _initialize_sec_identity():
    """Configure edgartools from env, cache, or a one-time terminal prompt."""
    global _SEC_IDENTITY_INITIALIZED, _SEC_IDENTITY_VALUE
    if _SEC_IDENTITY_INITIALIZED:
        return

    env_identity = os.environ.get("SEC_IDENTITY", "").strip()
    if env_identity:
        # Keep compatibility with edgartools' accepted free-form identity
        # string for CI, containers, and other non-interactive environments.
        set_identity(env_identity)
        _SEC_IDENTITY_VALUE = env_identity
        _SEC_IDENTITY_INITIALIZED = True
        return

    cached = _load_cached_sec_identity()
    if cached is None:
        cached = _prompt_for_sec_identity()

    _name, _email, identity = cached
    set_identity(identity)
    _SEC_IDENTITY_VALUE = identity
    _SEC_IDENTITY_INITIALIZED = True


def _sec_user_agent():
    """Return the configured identity for direct SEC HTTP requests."""
    _initialize_sec_identity()
    return _SEC_IDENTITY_VALUE


def _reset_cached_sec_identity():
    """Delete the saved identity so the next run displays setup again."""
    global _SEC_IDENTITY_INITIALIZED, _SEC_IDENTITY_VALUE
    path = _sec_identity_cache_path()
    try:
        os.remove(path)
        print(f"Removed saved SEC identity: {path}")
    except FileNotFoundError:
        print(f"No saved SEC identity found at: {path}")
    _SEC_IDENTITY_INITIALIZED = False
    _SEC_IDENTITY_VALUE = None


# Optional persistent on-disk cache of downloaded filings. SEC filings are
# immutable, so enabling this makes repeat runs substantially faster without
# changing parsed filing bytes. It remains opt-in to preserve the original
# live-network/index behavior exactly. Enable with SEC_LOCAL_STORAGE=1.
if os.environ.get("SEC_LOCAL_STORAGE", "").strip().lower() in ("1", "true", "yes", "on"):
    try:
        from edgar import use_local_storage
        use_local_storage()
        print("[Cache] edgartools local storage ON -- re-runs read cached filings from disk.")
    except Exception as _cache_err:
        print(f"[Cache] Local storage unavailable ({type(_cache_err).__name__}); using network.")

# ---------------------------------------------------------------------------
# Network Retry Logic for SEC EDGAR
# ---------------------------------------------------------------------------
def retry_sec_request(retries=3, delay=5):
    """
    Decorator to enforce global rate limits and retry on network failures.

    Catches the full httpx timeout family (ReadTimeout, ConnectTimeout,
    WriteTimeout, PoolTimeout) via the base TimeoutException, plus
    low-level connection/OS errors.  Uses exponential backoff so a
    struggling SEC server gets progressively more breathing room.
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(1, retries + 1):
                sec_limiter.wait()
                try:
                    return func(*args, **kwargs)
                except httpx.TimeoutException as e:
                    # Covers ReadTimeout, ConnectTimeout, WriteTimeout, PoolTimeout
                    wait = delay * attempt   # exponential-ish backoff: 5s, 10s, 15s...
                    print(f"  [Timeout] {func.__name__} timed out (attempt {attempt}/{retries}). "
                          f"Retrying in {wait}s... ({type(e).__name__})")
                    if attempt == retries:
                        raise
                    time.sleep(wait)
                except httpx.HTTPStatusError as e:
                    if e.response.status_code == 429:
                        wait = delay * (2 ** attempt)   # aggressive backoff on 429
                        print(f"  [429 Throttled] {func.__name__} rate-limited. "
                              f"Backing off {wait}s (attempt {attempt}/{retries})...")
                        time.sleep(wait)
                        if attempt == retries:
                            raise
                    else:
                        raise
                except (ConnectionError, OSError) as e:
                    wait = delay * attempt
                    print(f"  [Network Error] {func.__name__}: {type(e).__name__}. "
                          f"Retry {attempt}/{retries} in {wait}s...")
                    if attempt == retries:
                        raise
                    time.sleep(wait)
        return wrapper
    return decorator

# Successful SEC objects/documents are immutable within one extraction run.
# Cache only successful fetches; exceptions still follow the original retry path.
_FETCH_CACHE_LOCK = threading.RLock()
_COMPANY_OBJECT_CACHE = {}
_FILING_HTML_CACHE = {}
_FILINGS_CACHE = {}
_FILING_XBRL_CACHE = {}
_FILING_LOCAL_METADATA_CACHE = {}
_FILING_PERIOD_REPORT_CACHE = {}
_FILING_PERIOD_ESTIMATE_CACHE = {}
_FILING_ACCESSION_KEY_CACHE = {}
_FILING_URL_CACHE = {}
_TEXTBLOCK_HTML_TABLES_CACHE = {}


def _company_cache_key(company):
    try:
        cik = getattr(company, "cik", None)
        if cik is not None:
            return ("cik", str(cik))
    except Exception:
        pass
    return ("object", id(company))


def _filing_cache_key(filing):
    # Once an immutable accession is discovered for a Filing object, reuse it.
    # Do not cache the object-id fallback: a lazy object might expose accession
    # later, and the old behavior would then switch to the accession key.
    obj_id = id(filing)
    cached = _FILING_ACCESSION_KEY_CACHE.get(obj_id)
    if cached is not None:
        _profile_count("filing_cache_key_accession_hits")
        return cached
    for attr in ("accession_no", "accession_number"):
        try:
            value = getattr(filing, attr, None)
            if value:
                key = ("accession", str(value))
                _FILING_ACCESSION_KEY_CACHE[obj_id] = key
                _profile_count("filing_cache_key_accession_misses")
                return key
        except Exception:
            pass
    _profile_count("filing_cache_key_object_fallbacks")
    return ("object", obj_id)



def _get_filing_local_metadata(filing):
    """Return stable, locally available filing metadata cached per accession.

    This helper deliberately reads only the same attributes the old code
    already touched in normal/error paths.  It never requests period_of_report,
    because that property can trigger an SGML header fetch in edgartools.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_LOCAL_METADATA_CACHE.get(key)
    if cached is not None:
        _profile_count("filing_local_metadata_cache_hits")
        return cached

    _profile_count("filing_local_metadata_cache_misses")
    meta = {}
    for attr in ("form", "filing_date", "accession_no", "accession_number"):
        try:
            meta[attr] = getattr(filing, attr, None)
        except Exception:
            meta[attr] = None
    with _FETCH_CACHE_LOCK:
        return _FILING_LOCAL_METADATA_CACHE.setdefault(key, meta)


def _filing_metadata_value(filing, attr, default=None):
    meta = _get_filing_local_metadata(filing)
    value = meta.get(attr)
    return default if value is None else value


def _get_filing_url_cached(filing):
    """Return a filing URL using the same lazy sources as the old local helper.

    Cache only successful, non-empty values.  A transient failure or missing URL
    is not cached, so later calls retain the original chance to recover.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_URL_CACHE.get(key)
    if cached is not None:
        _profile_count("filing_url_cache_hits")
        return cached

    _profile_count("filing_url_cache_misses")
    try:
        url = (filing.document.url if hasattr(filing, 'document')
               else getattr(filing, 'url', None))
    except Exception:
        try:
            url = getattr(filing, 'url', None)
        except Exception:
            url = None
    if url:
        with _FETCH_CACHE_LOCK:
            return _FILING_URL_CACHE.setdefault(key, url)
    return url

def _normalize_form_key(form):
    if isinstance(form, (list, tuple, set, frozenset)):
        return tuple(form)
    return str(form)


@retry_sec_request(retries=3, delay=5)
def _fetch_company_uncached(ticker):
    return Company(ticker)


def fetch_company(ticker):
    with _ProfileTimer("fetch_company"):
        key = str(ticker).strip().upper()
        with _FETCH_CACHE_LOCK:
            cached = _COMPANY_OBJECT_CACHE.get(key)
        if cached is not None:
            return cached
        company = _fetch_company_uncached(ticker)
        with _FETCH_CACHE_LOCK:
            return _COMPANY_OBJECT_CACHE.setdefault(key, company)


@retry_sec_request(retries=3, delay=5)
def _fetch_filings_uncached(company, form):
    return company.get_filings(form=form)


def get_company_filings(company, form):
    with _ProfileTimer("get_company_filings"):
        key = (_company_cache_key(company), _normalize_form_key(form))
        with _FETCH_CACHE_LOCK:
            cached = _FILINGS_CACHE.get(key)
        if cached is not None:
            return cached

        filings = _fetch_filings_uncached(company, form)
        with _FETCH_CACHE_LOCK:
            return _FILINGS_CACHE.setdefault(key, filings)


def fetch_filings(company, limit):
    return get_company_filings(company, ["10-K", "10-Q"]).head(limit + 4)


def _native_annual_filing_fy(filing, ye_month=12):
    """Infer a native 10-K fiscal year using only local filing metadata.

    This intentionally avoids calling ``fetch_filings``: default quarterly mode
    keeps the existing 10-K + 10-Q fetch path, while ``--annual`` gets its own
    10-K / 10-K/A listing.
    """
    period = None
    for attr in ("period_of_report", "period_end", "report_date"):
        try:
            period = getattr(filing, attr, None)
        except Exception:
            period = None
        if period:
            break
    if not period:
        return None
    try:
        end_dt = pd.to_datetime(period, errors="coerce")
        if pd.isna(end_dt):
            return None
        fy, _q = get_period_info(end_dt, ye_month)
        return int(fy)
    except Exception:
        return None


def fetch_native_annual_filings(company, limit, ye_month=12):
    """Fetch native annual filings for ``--annual`` mode only.

    The normal native quarterly fetch must remain ``fetch_filings(...)``.  This
    function scans 10-K and 10-K/A filings until it has enough distinct fiscal
    years, allowing for amendments/transition filings, then the annual builder
    trims final columns to ``limit``.
    """
    target = max(1, int(limit or 1))
    filings = get_company_filings(company, ["10-K", "10-K/A"])
    selected, usable_fys = [], set()
    max_scan = max(target + 12, target * 4)

    for filing in filings:
        selected.append(filing)
        fy = _native_annual_filing_fy(filing, ye_month=ye_month)
        if fy is not None:
            usable_fys.add(fy)
        if len(usable_fys) >= target + 1 or len(selected) >= max_scan:
            break

    return selected


@retry_sec_request(retries=5, delay=8)
def _fetch_xbrl_uncached(filing):
    return filing.xbrl()


def fetch_xbrl(filing):
    # XBRL objects for a SEC filing are immutable for the duration of one
    # extraction run.  Cache only successful parses; exceptions still follow
    # the original retry/backoff behavior.
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_XBRL_CACHE.get(key)
    if cached is not None:
        _profile_count("xbrl_cache_hits")
        return cached
    _profile_count("xbrl_cache_misses")
    xbrl = _fetch_xbrl_uncached(filing)
    if xbrl is not None:
        with _FETCH_CACHE_LOCK:
            return _FILING_XBRL_CACHE.setdefault(key, xbrl)
    return xbrl


@retry_sec_request(retries=5, delay=8)
def _fetch_html_uncached(filing):
    return filing.html()


def fetch_html(filing):
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_HTML_CACHE.get(key)
    if cached is not None:
        _profile_count("html_cache_hits")
        return cached
    _profile_count("html_cache_misses")
    html = _fetch_html_uncached(filing)
    if html is not None:
        with _FETCH_CACHE_LOCK:
            return _FILING_HTML_CACHE.setdefault(key, html)
    return html

# Parsed filing HTML caches. These are behavior-preserving speed caches:
# they reuse the exact same hidden-node removal and pd.read_html output that
# the old recovery/rescue paths created repeatedly for the same immutable SEC
# filing. Treat cached tables/rows as read-only.
_FILING_HTML_TABLES_CACHE = {}
_FILING_HTML_RESCUE_ROWS_CACHE = {}

# Optional persistent parsed-HTML cache. This is separate from raw EDGAR local
# storage: it avoids repeating BeautifulSoup + pd.read_html for immutable
# filings/TextBlocks across separate CLI re-runs. Values are keyed by stable
# accession/TextBlock hash plus parser/source fingerprint, so code/parser edits
# invalidate old parsed tables automatically.
_PERSISTENT_HTML_CACHE_VERSION = "2026-07-07.html-parse.v1"
_PERSISTENT_HTML_CACHE_DISABLED = {"0", "false", "no", "off", "disable", "disabled"}
_PERSISTENT_HTML_CACHE_ENABLED = (
    os.environ.get("SEC_PERSISTENT_HTML_CACHE", "1").strip().lower()
    not in _PERSISTENT_HTML_CACHE_DISABLED
)

# All script-managed persistent caches default to one project-local cache
# directory instead of creating separate cache folders under output/.
# Individual legacy cache-dir environment variables still override their own
# locations, and SEC_CACHE_DIR can move the whole cache bundle at once.
_SCRIPT_CACHE_ROOT = os.path.abspath(os.environ.get("SEC_CACHE_DIR", ".cache"))


def _script_cache_dir(*parts):
    return os.path.join(_SCRIPT_CACHE_ROOT, *parts)


_PERSISTENT_HTML_CACHE_ROOT = os.path.abspath(
    os.environ.get("SEC_PERSISTENT_HTML_CACHE_DIR",
                   _script_cache_dir("sec_html_parse"))
)
_PERSISTENT_HTML_CACHE_MISS = object()
_PERSISTENT_HTML_CACHE_LOCK = threading.RLock()


def _persistent_cache_source_fingerprint():
    try:
        return _fx_code_fingerprint()
    except Exception:
        try:
            with open(__file__, "rb") as fh:
                return hashlib.sha256(fh.read()).hexdigest()
        except Exception:
            return _PERSISTENT_HTML_CACHE_VERSION


def _stable_filing_identity_for_persistent_cache(filing):
    try:
        meta = _get_filing_local_metadata(filing)
        acc = meta.get("accession_no") or meta.get("accession_number")
        if not acc:
            return None
        return (
            str(meta.get("form") or ""),
            str(acc),
            str(meta.get("filing_date") or ""),
        )
    except Exception:
        return None


def _persistent_html_cache_path(kind, key):
    if not _PERSISTENT_HTML_CACHE_ENABLED or key is None:
        return None
    try:
        key_blob = pickle.dumps((_PERSISTENT_HTML_CACHE_VERSION, kind, key), protocol=4)
    except Exception:
        key_blob = repr((_PERSISTENT_HTML_CACHE_VERSION, kind, key)).encode("utf-8", "replace")
    digest = hashlib.sha256(key_blob).hexdigest()
    return os.path.join(_PERSISTENT_HTML_CACHE_ROOT, str(kind), digest[:2], f"{digest}.pkl")


def _persistent_html_cache_get(kind, key):
    path = _persistent_html_cache_path(kind, key)
    if not path:
        return _PERSISTENT_HTML_CACHE_MISS
    try:
        with _PERSISTENT_HTML_CACHE_LOCK:
            if not os.path.exists(path):
                return _PERSISTENT_HTML_CACHE_MISS
            with open(path, "rb") as fh:
                payload = pickle.load(fh)
        if not isinstance(payload, dict):
            return _PERSISTENT_HTML_CACHE_MISS
        if payload.get("version") != _PERSISTENT_HTML_CACHE_VERSION:
            return _PERSISTENT_HTML_CACHE_MISS
        _profile_count(f"persistent_html_{kind}_hits")
        return payload.get("value")
    except Exception:
        _profile_count(f"persistent_html_{kind}_read_failures")
        return _PERSISTENT_HTML_CACHE_MISS


def _persistent_html_cache_set(kind, key, value):
    path = _persistent_html_cache_path(kind, key)
    if not path:
        return
    tmp = f"{path}.{os.getpid()}.{threading.get_ident()}.tmp"
    try:
        with _PERSISTENT_HTML_CACHE_LOCK:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(tmp, "wb") as fh:
                pickle.dump({"version": _PERSISTENT_HTML_CACHE_VERSION, "value": value},
                            fh, protocol=pickle.HIGHEST_PROTOCOL)
            os.replace(tmp, path)
        _profile_count(f"persistent_html_{kind}_writes")
    except Exception:
        _profile_count(f"persistent_html_{kind}_write_failures")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


def _filing_html_persistent_cache_key(filing, kind):
    ident = _stable_filing_identity_for_persistent_cache(filing)
    if ident is None:
        return None
    return (
        "filing-html",
        str(kind),
        _persistent_cache_source_fingerprint(),
        ident,
    )


def _textblock_persistent_cache_key(value):
    val_text = str(value)
    digest = hashlib.sha256(val_text.encode('utf-8', 'surrogatepass')).hexdigest()
    return (
        "textblock-html",
        _persistent_cache_source_fingerprint(),
        digest,
    )


def _get_filing_html_tables_cached(filing):
    """Return hidden-cleaned pd.read_html tables for a filing, cached per accession.

    This does not change table discovery logic: it uses the same fetch_html ->
    BeautifulSoup -> hidden style removal -> pd.read_html pipeline that the
    callers previously ran inline. The cache only avoids rebuilding the same
    table list when multiple repair paths inspect the same filing.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_HTML_TABLES_CACHE.get(key)
    if cached is not None:
        _profile_count("filing_html_tables_cache_hits")
        return cached

    persistent_key = _filing_html_persistent_cache_key(filing, "tables")
    persistent = _persistent_html_cache_get("filing_tables", persistent_key)
    if persistent is not _PERSISTENT_HTML_CACHE_MISS:
        tables = tuple(persistent)
        with _FETCH_CACHE_LOCK:
            return _FILING_HTML_TABLES_CACHE.setdefault(key, tables)

    _profile_count("filing_html_tables_cache_misses")
    html_content = fetch_html(filing)
    soup = BeautifulSoup(html_content, 'html.parser')
    for hidden in soup.find_all(style=_DISPLAY_NONE_RE):
        hidden.decompose()
    tables = tuple(pd.read_html(StringIO(str(soup))))
    _persistent_html_cache_set("filing_tables", persistent_key, tables)

    with _FETCH_CACHE_LOCK:
        return _FILING_HTML_TABLES_CACHE.setdefault(key, tables)


def _normalize_rescue_row_values(row_values):
    """Mirror the old HTML-rescue row-value cleanup exactly."""
    row_vals = [
        str(x).strip()
        for x in row_values
        if str(x) != 'nan' and str(x).strip() not in ('', '$')
    ]
    if len(row_vals) < 2:
        return None
    return tuple(row_vals)


class _HtmlRescueRow:
    """Immutable row prepared for HTML segment rescue.

    ``values`` preserves the old cleaned cell strings and ordering, while
    ``nums`` caches the numeric parse that the rescue loop previously repeated
    for every category/quarter scan.
    """
    __slots__ = ("values", "nums", "name0_lower", "name01_lower")

    def __init__(self, values):
        self.values = tuple(values)
        self.nums = tuple(_parse_rescue_numeric_cells(self.values[1:]))
        self.name0_lower = self.values[0].lower() if self.values else ""
        self.name01_lower = " ".join(self.values[:2]).lower() if len(self.values) >= 2 else self.name0_lower

    def __getitem__(self, idx):
        return self.values[idx]

    def __iter__(self):
        return iter(self.values)

    def __len__(self):
        return len(self.values)


def _parse_rescue_numeric_cells(cells):
    nums = []
    for cell in cells:
        clean = re.sub(r'[^\d\.\(\)-]', '', cell)
        if clean:
            if '(' in cell or (len(cell) > 1 and cell.startswith('-')):
                clean = '-' + clean.replace('(', '').replace(')', '').replace('-', '')
            try:
                nums.append(float(clean))
            except ValueError:
                continue
    return nums


def _get_filing_html_rescue_rows_cached(filing):
    """Return pre-cleaned table rows for segment HTML rescue, cached per filing.

    The segment rescue scans the same tables once per category. This helper
    keeps the table order, row order, and row-value cleanup identical to
    the original inline loop, but computes the reusable row values once per filing.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_HTML_RESCUE_ROWS_CACHE.get(key)
    if cached is not None:
        return cached

    persistent_key = _filing_html_persistent_cache_key(filing, "rescue_rows")
    persistent = _persistent_html_cache_get("rescue_rows", persistent_key)
    if persistent is not _PERSISTENT_HTML_CACHE_MISS:
        prepared_tables = tuple(persistent)
        with _FETCH_CACHE_LOCK:
            return _FILING_HTML_RESCUE_ROWS_CACHE.setdefault(key, prepared_tables)

    prepared_tables = []
    for t_df in _get_filing_html_tables_cached(filing):
        t_df = t_df.dropna(how='all', axis=0).dropna(how='all', axis=1)
        if t_df.empty:
            continue

        prepared_rows = []
        for row_values in t_df.itertuples(index=False, name=None):
            normalized = _normalize_rescue_row_values(row_values)
            if normalized is not None:
                prepared_rows.append(_HtmlRescueRow(normalized))
        if prepared_rows:
            prepared_tables.append(tuple(prepared_rows))

    prepared_tables = tuple(prepared_tables)
    _persistent_html_cache_set("rescue_rows", persistent_key, prepared_tables)
    with _FETCH_CACHE_LOCK:
        return _FILING_HTML_RESCUE_ROWS_CACHE.setdefault(key, prepared_tables)


def _get_textblock_html_tables_cached(value):
    """Parse TextBlock HTML tables once per exact TextBlock string.

    The cache key is the SHA-256 of the exact TextBlock text.  Returned
    DataFrames are copies so callers can keep using the same drop/filter code
    without ever mutating the cached table objects.
    """
    val_text = str(value)
    key = hashlib.sha256(val_text.encode('utf-8', 'surrogatepass')).hexdigest()
    with _FETCH_CACHE_LOCK:
        cached = _TEXTBLOCK_HTML_TABLES_CACHE.get(key)
    if cached is not None:
        _profile_count("textblock_html_table_cache_hits")
        return tuple(t.copy(deep=True) for t in cached)

    persistent_key = _textblock_persistent_cache_key(value)
    persistent = _persistent_html_cache_get("textblock_tables", persistent_key)
    if persistent is not _PERSISTENT_HTML_CACHE_MISS:
        tables = tuple(persistent)
        with _FETCH_CACHE_LOCK:
            stored = _TEXTBLOCK_HTML_TABLES_CACHE.setdefault(key, tables)
        return tuple(t.copy(deep=True) for t in stored)

    _profile_count("textblock_html_table_cache_misses")
    with _ProfileTimer("textblock_html_parse"):
        soup = BeautifulSoup(val_text, 'html.parser')
        for hidden in soup.find_all(style=_DISPLAY_NONE_RE):
            hidden.decompose()
        tables = tuple(pd.read_html(StringIO(str(soup))))
    _profile_count("textblock_html_tables_parsed", len(tables))
    _persistent_html_cache_set("textblock_tables", persistent_key, tables)
    with _FETCH_CACHE_LOCK:
        stored = _TEXTBLOCK_HTML_TABLES_CACHE.setdefault(key, tables)
    return tuple(t.copy(deep=True) for t in stored)


def _build_filing_period_lookup(filings, ye_month, wanted_periods=None):
    """Map period column (YYYY-Qn) to the first matching filing.

    Preserves the old rescue behavior where the first filing encountered for a
    matching fiscal quarter wins; it just avoids repeating the same scan for
    every failed quarter.  When ``wanted_periods`` is supplied, only those
    period keys are material to the caller, so the scan stops once they are all
    found.
    """
    by_period = {}
    wanted = set(wanted_periods) if wanted_periods is not None else None
    for f in filings:
        f_end = pd.to_datetime(f.period_of_report)
        f_fy, f_q = get_period_info(f_end, ye_month)
        key = f"{f_fy}-{f_q}"
        if wanted is not None and key not in wanted:
            continue
        if key not in by_period:
            by_period[key] = f
            if wanted is not None and wanted.issubset(by_period):
                break
    return by_period

# ---------------------------------------------------------------------------
# Safe filing-identity helpers (zero additional network calls in error paths)
# ---------------------------------------------------------------------------
def safe_filing_id(filing):
    """
    Return a human-readable filing identifier using ONLY locally cached
    attributes -- no network calls.  Used in error/retry handlers so that
    a timeout on one fetch never cascades into another.
    """
    try:
        meta = _get_filing_local_metadata(filing)
        form = meta.get('form') or '?'
        filed = meta.get('filing_date') or '?'
        acc = meta.get('accession_no') or meta.get('accession_number') or '?'
        return f"{form} acc={acc} filed={filed}"
    except Exception:
        return repr(filing)


def fetch_period_of_report(filing, retries=4, base_delay=10):
    """
    Fetch filing.period_of_report with its own dedicated retry loop.

    This attribute triggers a lightweight SGML header fetch in the edgar
    library.  Retrying it here (with back-off) gives permanently-failed
    filings a last chance to at least contribute their period date, so
    the output CSV shows an (empty) column rather than silently omitting
    the quarter entirely.

    Returns the period string on success, or None after all retries.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_PERIOD_REPORT_CACHE.get(key)
    if cached is not None:
        return cached

    with _ProfileTimer("fetch_period_of_report"):
        for attempt in range(1, retries + 1):
            try:
                sec_limiter.wait()
                period = filing.period_of_report
                if period:
                    with _FETCH_CACHE_LOCK:
                        _FILING_PERIOD_REPORT_CACHE.setdefault(key, period)
                return period
            except Exception as e:
                if attempt == retries:
                    return None
                wait = base_delay * attempt
                print(f"  [Period Retry {attempt}/{retries}] "
                      f"period_of_report failed ({type(e).__name__}), retrying in {wait}s...")
                time.sleep(wait)
    return None


def estimate_period_from_filing_date(filing):
    """
    Last-resort fallback: estimate fiscal-quarter end from the filing date.

    10-K filers must file within 60-90 days of FY end.
    10-Q filers must file within 40-45 days of quarter end.

    Returns (date_string, True) on success, or (None, False).
    The True flag signals to callers that the date is an estimate only.
    """
    key = _filing_cache_key(filing)
    with _FETCH_CACHE_LOCK:
        cached = _FILING_PERIOD_ESTIMATE_CACHE.get(key)
    if cached is not None:
        return cached

    try:
        meta = _get_filing_local_metadata(filing)
        filing_dt = pd.to_datetime(meta.get('filing_date'))
        form = meta.get('form') or ''
        offset = 75 if '10-K' in form else 45
        estimated = filing_dt - pd.DateOffset(days=offset)
        result = (estimated.strftime('%m/%d/%y'), True)
    except Exception:
        result = (None, False)
    with _FETCH_CACHE_LOCK:
        return _FILING_PERIOD_ESTIMATE_CACHE.setdefault(key, result)


def extract_l2_segments_from_html(facts_df, expected_segments):
    with _ProfileTimer("extract_l2_segments_from_html"):
        return _extract_l2_segments_from_html_impl(facts_df, expected_segments)


def _extract_l2_segments_from_html_impl(facts_df, expected_segments):
    '''
    Parse HTML TextBlock XBRL facts to extract the human-readable segment
    revenue table.  Returns a dict of {segment_name: [col0, col1, ...]} where
    col0 is always the *current reporting period* value.

    Design notes
    ------------
    *  Searches segment-specific TextBlock facts first, then all TextBlocks.
    *  For each HTML table found, scans rows for a 'Revenue' header then
       collects segment rows until a different metric header is encountered.
    *  Deduplication strategy: adjacent identical values in a row are
       collapsed ONLY when there are more numeric cells than expected columns
       (SEC tables sometimes use merged/repeated header cells). We keep up to
       4 distinct values, preserving col0 (current-period) accuracy.
    *  Requires â‰¥ 2 expected segments to be present in the same table before
       accepting results.
    '''
    if not expected_segments:
        return {}

    text_blocks = facts_df[facts_df['concept'].str.contains('TextBlock', case=False, na=False)]
    _profile_count("textblock_rows_seen", len(text_blocks))
    segment_blocks = text_blocks[text_blocks['concept'].str.contains('Segment', case=False)]
    _profile_count("segment_textblock_rows_seen", len(segment_blocks))
    blocks_to_search = pd.concat([segment_blocks, text_blocks.drop(segment_blocks.index)])

    expected_lower = {s.lower(): s for s in expected_segments}
    best_results: dict[str, list[float]] = {}
    # Invariant across all candidate tables (depends only on expected_segments).
    min_matches = max(2, len(expected_segments) // 2)

    REVENUE_HEADERS   = {'revenue', 'net revenue', 'net sales', 'total revenue', 'revenues'}
    STOP_HEADERS      = {'operating income', 'operating profit', 'gross profit',
                         'segment income', 'operating loss', 'total'}

    for raw_val in blocks_to_search['value'].to_numpy(copy=False):
        val = str(raw_val)
        if '<table' not in val.lower():
            continue

        # Quick pre-screen: need â‰¥ 2 expected segment names in this block
        matches_pre = sum(1 for s in expected_lower if s in val.lower())
        if matches_pre < min(2, len(expected_segments)):
            continue

        try:
            tables = _get_textblock_html_tables_cached(val)
        except Exception:
            continue

        for tbl in tables:
            tbl = tbl.dropna(how='all', axis=0).dropna(how='all', axis=1)
            if tbl.empty:
                continue

            in_revenue_section = False
            results: dict[str, list[float]] = {}

            for row_values in tbl.itertuples(index=False, name=None):
                cells = [
                    str(x).strip()
                    for x in row_values
                    if str(x) not in ('nan', 'NaN', '') and str(x).strip() not in ('$', '')
                ]
                if not cells:
                    continue

                row_str = ' '.join(cells).lower()

                # -- Section header detection ------------------------------
                # Enter revenue section when the first cell matches a revenue keyword
                # and there are no numeric values in the row (it's a header row).
                first_lower = cells[0].lower()
                if any(h == first_lower or first_lower.startswith(h) for h in REVENUE_HEADERS):
                    # Confirm it's a header (no digits in numeric cells)
                    has_nums = any(re.search(r'\d', c) for c in cells[1:])
                    if not has_nums:
                        in_revenue_section = True
                        continue

                if in_revenue_section:
                    # Stop when we hit a different metric section
                    if any(kw in row_str for kw in STOP_HEADERS):
                        if not any(s in row_str for s in expected_lower):
                            in_revenue_section = False
                            continue

                    # Try to match segment name in first cell or first two cells
                    matched_seg: str | None = None
                    for name_candidate in [cells[0], ' '.join(cells[:2])]:
                        nl = name_candidate.lower()
                        for lower_seg, orig_seg in expected_lower.items():
                            if (
                                nl == lower_seg
                                or nl.startswith(lower_seg + ' ')
                                or nl.endswith(' ' + lower_seg)
                                or lower_seg in nl
                            ):
                                matched_seg = orig_seg
                                break
                        if matched_seg:
                            break

                    if matched_seg and matched_seg not in results:
                        nums: list[float] = []
                        for cell in cells[1:]:
                            clean = re.sub(r'[^\d\.\(\)-]', '', cell)
                            if not clean:
                                continue
                            if '(' in cell or (len(cell) > 1 and cell.startswith('-')):
                                clean = '-' + clean.replace('(', '').replace(')', '').replace('-', '')
                            try:
                                nums.append(float(clean))
                            except ValueError:
                                continue

                        if nums:
                            # Smart dedup: only collapse *adjacent equal* values when
                            # there are more raw values than typical column count (4).
                            # This preserves cases where prior-year QTD == current-year
                            # QTD by coincidence.
                            if len(nums) > 4:
                                deduped: list[float] = []
                                for n in nums:
                                    if not deduped or deduped[-1] != n:
                                        deduped.append(n)
                                nums = deduped

                            results[matched_seg] = nums[:4]

            # Accept table if it contains â‰¥ min_matches segments
            # (min_matches is invariant; computed once above).
            if len(results) >= min_matches:
                if len(results) > len(best_results):
                    best_results = results
                if len(best_results) >= len(expected_segments):
                    return best_results

    return best_results

# ---------------------------------------------------------------------------
# Comprehensive Concept Mapping
# ---------------------------------------------------------------------------
CONCEPT_MAP = {
    # 1. INCOME STATEMENT
    'Revenue': {'tags': ['SalesRevenueNet', 'Revenues', 'RevenuesNetOfInterestExpense', 'RevenueFromContractWithCustomerExcludingAssessedTax', 'RevenueFromContractWithCustomerIncludingAssessedTax', 'SalesRevenueGross', 'RealEstateRevenueNet', 'OperatingLeaseLeaseIncome'], 'cat': '1_Income_Statement'},
    'Product Revenue': {'tags': ['SalesRevenueGoodsNet'], 'cat': '1_Income_Statement'},
    'Service Revenue': {'tags': ['ServiceRevenue'], 'cat': '1_Income_Statement'},
    'Subscription Revenue': {'tags': ['SubscriptionRevenue'], 'cat': '1_Income_Statement'},
    'Licenses Revenue': {'tags': ['LicensesRevenue'], 'cat': '1_Income_Statement'},
    'Lease & Other Revenue': {'tags': ['OperatingLeasesIncomeStatementLeaseRevenue', 'RevenueFromRelatedParties', 'NetInterestIncome'], 'cat': '1_Income_Statement'},
    'Cost of Revenue': {'tags': ['CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfGoodsAndServiceExcludingDepreciationDepletionAndAmortization', 'ProvisionForLoanAndLeaseLosses'], 'cat': '1_Income_Statement'},
    'Cost of Goods Sold': {'tags': ['CostOfGoodsSold', 'CostOfGoodsSoldRelatedParty'], 'cat': '1_Income_Statement'},
    'Cost of Services': {'tags': ['CostOfServices'], 'cat': '1_Income_Statement'},
    'Cost of Lease & Other Revenue': {'tags': ['OperatingLeasesCostOfLeaseRevenue'], 'cat': '1_Income_Statement'},
    'Gross Profit': {'tags': ['GrossProfit', 'RealEstateGrossProfit', 'GrossProfitRelatedParty'], 'cat': '1_Income_Statement'},
    'Premiums Earned': {'tags': ['PremiumsEarnedNet', 'LifeInsurancePremiums', 'HealthCarePremiumsNet', 'PremiumsAndOtherConsideration'], 'cat': '1_Income_Statement'},
    'Net Investment Income': {'tags': ['InvestmentIncomeNet', 'NetInvestmentIncome', 'InvestmentIncomeInterest'], 'cat': '1_Income_Statement'},
    'Policyholder Claims/Benefits': {'tags': ['PolicyholderBenefitsAndClaimsIncurredNet', 'BenefitsLossesAndExpenses', 'HealthCareMedicalBenefits'], 'cat': '1_Income_Statement'},
    'Amortization of DAC': {'tags': ['AmortizationOfDeferredAcquisitionCosts', 'DeferredAcquisitionCostsAmortization'], 'cat': '1_Income_Statement'},
    'Research & Development': {'tags': ['TechnologyAndDevelopmentExpense', 'ResearchAndDevelopmentExpense', 'ResearchAndDevelopmentExpenseExcludingAcquiredInProcessCost'], 'cat': '1_Income_Statement'},
    'In-Process R&D': {'tags': ['ResearchAndDevelopmentInProcess', 'ResearchAndDevelopmentExpenseSoftwareExcludingAcquiredInProcessCost'], 'cat': '1_Income_Statement'},
    'Related Party R&D': {'tags': ['ResearchAndDevelopmentExpenseRelatedParty'], 'cat': '1_Income_Statement'},
    'Other Operating Expenses': {'tags': ['OtherCostAndExpenseOperating', 'OtherOperatingIncomeExpenseNet'], 'cat': '1_Income_Statement'},
    # -- SGA: Split into components so dedup doesn't lose one --------------
    # Companies that file the combined tag get "Selling, General & Admin".
    # Companies that split (e.g. Google) get separate S&M + G&A rows.
    # The OpInc bridge sums intelligently to avoid double-counting.
    'Selling, General & Admin': {'tags': ['SellingGeneralAndAdministrativeExpense'], 'cat': '1_Income_Statement'},
    'Exploration Expense': {'tags': ['ExplorationExpense', 'ExplorationAbandonmentAndDryHoleCostsExpense', 'ExplorationAbandonmentAndDryHoleCosts', 'ResultsOfOperationsExplorationExpenses', 'ExplorationCosts'], 'cat': '1_Income_Statement'},
    'Depreciation, Depletion & Amortization': {'tags': [], 'cat': '1_Income_Statement'},
    'Depreciation & Amortization Expense': {'tags': [], 'cat': '1_Income_Statement'},
    'Sales & Marketing': {'tags': ['SellingAndMarketingExpense', 'MarketingExpense', 'AdvertisingExpense', 'SellingExpense'], 'cat': '1_Income_Statement'},
    'General & Administrative': {'tags': ['GeneralAndAdministrativeExpense', 'OtherGeneralAndAdministrativeExpense', 'OtherNoninterestExpense'], 'cat': '1_Income_Statement'},
    'Customer Bad Debt': {'tags': ['CustomerBadDebt'], 'cat': '1_Income_Statement'},
    'Salaries & Employee Benefits': {'tags': ['LaborAndRelatedExpense'], 'cat': '1_Income_Statement'},
    'Marketing Expense': {'tags': ['MarketingExpense', 'MarketingAndAdvertisingExpense'], 'cat': '1_Income_Statement'},
    # -- Additional operating expense line items (between SGA and OpInc) ---
    'Amortization of Intangibles': {'tags': ['AmortizationOfIntangibleAssets', 'AmortizationOfAcquisitionCosts', 'AcquiredInPlaceLeasesAmortizationExpense'], 'cat': '1_Income_Statement'},
    'Restructuring & Related Charges': {'tags': ['RestructuringCharges', 'RestructuringSettlementAndImpairmentProvisions', 'RestructuringAndRelatedCostIncurredCost', 'SeveranceCosts1', 'RestructuringCostsAndAssetImpairmentCharges'], 'cat': '1_Income_Statement'},
    'Impairment Charges': {'tags': ['AssetImpairmentCharges', 'GoodwillImpairmentLoss', 'ImpairmentOfIntangibleAssetsExcludingGoodwill', 'ImpairmentOfLongLivedAssetsHeldForUse', 'TangibleAssetImpairmentCharges', 'OtherAssetImpairmentCharges'], 'cat': '1_Income_Statement'},
    'Acquisition-Related Costs': {'tags': ['BusinessCombinationAcquisitionRelatedCosts', 'MergerRelatedCosts', 'BusinessCombinationIntegrationRelatedCosts'], 'cat': '1_Income_Statement'},
    'Litigation & Settlement Charges': {'tags': ['LitigationSettlementAmountAwardedToOtherParty', 'LitigationSettlementExpense', 'LossContingencyLossInPeriod', 'GainLossRelatedToLitigationSettlement'], 'cat': '1_Income_Statement'},
    # CostsAndExpenses / OperatingCostsAndExpenses are COGS-inclusive total-cost
    # subtotals for many filers (META, AMZN-style P&Ls). Keep them separate from
    # true operating-expense-only totals so the output label matches the math.
    'Total Costs and Expenses': {'tags': ['OperatingCostsAndExpenses', 'CostsAndExpenses'], 'cat': '1_Income_Statement'},
    'Total Operating Expenses': {'tags': ['OperatingExpenses', 'NoninterestExpense', 'ExpenseAndIncomeOther', 'OtherExpenseAndIncome'], 'cat': '1_Income_Statement'},
    'Operating Income': {'tags': ['OperatingIncomeLoss', 'OperatingIncomeLossFromContinuingOperations', 'RealEstateOperatingIncome'], 'cat': '1_Income_Statement'},
    # -- Non-operating: separate entries to prevent dedup loss -------------
    # Interest Income and Interest Expense are standalone rows.
    # "Other Non-operating Income" captures the OTHER subcomponent.
    # "Total Non-operating Income" captures the TOTAL if filed separately.
    # The Pretax bridge sums: IntInc - IntExp + OtherNonOp (or TotalNonOp)
    'Interest Expense': {'tags': ['InterestExpense', 'InterestExpenseNonoperating', 'InterestExpenseDebt', 'InterestAndDebtExpense', 'InterestExpenseNet', 'InterestExpenseAndAmortizationOfDebtDiscountPremium', 'InterestExpenseOperating', 'InterestExpenseBorrowings'], 'cat': '1_Income_Statement'},
    'Interest Income': {'tags': ['InvestmentIncomeInterest', 'InterestIncomeOperating', 'InterestIncomeInterestEarningAsset', 'InterestAndDividendIncomeOperating', 'InterestIncomeNonoperating', 'InvestmentIncomeNet', 'InvestmentIncomeDividend', 'InvestmentIncomeNonoperatingNet', 'InvestmentIncomeNonoperating', 'TotalInterestAndDividendIncome'], 'cat': '1_Income_Statement'},
    'Net Interest Income (Expense)': {'tags': ['InterestIncomeExpenseNet'], 'cat': '1_Income_Statement'},
    'Gain/Loss on Investments': {'tags': ['GainLossOnInvestments', 'UnrealizedGainLossOnInvestments', 'EquitySecuritiesFvNiGainLoss', 'DebtSecuritiesGainLoss', 'GainLossOnSaleOfInvestments', 'RealizedInvestmentGainsLosses', 'GainLossOnSaleOfDerivatives', 'MarketableSecuritiesUnrealizedGainLoss', 'UnrealizedGainLossOnMarketableAndNonmarketableEquityInvestments', 'EquitySecuritiesFvNiUnrealizedGainLoss'], 'cat': '1_Income_Statement'},
    'Equity Method Income': {'tags': ['IncomeLossFromEquityMethodInvestments', 'IncomeLossFromEquityMethodInvestmentsNetOfDividendsOrDistributions', 'IncomeLossFromEquityMethodInvestmentsAndFairMarketValueAlternativeInvestments'], 'cat': '1_Income_Statement'},
    'Other Income / (Expense)': {'tags': ['OtherNonoperatingIncomeExpense', 'OtherNonoperatingIncome', 'MiscellaneousNonoperatingIncomeExpense', 'ForeignCurrencyTransactionGainLossBeforeTax', 'GainLossOnSaleOfPropertyPlantAndEquipment', 'GainLossOnSaleOfBusiness', 'GainLossOnExtinguishmentOfDebt'], 'cat': '1_Income_Statement'},
    'Total Non-operating Income': {'tags': ['NonoperatingIncomeExpense', 'NonoperatingIncomePlusOtherNonoperatingIncomeExpense'], 'cat': '1_Income_Statement'},
    'Pretax Income': {'tags': ['IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesNoncontrollingInterest', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesForeign', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesDomestic', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesAndAdoptionOfNewAccountingPrinciple'], 'cat': '1_Income_Statement'},
    'Income Tax Expense': {'tags': ['IncomeTaxExpenseBenefit', 'IncomeTaxExpenseBenefitContinuingOperations', 'CurrentIncomeTaxExpenseBenefit', 'DeferredIncomeTaxExpenseBenefit'], 'cat': '1_Income_Statement'},
    'Net Income': {'tags': ['NetIncomeLoss', 'ProfitLoss', 'NetIncomeLossAvailableToCommonStockholdersBasic', 'NetIncomeLossAttributableToParent', 'IncomeLossFromContinuingOperationsNetOfTax'], 'cat': '1_Income_Statement'},
    'Net Income to Noncontrolling Interest': {'tags': ['NetIncomeLossAttributableToNoncontrollingInterest'], 'cat': '1_Income_Statement'},
    'Income from Discontinued Operations': {'tags': ['IncomeLossFromDiscontinuedOperationsNetOfTax'], 'cat': '1_Income_Statement'},
    'EPS Basic': {'tags': ['EarningsPerShareBasic', 'IncomeLossFromContinuingOperationsNetOfTaxPerBasicShare', 'IncomeLossFromDiscontinuedOperationsNetOfTaxPerBasicShare', 'ExtraordinaryItemNetOfTaxPerBasicShare', 'CumulativeEffectOfNetOfTaxPerBasicShare', 'EarningsPerShareBasicAndDilutedOtherThanCommonStock'], 'cat': '1_Income_Statement'},
    'EPS Diluted': {'tags': ['EarningsPerShareDiluted', 'IncomeLossFromContinuingOperationsNetOfTaxPerDilutedShare', 'IncomeLossFromDiscontinuedOperationsNetOfTaxPerDilutedShare'], 'cat': '1_Income_Statement'},
    'Shares Outstanding Basic': {'tags': ['WeightedAverageNumberOfSharesOutstandingBasic', 'EntityCommonStockSharesOutstanding'], 'cat': '1_Income_Statement'},
    'Shares Outstanding Diluted': {'tags': ['WeightedAverageNumberOfDilutedSharesOutstanding', 'WeightedAverageNumberOfSharesOutstandingBasic', 'EntityCommonStockSharesOutstanding'], 'cat': '1_Income_Statement'},

    # 2. BALANCE SHEET
    'Cash & Equivalents': {'tags': ['CashAndCashEquivalentsAtCarryingValue', 'Cash', 'CashEquivalentsAtCarryingValue', 'RestrictedCashAndCashEquivalentsAtCarryingValue', 'RestrictedCash', 'RestrictedCashCurrent', 'RestrictedCashNoncurrent', 'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents', 'CashReserveDepositRequiredAndMade'], 'cat': '2_Balance_Sheet'},
    'Segregated Cash & Securities': {'tags': ['CashAndSecuritiesSegregatedForRegulatoryPurposesOrDepositedWithClearingOrganizations', 'SecuritiesSegregatedUnderRegulations', 'SecuritiesSegregatedForRegulatoryPurposesPurchasedUnderAgreementsToResell', 'SegregatedSecuritiesPurchasedUnderAgreementsToResell', 'SecuritiesBorrowedThatWereSegregatedToSatisfyRegulatoryRequirements'], 'cat': '2_Balance_Sheet'},
    'Securities Segregated': {'tags': ['CashAndSecuritiesSegregatedUnderFederalAndOtherRegulations'], 'cat': '2_Balance_Sheet'},
    'Marketable debt securities': {'tags': ['DebtSecuritiesCurrent', 'DebtSecuritiesAvailableForSaleExcludingAccruedInterestCurrent'], 'cat': '2_Balance_Sheet'},
    'Marketable equity securities': {'tags': ['EquitySecuritiesFvNi'], 'cat': '2_Balance_Sheet'},
    'Short-term Investments': {'tags': ['MarketableSecuritiesCurrent', 'ShortTermInvestments', 'AvailableForSaleSecuritiesCurrent', 'AvailableForSaleSecuritiesDebtSecuritiesCurrent', 'HeldToMaturitySecuritiesCurrent', 'TradingSecuritiesCurrent', 'Investments'], 'cat': '2_Balance_Sheet'},
    'Accounts Receivable': {'tags': ['AccountsReceivableNetCurrent', 'ReceivablesNetCurrent', 'AccountsAndNotesReceivableNetCurrent', 'AccountsReceivableNet', 'ReceivablesNet'], 'cat': '2_Balance_Sheet'},
    'Accounts Receivable (Gross)': {'tags': ['AccountsReceivableGrossCurrent'], 'cat': '2_Balance_Sheet'},
    'Unbilled Receivables': {'tags': ['UnbilledReceivablesCurrent'], 'cat': '2_Balance_Sheet'},
    'Related Party Receivables': {'tags': ['AccountsReceivableRelatedPartiesCurrent'], 'cat': '2_Balance_Sheet'},
    'Trade Receivables Held For Sale': {'tags': ['TradeReceivablesHeldForSaleAmount'], 'cat': '2_Balance_Sheet'},
    'Prepaid Expenses': {'tags': ['PrepaidExpenseAndOtherAssetsCurrent'], 'cat': '2_Balance_Sheet'},
    'Restricted Cash': {'tags': ['RestrictedCashAndCashEquivalentsAtCarryingValue', 'RestrictedCashAndCashEquivalents', 'RestrictedCashAndInvestments', 'RestrictedCashAndCashEquivalentsU.S.LitigationEscrow'], 'cat': '2_Balance_Sheet'},
    'Inventory': {'tags': ['InventoryNet', 'InventoryGross', 'InventoryNetCurrent', 'InventoryNetNoncurrent', 'InventoryNetOfAllowancesCustomerAdvancesAndProgressBillings'], 'cat': '2_Balance_Sheet'},
    'Inventory: Raw Materials': {'tags': ['InventoryRawMaterials', 'InventoryRawMaterialsAndSupplies', 'InventoryRawMaterialsNetOfReserves'], 'cat': '2_Balance_Sheet'},
    'Inventory: Work in Process': {'tags': ['InventoryWorkInProcess', 'InventoryWorkInProcessNetOfReserves'], 'cat': '2_Balance_Sheet'},
    'Inventory: Finished Goods': {'tags': ['InventoryFinishedGoods', 'InventoryFinishedGoodsNetOfReserves'], 'cat': '2_Balance_Sheet'},
    'Total Current Assets': {'tags': ['AssetsCurrent'], 'cat': '2_Balance_Sheet'},
    'Real Estate, Net': {'tags': ['RealEstateInvestmentPropertyNet', 'RealEstateInvestmentPropertyAtCost', 'RealEstateGrossAtCarryingValue', 'RealEstateAndAccumulatedDepreciationNetOfAccumulatedDepreciation'], 'cat': '2_Balance_Sheet'},
    'Property, Plant & Equipment': {'tags': ['PropertyPlantAndEquipmentNet', 'PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetAfterAccumulatedDepreciationAndAmortization', 'PropertyPlantandEquipmentandFinanceLeaseRightofUseAssetafterAccumulatedDepreciationandAmortization', 'PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetNet', 'OilAndGasPropertySuccessfulEffortMethodNet', 'OilAndGasPropertyFullCostMethodNet', 'PropertyPlantAndEquipmentGross', 'OilAndGasPropertySuccessfulEffortMethodGross', 'OilAndGasPropertyFullCostMethodGross'], 'cat': '2_Balance_Sheet'},
    'Operating Lease ROU Asset': {'tags': ['OperatingLeaseRightOfUseAsset'], 'cat': '2_Balance_Sheet'},
    'Long-term Investments': {'tags': ['MarketableSecuritiesNoncurrent', 'LongTermInvestments', 'OtherLongTermInvestments', 'EquitySecuritiesFVNINoncurrent', 'AvailableForSaleDebtSecuritiesNoncurrent', 'AvailableForSaleSecuritiesNoncurrent', 'NonMarketableSecurities', 'EquitySecuritiesFvNiCurrentAndNoncurrent', 'InvestmentsInAffiliatesSubsidiariesAssociatesAndJointVentures'], 'cat': '2_Balance_Sheet'},
    'Other Non-Current Assets': {'tags': ['OtherAssetsNoncurrent'], 'cat': '2_Balance_Sheet'},
    'Intangible Assets (Net)': {'tags': ['IntangibleAssetsNetExcludingGoodwill'], 'cat': '2_Balance_Sheet'},
    'Intangible Assets & Goodwill': {'tags': ['IntangibleAssetsNetIncludingGoodwill'], 'cat': '2_Balance_Sheet'},
    'Receivables from Brokers & Dealers': {'tags': ['ReceivablesFromBrokerDealersAndClearingOrganizations', 'ReceivablesFromBrokersDealersAndClearingOrganizations', 'BrokerageReceivables', 'ReceivablesFromClearingOrganizations', 'OtherReceivablesFromBrokerDealersAndClearingOrganizations', 'IncreaseDecreaseInBrokerageReceivables'], 'cat': '2_Balance_Sheet'},
    'Customer Receivables': {'tags': ['CustomerReceivables', 'ReceivablesFromCustomers', 'CustomerAccountsReceivable', 'ReceivablesFromCustomerNet', 'ReceivablesFromAndPayablesToBrokerageClientsTextBlock'], 'cat': '2_Balance_Sheet'},
    'Other Receivables': {'tags': ['OtherReceivables', 'OtherReceivablesNet', 'CustomerAndOtherReceivables', 'IncreaseDecreaseInOtherReceivables'], 'cat': '2_Balance_Sheet'},
    'Loans & Leases (Net)': {'tags': ['LoansAndLeasesReceivableNetOfAllowance', 'LoansAndLeasesReceivableNetReportedAmount', 'LoansReceivableNet', 'LoansAndLeasesReceivableNetOfDeferredIncome', 'LoansReceivableNetOfAllowance'], 'cat': '2_Balance_Sheet'},
    'Securities Borrowed': {'tags': ['SecuritiesBorrowed', 'SecuritiesBorrowedGross', 'SecuritiesBorrowedLiability'], 'cat': '2_Balance_Sheet'},
    'Securities Purchased under Agreements to Resell': {'tags': ['SecuritiesPurchasedUnderAgreementsToResell', 'SecuritiesPurchasedUnderAgreementsToResellGross', 'SecuritiesPurchasedUnderAgreementsToResellLiability'], 'cat': '2_Balance_Sheet'},
    'Margin Loans': {'tags': ['MarginLoansReceivable', 'MarginReceivables'], 'cat': '2_Balance_Sheet'},
    'Financial Instruments Owned': {'tags': ['FinancialInstrumentsOwnedAtFairValue', 'TradingSecurities'], 'cat': '2_Balance_Sheet'},
    'Total Assets': {'tags': ['Assets'], 'cat': '2_Balance_Sheet'},
    'Accounts Payable': {'tags': ['AccountsPayableCurrent', 'AccountsPayableAndAccruedLiabilitiesCurrent', 'AccountsPayableAndAccruedLiabilitiesFairValueDisclosure'], 'cat': '2_Balance_Sheet'},
    'Accounts Payable (Trade)': {'tags': ['AccountsPayableTradeCurrent'], 'cat': '2_Balance_Sheet'},
    'Accounts Payable (Other)': {'tags': ['AccountsPayableOtherCurrent', 'AccountsPayableOtherCurrentAndNoncurrent'], 'cat': '2_Balance_Sheet'},
    'Accounts Payable (Related Parties)': {'tags': ['AccountsPayableRelatedPartiesCurrent'], 'cat': '2_Balance_Sheet'},
    'Other Payables': {'tags': ['OtherAccountsPayableAndAccruedLiabilities', 'OtherAccountsPayable', 'CustomerAndOtherPayables', 'OtherPayablesToBrokerDealersAndClearingOrganizations'], 'cat': '2_Balance_Sheet'},
    'Total Payables': {'tags': ['AccountsPayableAndAccruedLiabilities', 'AccountsPayable'], 'cat': '2_Balance_Sheet'},
    'Accrued Expenses': {'tags': ['AccruedLiabilitiesCurrent', 'AccruedExpensesAndOtherLiabilitiesCurrent', 'AccruedExpensesAndOtherCurrent', 'AccruedLiabilitiesAndOtherCurrent'], 'cat': '2_Balance_Sheet'},
    'Accrued Employee Liabilities': {'tags': ['EmployeeRelatedLiabilitiesCurrent'], 'cat': '2_Balance_Sheet'},
    'Deferred Revenue': {'tags': ['DeferredRevenueCurrent', 'ContractWithCustomerLiabilityCurrent', 'ContractLiabilityCurrent', 'DeferredRevenue', 'DeferredRevenueNoncurrent', 'ContractWithCustomerLiability'], 'cat': '2_Balance_Sheet'},
    'Short-term Debt': {'tags': ['DebtCurrent'], 'cat': '2_Balance_Sheet'},
    'Short-term Borrowings': {'tags': ['ShortTermBorrowings', 'CommercialPaper', 'LinesOfCreditCurrent', 'NotesPayableCurrent', 'ShortTermBankLoansAndNotesPayable', 'OtherShortTermBorrowings'], 'cat': '2_Balance_Sheet'},
    'Current Portion of Long-term Debt': {'tags': ['LongTermDebtCurrent', 'ConvertibleDebtCurrent'], 'cat': '2_Balance_Sheet'},
    'Operating Lease Liability (Current)': {'tags': ['OperatingLeaseLiabilityCurrent'], 'cat': '2_Balance_Sheet'},
    'Finance Lease Liability (Current)': {'tags': ['FinanceLeaseLiabilityCurrent'], 'cat': '2_Balance_Sheet'},
    'Total Current Liabilities': {'tags': ['LiabilitiesCurrent'], 'cat': '2_Balance_Sheet'},
    'Goodwill': {'tags': ['Goodwill', 'GoodwillAcquiredDuringPeriod'], 'cat': '2_Balance_Sheet'},
    'Long-term Debt': {'tags': ['LongTermDebtNoncurrent', 'LongTermDebt', 'LongTermDebtAndCapitalLeaseObligations'], 'cat': '2_Balance_Sheet'},
    'Senior Notes': {'tags': ['SeniorLongTermNotes', 'SeniorNotes'], 'cat': '2_Balance_Sheet'},
    'Convertible Debt': {'tags': ['ConvertibleDebtNoncurrent', 'ConvertibleLongTermNotesPayable'], 'cat': '2_Balance_Sheet'},
    'Other Long-term Borrowings': {'tags': ['LinesOfCreditNoncurrent', 'LongTermLoansPayable', 'NotesPayable'], 'cat': '2_Balance_Sheet'},
    'Operating Lease Liability (Non-current)': {'tags': ['OperatingLeaseLiabilityNoncurrent', 'OperatingLeaseLiability'], 'cat': '2_Balance_Sheet'},
    'Finance Lease Liability (Non-current)': {'tags': ['FinanceLeaseLiabilityNoncurrent'], 'cat': '2_Balance_Sheet'},
    'Other Non-Current Liabilities': {'tags': ['OtherLiabilitiesNoncurrent', 'DeferredTaxLiabilitiesNoncurrent', 'OperatingLeaseLiabilityNoncurrent', 'ContractWithCustomerLiabilityNoncurrent', 'DeferredRevenueNoncurrent', 'PensionAndOtherPostretirementDefinedBenefitPlansLiabilitiesNoncurrent', 'DefinedBenefitPensionPlanLiabilitiesNoncurrent', 'AssetRetirementObligation', 'AssetRetirementObligationsNoncurrent', 'EnvironmentalRemediationLiabilityNoncurrent', 'DerivativeLiabilitiesNoncurrent', 'UnrecognizedTaxBenefits'], 'cat': '2_Balance_Sheet'},
    'Payables to Brokers & Dealers': {'tags': ['PayablesToBrokerDealersAndClearingOrganizations', 'PayablesToBrokersDealersAndClearingOrganizations', 'BrokeragePayables', 'OtherPayablesToBrokerDealersAndClearingOrganizations', 'BrokerDealerPayableToOtherBrokerDealerAndClearingOrganizationFairValueDisclosure', 'CommissionsPayableToBrokerDealersAndClearingOrganizations'], 'cat': '2_Balance_Sheet'},
    'Customer Payables': {'tags': ['CustomerPayables', 'PayablesToCustomers', 'CustomerAccountsPayable', 'BrokerDealerPayableToCustomerFairValueDisclosure', 'BusinessCombinationRecognizedIdentifiableAssetsAcquiredAndLiabilitiesAssumedCurrentLiabilitiesPayableToUsers'], 'cat': '2_Balance_Sheet'},
    'Deposits': {'tags': ['Deposits', 'InterestBearingDeposits', 'NoninterestBearingDeposits', 'TimeDeposits'], 'cat': '2_Balance_Sheet'},
    'Securities Loaned': {'tags': ['SecuritiesLoaned', 'SecuritiesLoanedGross', 'SecuritiesLoanedAsset'], 'cat': '2_Balance_Sheet'},
    'Securities Sold under Agreements to Repurchase': {'tags': ['SecuritiesSoldUnderAgreementsToRepurchase'], 'cat': '2_Balance_Sheet'},
    'Financial Instruments Sold, Not Yet Purchased': {'tags': ['FinancialInstrumentsSoldNotYetPurchasedAtFairValue', 'TradingLiabilities'], 'cat': '2_Balance_Sheet'},
    'Total Liabilities': {'tags': ['Liabilities'], 'cat': '2_Balance_Sheet'},
    'Common Stock': {'tags': ['CommonStocksIncludingAdditionalPaidInCapital', 'CommonStockValue'], 'cat': '2_Balance_Sheet'},
    'Additional Paid-In Capital': {'tags': ['AdditionalPaidInCapital', 'AdditionalPaidInCapitalCommonStock'], 'cat': '2_Balance_Sheet'},
    'Retained Earnings': {'tags': ['RetainedEarningsAccumulatedDeficit', 'RetainedEarnings', 'RetainedEarningsAppropriated', 'RetainedEarningsUnappropriated'], 'cat': '2_Balance_Sheet'},
    'Comprehensive Income': {'tags': ['AccumulatedOtherComprehensiveIncomeLossNetOfTax'], 'cat': '2_Balance_Sheet'},
    'Total Equity': {'tags': ['StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', 'StockholdersEquity', 'CommonStockholdersEquity'], 'cat': '2_Balance_Sheet'},
    'Marketable Securities And Equities': {'tags': ['MarketableSecuritiesAndEquitySecuritiesFVNI', 'DebtSecuritiesAvailableForSaleAndEquitySecuritiesFVNI', 'AvailableForSaleDebtSecuritiesAmortizedCostBasis', 'AvailableForSaleSecuritiesDebtSecurities', 'MarketableSecurities', 'AvailableForSaleSecurities', 'EquitySecuritiesFvni', 'EquitySecurities', 'TradingSecurities', 'InvestmentsInDebtAndMarketableEquitySecurities'], 'cat': '2_Balance_Sheet'},


    # 3. CASH FLOW
    'Net Income (CF)': {'tags': ['NetIncomeLoss', 'ProfitLoss', 'NetIncomeLossAvailableToCommonStockholdersBasic', 'NetIncomeLossAttributableToParent', 'IncomeLossFromContinuingOperationsNetOfTax'], 'cat': '3_Cash_Flow'},
    'Depreciation & Amortization': {'tags': ['Depreciation', 'DepreciationAmortizationAndOther', 'DepreciationDepletionAndAmortization', 'DepreciationAndAmortization', 'Depletion', 'OtherDepreciationAndAmortization', 'DepreciationAndAmortizationOfFinanceLeaseRightOfUseAssets', 'DepreciationAndAmortizationOfPropertyPlantAndEquipment', 'AmortizationOfComputerSoftwareCosts', 'AmortizationOfDeferredCharges', 'OperatingLeaseRightOfUseAssetAmortization', 'DepreciationAndImpairmentOnDispositionOfPropertyAndEquipment', 'AmortizationOfIntangibleAssets', 'AmortizationOfAcquisitionCosts', 'AcquiredInPlaceLeasesAmortizationExpense'], 'cat': '3_Cash_Flow'},
    'Amortization of Intangibles (CF)': {'tags': ['AmortizationOfIntangibleAssets', 'AmortizationAndImpairmentOfIntangibleAssets'], 'cat': '3_Cash_Flow'},
    'Stock-Based Compensation': {'tags': ['ShareBasedCompensation', 'AllocatedShareBasedCompensationExpense', 'ShareBasedCompensationExpense', 'NoncashOrPartNoncashExpenseShareBasedCompensation'], 'cat': '3_Cash_Flow'},
    'Deferred Income Taxes': {'tags': ['DeferredIncomeTaxExpenseBenefit', 'DeferredIncomeTaxesAndTaxCredits', 'DeferredIncomeTaxAssetsNet'], 'cat': '3_Cash_Flow'},
    'Gain/Loss on Investments (CF)': {'tags': ['GainLossOnInvestments', 'GainLossOnSaleOfInvestments', 'UnrealizedGainLossOnInvestments', 'EquitySecuritiesFvNiGainLoss', 'DebtSecuritiesGainLoss', 'RealizedInvestmentGainsLosses', 'MarketableSecuritiesUnrealizedGainLoss', 'DebtAndEquitySecuritiesUnrealizedGainLoss', 'UnrealizedGainLossOnMarketableAndNonmarketableEquityInvestments', 'GainLossOnSaleOfDerivatives', 'EquitySecuritiesFvNiUnrealizedGainLoss'], 'cat': '3_Cash_Flow'},
    'Other Non-Cash Items': {'tags': ['OtherNoncashIncomeExpense', 'OtherNoncashIncome'], 'cat': '3_Cash_Flow'},
    'Change in AR': {'tags': ['IncreaseDecreaseInAccountsReceivable', 'IncreaseDecreaseInAccountsAndOtherReceivables', 'IncreaseDecreaseInAccountsReceivableAndOtherOperatingAssets', 'IncreaseDecreaseInReceivables'], 'cat': '3_Cash_Flow'},
    'Change in Inventory': {'tags': ['IncreaseDecreaseInInventories', 'IncreaseDecreaseInInventoriesAndOtherOperatingAssets'], 'cat': '3_Cash_Flow'},
    'Change in AP': {'tags': ['IncreaseDecreaseInAccountsPayable', 'IncreaseDecreaseInAccountsPayableAndAccruedLiabilities', 'IncreaseDecreaseInAccountsPayableTrade', 'IncreaseDecreaseInPayablesToCustomers', 'IncreaseDecreaseInPayablesToCustomersSCF', 'IncreaseDecreaseInOtherAccountsPayable'], 'cat': '3_Cash_Flow'},
    'Change in Deferred Revenue': {'tags': ['IncreaseDecreaseInDeferredRevenue', 'IncreaseDecreaseInContractWithCustomerLiability', 'IncreaseDecreaseInContractLiability'], 'cat': '3_Cash_Flow'},
    'Change in Accrued Expenses': {'tags': ['IncreaseDecreaseInAccruedLiabilities', 'IncreaseDecreaseInEmployeeRelatedLiabilities', 'IncreaseDecreaseInAccruedLiabilitiesAndOtherOperatingLiabilities'], 'cat': '3_Cash_Flow'},
    'Change in Income Taxes': {'tags': ['IncreaseDecreaseInIncomeTaxesPayable', 'IncreaseDecreaseInIncomeTaxes', 'IncreaseDecreaseInAccruedIncomeTaxesPayable'], 'cat': '3_Cash_Flow'},
    'Change in Prepaid & Other Assets': {'tags': ['IncreaseDecreaseInPrepaidDeferredExpenseAndOtherAssets', 'IncreaseDecreaseInOtherOperatingAssets', 'IncreaseDecreaseInOtherCurrentAssets', 'IncreaseDecreaseInOtherNoncurrentAssets'], 'cat': '3_Cash_Flow'},
    'Change in Other Liabilities': {'tags': ['IncreaseDecreaseInOtherOperatingLiabilities', 'IncreaseDecreaseInOtherCurrentLiabilities', 'IncreaseDecreaseInOtherNoncurrentLiabilities'], 'cat': '3_Cash_Flow'},
    'Operating Cash Flow': {'tags': ['NetCashProvidedByUsedInOperatingActivities', 'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations', 'NetCashProvidedByUsedInOperatingActivitiesDiscontinuedOperations'], 'cat': '3_Cash_Flow'},
    'Capital Expenditures': {'tags': ['PaymentsToAcquireProductiveAssets', 'PaymentsToAcquirePropertyPlantAndEquipment', 'PaymentsForCapitalImprovements', 'PurchasesOfPropertyAndEquipment', 'PropertyPlantAndEquipmentAdditions', 'AcquisitionsOfPropertyPlantAndEquipment', 'PaymentsToAcquirePropertyAndEquipment', 'PurchaseOfPropertyPlantAndEquipment', 'PaymentsToAcquirePropertyPlantAndEquipmentAndIntangibleAssets', 'PurchasesOfPropertyAndEquipmentAndIntangibleAssets'], 'cat': '3_Cash_Flow'},
    'Capital Expenditures (Software)': {'tags': ['PaymentsToAcquireSoftware', 'PaymentsToDevelopSoftware'], 'cat': '3_Cash_Flow'},
    'Capital Expenditures (Intangibles)': {'tags': ['PaymentsToAcquireIntangibleAssets'], 'cat': '3_Cash_Flow'},
    'Capital Expenditures (Equipment & Buildings)': {'tags': ['PaymentsToAcquireOtherPropertyPlantAndEquipment', 'PaymentsToAcquireMachineryAndEquipment', 'PaymentsToAcquireBuildings', 'PaymentsToAcquireLandAndBuildingsAndImprovements', 'PaymentsForConstructionInProcess', 'PaymentsToDevelopRealEstate', 'CapitalExpendituresIncurredButNotYetPaid'], 'cat': '3_Cash_Flow'},
    'Acquisitions': {'tags': ['PaymentsToAcquireBusinessesNetOfCashAcquired', 'PaymentsToAcquireBusinessesGross', 'PaymentsToAcquireBusinessesAndIntangibles'], 'cat': '3_Cash_Flow'},
    'Divestitures': {'tags': ['ProceedsFromDivestitureOfBusinesses', 'ProceedsFromSalesOfBusinessesNetOfCashDivested', 'ProceedsFromDivestitureOfBusinessesNetOfCashDivested', 'ProceedsFromSaleOfBusiness', 'ProceedsFromDivestitures', 'ProceedsFromSaleOfSubsidiaries', 'ProceedsFromSaleOfBusinessesAndInterestsInAffiliates', 'ProceedsFromDivestitureOfBusinessesAndInterestsInAffiliates', 'ProceedsFromSaleOfBusinessSegment'], 'cat': '3_Cash_Flow'},
    'Purchases of Investments': {'tags': ['PaymentsToAcquireMarketableSecurities', 'PaymentsToAcquireAvailableForSaleSecurities', 'PaymentsToAcquireShortTermInvestments', 'PaymentsToAcquireOtherInvestments', 'PaymentsToAcquireInvestments', 'PaymentsToAcquireLongTermInvestments', 'PaymentsToAcquireAvailableForSaleSecuritiesDebt'], 'cat': '3_Cash_Flow'},
    'Proceeds from Investments': {'tags': ['ProceedsFromSaleAndMaturityOfMarketableSecurities', 'ProceedsFromSaleAndMaturityOfAvailableForSaleSecurities', 'ProceedsFromSaleAndMaturityOfOtherInvestments', 'ProceedsFromMaturitiesPrepaymentsAndCallsOfAvailableForSaleSecurities', 'ProceedsFromSaleOfMarketableSecurities', 'ProceedsFromMaturitiesOfMarketableSecurities', 'ProceedsFromSaleOfInvestments', 'ProceedsFromSaleAndMaturityOfInvestments', 'ProceedsFromSaleAndMaturityOfAvailableForSaleSecuritiesDebt'], 'cat': '3_Cash_Flow'},
    'Proceeds from Asset Sales': {'tags': ['ProceedsFromSaleOfPropertyPlantAndEquipment', 'ProceedsFromSaleOfIntangibleAssets', 'ProceedsFromSaleOfOtherAssets', 'ProceedsFromSaleOfProductiveAssets', 'ProceedsFromSaleOfOperatingLeaseAssets', 'ProceedsFromSaleOfRealEstate', 'PaymentsForProceedsFromProductiveAssets', 'PaymentsForProceedsFromOtherInvestingActivities', 'ProceedsFromPropertyPlantAndEquipmentSalesAndIncentives', 'ProceedsFromRebatesOnPurchasesOfProductiveAssets'], 'cat': '3_Cash_Flow'},
    'Investing Cash Flow': {'tags': ['NetCashProvidedByUsedInInvestingActivities', 'NetCashProvidedByUsedInInvestingActivitiesContinuingOperations', 'NetCashProvidedByUsedInInvestingActivitiesDiscontinuedOperations'], 'cat': '3_Cash_Flow'},
    'Short-term Debt Issued': {'tags': ['ProceedsFromIssuanceOfShortTermDebt', 'ProceedsFromShortTermDebt', 'ProceedsFromShortTermDebtMaturingInThreeMonthsOrLess', 'ProceedsFromIssuanceOfCommercialPaper'], 'cat': '3_Cash_Flow'},
    'Short-term Debt Repaid': {'tags': ['RepaymentsOfShortTermDebt', 'PaymentsForRepaymentsOfShortTermDebt', 'RepaymentsOfCommercialPaper'], 'cat': '3_Cash_Flow'},
    'Net Change in Short-term Debt': {'tags': ['NetCashProvidedByUsedInShortTermDebt', 'IncreaseDecreaseInShortTermDebt', 'IncreaseDecreaseInLinesOfCredit', 'ProceedsFromRepaymentsOfCommercialPaper', 'ProceedsFromRepaymentsOfShortTermDebtMaturingInMoreThanThreeMonths', 'ProceedsFromRepaymentsOfShortTermDebt'], 'cat': '3_Cash_Flow'},
    'Net Short-Term Debt Issued (Repaid)': {'tags': [], 'cat': '3_Cash_Flow'},
    'Long-term Debt Issued': {'tags': ['ProceedsFromIssuanceOfLongTermDebt', 'ProceedsFromIssuanceOfLongTermDebtAndCapitalSecuritiesNet', 'ProceedsFromLongTermLinesOfCredit', 'ProceedsFromIssuanceOfSeniorLongTermDebt', 'ProceedsFromIssuanceOfMediumTermNotes', 'ProceedsFromIssuanceOfOtherLongTermDebt', 'ProceedsFromIssuanceOfUnsecuredDebt', 'ProceedsFromIssuanceOfSecuredDebt', 'ProceedsFromIssuanceOfSubordinatedDebt'], 'cat': '3_Cash_Flow'},
    'Long-term Debt Repaid': {'tags': ['RepaymentsOfLongTermDebt', 'RepaymentsOfLongTermDebtAndCapitalSecurities', 'RepaymentsOfLongTermLinesOfCredit', 'RepaymentsOfSeniorDebt', 'RepaymentsOfMediumTermNotes', 'RepaymentsOfOtherLongTermDebt', 'RepaymentsOfUnsecuredDebt', 'RepaymentsOfSecuredDebt', 'RepaymentsOfSubordinatedDebt', 'FinanceLeasePrincipalPayments'], 'cat': '3_Cash_Flow'},
    'Net Long-Term Debt Issued (Repaid)': {'tags': [], 'cat': '3_Cash_Flow'},
    'Total Debt Issued': {'tags': ['ProceedsFromIssuanceOfDebt', 'ProceedsFromDebtNetOfIssuanceCosts', 'ProceedsFromConvertibleDebt', 'ProceedsFromLinesOfCredit', 'ProceedsFromIssuanceOfUnsecuredDebt', 'ProceedsFromIssuanceOfSecuredDebt', 'ProceedsFromDebtMaturingInMoreThanThreeMonths', 'ProceedsFromNotesPayable', 'ProceedsFromIssuanceOfSubordinatedDebt'], 'cat': '3_Cash_Flow'}, 
    'Total Debt Repaid': {'tags': ['RepaymentsOfDebt', 'RepaymentsOfConvertibleDebt', 'RepaymentsOfLinesOfCredit', 'RepaymentsOfUnsecuredDebt', 'RepaymentsOfSecuredDebt', 'RepaymentsOfNotesPayable', 'RepaymentsOfOtherDebt', 'RepaymentsOfDebtMaturingInMoreThanThreeMonths', 'RepaymentsOfSubordinatedDebt', 'RepaymentsOfDebtAndCapitalLeaseObligations'], 'cat': '3_Cash_Flow'},
    'Total Net Debt Issued (Repaid)': {'tags': [], 'cat': '3_Cash_Flow'},
    'Cash Interest Paid': {'tags': ['InterestPaid', 'InterestPaidNet', 'InterestPaidNetOfCapitalizedInterest'], 'cat': '3_Cash_Flow'},
    'Share Repurchases': {'tags': ['PaymentsForRepurchaseOfCommonStock', 'StockRepurchasedDuringPeriodValue', 'StockRepurchasedAndRetiredDuringPeriodValue', 'TreasuryStockValueAcquiredCostMethod', 'TreasuryStockValueAcquiredParValueMethod', 'PaymentsForRepurchaseOfPreferredStock', 'PaymentsForRepurchaseOfEquity', 'StockRepurchasedAndRetiredDuringPeriodShares'], 'cat': '3_Cash_Flow'},
    'Shares Issued': {'tags': ['ProceedsFromIssuanceOfCommonStock', 'ProceedsFromIssuanceOfPreferredStock', 'ProceedsFromIssuanceOfCommonStockAndPreferredStock', 'StockIssuedDuringPeriodSharesNewIssues'], 'cat': '3_Cash_Flow'},
    'Shares Issued (Stock Plans)': {'tags': ['ProceedsFromIssuanceOfSharesUnderStockOptionAndStockPurchasePlans', 'ProceedsFromStockOptionsExercised', 'ProceedsFromIssuanceOfCommonStockUnderEmployeeStockPurchasePlan', 'StockIssuedDuringPeriodValueEmployeeStockPurchasePlan', 'ProceedsFromIssuanceOfSharesUnderIncentiveAndShareBasedCompensationPlansIncludingStockOptions', 'ProceedsFromStockPlans', 'StockIssuedDuringPeriodSharesRestrictedStockAwardsAndStockOptionsExercised'], 'cat': '3_Cash_Flow'},
    'Net Shares Issued (Repurchased)': {'tags': [], 'cat': '3_Cash_Flow'},
    'Dividends Paid': {'tags': ['PaymentsOfDividends', 'PaymentsOfDividendsCommonStock', 'PaymentsOfDividendsPreferredStock', 'DividendsCommonStockCash', 'DividendsCash'], 'cat': '3_Cash_Flow'},
    'Taxes Paid on Stock Awards': {'tags': ['PaymentsRelatedToTaxWithholdingForShareBasedCompensation', 'PaymentsOfTaxWithholdingForShareBasedCompensation'], 'cat': '3_Cash_Flow'},
    'Financing Cash Flow': {'tags': ['NetCashProvidedByUsedInFinancingActivities', 'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations', 'NetCashProvidedByUsedInFinancingActivitiesDiscontinuedOperations'], 'cat': '3_Cash_Flow'},
    'Net Cash Flow': {'tags': [], 'cat': '3_Cash_Flow'},
    'FX Effect on Cash': {'tags': ['EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents', 'EffectOfExchangeRateOnCashAndCashEquivalents', 'EffectOfExchangeRateOnCash'], 'cat': '3_Cash_Flow'},
    'Cash Reconciliation: Total': {'tags': ['CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents'], 'cat': '3_Cash_Flow'},
    
    # Remaining Performance Obligations
    'RPO - Totals': {'tags': ['RevenueRemainingPerformanceObligation'], 'cat': '6_Disclosures'},
    'RPO - Next 12 Months': {'tags': ['RevenueRemainingPerformanceObligationExpectedTimingOfSatisfactionPeriod1'], 'cat': '6_Disclosures'},
    'RPO - Next 12 Months (%)': {'tags': ['RevenueRemainingPerformanceObligationPercentage'], 'cat': '6_Disclosures'},

    # 7. CONCENTRATION RISK (Standard Tags)
    'Customer Concentration %': {'tags': ['ConcentrationRiskPercentage1', 'ConcentrationRiskRevenue', 'ConcentrationRiskPercentage', 'ConcentrationRiskNumberOfMajorCustomers', 'ConcentrationRiskPercentageOfAccountsReceivable', 'ConcentrationRiskAccountsReceivable', 'ConcentrationRiskSupplierPercentage', 'ConcentrationRiskGeographicRegion', 'RevenueFromExternalCustomersByReportableSegment', 'MajorCustomerMember'], 'cat': '7_Concentration_Risk'},
}

# ---------------------------------------------------------------------------
# Income-Statement Sort Infrastructure
# ---------------------------------------------------------------------------
# One canonical ordering that covers every known IS concept in P&L flow order.
# Insurance- and financial-specific items are included at the top so that a
# single list serves all company types.  _build_item_order() then applies
# type-specific overrides (e.g. moving Interest Income/Expense to non-operating
# for standard/insurance companies) and fills any remaining gaps via
# keyword-bucket fallback â€” so no label ever ends up at position 999.
# ---------------------------------------------------------------------------

# Positions 0-4   : primary revenue / top-line items
# Positions 5-9   : cost of revenue / claims / gross profit
# Positions 10-19 : operating expense detail lines
# Positions 20-21 : total opex / operating income subtotals
# Positions 22-26 : non-operating items
# Positions 27-31 : pre/post-tax and net income
# Positions 32+   : per-share metrics
_IS_CANONICAL_ORDER: list[str] = [
    # â”€â”€ Primary revenue / top-line â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Premiums Earned',                    # insurance
    'Net Investment Income',              # insurance / financial
    'Net Interest Income (Expense)',      # financial primary metric
    'Interest Income',                    # financial primary revenue
    'Revenue',                            # standard + insurance revenue total
    # â”€â”€ Cost of revenue / claims â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Policyholder Claims/Benefits',       # insurance COGS equivalent
    'Amortization of DAC',                # insurance acquisition-cost amortization
    'Cost of Revenue',                    # standard COGS
    'Gross Profit',                       # standard / insurance subtotal
    # â”€â”€ Operating expense detail â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Research & Development',
    'Selling, General & Admin',
    'Sales & Marketing',
    'General & Administrative',
    'Marketing Expense',
    'Amortization of Intangibles',        # â† was falling to 99; now explicit
    'Depreciation, Depletion & Amortization',  # oil & gas / extractive IS line
    'Depreciation & Amortization Expense',     # REIT / real estate IS line
    'Exploration Expense',                # oil & gas exploration & dry-hole costs
    'Restructuring & Related Charges',
    'Impairment Charges',
    'Acquisition-Related Costs',
    'Litigation & Settlement Charges',
    # â”€â”€ Operating subtotals â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Total Operating Expenses',
    'Total Costs and Expenses',
    'Operating Income',
    # â”€â”€ Non-operating â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Interest Expense',                   # non-op for standard/insurance
    'Gain/Loss on Investments',
    'Equity Method Income',
    'Other Income / (Expense)',
    'Total Non-operating Income',
    # â”€â”€ Pre/post-tax â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'Pretax Income',
    'Income Tax Expense',
    'Net Income to Noncontrolling Interest',
    'Income from Discontinued Operations',
    'Net Income',
    # â”€â”€ Per-share â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    'EPS Basic',
    'EPS Diluted',
    'Shares Outstanding Basic',
    'Shares Outstanding Diluted',
]

_IS_BASE_ORDER: dict[str, int] = {n: i for i, n in enumerate(_IS_CANONICAL_ORDER)}


def _keyword_is_pos(label: str) -> float:
    """
    Keyword-bucket fallback for IS labels not in _IS_CANONICAL_ORDER.
    Returns a structural income-statement slot instead of dumping unknown
    filer-specific rows at the bottom.  The buckets are intentionally generic:
    they use accounting nature words (revenue, cost, opex, non-op, tax, EPS),
    not company names or ticker-specific labels.
    """
    l = re.sub(r"\s+", " ", str(label or "").lower()).strip()

    # Preferred-stock deductions are below net income and above EPS/share rows.
    # Check this before the revenue block because redemption-premium labels
    # contain the word "premium" but are not top-line revenue.
    if any(k in l for k in ('preferred stock dividend', 'preferred dividends',
                            'redemption premium')):
        return 31.3

    # Cost labels often contain the word "revenue" ("Cost of Revenue",
    # "Cost of Lease & Other Revenue").  Classify them before the revenue
    # block so they do not get mistaken for top-line revenue details.
    if any(k in l for k in (
        'cost of', 'costs of', 'direct cost', 'direct costs', 'cost of goods',
        'cost of services', 'cost of products', 'leased and rented property',
    )):
        if l == 'cost of revenue':
            return 7.8
        if l.startswith(('cost of ', 'costs of ', 'direct cost', 'direct costs')):
            return 7.1
        if 'amortiz' in l:
            return 7.2
        if 'restructur' in l:
            return 7.3
        return 7.1

    # Explicit non-operating wording wins before generic benefits/costs/charges.
    if 'non-operat' in l or 'nonoperat' in l:
        return 24

    # Revenue / top-line zone.  Include common operating-income variants used
    # by insurers, REITs, brokers, and service/product filers.  Exact totals
    # keep the main Revenue row first; revenue detail rows receive a small
    # offset so they stay below the total even without a "Revenue - ..." name.
    if any(k in l for k in (
        'revenue', 'revenues', 'net sales', 'sales net', 'sales revenue',
        'premium', 'premiums earned', 'subscription', 'service income',
        'fee income', 'rental income', 'lease income', 'interest and fee income',
    )):
        if l in ('revenue', 'revenues', 'net sales'):
            return 4
        return 4.2

    # Employee-benefit compensation is an operating expense; policyholder or
    # healthcare benefits are the insurance COGS equivalent.
    if any(k in l for k in ('employee benefit', 'employee benefits', 'salaries',
                            'compensation and benefit')):
        return 13

    # Claims / policyholder benefits are the insurance COGS equivalent. Avoid
    # a bare employee-benefit match here.
    if (any(k in l for k in ('claim', 'policyholder', 'loss ratio', 'medical benefit',
                             'insurance benefit'))
            or ('benefit' in l and 'employee' not in l and 'salar' not in l
                and 'compensation' not in l)):
        return 6

    # Cost of revenue / direct operating costs.  Plural/direct-cost phrasing is
    # common in REIT and leasing filers.
    if any(k in l for k in (
        'cost of', 'costs of', 'direct cost', 'direct costs', 'cost of goods',
        'cost of services', 'cost of products', 'leased and rented property',
    )):
        if l == 'cost of revenue':
            return 7.8
        if l.startswith(('cost of ', 'costs of ', 'direct cost', 'direct costs')):
            return 7.1
        if 'amortiz' in l:
            return 7.2
        if 'restructur' in l:
            return 7.3
        return 7.1

    # Gross profit subtotal.
    if 'gross' in l and 'profit' in l:
        return 8

    # R&D / technology development.
    if any(k in l for k in ('research', 'development', 'r&d')):
        return 9

    # Sales & marketing / marketing-only operating expense.
    if any(k in l for k in ('sales and marketing', 'sales & marketing',
                            'sales marketing', 'marketing')):
        return 10

    # SG&A / selling / admin.
    if any(k in l for k in ('selling', 'general', 'admin', 'sg&a', 'sga')):
        return 11

    # Recurring operating-expense face lines used by banks, brokers, REITs and
    # other services companies.  These should live above Total Operating
    # Expenses, even when the filer uses a custom concept.
    if any(k in l for k in (
        'brokerage', 'clearance', 'clearing', 'execution', 'regulatory',
        'occupancy', 'operations and support', 'support expense',
        'compensation', 'employee benefit', 'employee benefits',
        'salaries', 'data processing', 'professional fee', 'professional fees',
        'advertising', 'bad debt', 'credit loss', 'credit losses',
        'operating expense', 'operating expenses', 'other operating',
    )):
        return 13

    # Amortization / depreciation lines.
    if any(k in l for k in ('amortiz', 'depreciat', 'depletion')):
        return 14

    # Operating special charges. Some filers present loss on held-for-sale
    # subsidiaries inside operating costs, not below operating income.
    if 'subsidiar' in l and 'held for sale' in l:
        return 16

    # Non-operating finance items must be checked before one-time "charges";
    # e.g. GE-style "Interest and other financial charges" is not a
    # restructuring charge.
    if any(k in l for k in (
        'interest expense', 'interest income', 'financial charge',
        'financial charges', 'finance cost', 'finance costs', 'gain/loss',
        'gain (loss)', 'loss on sale', 'gain on sale', 'loss from sale',
        'gain from sale', 'equity method', 'investment income', 'other income',
        'other expense', 'foreign currency', 'derivative', 'non-operat',
        'nonoperat',
    )):
        return 24

    # Restructuring, impairment, litigation and other one-time operating charges.
    # Avoid a bare "charge" keyword: it misclassifies finance charges.
    if any(k in l for k in (
        'restructur', 'impairment', 'write-off', 'write off',
        'acquisition-related', 'litigation', 'settlement', 'severance',
    )):
        return 16

    # Total operating-expense and COGS-inclusive total-cost subtotals.
    if 'total cost' in l or 'costs and expenses' in l:
        return 20.5
    if 'total operating' in l:
        return 20

    # Operating income / loss subtotal.
    if any(k in l for k in ('operating income', 'operating profit', 'operating loss')):
        return 21

    # Pretax.
    if any(k in l for k in ('pretax', 'pre-tax', 'income before tax', 'before income tax')):
        return 27

    # Tax.
    if 'tax' in l:
        return 28

    # Noncontrolling / minority interest.
    if any(k in l for k in ('noncontrolling', 'minority interest')):
        return 29

    # Net income / earnings.
    if any(k in l for k in ('net income', 'net loss', 'net earnings')):
        return 30

    if 'discontinued operation' in l:
        return 31.2

    # EPS support disclosures come after the main EPS/share-count rows.
    if 'antidilutive securities' in l:
        return 33.2

    # Per-share/share-count rows.
    if any(k in l for k in (
        'eps', 'earnings per share', 'per share', 'diluted share',
        'basic share', 'shares outstanding', 'weighted average share',
    )):
        return 32

    # Unknown IS item â€” keep it inside the income statement but below the main
    # P&L flow, before category-level catch-alls.
    return 95

# ---------------------------------------------------------------------------
# Layered income-statement row-sort evidence
# ---------------------------------------------------------------------------
# The final sorter deliberately follows the most objective evidence available:
#   1) company presentation anchors for auto-learned face rows,
#   2) calculation-linkbase roll-up parents,
#   3) curated CONCEPT_MAP / canonical row buckets,
#   4) generic semantic label fallback.
# This keeps new/custom company labels out of the bottom catch-all without
# adding ticker-specific rules or touching any numeric values.

def _is_totalish_label(label: str) -> bool:
    l = re.sub(r"\s+", " ", str(label or "").lower()).strip()
    return (
        l.startswith('total ')
        or l in ('operating expenses', 'costs and expenses', 'total costs and expenses')
        or 'total operating' in l
        or 'total cost' in l
    )


def _is_main_revenue_label(label: str) -> bool:
    l = re.sub(r"\s+", " ", str(label or "").lower()).strip()
    return l in ('revenue', 'revenues', 'net sales')


def _local_concept_name(concept: str) -> str:
    c = str(concept or '')
    if ':' in c:
        c = c.split(':')[-1]
    if '_' in c:
        c = c.split('_')[-1]
    return c


def _semantic_is_pos(label: str) -> float:
    """Semantic fallback only; kept separate so calc/presentation can win."""
    return float(_keyword_is_pos(label))


def _calc_parent_to_is_pos(parent: str, eff_weight: float, label: str):
    """Map a calculation ancestor to a structural P&L sort bucket."""
    p = _local_concept_name(parent)
    l = re.sub(r"\s+", " ", str(label or "").lower()).strip()

    if p in _REVENUE_ROLLUP_PARENTS:
        return 4.0 if _is_main_revenue_label(label) else 4.2

    if p in ('CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfGoodsSold', 'CostOfServices'):
        # Cost detail rows should appear before the aggregate Cost of Revenue
        # subtotal when both are present, but all still stay before Gross Profit.
        return 7.8 if l == 'cost of revenue' else 7.1

    if p == 'GrossProfit':
        if 'gross' in l and 'profit' in l:
            return 8.0
        # In calc links, revenue children usually contribute + and cost lines -.
        # Keep the aggregate Cost of Revenue below cost detail rows, but before GP.
        if eff_weight < 0:
            return 7.8 if l == 'cost of revenue' else 7.1
        return 4.2

    if p in _OPEX_ROLLUP_PARENTS:
        return 20.0 if _is_totalish_label(label) else 13.0

    if p in ('OperatingIncomeLoss', 'OperatingIncomeLossFromContinuingOperations',
             'RealEstateOperatingIncome'):
        if any(k in l for k in ('operating income', 'operating profit', 'operating loss')):
            return 21.0
        # Expense children reduce operating income; income children add to it.
        return 13.0 if eff_weight < 0 else 18.0

    if p in ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest',
             'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'):
        if any(k in l for k in ('pretax', 'pre-tax', 'income before tax', 'before income tax')):
            return 27.0
        return 24.0

    if p in ('IncomeTaxExpenseBenefit', 'IncomeTaxExpenseBenefitContinuingOperations'):
        return 28.0

    if p in ('NetIncomeLoss', 'ProfitLoss'):
        if any(k in l for k in ('net income', 'net loss', 'net earnings')):
            return 30.0
        if 'tax' in l:
            return 28.0
        if any(k in l for k in ('noncontrolling', 'minority interest')):
            return 29.0
        return 24.0

    return None


def _calc_lineage_is_pos_for_concept(concept: str, label: str, max_hops: int = 5):
    """Infer P&L bucket by walking from a concept to known calc ancestors."""
    start = _local_concept_name(concept)
    if not start:
        return None

    direct = _calc_parent_to_is_pos(start, 1.0, label)
    if direct is not None:
        return direct

    frontier = {(start, 1.0)}
    seen = set()
    best = None
    for _ in range(max_hops):
        nxt = set()
        for c, w in frontier:
            if c in seen:
                continue
            seen.add(c)
            for p, pw in GLOBAL_CALC_PARENT.get(c, ()):
                try:
                    eff = w * (float(pw) if pw is not None else 1.0)
                except Exception:
                    eff = w
                pos = _calc_parent_to_is_pos(p, eff, label)
                if pos is not None:
                    # Prefer the closest evidence; if multiple parents exist at
                    # the same distance, choose the earlier structural bucket.
                    best = pos if best is None else min(best, pos)
                nxt.add((_local_concept_name(p), eff))
        if best is not None:
            return float(best)
        frontier = nxt
        if not frontier:
            break
    return None


def _calc_lineage_is_pos(label: str):
    """Return a calculation-linkbase P&L bucket for a CONCEPT_MAP label."""
    info = CONCEPT_MAP.get(label)
    if not isinstance(info, dict):
        return None
    tags = info.get('tags') or []
    best = None
    for tag in tags:
        pos = _calc_lineage_is_pos_for_concept(tag, label)
        if pos is None:
            continue
        best = pos if best is None else min(best, pos)
    return None if best is None else float(best)


def _evidence_is_pos(label: str, is_financial: bool = False,
                     is_insurance: bool = False, allow_weak_semantic: bool = True) -> float:
    """Best available generic bucket for an income-statement label.

    This function never uses ticker/company names. It is only about accounting
    structure.  Known canonical rows keep their canonical/type-specific order;
    custom/auto/unknown rows get calc evidence first and semantic fallback last.
    """
    label = str(label or '')
    base = label.split(' - ')[0]

    # Financial core interest lines are operating/top-line in broker/bank-style
    # statements.  Standard/insurance filers keep interest below operating income.
    if is_financial:
        fin_core = {
            'Interest Income': 0.0,
            'Interest Expense': 1.0,
            'Net Interest Income (Expense)': 2.0,
        }
        if base in fin_core:
            return fin_core[base]

    # Calculation linkbase is stronger than text semantics for custom rows.
    calc_pos = _calc_lineage_is_pos(base)
    if calc_pos is not None:
        # In standard filers, calc links can show interest expense under a broad
        # cost/expense parent; the canonical mapping is safer for known interest.
        if not (not is_financial and base in ('Interest Expense', 'Interest Income',
                                             'Net Interest Income (Expense)')):
            return float(calc_pos)

    semantic = _semantic_is_pos(base)
    if not allow_weak_semantic and semantic >= 90:
        return 95.0
    return float(semantic)


def _is_sort_section(pos: float) -> int:
    """Coarse section number used to detect bad presentation anchors."""
    try:
        pos = float(pos)
    except Exception:
        return 99
    if pos < 5:
        return 0      # revenue/top-line
    if pos < 9:
        return 1      # cost/gross profit
    if pos < 20:
        return 2      # operating expense detail
    if pos < 22:
        return 3      # operating subtotal
    if pos < 27:
        return 4      # non-operating
    if pos < 32:
        return 5      # tax/net income
    if pos < 34:
        return 6      # EPS/shares
    return 7


def _presentation_anchor_pos(label: str, item_order: dict,
                             is_financial: bool = False,
                             is_insurance: bool = False):
    """Company-presentation placement for an auto-learned label, if reliable."""
    info = CONCEPT_MAP.get(label)
    if not (isinstance(info, dict) and info.get('auto')):
        return None
    seq = info.get('anchor_seq', 0)
    anchor = info.get('anchor')
    if anchor is not None and anchor in item_order:
        anchored = float(item_order[anchor]) + 0.001 * (seq + 1)
        evidence = _evidence_is_pos(label, is_financial, is_insurance,
                                    allow_weak_semantic=False)
        # If the anchor would throw an obvious P&L line across sections,
        # let calc/semantic evidence rescue it.  Be stricter around the top of
        # the statement: a revenue detail anchored after a cost row (or a cost
        # detail anchored in revenue/non-op territory) is visibly wrong even
        # though the coarse section distance is only one bucket.
        if evidence < 90:
            if evidence < 5 and anchored >= 5:
                return evidence + 0.001 * (seq + 1)
            if 5 <= evidence < 9 and (anchored < 5 or anchored >= 9):
                return evidence + 0.001 * (seq + 1)
            if 9 <= evidence < 20 and (anchored < 5 or anchored >= 20):
                return evidence + 0.001 * (seq + 1)
            if abs(_is_sort_section(anchored) - _is_sort_section(evidence)) >= 2:
                return evidence + 0.001 * (seq + 1)
        if info.get('is_dyn_opex') and 'Gross Profit' in item_order:
            anchored = max(anchored, float(item_order['Gross Profit']) + 0.001 * (seq + 1))
        return anchored
    if anchor is None and 'anchor' in info:
        evidence = _evidence_is_pos(label, is_financial, is_insurance,
                                    allow_weak_semantic=False)
        if evidence < 90:
            return evidence + 0.001 * (seq + 1)
        return -1 + 0.001 * (seq + 1)
    return None


def _build_item_order(is_insurance: bool, is_financial: bool) -> dict[str, float]:
    """
    Build a complete item_order dict for sorting income-statement rows.

    This is the single generic IS sorter used by the final output.  It avoids
    ticker-specific row fixes by assigning every known/learned IS label to a
    structural P&L bucket.  Company presentation anchors can still nudge
    auto-learned rows within those buckets later.
    """
    if is_financial:
        # Broker/bank-style statements treat interest income/expense and net
        # interest income as core operations, so keep them at the top instead
        # of below Operating Income.
        fin_order = [
            'Interest Income',
            'Interest Expense',
            'Net Interest Income (Expense)',
            'Revenue',
            'Gross Profit',          # anchor for dynamic opex; often not filed
            'Research & Development',
            'Sales & Marketing',
            'Marketing Expense',
            'Selling, General & Admin',
            'General & Administrative',
            'Salaries & Employee Benefits',
            'Customer Bad Debt',
            'Other Operating Expenses',
            'Amortization of Intangibles',
            'Depreciation, Depletion & Amortization',
            'Depreciation & Amortization Expense',
            'Restructuring & Related Charges',
            'Impairment Charges',
            'Acquisition-Related Costs',
            'Litigation & Settlement Charges',
            'Total Operating Expenses',
            'Total Costs and Expenses',
            'Operating Income',
            'Gain/Loss on Investments',
            'Equity Method Income',
            'Other Income / (Expense)',
            'Total Non-operating Income',
            'Pretax Income',
            'Income Tax Expense',
            'Net Income to Noncontrolling Interest',
            'Income from Discontinued Operations',
            'Net Income',
            'EPS Basic',
            'EPS Diluted',
            'Shares Outstanding Basic',
            'Shares Outstanding Diluted',
        ]
        order: dict[str, float] = {name: float(pos) for pos, name in enumerate(fin_order)}
    else:
        order = {name: float(pos) for name, pos in _IS_BASE_ORDER.items()}
        # For standard and insurance P&Ls, interest lines are non-operating,
        # but still below Operating Income.  The previous gain/loss-minus
        # arithmetic could place Interest Expense above Operating Income when
        # canonical positions were compacted.
        non_op_anchor = order['Operating Income'] + 1
        order['Interest Expense']              = non_op_anchor
        order['Interest Income']               = non_op_anchor + 0.1
        order['Net Interest Income (Expense)'] = non_op_anchor + 0.2
        # When a filer reports cost-detail rows plus an aggregate Cost of
        # Revenue line, place details first, then the total, then Gross Profit.
        order['Cost of Revenue'] = 7.8
        order['Gross Profit'] = 8.0
        order['Total Costs and Expenses'] = 20.5

    # Keep attribution/discontinued rows below the main Net Income subtotal
    # and above EPS/share rows across all statement types.
    if 'Net Income' in order:
        order['Income Tax Expense'] = 30.0
        order['Net Income'] = 31.0
        order['Net Income to Noncontrolling Interest'] = 31.1
        order['Income from Discontinued Operations'] = 31.2
        order['EPS Basic'] = 32.0
        order['EPS Diluted'] = 32.1
        order['Shares Outstanding Basic'] = 33.0
        order['Shares Outstanding Diluted'] = 33.1

    # Fill every remaining CONCEPT_MAP IS item that is not explicitly covered.
    # This covers both curated rows and auto-learned face-statement rows.
    for name, info in CONCEPT_MAP.items():
        if isinstance(info, dict) and info.get('cat') == '1_Income_Statement':
            if name not in order:
                order[name] = float(_evidence_is_pos(name, is_financial, is_insurance))

    return order



# ---------------------------------------------------------------------------
# Balance-sheet and cash-flow final row-sort evidence
# ---------------------------------------------------------------------------
# The income statement already has a layered evidence sorter.  These helpers
# extend the same approach to the other two face statements:
#   1) company XBRL presentation positions captured from the filing,
#   2) calculation-linkbase roll-up parents,
#   3) curated CONCEPT_MAP / canonical buckets,
#   4) generic semantic label fallback.
# The final sort guard still verifies that this changes row order only, never
# values.  When final-pivot cache is loaded without live XBRL state, this falls
# back deterministically to calc/semantic buckets.

def _norm_sort_label(label: str) -> str:
    return re.sub(r"\s+", " ", str(label or "").replace('_', ' ')).strip().lower()


def _semantic_bs_pos(label: str) -> float:
    """Generic structural balance-sheet bucket for known or custom labels."""
    l = _norm_sort_label(label)

    # Totals/subtotals first because they contain broad words like assets/liabs.
    if 'total current assets' in l:
        return 19.0
    if l in ('total assets',) or ('total assets' in l and 'liabilities' not in l):
        return 29.0
    if 'total current liabilities' in l:
        return 39.0
    if l == 'total liabilities' or ('total liabilities' in l and 'equity' not in l):
        return 49.0
    if 'total equity' in l or 'stockholders equity' in l or "shareholders' equity" in l:
        return 59.0

    # Current / liquid assets.
    if any(k in l for k in ('cash', 'cash equivalent')):
        if any(k in l for k in ('restricted', 'segregated', 'clearing organization', 'regulatory')):
            return 1.5
        return 0.0
    if 'segregated' in l or ('clearing organization' in l and 'deposit' in l):
        return 1.6
    if any(k in l for k in ('marketable debt', 'debt securities', 'held-to-maturity', 'available-for-sale')) and 'liabil' not in l:
        return 2.0
    if any(k in l for k in ('marketable equity', 'equity securities', 'marketable securities', 'short-term investment', 'short term investment', 'trading securities')) and 'sold' not in l and 'liabil' not in l:
        return 2.3
    if 'accrued' in l and not any(k in l for k in ('receivable', 'asset')):
        return 31.0
    if any(k in l for k in ('accounts receivable', 'receivable', 'contract with customer', 'unbilled', 'revenue share')) and not any(k in l for k in ('payable', 'liabil')):
        if any(k in l for k in ('broker', 'dealer', 'customer receivable', 'margin')):
            return 13.5
        return 4.0
    if any(k in l for k in ('prepaid', 'capitalized contract cost')) and 'noncurrent' not in l and 'non-current' not in l:
        return 5.0
    if 'inventory' in l:
        if 'raw material' in l:
            return 6.1
        if 'work in process' in l:
            return 6.2
        if 'finished goods' in l:
            return 6.3
        return 6.0
    if 'other current assets' in l:
        return 8.0

    # Non-current assets.
    if any(k in l for k in ('real estate', 'land', 'building improvement', 'commercial real estate')) and not any(k in l for k in ('payable', 'liabil', 'debt')):
        return 20.0
    if any(k in l for k in ('property, plant', 'property plant', 'equipment', 'rou asset', 'right-of-use asset')) and not any(k in l for k in ('liabil', 'payable', 'debt')):
        return 21.0
    if any(k in l for k in ('equity in net assets', 'investment in subsidiaries', 'net assets of subsidiaries')):
        return 98.0
    if any(k in l for k in ('long-term investment', 'long term investment', 'nonmarketable', 'joint venture', 'affiliates')) and not any(k in l for k in ('liabil', 'payable')):
        return 22.0
    if 'deferred income tax assets' in l or 'deferred tax assets' in l:
        return 23.0
    if any(k in l for k in ('other non-current assets', 'other noncurrent assets', 'prepaid expense, noncurrent', 'disposal group')) and 'liabilit' not in l:
        return 23.5
    if any(k in l for k in ('intangible', 'finite-lived intangible')) and 'goodwill' not in l and 'liabil' not in l:
        return 24.0
    if 'goodwill' in l:
        return 24.5
    if any(k in l for k in ('loans', 'leases', 'securities borrowed', 'purchased under agreements to resell', 'financial instruments owned', 'user-held fractional shares')) and not any(k in l for k in ('liabil', 'obligation', 'sold', 'repurchase obligation')):
        return 25.0

    # Current liabilities.
    if any(k in l for k in ('accounts payable', 'payable', 'accrued', 'taxes payable', 'unrecognized tax benefits', 'claims', 'claim adjustment', 'intercompany payable', 'sales discounts and allowances')):
        if any(k in l for k in ('dividends payable', 'customer payables', 'broker', 'dealer')):
            return 34.0
        return 31.0
    if any(k in l for k in ('deferred revenue', 'contract liability')):
        return 32.0
    if any(k in l for k in ('short-term debt', 'short term debt', 'short-term borrowings', 'short term borrowings', 'current portion', 'current lease liability', 'operating lease liability (current)', 'finance lease liability (current)')):
        return 35.0

    # Non-current liabilities.
    if any(k in l for k in ('long-term debt', 'long term debt', 'senior notes', 'convertible debt', 'notes payable', 'borrowings', 'revolving credit', 'commercial paper')):
        return 40.0
    if any(k in l for k in ('non-current liabilities', 'noncurrent liabilities', 'deferred compensation liability', 'deferred tax liabilities', 'insurance liabilities', 'annuity benefits', 'below market lease')):
        return 42.0
    if any(k in l for k in ('operating lease liability (non-current)', 'operating lease liability (noncurrent)', 'finance lease liability (non-current)', 'finance lease liability (noncurrent)')):
        return 42.5
    if any(k in l for k in ('deposits', 'securities loaned', 'sold under agreements to repurchase', 'financial instruments sold', 'repurchase obligation', 'disposal group')) and 'asset' not in l:
        return 43.0

    # Mezzanine/equity.
    if any(k in l for k in ('temporary equity', 'redeemable noncontrolling', 'preferred stock', 'nonredeemable or redeemable')):
        return 50.0
    if any(k in l for k in ('common stock', 'additional paid-in', 'paid in capital', 'retained earnings', 'accumulated distributions', 'comprehensive income', 'noncontrolling interest', 'attributable to noncontrolling interest')):
        return 52.0

    return 89.0


def _semantic_cf_pos(label: str) -> float:
    """Generic structural cash-flow bucket for known or custom labels.

    The cash-flow sorter is XBRL-first, but presentation order is allowed to
    win only inside sensible accounting sub-blocks.  These positions therefore
    intentionally distinguish debt financing from equity financing: otherwise a
    company's presentation order can place a stock-issuance row between debt
    issuance and debt repayment, which is technically XBRL-derived but visually
    confusing in a normalized statement.
    """
    l = _norm_sort_label(label)

    # Supplemental cash disclosures belong after the statement cash-flow bridge.
    if any(k in l for k in ('income taxes paid', 'cash taxes paid', 'cash paid', 'interest paid')):
        return 96.0
    if 'cash reconciliation' in l:
        return 94.0
    if 'fx effect' in l or 'exchange rate' in l:
        return 93.0
    if l == 'net cash flow' or 'net cash flow' in l or 'increase in cash' in l or 'cash within businesses held for sale' in l:
        return 92.0
    if any(k in l for k in ('noncash financing', 'non-cash financing', 'noncash investing', 'non-cash investing',
                            'temporary equity', 'accretion to redemption value',
                            'redeemable convertible preferred stock',
                            'accrued taxes related to net share settlement')):
        return 97.0

    if l == 'net income (cf)' or l.startswith('net income'):
        return 0.0

    # Operating adjustments and working-capital changes.
    if any(k in l for k in ('depreciation', 'amortization', 'stock-based compensation', 'share-based compensation', 'deferred income tax', 'deferred tax', 'gain/loss', 'gain (loss)', 'pension', 'noncash', 'non-cash', 'debt issuance costs', 'extinguishment of debt')):
        return 4.0
    if any(k in l for k in ('adjustment to reconcile', 'other non-cash', 'other noncash')):
        return 8.0

    # Some financing rows start with "change in" / "increase (decrease)";
    # classify obvious debt/equity financing before the generic working-capital
    # change test below.  Keep debt and equity in separate financing sub-blocks
    # so presentation order can sort within each group but cannot interleave them.
    # Debt financing sub-blocks.  Check net labels first because labels such
    # as "Net Short-Term Debt Issued (Repaid)" contain both "issued" and
    # "repaid".  Normalized display order should be:
    #   short-term issued -> short-term repaid -> short-term net
    #   long-term issued  -> long-term repaid  -> long-term net
    #   total issued      -> total repaid      -> total net
    if ('net change in short-term debt' in l or 'net change in short term debt' in l):
        # Raw SEC net-change source row.  If it exactly duplicates the normalized
        # Net Short-Term Debt row, a later cleanup pass preserves it in
        # disclosures.  If it is kept in cash flow, show it after gross
        # issued/repaid but before the normalized net row.
        return 60.6
    if ('net short-term debt' in l or 'net short term debt' in l):
        return 60.8
    if any(k in l for k in ('short-term debt', 'short term debt', 'commercial paper')):
        if any(k in l for k in ('proceeds', 'issued', 'issuance', 'borrowings')):
            return 60.0
        if any(k in l for k in ('repay', 'repayment', 'repaid', 'payment')):
            return 60.4
        return 60.8
    if ('net long-term debt' in l or 'net long term debt' in l
            or 'net change in long-term debt' in l or 'net change in long term debt' in l):
        return 61.8
    if 'total net debt' in l or ('net debt' in l and any(k in l for k in ('issued', 'issuance', 'repaid', 'repay', 'repayment'))):
        return 63.0
    if any(k in l for k in ('long-term debt', 'long term debt', 'revolving credit')):
        if any(k in l for k in ('proceeds', 'issued', 'issuance', 'borrowings')):
            return 61.0
        if any(k in l for k in ('repay', 'repayment', 'repaid', 'payment')):
            return 61.4
        return 61.8
    if any(k in l for k in ('total debt issued', 'debt issued', 'proceeds from debt', 'proceeds from issuance of debt', 'borrowings', 'revolving credit proceeds')):
        return 62.2
    if any(k in l for k in ('total debt repaid', 'debt repaid', 'repayments and other debt', 'repayments of debt', 'repayment of debt', 'revolving credit repayment')):
        return 62.6
    if any(k in l for k in ('total debt', 'net debt')):
        return 63.0
    if 'net shares issued' in l or 'net shares repurchased' in l or 'net share' in l:
        return 72.5
    if any(k in l for k in ('shares issued', 'stock plans', 'issuance of common stock', 'initial public offering', 'proceeds from stock', 'stock option', 'employee stock purchase')):
        return 70.0
    if any(k in l for k in ('share repurchase', 'repurchase of common stock', 'repurchases of common stock', 'unsettled repurchases')):
        return 71.0
    if any(k in l for k in ('taxes paid on stock awards', 'tax withholding', 'stock based award', 'stock-based award')):
        return 72.0
    if any(k in l for k in ('noncontrolling interests', 'redeemable noncontrolling', 'customer funds', 'hedge, financing')):
        return 74.0

    if any(k in l for k in ('change in', 'increase (decrease)', 'decrease (increase)', 'progress collections', 'sales discounts', 'health care insurance liabilities', 'securities loaned transactions', 'prepaid', 'accounts receivable', 'inventory', 'accounts payable', 'deferred revenue', 'accrued expenses', 'income taxes', 'other liabilities')):
        return 12.0
    if 'other operating adjustments' in l:
        return 18.0
    if 'operating cash flow' in l or 'operating activities' in l:
        return 19.0

    # Investing activities.
    if any(k in l for k in ('capital expenditure', 'property and equipment', 'commercial real estate', 'software')):
        return 30.0
    if any(k in l for k in ('acquisition', 'merger', 'contingent consideration', 'asset acquisition', 'cash acquired', 'payments to acquire', 'payment to acquire', 'acquire interest in subsidiaries')):
        return 34.0
    if any(k in l for k in ('purchases of investments', 'purchase of investment', 'originations', 'loans to', 'credit card receivables', 'held-for-sale assets', 'loans')):
        return 36.0
    if any(k in l for k in ('proceeds from investments', 'repayments and maturities', 'collection of notes', 'venture distribution', 'distribution')):
        return 40.0
    if any(k in l for k in ('divestiture', 'divestitures', 'sale of asset', 'asset sales', 'insurance settlement')):
        return 43.0
    if 'other investing adjustments' in l:
        return 48.0
    if 'investing cash flow' in l or 'investing activities' in l:
        return 49.0

    # Financing activities.
    # Debt financing sub-blocks.  Check net labels first because labels such
    # as "Net Short-Term Debt Issued (Repaid)" contain both "issued" and
    # "repaid".  Normalized display order should be:
    #   short-term issued -> short-term repaid -> short-term net
    #   long-term issued  -> long-term repaid  -> long-term net
    #   total issued      -> total repaid      -> total net
    if ('net change in short-term debt' in l or 'net change in short term debt' in l):
        # Raw SEC net-change source row.  If it exactly duplicates the normalized
        # Net Short-Term Debt row, a later cleanup pass preserves it in
        # disclosures.  If it is kept in cash flow, show it after gross
        # issued/repaid but before the normalized net row.
        return 60.6
    if ('net short-term debt' in l or 'net short term debt' in l):
        return 60.8
    if any(k in l for k in ('short-term debt', 'short term debt', 'commercial paper')):
        if any(k in l for k in ('proceeds', 'issued', 'issuance', 'borrowings')):
            return 60.0
        if any(k in l for k in ('repay', 'repayment', 'repaid', 'payment')):
            return 60.4
        return 60.8
    if ('net long-term debt' in l or 'net long term debt' in l
            or 'net change in long-term debt' in l or 'net change in long term debt' in l):
        return 61.8
    if 'total net debt' in l or ('net debt' in l and any(k in l for k in ('issued', 'issuance', 'repaid', 'repay', 'repayment'))):
        return 63.0
    if any(k in l for k in ('long-term debt', 'long term debt', 'revolving credit')):
        if any(k in l for k in ('proceeds', 'issued', 'issuance', 'borrowings')):
            return 61.0
        if any(k in l for k in ('repay', 'repayment', 'repaid', 'payment')):
            return 61.4
        return 61.8
    if any(k in l for k in ('total debt issued', 'debt issued', 'proceeds from debt', 'proceeds from issuance of debt', 'borrowings', 'revolving credit proceeds')):
        return 62.2
    if any(k in l for k in ('total debt repaid', 'debt repaid', 'repayments and other debt', 'repayments of debt', 'repayment of debt', 'revolving credit repayment')):
        return 62.6
    if any(k in l for k in ('total debt', 'net debt')):
        return 63.0
    if 'net shares issued' in l or 'net shares repurchased' in l or 'net share' in l:
        return 72.5
    if any(k in l for k in ('shares issued', 'stock plans', 'issuance of common stock', 'initial public offering', 'proceeds from stock', 'stock option', 'employee stock purchase')):
        return 70.0
    if any(k in l for k in ('share repurchase', 'repurchase of common stock', 'repurchases of common stock', 'unsettled repurchases')):
        return 71.0
    if any(k in l for k in ('taxes paid on stock awards', 'tax withholding', 'stock based award', 'stock-based award')):
        return 72.0
    if any(k in l for k in ('dividends paid', 'distribution made', 'limited partner')):
        return 73.0
    if any(k in l for k in ('noncontrolling interests', 'redeemable noncontrolling', 'customer funds', 'hedge, financing')):
        return 74.0
    if 'other financing adjustments' in l:
        return 78.0
    if 'financing cash flow' in l or 'financing activities' in l:
        return 79.0

    return 89.0


def _semantic_statement_pos(cat: str, label: str) -> float:
    if cat == '2_Balance_Sheet':
        return _semantic_bs_pos(label)
    if cat == '3_Cash_Flow':
        return _semantic_cf_pos(label)
    return 999.0


def _stmt_section(cat: str, pos: float) -> int:
    """Coarse sections for BS/CF so XBRL order cannot cross major subtotals."""
    try:
        pos = float(pos)
    except Exception:
        return 99
    if cat == '2_Balance_Sheet':
        if pos < 19:
            return 0     # current/liquid assets
        if pos < 29:
            return 1     # non-current assets
        if pos < 30:
            return 2     # total assets
        if pos < 39:
            return 3     # current liabilities
        if pos < 49:
            return 4     # non-current liabilities
        if pos < 50:
            return 5     # total liabilities
        if pos < 60:
            return 6     # equity / mezzanine
        return 7
    if cat == '3_Cash_Flow':
        # Finer than the three classic CF sections on purpose.  XBRL
        # presentation order is still used, but only within these accounting
        # sub-blocks.  This keeps debt financing rows together and equity
        # financing rows together even when a filer's presentation tree
        # interleaves them.
        if pos < 9:
            return 0     # net income and non-cash adjustments
        if pos < 19:
            return 1     # working-capital changes / other operating bridge
        if pos < 20:
            return 2     # operating cash flow subtotal
        if pos < 31:
            return 3     # capex / PP&E investing
        if pos < 35:
            return 4     # acquisitions / business combinations
        if pos < 39:
            return 5     # investment purchases / loan originations
        if pos < 44:
            return 6     # investment proceeds / asset sales
        if pos < 49:
            return 7     # other investing details
        if pos < 50:
            return 8     # investing cash flow subtotal
        if pos < 61:
            return 9     # short-term debt issued/repaid/net
        if pos < 62:
            return 10    # long-term debt issued/repaid/net
        if pos < 63:
            return 11    # total debt issued/repaid
        if pos < 66:
            return 12    # total net debt and other debt activity
        if pos < 71:
            return 13    # equity issuance / stock plans
        if pos < 72:
            return 14    # share repurchases
        if pos < 73:
            return 15    # stock-award tax withholding / net equity activity
        if pos < 75:
            return 16    # dividends / NCI / customer-fund financing
        if pos < 79:
            return 17    # other financing details
        if pos < 80:
            return 18    # financing cash flow subtotal
        if pos < 95:
            return 19    # net cash / FX / reconciliation
        return 20        # supplemental cash disclosures
    return 99


def _calc_parent_to_statement_pos(cat: str, parent: str, eff_weight: float, label: str):
    """Map a calculation ancestor to a structural BS/CF sort bucket."""
    p = _local_concept_name(parent)
    l = _norm_sort_label(label)
    sem = _semantic_statement_pos(cat, label)

    if cat == '2_Balance_Sheet':
        if p == 'AssetsCurrent':
            return 19.0 if 'total current assets' in l else (sem if sem < 19 else 8.0)
        if p == 'Assets':
            if 'total assets' in l and 'liabilities' not in l:
                return 29.0
            return sem if sem < 29 else 20.0
        if p == 'LiabilitiesCurrent':
            return 39.0 if 'total current liabilities' in l else (sem if 30 <= sem < 39 else 31.0)
        if p == 'Liabilities':
            if 'total liabilities' in l and 'equity' not in l:
                return 49.0
            return sem if 30 <= sem < 49 else 40.0
        if p in ('CommitmentsAndContingencies',):
            return 48.5
        if p in ('TemporaryEquityCarryingAmount', 'RedeemableNoncontrollingInterestEquityCarryingAmount'):
            return 50.0
        if p in ('StockholdersEquity', 'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest',
                 'PartnersCapital', 'CommonStocksIncludingAdditionalPaidInCapital'):
            return 59.0 if any(k in l for k in ('total equity', 'stockholders equity', 'shareholders equity')) else 52.0
        if p == 'LiabilitiesAndStockholdersEquity':
            if any(k in l for k in ('equity', 'stockholders', 'shareholders', 'retained earnings', 'paid-in capital')):
                return 52.0
            if any(k in l for k in ('liabil', 'payable', 'debt', 'borrowings')):
                return sem if 30 <= sem < 50 else 40.0
            if 'asset' in l:
                return sem if sem < 30 else 29.0
            return sem

    if cat == '3_Cash_Flow':
        if p in _CF_OPERATING_PARENTS:
            if 'operating cash flow' in l or 'operating activities' in l:
                return 19.0
            return sem if sem < 19 else 12.0
        if p in _CF_INVESTING_PARENTS:
            if 'investing cash flow' in l or 'investing activities' in l:
                return 49.0
            return sem if 30 <= sem < 49 else 40.0
        if p in _CF_FINANCING_PARENTS:
            if 'financing cash flow' in l or 'financing activities' in l:
                return 79.0
            return sem if 60 <= sem < 79 else 70.0
        if p in ('NetIncreaseDecreaseInCashAndCashEquivalents',
                 'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect'):
            return 92.0
        if p in ('EffectOfExchangeRateOnCashAndCashEquivalents',
                 'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents'):
            return 93.0
        if p in ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents',
                 'CashAndCashEquivalentsAtCarryingValue'):
            return 94.0

    return None


def _calc_lineage_statement_pos_for_concept(cat: str, concept: str, label: str, max_hops: int = 5):
    """Infer BS/CF bucket by walking from a concept to known calc ancestors."""
    start = _local_concept_name(concept)
    if not start:
        return None

    direct = _calc_parent_to_statement_pos(cat, start, 1.0, label)
    if direct is not None:
        return float(direct)

    frontier = {(start, 1.0)}
    seen = set()
    best = None
    for _ in range(max_hops):
        nxt = set()
        for c, w in frontier:
            if c in seen:
                continue
            seen.add(c)
            for p, pw in GLOBAL_CALC_PARENT.get(c, ()):  # child -> [(parent, weight)]
                try:
                    eff = w * (float(pw) if pw is not None else 1.0)
                except Exception:
                    eff = w
                pos = _calc_parent_to_statement_pos(cat, p, eff, label)
                if pos is not None:
                    best = pos if best is None else min(best, pos)
                nxt.add((_local_concept_name(p), eff))
        if best is not None:
            return float(best)
        frontier = nxt
        if not frontier:
            break
    return None


def _calc_lineage_statement_pos(cat: str, label: str):
    info = CONCEPT_MAP.get(label)
    if not isinstance(info, dict):
        return None
    if info.get('cat') != cat:
        return None
    best = None
    for tag in (info.get('tags') or []):
        pos = _calc_lineage_statement_pos_for_concept(cat, tag, label)
        if pos is None:
            continue
        best = pos if best is None else min(best, pos)
    return None if best is None else float(best)


def _presentation_statement_pos(cat: str, label: str):
    """Return the median XBRL presentation ordinal for this label on `cat`."""
    info = CONCEPT_MAP.get(str(label).split(' - ')[0])
    if not isinstance(info, dict) or info.get('cat') != cat:
        return None
    vals = []
    for tag in (info.get('tags') or []):
        cmap = _FACE_PRESENTATION_POS.get(_local_concept_name(tag))
        if not isinstance(cmap, dict):
            continue
        for value in cmap.get(cat, ()):
            try:
                vals.append(float(value))
            except Exception:
                pass
    if not vals:
        return None
    try:
        return float(np.median(vals))
    except Exception:
        return float(sorted(vals)[len(vals) // 2])


def _evidence_statement_pos(cat: str, label: str, _seen=None) -> float:
    """Calc/canonical/semantic bucket, with auto anchor rescue for vague labels."""
    base = str(label).split(' - ')[0]
    calc_pos = _calc_lineage_statement_pos(cat, base)
    if calc_pos is not None:
        return float(calc_pos)

    sem = float(_semantic_statement_pos(cat, base))
    info = CONCEPT_MAP.get(base)
    if sem >= 89 and isinstance(info, dict) and info.get('auto') and info.get('anchor'):
        anchor = info.get('anchor')
        if anchor and anchor != base:
            _seen = set(_seen or ())
            if anchor not in _seen:
                _seen.add(base)
                anchor_pos = _evidence_statement_pos(cat, anchor, _seen)
                if anchor_pos < 89:
                    return float(anchor_pos) + 0.25 + 0.001 * float(info.get('anchor_seq', 0) or 0)
    return sem


def _layered_statement_order_pos(cat: str, label: str, seq: int = 0) -> float:
    """Final XBRL-aware sort coordinate for BS/CF labels.

    The sorter is XBRL-first *inside* sensible accounting buckets, not across
    subtotals.  The structural evidence bucket remains ahead of presentation
    order, so totals such as Total Equity stay after APIC/AOCI/retained
    earnings even when a filer presentation tree lists the subtotal early.
    """
    evidence = float(_evidence_statement_pos(cat, label))
    section = _stmt_section(cat, evidence)
    pres = _presentation_statement_pos(cat, label)
    # Evidence controls the accounting sub-block; XBRL presentation controls
    # fine order within that sub-block.  This keeps BS/CF presentation-aware
    # while preventing subtotal/component inversions.
    if pres is not None:
        return float(section * 1000.0 + evidence * 10.0 + min(max(pres, 0.0), 999.0) * 0.001 + seq * 0.000001)
    return float(section * 1000.0 + evidence * 10.0 + seq * 0.001)


def _build_statement_order(cat: str) -> dict[str, float]:
    """Build a category-specific XBRL-first order map for BS/CF."""
    if cat not in ('2_Balance_Sheet', '3_Cash_Flow'):
        return {name: float(i) for i, name in enumerate(CONCEPT_MAP.keys())}

    order: dict[str, float] = {}
    same_cat_names = [
        name for name, info in CONCEPT_MAP.items()
        if isinstance(info, dict) and info.get('cat') == cat
    ]
    for i, name in enumerate(same_cat_names):
        order[name] = _layered_statement_order_pos(cat, name, i)
    return order


def _statement_order_for_label(cat: str, label: str, order_maps: dict[str, dict[str, float]]) -> float:
    base = str(label).split(' - ')[0]
    if cat in ('2_Balance_Sheet', '3_Cash_Flow'):
        order = order_maps.setdefault(cat, _build_statement_order(cat))
        if base not in order:
            order[base] = _layered_statement_order_pos(cat, base, len(order))
        return order[base]
    order = order_maps.setdefault(cat, {name: float(i) for i, name in enumerate(CONCEPT_MAP.keys())})
    return order.get(base, 999.0)


# ---------------------------------------------------------------------------
# Business-segment final row sort
# ---------------------------------------------------------------------------
# Segment rows are not face-statement rows, so they do not always have a clean
# concept-only presentation coordinate.  The final business-segment sorter uses
# accounting structure first (metric type), then keeps genuine reportable
# segment members ahead of accounting adjustments / legal-entity / tax /
# receivable spillovers.  It is deterministic and row-order-only; the final
# sort guard below still proves the values are untouched.

_SEGMENT_MEMBER_NOISE_TERMS = (
    'gain loss', 'gain/loss', 'asset impairment', 'impairment charge',
    'reclass', 'reclassified', 'reclassification', 'acquisition',
    'deferred income tax', 'noncash tax', 'non-cash tax', 'income tax',
    'accounts receivable', 'receivable sale', 'payments to acquire',
    'variable interest entity', 'primary beneficiary', 'svb', 'barclays',
    'wf arrangement', 'common class', 'available to common stockholders',
    'common stockholders diluted', 'chief executive officer',
    'investment agreement', 'unallocated', 'elimination', 'corporate',
    'other', 'all other', 'restructuring', 'severance', 'legal entity',
)

_SEGMENT_MEMBER_ACCOUNTING_TERMS = (
    'accounts receivable', 'accounts payable', 'income tax', 'deferred tax',
    'stockholders', 'common class', 'variable interest entity', 'primary beneficiary',
    'asset impairment', 'gain loss', 'reclass', 'unallocated expenses',
)


def _split_segment_display_label(label: str):
    parts = [p.strip() for p in str(label or '').split(' - ')]
    metric = parts[0] if parts else ''
    member = ' - '.join(p for p in parts[1:] if p)
    return metric, member


def _semantic_segment_metric_pos(metric: str) -> float:
    """Financial-statement-like order for segment metric prefixes."""
    l = _norm_sort_label(metric)

    if any(k in l for k in ('customer concentration', 'concentration')):
        return 80.0
    if 'customer revenue' in l:
        return 0.8

    # Cost before generic revenue because labels can contain both words.
    if any(k in l for k in ('cost of revenue', 'cost of sales', 'cost of goods', 'cost of service')):
        return 10.0
    if any(k in l for k in ('revenue', 'revenues', 'sales')):
        return 0.0
    # Company-defined monetary drivers recovered from source HTML are related
    # to the revenue discussion but are not necessarily revenue themselves.
    # Keep their explicitly qualified block beside revenue without conflating
    # the two metric definitions.
    if l == 'operating measure':
        return 1.0
    if 'gross profit' in l or 'gross margin' in l:
        return 20.0
    if any(k in l for k in ('operating expenses', 'operating expense', 'total operating expenses')):
        return 30.0
    if any(k in l for k in ('operating income', 'operating loss', 'income from operations', 'operating profit')):
        return 40.0
    if any(k in l for k in ('pretax income', 'pre-tax income', 'income before tax')):
        return 50.0
    if any(k in l for k in ('net income', 'net loss', 'net earnings')):
        return 60.0
    if l == 'income' or l.startswith('income '):
        return 55.0

    if 'capital expenditure' in l or 'capital expenditures' in l:
        return 70.0
    if 'depreciation' in l or 'amortization' in l:
        return 72.0
    if 'assets' in l or l == 'assets':
        return 74.0
    if 'liabilities' in l or l == 'liabilities':
        return 76.0
    if 'inventory' in l:
        return 78.0
    if 'accounts receivable' in l:
        return 82.0
    if 'accounts payable' in l:
        return 84.0
    if 'property, plant' in l or 'property plant' in l or 'equipment' in l:
        return 86.0
    if 'deferred revenue' in l:
        return 88.0
    return 95.0


def _segment_member_noise_rank(member: str) -> int:
    """0=real reportable member, higher=adjustment/accounting spillover."""
    l = _norm_sort_label(member)
    if not l:
        return 0
    # Combined segment labels are still useful segment aggregates, but should
    # follow the individual reportable segments.
    if ' and ' in l and not any(k in l for k in _SEGMENT_MEMBER_ACCOUNTING_TERMS):
        return 1
    if 'acquisition' in l:
        return 2
    if any(k in l for k in ('unallocated', 'elimination', 'corporate', 'all other')):
        return 3
    if l in ('other', 'service other') or l.endswith(' other'):
        return 4
    if any(k in l for k in _SEGMENT_MEMBER_NOISE_TERMS):
        return 5
    return 0


def _segment_member_sort_text(text: str) -> str:
    text = re.sub(r"\s+", " ", str(text or '')).strip()
    return text.casefold()


def _business_segment_core_members(labels) -> list[str]:
    """Infer recurring reportable business members from the displayed rows."""
    seen_order = []
    metric_hits = {}
    for label in labels or []:
        metric, member = _split_segment_display_label(label)
        if not member:
            continue
        if _segment_member_noise_rank(member) != 0:
            continue
        if ' - ' in member:
            # Fine-grained member such as "Advertising - Family Of Apps"; it
            # should group under a core segment if one exists, but should not
            # define the core set by itself.
            continue
        mpos = _semantic_segment_metric_pos(metric)
        if mpos >= 65:
            continue
        key = _segment_member_sort_text(member)
        if key not in metric_hits:
            seen_order.append(member)
            metric_hits[key] = set()
        metric_hits[key].add(_norm_sort_label(metric))

    recurring = [m for m in seen_order if len(metric_hits.get(_segment_member_sort_text(m), ())) >= 2]
    if recurring:
        return recurring

    # Some companies disclose only revenue by product/business.  In that case,
    # fall back to the non-noisy revenue members.
    revenue_only = []
    for label in labels or []:
        metric, member = _split_segment_display_label(label)
        if not member or _segment_member_noise_rank(member) != 0:
            continue
        if _semantic_segment_metric_pos(metric) <= 1.0 and member not in revenue_only:
            revenue_only.append(member)
    return revenue_only


def _business_segment_context(labels) -> dict:
    core = _business_segment_core_members(labels)
    core_sorted = sorted(core, key=_segment_member_sort_text)
    return {
        'core_members': core_sorted,
        'core_order': {_segment_member_sort_text(m): i for i, m in enumerate(core_sorted)},
    }


def _segment_member_group(member: str, context: dict):
    """Return (core_order, detail_level, detail_text) for member grouping."""
    member_text = str(member or '').strip()
    m_norm = _segment_member_sort_text(member_text)
    core_order = (context or {}).get('core_order', {})
    core_members = (context or {}).get('core_members', [])

    if m_norm in core_order:
        return core_order[m_norm], 0, ''

    best = None
    for core in core_members:
        c_norm = _segment_member_sort_text(core)
        if not c_norm:
            continue
        if m_norm.startswith(c_norm + ' - ') or m_norm.endswith(' - ' + c_norm) or c_norm in m_norm:
            cand = (core_order.get(c_norm, 999), 1, m_norm.replace(c_norm, '').strip(' -'))
            if best is None or cand < best:
                best = cand
    if best is not None:
        return best
    return 999, 0, m_norm


def _segment_business_sort_key(label: str, context: dict | None = None):
    metric, member = _split_segment_display_label(label)
    metric_pos = _semantic_segment_metric_pos(metric)
    noise_rank = _segment_member_noise_rank(member)
    group_ord, detail_level, detail_text = _segment_member_group(member, context or {})

    # Unknown/non-core rows should not jump ahead of obvious reportable segment
    # rows solely because their member text sorts alphabetically.
    if group_ord == 999 and noise_rank == 0 and member:
        group_ord = 900

    return (
        metric_pos,
        group_ord,
        detail_level,
        noise_rank,
        detail_text,
        _segment_member_sort_text(member),
        str(label).casefold(),
    )


def _audit_business_segment_sort(final_pivot, max_warnings: int = 8):
    if final_pivot is None or final_pivot.empty:
        return final_pivot
    try:
        labels = [idx[1] for idx in final_pivot.index if idx[0] == '4a_Segments_Business']
    except Exception:
        return final_pivot
    if not labels:
        return final_pivot
    metric_positions = [_semantic_segment_metric_pos(_split_segment_display_label(l)[0]) for l in labels]
    bad = []
    best_seen = -1.0
    for label, pos in zip(labels, metric_positions):
        if pos + 1e-9 < best_seen and pos < 90:
            bad.append(label)
        best_seen = max(best_seen, pos)
    if bad:
        shown = ', '.join(dict.fromkeys(map(str, bad[:max_warnings])))
        extra = '' if len(bad) <= max_warnings else f" (+{len(bad)-max_warnings} more)"
        print(f"  [Sort Audit] Possible business-segment row(s) outside expected metric order: {shown}{extra}")
    return final_pivot


def _audit_balance_sheet_sort(final_pivot, max_warnings: int = 8):
    if final_pivot is None or final_pivot.empty:
        return final_pivot
    try:
        labels = [idx[1] for idx in final_pivot.index if idx[0] == '2_Balance_Sheet']
    except Exception:
        return final_pivot
    positions = {label: i for i, label in enumerate(labels)}
    suspicious = []
    for label in labels:
        pos = _evidence_statement_pos('2_Balance_Sheet', label)
        if pos < 29 and 'Total Assets' in positions and positions[label] > positions['Total Assets']:
            suspicious.append(label)
        elif 30 <= pos < 49 and 'Total Assets' in positions and positions[label] < positions['Total Assets']:
            suspicious.append(label)
        elif 30 <= pos < 49 and 'Total Liabilities' in positions and positions[label] > positions['Total Liabilities']:
            suspicious.append(label)
        elif pos >= 50 and 'Total Liabilities' in positions and positions[label] < positions['Total Liabilities']:
            suspicious.append(label)
    if suspicious:
        shown = ', '.join(dict.fromkeys(suspicious[:max_warnings]))
        extra = '' if len(suspicious) <= max_warnings else f" (+{len(suspicious)-max_warnings} more)"
        print(f"  [Sort Audit] Possible balance-sheet row(s) outside expected section: {shown}{extra}")
    return final_pivot


def _audit_cash_flow_sort(final_pivot, max_warnings: int = 8):
    if final_pivot is None or final_pivot.empty:
        return final_pivot
    try:
        labels = [idx[1] for idx in final_pivot.index if idx[0] == '3_Cash_Flow']
    except Exception:
        return final_pivot
    positions = {label: i for i, label in enumerate(labels)}
    suspicious = []
    for label in labels:
        pos = _evidence_statement_pos('3_Cash_Flow', label)
        if pos < 19 and 'Operating Cash Flow' in positions and positions[label] > positions['Operating Cash Flow']:
            suspicious.append(label)
        elif 30 <= pos < 49 and 'Investing Cash Flow' in positions and positions[label] > positions['Investing Cash Flow']:
            suspicious.append(label)
        elif 60 <= pos < 79 and 'Financing Cash Flow' in positions and positions[label] > positions['Financing Cash Flow']:
            suspicious.append(label)
        elif pos >= 92 and 'Net Cash Flow' in positions and positions[label] < positions['Net Cash Flow']:
            suspicious.append(label)
    if suspicious:
        shown = ', '.join(dict.fromkeys(suspicious[:max_warnings]))
        extra = '' if len(suspicious) <= max_warnings else f" (+{len(suspicious)-max_warnings} more)"
        print(f"  [Sort Audit] Possible cash-flow row(s) outside expected section: {shown}{extra}")
    return final_pivot


def _audit_face_statement_sort(final_pivot, is_item_order=None, is_financial=False, is_insurance=False):
    final_pivot = _audit_income_statement_sort(final_pivot, is_item_order, is_financial, is_insurance)
    final_pivot = _audit_balance_sheet_sort(final_pivot)
    final_pivot = _audit_cash_flow_sort(final_pivot)
    final_pivot = _audit_business_segment_sort(final_pivot)
    return final_pivot

# ---------------------------------------------------------------------------
# Smart Concept Resolution System
# Two-tier: Exact Lookup  ->  Max-Pairwise IDF-weighted Jaccard
#
# Problem this solves
# -------------------
# The scope-mismatch guard compares annual vs. quarterly XBRL concept strings.
# Many pairs are *semantically identical* but use different taxonomy tags:
#   Annual   = 'PurchasesOfPropertyAndEquipmentAndIntangibleAssets'
#   Quarterly = 'PaymentsToAcquirePropertyPlantAndEquipment'
# A raw string comparison flags these as a mismatch and drops Q4.
#
# Two-tier design
# ---------------
# Tier 1 â€” Exact lookup (O(1)):
#   Reverse map of every tag in CONCEPT_MAP -> its label.
#   Handles all registered aliases instantly.
#
# Tier 2 â€” Max pairwise IDF-weighted Jaccard (O(labels Ã— tags_per_label)):
#   For any concept NOT in the exact map (company extension tags, obscure
#   US-GAAP variants, post-cutoff taxonomy additions), split the CamelCase
#   name into tokens, normalize synonyms, then find the MOST SIMILAR
#   individually registered tag across all labels.  The score is an
#   IDF-weighted Jaccard coefficient between the input token set and the
#   candidate tag's token set.
#
#   Why max-pairwise (not pool-union Dice):
#     CapEx has 17 registered tags; Amortization of Intangibles has 3.
#     Pool-union Dice penalises CapEx because its pool denominator is much
#     larger, causing false positives toward smaller labels.  Max-pairwise
#     always compares the input to its nearest registered sibling, so label
#     richness does not bias the score.
#
#   Results are cached â€” each unique concept string is resolved only once.
#
# Public interface
# ----------------
#   resolve_concept_to_label(concept: str | None) -> str | None
#   Always call this instead of _CONCEPT_TAG_TO_LABEL directly.
# ---------------------------------------------------------------------------
import math as _math
from collections import Counter as _Counter

# -- Tier 1: exact reverse lookup -------------------------------------------
_CONCEPT_TAG_TO_LABEL: dict[str, str] = {
    tag: label
    for label, info in CONCEPT_MAP.items()
    if isinstance(info, dict) and 'tags' in info
    for tag in info['tags']
}

# -- Tier 2: fuzzy token infrastructure -------------------------------------

_CAMEL_RE = re.compile(r'(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])')

# Generic words that appear in almost every concept â€” no discriminating signal.
_XBRL_STOP = frozenset({
    'of', 'to', 'and', 'or', 'in', 'by', 'for', 'from', 'the', 'a', 'an',
    'with', 'on', 'at', 'into', 'upon', 'per', 'over', 'under', 'during',
    'us',   # leaks from 'us-gaap' namespace prefixes
})

# Synonym normalization for XBRL financial tokens.
# Maps raw lowercase token -> canonical form.
# Normalises action verbs that vary across taxonomy versions; leaves nouns
# (which carry real discriminating signal) mostly intact.
_XBRL_SYNONYMS: dict[str, str] = {
    # Cash outflow action verbs â€” all mean "we paid cash for something"
    'payments':      'payments',
    'payment':       'payments',
    'purchases':     'payments',
    'purchase':      'payments',
    'expenditures':  'payments',
    'expenditure':   'payments',
    'additions':     'payments',   # "Additions to PP&E" is CapEx phrasing
    'addition':      'payments',
    'acquisition':   'payments',   # "Acquisition of PP&E" â‰  business acquisition
    'acquisitions':  'payments',   # (business-acquisition tags contain 'Businesses')
    'spending':      'payments',
    'outlays':       'payments',
    # Cash inflow action verbs
    'proceeds':      'proceeds',
    'receipts':      'proceeds',
    'collections':   'proceeds',
    # Debt issuance verbs
    'issuance':      'issuance',
    'borrowings':    'issuance',
    # PP&E noun synonyms
    'machinery':     'equipment',
    'machines':      'equipment',
    # Intangibles
    'intangibles':   'intangible',
    # Investments
    'investments':   'investment',
    'investment':    'investment',
    'investing':     'investment',
    # Revenue
    'revenues':      'revenue',
    'sales':         'revenue',
    # Repayments
    'repayments':    'repayments',
    'reimbursements':'repayments',
    # Securities
    'securities':    'securities',
    # Leases
    'leases':        'lease',
    'leased':        'lease',
}

@lru_cache(maxsize=65536, typed=True)
def _tokenize_xbrl_concept(concept: str) -> frozenset[str]:
    """
    Split a CamelCase XBRL concept string into a normalized token set.

    Pipeline
    --------
    1. Typo Normalization: 'PlantandEquipment' -> 'PlantAndEquipment'
    2. CamelCase split : 'PaymentsToAcquirePropertyPlantAndEquipment'
                         -> ['Payments','To','Acquire','Property','Plant','Equipment']
    3. Lowercase       : ['payments','to','acquire','property','plant','equipment']
    4. Stop removal    : ['payments','acquire','property','plant','equipment']
    5. Synonym norm    : ['payments','acquire','property','plant','equipment']
    6. Return frozenset: {'payments','acquire','property','plant','equipment'}
    """
    # Fix common missing capitalization typos (e.g., 'and', 'of', 'after')
    # that glue words together and break the CamelCase splitter.
    concept = re.sub(
        r'(?<=[a-z])(and|of|to|for|the|in|after|before|net|gross)(?=[A-Z])',
        lambda m: m.group(1).capitalize(),
        concept
    )
    raw = [t.lower() for t in _CAMEL_RE.sub(r' ', concept).split()]
    filtered = [t for t in raw if t not in _XBRL_STOP and len(t) > 1]
    normalized = [_XBRL_SYNONYMS.get(t, t) for t in filtered]
    return frozenset(normalized)

# Build per-label list of individually tokenized tags.
# Kept separate (NOT merged into a pool) so scoring is max-pairwise:
# each input concept is compared to its nearest registered sibling,
# not against a diluted union that grows with the number of tags.
_LABEL_TAG_TOKEN_LISTS: dict[str, list[frozenset[str]]] = {}
for _lbl, _lbl_info in CONCEPT_MAP.items():
    if not isinstance(_lbl_info, dict) or 'tags' not in _lbl_info:
        continue
    _LABEL_TAG_TOKEN_LISTS[_lbl] = [
        _tokenize_xbrl_concept(_tag) for _tag in _lbl_info['tags']
    ]

# Token -> candidate labels index for fuzzy matching.  The resolver still
# iterates _LABEL_TAG_TOKEN_LISTS in insertion order, preserving tie behavior;
# this index only lets it skip labels that share no token and therefore have
# an exact Jaccard score of 0.
_TOKEN_TO_LABELS: dict[str, set[str]] = {}
for _lbl, _tag_token_list in _LABEL_TAG_TOKEN_LISTS.items():
    for _tset in _tag_token_list:
        for _tok in _tset:
            _TOKEN_TO_LABELS.setdefault(_tok, set()).add(_lbl)

# IDF: document frequency â€” how many labels contain each token.
# Computed over the union pool per label (for IDF purposes only;
# scoring still uses individual tag comparisons, not the pool).
_TOKEN_DF: _Counter = _Counter()
for _tag_list in _LABEL_TAG_TOKEN_LISTS.values():
    _lbl_union: set[str] = set()
    for _tset in _tag_list:
        _lbl_union.update(_tset)
    for _tok in _lbl_union:
        _TOKEN_DF[_tok] += 1
_N_LABELS_FOR_IDF = max(len(_LABEL_TAG_TOKEN_LISTS), 1)

@lru_cache(maxsize=4096, typed=True)
def _idf(token: str) -> float:
    """IDF weight: log(N / df + 1). Rare tokens -> high weight."""
    return _math.log(_N_LABELS_FOR_IDF / max(_TOKEN_DF.get(token, 1), 1) + 1.0)

@lru_cache(maxsize=131072, typed=True)
def _idf_jaccard(set_a: frozenset[str], set_b: frozenset[str]) -> float:
    """
    IDF-weighted Jaccard similarity between two token sets.
    Score = Î£idf(intersection) / Î£idf(union).
    Range [0, 1]. Returns 0.0 for empty inputs or disjoint sets.
    """
    if not set_a or not set_b:
        return 0.0
    inter = set_a & set_b
    if not inter:
        return 0.0
    union = set_a | set_b
    inter_score = sum(_idf(t) for t in inter)
    union_score = sum(_idf(t) for t in union)
    return inter_score / union_score if union_score > 0 else 0.0

# Jaccard threshold for Tier 2 to return a match.
# Genuine extension-tag pairs typically score 0.45â€“0.90.
# Unrelated concepts score 0.0â€“0.15.
# 0.25 gives comfortable margin while accepting loose-but-valid matches
# (e.g. an extension tag sharing 2-3 highly specific tokens with a sibling).
_FUZZY_THRESHOLD = 0.45

# Cache: Tier 2 results only (Tier 1 is already O(1)).
_FUZZY_CACHE: dict[str, str | None] = {}


def resolve_concept_to_label(concept: str | None) -> str | None:
    """
    Resolve an XBRL concept tag to its CONCEPT_MAP label name.

    Tier 1 â€” Exact match (O(1)):
        Every tag explicitly registered in CONCEPT_MAP.
        E.g. 'PaymentsToAcquirePropertyPlantAndEquipment' -> 'Capital Expenditures'

    Tier 2 â€” Max pairwise IDF-weighted Jaccard (O(labels Ã— tags_per_label)):
        For unregistered tags, tokenize the concept and compare it pairwise
        against every individually registered tag in CONCEPT_MAP.  The label
        whose single best-matching registered tag yields the highest
        IDF-weighted Jaccard score wins (if above _FUZZY_THRESHOLD).

        E.g. 'AdditionsToPropertyPlantAndEquipmentAndIntangibleAssets'
             (company extension, not in CONCEPT_MAP) ->
             nearest registered sibling:
             'PaymentsToAcquirePropertyPlantAndEquipmentAndIntangibleAssets'
             (Jaccard â‰ˆ 0.86) -> 'Capital Expenditures'

    Returns the matched label, or None if concept is empty or below threshold.
    Repeated calls for the same concept string are O(1) via cache.
    """
    if not concept:
        return None

    # Tier 1: exact
    exact = _CONCEPT_TAG_TO_LABEL.get(concept)
    if exact is not None:
        return exact

    # Cache check
    if concept in _FUZZY_CACHE:
        return _FUZZY_CACHE[concept]

    # Tier 2: max pairwise IDF-weighted Jaccard
    input_tokens = _tokenize_xbrl_concept(concept)
    if not input_tokens:
        _FUZZY_CACHE[concept] = None
        return None

    best_label: str | None = None
    best_score = 0.0

    candidate_labels: set[str] = set()
    for _tok in input_tokens:
        candidate_labels.update(_TOKEN_TO_LABELS.get(_tok, ()))

    if not candidate_labels:
        _FUZZY_CACHE[concept] = None
        return None

    for lbl, tag_token_list in _LABEL_TAG_TOKEN_LISTS.items():
        if lbl not in candidate_labels:
            continue
        for tag_tokens in tag_token_list:
            if not (input_tokens & tag_tokens):
                continue
            score = _idf_jaccard(input_tokens, tag_tokens)
            if score > best_score:
                best_score = score
                best_label = lbl

    result = best_label if best_score >= _FUZZY_THRESHOLD else None

    # Guard: a 'Noncash...' concept is a cash-flow reconciling add-back and must
    # never be fuzzy-matched onto an income-statement revenue/cost line just
    # because they share lease/cost tokens (PLTR 'NoncashOperatingLeaseCost'
    # was landing on 'Cost of Lease & Other Revenue').
    if result is not None and concept.lower().startswith('noncash'):
        _info = CONCEPT_MAP.get(result)
        if isinstance(_info, dict) and _info.get('cat') == '1_Income_Statement':
            result = None

    if result is not None:
        print(f"  [ConceptResolver] Fuzzy: '{concept}' -> '{result}' "
              f"(Jaccard={best_score:.3f})")

    _FUZZY_CACHE[concept] = result
    return result


# ---------------------------------------------------------------------------
# Strict Standard Tag Mapping (Segment & Core)
# ---------------------------------------------------------------------------
# These tags are mapped strictly to core metrics to avoid greedy prefix matching
# and ensure institutional-grade accuracy.
STANDARD_TAG_MAP = {
    'RevenueFromContractWithCustomerExcludingAssessedTax': 'Revenue',
    'ExternalRevenueFromContractWithCustomerByReportableSegment': 'Revenue',
    'SalesRevenueNet': 'Revenue',
    'Revenues': 'Revenue',
    'OperatingIncomeLoss': 'Operating Income',
    'OperatingProfitLoss': 'Operating Income',
    'GrossProfit': 'Gross Profit',
    'NetIncomeLoss': 'Net Income',
    'Assets': 'Assets',
    'Liabilities': 'Liabilities',
    'PropertyPlantAndEquipmentNet': 'Property, Plant & Equipment',
    'DepreciationDepletionAndAmortization': 'Depreciation & Amortization',
    'DepreciationAndAmortization': 'Depreciation & Amortization',
}

NO_SUBTRACT = ['Shares Outstanding', 'EPS', 'Margin', 'Ratio', 'ROE', 'ROA', 'Percentage', 'Percent', '%', 'Useful Life', 'RPO', 'Per Share', 'per share', 'Per Basic', 'Per Diluted']

SEGMENT_PREFIXES = [
    ('propertyplantandequipmentusefullife', 'Useful Life - PPE'),
    ('finitelivedintangibleassetusefullife', 'Useful Life - Intangibles'),
    ('revenueremainingperformanceobligationpercentage', 'RPO %'), 
    ('revenueremainingperformanceobligation', 'RPO Timing'),
    ('operatingincome', 'Operating Income'),
    ('operatingprofit', 'Operating Income'),
    ('segmentreportingotheritemamount', 'Operating Expenses'),
    ('operatingexpenses', 'Operating Expenses'),
    ('operatingcostsandexpenses', 'Operating Expenses'),
    ('netincome', 'Net Income'),
    ('contribution', 'Operating Income'),
    ('grossprofit', 'Gross Profit'),
    ('income', 'Income'), 
    ('revenue', 'Revenue'),
    ('sales', 'Revenue'),
    ('assets', 'Assets'),
    ('liabilities', 'Liabilities'),
    ('depreciation', 'Depreciation'),
    ('amortization', 'Amortization'),
    ('capitalexpenditure', 'Capital Expenditures'),
    ('inventory', 'Inventory'),
    ('rawmaterials', 'Inventory: Raw Materials'),
    ('workinprocess', 'Inventory: Work in Process'),
    ('finishedgoods', 'Inventory: Finished Goods'),
    ('operatingleaserightofuseasset', 'Operating Lease ROU Asset'),
    ('operatingleaseliability', 'Operating Lease Liability'),
    ('receivable', 'Accounts Receivable'),
    ('payable', 'Accounts Payable'),
    ('propertyplantandequipment', 'Property, Plant & Equipment'),
    ('deferredrevenue', 'Deferred Revenue'),
    ('contractliability', 'Deferred Revenue'),
    ('concentrationriskpercentage', 'Customer Concentration %'),
    ('concentrationriskrevenue', 'Customer Revenue'),
]


# ---------------------------------------------------------------------------
# Revenue Sub-Type Concepts (dim_count=0 passthrough)
# ---------------------------------------------------------------------------
# These CONCEPT_MAP Revenue tags represent a *portion* of total revenue, not
# the total itself (e.g. SubscriptionRevenue = $4.9B out of total $5.2B for ADBE).
# At dim_count=0 they normally get consumed by the CONCEPT_MAP as consolidated
# Revenue, then lose to higher-priority total-revenue tags during dedup.
# The real sub-total value disappears entirely from the segment output.
REVENUE_SUBTYPE_CONCEPTS = {
    'SubscriptionRevenue',
    'LicensesRevenue',
    'ServiceRevenue',
    'SalesRevenueGoodsNet',
    'OperatingLeasesIncomeStatementLeaseRevenue',
    'RevenueFromRelatedParties',
}

# ---------------------------------------------------------------------------
# Standard XBRL Axes for Segment Extraction
# ---------------------------------------------------------------------------
BUSINESS_SEGMENT_AXES = {
    'us-gaap:StatementBusinessSegmentsAxis',
    'us-gaap:SegmentReportingInformationBySegmentAxis',
    'us-gaap:SubsegmentsConsolidationItemsAxis', # For sub-segment granularity
}

GEOGRAPHIC_SEGMENT_AXES = {
    'srt:StatementGeographicalAxis',
    'us-gaap:StatementGeographicalAxis',
}

PRODUCT_AXIS = 'us-gaap:ProductOrServiceAxis'
# Income-statement face items that filers commonly disaggregate by nature of
# revenue (ProductOrServiceAxis) DIRECTLY on the face (ASC 606), e.g. IBM's
# Services/Sales/Financing or Microsoft's Product/Service split. Such a pure
# product/service breakdown of these lines belongs on the income statement,
# not in the business-segment block.
_FACE_DISAGG_PREFIXES = {'Revenue', 'Cost of Revenue'}
CONSOLIDATION_AXIS = 'us-gaap:ConsolidationItemsAxis'
# Statement-of-equity / AOCI rollforward axes. Dimensional facts on these axes
# are equity components (retained earnings, AOCI cash-flow-hedge & defined-benefit
# adjustments, treasury stock, NCI), not reportable business segments.
EQUITY_COMPONENT_AXES = ('StatementEquityComponentsAxis', 'EquityComponentsAxis')

# ---------------------------------------------------------------------------
# Geographic Segment Detection (Regions vs Countries)
# ---------------------------------------------------------------------------
# XBRL axis names that indicate a geographic breakdown dimension
GEOGRAPHIC_AXIS_KEYWORDS = {'geograph', 'country', 'region', 'territory', 'area', 'location'}

# Known Region keywords (regex-based detection)
REGION_KEYWORDS = r'\b(emea|apac|americas|international|worldwide|europe|asia|africa|north america|latin america|pacific|foreign|greater china|middle east|domestic|u\.s\.)\b'
# Precompiled form of the above, used on hot per-fact/per-member paths.  Compiling
# once and calling the pattern's methods is identical to re.search/re.sub with the
# string pattern (which the re module compiles-and-caches internally) but avoids the
# per-call cache lookup.
_REGION_KEYWORDS_RE = re.compile(REGION_KEYWORDS)

# Known geographic member names (lowercased, after clean_name processing)
# Includes all ISO 3166 country names and codes.
GEOGRAPHIC_COUNTRIES = {
    'afghanistan','albania','algeria','andorra','andorra','angola','antigua','argentina','armenia','australia','austria','azerbaijan',
    'bahamas','bahrain','bangladesh','barbados','belarus','belgium','belize','benin','bhutan','bolivia','bosnia','botswana',
    'brazil','brunei','bulgaria','burkina faso','burundi','cambodia','cameroon','canada','cape verde','central african republic',
    'chad','chile','china','colombia','comoros','congo','costa rica','croatia','cuba','cyprus','czech republic','denmark',
    'djibouti','dominica','dominican republic','ecuador','egypt','el salvador','equatorial guinea','eritrea','estonia','eswatini',
    'ethiopia','fiji','finland','france','gabon','gambia','georgia','germany','ghana','greece','grenada','guatemala','guinea',
    'guyana','haiti','honduras','hungary','iceland','india','indonesia','iran','iraq','ireland','israel','italy','jamaica',
    'japan','jordan','kazakhstan','kenya','kiribati','korea','kuwait','kyrgyzstan','laos','latvia','lebanon','lesotho',
    'liberia','libya','liechtenstein','lithuania','luxembourg','madagascar','malawi','malaysia','maldives','mali','malta',
    'marshall islands','mauritania','mauritius','mexico','micronesia','moldova','monaco','mongolia','montenegro','morocco',
    'mozambique','myanmar','namibia','nauru','nepal','netherlands','new zealand','nicaragua','niger','nigeria','north macedonia',
    'norway','oman','pakistan','palau','panama','papua new guinea','paraguay','peru','philippines','poland','portugal','qatar',
    'romania','russia','rwanda','saint kitts','saint lucia','saint vincent','samoa','san marino','sao tome','saudi arabia',
    'senegal','serbia','seychelles','sierra leone','singapore','slovakia','slovenia','solomon islands','somalia','south africa',
    'south sudan','spain','sri lanka','sudan','suriname','sweden','switzerland','syria','taiwan','tajikistan','tanzania',
    'thailand','timor-leste','togo','tonga','trinidad','tunisia','turkey','turkmenistan','tuvalu','uganda','ukraine','united arab emirates',
    'united kingdom','united states','uruguay','uzbekistan','vanuatu','vatican','venezuela','vietnam','yemen','zambia','zimbabwe',
    'u.s.','uk','usa','hong kong','taiwan','macau','uae',
}

# ISO Alpha-2 codes
GEOGRAPHIC_CODES = {
    'af','ax','al','dz','as','ad','ao','ai','aq','ag','ar','am','aw','au','at',
    'az','bs','bh','bd','bb','by','be','bz','bj','bm','bt','bo','bq','ba','bw',
    'bv','br','io','bn','bg','bf','bi','cv','kh','cm','ca','ky','cf','td','cl',
    'cn','cx','cc','co','km','cg','cd','ck','cr','ci','hr','cu','cw','cy','cz',
    'dk','dj','dm','do','ec','eg','sv','gq','er','ee','sz','et','fk','fo','fj',
    'fi','fr','gf','pf','tf','ga','gm','ge','de','gh','gi','gr','gl','gd','gp',
    'gu','gt','gg','gn','gw','gy','ht','hm','va','hn','hk','hu','is','in','id',
    'ir','iq','ie','im','il','it','jm','jp','je','jo','kz','ke','ki','kp','kr',
    'kw','kg','la','lv','lb','ls','lr','ly','li','lt','lu','mo','mg','mw','my',
    'mv','ml','mt','mh','mq','mr','mu','yt','mx','fm','md','mc','mn','me','ms',
    'ma','mz','mm','na','nr','np','nl','nc','nz','ni','ne','ng','nu','nf','mk',
    'mp','no','om','pk','pw','ps','pa','pg','py','pe','ph','pn','pl','pt','pr',
    'qa','re','ro','ru','rw','bl','sh','kn','lc','mf','pm','vc','ws','sm','st',
    'sa','sn','rs','sc','sl','sg','sx','sk','si','sb','so','za','gs','ss','es',
    'lk','sd','sur','sj','se','ch','sy','tw','tj','tz','th','tl','tg','tk','to',
    'tt','tn','tr','tm','tc','tv','ug','ua','ae','gb','um','us','uy','uz','vu',
    've','vn','vg','vi','wf','eh','ye','zm','zw'
}

# Member names to ignore (consolidated totals or placeholder dimensions)
IGNORE_MEMBERS = {
    'Operating', 'Reportable', 'Products And Services', 'Consolidated Entities', 
    'Segment', 'Scenario Unspecified', 'Operating Segments', 'Reportable Segments',
    'Intersegment Eliminations',
    'Eliminations', 'Corporate and Other', 'Corporate And Other', 'Adjustments'
}

# ---------------------------------------------------------------------------
# Segment Noise Concepts - XBRL tags that carry a segment dimension but are
# NOT financial performance metrics.  They leak into 4a_Segments_Business
# and pollute the output with ratios, HR-policy metrics, and EPS per share.
# ---------------------------------------------------------------------------
# Concept fragments that indicate a NON-segment disclosure even when the
# concept name happens to match a greedy segment prefix (e.g. 'sales').
# These are equity (treasury stock) or financing-receivable credit-quality
# disclosures, never business-segment performance metrics.
_NON_SEGMENT_CONCEPT_PATTERNS = ('treasurystock', 'financingreceivable')

SEGMENT_NOISE_CONCEPTS = {
    # HR / benefits ratios
    'DefinedContributionPlanEmployerMatchingContributionPercentOfMatch',
    'DefinedContributionPlanEmployerMatchingContributionPercent',
    'DefinedContributionPlanMaximumAnnualContributionsPerEmployeePercent',
    'DefinedContributionPlanEmployerMatchingContributionPercentOfEmployeesGrossPay',
    # EPS per-share metrics -- already in 1_Income_Statement; never a segment metric
    'IncomeLossFromContinuingOperationsPerBasicShare',
    'IncomeLossFromContinuingOperationsPerDilutedShare',
    'IncomeLossFromDiscontinuedOperationsNetOfTaxPerBasicShare',
    'IncomeLossFromDiscontinuedOperationsNetOfTaxPerDilutedShare',
    'EarningsPerShareBasic',
    'EarningsPerShareDiluted',
    # Tax jurisdiction disclosures mis-classified as segments
    'IncomeLossFromContinuingOperationsBeforeIncomeTaxesForeign',
    'IncomeLossFromContinuingOperationsBeforeIncomeTaxesDomestic',
    # Lease-extension options and practical expedient disclosures
    'LesseeOperatingLeaseRenewalTerm',
    'LesseeOperatingLeaseTermOfContract',
    'LesseeFinanceLeaseTermOfContract',
    # RSU / equity plan mechanics
    'ShareBasedCompensationArrangementByShareBasedPaymentAwardEquityInstrumentsOtherThanOptionsNonvestedNumber',
    # Other non-performance items
    'NumberOfOperatingSegments',
    'NumberOfReportableSegments',
    'RevenueRemainingPerformanceObligationPercentage',
    'RevenueReclassedAmount',
    'RevenueFromContractWithCustomerExpectedTimingOfSatisfactionPeriod1',
}

# Segment-metric value-range guards:
# If a fact has a segment dimension but its numeric value looks like a ratio
# (0 < val â‰¤ 2) AND the concept is NOT a known ratio metric, treat it as
# noise and skip it.  Ratios > 2 are allowed (e.g. useful life in years).
_SEG_RATIO_PREFIXES = {
    'concentrationrisk', 'percentage', 'percent', 'ratio',
    'revenueremainingperformanceobligationpercentage',
    'effectiveincometaxrate',
}

def _is_segment_noise_value(concept, val):
    """Return True if this value looks like a leaked ratio / percentage in a segment row."""
    if not isinstance(val, (int, float)) or pd.isna(val):
        return False
    c_lower = concept.lower()
    # If it's a known ratio-metric concept, allow it (handled elsewhere)
    if any(p in c_lower for p in _SEG_RATIO_PREFIXES):
        return False
    # Values between -2 and 2 (exclusive of 0) that are NOT integers are ratios
    if 0 < abs(val) <= 2 and val != int(val):
        return True
    return False

# ---------------------------------------------------------------------------
# Category sort order
# ---------------------------------------------------------------------------
CAT_ORDER = {
    # The period-date metadata row must be the first data row in CSV/XLSX,
    # immediately below the Category/Label column header.  The authoritative
    # pre-write sorter also runs on cached pivots, so assigning it the only
    # negative category rank keeps this invariant across fresh and cached
    # quarterly/annual native outputs.
    '0_Period_Header':      -1,
    '1_Income_Statement':    0,
    '2_Balance_Sheet':       1,
    '3_Cash_Flow':           2,
    '4a_Segments_Business':  3,
    '4b_Segments_Geographic_Regions': 4,
    '4c_Segments_Geographic_Countries': 4,
    '5_KPI_Metrics':         5,
    '6_Disclosures':         6,
    '6_Disclosures_Cross_Tabulated': 6,
    '7_Concentration_Risk':  7,
    '8_Integrity_Checks':    8,
}

# Preferred metric-type order within segment sections
SEGMENT_METRIC_ORDER = [
    'Revenue', 'Operating Measure', 'Gross Profit', 'Operating Income', 'Operating Expenses', 'Net Income',
    'Assets', 'Liabilities', 'Depreciation', 'Amortization',
    'Capital Expenditures', 'Inventory', 'Inventory: Raw Materials', 'Inventory: Work in Process', 'Inventory: Finished Goods',
    'Operating Lease ROU Asset', 'Operating Lease Liability',
    'Accounts Receivable',
    'Accounts Payable', 'Property, Plant & Equipment', 'Deferred Revenue', 'Income',
    'Customer Concentration %', 'Customer Revenue',
]

# Segment category markers for easy filtering
SEG_CATS = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'}

# Genuine segment metrics that are considered core business segments
GENUINE_SEGMENT_METRICS = {
    'Revenue', 'Operating Measure', 'Gross Profit', 'Operating Income', 'Operating Expenses', 
    'Total Operating Expenses', 'Net Income', 'Cost of Revenue',
    'Depreciation', 'Amortization', 'Depreciation & Amortization',
    'Capital Expenditures', 'Inventory', 'Inventory: Raw Materials', 
    'Inventory: Work in Process', 'Inventory: Finished Goods',
    'Accounts Receivable', 'Accounts Payable', 'Property, Plant & Equipment', 
    'Deferred Revenue', 'Income', 'Customer Revenue'
}

# Logical display order for KPI Metrics
KPI_ORDER = [
    'Metric: Gross Margin %',
    'Metric: Net Margin %',
    'Metric: EBIT Margin %',
    'Metric: EBITDA',
    'Metric: EBITDA Margin %',
    'Metric: Free Cash Flow',
    'Metric: FCF Margin %',
    'Metric: Unlevered Free Cash Flow',
    'Metric: Total Debt',
    'Metric: Total Lease Liabilities',
    'Metric: Net Cash (Debt)',
    'Metric: ROE % (Annualised)',
    'Metric: Effective Tax Rate %',
]

# Public/export labels created by _normalize_output_margin_rows must retain the
# same sort position as their internal calculation labels.  Without this alias,
# FCF Margin (%) is treated as an unknown KPI after normalization and falls to
# the bottom of the KPI section.
KPI_ORDER_ALIASES = {
    'FCF Margin (%)': 'Metric: FCF Margin %',
}

# Disclosure noise patterns: labels whose sub-part (after " - ") looks like a raw
# XBRL tag rather than a real segment name.  Matched case-insensitively.
DISCLOSURE_PATTERNS = [
    'effective income tax rate reconciliation',
    'income tax reconciliation',
    'income tax paid',
    'income taxes paid',
    'deferred federal income tax',
    'deferred foreign income tax',
    'deferred state and local income tax',
    'deferred income tax expense',
    'federal income tax expense',
    'foreign income tax expense',
    'state and local income tax expense',
    'state and local jurisdiction',
    'other comprehensive income',
    'accumulated net unrealized',
    'accumulated translation adjustment',
    'aoci',
    'comprehensive income net of tax',
    'unrecognized tax benefits',
    'income loss from continuing operations before income taxes',
    'income loss from equity method investments',
    'investment income interest',
    'miscellaneous income expense',
    'other noncash income expense',
    'payments to acquire available for sale',
    'proceeds from property plant and equipment',
    'contract with customer liability revenue recognized',
    'equity method investment nonconsolidated',
    'increase decrease in',
    'allowance for doubtful',
    'capital expenditures incurred but not yet paid',
    'amortization of intangible assets',
    'segment expenditure addition',
    'payments to acquire productive assets',
    'nonoperating income expense',
    'operating income loss',
    'retained earnings',
    'service life',
    'change in property and equipment useful life',
    'technology and computer equipment',
    'depreciation depletion and amortization',
    'finance lease right of use asset amortization',
    'corporate non',
    'effective income tax rate continuing operations',
    'income tax expense benefit',
    'material reconciling items',
    'adjustment for amortization',
    'net inventory and related charges',
    'geographic concentration risk',
    'proceeds from sale of available for sale',   # investment proceeds â‰  revenue
    'other comprehensive income loss available for sale',
    'contract with customer liability revenue recognized',
    'inventory purchase obligations',              # purchase obligation charge
    'h20product',                                  # H20 chip inventory charge
    'excess inventory and purchase',
    'defined contribution plan',
    'employer matching contribution',
    'maximum annual contributions',
    'percentage of',
    'percent of',
    'fair value inputs',
    'hierarchy',
    'level 1',
    'level 2',
    'level 3',
    'level1',
    'level2',
    'level3',
    'weighted average',
    'useful life',
    'service life',
    'remaining performance obligation',
    'operating lease liability'

    # -- Tax jurisdiction detail (state / foreign) -------------------------
    # Catches "Income - Foreign Tax Jurisdiction Other" / "Income - State And Local
    # Tax Jurisdiction Other" that leak into 4a_Segments_Business.
    'tax jurisdiction',

    # -- AMD annual-only XBRL extension tags (10-K only, never quarterly) --
    # "Depreciation - Cost Depreciation Amortization And Depletion"
    'cost depreciation amortization and depletion',
]

_AND_PROTECT_WORDS = (
    # place names whose stem ends in '...and' (mostly '-land')
    'ireland', 'iceland', 'finland', 'poland', 'thailand', 'zealand', 'switzerland',
    'greenland', 'england', 'scotland', 'holland', 'swaziland', 'newfoundland',
    'queensland', 'rhineland', 'maryland', 'portland', 'oakland', 'cleveland',
    'auckland', 'sunderland', 'lapland', 'gotland', 'jutland', 'saarland',
    'island', 'highland', 'lowland', 'homeland', 'mainland', 'wetland', 'grassland',
    'woodland', 'farmland', 'wasteland', 'wonderland', 'motherland', 'fatherland',
    'hinterland', 'borderland', 'headland', 'parkland', 'moorland', 'marshland',
    'peatland', 'rangeland', 'cropland', 'midland', 'flatland', 'inland', 'upland',
    'overland', 'dryland', 'heathland', 'badland',
    # ordinary words ending in '...and'
    'thousand', 'husband', 'demand', 'command', 'expand', 'disband', 'errand',
    'strand', 'brand', 'grand', 'gland', 'bland', 'reprimand', 'contraband',
    'ampersand', 'viand', 'riband', 'remand', 'garland', 'wetland',
)

def _split_glued_and(s: str) -> str:
    """Split a lowercased 'and' connector joining two TitleCase fragments.

    'ProfessionalServicesandOther' -> 'ProfessionalServices and Other'
    'SubscriptionandSupport'       -> 'Subscription and Support'
    while protecting words that genuinely end in '...and' (Ireland, Demand,
    Thousand, Brand, ...). Must run before the lowercase->uppercase camelCase
    splitter, which would otherwise break the 'and'+Uppercase adjacency.
    """
    def repl(mo):
        stem = mo.group(1)
        word = (stem + 'and').lower()
        if len(stem) < 3 or any(word.endswith(w) for w in _AND_PROTECT_WORDS):
            return mo.group(0)
        return stem + ' and ' + mo.group(2)
    return re.sub(r'([A-Za-z]+?)and([A-Z][a-z])', repl, s)

def clean_name(name):
    """Normalize label strings to ensure 10-Q custom tags and 10-K dimension members perfectly align."""
    m = str(name).replace('Member', '').replace('Domain', '').split(':')[-1]
    m = m.replace('StatementBusinessSegmentsAxis', 'OperatingSegments')
    m = m.replace('StatementOperatingSegmentsAxis', 'OperatingSegments')
    
    # Pass 1: split where an uppercase-sequence is immediately followed by
    # an uppercase+lowercase pair, e.g. "OEMAnd" -> "OEM And", "GPUChip" -> "GPU Chip".
    m = re.sub(r'([A-Z]{2,})([A-Z][a-z])', r'\1 \2', m)
    # Pass 1b: split a lowercased 'and' connector (Salesforce-style element names
    # 'SubscriptionandSupport' / 'ProfessionalServicesandOther'). Runs before the
    # camelCase split so the 'and'+Uppercase boundary it relies on is intact.
    m = _split_glued_and(m)
    # Pass 2: standard camelCase split, e.g. "DataCenter" -> "Data Center"
    m = re.sub(r'([a-z])([A-Z])', r'\1 \2', m)
    # Pass 3 (FIXED): split 'And' as a conjunction ONLY when followed by uppercase â€”
    # that means it joins two CamelCase word fragments, not part of an English word.
    #   Conjunction  : 'ComputeAndNetworking' -> 'And'+'N'(upper) -> 'Compute And Networking' âœ“
    #   Compound word: 'HAndsets'             -> 'And'+'s'(lower) -> skip (handled by Pass 4) âœ“
    m = re.sub(r'(\w)([Aa]nd)([A-Z])', r'\1 And \3', m)
    # Pass 4 (NEW): de-encode compound English words where a single uppercase letter
    # is immediately followed by 'And' + lowercase continuation â€” the CamelCase
    # encoding of words like 'Handsets' (H+and+sets = HAndsets) and 'Bandwidth'.
    # Examples: HAndsets -> Handsets,  BAndwidth -> Bandwidth
    m = re.sub(r'\b([A-Z])And([a-z])', lambda x: x.group(1) + 'and' + x.group(2), m)
    m = m.replace('Asia Pacific', 'Asia')
    # Standardize common noisy delimiters
    m = m.replace(':', '').strip()
    m = m.replace('Datacenter', 'Data Center')
    
    # Strip off common suffixes that create mismatch
    m = re.sub(r'\s+(Segment|Revenue|Sales|Net|Amount|Total)$', '', m, flags=re.IGNORECASE).strip()
    return m


@lru_cache(maxsize=65536, typed=True)
def _normalize_member_label(m: str) -> str:
    """
    Normalize punctuation-variant abbreviation forms so that labels
    representing the same entity resolve to the same canonical string.

    Problem this solves
    -------------------
    Companies that file both 10-K and 10-Q reports sometimes use different
    typographic conventions for the same member label across filing types:

      10-K XBRL label : "U.S. Government"   (dotted abbreviation)
      10-Q XBRL label : "U S Government"    (space-separated letters)
      tag fallback     : "USGovernment"      (compact CamelCase â€” no separator)

    Because xbrl.get_labels_for_concept() returns the raw human-readable
    label, the dots (or spaces) are preserved verbatim.  In some filings the
    XBRL taxonomy does not define a human-readable label and the library
    returns the tag name itself (e.g. "USGovernment [Member]"), which after
    stripping "[Member]" yields a compact CamelCase form with no separator.
    All three forms must resolve to the same canonical string so that the
    10-K annual row and the 10-Q quarterly rows share a label and Q4
    derivation (annual YTD minus YTD9 baseline) can succeed.

    Rules applied (in order)
    ------------------------
    Rule 1 â€“ Dotted abbreviations:
        Strip internal dots from sequences where each character is a single
        uppercase letter followed by a dot.
        Examples:  'U.S.' -> 'US',  'U.K.' -> 'UK',  'U.A.E.' -> 'UAE'
        Also handles the no-trailing-dot variant: 'U.S' -> 'US'

    Rule 2 â€“ Spaced abbreviations:
        Collapse a run of space-separated single uppercase letters into one
        token.
        Examples:  'U S Government' -> 'US Government',  'R O W' -> 'ROW'

    Rule 3 â€“ Compact CamelCase abbreviations:
        Split a run of 2+ uppercase letters that is immediately followed by
        an uppercase+lowercase pair (i.e. a new title-case word).
        Examples:  'USGovernment' -> 'US Government'
                   'USMilitary'   -> 'US Military'
                   'OEMAnd...'    -> handled correctly via backtracking
        This is the same as Pass 1 in clean_name() and covers the case where
        xbrl.get_labels_for_concept() returns the raw tag name (no spaces).
        Rules 1 and 2 are applied first so a dotted/spaced form never reaches
        Rule 3 in a broken intermediate state.

    Safety
    ------
    All three rules are no-ops for normal multi-word segment names:
      'Data Center', 'Gaming', 'Professional Visualization', 'Automotive',
      'OEM', 'North America', 'Asia Pacific', 'Middle East And Africa'.
    This fix is therefore transparent for NVDA, AMD, MSFT, GOOGL, and any
    other company whose segments do not use dotted/spaced/compact abbreviations.
    """
    # Rule 1a: Remove dots from 2+-letter dotted abbreviations, e.g. 'U.S.' -> 'US'
    m = re.sub(r'\b((?:[A-Z]\.){2,})', lambda mo: mo.group(0).replace('.', ''), m)
    # Rule 1b: Handle two-letter without trailing dot, e.g. 'U.S' -> 'US'
    m = re.sub(r'\b([A-Z])\.([A-Z])\b', r'\1\2', m)
    # Rule 2: Collapse space-separated single uppercase letters, e.g. 'U S' -> 'US'
    m = re.sub(r'\b((?:[A-Z] )+[A-Z])\b', lambda mo: mo.group(0).replace(' ', ''), m)
    # Rule 3: Split compact CamelCase abbreviation+word, e.g. 'USGovernment' -> 'US Government'
    # Matches a sequence of 2+ uppercase letters immediately followed by an uppercase+lowercase
    # pair (start of a new title-case word).  Identical to clean_name() Pass 1.
    m = re.sub(r'([A-Z]{2,})([A-Z][a-z])', r'\1 \2', m)
    return m


def parse_duration(val):
    """Convert ISO 8601 duration strings (e.g. P5Y, P10Y6M) to numeric years."""
    if not isinstance(val, str) or not val.startswith('P'):
        return val
    years = 0.0
    y_match = re.search(r'(\d+)Y', val)
    if y_match:
        years += float(y_match.group(1))
    m_match = re.search(r'(\d+)M', val)
    if m_match:
        years += float(m_match.group(1)) / 12.0
    return years if (y_match or m_match) else val

def get_period_info(end_date, ye_month, dur=0):
    dt = pd.to_datetime(end_date)
    # Normalizes 52/53 week filers (shifts a late month-end closing into the proper bucket)
    if dt.day < 15:
        dt = dt - pd.DateOffset(months=1)
        
    fy = dt.year + (1 if dt.month > ye_month else 0)
    
    m_into_fy = (dt.month - ye_month) % 12
    if m_into_fy == 0: q = 'Q4'
    elif m_into_fy <= 3: q = 'Q1'
    elif m_into_fy <= 6: q = 'Q2'
    elif m_into_fy <= 9: q = 'Q3'
    else: q = 'Q4'
    
    return fy, q

def extract_period_end_date(filing, facts_df):
    """Extract the period-ending date from DEI XBRL tags.
    - 10-K (Annual / Q4): uses CurrentFiscalYearEndDate
    - 10-Q (Quarterly)  : uses DocumentPeriodEndDate
    Falls back to filing.period_of_report when the tag is absent.
    """
    tag_name = ('CurrentFiscalYearEndDate' if '10-K' in getattr(filing, 'form', '')
                else 'DocumentPeriodEndDate')

    concept_short = (facts_df['_concept_short'] if '_concept_short' in facts_df.columns
                     else facts_df['concept'].str.rsplit(':', n=1).str[-1])
    mask = concept_short == tag_name
    matches = facts_df[mask]

    if not matches.empty:
        raw = matches.iloc[0]['value']
        if pd.notna(raw):
            raw_str = str(raw).strip()
            # CurrentFiscalYearEndDate is sometimes stored as --MM-DD (no year)
            if raw_str.startswith('--'):
                filing_year = pd.to_datetime(filing.period_of_report).year
                try:
                    return pd.to_datetime(f"{filing_year}-{raw_str[2:]}",
                                          format='%Y-%m-%d').strftime('%m/%d/%y')
                except Exception:
                    pass
            else:
                try:
                    return pd.to_datetime(raw_str).strftime('%m/%d/%y')
                except Exception:
                    pass

    # Fallback: filing.period_of_report
    try:
        return pd.to_datetime(filing.period_of_report).strftime('%m/%d/%y')
    except Exception:
        return None


def augment_facts_with_calculations(facts_df, calc_trees, max_passes=2):
    import pandas as pd
    if facts_df is None or facts_df.empty or not calc_trees:
        return facts_df

    calc_relationships = []
    for role_uri, tree in calc_trees.items():
        if hasattr(tree, 'all_nodes'):
            for elem_id, node in tree.all_nodes.items():
                parent = node.parent
                if parent:
                    p = parent.replace('_', ':', 1) if '_' in parent else parent
                    c = elem_id.replace('_', ':', 1) if '_' in elem_id else elem_id
                    
                    p_clean = p.split(':')[-1]
                    c_clean = c.split(':')[-1]
                    GLOBAL_CALC_PARENT.setdefault(c_clean, set()).add((p_clean, node.weight))
                    
                    calc_relationships.append({
                        'parent': p,
                        'child': c,
                        'weight': node.weight
                    })
    if not calc_relationships:
        return facts_df

    calc_df = pd.DataFrame(calc_relationships).drop_duplicates()
    
    dim_cols = [c for c in facts_df.columns if c.startswith('dim_')]
    context_cols = ['period_start', 'period_end', 'period_instant', 'unit_ref'] + dim_cols
    
    facts = facts_df.dropna(subset=['value']).copy()
    facts['value'] = pd.to_numeric(facts['value'], errors='coerce')
    facts = facts.dropna(subset=['value'])
    if 'is_calculated' not in facts.columns:
        facts['is_calculated'] = False
    
    new_facts = []
    grouped = facts.groupby(context_cols, dropna=False)
    
    # Pre-map parent to children for faster lookup.  This reproduces
    # ``calc_df.groupby('parent').apply(lambda x: list(zip(x['child'],
    # x['weight']))).to_dict()`` exactly -- ``groupby`` sorts its keys and
    # preserves original row order within each group -- while avoiding the
    # per-group Python overhead of groupby/apply on every filing.
    parent_to_children = {}
    for _parent, _child, _weight in zip(calc_df['parent'], calc_df['child'], calc_df['weight']):
        parent_to_children.setdefault(_parent, []).append((_child, _weight))
    if parent_to_children:
        parent_to_children = {_k: parent_to_children[_k] for _k in sorted(parent_to_children)}
    
    for context_vals, group in grouped:
        concept_to_value = dict(zip(group['concept'], group['value']))
        added_in_context = {}
        
        for _ in range(max_passes):
            changed = False
            for parent, children_info in parent_to_children.items():
                # --- RULE 1: Derive Parent from Children ---
                if parent not in concept_to_value:
                    all_present = True
                    p_val = 0
                    for child, weight in children_info:
                        if child in concept_to_value:
                            p_val += concept_to_value[child] * weight
                        else:
                            all_present = False
                            break
                    if all_present and children_info:
                        concept_to_value[parent] = p_val
                        added_in_context[parent] = p_val
                        changed = True
                
                # --- RULE 2: Derive Missing Child from Parent ---
                else:
                    parent_val = concept_to_value[parent]
                    missing_children = []
                    present_sum = 0
                    for child, weight in children_info:
                        if child in concept_to_value:
                            present_sum += concept_to_value[child] * weight
                        else:
                            missing_children.append((child, weight))
                    
                    if len(missing_children) == 1:
                        child, weight = missing_children[0]
                        if weight != 0:
                            missing_val = (parent_val - present_sum) / weight
                            
                            # Phantom Zero Guard: If a derived child is 0 but the parent is materially non-zero, 
                            # it means the present children already perfectly explain the parent. This happens 
                            # when the calculation linkbase contains mutually exclusive presentation paths. 
                            # We skip adding these phantom zeros to prevent them from overriding real reported facts.
                            if abs(missing_val) < 1e-5 and abs(parent_val) > 1e-5:
                                continue
                            concept_to_value[child] = missing_val
                            added_in_context[child] = missing_val
                            changed = True
            if not changed:
                break
                
        if added_in_context:
            for concept, val in added_in_context.items():
                row = dict(zip(context_cols, context_vals))
                row['concept'] = concept
                row['value'] = val
                row['is_calculated'] = True
                new_facts.append(row)
                
    if new_facts:
        new_df = pd.DataFrame(new_facts)
        return pd.concat([facts_df, new_df], ignore_index=True)
    return facts_df

def _infer_html_scale(extracted, html_nums):
    """
    Infer the correct multiplier for raw HTML table numbers by comparing
    against a known XBRL reference value from the same filing.

    Strategy:
      1. Find a reliable XBRL-extracted value (Revenue preferred, then
         Operating Income, Net Income) from the same quarter.
      2. For each candidate scale (1, 1_000, 1_000_000), check which one
         makes the HTML number closest to the XBRL reference.
      3. If no XBRL reference is available, fall back to magnitude-based
         inference: numbers < 50_000 are likely "in millions" already;
         numbers > 1_000_000 are likely raw (scale = 1).

    Returns the scale factor (int).
    """
    CANDIDATE_SCALES = [1, 1_000, 1_000_000]

    # Step 1: Collect XBRL reference values for calibration
    ref_labels = ['Revenue', 'Operating Income', 'Net Income', 'Cost of Revenue']
    xbrl_refs = {}
    for fact in extracted:
        if fact['Label'] in ref_labels and fact.get('DimCount', 0) == 0:
            lbl = fact['Label']
            try:
                val = abs(float(fact['Value']))
                if val > 0 and (lbl not in xbrl_refs or val > xbrl_refs[lbl]):
                    xbrl_refs[lbl] = val
            except (ValueError, TypeError):
                continue

    # Step 2: Try to match against XBRL reference
    if xbrl_refs and html_nums:
        # Use the largest available reference (typically Revenue)
        ref_val = max(xbrl_refs.values())
        html_raw = max(abs(n) for n in html_nums[:4] if n != 0) if html_nums else 0

        if html_raw > 0:
            best_scale, best_err = 1_000_000, float('inf')
            for sc in CANDIDATE_SCALES:
                scaled = html_raw * sc
                # Compare order of magnitude
                err = abs(scaled - ref_val) / max(ref_val, 1)
                if err < best_err:
                    best_err = err
                    best_scale = sc
            # Accept if the best scale puts us within 100x of the reference
            if best_err < 100:
                return best_scale

    # Step 3: Magnitude-based fallback
    # SEC filings typically use "in millions" or "in thousands"
    if html_nums:
        max_html = max(abs(n) for n in html_nums[:4] if n != 0) if html_nums else 0
        if max_html > 1_000_000:
            return 1           # already in raw units
        elif max_html > 10_000:
            return 1_000       # likely "in thousands"
        else:
            return 1_000_000   # likely "in millions"

    return 1_000_000  # conservative default


# ---------------------------------------------------------------------------
# Dynamic Face-Statement Concept Learning
# ---------------------------------------------------------------------------
# Problem this solves: any XBRL tag that is not in CONCEPT_MAP (exactly or via
# the fuzzy resolver / linkbase heuristics / Arelle anchoring) used to be
# silently dropped, so company-specific face-statement lines (e.g. Amazon's
# "Fulfillment", Alphabet's "Cost of revenues TAC") never reached the CSV.
#
# Strategy: the filing itself tells us which concepts sit on the face
# financial statements via its presentation linkbase.  edgartools classifies
# each presentation role (IncomeStatement / BalanceSheet / CashFlowStatement).
# For every non-abstract, numeric concept on a face statement that NO existing
# tier can resolve, we dynamically register a new CONCEPT_MAP entry using the
# company's own human-readable label for the line.  From that point on the
# normal extraction / dedup / quarterly-derivation machinery treats it exactly
# like a hand-curated concept -- so the output is complete for ANY filer,
# without flooding the CSV with footnote-level disclosures (only concepts
# actually presented on the face statements are learned).
#
# Thread safety: filings are processed in a ThreadPoolExecutor.  Unlike the
# existing learners (which only append to tag lists), this adds NEW KEYS to
# CONCEPT_MAP, so all mutations happen under _CONCEPT_LEARN_LOCK and the
# extraction loop iterates over a per-filing snapshot (see extract_from_filing).
# ---------------------------------------------------------------------------

_CONCEPT_LEARN_LOCK = threading.Lock()

# edgartools statement type -> our output category.
# 'ComprehensiveIncome' is included because filers that combine the income
# statement and OCI into one statement get that type for their PRIMARY P&L.
_STMT_TYPE_TO_CAT = {
    'IncomeStatement':     '1_Income_Statement',
    'ComprehensiveIncome': '1_Income_Statement',
    'BalanceSheet':        '2_Balance_Sheet',
    'CashFlowStatement':   '3_Cash_Flow',
}

_AUTO_SKIP_SUFFIXES = ('Abstract', 'Axis', 'Member', 'Domain', 'Table',
                       'LineItems', 'Hierarchy', 'RollUp', 'Rollforward')
_AUTO_SKIP_SUBSTR   = ('TextBlock', 'PolicyText', 'TableText')

# Concepts already evaluated once -- avoids re-running label resolution for
# every filing of the same company (concepts repeat across quarters).
_AUTO_SEEN_CONCEPTS: set = set()


def _normalize_label_key(label: str) -> str:
    """
    Canonical key for matching a filer's human-readable line label against
    existing CONCEPT_MAP labels, so e.g. 'Selling, general and administrative'
    merges into 'Selling, General & Admin'-adjacent labels, and era variants
    like 'Trading Gain (Loss)' / 'Trading gains' converge to one key instead
    of creating duplicate rows.  Steps: drop parentheticals, lowercase,
    '&'->'and', strip punctuation, drop leading 'total ' and trailing
    ' net'/' total', then per-token depluralization (except '-ss' words).
    """
    l = re.sub(r'\(.*?\)', ' ', str(label).lower()).replace('&', 'and')
    l = re.sub(r'[^\w\s]', ' ', l)
    l = re.sub(r'\s+', ' ', l).strip()
    if l.startswith('total '):
        l = l[6:]
    for suffix in (' net', ' total'):
        if l.endswith(suffix):
            l = l[:-len(suffix)]
    toks = []
    for t in l.split():
        if len(t) > 3 and t.endswith('s') and not t.endswith('ss'):
            t = t[:-1]
        toks.append(t)
    return ' '.join(toks)


_NORMALIZED_LABEL_INDEX: dict[str, str] = {
    _normalize_label_key(_lbl): _lbl for _lbl in CONCEPT_MAP
}

_CAT_SUFFIX = {'1_Income_Statement': 'IS', '2_Balance_Sheet': 'BS', '3_Cash_Flow': 'CF'}


def _classify_statement_roles(xbrl) -> dict:
    """
    Map presentation-role URI -> output category for the face financial
    statements of this filing.  Prefers edgartools' own classification
    (get_all_statements), falling back to keyword matching on the role
    definition for older/odd taxonomies.  Parenthetical roles are excluded.
    """
    trees = getattr(xbrl, 'presentation_trees', None) or {}
    role_to_cat = {}

    try:
        for stmt in (xbrl.get_all_statements() or []):
            stype = stmt.get('type') or ''
            if not stype or 'Parenthetical' in stype:
                continue
            cat = _STMT_TYPE_TO_CAT.get(stype)
            if cat:
                role_to_cat[stmt.get('role')] = cat
    except Exception:
        pass

    if not role_to_cat:
        for role, tree in trees.items():
            d = (getattr(tree, 'definition', '') or '').lower()
            if not d or 'parenthetical' in d:
                continue
            if 'balance sheet' in d or 'financial position' in d:
                role_to_cat[role] = '2_Balance_Sheet'
            elif 'cash flow' in d:
                role_to_cat[role] = '3_Cash_Flow'
            elif 'statement' in d and any(k in d for k in ('operation', 'income', 'earnings')):
                # Pure-OCI statements are skipped; combined P&L+OCI is kept.
                if 'comprehensive' in d and not any(k in d for k in ('operation', 'earnings', 'and comprehensive')):
                    continue
                role_to_cat[role] = '1_Income_Statement'
    return role_to_cat


def _is_numericish_element(xbrl, elem_id, concept) -> bool:
    """True unless the taxonomy catalog says this element is non-numeric/abstract."""
    try:
        catalog = getattr(xbrl, 'element_catalog', None) or {}
        item = catalog.get(elem_id) or catalog.get(concept)
        if item is None:
            return True  # unknown -> let the numeric coercion downstream decide
        if getattr(item, 'abstract', False):
            return False
        dt = (getattr(item, 'data_type', '') or '').lower()
        if not dt:
            return True
        return any(k in dt for k in ('monetary', 'pershare', 'shares', 'percent',
                                     'decimal', 'integer', 'pure'))
    except Exception:
        return True


# OCI components from combined P&L+OCI statements, and capital-stock
# boilerplate (shares authorized/issued, par values) from the face balance
# sheet, add noise without analytical value -- institutional sheets omit them.
_OCI_LABEL_PATTERNS = (
    'other comprehensive income', 'oci,', 'oci ', 'comprehensive income',
    'cumulative translation', 'reclassification adjustment',
    'items of other comprehensive income', 'unrealized holding gain',
)
_CAPSTOCK_NOISE_SUBSTR = ('SharesAuthorized', 'SharesIssued', 'SharesOutstanding',
                          'ParOrStatedValuePerShare', 'ParValuePerShare')
_CAPSTOCK_NOISE_PREFIX = ('CommonStock', 'PreferredStock', 'TreasuryStock',
                          'TemporaryEquity', 'ConvertiblePreferredStock')


def _is_learning_noise(concept: str, label: str) -> bool:
    cl = concept.lower()
    if 'comprehensiveincome' in cl or cl.startswith('oci'):
        return True
    ll = (label or '').lower()
    if any(p in ll for p in _OCI_LABEL_PATTERNS):
        return True
    if (any(s in concept for s in _CAPSTOCK_NOISE_SUBSTR)
            and concept.startswith(_CAPSTOCK_NOISE_PREFIX)):
        return True
    return False


def _presentation_order(tree) -> list:
    """
    Walk a presentation tree in document order (roots first, children sorted
    by their 'order' attribute), so learned concepts can be anchored to the
    position the company itself presents them at.  Falls back to dict
    insertion order for any disconnected nodes.
    """
    nodes = getattr(tree, 'all_nodes', {}) or {}
    visited, out = set(), []

    def _kid_ids(n):
        ids = []
        for c in (getattr(n, 'children', None) or []):
            cid = c if isinstance(c, str) else getattr(c, 'element_id', None)
            if cid:
                ids.append(cid)
        ids.sort(key=lambda cid: float(getattr(nodes.get(cid), 'order', 0) or 0)
                 if nodes.get(cid) is not None else 0.0)
        return ids

    def _walk(eid):
        if eid in visited or eid not in nodes:
            return
        visited.add(eid)
        out.append(eid)
        for cid in _kid_ids(nodes[eid]):
            _walk(cid)

    roots = [eid for eid, n in nodes.items() if not getattr(n, 'parent', None)]
    try:
        roots.sort(key=lambda eid: float(getattr(nodes[eid], 'order', 0) or 0))
    except Exception:
        pass
    for r in roots:
        _walk(r)
    for eid in nodes:
        if eid not in visited:
            out.append(eid)
    return out


# Per-anchor sequence counters so multiple lines learned after the same
# curated anchor keep their relative presentation order.
_ANCHOR_SEQ: dict = {}


def _apply_presentation_anchors(item_order: dict, is_financial: bool = False,
                                is_insurance: bool = False) -> dict:
    """
    Re-position auto-learned labels using company presentation order first,
    but validate the anchor against calculation/semantic evidence so a custom
    opex/revenue line is not dragged below Net Income by a weak fallback anchor.
    """
    item_order = dict(item_order)
    for lbl, info in list(CONCEPT_MAP.items()):
        if not (isinstance(info, dict) and info.get('auto')):
            continue
        pos = _presentation_anchor_pos(lbl, item_order, is_financial, is_insurance)
        if pos is not None:
            item_order[lbl] = float(pos)
    return item_order


def learn_statement_concepts(xbrl) -> int:
    """
    Scan this filing's face financial statements (via the presentation
    linkbase) and dynamically register every concept that the existing
    resolution tiers (exact CONCEPT_MAP, fuzzy resolver, STANDARD_TAG_MAP)
    cannot already place.  Labels reuse the filer's own human-readable line
    names, merging into existing CONCEPT_MAP labels when the names match.

    Returns the number of concepts newly registered or merged.
    """
    role_to_cat = _classify_statement_roles(xbrl)
    if not role_to_cat:
        return 0

    trees = getattr(xbrl, 'presentation_trees', None) or {}
    learned = 0

    # Pass 0 (inversion): record where this filing ACTUALLY presents each
    # face concept. The static map is only trusted when the company itself
    # presents the concept on that statement somewhere in its filings.
    for role0, cat0 in role_to_cat.items():
        tree0 = trees.get(role0)
        nodes0 = getattr(tree0, 'all_nodes', None)
        if not nodes0:
            continue
        try:
            if sum(1 for _n in nodes0.values()
                   if not getattr(_n, 'is_abstract', False)) < 5:
                continue
        except Exception:
            pass
        for _pos0, _eid in enumerate(_presentation_order(tree0)):
            _nd = nodes0.get(_eid)
            if _nd is None:
                continue
            try:
                if getattr(_nd, 'is_abstract', False):
                    continue
                _c0 = _eid.replace('_', ':', 1).split(':')[-1]
                if _c0:
                    _FACE_PRESENTED.setdefault(_c0, set()).add(cat0)
                    _FACE_PRESENTATION_POS.setdefault(_c0, {}).setdefault(cat0, []).append(float(_pos0))
            except Exception:
                continue

    for role, cat in role_to_cat.items():
        tree = trees.get(role)
        nodes = getattr(tree, 'all_nodes', None)
        if not nodes:
            continue
        try:
            _n_concrete = sum(1 for _n in nodes.values()
                              if not getattr(_n, 'is_abstract', False))
        except Exception:
            _n_concrete = 99
        if _n_concrete < 5:
            continue  # disclosure schedule, not a face statement

        last_anchor = None
        for elem_id in _presentation_order(tree):
            node = nodes.get(elem_id)
            if node is None:
                continue
            try:
                if getattr(node, 'is_abstract', False):
                    continue
                concept = elem_id.replace('_', ':', 1).split(':')[-1]
                if not concept:
                    continue
                if concept.endswith(_AUTO_SKIP_SUFFIXES):
                    continue
                if any(s in concept for s in _AUTO_SKIP_SUBSTR):
                    continue

                # Exact-mapped concept: normally nothing to learn (it
                # anchors later learned lines) -- UNLESS the company never
                # presents it on the statement our static map claims. Then
                # the filing's own linkbase wins (linkbase-first inversion)
                # and the concept is re-learned as an as-filed line in the
                # category the company actually uses.
                _exact = _CONCEPT_TAG_TO_LABEL.get(concept)
                if _exact is not None:
                    _einfo = CONCEPT_MAP.get(_exact)
                    _ecat = _einfo.get('cat') if isinstance(_einfo, dict) else None
                    if (isinstance(_einfo, dict) and _ecat == cat
                            and not _einfo.get('auto')):
                        last_anchor = _exact
                        continue
                    _face_cats = _FACE_PRESENTED.get(concept, set())
                    # Direction whitelist: never displace an IS/BS mapping
                    # because the concept shows up on the cash-flow face --
                    # balance elements appear there as reconciliation
                    # footers and expense elements as add-backs. Only a
                    # static *CF* claim contradicted by IS/BS presentation,
                    # or an IS<->BS contradiction, is genuine evidence.
                    _dir_ok = (cat != '3_Cash_Flow')
                    # Twin guard: a curated label in the presented category
                    # already carrying this concept means the concept is
                    # correctly represented there (e.g. 'Net Income' on the
                    # IS vs 'Net Income (CF)' on the CF).
                    if _dir_ok:
                        for _l2, _i2 in list(CONCEPT_MAP.items()):
                            if (isinstance(_i2, dict) and _i2.get('cat') == cat
                                    and not _i2.get('auto')
                                    and concept in (_i2.get('tags') or [])):
                                _dir_ok = False
                                last_anchor = _l2
                                break
                    # Primary-tag guard: only stray aliases may be displaced.
                    if _dir_ok and isinstance(_einfo, dict):
                        try:
                            if concept in (_einfo.get('tags') or [])[:2]:
                                _dir_ok = False
                        except Exception:
                            pass
                    if (_dir_ok and _ecat and _ecat != cat
                            and _ecat in ('1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow')
                            and cat in ('1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow')
                            and _ecat not in _face_cats
                            and not _einfo.get('auto')
                            and concept not in _RESOLVE_OVERRIDDEN):
                        _calc_cat = _calc_section_of(concept)
                        if _calc_cat is None or _calc_cat == cat:
                            _RESOLVE_OVERRIDDEN.add(concept)
                            try:
                                if concept in (_einfo.get('tags') or []):
                                    _einfo['tags'].remove(concept)
                            except Exception:
                                pass
                            print(f"  [Resolve] Linkbase override: '{concept}' is presented on "
                                  f"{_CAT_SUFFIX.get(cat, cat)} but statically mapped to "
                                  f"{_CAT_SUFFIX.get(_ecat, _ecat)} ('{_exact}'); following the filing.")
                            # fall through to learn it as an as-filed row
                        else:
                            continue
                    else:
                        continue

                if concept in _AUTO_SEEN_CONCEPTS:
                    continue
                # Mark seen early so failures aren't retried every filing.
                _AUTO_SEEN_CONCEPTS.add(concept)

                if concept in STANDARD_TAG_MAP and concept not in _RESOLVE_OVERRIDDEN:
                    continue
                # Fuzzy-resolvable: the pipeline captures it; also a valid anchor.
                _fz = None if concept in _RESOLVE_OVERRIDDEN else resolve_concept_to_label(concept)
                if _fz is not None:
                    _finfo = CONCEPT_MAP.get(_fz)
                    if (isinstance(_finfo, dict) and _finfo.get('cat') == cat
                            and not _finfo.get('auto')):
                        last_anchor = _fz
                    continue
                if not _is_numericish_element(xbrl, elem_id, concept):
                    continue

                # Human-readable label: filer's standard label, else the
                # presentation display label, else a prettified tag name.
                label = (getattr(node, 'standard_label', None)
                         or getattr(node, 'display_label', None)
                         or clean_name(concept) or '').strip()
                label = label.replace(' [Member]', '').strip().rstrip(':').strip()
                if len(label) < 3:
                    continue
                if _is_learning_noise(concept, label):
                    continue
                # ' - ' is the segment-label delimiter throughout the pipeline;
                # keep learned core labels unambiguous.
                label = label.replace(' - ', ' \u2013 ')

                with _CONCEPT_LEARN_LOCK:
                    # Re-check under the lock (another worker may have won).
                    if concept in _CONCEPT_TAG_TO_LABEL and concept not in _RESOLVE_OVERRIDDEN:
                        continue

                    norm = _normalize_label_key(label)
                    existing = _NORMALIZED_LABEL_INDEX.get(norm)

                    if existing is not None and CONCEPT_MAP.get(existing, {}).get('cat') == cat:
                        # Same line name, same statement -> alias of an
                        # existing row.  Appended last = lowest TagRank
                        # priority, so curated tags still win dedup.
                        if concept not in CONCEPT_MAP[existing]['tags']:
                            CONCEPT_MAP[existing]['tags'].append(concept)
                        final_label = existing
                        print(f"  [FaceStmt] '{concept}' merged into existing '{existing}'")
                    else:
                        final_label = label
                        if final_label in CONCEPT_MAP:
                            # Name collision with a different statement's row
                            # (e.g. 'Net Income' on the cash-flow statement).
                            final_label = f"{label} ({_CAT_SUFFIX.get(cat, cat)})"
                        if final_label in CONCEPT_MAP:
                            if concept not in CONCEPT_MAP[final_label]['tags']:
                                CONCEPT_MAP[final_label]['tags'].append(concept)
                        else:
                            _akey = (cat, last_anchor)
                            _seq = _ANCHOR_SEQ.get(_akey, 0)
                            _ANCHOR_SEQ[_akey] = _seq + 1
                            CONCEPT_MAP[final_label] = {
                                'tags': [concept], 'cat': cat, 'auto': True,
                                'anchor': last_anchor, 'anchor_seq': _seq,
                            }
                            _NORMALIZED_LABEL_INDEX.setdefault(norm, final_label)
                            print(f"  [FaceStmt] Learned new line '{final_label}' "
                                  f"<- '{concept}' ({_CAT_SUFFIX.get(cat, cat)}, "
                                  f"after '{last_anchor}')")

                    _CONCEPT_TAG_TO_LABEL[concept] = final_label
                    _FUZZY_CACHE[concept] = final_label
                    learned += 1
            except Exception:
                continue

    return learned


def resolve_custom_tags(calc_trees):
    """
    Parses the calculation linkbase to find custom extension tags that roll up 
    into standard Cash Flow / Income Statement totals, and maps them to our CONCEPT_MAP
    based on naming heuristics.
    """
    if not calc_trees: return
    
    investing_parents = {'NetCashProvidedByUsedInInvestingActivities','NetCashProvidedByUsedInInvestingActivitiesContinuingOperations','PaymentsToAcquirePropertyPlantAndEquipment','PaymentsToAcquireProductiveAssets','PurchaseOfPropertyPlantAndEquipment','PaymentsToAcquirePropertyAndEquipment','PurchasesOfPropertyAndEquipment'}
    financing_parents = {'NetCashProvidedByUsedInFinancingActivities', 'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations'}
    operating_parents = {'NetCashProvidedByUsedInOperatingActivities', 'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations'}
    
    for role_uri, tree in calc_trees.items():
        if hasattr(tree, 'all_nodes'):
            for elem_id, node in tree.all_nodes.items():
                parent_id = node.parent
                if not parent_id: continue
                parent_clean = parent_id.replace('_', ':', 1).split(':')[-1] if '_' in parent_id else parent_id.split(':')[-1]
                child_ns = elem_id.split('_')[0] if '_' in elem_id else ''
                child_clean = elem_id.replace('_', ':', 1).split(':')[-1] if '_' in elem_id else elem_id.split(':')[-1]
                
                # Only look at custom tags
                if child_ns != 'us-gaap':
                    lbl_lower = child_clean.lower()
                    target_label = None
                    
                    if parent_clean in investing_parents:
                        if 'property' in lbl_lower or 'equipment' in lbl_lower or 'capital' in lbl_lower:
                            if node.weight < 0:
                                target_label = 'Capital Expenditures'
                            else:
                                target_label = 'Proceeds from Asset Sales'
                        elif 'acquisition' in lbl_lower or 'business' in lbl_lower or 'divest' in lbl_lower:
                            if node.weight < 0:
                                target_label = 'Acquisitions'
                            else:
                                target_label = 'Divestitures'
                        elif 'investment' in lbl_lower or 'securit' in lbl_lower:
                            if node.weight > 0: # Proceeds are positive to investing
                                target_label = 'Proceeds from Investments'
                            else: # Purchases are negative to investing
                                target_label = 'Purchases of Investments'
                    elif parent_clean in ('InterestIncomeExpenseNet', 'InterestIncomeExpenseOperatingNet', 'InterestIncomeExpenseAfterProvisionForLoanLoss', 'InterestIncomeExpenseNonoperatingNet'):
                        if node.weight > 0:
                            target_label = 'Interest Income'
                        else:
                            target_label = 'Interest Expense'
                                
                    if target_label and child_clean not in CONCEPT_MAP[target_label]['tags']:
                        print(f"  [Linkbase] Auto-mapped custom tag '{child_clean}' -> '{target_label}'")
                        CONCEPT_MAP[target_label]['tags'].append(child_clean)
                        # Register so the concept-learner treats it as
                        # exact-mapped and does not spawn a duplicate face row.
                        _ann = isinstance(CONCEPT_MAP[target_label], dict) and CONCEPT_MAP[target_label].get('auto')
                        if not _ann:
                            _CONCEPT_TAG_TO_LABEL.setdefault(child_clean, target_label)

# Geographic member classifier shared by XBRL, HTML, and final routing.
#
# A member is geographic only when its text can be fully explained by:
#   * one or more country/region names;
#   * geographic connectors such as "including" or "other countries"; and
#   * harmless accounting/display words such as "revenue" or "segment".
# Any remaining business word ("Gaming", "Cloud Services", "Government")
# makes the result non-geographic. This keeps the fix generic and prevents
# country words inside real business-segment names from being reclassified.
_GEO_MEMBER_STOPWORDS = frozenset({
    'revenue', 'revenues', 'sale', 'sales', 'fee', 'fees', 'income', 'loss',
    'profit', 'earnings', 'margin', 'expense', 'expenses', 'cost', 'costs',
    'asset', 'assets', 'liability', 'liabilities', 'debt', 'equity', 'cash',
    'flow', 'flows', 'operations', 'operating', 'net', 'gross', 'total',
    'segment', 'segments', 'member', 'group', 'company', 'corp', 'inc',
    'ltd', 'llc', 'plc', 'division', 'subsidiary', 'consolidated',
})
_GEO_CONNECTOR_WORDS = frozenset({
    'and', 'or', 'of', 'for', 'the', 'in', 'at', 'by', 'from',
    'including', 'includes', 'included', 'excluding', 'excludes', 'excluded',
    'except', 'plus', 'with', 'without', 'other', 'all', 'rest', 'remainder',
    'greater', 'mainland', 'not', 'separately', 'disclosed', 'countries',
    'country', 'regions', 'region', 'territories', 'territory', 'locations',
    'location', 'international', 'domestic', 'foreign', 'worldwide',
})
_GEO_MULTI_SCOPE_WORDS = frozenset({
    'including', 'excluding', 'and', 'or', 'other', 'all', 'rest',
    'remainder', 'greater', 'mainland', 'countries', 'regions',
    'territories', 'international', 'worldwide',
})
# Longest first so "papua new guinea" is consumed before "guinea" and
# "united states" before a shorter overlapping phrase.
_GEO_COUNTRY_PATTERNS = tuple(
    (country, re.compile(r'(?<![a-z])' + re.escape(country) + r'(?![a-z])', re.I))
    for country in sorted(GEOGRAPHIC_COUNTRIES, key=lambda value: (-len(value), value))
)


def _geo_clean_member_text(label: str) -> str:
    text = html_lib.unescape(str(label or ''))
    text = re.sub(r'\s+', ' ', text).strip()
    return text


@lru_cache(maxsize=65536, typed=True)
def _classify_geographic_member(label: str, is_geographic_axis: bool = False):
    """Return ``country``, ``region``, or ``None`` for a display member.

    Full country names are strong evidence even without a geographic axis.
    ISO alpha-2 codes remain weak and are accepted only on a geographic axis.
    Composite labels such as ``China (including Hong Kong)`` are regions,
    because they combine more than one jurisdiction and should not be emitted
    as a single country row.
    """
    original = _geo_clean_member_text(label)
    if not original:
        return None
    lowered = original.casefold().strip(' .,:;()[]{}')
    if lowered in GEOGRAPHIC_COUNTRIES:
        return 'country'
    if is_geographic_axis and lowered in GEOGRAPHIC_CODES:
        return 'country'

    # Standardize separators before token analysis. Parentheses are structural
    # punctuation in labels such as "China (including Hong Kong)", not a
    # signal that the member is non-geographic.
    working = lowered.replace('&', ' and ')
    working = re.sub(r'[/|,+;:()\[\]{}]', ' ', working)
    working = re.sub(r'\s+', ' ', working).strip()

    countries = []
    country_stripped = working
    for country, pattern in _GEO_COUNTRY_PATTERNS:
        if pattern.search(country_stripped):
            countries.append(country)
            country_stripped = pattern.sub(' ', country_stripped)

    had_region_keyword = bool(_REGION_KEYWORDS_RE.search(working))
    country_stripped = _REGION_KEYWORDS_RE.sub(' ', country_stripped)
    words = re.findall(r'[a-z]+', country_stripped)
    leftovers = [
        word for word in words
        if word not in _GEO_MEMBER_STOPWORDS
        and word not in _GEO_CONNECTOR_WORDS
    ]
    if leftovers:
        return None

    # A label must contain positive geographic evidence. Connector-only or
    # accounting-only labels such as "Other Revenue" are not geography.
    generic_geo_phrase = bool(re.search(
        r'\b(?:other|all|rest|remainder)\s+(?:countries|regions|territories)\b',
        working,
    ))
    if not countries and not had_region_keyword and not generic_geo_phrase:
        return None

    tokens = set(re.findall(r'[a-z]+', working))
    is_composite = (
        len(set(countries)) > 1
        or had_region_keyword
        or generic_geo_phrase
        or bool(tokens & _GEO_MULTI_SCOPE_WORDS)
    )
    return 'region' if is_composite else 'country'


# Geographic alias/basis consolidation v2.
_GEO_DISPLAY_ALIASES = {
    'us': 'United States', 'u s': 'United States', 'usa': 'United States',
    'united states of america': 'United States',
    'tw': 'Taiwan', 'sg': 'Singapore', 'cn': 'China', 'hk': 'Hong Kong',
    'uk': 'United Kingdom', 'u k': 'United Kingdom',
    'all other countries': 'Other', 'other countries': 'Other',
    'all other countries not separately disclosed': 'Other',
    'other countries not separately disclosed': 'Other',
}
_GEO_EXPLICIT_BASES = (
    'Customer Billing Location',
    'Customer Headquarters Location',
    'Shipment Destination',
)


def _geo_alias_key(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').casefold()).strip()


def _split_geographic_basis_suffix(member: str):
    text = _geo_clean_member_text(member)
    for basis in _GEO_EXPLICIT_BASES:
        suffix = f' ({basis})'
        if text.casefold().endswith(suffix.casefold()):
            return text[:-len(suffix)].strip(), basis
    return text, None


def _canonical_geographic_member_label(label: str, is_geographic_axis: bool = False) -> str:
    """Canonicalize proven geography and common ISO/display aliases."""
    text = _geo_clean_member_text(label)
    base_text, basis = _split_geographic_basis_suffix(text)
    alias = _GEO_DISPLAY_ALIASES.get(_geo_alias_key(base_text))
    if alias:
        canonical = alias
    else:
        if _classify_geographic_member(base_text, is_geographic_axis) is None:
            return text
        canonical = base_text.replace('&', ' And ')
        canonical = re.sub(r'[/|,+;:()\[\]{}]', ' ', canonical)
        canonical = re.sub(r'\s+', ' ', canonical).strip()
        for word in (
            'including', 'excluding', 'except', 'and', 'or', 'other', 'all',
            'rest', 'remainder', 'greater', 'mainland', 'not', 'separately',
            'disclosed', 'countries', 'country', 'regions', 'region',
            'territories', 'territory',
        ):
            canonical = re.sub(rf'\b{word}\b', word.title(), canonical, flags=re.I)
    return f'{canonical} ({basis})' if basis else canonical


@lru_cache(maxsize=65536, typed=True)
def _is_pure_geographic_label(label: str, is_geographic_axis: bool = False) -> bool:
    return _classify_geographic_member(label, is_geographic_axis) is not None

# ---------------------------------------------------------------------------
# Generic table-structured business-breakdown rescue
# ---------------------------------------------------------------------------
# Some issuers disclose recurring business/segment amounts only inside HTML
# TextBlock tables rather than as individually tagged XBRL facts.  This rescue
# is deliberately conservative and additive:
#   * no ticker, industry, or metric-name registry;
#   * exact current-period column binding from table headers;
#   * monetary tables only;
#   * source row labels are preserved;
#   * at least two coherent rows are required;
#   * existing XBRL facts always win (exact-key skip + TagRank 998).
# It is not an operating-metrics extractor and does not scan free-form prose.

_GENERIC_BUSINESS_POSITIVE_CONTEXT_RE = tuple(re.compile(p, re.I) for p in (
    r'\brevenues?\s+by\s+(?:business|segment|product|service|market|geograph)',
    r'\bnet\s+sales\s+(?:information|by\s+groups?|by\s+segment)',
    r'\breportable\s+segment\s+financial\s+information',
    r'\boperating\s+income\s*\(loss\)?\s+by\s+segment',
    r'\bsummary\s+of\s+.{0,120}\b(?:business|segment|product|service)\b',
))

_GENERIC_BUSINESS_DRIVER_CONTEXT_RE = tuple(re.compile(p, re.I) for p in (
    r'\bgrowth\s+trends?\s+in\s+(?:our|the)\s+key\s+drivers\b',
    r'\bsummary\s+of\s+(?:our|the)\s+key\s+(?:metrics|drivers)\b',
    r'\bkey\s+(?:operating|performance)\s+(?:metrics|drivers)\b',
))

_GENERIC_BUSINESS_NEGATIVE_CONTEXT_RE = tuple(re.compile(p, re.I) for p in (
    r'\bother\s+income\s*\(expense\)',
    r'\bcash\s+flows?\b',
    r'\bgeneral\s+and\s+administrative\s+expenses?\b',
    r'\bincome\s+tax(?:es)?\b',
    r'\bfair\s+value\b',
    r'\bmedical\s+costs?\s+payable\b',
    r'\bdividends?\b',
    r'\blease\s+cost\b',
))

_GENERIC_BUSINESS_EXCLUDED_HEADER_TERMS = (
    'increase', 'decrease', 'change', 'growth', 'currency-neutral',
    'currency neutral', 'local currency', 'as reported', 'variance', '%',
    'percentage', 'margin', 'mix', 'rate',
)

_GENERIC_BUSINESS_STOP_SECTION_TERMS = (
    'year-over-year percentage', 'net sales mix', 'growth (decline)',
    'operating margin', 'increase/(decrease)', 'increase / (decrease)',
)

_GENERIC_BUSINESS_CORE_ACCOUNTING_LABELS = (
    'investment income', 'interest expense', 'other income',
    'net cash provided', 'net cash used', 'income tax', 'cash flow',
    'accounts payable', 'accounts receivable', 'depreciation',
    'general and administrative',
)

_GENERIC_BUSINESS_NON_MONETARY_LABEL_RE = re.compile(
    r'\b(?:margin|percentage|percent|rate|ratio|per\s+share|eps|headcount|'
    r'employees?|monthly\s+active|active\s+users?|users?|trips?|nights?\s+booked|'
    r'units?\s+(?:sold|shipped)|subscribers?|customer\s+count)\b',
    re.I,
)

_HTML_BUSINESS_CONCEPT = 'HTMLBusinessBreakdown'


def _gb_clean_text(value) -> str:
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    return re.sub(r'\s+', ' ', str(value)).strip()


def _gb_flatten_column(column) -> str:
    if isinstance(column, tuple):
        parts = []
        for part in column:
            text = _gb_clean_text(part)
            if not text or text.lower().startswith('unnamed:'):
                continue
            if not parts or parts[-1] != text:
                parts.append(text)
        return ' | '.join(parts)
    return _gb_clean_text(column)


def _gb_parse_number(value):
    raw = _gb_clean_text(value)
    if not raw or raw.lower() in {'nan', 'nm', 'n/m', 'â€”', '-', '--'}:
        return None, raw
    if '%' in raw:
        return None, raw
    negative = raw.startswith('(') and raw.endswith(')')
    normalized = (raw.replace('$', '').replace('â‚¬', '').replace('Â£', '')
                  .replace(',', '').replace('(', '').replace(')', ''))
    normalized = re.sub(r'(?<=\d)[a-zA-Z]+$', '', normalized)
    normalized = re.sub(r'[^0-9.\-]', '', normalized)
    if not normalized or normalized in {'-', '.', '-.'}:
        return None, raw
    try:
        parsed = float(normalized)
    except (TypeError, ValueError):
        return None, raw
    return (-abs(parsed) if negative else parsed), raw


def _gb_is_decorator_column(series: pd.Series) -> bool:
    values = [_gb_clean_text(v) for v in series.tolist() if _gb_clean_text(v)]
    if not values:
        return True
    return all(value in {'$', '%', 'â‚¬', 'Â£'} for value in values)


def _gb_row_unit_evidence(row, positions) -> str:
    """Return money/percent/unknown evidence for one displayed period row."""
    texts = []
    for position in positions:
        if position < len(row):
            text = _gb_clean_text(row[position])
            if text:
                texts.append(text)
        # pandas may deduplicate identical SEC period headers ("2026",
        # "2026.1") even though the first cell is only a currency decorator.
        # Inspect immediate decorator neighbors, but never borrow arbitrary
        # values from growth/comparison columns elsewhere in the row.
        for neighbor in (position - 1, position + 1):
            if 0 <= neighbor < len(row):
                text = _gb_clean_text(row[neighbor])
                if text in {'$', '%', 'â‚¬', 'Â£'}:
                    texts.append(text)
    if any('%' in text for text in texts):
        return 'percent'
    if any(any(symbol in text for symbol in ('$', 'â‚¬', 'Â£')) for text in texts):
        return 'money'
    return 'unknown'


def _gb_table_fingerprint(table: pd.DataFrame) -> str:
    """Stable content identity used to detect conflicting parses of one table."""
    if table is None or table.empty:
        return ''
    parts = ['\x1f'.join(_gb_flatten_column(column) for column in table.columns)]
    for row in table.itertuples(index=False, name=None):
        parts.append('\x1f'.join(_gb_clean_text(value) for value in row))
    return hashlib.sha256('\x1e'.join(parts).encode('utf-8', 'replace')).hexdigest()


def _gb_normalize_table_grid(table: pd.DataFrame, current_year: int) -> pd.DataFrame:
    """Normalize proper headers and SEC td-only flattened table grids."""
    if table is None or table.empty:
        return pd.DataFrame()
    df = table.copy().dropna(how='all', axis=0).dropna(how='all', axis=1)
    if df.empty:
        return df

    # Retain decorator columns.  SEC grids commonly store "$" or "%" in a
    # separate cell under the same period header.  Value selection coalesces
    # duplicate period columns safely, while retaining these cells preserves
    # the row-level unit evidence needed to distinguish money from counts and
    # percentages.
    year_pattern = re.compile(rf'(?<!\d){int(current_year)}(?!\d)')
    if any(year_pattern.search(_gb_flatten_column(column)) for column in df.columns):
        return df

    # pd.read_html(header=None) is common for SEC tables whose header cells are
    # rendered as ordinary <td> elements.  Promote the rows before the first
    # real data row into synthetic column headers.
    data_start = None
    for position in range(min(len(df), 10)):
        row = df.iloc[position].tolist()
        first = _gb_clean_text(row[0]) if row else ''
        numeric_count = sum(_gb_parse_number(value)[0] is not None for value in row[1:])
        if not first or numeric_count < 2:
            continue
        if re.fullmatch(r'[\d\s/().-]+', first):
            continue
        if any(term in first.lower() for term in (
                'three months ended', 'six months ended', 'nine months ended',
                'year ended', 'increase', 'decrease', 'in millions',
                'in thousands', 'in billions')):
            continue
        data_start = position
        break
    if data_start is None or data_start == 0:
        return df

    new_columns = []
    for column_position in range(len(df.columns)):
        parts = []
        for row_position in range(data_start):
            text = _gb_clean_text(df.iloc[row_position, column_position])
            if not text or text in {'$', '%', 'â‚¬', 'Â£'}:
                continue
            if not parts or parts[-1] != text:
                parts.append(text)
        new_columns.append(' | '.join(parts) or f'column_{column_position}')
    df = df.iloc[data_start:].copy()
    df.columns = new_columns
    return df


def _gb_detect_scale(context: str, table: pd.DataFrame) -> float:
    corpus = str(context or '') + ' ' + ' '.join(
        _gb_flatten_column(column) for column in table.columns)
    for row in table.head(5).itertuples(index=False, name=None):
        corpus += ' ' + ' '.join(_gb_clean_text(value) for value in row)
    lowered = corpus.lower()
    if re.search(r'\bin\s+billions?\b', lowered):
        return 1e9
    if re.search(r'\bin\s+millions?\b', lowered):
        return 1e6
    if re.search(r'\bin\s+thousands?\b', lowered):
        return 1e3
    return 1.0


def _gb_has_money_evidence(context: str, table: pd.DataFrame) -> bool:
    corpus = str(context or '') + ' ' + ' '.join(
        _gb_flatten_column(column) for column in table.columns)
    for row in table.head(20).itertuples(index=False, name=None):
        corpus += ' ' + ' '.join(_gb_clean_text(value) for value in row)
    lowered = corpus.lower()
    if any(symbol in corpus for symbol in ('$', 'â‚¬', 'Â£')):
        return True
    return any(term in lowered for term in (
        'revenue', 'net sales', 'operating income',
        'earnings from operations'))


def _gb_has_positive_context(context: str) -> bool:
    normalized = _gb_clean_text(context)
    return (
        any(pattern.search(normalized)
            for pattern in _GENERIC_BUSINESS_POSITIVE_CONTEXT_RE)
        or any(pattern.search(normalized)
               for pattern in _GENERIC_BUSINESS_DRIVER_CONTEXT_RE)
    )


def _gb_geographic_basis_from_context(context: str):
    """Return the strongest explicit geographic reporting basis in context.

    Customer headquarters and customer billing are accounting/reporting bases;
    shipment language is often only explanatory narrative near the same table.
    Therefore an explicit customer basis outranks incidental shipment text.
    Shipment Destination is returned only when no customer-location basis is
    present. This is generic and fail-closed for unrelated prose.
    """
    raw_context = _gb_clean_text(context)
    raw_context = re.sub(r'(?<=[a-z0-9])(?=[A-Z])', ' ', raw_context)
    lowered = raw_context.replace('_', ' ').replace(':', ' ').casefold()
    headquarters = bool(re.search(
        r'\bcustomer(?:s|\'s)?\s+(?:headquarters|headquarter)(?:\s+location)?\b|'
        r'\blocation\s+of\s+(?:the\s+)?customer(?:s|\'s)?\s+(?:headquarters|headquarter)\b|'
        r'\bcustomer(?:s)?\s+(?:are|is)\s+headquartered\b|'
        r'\bheadquarters(?:\s+location)?\s+of\s+(?:the\s+)?customer(?:s)?\b',
        lowered,
    ))
    if headquarters:
        return 'Customer Headquarters Location'

    billing = bool(re.search(
        r'\bcustomer(?:s|\'s)?\s+billing(?:\s+location)?\b|'
        r'\bbilling\s+location(?:s)?\b|'
        r'\blocation\s+where\s+(?:the\s+)?customer(?:s)?\s+(?:is|are)\s+billed\b',
        lowered,
    ))
    if billing:
        return 'Customer Billing Location'

    shipment = bool(re.search(
        r'\bshipment\s+destination(?:s)?\b|\bship-to(?:\s+location)?\b|'
        r'\bdestination\s+to\s+which\s+(?:the\s+)?products?\s+(?:is|are|were)\s+shipped\b',
        lowered,
    ))
    return 'Shipment Destination' if shipment else None


def _gb_context_class(context: str):
    normalized = _gb_clean_text(context).lower()
    if any(pattern.search(normalized)
           for pattern in _GENERIC_BUSINESS_NEGATIVE_CONTEXT_RE):
        return False, None, 0.0
    if not _gb_has_positive_context(normalized):
        return False, None, 0.0
    # Recurring monetary key-driver tables are not necessarily revenue tables.
    # A nearby sentence can even explicitly say that the measures are not net
    # revenue.  Qualify, but otherwise preserve, their source labels instead of
    # letting an incidental mention of "revenue" manufacture a Revenue prefix.
    # The generic prefix also keeps these rows visibly grouped beside revenue
    # while making clear that they are a different type of series.
    if any(pattern.search(normalized)
           for pattern in _GENERIC_BUSINESS_DRIVER_CONTEXT_RE):
        return True, 'Operating Measure', 0.93
    if 'operating income' in normalized or 'earnings from operations' in normalized:
        return True, 'Operating Income', 0.96
    if 'revenue' in normalized or 'net sales' in normalized:
        return True, 'Revenue', 0.98
    # Company-defined recurring monetary-driver table.  Preserve the source
    # row label instead of encoding an industry-specific metric vocabulary.
    return True, None, 0.93


def _gb_period_header_score(header: str, current_year: int,
                            preferred_duration: int) -> float | None:
    lower = str(header or '').lower()
    if not re.search(rf'(?<!\d){int(current_year)}(?!\d)', str(header or '')):
        return None
    if any(term in lower for term in _GENERIC_BUSINESS_EXCLUDED_HEADER_TERMS):
        return None

    score = 100.0
    if preferred_duration <= 120:
        if 'three months' in lower or 'quarter' in lower:
            score += 40.0
        if any(term in lower for term in ('six months', 'nine months', 'year to date', 'ytd')):
            score -= 50.0
        if any(term in lower for term in ('year ended', 'twelve months')):
            score -= 70.0
    else:
        if any(term in lower for term in ('year ended', 'twelve months', 'fiscal year')):
            score += 40.0
        if any(term in lower for term in ('three months', 'quarter', 'six months', 'nine months')):
            score -= 60.0
    return score


def _gb_select_current_value_column(table: pd.DataFrame, current_year: int,
                                    preferred_duration: int):
    candidates = []
    for position, column in enumerate(table.columns):
        header = _gb_flatten_column(column)
        score = _gb_period_header_score(header, current_year, preferred_duration)
        if score is not None:
            series = table.iloc[:, position]
            numeric_count = sum(
                _gb_parse_number(value)[0] is not None
                for value in series.tolist()
            )
            # Prefer the actual numeric cell over a deduplicated "$"/"%"
            # decorator column.  Earlier numeric columns still win true ties,
            # which preserves the QTD-before-YTD safeguard.
            candidates.append((score, numeric_count, -position, position, header))
    if not candidates:
        return None, None
    _, _, _, position, header = max(candidates)
    return position, header


def _gb_equivalent_period_columns(table: pd.DataFrame, selected_position: int,
                                  selected_header: str):
    """Return duplicate columns representing the same displayed period.

    Flattened SEC grids frequently repeat a header over separate currency,
    value, and spacer cells.  Only exact normalized-header matches are grouped;
    differing analytical or comparison headers remain separate.
    """
    positions = []
    for position, column in enumerate(table.columns):
        if _gb_flatten_column(column) == selected_header:
            positions.append(position)
    return positions or [selected_position]


def _gb_coalesced_period_value(row, positions):
    """Read one unambiguous number from duplicate same-period cells."""
    parsed = []
    for position in positions:
        if position >= len(row):
            continue
        value, raw = _gb_parse_number(row[position])
        if value is not None:
            parsed.append((float(value), raw))
    if not parsed:
        return None, ''
    distinct = {value for value, _ in parsed}
    if len(distinct) != 1:
        # Exact same-period headers carrying different values are structurally
        # ambiguous.  Fail closed rather than choosing a positional guess.
        return None, ''
    return parsed[0]


def _gb_duration_from_header(header: str, fallback: int) -> int:
    lower = str(header or '').lower()
    if 'three months' in lower or 'quarter' in lower:
        return 90
    if 'six months' in lower:
        return 180
    if 'nine months' in lower:
        return 270
    if any(term in lower for term in ('year ended', 'twelve months', 'fiscal year')):
        return 365
    return int(fallback)


def _gb_row_label(value) -> str:
    text = _gb_clean_text(value)
    text = re.sub(r'\s*\(\d+\)\s*$', '', text).strip()
    text = re.sub(r'\s+[Â¹Â²Â³â´âµâ¶â·â¸â¹]+$', '', text).strip()
    return text


def _gb_section_from_header(label: str):
    normalized = str(label or '').lower().rstrip(':').strip()
    if normalized in {'revenue', 'revenues', 'net sales'}:
        return True, 'Revenue'
    if normalized in {
            'earnings from operations', 'operating income',
            'operating income (loss)', 'operating profit'}:
        return True, 'Operating Income'
    if any(term in normalized for term in _GENERIC_BUSINESS_STOP_SECTION_TERMS):
        return False, None
    return None


def _extract_generic_business_breakdown_from_table(
        table: pd.DataFrame, context: str, current_year: int,
        preferred_duration: int = 90):
    """Return conservative source-backed business-table candidates.

    This pure helper is intentionally unit-testable without SEC/network access.
    """
    eligible, default_prefix, base_confidence = _gb_context_class(context)
    if not eligible or table is None or table.empty or len(table.columns) < 2:
        return []
    if not _gb_has_money_evidence(context, table):
        return []

    normalized = _gb_normalize_table_grid(table, current_year)
    if normalized.empty or len(normalized.columns) < 2:
        return []
    value_position, value_header = _gb_select_current_value_column(
        normalized, current_year, preferred_duration)
    if value_position is None:
        return []
    value_positions = _gb_equivalent_period_columns(
        normalized, value_position, value_header)

    duration = _gb_duration_from_header(value_header, preferred_duration)
    # Detect unit scale from the original grid because td-only normalization
    # removes header/unit rows before candidate extraction.
    scale = _gb_detect_scale(context, table)
    period_has_money = any(
        _gb_row_unit_evidence(row, value_positions) == 'money'
        for row in normalized.itertuples(index=False, name=None)
    )
    active = True
    active_prefix = default_prefix
    candidates = []
    seen_labels = set()

    for row in normalized.itertuples(index=False, name=None):
        label = _gb_row_label(row[0])
        if not label:
            continue
        lowered = label.lower()
        numeric_count = sum(_gb_parse_number(value)[0] is not None for value in row[1:])
        if numeric_count == 0:
            section_change = _gb_section_from_header(label)
            if section_change is not None:
                active, active_prefix = section_change
            elif lowered.endswith(':'):
                active = False
            continue

        if not active or value_position >= len(row):
            continue
        if (lowered.startswith('total ') or lowered.endswith(' - total')
                or 'consolidated' in lowered):
            continue
        if 'elimination' in lowered:
            continue
        if any(term in lowered for term in _GENERIC_BUSINESS_CORE_ACCOUNTING_LABELS):
            continue

        row_unit = _gb_row_unit_evidence(row, value_positions)
        # A detached percent cell is authoritative even when the adjacent
        # numeric cell contains no "%" character.  Never apply a monetary
        # table scale to that row.
        if row_unit == 'percent':
            continue
        # Recurring key-driver tables frequently mix dollars with counts.  For
        # this broad context class, require currency evidence on the row itself
        # instead of borrowing a "$ in millions" caption from another row.
        non_monetary_label = bool(
            _GENERIC_BUSINESS_NON_MONETARY_LABEL_RE.search(label)
        )
        if active_prefix == 'Operating Measure' and row_unit != 'money':
            # SEC tables often print "$" only on the first monetary row.  A
            # later row may inherit that unit only when the selected period has
            # direct currency evidence elsewhere and the row label is not a
            # count/ratio/usage metric.
            if not period_has_money or non_monetary_label:
                continue
        if row_unit != 'money' and non_monetary_label:
            continue

        value, raw_value = _gb_coalesced_period_value(row, value_positions)
        if value is None:
            continue
        display_label = f'{active_prefix} - {label}' if active_prefix else label
        normalized_label = re.sub(r'\s+', ' ', display_label).strip().casefold()
        if normalized_label in seen_labels:
            continue
        seen_labels.add(normalized_label)
        candidates.append({
            'SourceLabel': label,
            'Label': display_label,
            'Value': float(value) * float(scale),
            'RawValue': raw_value,
            'Scale': float(scale),
            'Prefix': active_prefix,
            'Confidence': float(base_confidence),
            'Duration': int(duration),
            'ValueHeader': value_header,
        })

    # Multiple coherent rows are required.  This is the principal guard against
    # one-off nearby-number matches and unrelated accounting tables.
    return candidates if len(candidates) >= 2 else []


def _gb_nearby_table_context(table_tag, max_chars: int = 1800) -> str:
    """Collect nearby preceding block text without swallowing the whole filing."""
    pieces = []
    seen = set()
    total = 0
    caption = table_tag.find('caption') if hasattr(table_tag, 'find') else None
    if caption is not None:
        text = _gb_clean_text(caption.get_text(' ', strip=True))
        if text:
            pieces.append(text)
            seen.add(text)
            total += len(text)

    try:
        previous = table_tag.find_all_previous(
            ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'div', 'strong', 'b'],
            limit=36)
    except Exception:
        previous = []
    for node in previous:
        try:
            if node.find_parent('table') is not None:
                continue
            # Avoid broad container divs that aggregate unrelated neighboring
            # sections and can contaminate an otherwise local table context.
            if node.name == 'div' and node.find(['div', 'p', 'table']) is not None:
                continue
            text = _gb_clean_text(node.get_text(' ', strip=True))
        except Exception:
            continue
        if not text or text in seen or len(text) > 1200:
            continue
        seen.add(text)
        pieces.append(text)
        total += len(text)
        if total >= max_chars:
            break
    # find_all_previous is nearest-first.  Keep nearest context first because
    # positive/negative gates are based on presence, not narrative ordering.
    return _gb_clean_text(' '.join(pieces))[:max_chars]


def _extract_generic_business_breakdowns_from_textblocks(
        facts_df: pd.DataFrame, filing_form: str, period_end_date,
        ye_month: int, filed, filing_url=None, existing_facts=None,
        filing=None):
    """Extract additive 4a facts from eligible filing HTML tables.

    XBRL TextBlocks remain the first source.  The visible filing HTML is a
    second source because SEC inline-XBRL presentation tables are not required
    to be contained in a TextBlock fact.
    """
    form = str(filing_form or '').upper().strip()
    if form not in {'10-Q', '10-Q/A', '10-K', '10-K/A'}:
        return []
    if facts_df is None or facts_df.empty or not {'concept', 'value'}.issubset(facts_df.columns):
        return []
    end_dt = pd.to_datetime(period_end_date, errors='coerce')
    if pd.isna(end_dt):
        return []

    current_year = int(end_dt.year)
    preferred_duration = 365 if '10-K' in form else 90
    existing_keys = set()
    for fact in existing_facts or ():
        if fact.get('Category') != '4a_Segments_Business':
            continue
        key = (re.sub(r'\s+', ' ', str(fact.get('Label') or '')).strip().casefold(),
               fact.get('FY'), fact.get('Q'))
        existing_keys.add(key)

    text_blocks = facts_df[
        facts_df['concept'].astype(str).str.contains('TextBlock', case=False, na=False)
    ]
    accepted = []
    accepted_keys = set()

    def _accept_candidates(candidates, candidate_end_dt, source_period_role,
                           source_table_fingerprint, source_geographic_basis=None):
        candidate_fy, candidate_q = get_period_info(candidate_end_dt, ye_month)
        candidate_end_text = candidate_end_dt.strftime('%Y-%m-%d')
        for candidate in candidates:
            label_key = re.sub(r'\s+', ' ', candidate['Label']).strip().casefold()
            fact_key = (label_key, candidate_fy, candidate_q)
            if fact_key in existing_keys or fact_key in accepted_keys:
                continue
            duration = int(candidate.get('Duration') or preferred_duration)
            start_dt = candidate_end_dt - pd.Timedelta(days=duration)
            accepted.append({
                'Category': '4a_Segments_Business',
                'Label': candidate['Label'],
                'Value': candidate['Value'],
                'FY': candidate_fy,
                'Q': candidate_q,
                'End': candidate_end_text,
                'Start': start_dt.strftime('%Y-%m-%d'),
                'Duration': duration,
                'Filed': filed,
                'TagRank': 998,
                'DimCount': 0,
                'IsCalculated': False,
                'Concept': 'HTMLBusinessBreakdown',
                'FilingUrl': filing_url,
                'SourceLabel': candidate.get('SourceLabel'),
                'SourceValueHeader': candidate.get('ValueHeader'),
                'SourceConfidence': candidate.get('Confidence'),
                'SourcePeriodRole': source_period_role,
                'SourceTableFingerprint': source_table_fingerprint,
                'SourceGeographicBasis': source_geographic_basis,
            })
            accepted_keys.add(fact_key)

    def _accept_table(table, context):
        table_fingerprint = _gb_table_fingerprint(table)
        candidates = _extract_generic_business_breakdown_from_table(
            table, context, current_year,
            preferred_duration=preferred_duration)
        if not candidates:
            return
        geographic_basis = _gb_geographic_basis_from_context(context)
        _accept_candidates(
            candidates, end_dt, 'current', table_fingerprint, geographic_basis)

        # Comparative columns carry an explicit prior-year header and the same
        # quarter/annual duration.  They safely backfill an earlier period when
        # that period's own filing table is absent or malformed.  Never use a
        # comparative column unless the same table first proved its current-
        # year column; this prevents a stale historical table from being
        # relabeled as a newer period.
        comparative_year = current_year - 1
        comparative_candidates = _extract_generic_business_breakdown_from_table(
            table, context, comparative_year,
            preferred_duration=preferred_duration)
        if comparative_candidates:
            comparative_end_dt = end_dt - pd.DateOffset(years=1)
            _accept_candidates(
                comparative_candidates, comparative_end_dt, 'comparative',
                table_fingerprint, geographic_basis)

    for raw_value in text_blocks['value'].to_numpy(copy=False):
        html_text = str(raw_value or '')
        if '<table' not in html_text.lower():
            continue
        # Cheap source-level gate before BeautifulSoup/pandas work.
        plain_preview = _gb_clean_text(re.sub(r'<[^>]+>', ' ', html_lib.unescape(html_text)))
        if not _gb_has_positive_context(plain_preview):
            continue
        try:
            soup = BeautifulSoup(html_text, 'html.parser')
        except Exception:
            continue
        for table_tag in soup.find_all('table'):
            context = _gb_nearby_table_context(table_tag)
            if not context:
                context = plain_preview[:1800]
            try:
                parsed_tables = pd.read_html(StringIO(str(table_tag)))
            except Exception:
                continue
            for table in parsed_tables[:1]:
                _accept_table(table, context)

    # Inline-XBRL presentation tables can live directly in the filing HTML
    # without being wrapped by an XBRL TextBlock fact.  Scan only tables whose
    # local context passes the same conservative gate, and parse at most the
    # first DataFrame produced for each HTML table.
    if filing is not None:
        try:
            filing_soup = BeautifulSoup(fetch_html(filing), 'html.parser')
            for hidden in filing_soup.find_all(style=_DISPLAY_NONE_RE):
                hidden.decompose()
            for table_tag in filing_soup.find_all('table'):
                context = _gb_nearby_table_context(table_tag)
                if not _gb_has_positive_context(context):
                    continue
                try:
                    parsed_tables = pd.read_html(StringIO(str(table_tag)))
                except Exception:
                    continue
                for table in parsed_tables[:1]:
                    _accept_table(table, context)
        except Exception as filing_html_error:
            _debug_print(
                f"  [Business Table] Full filing HTML scan skipped "
                f"({type(filing_html_error).__name__}: {filing_html_error})")

    if accepted:
        print(f"  [Business Table] Recovered {len(accepted)} source-bound business breakdown fact(s) from filing HTML tables.")
    return accepted


def _gb_fact_period_key(fact):
    """Period identity that keeps annual, quarterly, and YTD facts separate."""
    try:
        duration = int(float(fact.get('Duration') or 0))
    except (TypeError, ValueError):
        duration = 0
    if duration >= 300:
        scope = 'FY'
    elif duration >= 220:
        scope = 'YTD9'
    elif duration >= 150:
        scope = 'YTD6'
    elif duration > 0:
        scope = 'Q'
    else:
        scope = 'UNKNOWN'
    return (fact.get('FY'), str(fact.get('Q') or ''), scope)


def _gb_close_values(left, right) -> bool:
    try:
        left = float(left)
        right = float(right)
    except (TypeError, ValueError):
        return False
    if not (np.isfinite(left) and np.isfinite(right)):
        return False
    return abs(left - right) <= max(1e-6, 1e-9 * max(abs(left), abs(right), 1.0))


def _gb_clean_member_footnote(member: str) -> str:
    text = re.sub(r'\s+', ' ', str(member or '')).strip()
    # SEC table footnote markers are often flattened into labels as a trailing
    # digit ("International1").  Strip only digits attached to a word; retain
    # legitimate standalone numbered business names.
    return re.sub(r'(?<=[A-Za-z])\d+$', '', text).strip()


def _gb_normalized_display_parts(label: str):
    metric, member = _split_segment_display_label(label)
    member = _gb_clean_member_footnote(member)
    return (_normalize_label_key(metric), _normalize_label_key(member))


def _gb_route_html_business_fact(fact: dict) -> dict:
    """Route proven geographic HTML revenue members before segment cleanup."""
    routed = dict(fact)
    metric, member = _split_segment_display_label(routed.get('Label'))
    member = _gb_clean_member_footnote(member)
    if not member or _normalize_label_key(metric) != 'revenue':
        return routed
    geo_kind = _classify_geographic_member(member, False)
    if geo_kind is None:
        return routed
    member = _canonical_geographic_member_label(member, False)
    routed['Category'] = (
        '4c_Segments_Geographic_Countries'
        if geo_kind == 'country'
        else '4b_Segments_Geographic_Regions'
    )
    routed['Label'] = f'{metric} - {member}'
    return routed


def _normalize_extracted_geographic_facts(extracted):
    """Normalize geographic facts without leaking one table basis across a filing.

    A filing may contain comparative billing-location XBRL and a new
    headquarters-location table simultaneously. Basis is therefore attached
    only when carried by the exact fact/table/axis/concept source.
    """
    if not extracted:
        return extracted
    changed = 0
    for fact in extracted:
        if fact.get('Category') not in {
            '4a_Segments_Business',
            '4b_Segments_Geographic_Regions',
            '4c_Segments_Geographic_Countries',
        }:
            continue
        metric, member = _split_segment_display_label(fact.get('Label'))
        if _normalize_label_key(metric) != 'revenue' or not member:
            continue
        source_basis = str(fact.get('SourceGeographicBasis') or '').strip()
        axis_hint = fact.get('Category') in {
            '4b_Segments_Geographic_Regions',
            '4c_Segments_Geographic_Countries',
        }
        geo_kind = _classify_geographic_member(member, axis_hint)
        if geo_kind is None and source_basis and _geo_alias_key(member) in {'other', 'all other'}:
            geo_kind = 'region'
        if geo_kind is None:
            continue
        canonical_member = (
            'Other' if _geo_alias_key(member) in {'other', 'all other'}
            else _canonical_geographic_member_label(member, axis_hint)
        )
        target_category = (
            '4c_Segments_Geographic_Countries'
            if geo_kind == 'country' else '4b_Segments_Geographic_Regions'
        )
        basis_suffix = f' ({source_basis})' if source_basis else ''
        target_label = f'{metric} - {canonical_member}{basis_suffix}'
        if fact.get('Category') != target_category or fact.get('Label') != target_label:
            fact['Category'] = target_category
            fact['Label'] = target_label
            changed += 1
    if changed:
        print(f"  [Geography] Normalized/routed {changed} geographic revenue fact(s).")
    return extracted


def _geo_values_scale_equivalent(left, right) -> bool:
    """True only for exact/effectively exact 1x, 1,000x or 1,000,000x copies."""
    try:
        left = float(left); right = float(right)
    except (TypeError, ValueError):
        return False
    if not (np.isfinite(left) and np.isfinite(right)):
        return False
    if left == right:
        return True
    if left == 0 or right == 0 or (left < 0) != (right < 0):
        return False
    hi, lo = max(abs(left), abs(right)), min(abs(left), abs(right))
    ratio = hi / lo
    return any(abs(ratio - scale) <= 1e-8 * scale for scale in (1_000.0, 1_000_000.0))


def _consolidate_geographic_alias_facts(df: pd.DataFrame) -> pd.DataFrame:
    """Canonicalize geographic aliases and quarantine exact scale duplicates.

    Runs before generic fact dedup. It may change Category/Label metadata for
    proven geographic revenue facts and may drop only a smaller value that is
    an exact 1,000x/1,000,000x unit copy of a larger candidate for the identical
    label and period. No value is rescaled, derived, copied between periods, or
    moved between different reporting bases.
    """
    required = {'Category', 'Label', 'Value'}
    if df is None or df.empty or not required.issubset(df.columns):
        return df
    geo_categories = {
        '4a_Segments_Business',
        '4b_Segments_Geographic_Regions',
        '4c_Segments_Geographic_Countries',
    }
    work = df.copy()
    row_meta = {}

    # Pass 1: identify explicit bases and canonical member names.
    for idx, row in work.iterrows():
        category = row.get('Category')
        if category not in geo_categories:
            continue
        metric, member = _split_segment_display_label(row.get('Label'))
        if _normalize_label_key(metric) != 'revenue' or not member:
            continue
        member_base, suffix_basis = _split_geographic_basis_suffix(member)
        source_basis = str(row.get('SourceGeographicBasis') or '').strip()
        explicit_basis = suffix_basis or (source_basis if source_basis in _GEO_EXPLICIT_BASES else None)
        axis_hint = category in {
            '4b_Segments_Geographic_Regions',
            '4c_Segments_Geographic_Countries',
        }
        geo_kind = _classify_geographic_member(member_base, axis_hint)
        # "Other" is geographic only when a source explicitly establishes a
        # geographic reporting basis. Never route a generic business "Other".
        if geo_kind is None and explicit_basis and _geo_alias_key(member_base) in {
                'other', 'all other'}:
            geo_kind = 'region'
        if geo_kind is None:
            continue
        canonical_member = _canonical_geographic_member_label(member_base, axis_hint)
        target_category = (
            '4c_Segments_Geographic_Countries'
            if geo_kind == 'country' else '4b_Segments_Geographic_Regions'
        )
        row_meta[idx] = {
            'member': canonical_member,
            'basis': explicit_basis,
            'category': target_category,
        }

    if not row_meta:
        return work

    # Basis is never propagated across a whole filing; one filing may contain
    # multiple geographic reporting bases.

    # Pass 3: infer an unqualified alias's basis only from repeated identical
    # overlaps with one explicit-basis series and zero material conflicts.
    period_cols = [c for c in ('End', 'Duration', 'FY', 'Q') if c in work.columns]
    explicit_series = defaultdict(lambda: defaultdict(list))
    unqualified = defaultdict(list)
    for idx, meta in row_meta.items():
        period_key = tuple(work.at[idx, c] for c in period_cols)
        value = work.at[idx, 'Value']
        member = meta['member']
        if meta['basis']:
            explicit_series[(member, meta['basis'])][period_key].append(value)
        else:
            unqualified[member].append((idx, period_key, value))

    for member, rows in unqualified.items():
        candidate_scores = []
        for (series_member, basis), values_by_period in explicit_series.items():
            if series_member != member:
                continue
            matches = conflicts = 0
            for _, period_key, value in rows:
                candidates = values_by_period.get(period_key, ())
                if not candidates:
                    continue
                if any(_geo_values_scale_equivalent(value, candidate)
                       for candidate in candidates):
                    matches += 1
                else:
                    conflicts += 1
            if matches >= 2 and conflicts == 0:
                candidate_scores.append((matches, basis))
        if candidate_scores:
            candidate_scores.sort(reverse=True)
            if len(candidate_scores) == 1 or candidate_scores[0][0] > candidate_scores[1][0]:
                winning_basis = candidate_scores[0][1]
                for idx, _, _ in rows:
                    row_meta[idx]['basis'] = winning_basis

    # Pass 4: source-cohort inference. Assign an unqualified member only when
    # at least two *other* geographic members from the same filing/period agree
    # on one basis, no other basis is present, and there is no same-member
    # explicit-basis value conflict. This fills omitted basis metadata without
    # repeating the unsafe whole-filing propagation used by v1.
    source_cols = [c for c in ('Filed', 'End', 'Duration', 'Start') if c in work.columns]
    if source_cols:
        source_groups = defaultdict(list)
        for idx, meta in row_meta.items():
            source_key = tuple(work.at[idx, c] for c in source_cols)
            source_groups[source_key].append(idx)
        for indices in source_groups.values():
            bases = {row_meta[idx]['basis'] for idx in indices if row_meta[idx]['basis']}
            evidence_members = {
                row_meta[idx]['member'] for idx in indices if row_meta[idx]['basis']
            }
            if len(bases) != 1 or len(evidence_members) < 2:
                continue
            winning_basis = next(iter(bases))
            for idx in indices:
                meta = row_meta[idx]
                if meta['basis']:
                    continue
                same_member_explicit = [
                    other for other in indices
                    if row_meta[other]['member'] == meta['member']
                    and row_meta[other]['basis']
                ]
                if same_member_explicit:
                    value = work.at[idx, 'Value']
                    if not any(_geo_values_scale_equivalent(value, work.at[other, 'Value'])
                               for other in same_member_explicit):
                        continue
                meta['basis'] = winning_basis

    changed = 0
    for idx, meta in row_meta.items():
        suffix = f" ({meta['basis']})" if meta['basis'] else ''
        target_label = f"Revenue - {meta['member']}{suffix}"
        if work.at[idx, 'Category'] != meta['category']:
            work.at[idx, 'Category'] = meta['category']; changed += 1
        if work.at[idx, 'Label'] != target_label:
            work.at[idx, 'Label'] = target_label; changed += 1

    # Pass 5: remove only exact unit-scale duplicates for the identical series
    # and period. The larger value is retained because sec-data's output unit is
    # raw currency units; the smaller copy is an unscaled thousands/millions
    # table value. Materially different values remain for normal restatement
    # arbitration.
    group_cols = ['Category', 'Label'] + period_cols
    drop_idx = []
    for _, group in work.loc[list(row_meta)].groupby(group_cols, dropna=False, sort=False):
        numeric = pd.to_numeric(group['Value'], errors='coerce').dropna()
        distinct = sorted(set(float(v) for v in numeric))
        if len(distinct) < 2:
            continue
        max_abs = max(abs(v) for v in distinct)
        if max_abs < 1_000_000:
            continue
        if not all(_geo_values_scale_equivalent(v, max_abs if v >= 0 else -max_abs)
                   for v in distinct):
            continue
        for idx in group.index:
            value = pd.to_numeric(pd.Series([work.at[idx, 'Value']]), errors='coerce').iloc[0]
            if pd.notna(value) and abs(float(value)) < max_abs:
                drop_idx.append(idx)

    if drop_idx:
        work = work.drop(index=sorted(set(drop_idx)))
        print(f"  [Geography] Removed {len(set(drop_idx))} exact unit-scale duplicate fact(s).")
    if changed:
        print(f"  [Geography] Consolidated {changed} geographic alias/basis metadata field(s).")
    return work


def _gb_series_categories_compatible(html_category: str, xbrl_category: str) -> bool:
    geo = {'4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries'}
    if html_category in geo:
        return xbrl_category in geo
    if html_category == '4a_Segments_Business':
        # Product/service revenue can be presented on the face statement and
        # moved to 4a later.  Reconcile it before that move creates duplicates.
        return xbrl_category in {'4a_Segments_Business', '1_Income_Statement'}
    return html_category == xbrl_category


def _reconcile_html_business_breakdown_facts(all_facts):
    """Quarantine and reconcile low-priority HTML table facts against XBRL.

    Rules are deliberately asymmetric: XBRL always wins.  HTML can only fill
    missing periods after one exact normalized label match, or after at least
    two value-identical overlapping periods prove an alias.  Conflicts and
    ambiguous matches are dropped rather than guessed.
    """
    if not all_facts:
        return list(all_facts or [])

    html_rows = []
    xbrl_rows = []
    passthrough = []
    relevant_categories = {
        '1_Income_Statement', '4a_Segments_Business',
        '4b_Segments_Geographic_Regions',
        '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated',
    }
    for fact in all_facts:
        if fact.get('Concept') == _HTML_BUSINESS_CONCEPT:
            html_rows.append(_gb_route_html_business_fact(fact))
        else:
            passthrough.append(fact)
            concept = str(fact.get('Concept') or '')
            if (fact.get('Category') in relevant_categories
                    and concept and not concept.startswith('HTML')):
                xbrl_rows.append(fact)

    if not html_rows:
        return list(all_facts)

    # If one source row is classified with more than one metric prefix, the
    # local context was contaminated.  Fail closed for every conflicting copy.
    conflict_groups = defaultdict(list)
    for position, fact in enumerate(html_rows):
        metric_key, member_key = _gb_normalized_display_parts(fact.get('Label'))
        try:
            numeric_value = round(float(fact.get('Value')), 6)
        except (TypeError, ValueError):
            numeric_value = None
        source_identity = (
            fact.get('Accession'), fact.get('SourceTableFingerprint') or '',
            member_key, _gb_fact_period_key(fact), numeric_value,
            _normalize_label_key(fact.get('SourceValueHeader') or ''),
        )
        conflict_groups[source_identity].append((position, metric_key))

    conflict_positions = set()
    for entries in conflict_groups.values():
        if len({metric for _, metric in entries}) > 1:
            conflict_positions.update(position for position, _ in entries)

    if conflict_positions:
        html_rows = [fact for position, fact in enumerate(html_rows)
                     if position not in conflict_positions]

    # Build XBRL-backed series.  Each period can contain several duplicate
    # facts; an HTML value only needs to match one authoritative candidate.
    xbrl_series = {}
    for fact in xbrl_rows:
        key = (fact.get('Category'), str(fact.get('Label') or ''))
        series = xbrl_series.setdefault(key, {
            'Category': key[0], 'Label': key[1],
            'Parts': _gb_normalized_display_parts(key[1]),
            'Values': defaultdict(list),
        })
        series['Values'][_gb_fact_period_key(fact)].append(fact.get('Value'))

    html_series = defaultdict(list)
    for fact in html_rows:
        html_series[(fact.get('Category'), str(fact.get('Label') or ''))].append(fact)

    kept_html = []
    dropped_overlap = 0
    dropped_conflict = len(conflict_positions)
    canonicalized = 0

    for (html_category, html_label), facts in html_series.items():
        html_parts = _gb_normalized_display_parts(html_label)
        html_values = defaultdict(list)
        for fact in facts:
            html_values[_gb_fact_period_key(fact)].append(fact.get('Value'))

        candidates = []
        exact_conflict = False
        for series in xbrl_series.values():
            exact_label = series['Parts'] == html_parts
            same_metric = series['Parts'][0] == html_parts[0]
            if not same_metric:
                continue
            if not (exact_label or _gb_series_categories_compatible(
                    html_category, series['Category'])):
                continue
            overlap = set(html_values) & set(series['Values'])
            if not overlap:
                continue
            matching_periods = 0
            for period in overlap:
                if any(_gb_close_values(hv, xv)
                       for hv in html_values[period]
                       for xv in series['Values'][period]):
                    matching_periods += 1
            all_match = matching_periods == len(overlap)
            if exact_label and not all_match:
                exact_conflict = True
                continue
            # A normalized label match needs one verified overlap.  A label
            # alias (AWS/Amazon Web Services) needs two identical overlaps.
            if all_match and (exact_label or matching_periods >= 2):
                candidates.append((
                    1 if exact_label else 0,
                    matching_periods,
                    1 if html_category == series['Category'] else 0,
                    series,
                ))

        if exact_conflict:
            dropped_conflict += len(facts)
            continue
        if not candidates:
            kept_html.extend(facts)
            continue

        candidates.sort(key=lambda item: item[:3], reverse=True)
        best_score = candidates[0][:3]
        best = [item[3] for item in candidates if item[:3] == best_score]
        if len(best) != 1:
            # Multiple XBRL series fit equally well.  The HTML series is
            # redundant but cannot be safely assigned to one canonical row.
            dropped_conflict += len(facts)
            continue
        canonical = best[0]
        for fact in facts:
            period = _gb_fact_period_key(fact)
            if period in canonical['Values']:
                dropped_overlap += 1
                continue
            merged = dict(fact)
            if (merged.get('Category'), merged.get('Label')) != (
                    canonical['Category'], canonical['Label']):
                merged['SourceCanonicalizedFrom'] = merged.get('Label')
                merged['Category'] = canonical['Category']
                merged['Label'] = canonical['Label']
                canonicalized += 1
            kept_html.append(merged)

    if conflict_positions or dropped_overlap or dropped_conflict or canonicalized:
        print(
            f"  [Business Table Cleanup] kept={len(kept_html)}, "
            f"dropped_xbrl_overlap={dropped_overlap}, "
            f"dropped_conflict={dropped_conflict}, "
            f"canonicalized_missing_periods={canonicalized}."
        )
    return passthrough + kept_html


def extract_from_filing(filing, ye_month, ticker=None, use_arelle=False):
    with _ProfileTimer("extract_from_filing_total"):
        return _extract_from_filing_impl(filing, ye_month, ticker=ticker, use_arelle=use_arelle)


def _extract_from_filing_impl(filing, ye_month, ticker=None, use_arelle=False):
    # Resolve lazy filing metadata once and retain the original fallback path.
    _filing_meta = _get_filing_local_metadata(filing)
    filing_form = _filing_meta.get('form') or ''
    filing_date = _filing_meta.get('filing_date')
    _filing_url_loaded = False
    _filing_url_value = None

    def _get_filing_url_once():
        nonlocal _filing_url_loaded, _filing_url_value
        if not _filing_url_loaded:
            _filing_url_value = _get_filing_url_cached(filing)
            _filing_url_loaded = True
        return _filing_url_value

    if use_arelle and filing_form == '10-K':
        url = _get_filing_url_once()
        if url:
            arelle_update_concept_map(url)

    try:
        with _ProfileTimer("fetch_xbrl"):
            xbrl = fetch_xbrl(filing)
        if not xbrl: return [], None
        with _ProfileTimer("xbrl_facts_to_dataframe"):
            facts_df = xbrl.facts.to_dataframe()
        _profile_count("xbrl_fact_rows", len(facts_df))
        if hasattr(xbrl, 'calculation_trees'):
            resolve_custom_tags(xbrl.calculation_trees)
            with _ProfileTimer("augment_facts_with_calculations"):
                facts_df = augment_facts_with_calculations(facts_df, xbrl.calculation_trees)
    except Exception as e:
        print(f"  Warning: Could not parse XBRL: {e}")
        return [], None

    # --- Learn unmapped face-statement concepts from this filing's own
    #     presentation linkbase, so no face line item is ever dropped. ---
    try:
        with _ProfileTimer("learn_statement_concepts"):
            _n_learned = learn_statement_concepts(xbrl)
        if _n_learned:
            print(f"  [FaceStmt] Registered/merged {_n_learned} face-statement concept(s)")
    except Exception as _e:
        print(f"  Warning: face-statement concept learning failed: {_e}")

    # Per-filing snapshot of CONCEPT_MAP as a reverse map (tag -> [labels]).
    # 1) O(1) lookup per fact instead of scanning every label per row.
    # 2) Thread-safe: learners in parallel workers add NEW keys to
    #    CONCEPT_MAP; iterating the live dict per row could raise
    #    'dictionary changed size during iteration'.
    _tag_to_labels = {}
    _tag_rank_lookup = {}
    for _lbl, _info in list(CONCEPT_MAP.items()):
        for _rank, _t in enumerate(_info.get('tags', [])):
            _tag_to_labels.setdefault(_t, []).append(_lbl)
            _tag_rank_lookup[(_lbl, _t)] = _rank

    extracted = []
    with _ProfileTimer("extract_prepare_columns"):
        dim_cols = [c for c in facts_df.columns if c.startswith('dim_')]
    _profile_count("dimension_columns_seen", len(dim_cols))
    is_calc_col = 'is_calculated' in facts_df.columns

    _filing_date_raw = filing_date
    filing_dt = pd.to_datetime(_filing_date_raw)

    import re as _re
    
    # Pre-calculate company-specific keywords for dynamic dimension-member filtering.
    _co_ignore_keywords = set()
    if ticker: _co_ignore_keywords.add(ticker.lower())
    _noise = {'inc', 'corp', 'co', 'ltd', 'company', 'corporation', 'incorporated', 'holdings', 'holding', 'group', 'the'}
    try:
        _raw_co = getattr(filing, 'company', '')
        if isinstance(_raw_co, str):
            _clean_co = _re.sub(r'[^\w\s]', ' ', _raw_co.lower())
            for part in _clean_co.split():
                if part not in _noise and len(part) > 1:
                    _co_ignore_keywords.add(part)
                    
        # Check for former names
        company_obj = getattr(filing, 'company_obj', None)
        if hasattr(company_obj, 'former_names'):
            for fn in company_obj.former_names:
                name_str = str(fn.name) if hasattr(fn, 'name') else str(fn)
                fn_clean = _re.sub(r'[^\w\s]', ' ', name_str.lower())
                for part in fn_clean.split():
                    if part not in _noise and len(part) > 1:
                        _co_ignore_keywords.add(part)
    except:
        pass
    
    @lru_cache(maxsize=32768)
    def _is_ignore_mem(m_clean, concept_name):
        ml = m_clean.lower()
        if m_clean in IGNORE_MEMBERS:
            return True
        if ml == concept_name.lower() or ml.replace(' ', '') == concept_name.lower():
            return True
        tokens = _re.sub(r'[^\w\s]', ' ', ml).split()
        if not tokens: return False
        all_ignored = True
        for t in tokens:
            if t not in _noise and t not in _co_ignore_keywords:
                all_ignored = False
                break
        return all_ignored

    # --- Vectorized per-fact derivations, hoisted out of the row loop --------
    # The loop previously parsed period dates with a scalar pd.to_datetime on
    # every fact -- the single largest per-fact cost. Parsing in bulk here is
    # ~10x faster and yields identical Timestamps. Helper columns are '_'-
    # prefixed so they never match the 'dim_' dimension-column filter.
    if 'period_end' in facts_df.columns:
        _end_raw = facts_df['period_end']
        if 'period_instant' in facts_df.columns:
            _end_raw = _end_raw.where(_end_raw.notna(), facts_df['period_instant'])
    elif 'period_instant' in facts_df.columns:
        _end_raw = facts_df['period_instant']
    else:
        _end_raw = pd.Series(index=facts_df.index, dtype=object)
    facts_df['_end_dt'] = pd.to_datetime(_end_raw, errors='coerce')
    facts_df['_start_dt'] = (pd.to_datetime(facts_df['period_start'], errors='coerce')
                             if 'period_start' in facts_df.columns
                             else pd.Series(pd.NaT, index=facts_df.index))
    facts_df['_concept_short'] = facts_df['concept'].astype(str).str.rsplit(':', n=1).str[-1]
    facts_df['_dim_count'] = (facts_df[dim_cols].notna().sum(axis=1)
                              if dim_cols else pd.Series(0, index=facts_df.index, dtype='int64'))

    # Fiscal year/quarter depends only on the end date and fiscal year-end
    # month (``dur`` is intentionally ignored by get_period_info). Computing
    # it once per row in vectorized form avoids tens of thousands of scalar
    # Timestamp/DateOffset operations.
    _yr = facts_df['_end_dt'].dt.year
    _mo = facts_df['_end_dt'].dt.month
    _dy = facts_df['_end_dt'].dt.day
    _early = _dy < 15
    _eff_mo = _mo.where(~_early, ((_mo - 2) % 12) + 1)
    _eff_yr = _yr.where(~_early, _yr - (_mo == 1).astype('int64'))
    facts_df['_fy_fast'] = _eff_yr + (_eff_mo > ye_month).astype('int64')
    _m_into = (_eff_mo - ye_month) % 12
    facts_df['_q_fast'] = np.select(
        [_m_into == 0, _m_into <= 3, _m_into <= 6, _m_into <= 9],
        ['Q4', 'Q1', 'Q2', 'Q3'],
        default='Q4',
    )

    _outflow_prefixes = ('PaymentsTo', 'RepaymentsOf', 'CostOf', 'InterestExpense',
                         'SellingGeneral', 'ResearchAndDevelopment', 'PurchasesOf',
                         'PurchaseOf', 'PaymentsFor', 'PaymentsOf', 'AcquisitionsOf',
                         'PropertyPlantAndEquipmentAdditions')
    _norm_business_axes = {ax.replace(':', '_') for ax in BUSINESS_SEGMENT_AXES}
    _norm_geographic_axes = {ax.replace(':', '_') for ax in GEOGRAPHIC_SEGMENT_AXES}
    _consolidation_axis_norm = CONSOLIDATION_AXIS.replace(':', '_')
    _dim_col_roles = {
        _c: {
            'business': any(ax in _c for ax in _norm_business_axes),
            'geographic': any(ax in _c for ax in _norm_geographic_axes),
            'product': 'ProductOrServiceAxis' in _c,
            'reconciliation': _consolidation_axis_norm in _c,
            'equity_component': any(ax in _c for ax in EQUITY_COMPONENT_AXES),
            'supplemental_axis': (
                'Supplemental' in _c
                or 'Proforma' in _c
                or 'NonGAAP' in _c
                or 'AdjustmentsFor' in _c
            ),
        }
        for _c in dim_cols
    }
    _categorical_noise = {'geographic concentration risk', 'concentration risk',
                          'operating segments', 'reportable segments',
                          'business segments', 'product or service'}
    _member_label_cache = {}
    _segment_prefix_cache = {}
    _segment_prefix_hint_cache = {}

    def _matched_segment_prefix(concept, c_lower):
        if concept in _segment_prefix_cache:
            return _segment_prefix_cache[concept]
        matched = STANDARD_TAG_MAP.get(concept)
        if not matched:
            exact_core = _CONCEPT_TAG_TO_LABEL.get(concept)
            if exact_core:
                matched = exact_core.replace('Total ', '')
            else:
                for kw, nice_label in SEGMENT_PREFIXES:
                    if kw in c_lower:
                        matched = nice_label
                        break
        if matched == 'Revenue' and c_lower.startswith('costof'):
            matched = None
        if matched and any(p in c_lower for p in _NON_SEGMENT_CONCEPT_PATTERNS):
            matched = None
        _segment_prefix_cache[concept] = matched
        return matched

    def _has_segment_prefix_hint(c_lower):
        cached = _segment_prefix_hint_cache.get(c_lower)
        if cached is None:
            cached = any(kw == c_lower or kw in c_lower for kw, _ in SEGMENT_PREFIXES[:4])
            _segment_prefix_hint_cache[c_lower] = cached
        return cached

    def _clean_member_tag(member_tag):
        key = str(member_tag)
        cached = _member_label_cache.get(key)
        if cached is not None:
            return cached
        m_clean = None
        try:
            m_labels = xbrl.get_labels_for_concept(member_tag)
            if m_labels:
                m_clean = max(m_labels, key=len)
                m_clean = m_clean.replace(' [Member]', '').replace(' Member', '')
                m_clean = _normalize_member_label(m_clean)
        except Exception:
            # Preserve the original behavior under a transient taxonomy lookup
            # failure: use the fallback now, but allow a later fact to retry.
            return _normalize_member_label(clean_name(member_tag))
        if not m_clean:
            m_clean = _normalize_member_label(clean_name(member_tag))
        _member_label_cache[key] = m_clean
        return m_clean

    _column_positions = {name: pos for pos, name in enumerate(facts_df.columns)}
    _dim_col_positions = tuple((c, _column_positions[c]) for c in dim_cols)
    row = _FastTupleRow(None, _column_positions)
    for _row_values in facts_df.itertuples(index=False, name=None):
        row.reset(_row_values)
        concept = row['_concept_short']
        c_lower = concept.lower()
        end_dt = row['_end_dt']
        if pd.isna(end_dt): continue

        if end_dt > filing_dt and 'PerformanceObligation' not in concept:
            continue

        _start_dt = row['_start_dt']
        dur = (end_dt - _start_dt).days if pd.notna(_start_dt) else 0
        fy, q = int(row['_fy_fast']), row['_q_fast']
        # Date strings are only needed for rows that survive extraction.  Keep
        # the exact previous formatting while avoiding per-fact strftime work
        # for facts that are discarded.
        end_str = None
        start_str = None
        
        val = row['value']
        if isinstance(val, str) and val.startswith('P') and 'useful' in c_lower:
            try: val = parse_duration(val)
            except: pass

        dim_count = row['_dim_count']
        is_calculated = bool(row.get('is_calculated', False)) if is_calc_col else False
        # Build the active-dimension list only for facts that actually need
        # dimension inspection (standard consolidated checks or segment rows).
        active_dim_cols = None

        if isinstance(val, (int, float)):
            if concept.startswith(_outflow_prefixes) and val < 0:
                if 'Net' not in concept or concept.startswith(('Payments', 'Purchases', 'Purchase', 'Acquisitions', 'Repayments')): 
                    val = abs(val)

        # 1. Check for Standard Concepts (Consolidated)
        captured_by_concept_map = False
        target_labels = _tag_to_labels.get(concept)
                
        # If not matched directly, use fuzzy matcher for numeric facts (ignore text blocks)
        if not target_labels and 'TextBlock' not in concept and isinstance(val, (int, float)):
            # Do not fuzzy match if the concept is explicitly mapped for segment/disclosure extraction
            is_segment_prefix = _has_segment_prefix_hint(c_lower)
            _is_flow_recon = concept.startswith('IncreaseDecreaseIn') or concept.startswith('DecreaseIncreaseIn')
            if (concept not in STANDARD_TAG_MAP and concept not in _CONCEPT_TAG_TO_LABEL
                    and (not is_segment_prefix or _is_flow_recon)):
                fuzzy_lbl = resolve_concept_to_label(concept)
                if fuzzy_lbl and fuzzy_lbl in CONCEPT_MAP:
                    target_labels = (fuzzy_lbl,)
        if not target_labels:
            target_labels = ()

        is_consolidated_fact = (dim_count == 0)
        if target_labels and dim_count > 0:
            if active_dim_cols is None:
                active_dim_cols = [c for c, pos in _dim_col_positions if pd.notna(_row_values[pos])]
            # Check for explicit supplemental / pro-forma axes
            has_supplemental = any(_dim_col_roles[c]['supplemental_axis'] for c in active_dim_cols)
            if not has_supplemental:
                # Check if ALL active dimensions are just restating that it's consolidated
                all_consolidated = True
                for ad in active_dim_cols:
                    if 'Consolidated' not in str(row[ad]) and 'Consolidated' not in str(ad):
                        all_consolidated = False
                        break
                if all_consolidated:
                    is_consolidated_fact = True

        for target_label in target_labels:
            info = CONCEPT_MAP[target_label]
            if is_consolidated_fact:
                rank = _tag_rank_lookup.get((target_label, concept), 999)
                if end_str is None:
                    end_str = end_dt.strftime('%Y-%m-%d')
                    start_str = _start_dt.strftime('%Y-%m-%d') if pd.notna(_start_dt) else None
                extracted.append({
                    'Category': info['cat'], 'Label': target_label, 'Value': val,
                    'FY': fy, 'Q': q, 'End': end_str, 'Start': start_str, 'Duration': dur,
                    'Filed': _filing_date_raw, 'TagRank': rank,
                    'DimCount': 0, 'IsCalculated': is_calculated, 'Concept': concept,
                    'FilingUrl': _get_filing_url_once()
                })
                captured_by_concept_map = True

        # 2. Check for Segment/Dimensional Breakdowns
        matched_prefix = _matched_segment_prefix(concept, c_lower)

        if matched_prefix and dim_count > 0:
            if active_dim_cols is None:
                active_dim_cols = [c for c, pos in _dim_col_positions if pd.notna(_row_values[pos])]
            # -- Noise guard: skip concepts that are never segment performance metrics --
            if concept in SEGMENT_NOISE_CONCEPTS:
                continue
            # Skip ratio/percentage values that leaked into segment rows
            if _is_segment_noise_value(concept, val):
                continue

            is_business = any(_dim_col_roles[c]['business'] for c in active_dim_cols)
            is_geographic = any(_dim_col_roles[c]['geographic'] for c in active_dim_cols)
            is_reconciliation = any(_dim_col_roles[c]['reconciliation'] for c in active_dim_cols)
            is_product = any(_dim_col_roles[c]['product'] for c in active_dim_cols)
            is_equity_component = any(_dim_col_roles[c]['equity_component'] for c in active_dim_cols)

            label_mems = []
            seen_mems = set()
            is_region = False
            is_country = False
            is_definitely_geo = False

            for c in active_dim_cols:
                member_tag = row[c]
                m_clean = _clean_member_tag(member_tag)

                if _is_ignore_mem(m_clean, concept): continue
                
                ml = m_clean.lower()
                # Geographic classification rules:
                # - Full country names + region keywords = strong signal (no axis needed)
                # - ISO alpha-2 codes = WEAK signal. Only trust when the XBRL axis
                #   is a known geographic axis (is_geographic=True).  Without that
                #   confirmation an ISO code is almost certainly an internal abbreviation
                #   (e.g. NVDA labels their Networking sub-segment as TW internally).
                if _REGION_KEYWORDS_RE.search(ml) or ml in GEOGRAPHIC_COUNTRIES:
                    is_definitely_geo = True
                elif ml in GEOGRAPHIC_CODES and is_geographic:
                    is_definitely_geo = True

                if is_geographic or is_definitely_geo:
                    if ml in GEOGRAPHIC_COUNTRIES:
                        is_country = True
                    elif ml in GEOGRAPHIC_CODES and is_geographic:
                        is_country = True
                    elif _REGION_KEYWORDS_RE.search(ml):
                        is_region = True
                    else:
                        if len(ml) == 2: is_country = True
                        else: is_region = True

                if ml in seen_mems: continue
                label_mems.append(m_clean)
                seen_mems.add(ml)

            if not label_mems: continue

            # SPECIFICITY FILTER: If multiple members exist, keep only the most specific.
            if len(label_mems) > 1:
                # 1. Drop known categorical "noise" members if more specific ones exist
                specific_mems = [m for m in label_mems if m.lower() not in _categorical_noise]
                if specific_mems: label_mems = specific_mems

                # 2. Container-Child Filter
                if len(label_mems) > 1:
                    to_drop = set()
                    mems = list(label_mems)
                    for i in range(len(mems)):
                        for j in range(len(mems)):
                            if i == j: continue
                            m1l, m2l = mems[i].lower(), mems[j].lower()
                            if _re.search(rf'\b{_re.escape(m2l)}\b', m1l) and len(m1l) > len(m2l):
                                conjuncts = [c.strip() for c in _re.split(r'\band\b|\bor\b|,', m1l)]
                                if m2l in conjuncts:
                                    to_drop.add(mems[i])
                                else:
                                    to_drop.add(mems[j])
                    filtered = [m for m in mems if m not in to_drop]
                    label_mems = filtered if filtered else label_mems

            if is_equity_component and not is_business and not is_geographic and not is_product:
                seg_category = '6_Disclosures'
            elif matched_prefix in ('Customer Concentration %', 'Customer Revenue'): 
                seg_category = '7_Concentration_Risk'
            elif 'Useful Life' in matched_prefix or 'RPO' in matched_prefix: 
                seg_category = '6_Disclosures'
            else:
                final_geo_count = 0
                final_biz_count = 0
                for m in label_mems:
                    ml = m.lower()
                    # A dynamically intelligent check that accurately separates 
                    # purely geographic segment groupings ("International Revenues")
                    # from localized business segments ("International Transaction Revenues")
                    is_geo_mem = _is_pure_geographic_label(m, is_geographic)
                    
                    if is_geo_mem:
                        final_geo_count += 1
                    else:
                        final_biz_count += 1
                
                is_cross_tabulated = (is_business and is_geographic) or (final_geo_count > 0 and final_biz_count > 0)
                is_pure_geo = (is_geographic or final_geo_count > 0) and not is_cross_tabulated
                
                is_pure_product = (is_product and not is_business
                                   and not is_geographic and not is_definitely_geo
                                   and final_geo_count == 0
                                   and len(active_dim_cols) == 1)
                if is_cross_tabulated:
                    seg_category = '4d_Segments_Cross_Tabulated'
                elif is_pure_geo:
                    seg_category = '4c_Segments_Geographic_Countries' if is_country else '4b_Segments_Geographic_Regions'
                elif is_pure_product and matched_prefix in _FACE_DISAGG_PREFIXES:
                    # Nature-of-revenue disaggregation presented on the income
                    # statement face -- keep it there so the granular cost /
                    # revenue lines (IBM: Services/Sales/Financing; MSFT/AAPL:
                    # Product/Service) continue into modern filings instead of
                    # being buried as pseudo-segments in the 4a block.
                    seg_category = '1_Income_Statement'
                else:
                    seg_category = '4a_Segments_Business'

            joined_mems = " - ".join(sorted(label_mems))
            label = f"{matched_prefix} - {joined_mems}"
            
            rank = 999
            core_label = _CONCEPT_TAG_TO_LABEL.get(concept)
            if core_label and core_label in CONCEPT_MAP:
                info = CONCEPT_MAP[core_label]
                if isinstance(info, dict) and 'tags' in info and concept in info['tags']:
                    rank = _tag_rank_lookup.get((core_label, concept), 999)

            if end_str is None:
                end_str = end_dt.strftime('%Y-%m-%d')
                start_str = _start_dt.strftime('%Y-%m-%d') if pd.notna(_start_dt) else None
            extracted.append({
                'Category': seg_category, 'Label': label, 'Value': val,
                'FY': fy, 'Q': q, 'End': end_str, 'Start': start_str, 'Duration': dur,
                'Filed': _filing_date_raw, 'TagRank': rank,
                'DimCount': dim_count, 'IsCalculated': is_calculated, 'Concept': concept,
                'SourceGeographicBasis': (
                    _gb_geographic_basis_from_context(
                        ' '.join(str(c) for c in active_dim_cols) + ' ' + concept)
                    if seg_category in {
                        '4b_Segments_Geographic_Regions',
                        '4c_Segments_Geographic_Countries',
                    } else None
                ),
            })
        
        elif matched_prefix and not captured_by_concept_map and dim_count == 0:
            already_in_concept_map = concept in _tag_to_labels
            if already_in_concept_map and concept not in REVENUE_SUBTYPE_CONCEPTS: continue
            
            custom_name = clean_name(concept)
            if len(custom_name) < 3 or custom_name.lower() in [matched_prefix.lower(), 'total', 'consolidated']:
                continue

            if matched_prefix in ('Customer Concentration %', 'Customer Revenue'): seg_category = '7_Concentration_Risk'
            elif 'Useful Life' in matched_prefix or 'RPO' in matched_prefix: seg_category = '6_Disclosures'
            else: seg_category = '4a_Segments_Business'

            label = f"{matched_prefix} - {custom_name}"
            
            rank = 999
            core_label = _CONCEPT_TAG_TO_LABEL.get(concept)
            if core_label and core_label in CONCEPT_MAP:
                info = CONCEPT_MAP[core_label]
                if isinstance(info, dict) and 'tags' in info and concept in info['tags']:
                    rank = _tag_rank_lookup.get((core_label, concept), 999)

            if end_str is None:
                end_str = end_dt.strftime('%Y-%m-%d')
                start_str = _start_dt.strftime('%Y-%m-%d') if pd.notna(_start_dt) else None
            extracted.append({
                'Category': seg_category, 'Label': label, 'Value': val,
                'FY': fy, 'Q': q, 'End': end_str, 'Start': start_str, 'Duration': dur,
                'Filed': _filing_date_raw, 'TagRank': rank,
                'DimCount': 0, 'IsCalculated': is_calculated, 'Concept': concept,
                'SourceGeographicBasis': _gb_geographic_basis_from_context(concept),
            })

    has_interest = any(x['Label'] == 'Interest Expense' for x in extracted)
    if not has_interest:
        try:
            soup = BeautifulSoup(fetch_html(filing), 'html.parser')
            for tr in soup.find_all('tr'):
                text = tr.get_text(separator=' ', strip=True).lower()
                if 'interest expense' in text and not 'interest income' in text and not 'net' in text:
                    real_nums, tds = [], tr.find_all('td')
                    for j, td in enumerate(tds):
                        ct = td.get_text(strip=True)
                        if '%' in ct: continue
                        m = re.search(r'^[\(\$]*([\d,]+)[\)]*$', ct)
                        if m and not (j + 1 < len(tds) and '%' in tds[j+1].get_text(strip=True)):
                            real_nums.append(int(m.group(1).replace(',', '')))
                            
                    if len(real_nums) >= 1:
                        # -- Dynamic scale detection --------------------------
                        # Instead of hardcoding 1_000_000 (which is only correct
                        # for companies filing in millions), we infer the scale
                        # from a known XBRL reference value, then fall back to
                        # magnitude-based heuristics.
                        scale = _infer_html_scale(extracted, real_nums)
                        end_dt = pd.to_datetime(filing.period_of_report)
                        
                        if '10-K' in filing_form:
                            # 10-K tables show Annual values: [Current Year, Prior Year 1, Prior Year 2]
                            for offset, val in enumerate(real_nums[:3]):
                                past_dt = end_dt - pd.DateOffset(years=offset)
                                fy, q = get_period_info(past_dt, ye_month, 365)
                                extracted.append({
                                    'Category': '1_Income_Statement', 'Label': 'Interest Expense', 
                                    'Value': val * scale, 'FY': fy, 'Q': q, 'End': str(past_dt.date()), 
                                    'Duration': 365, 'Filed': filing_date, 'TagRank': 99, 
                                    'DimCount': 0, 'IsCalculated': False, 'Concept': 'HTMLFallback_Annual'
                                })
                        else:
                            if len(real_nums) >= 2:
                                fy, q = get_period_info(end_dt, ye_month, 90)
                                extracted.append({
                                    'Category': '1_Income_Statement', 'Label': 'Interest Expense', 
                                    'Value': real_nums[0] * scale, 'FY': fy, 'Q': q, 'End': str(end_dt.date()), 
                                    'Duration': 90, 'Filed': filing_date, 'TagRank': 99, 
                                    'DimCount': 0, 'IsCalculated': False, 'Concept': 'HTMLFallback'
                                })
                                if len(real_nums) >= 4:
                                    extracted.append({
                                        'Category': '1_Income_Statement', 'Label': 'Interest Expense', 
                                        'Value': real_nums[2]*scale, 'FY': fy, 'Q': q, 'End': str(end_dt.date()), 
                                        'Duration': 90*int(q[-1]), 'Filed': filing_date, 'TagRank': 99, 
                                        'DimCount': 0, 'IsCalculated': False, 'Concept': 'HTMLFallback'
                                    })
                        break
        except: pass

    # Reuse the already-computed short concept column for the DEI period tag
    # lookup, then remove all scratch columns before downstream HTML checks.
    period_end_date = extract_period_end_date(filing, facts_df)

    # Additive, low-priority rescue for coherent monetary business tables.
    # Existing XBRL facts are skipped explicitly and retain priority.
    try:
        _business_table_facts = _extract_generic_business_breakdowns_from_textblocks(
            facts_df=facts_df,
            filing_form=filing_form,
            period_end_date=period_end_date,
            ye_month=ye_month,
            filed=_filing_date_raw,
            filing_url=_get_filing_url_once(),
            existing_facts=extracted,
            filing=filing,
        )
        if _business_table_facts:
            extracted.extend(_business_table_facts)
    except Exception as _business_table_error:
        _debug_print(
            f"  [Business Table] Rescue skipped ({type(_business_table_error).__name__}: "
            f"{_business_table_error})")

    facts_df = facts_df.drop(columns=['_end_dt', '_start_dt', '_concept_short', '_dim_count',
                                     '_fy_fast', '_q_fast'],
                             errors='ignore')
    extracted = _correct_xbrl_segment_labels_from_html(extracted, facts_df, filing, ye_month)
    extracted = _normalize_extracted_geographic_facts(extracted)

    # Passive provenance fields used by native annual mode.  These fields are
    # added after every extraction/recovery path so existing quarterly logic is
    # not touched and synthetic/HTML rows receive the same metadata envelope.
    _accession_no = (_filing_meta.get('accession_no')
                     or _filing_meta.get('accession_number'))
    try:
        _filing_url_for_meta = _get_filing_url_once()
    except Exception:
        _filing_url_for_meta = None

    _profile_count("extracted_facts_before_metadata", len(extracted))
    for _fact in extracted:
        _fact.setdefault('Form', filing_form)
        _fact.setdefault('Accession', _accession_no)
        _fact.setdefault('FilingUrl', _filing_url_for_meta)
        _fact.setdefault('Filed', _filing_date_raw)
        if 'Start' not in _fact:
            _start_str = None
            try:
                _dur = float(_fact.get('Duration', 0) or 0)
                _end_dt_meta = pd.to_datetime(_fact.get('End'), errors='coerce')
                if pd.notna(_end_dt_meta) and _dur > 0:
                    _start_dt_meta = _end_dt_meta - pd.Timedelta(days=int(round(_dur)))
                    _start_str = _start_dt_meta.strftime('%Y-%m-%d')
            except Exception:
                _start_str = None
            _fact['Start'] = _start_str
            _fact.setdefault('StartEstimated', _start_str is not None)
        else:
            _fact.setdefault('StartEstimated', False)

    return extracted, period_end_date

def _recover_annual_cashflow_from_html(extracted, filing, ye_month):
    """
    10-K HTML fallback for annual cash-flow totals missed by XBRL.

    Problem this solves
    -------------------
    CapEx (and other CF items) disappear every Q4 because the 10-K XBRL uses
    an extension tag (e.g. a company-specific variant) that is not in our
    CONCEPT_MAP. Without the annual total, the engine cannot compute
    Q4_discrete = Annual - YTD9, so Q4 becomes NaN.

    Strategy
    --------
    1. Identify which cash-flow labels are MISSING an annual-duration (~365d)
       fact in the current `extracted` list.  We only chase:
         - Capital Expenditures
         - Operating Cash Flow
         - Investing Cash Flow
         - Financing Cash Flow
       These four are structurally verifiable (Investing = sum of sub-items,
       Financing = sum of sub-items) and are always present in the 10-K HTML.

    2. Parse the 10-K HTML filing's cash-flow statement table(s).
       For each missing label, look for a row whose text matches a curated set
       of keyword patterns and extract the first numeric column (current year).

    3. Scale the raw HTML number to match the XBRL reference magnitude.

    4. Inject a synthetic annual fact with Duration=365 so downstream
       Q4 derivation can compute Q4 = Annual - Q1 - Q2 - Q3.

    Design principles (universal safety)
    -------------------------------------
    * Only runs when the annual fact is actually missing -- never overwrites.
    * Scale is validated against the known OCF or Revenue XBRL value.
    * The injected fact has TagRank=998 (lower priority than real XBRL = 0-10)
      so any future genuine XBRL discovery wins in dedup.
    * Negative values are taken as-is (CF can be negative).
    * If the HTML parse fails for any reason, the function returns `extracted`
      unchanged -- graceful degradation.
    """
    # Labels we want to recover, with candidate row-text patterns (lowercase)
    RECOVERY_TARGETS = {
        'Capital Expenditures': [
            # Standard US-GAAP labels
            'purchase of property', 'purchases of property',
            'capital expenditure', 'capital expenditures',
            'acquisition of property', 'acquisitions of property',
            # "Additions to property" -- used by NVDA, INTC, others
            'additions to property', 'addition to property',
            'purchases related to property', 'payments to acquire property', 
            'capital spending', 'purchase of equipment', 'investment in property',
            'purchases of equipment', 'investing in property',
            'expenditures for property', 'spending on property',
        ],
        'Operating Cash Flow': [
            'net cash provided by operating', 'net cash used in operating',
            'cash flows from operating', 'net cash from operating',
            'operating activities',
        ],
        'Investing Cash Flow': [
            'net cash provided by investing', 'net cash used in investing',
            'cash flows from investing', 'net cash from investing',
            'investing activities',
        ],
        'Financing Cash Flow': [
            'net cash provided by financing', 'net cash used in financing',
            'cash flows from financing', 'net cash from financing',
            'financing activities',
        ],
    }

    filing_end = pd.to_datetime(filing.period_of_report)
    fy, q_label = get_period_info(filing_end, ye_month)
    annual_dur = (filing_end - (filing_end - pd.DateOffset(years=1))).days
    # Use a fixed 365 so downstream duration filter (350-385) always matches
    annual_dur = 365

    end_str = str(filing_end.date())

    # Step 1: Identify which labels already have an annual fact.
    # Circularity guard: if the only 365d annual for a label equals the sum of
    # its individual 91d quarterly facts (i.e. it is the supplemental rollup,
    # not an independently-audited consolidated figure), we do NOT treat it as
    # a valid existing annual.  This lets HTML recovery fetch the true audited
    # value from the CF statement narrative.
    #
    # Example: AMD FY2025 OCF.
    #   Supplemental quarterly data in the 10-K files:
    #     Q1=939M  Q2=2011M  Q3=2159M  Q4_supp=1384M  (each 91d)
    #     Full-year "YTD" = 6493M  (365d -- just the sum above, circular)
    #   Audited consolidated CF statement:  OCF = 7709M  (365d)
    #   Without this guard, existing_annual would contain 'Operating Cash Flow'
    #   (because 6493M is a valid 365d fact) and HTML recovery would skip it,
    #   leaving the audited 7709M undiscovered.

    # Collect individual 91d quarterly facts and YTD9 per label to detect circularity
    _qtly_sum = {}   # label -> sum of all 91d fact values for this FY
    _ytd9_val = {}
    _q4_val = {}
    for fact in extracted:
        if fact.get('Category') != '3_Cash_Flow' or fact.get('FY') != fy:
            continue
        dur = fact.get('Duration', 0)
        lbl = fact['Label']
        try:
            v = abs(float(fact['Value']))
            if 60 <= dur <= 120:
                _qtly_sum[lbl] = _qtly_sum.get(lbl, 0.0) + v
                if fact.get('Q') == 'Q4':
                    _q4_val[lbl] = max(_q4_val.get(lbl, 0.0), v)
            elif 260 <= dur <= 290:
                _ytd9_val[lbl] = max(_ytd9_val.get(lbl, 0.0), v)
        except (TypeError, ValueError):
            pass

    existing_annual = set()
    for fact in extracted:
        if fact['Category'] != '3_Cash_Flow':
            continue
        dur = fact.get('Duration', 0)
        if 350 <= dur <= 385 and fact['FY'] == fy and fact['Q'] == q_label:
            lbl = fact['Label']
            try:
                annual_val = abs(float(fact['Value']))
            except (TypeError, ValueError):
                continue
            # Circularity check: if annual â‰ˆ sum of the individual quarterly facts
            # or YTD9 + Q4_supp, it was constructed by summing the supplemental data â€”
            # not an independently-verified figure. Exclude it so we can recover
            # the true audited value from the 10-K HTML.
            q_sum = _qtly_sum.get(lbl, 0.0)
            ytd9 = _ytd9_val.get(lbl, 0.0)
            q4 = _q4_val.get(lbl, 0.0)
            
            implied_from_ytd = ytd9 + q4 if ytd9 > 0 and q4 > 0 else 0
            best_sum = max(q_sum, implied_from_ytd)
            
            if best_sum > 0:
                circular = abs(annual_val - best_sum) / max(best_sum, 1) < 0.03
                if circular:
                    print(f"  [HTML Recovery] Skipping circular annual for '{lbl}': "
                          f"{annual_val:,.0f} â‰ˆ quarterly sum/YTD {best_sum:,.0f} â€” will attempt HTML recovery")
                    continue   # do not add to existing_annual
            _debug_print(f"  [Debug] Adding {lbl} ({annual_val}) to existing_annual")
            existing_annual.add(lbl)

    missing = {lbl for lbl in RECOVERY_TARGETS if lbl not in existing_annual}
    if not missing:
        return extracted   # nothing to do

    # Step 2: Parse HTML
    try:
        tables = _get_filing_html_tables_cached(filing)
    except Exception:
        return extracted

    # Step 3: Scale calibration using XBRL reference values
    # Priority: Operating Cash Flow > Revenue (both large, reliably present)
    ref_val = None
    ref_label_order = ['Operating Cash Flow', 'Revenue']
    for ref_lbl in ref_label_order:
        for fact in extracted:
            if fact['Label'] != ref_lbl:
                continue
            dur = fact.get('Duration', 0)
            if 350 <= dur <= 385 and fact['FY'] == fy:
                try:
                    ref_val = abs(float(fact['Value']))
                    break
                except Exception:
                    continue
        if ref_val:
            break

    def _infer_scale_from_ref(raw_val):
        """Return the scale factor that brings raw_val closest to ref_val."""
        if not ref_val or raw_val == 0:
            # Magnitude fallback: SEC filings are almost always in millions
            if abs(raw_val) < 50_000:
                return 1_000_000
            elif abs(raw_val) < 50_000_000:
                return 1_000
            return 1
        best_scale, best_err = 1_000_000, float('inf')
        for sc in [1, 1_000, 1_000_000]:
            err = abs(raw_val * sc - ref_val) / max(ref_val, 1)
            if err < best_err:
                best_err = err
                best_scale = sc
        # Only trust scale if it brings us within 50x of reference
        if best_err > 50:
            return 1_000_000  # conservative default
        return best_scale

    # Step 4: Scan tables for missing labels
    recovered = {}   # label -> raw numeric value
    for tbl in tables:
        tbl = tbl.dropna(how='all', axis=0).dropna(how='all', axis=1)
        if tbl.empty:
            continue

        for row_values in tbl.itertuples(index=False, name=None):
            cells = [str(x).strip() for x in row_values
                     if str(x) not in ('nan', 'NaN', '') and str(x).strip() not in ('$', '')]
            if not cells:
                continue

            row_text = ' '.join(cells).lower()

            for lbl in list(missing):
                if lbl in recovered:
                    continue
                patterns = RECOVERY_TARGETS[lbl]
                if not any(pat in row_text for pat in patterns):
                    continue

                # Skip partial subtotals if we are looking for the main line
                if any(x in row_text for x in ['continuing operations', 'discontinued operations', 'before changes']):
                    continue

                # Extract numeric values from the row (skip the label cell)
                nums = []
                for cell in cells[1:]:
                    # Strip currency symbols, commas, whitespace
                    clean = re.sub(r'[^\d\.\(\)\-]', '', cell)
                    if not clean:
                        continue
                    negative = '(' in cell or (cell.strip().startswith('-') and len(cell) > 1)
                    clean = clean.replace('(', '').replace(')', '').replace('-', '')
                    try:
                        v = float(clean)
                        if negative:
                            v = -v
                        if v != 0:
                            nums.append(v)
                    except ValueError:
                        continue

                if not nums:
                    continue

                # First number is the current fiscal year (most recent column)
                raw = nums[0]
                scale = _infer_scale_from_ref(raw)
                recovered[lbl] = raw * scale
                missing.discard(lbl)

        if not missing:
            break   # all targets found

    # Collect max individual quarterly value per label for plausibility check below
    _max_qtly_abs = {}
    for fact in extracted:
        if fact.get('Category') != '3_Cash_Flow' or fact.get('FY') != fy:
            continue
        dur = fact.get('Duration', 0)
        if 60 <= dur <= 120:
            lbl_ = fact['Label']
            try:
                _max_qtly_abs[lbl_] = max(_max_qtly_abs.get(lbl_, 0.0),
                                          abs(float(fact['Value'])))
            except (TypeError, ValueError):
                pass

    # Step 5: Inject synthetic annual facts
    injected = 0
    for lbl, val in recovered.items():
        # CapEx and other outflows should be positive in our convention
        if lbl == 'Capital Expenditures' and val < 0:
            val = abs(val)

        # Plausibility guard for CapEx: the HTML-parsed annual must not be wildly
        # larger than the known individual quarters.  An oversized value signals
        # that the HTML scanner matched the wrong row (e.g. balance-sheet net PP&E,
        # or OCF). Threshold: annual â‰¤ max_quarterly Ã— 6.
        #
        # Examples this correctly rejects (NVDA wrong matches):
        #   FY2020: max_q=128M, limit=768M,  injected~1674M -> REJECT
        #   FY2019: max_q=150M, limit=900M,  injected~1404M -> REJECT
        # Examples this correctly accepts:
        #   FY2025: max_q=1077M, limit=6462M, injected~3236M -> ACCEPT
        #   FY2022: max_q=473M,  limit=2838M, injected~1128M -> ACCEPT
        if lbl == 'Capital Expenditures':
            max_q = _max_qtly_abs.get(lbl, 0.0)
            if max_q > 0 and abs(val) > max_q * 6:
                print(f"  [HTML Recovery] Rejected implausible CapEx annual {val:,.0f} "
                      f"(max quarterly={max_q:,.0f}, limit={max_q*6:,.0f}) for FY{fy}")
                continue

        extracted.append({
            'Category': '3_Cash_Flow',
            'Label': lbl,
            'Value': val,
            'FY': fy,
            'Q': q_label,
            'End': end_str,
            'Duration': annual_dur,
            'Filed': filing.filing_date,
            'TagRank': 998,   # below all real XBRL tags
            'DimCount': 0,
            'IsCalculated': False,
        })
        injected += 1
        print(f"  [HTML CF Recovery] {lbl} = {val:,.0f} (annual, 10-K HTML) for FY{fy}-{q_label}")

    return extracted


def _correct_xbrl_segment_labels_from_html(extracted, facts_df, filing, ye_month):
    """
    Compares XBRL-derived segment revenue labels against the authoritative HTML
    disclosure table in the same filing.  If the XBRL labels are swapped or
    mis-mapped (a known AMD / Nvidia issue), this corrects them.

    Redesigned algorithm (v2):
    --------------------------
    1.  Collect XBRL current-period segment revenue facts (QTD for 10-Q,
        annual for 10-K).  Require â‰¥ 2 distinct segments.
    2.  Extract HTML table values using the company's own segment names as
        search targets.
    3.  Calibrate the HTML scale against the XBRL values (or total revenue).
    4.  Build an assignment matrix: for each (XBRL segment, HTML segment) pair,
        check whether the XBRL value â‰ˆ the HTML value for the *current* period
        only (col 0 for QTD / col 0 for annual).  NO prior-period columns.
    5.  Use a greedy best-match to produce a rename_map.  Only apply the map
        if the number of actual renames stays below the total segment count
        (a full rotation is still accepted, but an all-same-name match is not).
    6.  Require that â‰¥ 50 % of unambiguous segments are confirmed by the HTML
        table before applying ANY correction.  This prevents false-positive
        swaps when the HTML parser returns noisy data.
    7.  Apply the rename across ALL facts (not just the current quarter) so
        that historical restated data in the same filing is also corrected.
    """
    is_10q = '10-Q' in filing.form
    is_10k = '10-K' in filing.form
    if not (is_10q or is_10k):
        return extracted

    filing_end = pd.to_datetime(filing.period_of_report)
    fy, q = get_period_info(filing_end, ye_month)
    SEG_CATS_LOCAL = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries'}

    # -- Step 1: Collect XBRL current-period BUSINESS segment revenue facts ----
    # CRITICAL: Only use 4a_Segments_Business here. Including geographic facts
    # (4b/4c) causes extract_l2_segments_from_html to pick up country rows from
    # the geographic revenue table. A coincidental value match between, e.g.,
    # NVIDIA's Networking segment and Taiwan's geographic revenue produces a
    # false rename (Networking -> TW).  Geographic data is self-consistent and
    # never needs HTML correction -- only business segments can have swapped tags.
    # For 10-Q: use QTD (duration ~90 days).  For 10-K: use annual (~365 days).
    target_dur    = 365 if is_10k else 91
    dur_tolerance = 40   # Â±40 days
    dur_lo, dur_hi = target_dur - dur_tolerance, target_dur + dur_tolerance

    xbrl_rev_facts = []
    xbrl_all_segs  = set()
    for fact in extracted:
        if fact['Category'] != '4a_Segments_Business':             continue  # <-- ONLY business segs
        if not fact['Label'].startswith('Revenue - '):             continue
        if fact['FY'] != fy or fact['Q'] != q:                    continue
        if fact.get('DimCount', 0) == 0:                           continue
        seg_name = fact['Label'][len('Revenue - '):]
        if ' - ' in seg_name:                                      continue   # skip multi-dim
        dur = fact.get('Duration', 0)
        if not (dur_lo <= dur <= dur_hi):                          continue   # wrong period type
        try:
            xbrl_rev_facts.append({'name': seg_name, 'val': float(fact['Value']), 'dur': dur})
            xbrl_all_segs.add(seg_name)
        except Exception:
            continue

    if len(xbrl_all_segs) < 2:
        return extracted   # nothing to cross-check

    # -- Step 2: Extract HTML table values -------------------------------------
    try:
        html_data = extract_l2_segments_from_html(facts_df, list(xbrl_all_segs))
    except Exception:
        return extracted
    if not html_data:
        return extracted

    # -- Step 3: Calibrate HTML scale ------------------------------------------
    # Use the ratio of max XBRL value to max HTML col-0 value.
    # Col 0 in the HTML table always corresponds to the current reporting period.
    html_col0_max = max((abs(vals[0]) for vals in html_data.values() if vals), default=0)
    xbrl_max      = max((abs(f['val']) for f in xbrl_rev_facts), default=0)

    scale = 1
    if html_col0_max > 0 and xbrl_max > 0:
        ratio = xbrl_max / html_col0_max
        if   500_000 < ratio < 2_000_000: scale = 1_000_000
        elif     500 < ratio <     2_000: scale = 1_000
        else:                             scale = 1

    # Scaled HTML col-0 values: segment_name -> current-period value
    html_current: dict[str, float] = {}
    for h_name, vals in html_data.items():
        if not vals:
            continue
        scaled = float(vals[0]) * scale
        if scaled != 0:
            html_current[h_name] = scaled

    if not html_current:
        return extracted

    # -- Step 4: Build XBRL -> HTML assignment matrix ---------------------------
    # For each XBRL segment name, find the HTML segment whose col-0 value is
    # within TOLERANCE of the XBRL value.  Strict current-period matching only.
    TOLERANCE = 0.025   # 2.5 % relative tolerance

    # xbrl_val_map: seg_name -> value (latest/best if duplicates)
    xbrl_val_map: dict[str, float] = {}
    for f in xbrl_rev_facts:
        n, v = f['name'], f['val']
        if n not in xbrl_val_map or abs(v) > abs(xbrl_val_map[n]):
            xbrl_val_map[n] = v

    TECHNICAL_KEYWORDS = {'debt securities', 'marketable', 'available for sale',
                          'equity securities', 'investments', 'level'}

    # potential_matches: list of {x, h, score} sorted by score desc
    potential_matches = []
    for x_name, x_val in xbrl_val_map.items():
        if any(kw in x_name.lower() for kw in TECHNICAL_KEYWORDS):
            continue
        for h_name, h_val in html_current.items():
            if h_val == 0:
                continue
            rel_err = abs(x_val - h_val) / max(abs(h_val), 1)
            if rel_err < TOLERANCE:
                # Score: exact name match = 2 pts, value match = 1 pt
                score = 1 + (1 if x_name.lower() == h_name.lower() else 0)
                potential_matches.append({'x': x_name, 'h': h_name, 'score': score})

    if not potential_matches:
        return extracted

    # -- Step 5: Greedy assignment (prefer exact name matches, then score) ------
    potential_matches.sort(key=lambda m: (m['score'], m['x'].lower() == m['h'].lower()), reverse=True)
    rename_map: dict[str, str] = {}
    used_x: set[str] = set()
    used_h: set[str] = set()
    for m in potential_matches:
        if m['x'] in used_x or m['h'] in used_h:
            continue
        if m['x'] != m['h']:
            rename_map[m['x']] = m['h']
        used_x.add(m['x'])
        used_h.add(m['h'])

    if not rename_map:
        return extracted   # no actual renames needed

    # -- Step 5b: Geographic-name guard ----------------------------------------
    # A business segment (4a) must NEVER be renamed to a geographic term.
    # This catches the NVDA case where Taiwan's geographic revenue row
    # value coincidentally equals Networking's XBRL value, producing a
    # false "Networking" -> "TW" rename that corrupts the business segment.
    def _is_geo_name(name: str) -> bool:
        nl = name.lower()
        return (
            bool(_REGION_KEYWORDS_RE.search(nl))
            or nl in GEOGRAPHIC_COUNTRIES
            or nl in GEOGRAPHIC_CODES
        )

    rename_map = {k: v for k, v in rename_map.items() if not _is_geo_name(v)}
    if not rename_map:
        return extracted   # all candidates were geographic names -- abort

    # -- Step 6: Confidence gate ------------------------------------------------
    # We must have matched at least half the segments.  This prevents a situation
    # where only one segment matched and we blindly rename it.
    n_matched   = len(used_x)
    n_total_seg = len(xbrl_all_segs)
    if n_matched < max(2, n_total_seg // 2):
        return extracted   # not enough evidence

    # -- Step 7: Handle displaced segments (rotation chains) -------------------
    # If segment A -> B via the rename, and B was an XBRL segment that wasn't
    # matched yet, try to find a HTML name for B as well.
    final_mapping = rename_map.copy()
    changed = True
    while changed:
        changed = False
        for target in list(final_mapping.values()):
            if target in xbrl_all_segs and target not in final_mapping:
                # Find best HTML match for the displaced segment
                best_h, best_score = None, 0
                already_used_h = set(final_mapping.values())
                for m in potential_matches:
                    if m['x'] == target and m['h'] not in already_used_h:
                        if m['score'] > best_score:
                            best_score = m['score']
                            best_h = m['h']
                if best_h:
                    final_mapping[target] = best_h
                else:
                    # No HTML match found -- mark with suffix to avoid collision
                    final_mapping[target] = f"{target} (Original)"
                changed = True

    # -- Apply the rename to ALL facts in this filing ---------------------------
    print(f"  [Segment HTML Fix] XBRL label swap corrected in {filing.form} FY{fy}-{q}:")
    for old, new in sorted(final_mapping.items()):
        print(f"    '{old}' -> '{new}'")

    renamed_count = 0
    for fact in extracted:
        if fact['Category'] not in SEG_CATS_LOCAL:
            continue
        parts = fact['Label'].split(' - ', 1)
        if len(parts) != 2:
            continue
        metric, seg_name = parts
        if seg_name in final_mapping:
            fact['Label'] = f"{metric} - {final_mapping[seg_name]}"
            renamed_count += 1

    if renamed_count:
        print(f"    [OK] Renamed {renamed_count} facts.")
    return extracted

def _presplit_segment_era_facts(df):
    if '_Filed_dt' not in df.columns:
        df['_Filed_dt'] = pd.to_datetime(df['Filed'], errors='coerce')
    SEG_CATS_LOCAL = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries'}
    seg_rev_mask = (df['Category'].isin(SEG_CATS_LOCAL) & df['Label'].str.startswith('Revenue - ') & (df['Duration'] >= 350) & (df['Duration'] <= 385))
    if not seg_rev_mask.any(): return df
    annual_seg = df[seg_rev_mask].copy()
    def is_single_component(label):
        seg_part = label.split(' - ', 1)[1] if ' - ' in label else ''
        return ' - ' not in seg_part
    annual_seg = annual_seg[annual_seg['Label'].apply(is_single_component)]
    if annual_seg.empty: return df
    divergences = []
    for (label, end), group in annual_seg.groupby(['Label', 'End']):
        if len(group) < 2: continue
        group_sorted = group.sort_values('_Filed_dt')
        found_divergence = False
        for i in range(len(group_sorted) - 1):
            old_row = group_sorted.iloc[i]
            for j in range(i + 1, len(group_sorted)):
                new_row = group_sorted.iloc[j]
                if (new_row['_Filed_dt'] - old_row['_Filed_dt']).days < 180: continue
                old_val, new_val = float(old_row['Value']), float(new_row['Value'])
                if old_val == 0: continue
                if abs(new_val - old_val) / abs(old_val) < 0.10: continue
                ratio = new_val / old_val
                if ratio < 0.33 or ratio > 3.0: continue
                seg_name = label.split(' - ', 1)[1] if ' - ' in label else label
                divergences.append({'label': label, 'seg_name': seg_name, 'end': end, 'old_filed_dt': old_row['_Filed_dt'], 'new_filed_dt': new_row['_Filed_dt'], 'old_val': old_val, 'new_val': new_val})
                found_divergence = True; break
            if found_divergence: break
    if len(divergences) < 2: return df
    from collections import defaultdict
    by_end = defaultdict(list)
    for d in divergences: by_end[d['end']].append(d)
    qualifying_divergences = []
    for end, end_divs in by_end.items():
        # Require â‰¥ 2 segments to diverge at the SAME period end date before
        # treating it as a real segment restructuring event.  A single-segment
        # divergence is more likely caused by a 10-K restatement of one small
        # line item and should NOT split the entire time series.
        if len(end_divs) >= 2:
            qualifying_divergences.extend(end_divs)
    if not qualifying_divergences: return df
    boundary_filed_dt = max(d['old_filed_dt'] for d in qualifying_divergences)
    affected_seg_names = {d['seg_name'] for d in qualifying_divergences}
    print(f"  [Segment Era Split] Detected segment definition change affecting {len(affected_seg_names)} segments")
    def is_affected(row):
        if row['Category'] not in SEG_CATS_LOCAL: return False
        label = row['Label']
        if ' - ' not in label: return False
        return label.split(' - ', 1)[1] in affected_seg_names
    split_mask = df.apply(is_affected, axis=1)
    if not split_mask.any(): return df
    df = df.copy()
    df.loc[split_mask & (df['_Filed_dt'] <= boundary_filed_dt), 'Label'] += ' (Pre Change)'
    df.loc[split_mask & (df['_Filed_dt'] > boundary_filed_dt), 'Label'] += ' (Post Change)'
    return df

# Above this many distinct members, a (category, prefix) group is a granular
# breakdown (tenant industries, dated debt instruments, per-property rows...),
# not a handful of segments that got renamed. Pairwise rename/merge detection
# there is both meaningless and O(K^2)/O(R*A), so it is skipped -- this keeps
# filers with very high member cardinality (e.g. Realty Income) from hanging.
_SEGMENT_GROUP_CAP = 100


def _reconcile_segment_labels(df, ticker=None, company_name=None):
    from collections import defaultdict
    import numpy as np

    # -- Pre-Layer: Rescue Income Statement dynamically learned segment items
    STANDARD_IS_LABELS = {
        'Revenue', 'Cost of Revenue', 'Gross Profit', 'Premiums Earned', 
        'Net Investment Income', 'Policyholder Claims/Benefits', 'Amortization of DAC', 
        'Research & Development', 'Selling, General & Admin', 'Sales & Marketing', 
        'General & Administrative', 'Customer Bad Debt', 'Salaries & Employee Benefits', 
        'Marketing Expense', 'Amortization of Intangibles', 'Restructuring & Related Charges', 
        'Impairment Charges', 'Acquisition-Related Costs', 'Litigation & Settlement Charges', 
        'Total Operating Expenses', 'Operating Income', 'Interest Expense', 'Interest Income', 
        'Net Interest Income (Expense)', 'Gain/Loss on Investments', 'Equity Method Income', 
        'Other Income / (Expense)', 'Total Non-operating Income', 'Pretax Income', 
        'Income Tax Expense', 'Net Income'
    }
    
    is_mask = df['Category'] == '1_Income_Statement'
    seg_4a_mask = df['Category'] == '4a_Segments_Business'
    if is_mask.any() and seg_4a_mask.any():
        is_labels = df.loc[is_mask, 'Label'].unique()
        seg_labels = df.loc[seg_4a_mask, 'Label'].unique()
        safe = {'services', 'service', 'products', 'product', 'net', 'gross', 'segment', 'segments', 'member', 'inc', 'co', 'corp', 'company', 'operating', 'reporting', 'concentration', 'risk', 'geographic', 'and', 'revenue', 'revenues', 'fee', 'fees'}
        
        parsed_segs = []
        for l in seg_labels:
            parts = l.split(' - ', 1)
            if len(parts) == 2:
                parsed_segs.append((parts[0], set(parts[1].lower().split()), l))
                
        _rescue_pvals_cache = {}

        def _rescue_pvals(lbl, cat_name):
            # newest-filed-first value per period (df is sorted Filed desc).
            # Cache repeated label/category scans during duplicate verification;
            # clear this cache after any in-place label/category move below.
            cache_key = (lbl, cat_name)
            cached = _rescue_pvals_cache.get(cache_key)
            if cached is not None:
                return cached
            sub = df[(df['Label'] == lbl) & (df['Category'] == cat_name)]
            out = {}
            for r in sub.itertuples(index=False):
                key = (r.FY, r.Q, r.Duration)
                if key not in out and pd.notna(r.Value):
                    try:
                        out[key] = float(r.Value)
                    except (TypeError, ValueError):
                        pass
            _rescue_pvals_cache[cache_key] = out
            return out

        for l in is_labels:
            if l in STANDARD_IS_LABELS: continue
            
            t1 = set(l.lower().split())
            t1_clean = t1 - safe
            if not t1_clean and not t1: continue
            
            matched_prefix = None
            for prefix, seg_t_raw, seg_full in parsed_segs:
                # Disclosure-style metrics (Stock-Based Compensation,
                # Acquisition-Related Costs, Useful Life...) re-use face
                # captions as members -- never pull face lines into those.
                if prefix not in GENUINE_SEGMENT_METRICS:
                    continue
                t2 = seg_t_raw
                t2_clean = t2 - safe
                if (t1_clean == t2_clean and len(t1_clean) > 0) or (t1 < t2 and (t2 - t1) <= safe) or (t2 < t1 and (t1 - t2) <= safe):
                    # Value gate: only a provable duplicate may move.
                    fp = _rescue_pvals(l, '1_Income_Statement')
                    sp = _rescue_pvals(seg_full, '4a_Segments_Business')
                    common = [k for k in fp if k in sp]
                    if len(common) < 3:
                        continue
                    if any(abs(fp[k] - sp[k]) > max(1e-6, 1e-4 * abs(sp[k])) for k in common):
                        continue
                    if any(k not in sp for k in fp):
                        continue   # face row has data the segment row lacks
                    matched_prefix = prefix
                    break
                    
            if matched_prefix:
                print(f"  [Segment Rescue] Moving IS line '{l}' to 4a_Segments_Business as '{matched_prefix} - {l}' (value-verified duplicate)")
                mask = is_mask & (df['Label'] == l)
                df.loc[mask, 'Category'] = '4a_Segments_Business'
                df.loc[mask, 'Label'] = f"{matched_prefix} - {l}"
                _rescue_pvals_cache.clear()
                
    SEG_CATS_LOCAL = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries'}
    seg_mask = df['Category'].isin(SEG_CATS_LOCAL)
    if not seg_mask.any(): return df

    def _norm_label(lbl: str) -> str:
        return _normalize_member_label(lbl)

    unique_labels_0 = df[seg_mask][['Category', 'Label']].drop_duplicates()
    norm_groups: dict = defaultdict(list)
    for r in unique_labels_0.itertuples(index=False):
        key = (r.Category, _norm_label(r.Label))
        norm_groups[key].append(r.Label)

    rename_map_0: dict = {}
    for (cat, norm_key), lbls in norm_groups.items():
        if len(lbls) < 2: continue
        unique_lbls = list(dict.fromkeys(lbls)) 
        if len(unique_lbls) < 2: continue
        def _recency(lbl):
            subset = df[(df['Category'] == cat) & (df['Label'] == lbl)]
            return pd.to_datetime(subset['Filed']).max() if not subset.empty else pd.Timestamp.min
        canonical = max(unique_lbls, key=lambda l: (_recency(l), -len(l)))
        for lbl in unique_lbls:
            if lbl != canonical: rename_map_0[lbl] = canonical

    if rename_map_0:
        for old_l, new_l in rename_map_0.items():
            print(f"  [Label Norm - Abbrev Variant] '{old_l}' -> '{new_l}'")
        df = df.copy()
        df['Label'] = df['Label'].replace(rename_map_0)

    # -------------------------------------------------------------------------
    # PERFORMANCE FIX: Pre-compute label lookups to avoid O(N) Pandas masking
    # -------------------------------------------------------------------------
    label_pvals = {}
    label_maxf = {}
    
    def refresh_lookups():
        label_pvals.clear()
        label_maxf.clear()
        for lbl, grp in df[df['Category'].isin(SEG_CATS_LOCAL)].groupby('Label'):
            pvals = {}
            for r in grp.itertuples(index=False):
                key = (r.FY, r.Q, r.Duration)
                # CRITICAL FIX: Only set if not exists to mimic original df.iloc[0]
                # Since df is sorted by _Filed_dt DESCENDING, the first row 
                # encountered is the newest/highest priority fact.
                if key not in pvals:
                    v = r.Value
                    pvals[key] = float(v) if pd.notna(v) else np.nan
            label_pvals[lbl] = pvals
            label_maxf[lbl] = pd.to_datetime(grp['Filed']).max()

    refresh_lookups()

    # --- Layer 1: 2-part label dynamic merge ---
    unique_labels = df[df['Category'].isin(SEG_CATS_LOCAL)][['Category', 'Label']].drop_duplicates()
    parsed = []
    for r in unique_labels.itertuples(index=False):
        parts = r.Label.split(' - ', 1)
        if len(parts) == 2: parsed.append((r.Category, parts[0], parts[1], r.Label))
        
    groups = defaultdict(list)
    for cat, prefix, suffix, lbl in parsed: groups[(cat, prefix)].append((suffix, lbl))
    rename_map = {}
    
    for (cat, prefix), items in groups.items():
        if len(items) < 2 or len(items) > _SEGMENT_GROUP_CAP: continue
        for i in range(len(items)):
            for j in range(i+1, len(items)):
                s1, l1 = items[i]; s2, l2 = items[j]
                if l1 in rename_map or l2 in rename_map: continue
                if l1.count(' - ') >= 2 or l2.count(' - ') >= 2: continue
                
                pvals1, pvals2 = label_pvals.get(l1, {}), label_pvals.get(l2, {})
                common = set(pvals1.keys()) & set(pvals2.keys())
                
                has_conflict = False
                for p in common:
                    v1, v2 = pvals1[p], pvals2[p]
                    if pd.notna(v1) and pd.notna(v2) and abs(v1 - v2) > 1e-6:
                        has_conflict = True; break
                if has_conflict: continue
                
                t1, t2 = set(s1.lower().split()), set(s2.lower().split())
                safe = {'services', 'service', 'products', 'product', 'net', 'gross', 'segment', 'segments', 'member', 'inc', 'co', 'corp', 'company', 'operating', 'reporting', 'concentration', 'risk', 'geographic', 'and', 'revenue', 'revenues', 'fee', 'fees'}
                t1_clean = t1 - safe
                t2_clean = t2 - safe
                if (t1_clean == t2_clean and len(t1_clean) > 0) or (t1 < t2 and (t2 - t1) <= safe) or (t2 < t1 and (t1 - t2) <= safe):
                    max_filed_1 = label_maxf.get(l1, pd.Timestamp.min)
                    max_filed_2 = label_maxf.get(l2, pd.Timestamp.min)
                    if max_filed_2 >= max_filed_1: rename_map[l1] = l2
                    else: rename_map[l2] = l1
                    
    if rename_map:
        for old_l, new_l in rename_map.items(): print(f"  [Dynamic Label Merge] Mapping '{old_l}' -> '{new_l}'")
        df['Label'] = df['Label'].replace(rename_map)
        refresh_lookups()

    # --- Layer 1c: Company-Name-Aware Segment Merge ---
    _company_tokens = set()
    MERGE_FILLER_SAFE = {'and', 'the', 'of', 'for', 'in', 'segment', 'segments', 'inc', 'corp', 'company', 'group', 'services', 'service', 'revenue', 'revenues', 'fee', 'fees'}
    _co_noise = {'inc', 'corp', 'co', 'ltd', 'company', 'corporation', 'incorporated', 'holdings', 'holding', 'group', 'the', 'plc', 'sa', 'nv', 'ag', 'se', 'lp', 'llc', 'limited'}
    if ticker: _company_tokens.add(ticker.lower())
    if company_name:
        for token in re.sub(r'[^\w\s]', ' ', company_name.lower()).split():
            if token not in _co_noise and len(token) > 1: _company_tokens.add(token)

    if _company_tokens:
        merge_safe_1c = _company_tokens | MERGE_FILLER_SAFE
        unique_labels_1c = df[df['Category'].isin(SEG_CATS_LOCAL)][['Category', 'Label']].drop_duplicates()
        groups_1c = defaultdict(list)
        for r in unique_labels_1c.itertuples(index=False):
            parts = r.Label.split(' - ', 1)
            if len(parts) == 2: groups_1c[(r.Category, parts[0])].append(r.Label)

        rename_map_1c = {}
        for (cat, prefix), labels in groups_1c.items():
            if len(labels) < 2 or len(labels) > _SEGMENT_GROUP_CAP: continue
            for i in range(len(labels)):
                for j in range(i + 1, len(labels)):
                    l1, l2 = labels[i], labels[j]
                    if l1 in rename_map_1c or l2 in rename_map_1c: continue

                    parts1, parts2 = l1.split(' - ')[1:], l2.split(' - ')[1:]
                    if len(parts1) != len(parts2): continue

                    diff_indices = [k for k in range(len(parts1)) if parts1[k].lower() != parts2[k].lower()]
                    if len(diff_indices) != 1: continue

                    di = diff_indices[0]
                    t1, t2 = set(parts1[di].lower().split()), set(parts2[di].lower().split())
                    intersection, sym_diff = t1 & t2, t1.symmetric_difference(t2)

                    if len(intersection) < 2 or len(sym_diff) > 3 or len(sym_diff) == 0: continue
                    if not sym_diff <= merge_safe_1c: continue

                    pvals1, pvals2 = label_pvals.get(l1, {}), label_pvals.get(l2, {})
                    common = set(pvals1.keys()) & set(pvals2.keys())

                    has_conflict = False
                    for p in common:
                        v1, v2 = pvals1[p], pvals2[p]
                        if pd.notna(v1) and pd.notna(v2) and v1 != 0 and v2 != 0:
                            if abs(v1 - v2) / max(abs(v1), abs(v2)) > 0.01:
                                has_conflict = True; break
                    if has_conflict: continue

                    max_filed_1 = label_maxf.get(l1, pd.Timestamp.min)
                    max_filed_2 = label_maxf.get(l2, pd.Timestamp.min)
                    if max_filed_2 >= max_filed_1: rename_map_1c[l1] = l2
                    else: rename_map_1c[l2] = l1

        if rename_map_1c:
            for old_l, new_l in rename_map_1c.items(): print(f"  [Label Merge - Company Name Variant] '{old_l}' -> '{new_l}'")
            df['Label'] = df['Label'].replace(rename_map_1c)
            refresh_lookups()

    # --- Layer 1b: 3-part label qualifier merge ---
    QUALIFIER_SAFE = {'mobile', 'enterprise', 'consumer', 'commercial', 'professional', 'digital', 'advanced', 'legacy', 'standard', 'traditional', 'core', 'base'}
    unique_labels_1b = df[df['Category'].isin(SEG_CATS_LOCAL)][['Category', 'Label']].drop_duplicates()
    three_part_groups = defaultdict(list)
    for r in unique_labels_1b.itertuples(index=False):
        parts = r.Label.split(' - ')
        if len(parts) == 3:
            three_part_groups[(r.Category, parts[0], parts[2])].append((parts[1], r.Label))

    rename_map_1b = {}
    for (cat, prefix, suffix), items in three_part_groups.items():
        if len(items) < 2 or len(items) > _SEGMENT_GROUP_CAP: continue
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                mid1, l1 = items[i]; mid2, l2 = items[j]
                if l1 in rename_map_1b or l2 in rename_map_1b: continue
                
                pvals1, pvals2 = label_pvals.get(l1, {}), label_pvals.get(l2, {})
                common = set(pvals1.keys()) & set(pvals2.keys())
                has_conflict = False
                for p in common:
                    v1, v2 = pvals1[p], pvals2[p]
                    if pd.notna(v1) and pd.notna(v2) and abs(v1 - v2) > 1e-6:
                        has_conflict = True; break
                if has_conflict: continue
                
                t1, t2 = set(mid1.lower().split()), set(mid2.lower().split())
                if t1 < t2 and (t2 - t1) <= QUALIFIER_SAFE: rename_map_1b[l2] = l1
                elif t2 < t1 and (t1 - t2) <= QUALIFIER_SAFE: rename_map_1b[l1] = l2

    if rename_map_1b:
        for old_l, new_l in rename_map_1b.items(): print(f"  [Dynamic Label Merge - Sub-Segment Qualifier] '{old_l}' -> '{new_l}'")
        df['Label'] = df['Label'].replace(rename_map_1b)
        refresh_lookups()

    # --- Layer 2: Value continuity ---
    unique_labels2 = df[df['Category'].isin(SEG_CATS_LOCAL)][['Category', 'Label']].drop_duplicates()
    groups2 = defaultdict(list)
    for r in unique_labels2.itertuples(index=False):
        parts = r.Label.split(' - ', 1)
        if len(parts) == 2 and ' - ' not in parts[1]: groups2[(r.Category, parts[0])].append((parts[1], r.Label))
        
    rename_map2 = {}
    for (cat, prefix), items2 in groups2.items():
        if len(items2) < 2 or len(items2) > _SEGMENT_GROUP_CAP: continue
        for i in range(len(items2)):
            for j in range(i+1, len(items2)):
                s1, l1 = items2[i]; s2, l2 = items2[j]
                if l1 in rename_map2 or l2 in rename_map2: continue
                
                pvals1, pvals2 = label_pvals.get(l1, {}), label_pvals.get(l2, {})
                l1_periods, l2_periods = set(pvals1.keys()), set(pvals2.keys())
                common = l1_periods & l2_periods
                if not common: continue

                # Distinct dated instruments (notes/bonds with different
                # maturities) commonly carry identical par balances and would
                # otherwise be fused on value continuity alone. If both suffixes
                # name years and those years are disjoint, they are different
                # instruments -- never merge.
                _yr1 = set(re.findall(r'(?:19|20)\d{2}', s1))
                _yr2 = set(re.findall(r'(?:19|20)\d{2}', s2))
                if _yr1 and _yr2 and _yr1.isdisjoint(_yr2):
                    continue
                
                non_zero_common = 0
                all_match = True
                for p in common:
                    v1, v2 = pvals1[p], pvals2[p]
                    
                    # Exact behavioral match of original code's float/NaN coercion
                    if pd.notna(v1) and pd.notna(v2):
                        if v1 != 0 and v2 != 0:
                            non_zero_common += 1
                            with np.errstate(invalid='ignore', divide='ignore'):
                                denom = max(abs(v1), abs(v2))
                                if denom != 0 and (abs(v1 - v2) / denom) > 0.0001:
                                    all_match = False; break
                        elif (v1 == 0 and v2 != 0) or (v1 != 0 and v2 == 0):
                            all_match = False; break
                            
                if all_match and non_zero_common >= 3:
                    overlap_ratio = non_zero_common / max(len(l1_periods), len(l2_periods))
                    if overlap_ratio >= 0.20:
                        max_filed_1 = label_maxf.get(l1, pd.Timestamp.min)
                        max_filed_2 = label_maxf.get(l2, pd.Timestamp.min)
                        if max_filed_2 >= max_filed_1: rename_map2[l1] = l2
                        else: rename_map2[l2] = l1
                        
    if rename_map2:
        for old_l, new_l in rename_map2.items(): print(f"  [Dynamic Label Merge - Value Continuity] Mapping '{old_l}' -> '{new_l}'")
        df['Label'] = df['Label'].replace(rename_map2)

    return df

def _detect_and_merge_renamed_segments(df: pd.DataFrame) -> pd.DataFrame:
    """
    Intelligently detect and merge segment renaming events (e.g., CAT renaming
    'Energy & Transportation' to 'Power & Energy').

    If a segment was renamed, the latest filing will have history under the NEW name.
    Older filings will have history under the OLD name. They will have overlapping
    historical periods with highly similar (or identical) values.
    """
    import datetime

    # 1. Isolate segment lines across all segment categories
    seg_mask = df['Category'].isin({'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'})
    if not seg_mask.any():
        return df

    seg_df = df[seg_mask].copy()
    if '_Filed_dt' not in seg_df.columns:
        seg_df['_Filed_dt'] = pd.to_datetime(seg_df['Filed'], errors='coerce')

    global_max_date = seg_df['_Filed_dt'].max()
    if pd.isna(global_max_date):
        return df

    # 2. Determine Active vs Retired segments
    seg_lifespans = seg_df.groupby('Label')['_Filed_dt'].max()
    cutoff_date = global_max_date - pd.Timedelta(days=90)

    active_segs = seg_lifespans[seg_lifespans >= cutoff_date].index.tolist()
    retired_segs = seg_lifespans[seg_lifespans < cutoff_date].index.tolist()

    if not active_segs or not retired_segs:
        return df

    def _tokenize(name):
        base = name.split(' - ', 1)[1] if ' - ' in name else name
        tokens = set(re.findall(r'[a-z0-9]+', base.lower()))
        return tokens - {'and', 'or', 'the', 'of', 'in', 'segment', 'segments', 'other', 'reportable', 'subsegments', 'excluding', 'intersegment', 'eliminations', 'elimination'}

    # Pre-compute everything per-LABEL in single passes. Doing these lookups
    # inside the retired x active loop (a per-pair seg_df scan) is O(R*A*N) and
    # hangs on filers with hundreds of segment-category labels (e.g. Realty
    # Income's tenant-industry / debt-instrument breakdowns).
    _all_seg_labels = set(active_segs) | set(retired_segs)
    _nz = seg_df[seg_df['Value'].notna() & (seg_df['Value'] != 0)]
    period_vals = {
        lbl: grp.groupby(['FY', 'Q'])['Value'].median().to_dict()
        for lbl, grp in _nz.groupby('Label') if lbl in _all_seg_labels
    }
    label_filings = {
        lbl: set(grp['Filed'].dropna())
        for lbl, grp in seg_df.groupby('Label') if lbl in _all_seg_labels
    }
    label_tokens = {lbl: _tokenize(lbl) for lbl in _all_seg_labels}
    label_prefix = {lbl: (lbl.split(' - ')[0] if ' - ' in lbl else lbl)
                    for lbl in _all_seg_labels}
    _active_by_prefix = {}
    for _a in active_segs:
        _active_by_prefix.setdefault(label_prefix[_a], []).append(_a)

    rename_map = {}

    # 3. Compare overlaps within a metric prefix (a rename never crosses metrics)
    for retired in retired_segs:
        ret_prefix = label_prefix[retired]
        candidates = _active_by_prefix.get(ret_prefix)
        if not candidates:
            continue
        if len(candidates) > _SEGMENT_GROUP_CAP:
            continue
        ret_vals = period_vals.get(retired, {})
        ret_tokens = label_tokens[retired]
        ret_filings = label_filings.get(retired, set())

        best_match = None
        best_diff = float('inf')

        for active in candidates:
            if active == retired:
                continue
            act_filings = label_filings.get(active, set())
            if ret_filings & act_filings:
                continue

            act_vals = period_vals.get(active, {})
            act_tokens = label_tokens[active]

            common_periods = ret_vals.keys() & act_vals.keys()

            if len(common_periods) >= 2:
                diffs = []
                for p in common_periods:
                    v_ret = ret_vals[p]
                    v_act = act_vals[p]
                    denom = max(abs(v_ret), abs(v_act))
                    if denom == 0:
                        diffs.append(0)
                    else:
                        diffs.append(abs(v_ret - v_act) / denom)

                avg_diff = sum(diffs) / len(diffs)
                jaccard = len(ret_tokens & act_tokens) / max(len(ret_tokens | act_tokens), 1)

                # A genuine rename re-reports the SAME number under a new member
                # name, so require BOTH near-exact values AND a real name
                # relationship: one token set contained in the other (a word was
                # added/removed, e.g. "Government" -> "Government Operating") or
                # high overlap. Never merge on numbers alone -- two distinct
                # same-prefix segments ("US Commercial" vs "US Government") can
                # sit within a few percent for a couple of quarters and would
                # otherwise be silently fused.
                if ret_tokens and act_tokens:
                    _subset = ret_tokens <= act_tokens or act_tokens <= ret_tokens
                    if (_subset or jaccard >= 0.6) and avg_diff < 0.05:
                        if avg_diff < best_diff:
                            best_diff = avg_diff
                            best_match = active

        if best_match:
            rename_map[retired] = best_match
            print(f"  [Segment Rename] Merging retired '{retired}' -> '{best_match}' (Avg Diff: {best_diff:.1%})")

    # 4. Apply mapping globally to all metrics and categories
    if rename_map:
        suffix_map = {}
        for old_l, new_l in rename_map.items():
            if ' - ' in old_l and ' - ' in new_l:
                old_suf = old_l.split(' - ', 1)[1]
                new_suf = new_l.split(' - ', 1)[1]
                if len(old_suf) > 4:
                    suffix_map[old_suf] = new_suf

        _rename_targets = set(rename_map.values())

        def rename_label(lbl):
            if lbl in rename_map:
                return rename_map[lbl]

            # Never re-process an active label that is itself a rename target,
            # and only propagate a suffix to OTHER metrics on an EXACT match.
            # Substring replacement double-appends when the old suffix is a
            # prefix of the new one ("Government" -> "Government Operating"
            # would turn the active "Government Operating" into "Government
            # Operating Operating", splitting the very series being merged).
            if lbl in _rename_targets:
                return lbl

            if ' - ' in lbl:
                parts = lbl.split(' - ')
                prefix = parts[0]
                rest = ' - '.join(parts[1:])
                for old_suf in sorted(suffix_map.keys(), key=len, reverse=True):
                    if rest == old_suf:
                        return f"{prefix} - {suffix_map[old_suf]}"
            return lbl

        df['Label'] = df['Label'].apply(rename_label)

    return df


def _merge_concurrent_member_variants(df: pd.DataFrame) -> pd.DataFrame:
    """Unify one segment member that a filer tags under two slightly different
    names in different filing types.

    Amazon is the motivating case: advertising revenue is tagged
    'AdvertisingServicesMember' in the 10-Q but 'AdvertisingMember' in the 10-K,
    so one real series is split into two labels -- 'Revenue - Advertising
    Services' carrying the quarterly facts and 'Revenue - Advertising' carrying
    the annual ones. The annual therefore never lands in the quarterly label's
    fact pool, the Q4 = Annual - 9M derivation in build_pivoted_data finds no
    annual, and Q4 goes blank for every year.

    This uses a dynamic, morphologically-aware subset matcher:
    1. Stems words to handle plurals/suffixes (services -> service).
    2. Dynamically learns company-specific generic words (any word appearing 
       in >= 40% of the company's segments is treated as generic filler).
    3. Merges when the distinctive tokens of one are a STRICT subset of the 
       other, differing ONLY by generic descriptors.
    """
    SEG = ('1_Income_Statement', '4a_Segments_Business', '4b_Segments_Geographic_Regions',
           '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated')
    if 'FY' not in df.columns or 'Q' not in df.columns:
        return df
    seg_mask = df['Category'].isin(SEG) & df['Label'].str.startswith('Revenue - ')
    if not seg_mask.any():
        return df
    seg_df = df[seg_mask]
    labels = list(seg_df['Label'].unique())

    def _stem(word):
        w = word.lower()
        if w.endswith('ies') and len(w) > 4: 
            w = w[:-3] + 'y'
        elif w.endswith('es') and len(w) > 3 and w[-3] in 'sxzh': 
            w = w[:-2]
        elif w.endswith('s') and len(w) > 3 and w[-2] not in 'su': 
            w = w[:-1]
        if w.endswith('ing') and len(w) > 5: 
            w = w[:-3]
        elif w.endswith('ion') and len(w) > 5: 
            w = w[:-3]
        elif w.endswith('ed') and len(w) > 4:
            w = w[:-2]
        return w

    def _toks(lbl):
        base = lbl.split(' - ', 1)[1] if ' - ' in lbl else lbl
        return set(_stem(w) for w in re.findall(r'[a-z0-9]+', base.lower()))

    tok = {l: _toks(l) for l in labels}
    
    # Core generic terms across all accounting
    core_generic = {
        'service', 'net', 'gross', 'revenue', 'sale', 'fee', 'income', 
        'segment', 'member', 'total', 'operat', 'divis', 'group', 'unit', 
        'report', 'business', 'product', 'relat'
    }
    vague_tokens = {'other', 'misc', 'miscellaneous', 'rest', 'all', 'unallocat', 'eliminat', 'corporate'}

    # Dynamically discover company-specific generic words
    all_toks = [w for t_set in tok.values() for w in t_set]
    from collections import Counter
    tok_counts = Counter(all_toks)
    num_labels = len(labels)
    dynamic_generic = set(core_generic)
    if num_labels > 2:
        for w, c in tok_counts.items():
            if c / num_labels >= 0.4:
                dynamic_generic.add(w)

    pv = {}
    for lbl in labels:
        l = seg_df[(seg_df['Label'] == lbl) & seg_df['Value'].notna() & (seg_df['Value'] != 0)]
        pv[lbl] = l.groupby(['FY', 'Q'])['Value'].median().to_dict()

    def _conflicts(a, b):
        for p in set(pv[a]) & set(pv[b]):
            va, vb = pv[a][p], pv[b][p]
            denom = max(abs(va), abs(vb))
            if denom > 0 and abs(va - vb) / denom > 0.10:
                return True
        return False

    rename_map = {}
    for a in labels:
        ta = tok[a]
        if not ta or not (ta - dynamic_generic - vague_tokens):
            continue   # need a real distinctive token
        best = None
        for b in labels:
            if b == a:
                continue
            tb = tok[b]
            if ta < tb and (tb - ta) <= dynamic_generic and not _conflicts(a, b):
                if best is None or len(tok[b]) > len(tok[best]):
                    best = b
        if best is not None:
            rename_map[a] = best

    def _resolve(x):
        seen = set()
        while x in rename_map and x not in seen:
            seen.add(x)
            x = rename_map[x]
        return x
    rename_map = {a: _resolve(b) for a, b in rename_map.items() if _resolve(b) != a}
    if not rename_map:
        return df

    for a, b in rename_map.items():
        print(f"  [Member Variant] Unifying '{a}' -> '{b}' "
              f"(dynamic intelligent match)")
    suffix_map = {a.replace('Revenue - ', '', 1): b.replace('Revenue - ', '', 1)
                  for a, b in rename_map.items()}

    def rename_label(lbl):
        if ' - ' not in lbl:
            return lbl
        prefix, _, rest = lbl.partition(' - ')
        if rest in suffix_map:
            return f"{prefix} - {suffix_map[rest]}"
        return lbl
    df['Label'] = df['Label'].apply(rename_label)
    return df



_ARELLE_CACHE = {}


def arelle_update_concept_map(filing_url, foreign=False):
    print(f"  [Arelle Pre-Pass] Learning custom tags from {filing_url}...")
    try:
        from arelle import CntlrCmdLine, XbrlConst
        ctrl = CntlrCmdLine.CntlrCmdLine()
        ctrl.startLogging(logFileName=os.devnull)
        modelXbrl = ctrl.modelManager.load(filing_url)
        relSet = modelXbrl.relationshipSet(XbrlConst.summationItem)
        
        tree_map = {}
        for rel in relSet.modelRelationships:
            parent_name = rel.fromModelObject.qname.localName if rel.fromModelObject and hasattr(rel.fromModelObject, "qname") else ""
            child_name = rel.toModelObject.qname.localName if rel.toModelObject and hasattr(rel.toModelObject, "qname") else ""
            if parent_name and child_name:
                if parent_name not in tree_map:
                    tree_map[parent_name] = []
                tree_map[parent_name].append({'concept': child_name, 'weight': rel.weight})

        # FALLBACK: Also parse the Presentation Linkbase (Visual Tree) to discover orphaned tags
        # that were removed from the math tree but still grouped under logical abstract headers!
        relSet_pres = modelXbrl.relationshipSet(XbrlConst.parentChild)
        for rel in relSet_pres.modelRelationships:
            parent_name = rel.fromModelObject.qname.localName if rel.fromModelObject and hasattr(rel.fromModelObject, "qname") else ""
            child_name = rel.toModelObject.qname.localName if rel.toModelObject and hasattr(rel.toModelObject, "qname") else ""
            if parent_name and child_name:
                # Abstract presentation headers often append "Abstract", "LineItems", or "Table"
                clean_parent = parent_name.replace('Abstract', '').replace('LineItems', '').replace('Table', '')
                if clean_parent not in tree_map:
                    tree_map[clean_parent] = []
                
                # Check if we already found this child via Calculation Linkbase
                already_mapped = any(x['concept'] == child_name for x in tree_map[clean_parent])
                if not already_mapped:
                    # We don't have mathematical weights in the presentation tree, 
                    # but we assign a positive weight (1.0) so it successfully triggers the categorization matcher below!
                    tree_map[clean_parent].append({'concept': child_name, 'weight': 1.0})
        
        # We map standard categories to a LIST of parent/abstract tags they roll up into.
        # This covers Calculation Linkbases and Visual Presentation Linkbases.
        _LABEL_TO_PARENT_CONCEPT = {
            # --- Income Statement ---
            'Revenue': ['GrossProfit', 'OperatingIncomeLoss', 'Revenues'],
            'Cost of Revenue': ['GrossProfit', 'OperatingIncomeLoss', 'CostOfRevenue'],
            'Total Operating Expenses': ['OperatingIncomeLoss', 'OperatingExpenses'],
            'Operating Income': ['NetIncomeLoss', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxes'],
            'Net Income': ['NetIncomeLoss', 'NetIncomeLossAvailableToCommonStockholdersBasic', 'StatementOfIncomeAndComprehensiveIncome'],
            'Interest Income': ['InterestIncomeExpenseNet', 'NonoperatingIncomeExpense'],
            'Interest Expense': ['InterestIncomeExpenseNet', 'NonoperatingIncomeExpense'],
            'Research & Development': ['OperatingExpenses'],
            'Selling, General & Admin': ['OperatingExpenses'],
            
            # --- Cash Flow ---
            'Operating Cash Flow': ['NetCashProvidedByUsedInOperatingActivities'],
            'Investing Cash Flow': ['NetCashProvidedByUsedInInvestingActivities'],
            'Financing Cash Flow': ['NetCashProvidedByUsedInFinancingActivities'],
            'Capital Expenditures': ['NetCashProvidedByUsedInInvestingActivities'],
            'Share Repurchases': ['NetCashProvidedByUsedInFinancingActivities'],
            'Dividends Paid': ['NetCashProvidedByUsedInFinancingActivities'],
            
            # --- Balance Sheet ---
            'Total Assets': ['Assets'],
            'Total Current Assets': ['AssetsCurrent'],
            'Total Liabilities': ['Liabilities'],
            'Total Current Liabilities': ['LiabilitiesCurrent'],
            'Total Equity': ['StockholdersEquity'],
            'Inventory': ['InventoryNet', 'Inventory'],
            'Accounts Receivable': ['AssetsCurrent', 'ReceivablesNetCurrent'],
        }

        if foreign:
            # 20-F/40-F filers usually use IFRS concepts. Keep the original
            # native 10-K parent list unchanged, and only extend the Arelle
            # learner with IFRS-style parents when the foreign branch opts in.
            _LABEL_TO_PARENT_CONCEPT['Revenue'].extend([
                'Revenue', 'RevenueFromContractsWithCustomers',
                'RevenueFromContractsWithCustomersExcludingAssessedTax',
            ])
            _LABEL_TO_PARENT_CONCEPT['Cost of Revenue'].extend([
                'CostOfSales', 'CostOfGoodsAndServicesSold',
            ])
            _LABEL_TO_PARENT_CONCEPT['Total Operating Expenses'].extend([
                'ProfitLossFromOperatingActivities', 'OperatingCostsAndExpenses',
            ])
            _LABEL_TO_PARENT_CONCEPT['Operating Income'].extend([
                'ProfitLossFromOperatingActivities', 'OperatingProfitLoss',
            ])
            _LABEL_TO_PARENT_CONCEPT['Net Income'].extend([
                'ProfitLoss', 'ProfitLossAttributableToOwnersOfParent',
            ])
            _LABEL_TO_PARENT_CONCEPT['Interest Income'].extend([
                'FinanceIncome', 'InterestIncome',
            ])
            _LABEL_TO_PARENT_CONCEPT['Interest Expense'].extend([
                'FinanceCosts', 'InterestExpense',
            ])
            _LABEL_TO_PARENT_CONCEPT['Research & Development'].extend([
                'ProfitLossFromOperatingActivities', 'ResearchAndDevelopmentExpense',
            ])
            _LABEL_TO_PARENT_CONCEPT['Selling, General & Admin'].extend([
                'ProfitLossFromOperatingActivities', 'SellingGeneralAndAdministrativeExpense',
            ])
            _LABEL_TO_PARENT_CONCEPT['Operating Cash Flow'].append(
                'CashFlowsFromUsedInOperatingActivities')
            _LABEL_TO_PARENT_CONCEPT['Investing Cash Flow'].append(
                'CashFlowsFromUsedInInvestingActivities')
            _LABEL_TO_PARENT_CONCEPT['Financing Cash Flow'].append(
                'CashFlowsFromUsedInFinancingActivities')
            _LABEL_TO_PARENT_CONCEPT['Capital Expenditures'].append(
                'CashFlowsFromUsedInInvestingActivities')
            _LABEL_TO_PARENT_CONCEPT['Share Repurchases'].append(
                'CashFlowsFromUsedInFinancingActivities')
            _LABEL_TO_PARENT_CONCEPT['Dividends Paid'].append(
                'CashFlowsFromUsedInFinancingActivities')
            _LABEL_TO_PARENT_CONCEPT['Total Current Assets'].append('CurrentAssets')
            _LABEL_TO_PARENT_CONCEPT['Total Current Liabilities'].append('CurrentLiabilities')
            _LABEL_TO_PARENT_CONCEPT['Total Equity'].extend([
                'Equity', 'EquityAttributableToOwnersOfParent',
            ])
            _LABEL_TO_PARENT_CONCEPT['Inventory'].append('Inventories')
            _LABEL_TO_PARENT_CONCEPT['Accounts Receivable'].append(
                'TradeAndOtherCurrentReceivables')
        
        added_count = 0
        for label, parents in _LABEL_TO_PARENT_CONCEPT.items():
            for parent_concept in parents:
                discovered_tags = tree_map.get(parent_concept, [])
                for child in discovered_tags:
                    cname = child['concept']
                    weight = child['weight']
                    cname_lower = cname.lower()
                    
                    is_match = False
                    
                    # 1. Negative-weighted lines (Expenses, Outflows)
                    if label in ['Total Operating Expenses', 'Cost of Revenue', 'Interest Expense', 'Capital Expenditures', 'Share Repurchases', 'Dividends Paid']:
                        # Match if it officially subtracts (weight < 0) OR if it's forced to 1.0 by the presentation linkbase but sounds like an expense/outflow
                        if weight < 0 or (weight == 1.0 and any(kw in cname_lower for kw in ['expense', 'cost', 'payment', 'purchase', 'dividend', 'repurchase'])):
                            is_match = True
                            
                    # 2. Positive-weighted lines (Revenues, Incomes, Assets, Liabilities)
                    else:
                        if weight > 0:
                            # Prevent 'Interest Expense' (forced to weight 1.0 in visual tree) from accidentally mapping to 'Interest Income'
                            if 'expense' in cname_lower and label in ['Interest Income', 'Revenue', 'Operating Income', 'Net Income']:
                                pass # Skip
                            else:
                                is_match = True
                            
                    if is_match and cname not in CONCEPT_MAP[label]['tags']:
                        # Guard against adding text blocks or abstract headers as numeric data
                        if "abstract" not in cname_lower and "textblock" not in cname_lower:
                            CONCEPT_MAP[label]['tags'].append(cname)
                            print(f"    [Learned Tag] {cname} -> {label} (via {parent_concept})")
                            added_count += 1
                    
        # =====================================================================
        # SEC Anchoring (Definition Linkbase)
        # Automatically learn custom tags by seeing what standard US-GAAP 
        # tags the company explicitly anchored them to.
        # =====================================================================
        ANCHOR_ARCROLE = "http://fasb.org/us-gaap/arcrole/concept-wider-narrower"
        relSet_def = modelXbrl.relationshipSet(ANCHOR_ARCROLE)
        
        anchor_count = 0
        for rel in relSet_def.modelRelationships:
            narrow_obj = rel.fromModelObject
            wider_obj = rel.toModelObject
            
            if narrow_obj and wider_obj:
                narrow_tag = narrow_obj.qname.localName  # The custom extension tag
                wider_tag = wider_obj.qname.localName    # The standard US GAAP tag
                
                # Look up the wider (standard) tag in our CONCEPT_MAP
                mapped_category = None
                for category_name, category_info in CONCEPT_MAP.items():
                    if wider_tag in category_info['tags']:
                        mapped_category = category_name
                        break
                
                # If we know what the standard tag is, automatically learn the custom tag!
                if mapped_category and narrow_tag not in CONCEPT_MAP[mapped_category]['tags']:
                    # Special check: Don't map segment disclosures into core financial lines accidentally
                    if "Segment" not in narrow_tag and "Member" not in narrow_tag:
                        CONCEPT_MAP[mapped_category]['tags'].append(narrow_tag)
                        print(f"    [Anchoring] Learned '{narrow_tag}' is anchored to '{wider_tag}' -> Mapped to '{mapped_category}'")
                        anchor_count += 1
                        
        if anchor_count > 0:
            print(f"    [Anchoring Success] Auto-mapped {anchor_count} custom tags based on SEC definitions.")
            
        # Foreign 20-F/40-F extraction uses _FX_CONCEPT_ALIAS as its fast
        # concept lookup. Arelle mutates CONCEPT_MAP in-place, so refresh the
        # foreign alias table after each successful Arelle pre-pass. This does
        # not change any filing values by itself; it only makes newly learned
        # exact concept names eligible during the later XBRL fact selection.
        try:
            if "_fx_build_concept_alias" in globals():
                globals()["_FX_CONCEPT_ALIAS"] = _fx_build_concept_alias()
        except Exception:
            pass

        ctrl.modelManager.close()
        ctrl.close()
    except Exception as e:
        print(f"    [Arelle Error] Failed to learn from {filing_url}: {e}")


def arelle_global_rescue(filing_url, parent_concept):
    if not filing_url: return [], {}
    if filing_url not in _ARELLE_CACHE:
        try:
            print(f"    [Arelle Fallback] Spinning up Arelle for {filing_url}...")
            from arelle import CntlrCmdLine, XbrlConst
            ctrl = CntlrCmdLine.CntlrCmdLine()
            ctrl.startLogging(logFileName=os.devnull)
            modelXbrl = ctrl.modelManager.load(filing_url)
            relSet = modelXbrl.relationshipSet(XbrlConst.summationItem)
            
            tree_map = {}
            for rel in relSet.modelRelationships:
                parent_name = rel.fromModelObject.name if rel.fromModelObject else ""
                child_name = rel.toModelObject.name if rel.toModelObject else ""
                if parent_name and child_name:
                    if parent_name not in tree_map:
                        tree_map[parent_name] = []
                    tree_map[parent_name].append({
                        'concept': child_name,
                        'weight': rel.weight
                    })
            
            facts_map = {}
            for fact in modelXbrl.facts:
                if fact.context is not None and not fact.context.hasSegment:
                    cname = fact.qname.localName
                    if cname not in facts_map:
                        facts_map[cname] = []
                    try:
                        val = float(fact.value)
                        start = fact.context.startDatetime
                        end = fact.context.endDatetime if fact.context.endDatetime else fact.context.instantDatetime
                        dur = (end - start).days if start and end else 0
                        facts_map[cname].append({'val': val, 'dur': dur})
                    except: pass
                    
            _ARELLE_CACHE[filing_url] = {
                'tree': tree_map,
                'facts': facts_map
            }
            ctrl.modelManager.close()
            ctrl.close()
        except Exception as e:
            print(f"    [Arelle Error] {e}")
            _ARELLE_CACHE[filing_url] = None
            
    cache = _ARELLE_CACHE.get(filing_url)
    if not cache: return [], {}
    
    discovered_tags = cache['tree'].get(parent_concept, [])
    return discovered_tags, cache['facts']

def _neutralize_taxonomy_poisoning(df):
    """
    Late-filed granular sub-component facts (e.g. a 10-K footnote's
    'PaymentsToAcquireAvailableForSaleSecuritiesDebt' covering one interim
    quarter) can land on the same (Label, End, Duration) as the originally
    filed aggregate, and -- because dedup prefers the newest filing --
    silently replace a $17.8B aggregate with a $1.3B component.

    A blanket rank-over-recency rule is NOT the answer: it breaks legitimate
    cross-concept restatements (Amazon's refiled segment comparatives lost
    real values to stale zero facts under older concepts). The demotion is
    therefore value-aware and surgical: within one (Label, End, Duration)
    group, a fact is demoted only when ALL hold --
      * a lower-rank (more aggregate) concept also covers this exact period,
      * the fact's concept differs from that aggregate's concept,
      * its magnitude marks it as a strict sub-component (< 60% of the
        aggregate's value).
    Same-concept restatements and larger/equal re-tagged values are never
    touched, so recency keeps winning everywhere it should.
    """
    if df.empty or 'TagRank' not in df.columns or 'Concept' not in df.columns:
        return df
    dup = df.duplicated(['Label', 'End', 'Duration'], keep=False)
    if not dup.any():
        return df
    drop_idx = []
    for _key, g in df[dup].groupby(['Label', 'End', 'Duration'], sort=False):
        if g['Concept'].nunique() < 2 or g['TagRank'].nunique() < 2:
            continue
        min_rank = g['TagRank'].min()
        if min_rank >= 999:
            continue   # no curated aggregate present
        g_best = g[g['TagRank'] == min_rank].sort_values('_Filed_dt', ascending=False).iloc[0]
        bv = abs(g_best['Value']) if pd.notna(g_best['Value']) else 0.0
        if bv <= 0:
            continue
        losers = g[(g['TagRank'] > min_rank)
                   & (g['Concept'] != g_best['Concept'])
                   & (g['Value'].abs() < 0.6 * bv)]
        if len(losers):
            drop_idx.extend(losers.index.tolist())
    if drop_idx:
        print(f"  [Dedup] Demoted {len(drop_idx)} granular sub-component fact(s) "
              f"that would have overwritten same-period aggregates "
              f"(taxonomy-poisoning guard).")
        df = df.drop(index=drop_idx)
    return df

def _derive_crosstabulated_matrix(df: pd.DataFrame) -> pd.DataFrame:
    # =========================================================================
    # MATRIX DEDUCTION FOR CROSS-TABULATED SEGMENTS
    # Recovers missing 4d cells mathematically from known 4a/4b/4c totals.
    # =========================================================================
    df = df.copy()
    crosstab_labels = [lbl for (cat, lbl) in df.index if cat == '4d_Segments_Cross_Tabulated' and lbl.startswith('Revenue - ')]
    if crosstab_labels:
        geo_labels = {lbl: lbl.replace('Revenue - ', '') for (cat, lbl) in df.index if cat in ('4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries') and lbl.startswith('Revenue - ')}
        biz_labels = {lbl: lbl.replace('Revenue - ', '') for (cat, lbl) in df.index if cat == '4a_Segments_Business' and lbl.startswith('Revenue - ')}

        for q_col in df.columns:
            if not isinstance(q_col, str) or '-' not in q_col:
                continue

            # 1. Derive from Geography totals
            for geo_lbl, geo_name in geo_labels.items():
                child_4d = [lbl for lbl in crosstab_labels if lbl.startswith(f"Revenue - {geo_name} - ")]
                if not child_4d: continue

                geo_cat = '4b_Segments_Geographic_Regions' if ('4b_Segments_Geographic_Regions', geo_lbl) in df.index else '4c_Segments_Geographic_Countries'
                geo_total = pd.to_numeric(df.at[(geo_cat, geo_lbl), q_col], errors='coerce')
                if pd.isna(geo_total) or geo_total <= 0: continue

                missing_children = []
                known_sum = 0.0
                for child in child_4d:
                    v = pd.to_numeric(df.at[('4d_Segments_Cross_Tabulated', child), q_col], errors='coerce')
                    if pd.isna(v): missing_children.append(child)
                    else: known_sum += float(v)

                if len(missing_children) == 1:
                    missing_child = missing_children[0]
                    derived_val = geo_total - known_sum
                    if derived_val >= 0 or abs(derived_val) < geo_total * 0.2:
                        df.at[('4d_Segments_Cross_Tabulated', missing_child), q_col] = derived_val
                        print(f"  [Matrix Deduction] Derived missing cross-tabulated segment: {missing_child} {q_col} = {derived_val:,.0f} (from {geo_name} total)")

            # 2. Derive from Business/Product totals
            for biz_lbl, biz_name in biz_labels.items():
                child_4d = [lbl for lbl in crosstab_labels if lbl.endswith(f" - {biz_name}")]
                if not child_4d: continue

                biz_total = pd.to_numeric(df.at[('4a_Segments_Business', biz_lbl), q_col], errors='coerce')
                if pd.isna(biz_total) or biz_total <= 0: continue

                missing_children = []
                known_sum = 0.0
                for child in child_4d:
                    v = pd.to_numeric(df.at[('4d_Segments_Cross_Tabulated', child), q_col], errors='coerce')
                    if pd.isna(v): missing_children.append(child)
                    else: known_sum += float(v)

                if len(missing_children) == 1:
                    missing_child = missing_children[0]
                    derived_val = biz_total - known_sum
                    if derived_val >= 0 or abs(derived_val) < biz_total * 0.2:
                        df.at[('4d_Segments_Cross_Tabulated', missing_child), q_col] = derived_val
                        print(f"  [Matrix Deduction] Derived missing cross-tabulated segment: {missing_child} {q_col} = {derived_val:,.0f} (from {biz_name} total)")

    return df

def build_pivoted_data(all_facts, ticker, ye_month, company_name=None, is_financial=False, is_insurance=False, is_oil_gas=False, is_reit=False):
    with _ProfileTimer("build_pivoted_data_total"):
        return _build_pivoted_data_impl(
            all_facts, ticker, ye_month, company_name=company_name,
            is_financial=is_financial, is_insurance=is_insurance,
            is_oil_gas=is_oil_gas, is_reit=is_reit,
        )


def _build_pivoted_data_impl(all_facts, ticker, ye_month, company_name=None, is_financial=False, is_insurance=False, is_oil_gas=False, is_reit=False):
    df = pd.DataFrame(all_facts)
    if df.empty: return pd.DataFrame()
    df['Value'] = pd.to_numeric(df['Value'], errors='coerce')
    df = df.dropna(subset=['Value'])
    def normalize_segment_label(label):
        parts = label.split(' - ')
        if len(parts) > 1:
            prefix, mems = parts[0], [p.strip() for p in parts[1:]]
            unique_mems, seen = [], set()
            for m in mems:
                if m.lower() not in seen: unique_mems.append(m); seen.add(m.lower())
            final_mems = []
            for m in unique_mems:
                is_comp = False
                for o in unique_mems:
                    if m.lower() == o.lower(): continue
                    k, mc = o.lower(), m.lower()
                    if mc.startswith(k+' and ') or mc.endswith(' and '+k) or ' and '+k+' ' in mc or ' '+k+' and ' in mc:
                        is_comp = True; break
                if not is_comp: final_mems.append(m)
            if not final_mems: final_mems = unique_mems
            return prefix + ' - ' + ' - '.join(sorted(final_mems, key=str.lower))
        return label
    _label_norm_map = {
        _lbl: normalize_segment_label(_lbl)
        for _lbl in df['Label'].dropna().unique()
    }
    df['Label'] = df['Label'].map(lambda _lbl: _label_norm_map.get(_lbl, _lbl))
    df = _reconcile_segment_labels(df, ticker=ticker, company_name=company_name)
    df = _detect_and_merge_renamed_segments(df)
    df = _merge_concurrent_member_variants(df)
    df = _consolidate_geographic_alias_facts(df)

    df['_End_dt'] = pd.to_datetime(df['End'], errors='coerce')
    df['_Filed_dt'] = pd.to_datetime(df['Filed'], errors='coerce')
    df['_FilingDelay'] = (df['_Filed_dt'] - df['_End_dt']).dt.days.fillna(9999).astype(int)
    if 'IsCalculated' in df.columns:
        df['_IsCalcSort'] = df['IsCalculated'].astype(bool).astype('int8')
    else:
        df['_IsCalcSort'] = 0
    df['_DimSort'] = df['DimCount']
    is_seg = df['Category'].isin(SEG_CATS)
    df.loc[is_seg & df['DimCount'].eq(1), '_DimSort'] = 0
    df.loc[is_seg & df['DimCount'].eq(0), '_DimSort'] = 100
    
    df = _neutralize_taxonomy_poisoning(df)
    df = df.sort_values(['Label', 'End', 'Duration', '_Filed_dt', '_IsCalcSort', 'TagRank', '_DimSort'], ascending=[True, True, True, False, True, True, True])

    if 'Depreciation & Amortization' in df['Label'].values:
        total_tags = {
            'DepreciationDepletionAndAmortization',
            'DepreciationAndAmortization',
            'DepreciationAmortizationAndOther',
        }
        # Component concepts are grouped by economic family.  A broad family
        # total wins over its details, while independent families are added.
        # This derives D&A when no combined tag is filed without double-counting
        # (for example, a total intangible-amortization fact plus software detail).
        depreciation_total_tags = {'Depreciation'}
        depreciation_detail_families = (
            {'DepreciationAndAmortizationOfPropertyPlantAndEquipment'},
            {'DepreciationAndAmortizationOfFinanceLeaseRightOfUseAssets'},
        )
        depletion_tags = {'Depletion'}
        amortization_total_tags = {'AmortizationOfIntangibleAssets'}
        amortization_detail_families = (
            {'AmortizationOfComputerSoftwareCosts'},
            {'AmortizationOfDeferredCharges'},
            {'AmortizationOfAcquisitionCosts'},
            {'AcquiredInPlaceLeasesAmortizationExpense'},
        )
        other_da_tags = {'OtherDepreciationAndAmortization'}
        da_mask = df['Label'] == 'Depreciation & Amortization'
        da_df, new_da_rows = df[da_mask], []
        for (end, dur), group in da_df.groupby(['End', 'Duration']):
            for filed, f_group in group.groupby('Filed'):
                tags_list = CONCEPT_MAP['Depreciation & Amortization']['tags']
                f_group = f_group.copy()
                f_group['ConceptName'] = f_group['TagRank'].apply(lambda x: tags_list[int(x)] if x < len(tags_list) else 'Unknown')
                present_total = f_group[f_group['ConceptName'].isin(total_tags)].copy()
                if not present_total.empty:
                    # Prefer a clean D&A/DD&A total over the broader
                    # DepreciationAmortizationAndOther concept even when the
                    # latter happens to have an earlier configured tag rank.
                    total_priority = {
                        'DepreciationDepletionAndAmortization': 0,
                        'DepreciationAndAmortization': 1,
                        'DepreciationAmortizationAndOther': 2,
                    }
                    present_total['_DATotalPriority'] = present_total['ConceptName'].map(total_priority)
                    selected_total = present_total.sort_values(
                        ['_Filed_dt', '_DATotalPriority', 'TagRank'],
                        ascending=[False, True, True],
                    ).iloc[0]
                    new_da_rows.append(selected_total.drop(labels=['_DATotalPriority'], errors='ignore'))
                    continue

                def _best_component(tag_set):
                    candidates = f_group[f_group['ConceptName'].isin(tag_set)].copy()
                    if candidates.empty:
                        return None
                    candidates['_NumericValue'] = pd.to_numeric(candidates['Value'], errors='coerce')
                    candidates = candidates[candidates['_NumericValue'].notna()]
                    if candidates.empty:
                        return None
                    return candidates.sort_values(
                        ['_Filed_dt', 'TagRank'], ascending=[False, True]
                    ).iloc[0]

                selected_components = []
                dep_total = _best_component(depreciation_total_tags)
                if dep_total is not None:
                    selected_components.append(dep_total)
                else:
                    for family in depreciation_detail_families:
                        component = _best_component(family)
                        if component is not None:
                            selected_components.append(component)

                depletion = _best_component(depletion_tags)
                if depletion is not None:
                    selected_components.append(depletion)

                amort_total = _best_component(amortization_total_tags)
                if amort_total is not None:
                    selected_components.append(amort_total)
                else:
                    for family in amortization_detail_families:
                        component = _best_component(family)
                        if component is not None:
                            selected_components.append(component)

                other_da = _best_component(other_da_tags)
                if other_da is not None:
                    selected_components.append(other_da)

                if selected_components:
                    best_row = selected_components[0].copy()
                    best_row['Value'] = sum(float(r['_NumericValue']) for r in selected_components)
                    best_row['TagRank'] = 0
                    best_row['Concept'] = 'DerivedDepreciationAndAmortizationComponents'
                    new_da_rows.append(best_row.drop(labels=['_NumericValue'], errors='ignore'))
                else:
                    # Preserve an unusual mapped D&A concept rather than losing
                    # data merely because its taxonomy name is not yet classified.
                    new_da_rows.append(f_group.sort_values(['_Filed_dt', 'TagRank'], ascending=[False, True]).iloc[0])
        df = pd.concat([df[~da_mask], pd.DataFrame(new_da_rows)])
        df = df.sort_values(['Label', 'End', 'Duration', '_Filed_dt', '_IsCalcSort', 'TagRank', '_DimSort'], ascending=[True, True, True, False, True, True, True])

    df = df.drop(columns=[c for c in ['_End_dt', '_FilingDelay', '_DimSort', '_IsCalcSort'] if c in df.columns])
    df = _presplit_segment_era_facts(df)

    # =========================================================================
    # CHANGE 1: Cash-flow annual intelligence layer
    #
    # Two sub-fixes that run BEFORE the generic drop_duplicates:
    #
    # (A) Save ALL annual CF fact values per (Label, FY) so the Q4 derivation
    #     can try alternatives when the primary annual produces an implausible Q4.
    #
    # (B) Circularity-aware dedup: when two annual CF facts exist for the same
    #     (Label, End, Duration) and one matches the quarterly sum (circular
    #     supplemental rollup), drop it so the audited consolidated total wins.
    #     Example: AMD FY2025 OCF has 7709M (audited) and 6493M (Q1+Q2+Q3+Q4_supp).
    #     Without this, drop_duplicates keeps one arbitrarily.
    # =========================================================================

    # (A) Snapshot ALL annual CF values before dedup destroys alternatives
    _cf_ann_mask = (
        (df['Category'] == '3_Cash_Flow')
        & (df['Duration'] >= 350)
        & (df['Duration'] <= 385)
        & (df['Q'] == 'Q4')
    )
    _cf_annual_all = {
        key: sorted(set(float(v) for v in values))
        for key, values in df.loc[_cf_ann_mask].groupby(['Label', 'FY'], sort=False)['Value']
    }

    # (A-IS) Snapshot ALL annual income-statement REVENUE values before dedup.
    # A later restatement that moves a business to discontinued operations
    # restates only the ANNUAL income statement (continuing ops, ex-spinoff)
    # while the original interim quarters keep their as-reported (incl-spinoff)
    # basis.  Latest-filing-wins dedup then keeps the restated annual, so the
    # Q4 = Annual - 9M derivation subtracts incompatible bases and collapses
    # (IBM FY2019/FY2020 after the Kyndryl spin: derived Q4 ~$2B vs ~$21B real).
    # Keeping every annual value lets the derivation fall back to the original-
    # basis annual that is consistent with the quarters.
    _label_s = df['Label'].astype(str)
    _is_revenue_label = _label_s.eq('Revenue') | _label_s.str.startswith('Revenue - ')
    _is_ann_mask = (
        (df['Category'] == '1_Income_Statement')
        & _is_revenue_label
        & (df['Duration'] >= 350)
        & (df['Duration'] <= 385)
        & (df['Q'] == 'Q4')
    )
    _is_annual_all = {
        key: sorted(set(float(v) for v in values))
        for key, values in df.loc[_is_ann_mask].groupby(['Label', 'FY'], sort=False)['Value']
    }
    # (B) Circularity-aware dedup
    # Step 1: Compute per-quarter-deduplicated quarterly sums.
    #   CRITICAL: Before drop_duplicates, the same quarter may have facts from
    #   BOTH the 10-Q filing AND the 10-K supplemental note. Summing naively
    #   double-counts them. We first deduplicate per (Label, FY, Q) to get one
    #   value per quarter, then sum.
    _cf_qtly_mask = (
        (df['Category'] == '3_Cash_Flow')
        & (df['Duration'] >= 60)
        & (df['Duration'] <= 120)
    )
    _qtly_for_sum = df[_cf_qtly_mask].copy()
    if not _qtly_for_sum.empty:
        _qtly_for_sum = _qtly_for_sum.sort_values(
            ['Label', 'FY', 'Q', 'TagRank', '_Filed_dt'],
            ascending=[True, True, True, True, False]
        ).drop_duplicates(subset=['Label', 'FY', 'Q'], keep='first')
        _qtly_sums = _qtly_for_sum.groupby(['Label', 'FY'])['Value'].sum().to_dict()
    else:
        _qtly_sums = {}

    # Step 2: For annual CF facts with identical (Label, End, Duration),
    #   if one value matches the quarterly sum (circular) and another doesn't
    #   (audited), drop the circular one.
    _rows_to_drop = set()
    _cf_annual_df = df[_cf_ann_mask]
    if not _cf_annual_df.empty:
        for (lbl, end, dur), grp in _cf_annual_df.groupby(['Label', 'End', 'Duration']):
            if len(grp) < 2:
                continue
            fy_val = grp['FY'].iloc[0]
            q_sum = _qtly_sums.get((lbl, fy_val))
            if q_sum is None or q_sum == 0:
                continue

            circular_idxs = []
            non_circular_idxs = []
            # Equivalent to ``grp.iterrows()`` -- only ``idx`` and the 'Value'
            # cell are used -- without building a Series per row.
            for idx, _row_value in zip(grp.index, grp['Value']):
                try:
                    val = abs(float(_row_value))
                    rel_diff = abs(val - abs(q_sum)) / max(abs(q_sum), 1)
                    if rel_diff < 0.03:
                        circular_idxs.append(idx)
                    else:
                        non_circular_idxs.append(idx)
                except (TypeError, ValueError):
                    non_circular_idxs.append(idx)

            if non_circular_idxs and circular_idxs:
                _rows_to_drop.update(circular_idxs)
                try:
                    _circ_v = df.loc[circular_idxs[0], 'Value']
                    _aud_v = df.loc[non_circular_idxs[0], 'Value']
                    print(f"  [Circ-Dedup] {lbl[:45]} FY{fy_val}: "
                          f"dropping circular {float(_circ_v):,.0f} "
                          f"(â‰ˆ quarterly sum {q_sum:,.0f}), "
                          f"keeping audited {float(_aud_v):,.0f}")
                except Exception:
                    pass

    if _rows_to_drop:
        df = df.drop(index=_rows_to_drop)

    # CRITICAL: Keep 'first' after sorting by _Filed_dt DESCENDING ensures Latest Filing Wins (Restatement Priority)
    df = df.drop_duplicates(subset=['Label', 'End', 'Duration'], keep='first')

    # --- Pre-pivot Label Unification ---
    # Merge renamed segments/line-items BEFORE the derivation loop. This ensures that
    # the 365d annual fact and the 90d quarterly facts all land in the SAME group
    # for the Q4 = Annual - Q1 - Q2 - Q3 derivation logic to work.
    _MERGE_CATS = {'1_Income_Statement', '4a_Segments_Business', '4b_Segments_Geographic_Regions',
                   '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'}
    _rename_map = {}
    for cat in df['Category'].unique():
        if cat not in _MERGE_CATS: continue
        cat_df = df[df['Category'] == cat]
        lbls = cat_df['Label'].unique()
        rev_lbls = [l for l in lbls if l.startswith('Revenue - ') and l.count(' - ') == 1]
        if len(rev_lbls) < 2:
            continue

        # Pre-compute suffixes, token sets, row counts, and occupied periods
        # once per category.  The old pair loop rebuilt the same masks/groupbys
        # for every candidate pair; this preserves the exact overlap test while
        # avoiding repeated full cat_df scans.
        _rev_suffix = {l: l.split(' - ', 1)[1] for l in rev_lbls}
        _rev_lower = {l: l.lower() for l in rev_lbls}
        _rev_tokens = {
            l: set(re.findall(r'[a-z0-9]+', _rev_suffix[l].lower()))
            for l in rev_lbls
        }
        _label_counts = cat_df['Label'].value_counts().to_dict()
        _period_sets = {}
        _period_src = cat_df[cat_df['Label'].isin(rev_lbls) & cat_df['Value'].notna()]
        if not _period_src.empty:
            for _lbl, _grp in _period_src.groupby('Label', sort=False):
                _period_sets[_lbl] = set(zip(_grp['FY'], _grp['Q']))

        for a in rev_lbls:
            # These depend only on `a`; hoisted out of the inner `for b` loop.
            _a_has_other = 'other' in _rev_lower[a]
            a_clean = _rev_suffix[a]
            a_toks = _rev_tokens[a]
            for b in rev_lbls:
                if a == b: continue
                # Skip 'other' matching to prevent false positives
                if _a_has_other or 'other' in _rev_lower[b]: continue
                b_clean = _rev_suffix[b]
                if len(b_clean) < 4: continue
                
                b_toks = _rev_tokens[b]
                jaccard = len(a_toks & b_toks) / max(len(a_toks | b_toks), 1)
                
                if (a_clean.startswith(b_clean) or b_clean in a_clean) and jaccard > 0.6:
                    # Same behavior as the previous unstack/dropna check: any
                    # non-null shared (FY, Q) means the two live concurrently,
                    # so leave them to the post-pivot merge logic.
                    if _period_sets.get(a, set()) & _period_sets.get(b, set()):
                        continue

                    na = _label_counts.get(a, 0)
                    nb = _label_counts.get(b, 0)
                    survivor = a_clean if na >= nb else b_clean
                    donor = b_clean if na >= nb else a_clean
                    if donor not in _rename_map:
                        _rename_map[donor] = survivor

    def _resolve(x):
        seen = set()
        while x in _rename_map and x not in seen:
            seen.add(x)
            x = _rename_map[x]
        return x

    _rename_map = {k: _resolve(v) for k, v in _rename_map.items() if _resolve(v) != k}
    
    def _apply_rename(lbl):
        if not isinstance(lbl, str) or ' - ' not in lbl:
            return lbl
        prefix, suffix = lbl.split(' - ', 1)
        if suffix in _rename_map:
            return f"{prefix} - {_rename_map[suffix]}"
        return lbl

    for _donor, _surv in _rename_map.items():
        print(f"  [Pre-Pivot Merge] Unifying suffix '{_donor}' -> '{_surv}'")
        
    df['Label'] = df['Label'].apply(_apply_rename)

    final_rows = []

    # These three helpers and their per-group caches were previously
    # redefined on every loop iteration.  Defined once here, they capture
    # the `group` loop variable (read at call time) and the cache dicts,
    # which are cleared at the top of each iteration below, so behaviour is
    # identical to the previous per-iteration definition.
    _q_group_cache = {}
    _duration_filter_cache = {}
    _concept_resolve_cache = {}
    _best_val_cache = {}
    group = None

    def _resolve_concept_cached(_concept):
        if _concept in _concept_resolve_cache:
            return _concept_resolve_cache[_concept]
        _resolved = resolve_concept_to_label(_concept)
        _concept_resolve_cache[_concept] = _resolved
        return _resolved

    def _duration_candidates(q_label, target_dur, max_distance):
        _key = (q_label, target_dur, max_distance)
        cached = _duration_filter_cache.get(_key)
        if cached is not None:
            return cached
        q_group = _q_group_cache.get(q_label)
        if q_group is None:
            q_group = group[(group['Q'] == q_label)]
            _q_group_cache[q_label] = q_group
        if q_group.empty:
            _duration_filter_cache[_key] = q_group
            return q_group
        dist = (q_group['Duration'] - target_dur).abs()
        m = q_group.loc[dist <= max_distance].copy()
        if not m.empty:
            m['Dist'] = dist.loc[m.index].to_numpy()
        _duration_filter_cache[_key] = m
        return m

    def get_best_val(q_label, target_dur, max_distance=40, exclude_val=None, concept_filter=None):
        """
            Find the best-matching fact for (q_label, ~target_dur).

            The selection rules are unchanged.  The per-group caches only avoid
            re-filtering the same quarter/duration candidates and re-resolving
            concept aliases during repeated YTD/Q4 derivation probes.
            """
        _exclude_key = None if exclude_val is None else float(exclude_val)
        _cache_key = (q_label, target_dur, max_distance, _exclude_key, concept_filter)
        cached_best = _best_val_cache.get(_cache_key)
        if cached_best is not None:
            if cached_best is False:
                return None
            return cached_best[0], cached_best[1].copy()

        m = _duration_candidates(q_label, target_dur, max_distance)
        if m.empty:
            _best_val_cache[_cache_key] = False
            return None

        if concept_filter is not None:
            filter_label = _resolve_concept_cached(concept_filter)
            if filter_label is not None:
                # Semantic match: accept any concept resolving to same label
                m = m[m['Concept'].map(_resolve_concept_cached) == filter_label]
            else:
                # concept_filter not in CONCEPT_MAP at all â€” fall back to exact
                m = m[m['Concept'] == concept_filter]
            if m.empty:
                _best_val_cache[_cache_key] = False
                return None

        if exclude_val is not None:
            mask_excl = (m['Value'] - exclude_val).abs() > abs(exclude_val) * 0.001 + 1
            m = m[mask_excl]
            if m.empty:
                _best_val_cache[_cache_key] = False
                return None

        m = m.sort_values(['Dist', '_Filed_dt', 'IsCalculated', 'TagRank'], ascending=[True, False, True, True])
        result = (float(m.iloc[0]['Value']), m.iloc[0].copy())
        _best_val_cache[_cache_key] = result
        return result[0], result[1].copy()
    for (label, fy), group in df.groupby(['Label', 'FY']):
        _q_group_cache.clear()
        _duration_filter_cache.clear()
        _concept_resolve_cache.clear()
        _best_val_cache.clear()
        cat = group['Category'].iloc[0]
        is_bs, is_avg = cat == '2_Balance_Sheet', any(kw in label for kw in NO_SUBTRACT)
        is_segment_cat = cat in SEG_CATS
        final_q_vals, best_rows = {'Q1': None, 'Q2': None, 'Q3': None, 'Q4': None}, {}
        

        if is_bs or 'Useful Life' in label or 'RPO' in label:
            q_src = {qn: get_best_val(qn, 0, 15) for qn in ['Q1', 'Q2', 'Q3', 'Q4']}
        elif is_avg:
            # For Shares/EPS, allow annual (365d) duration if Q4 specific (90d) is missing
            q_src = {qn: get_best_val(qn, 91, 40) for qn in ['Q1', 'Q2', 'Q3']}
            q4_val = get_best_val('Q4', 91, 40)
            # North Star: Never use Annual (365d) total for a discrete Q4 EPS.
            if q4_val is None and 'EPS' not in label:
                q4_val = get_best_val('Q4', 365, 50)  # 50d tolerance for 52/53-week filers
            q_src['Q4'] = q4_val
        else:
            q_src = {qn: get_best_val(qn, 91, 40) for qn in ['Q1', 'Q2', 'Q3', 'Q4']}
        
        q_vals = {qn: (v[0] if v else None) for qn, v in q_src.items()}
        best_rows = {qn: (v[1] if v else None) for qn, v in q_src.items()}
        if label == 'Purchases of Investments' and fy == 2025:
            _debug_print(f"DEBUG q_vals for {label} {fy}: {q_vals}")

        # Track Q4 provenance for the scope-mismatch guard below.
        # _q4_directly_filed: True when the 10-Q/10-K filed an explicit
        #   ~91-day quarterly fact for Q4 (not via derivation).
        # _q4_derived_flag:   Set True by Case A/B/C/D when q_vals['Q4']
        #   is replaced with the derived value (Annual âˆ’ Q1 âˆ’ Q2 âˆ’ Q3).
        _q4_directly_filed = (q_src.get('Q4') is not None)
        _q4_derived_flag   = False

        ytd6 = get_best_val('Q2', 182, 40)
        ytd9 = get_best_val('Q3', 273, 40)
        ytd12 = get_best_val('Q4', 365, 50)  # 50d tolerance for 52/53-week filers

        v6_ytd  = ytd6[0]  if ytd6  else None
        v9_ytd  = ytd9[0]  if ytd9  else None
        v12_ytd = ytd12[0] if ytd12 else None

        if not is_bs and 'Useful Life' not in label and not is_avg:
            # Derive Q1 from YTD6 if missing
            if q_vals['Q1'] is None and ytd6 is not None:
                q2_qt_same_tag = get_best_val('Q2', 91, 40, concept_filter=ytd6[1].get('Concept'))
                if q2_qt_same_tag is not None:
                    q_vals['Q1'] = ytd6[0] - q2_qt_same_tag[0]
                    if best_rows['Q1'] is None: best_rows['Q1'] = best_rows['Q2'].copy() if best_rows['Q2'] is not None else ytd6[1].copy()
                elif q_vals['Q2'] is not None:
                    q_vals['Q1'] = ytd6[0] - q_vals['Q2']
                    if best_rows['Q1'] is None: best_rows['Q1'] = best_rows['Q2'].copy() if best_rows['Q2'] is not None else ytd6[1].copy()

            ALWAYS_POS = {'Capital Expenditures', 'Depreciation', 'Amortization', 'Stock-Based Compensation', 'Share Repurchases', 'Dividends Paid', 'Taxes Paid on Stock Awards', 'Total Debt Repaid', 'Short-term Debt Repaid', 'Long-term Debt Repaid', 'Purchases of Investments', 'Acquisitions'}

            # Derive Q2 from YTD6 if missing
            if q_vals['Q2'] is None and ytd6 is not None:
                q1_qt_same_tag = get_best_val('Q1', 91, 40, concept_filter=ytd6[1].get('Concept'))
                if q1_qt_same_tag is not None:
                    q_vals['Q2'] = ytd6[0] - q1_qt_same_tag[0]
                    if best_rows['Q2'] is None: best_rows['Q2'] = ytd6[1].copy()
                elif q_vals['Q1'] is not None:
                    q_vals['Q2'] = ytd6[0] - q_vals['Q1']
                    if best_rows['Q2'] is None: best_rows['Q2'] = ytd6[1].copy()
                elif cat == '3_Cash_Flow' and not is_segment_cat:
                    derived_q2 = ytd6[0]  # Assuming Q1 is 0
                    if derived_q2 >= 0 or label not in ALWAYS_POS:
                        q_vals['Q2'] = derived_q2
                        if best_rows['Q2'] is None: best_rows['Q2'] = ytd6[1].copy()

            # Derive Q3 from YTD9 if missing
            if q_vals['Q3'] is None and ytd9 is not None:
                ytd6_same_tag = get_best_val('Q2', 182, 40, concept_filter=ytd9[1].get('Concept'))
                if ytd6_same_tag is not None:
                    q_vals['Q3'] = ytd9[0] - ytd6_same_tag[0]
                    if best_rows['Q3'] is None: best_rows['Q3'] = ytd9[1].copy()
                elif v6_ytd is not None:
                    q_vals['Q3'] = ytd9[0] - v6_ytd
                    if best_rows['Q3'] is None: best_rows['Q3'] = ytd9[1].copy()
                elif cat == '3_Cash_Flow' and not is_segment_cat:
                    _q1 = q_vals['Q1'] if q_vals['Q1'] is not None else 0.0
                    _q2 = q_vals['Q2'] if q_vals['Q2'] is not None else 0.0
                    derived_q3 = ytd9[0] - (_q1 + _q2)
                    if derived_q3 >= 0 or label not in ALWAYS_POS:
                        q_vals['Q3'] = derived_q3
                        if best_rows['Q3'] is None: best_rows['Q3'] = ytd9[1].copy()

            _v9_synthetic = None
            if (
                v9_ytd is None
                and q_vals['Q1'] is not None
                and q_vals['Q2'] is not None
                and q_vals['Q3'] is not None
            ):
                _v9_synthetic = q_vals['Q1'] + q_vals['Q2'] + q_vals['Q3']
            elif cat == '3_Cash_Flow' and not is_segment_cat and v9_ytd is None:
                _q1 = q_vals['Q1'] if q_vals['Q1'] is not None else 0.0
                _q2 = q_vals['Q2'] if q_vals['Q2'] is not None else 0.0
                _q3 = q_vals['Q3'] if q_vals['Q3'] is not None else 0.0
                _v9_synthetic = _q1 + _q2 + _q3

            # --- SOLIDIFIED CASH FLOW Q2/Q3 FALLBACK SUBTRACTION ---
            if cat == '3_Cash_Flow' and not is_segment_cat:
                # Fallback for Q2: If Q2 is missing but YTD6 and Q1 are known
                if q_vals['Q2'] is None and v6_ytd is not None and q_vals['Q1'] is not None:
                    derived_q2 = v6_ytd - q_vals['Q1']
                    if derived_q2 >= 0 or label not in ALWAYS_POS:
                        q_vals['Q2'] = derived_q2
                        if best_rows['Q2'] is None:
                            best_rows['Q2'] = (ytd6[1] if ytd6 else (best_rows['Q1'].copy() if best_rows['Q1'] is not None else None))
                
                # Fallback for Q3: If Q3 is missing but YTD9 and YTD6 are known
                if q_vals['Q3'] is None and v9_ytd is not None and v6_ytd is not None:
                    derived_q3 = v9_ytd - v6_ytd
                    if derived_q3 >= 0 or label not in ALWAYS_POS:
                        q_vals['Q3'] = derived_q3
                        if best_rows['Q3'] is None:
                            best_rows['Q3'] = (ytd9[1] if ytd9 else (best_rows['Q2'].copy() if best_rows['Q2'] is not None else None))

    
        # --- DISCRETE Q4 CALCULATION ---
            # Priority 1: Q4 is missing -> derive from Annual âˆ’ YTD9.
            # Priority 2: Q4 equals the full annual total (filed as annual-duration
            #             XBRL by the 10-K, then picked up with too-loose tolerance)
            #             -> replace with the correct discrete Q4.
            # Safety: only subtract when we have a reliable 9-month baseline
            #         (either a real YTD9 fact, or at least 2 individual quarters).
            _has_baseline = (v9_ytd is not None) or (_v9_synthetic is not None) or (
                sum(v is not None for v in [q_vals['Q1'], q_vals['Q2'], q_vals['Q3']]) >= 2
            )

            if v12_ytd is not None and _has_baseline:
                _live_sum = (q_vals['Q1'] or 0) + (q_vals['Q2'] or 0) + (q_vals['Q3'] or 0)
                if v9_ytd is not None:
                    prev_9m = max(v9_ytd, _live_sum)
                elif _v9_synthetic is not None:
                    prev_9m = max(_v9_synthetic, _live_sum)
                else:
                    prev_9m = _live_sum
                    
                _CUMULATIVE_METRICS = (
                    'Revenue', 'Cost of Revenue', 'Gross Profit',
                    'Operating Expenses', 'Research & Development', 'Sales & Marketing',
                    'General & Administrative', 'Selling, General & Admin', 'Total Operating Expenses',
                    'Marketing Expense', 'Depreciation & Amortization',
                    'Amortization of Intangibles', 'Stock-Based Compensation',
                    'Capital Expenditures', 'Share Repurchases', 'Dividends Paid'
                )
                is_cumulative = label in _CUMULATIVE_METRICS or any(label.startswith(m + ' - ') for m in _CUMULATIVE_METRICS)

                # Concept mismatch guard: if the annual fact is wildly smaller
                # than the 9-month YTD fact, it is almost certainly a sub-component
                # tagged with a higher TagRank rather than the true annual sum.
                # We discard v12_ytd to prevent catastrophic negative Q4 derivation.
                if is_cumulative and prev_9m > 0 and v12_ytd is not None and v12_ytd < prev_9m * 0.5:
                    print(f"  [Concept Guard] Dropping invalid Annual fact for '{label}': "
                          f"Annual ({v12_ytd:,.0f}) < 50% of YTD9 ({prev_9m:,.0f})")
                    if ytd12:
                        _debug_print("DEBUG ytd12 df row:", ytd12[1].to_dict() if hasattr(ytd12[1], "to_dict") else ytd12[1])
                    v12_ytd = None

                if v12_ytd is not None:
                    discrete_q4 = v12_ytd - prev_9m

                    # Case A: Q4 is simply missing -- use the derived value.
                    if q_vals['Q4'] is None:
                        q_vals['Q4'] = discrete_q4
                        _q4_derived_flag = True
                        if best_rows['Q4'] is None:
                            best_rows['Q4'] = (
                                ytd12[1] if ytd12
                                else (best_rows['Q3'] or best_rows['Q2'] or best_rows['Q1'])
                            ).copy()

                    # Case B: Q4 was picked up as the full annual total (common when the
                    # only source is the 10-K XBRL where the annual tag sits on the Q4
                    # period boundary with ~365-day duration but duration-filter let it
                    # through).  Detect this by checking whether Q4 â‰ˆ v12_ytd AND
                    # discrete_q4 is materially smaller.
                    elif (
                        abs(q_vals['Q4'] - v12_ytd) / max(abs(v12_ytd), 1) < 0.005
                        and abs(discrete_q4) < abs(q_vals['Q4']) * 0.85
                    ):
                        q_vals['Q4'] = discrete_q4
                        _q4_derived_flag = True

                    # Case C: Q4 was derived earlier but equals the annual (e.g. when
                    # the YTD derivation produced the cumulative value accidentally).
                    # Same fix as Case B.
                    elif (
                        is_segment_cat
                        and q_vals['Q4'] is not None
                        and abs(q_vals['Q4'] - v12_ytd) / max(abs(v12_ytd), 1) < 0.02
                        and prev_9m > 0
                        and discrete_q4 > 0
                    ):
                        q_vals['Q4'] = discrete_q4
                        _q4_derived_flag = True

                    # Case D: Q4 direct fact doesn't reconcile with Annual - YTD9.
                    #
                    # Root cause this solves:
                    #   Some companies file TWO annual-duration XBRL facts for the same
                    #   label in their 10-K:
                    #     (a) Audited consolidated statement:  e.g. OCF = 7,709M
                    #     (b) Supplemental quarterly note YTD: e.g. OCF = 6,493M
                    #                                         (= YTD9 + Q4_supp, circular)
                    #
                    #   If (b) has a closer Duration to 365d than (a), get_best_val picks
                    #   (b) as v12_ytd.  Then:
                    #     implied = YTD9 + Q4_direct = 5109+1384 = 6493 = v12_ytd
                    #     recon_err = 0%  ->  Case D never fires
                    #
                    # Two-pass fix:
                    #   Pass 1: check if primary v12_ytd gives recon_err > 2%.
                    #   Pass 2: if recon_err <= 2% (circular condition detected),
                    #           ask get_best_val for an ALTERNATIVE annual that excludes
                    #           the circular value.  If it exists and gives recon_err > 2%,
                    #           it is the audited consolidated annual -- use it.
                    elif (
                        not is_segment_cat
                        and q_vals['Q4'] is not None
                        and (v9_ytd is not None or _v9_synthetic is not None)
                    ):
                        # Use the best available 9-month baseline for the reconciliation check.
                        # v9_ytd = real cumulative YTD9 from a 10-Q filing (preferred).
                        # _v9_synthetic = Q1+Q2+Q3 sum (reliable when all three are filed).
                        _baseline_9m = v9_ytd if v9_ytd is not None else _v9_synthetic
                        implied_annual = _baseline_9m + q_vals['Q4']
                        recon_err = abs(implied_annual - v12_ytd) / max(abs(v12_ytd), 1)

                        # Pass 2: the primary annual looks circular (recon_err near 0%
                        # means v12_ytd â‰ˆ baseline_9m + Q4_direct -- a tautology).
                        # Try an alternative annual that excludes the circular value.
                        if recon_err <= 0.02:
                            alt_ytd12 = get_best_val('Q4', 365, 50, exclude_val=v12_ytd)
                            if alt_ytd12 is not None:
                                alt_v12   = alt_ytd12[0]
                                alt_recon = abs(implied_annual - alt_v12) / max(abs(alt_v12), 1)
                                if alt_recon > 0.02:
                                    # Found the audited annual -- upgrade and recompute
                                    v12_ytd     = alt_v12
                                    discrete_q4 = alt_v12 - prev_9m
                                    recon_err   = alt_recon
                            else:
                                # No alternative found via exclude_val (only the circular
                                # supplemental existed as an XBRL 365d fact). Last resort:
                                # look directly for an HTML-injected annual (TagRank â‰¥ 990)
                                # in the group. HTML recovery injects TagRank=998 facts
                                # after detecting the circular annual via the circularity
                                # check in _recover_annual_cashflow_from_html.
                                html_rows = group[
                                    (group['Q'] == 'Q4') &
                                    (group['TagRank'] >= 990) &
                                    ((group['Duration'] - 365).abs() <= 50)
                                ]
                                if not html_rows.empty:
                                    html_val   = float(html_rows.iloc[0]['Value'])
                                    html_recon = abs(implied_annual - html_val) / max(abs(html_val), 1)
                                    if html_recon > 0.02:
                                        v12_ytd     = html_val
                                        discrete_q4 = html_val - prev_9m
                                        recon_err   = html_recon
                                        print(f"  [Q4 Recon Fix] {label}: "
                                            f"HTML-injected annual {html_val:,.0f} "
                                            f"resolves circular supplemental "
                                            f"({html_recon:.1%} gap) -> Q4={discrete_q4:,.0f}")

                        if recon_err > 0.02:
                            print(f"  [Q4 Recon Fix] {label}: "
                                f"baseline_9m({_baseline_9m:,.0f}) + Q4_direct({q_vals['Q4']:,.0f}) "
                                f"= {implied_annual:,.0f} vs Annual {v12_ytd:,.0f} "
                                f"({recon_err:.1%} gap) -- using derived Q4 = {discrete_q4:,.0f}")
                            q_vals['Q4'] = discrete_q4
                            _q4_derived_flag = True
        elif is_avg:
            if q_vals['Q2'] is None: q_vals['Q2'] = v6_ytd
            if q_vals['Q3'] is None: q_vals['Q3'] = v9_ytd
            if q_vals['Q4'] is None and 'EPS' not in label: q_vals['Q4'] = v12_ytd

        # =====================================================================
        # RESTATEMENT GUARD FOR INCOME-STATEMENT REVENUE Q4
        #
        # Mirror of the cash-flow scope fix, opposite direction.  When a filer
        # moves a business to discontinued operations and restates only the
        # ANNUAL prior-year income statement, the most-recent annual fact
        # (continuing ops, ex-spinoff) is subtracted from original-basis Q1-Q3
        # (incl-spinoff), so the derived Q4 collapses far below the quarterly
        # run-rate (or goes negative).  The true Q4 is recoverable: the ORIGINAL
        # annual is still on file -- it only lost latest-filing-wins dedup -- and
        # is preserved in _is_annual_all.  Pick the alternative annual whose
        # implied Q4 is plausible (positive and within [0.5x min(Q1-Q3),
        # 2.5x max]), preferring the largest such annual (the original, incl-
        # spinoff basis).  Fires ONLY on a derived Q4 that is anomalously small,
        # so a correct derivation is never disturbed.  (IBM Revenue FY2019 ->
        # $21.78B, FY2020 -> $20.37B, matching IBM's reported Q4 revenue.)
        # =====================================================================
        if (
            cat == '1_Income_Statement'
            and (label == 'Revenue' or str(label).startswith('Revenue - '))
            and not is_avg
            and not is_segment_cat
            and _q4_derived_flag
            and q_vals['Q4'] is not None
        ):
            _q123_is = [v for v in (q_vals['Q1'], q_vals['Q2'], q_vals['Q3'])
                        if v is not None]
            if len(_q123_is) >= 2:
                _min_is, _max_is = min(_q123_is), max(_q123_is)
                if _min_is > 0 and q_vals['Q4'] < 0.5 * _min_is:
                    _prev_9m_is = sum(v or 0 for v in
                                      (q_vals['Q1'], q_vals['Q2'], q_vals['Q3']))
                    _lo_is, _hi_is = 0.5 * _min_is, 2.5 * _max_is
                    _best_ann_is, _best_q4_is = None, None
                    for _alt_v in _is_annual_all.get((label, fy), []):
                        _alt_q4 = _alt_v - _prev_9m_is
                        if _lo_is <= _alt_q4 <= _hi_is:
                            if _best_ann_is is None or _alt_v > _best_ann_is:
                                _best_ann_is, _best_q4_is = _alt_v, _alt_q4
                    if _best_q4_is is not None:
                        print(f"  [Restatement Fix] {label} FY{fy}: derived "
                              f"Q4={q_vals['Q4']:,.0f} below run-rate "
                              f"(restated-annual basis mismatch). Original annual "
                              f"{_best_ann_is:,.0f} -> Q4={_best_q4_is:,.0f}")
                        q_vals['Q4'] = _best_q4_is

        # =====================================================================
        # SCOPE-MISMATCH GUARD FOR CASH FLOW Q4  â€”  Root-Cause-First Design
        #
        # Background â€” what a real scope mismatch looks like:
        #   A scope mismatch has one specific XBRL fingerprint: the 10-K
        #   annual fact uses a BROADER concept than the 10-Q quarterly facts.
        #   Example:
        #     Quarterly: PaymentsToAcquirePropertyPlantAndEquipment (PP&E only)
        #     Annual:    PaymentsToAcquireProductiveAssets  (PP&E+intangibles+sw)
        #   In that case Annual âˆ’ YTD9 produces a wildly inflated Q4 because
        #   the annual silently includes extra line items.
        #
        # Why the old ratio-threshold approach was flawed:
        #   Cash-flow lines are inherently lumpy. A company can legitimately
        #   spend 3-5Ã— more in Q4 than in any other quarter (large acquisition,
        #   year-end capex cycle, portfolio rebalancing). A Q4/max(Q1-Q3)
        #   ratio of 3x is NOT evidence of a data error â€” it is normal business
        #   behaviour. The ratio is a symptom of scope mismatch when scope
        #   mismatch exists, but it is equally a symptom of genuine volatility.
        #   Using it as the primary gate produces false positives.
        #
        # Decision tree  (executed only when Q4 was derived, not directly filed):
        #
        #   Gate â€” ratio must be >= 2.5Ã— to even start investigating.
        #          This keeps the common case fast and quiet.
        #
        #   Step 1 â€” Try ALL alternative annual values first.
        #     A plausible alt annual (from _cf_annual_all) that yields a
        #     Q4 inside the 2.5Ã— band is used immediately, regardless of
        #     whether a concept mismatch is detected.  This correctly handles:
        #       â€¢ Restated vs. original annual filings (same concept, different value)
        #       â€¢ Supplemental-note circularity (Case D analog at scope-fix level)
        #     Example: AMD Financing CF FY2021 â€” primary annual wrong, alt annual
        #     (-1,895M) gives plausible Q4 (-727M). âœ“
        #
        #   Step 2 â€” Concept-mismatch diagnostic (only reached if no alt annual).
        #     Compare the XBRL concept of the annual fact (ytd12[1]['Concept'])
        #     against the concepts of the Q1-Q3 facts (best_rows).
        #
        #     a) SAME concept: annual and quarterly use identical XBRL tag.
        #        -> By definition, no scope mismatch. The large Q4 is a real
        #          business event (acquisition, large investment, etc.).
        #        -> PRESERVE Q4. Log an informational note.
        #        Example: AMD Investing CF Q4-2024 (-1,214M is real). âœ“
        #
        #     b) DIFFERENT concept: annual uses a broader/narrower tag.
        #        -> Genuine scope mismatch is likely.
        #        -> DROP for scope-prone sub-component labels only
        #          (Capital Expenditures, Purchases of Investments).
        #        -> For top-level rollup labels (Investing CF, Financing CF,
        #          Operating CF) even a concept mismatch is rare and the
        #          rollup tag is standard; only drop if mismatch is explicit.
        #        Example: NVDA CapEx â€” PaymentsToAcquireProductiveAssets
        #          (annual) vs PaymentsToAcquirePropertyPlantAndEquipment
        #          (quarterly). âœ“
        #
        #     c) Concept UNKNOWN (HTMLFallback, None, or no quarterly facts):
        #        -> Insufficient evidence. Do NOT drop based on ratio alone.
        #        -> PRESERVE Q4 with a warning.
        # =====================================================================
        if (
            cat == '3_Cash_Flow'
            and not is_avg
            and not is_segment_cat
            and q_vals['Q4'] is not None
        ):
            _q123_known = [v for v in [q_vals['Q1'], q_vals['Q2'], q_vals['Q3']]
                           if v is not None and v != 0]
            if len(_q123_known) >= 2:
                _max_q123 = max(abs(v) for v in _q123_known)
                _q4_abs   = abs(q_vals['Q4'])

                if _max_q123 > 0 and _q4_abs > _max_q123 * 2.5:
                    _ratio_current = _q4_abs / _max_q123

                    # ----------------------------------------------------------
                    # Bypass: Q4 came from a directly-filed ~91-day quarterly
                    # XBRL fact.  Scope mismatch can only arise from derivation
                    # (Annual âˆ’ YTD9); a direct quarterly fact IS the real value.
                    # ----------------------------------------------------------
                    if _q4_directly_filed and not _q4_derived_flag:
                        # Informational only â€” no action needed.
                        print(f"  [Scope Fix] {label} FY{fy}: "
                              f"Q4={q_vals['Q4']:,.0f} is {_ratio_current:.1f}x max(Q1-Q3) "
                              f"but is a directly-filed quarterly fact â€” preserving.")

                    else:
                        # Q4 was derived (Annual âˆ’ YTD9). Apply root-cause logic.
                        _prev_9m = sum(v or 0 for v in
                                       [q_vals['Q1'], q_vals['Q2'], q_vals['Q3']])

                        # --------------------------------------------------
                        # Step 1: Try ALL alternative annual values.
                        # The best alt is the one whose implied Q4 falls
                        # within the 2.5Ã— plausibility band.  Among those,
                        # pick the one with the smallest implied Q4 magnitude
                        # (most conservative correction).
                        # --------------------------------------------------
                        _alts = _cf_annual_all.get((label, fy), [])
                        _best_q4     = None
                        _best_annual = None
                        for _alt_v in _alts:
                            _alt_q4 = _alt_v - _prev_9m
                            if abs(_alt_q4) <= _max_q123 * 2.5:
                                if _best_q4 is None or abs(_alt_q4) < abs(_best_q4):
                                    _best_q4     = _alt_q4
                                    _best_annual = _alt_v

                        if _best_q4 is not None:
                            # Alt annual resolves the anomaly â€” use it.
                            print(f"  [Scope Fix] {label} FY{fy}: "
                                  f"Q4={q_vals['Q4']:,.0f} is {_ratio_current:.1f}x "
                                  f"max(Q1-Q3) (derived). "
                                  f"Alt annual {_best_annual:,.0f} -> Q4={_best_q4:,.0f}")
                            q_vals['Q4'] = _best_q4

                        else:
                            # --------------------------------------------------
                            # Step 2: Concept-mismatch diagnostic.
                            # Gather the XBRL concept from the annual fact and
                            # from the Q1-Q3 quarterly facts (via best_rows).
                            # HTMLFallback and None are treated as "unknown".
                            # --------------------------------------------------
                            _UNKNOWN_CONCEPTS = {'HTMLFallback', 'HTMLFallback_Annual',
                                                 'Calculated', None}

                            _annual_concept = (
                                ytd12[1].get('Concept') if ytd12 is not None else None
                            )
                            if _annual_concept in _UNKNOWN_CONCEPTS:
                                _annual_concept = None

                            _qtly_concepts = set()
                            for _qn in ('Q1', 'Q2', 'Q3'):
                                _br = best_rows.get(_qn)
                                if _br is not None:
                                    _c = _br.get('Concept')
                                    if _c not in _UNKNOWN_CONCEPTS:
                                        _qtly_concepts.add(_c)

                            _concept_known = (
                                _annual_concept is not None
                                and len(_qtly_concepts) > 0
                            )
                            _concept_mismatch = (
                                _concept_known
                                and _annual_concept not in _qtly_concepts
                            )

                            # Alias guard: even when the raw concept strings differ,
                            # check whether both the annual and quarterly concepts
                            # resolve to the SAME CONCEPT_MAP label via the two-tier
                            # smart resolver (exact OR fuzzy token match).
                            # If they do, they represent the same financial metric â€”
                            # taxonomy variation only, not a real scope mismatch.
                            #
                            # Handles:
                            #   â€¢ Registered aliases  (Tier 1, exact)
                            #     e.g. NVDA FY2018: annual='PurchasesOfProperty...'
                            #          vs quarterly='PaymentsToAcquireProperty...'
                            #   â€¢ Extension/obscure tags  (Tier 2, fuzzy)
                            #     e.g. company-specific 'AdditionsToPropertyPlant...'
                            #          not yet registered in CONCEPT_MAP
                            if _concept_mismatch and _annual_concept is not None:
                                _annual_cm_label = resolve_concept_to_label(_annual_concept)
                                _qtly_cm_labels  = {
                                    resolve_concept_to_label(c) for c in _qtly_concepts
                                }
                                if (
                                    _annual_cm_label is not None
                                    and _annual_cm_label in _qtly_cm_labels
                                ):
                                    _concept_mismatch = False
                                    print(
                                        f"  [Scope Fix] {label} FY{fy}: "
                                        f"annual='{_annual_concept}' and quarterly="
                                        f"{sorted(_qtly_concepts)} both resolve to "
                                        f"'{_annual_cm_label}' â€” "
                                        f"treating as same scope, preserving Q4."
                                    )

                            if not _concept_mismatch:
                                # Concepts match (or are unknown) -> no evidence of
                                # scope mismatch.  The large Q4 is either a real
                                # business event or we lack the data to diagnose.
                                # In both cases: PRESERVE Q4.
                                if _concept_known:
                                    _reason = (f"annual concept '{_annual_concept}' "
                                               f"matches quarterly â€” real business event")
                                else:
                                    _reason = "concept data unavailable â€” no evidence of mismatch"
                                print(f"  [Scope Fix] {label} FY{fy}: "
                                      f"Q4={q_vals['Q4']:,.0f} is {_ratio_current:.1f}x "
                                      f"max(Q1-Q3) (derived) but {_reason} â€” preserving.")

                            else:
                                # Confirmed concept mismatch: annual uses a
                                # different XBRL concept than the quarterly facts.
                                # This is the canonical scope-mismatch signature.
                                # Drop ONLY for labels known to suffer this:
                                #   - Capital Expenditures (CapEx sub-components)
                                #   - Purchases of Investments
                                # Top-level CF lines (Investing/Operating/Financing)
                                # use standard rollup tags that rarely mismatch;
                                # for safety those are NOT auto-dropped.
                                _DROP_LABELS = {
                                    'Capital Expenditures',
                                    'Purchases of Investments',
                                }
                                if label in _DROP_LABELS:
                                    print(f"  [Scope Fix] {label} FY{fy}: "
                                          f"Q4={q_vals['Q4']:,.0f} is {_ratio_current:.1f}x "
                                          f"max(Q1-Q3) (derived). Concept mismatch confirmed: "
                                          f"annual='{_annual_concept}' vs quarterly="
                                          f"{sorted(_qtly_concepts)} â€” dropping Q4.")
                                    q_vals['Q4'] = None
                                else:
                                    # Even with a concept mismatch on a top-level
                                    # line, preserve and warn â€” the analyst can
                                    # investigate the raw filing if needed.
                                    print(f"  [Scope Fix] {label} FY{fy}: "
                                          f"Q4={q_vals['Q4']:,.0f} is {_ratio_current:.1f}x "
                                          f"max(Q1-Q3) (derived). Concept mismatch detected "
                                          f"(annual='{_annual_concept}' vs quarterly="
                                          f"{sorted(_qtly_concepts)}) but '{label}' is a "
                                          f"top-level CF line â€” preserving with caution.")

        for qn in ['Q1', 'Q2', 'Q3', 'Q4']:
            val = q_vals[qn]
            if val is not None and not pd.isna(val):
                r = best_rows[qn]
                if r is not None:
                    r['Value'] = val
                    r['Period'] = f"{fy}-{qn}"
                    final_rows.append(r)
    
    final_df = pd.DataFrame(final_rows)
    if final_df.empty: return pd.DataFrame()
    final_df = final_df.sort_values(['Category', 'Label', 'Period', 'Filed'], ascending=[True, True, True, False])
    final_df = final_df.drop_duplicates(subset=['Category', 'Label', 'Period'], keep='first')
    pivoted_temp = final_df.pivot(index=['Category', 'Label'], columns='Period', values='Value')
    # Inversion guarantee: every face-statement concept that produced facts
    # must yield an output row. Anything lost to dedup/merging is reported.
    try:
        if _FACE_PRESENTED and 'Concept' in df.columns:
            _face_mask = df['Concept'].isin(_FACE_PRESENTED.keys())
            if 'DimCount' in df.columns:
                _face_mask &= (pd.to_numeric(df['DimCount'], errors='coerce').fillna(0) == 0)
            _face_labels = set(df.loc[_face_mask, 'Label'].dropna().unique())
            _pivot_labels = set(pivoted_temp.index.get_level_values('Label'))
            _noncashish = ('noncash', 'non-cash', 'cashless', 'included in',
                           'paid by', 'obtained in exchange', 'reclassification',
                           'settlement of', 'period increase', 'in exchange for')
            _lost = sorted(
                l for l in _face_labels
                if l not in _pivot_labels
                and ' - ' not in str(l) and ':' not in str(l)
                and '(Net)' not in str(l)
                and not any(p in str(l).lower() for p in _noncashish))
            if _lost:
                print(f"  [Resolve] WARNING: face-statement lines produced no "
                      f"output row ({len(_lost)}): {', '.join(str(x) for x in _lost[:12])}"
                      + (" ..." if len(_lost) > 12 else ""))
            # Extraction-level: presented face concepts with no facts at all.
            _have_facts = set(df['Concept'].dropna().unique())
            _nofacts = sorted(
                c for c, cats in _FACE_PRESENTED.items()
                if c not in _have_facts and 'Abstract' not in c
                and not any(p in c.lower() for p in _noncashish))
            if _nofacts:
                print(f"  [Resolve] NOTE: presented face concept(s) produced no facts "
                      f"({len(_nofacts)}): {', '.join(_nofacts[:10])}"
                      + (" ..." if len(_nofacts) > 10 else ""))
    except Exception:
        pass
    kpi_long = calculate_kpis(pivoted_temp, is_reit=is_reit)
    final_long = pd.concat([final_df[['Category', 'Label', 'Period', 'Value']], kpi_long], ignore_index=True)
    final_long = final_long.drop_duplicates(subset=['Category', 'Label', 'Period'], keep='last')
    final_pivot = final_long.pivot(index=['Category', 'Label'], columns='Period', values='Value')
    import json
    final_pivot = _apply_accounting_engine(final_pivot, is_financial=is_financial, is_insurance=is_insurance, is_oil_gas=is_oil_gas, is_reit=is_reit)
    final_pivot = _apply_industry_kpis(final_pivot, is_financial=is_financial, is_insurance=is_insurance)
    final_pivot = _recompute_cf_residuals(final_pivot)
    final_pivot = _move_noisy_business_segment_rows_to_disclosures(final_pivot)
    final_pivot = _validate_and_repair_segment_data(final_pivot)
    final_pivot = _null_segment_total_leaks(final_pivot)
    final_pivot = _merge_prefix_continuation_members(final_pivot)
    final_pivot = _audit_segment_footing(final_pivot, ticker, ye_month)
    final_pivot = _derive_crosstabulated_matrix(final_pivot)
    final_pivot = _repair_always_positive(final_pivot)
    final_pivot = _repair_q4_from_annual_ytd(final_pivot)
    final_pivot = _remerge_false_positive_splits(final_pivot)
    final_pivot = _neutralize_ytd_undercapture(final_pivot)
    final_pivot = _recompute_cf_residuals(final_pivot)
    final_pivot = _refresh_balance_sheet_closure(final_pivot)
    sorted_cols = sorted(final_pivot.columns, key=lambda x: (x.split('-')[0], x.split('-')[1]), reverse=True)
    return final_pivot[sorted_cols]

_YTD_UNDERCAPTURE_LABELS = {
    'Purchases of Investments', 'Proceeds from Investments',
    'Capital Expenditures',
}


def _neutralize_ytd_undercapture(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    cols = [c for c in df.columns if isinstance(c, str) and len(c) == 7 and c[4] == '-']
    by_fy = {}
    for c in cols:
        by_fy.setdefault(c[:4], {})[c[5:]] = c
    fys = sorted(by_fy)
    touched = []
    for idx in list(df.index):
        if idx[1] not in _YTD_UNDERCAPTURE_LABELS:
            continue
        row = pd.to_numeric(df.loc[idx], errors='coerce')
        for i, fy in enumerate(fys):
            qmap = by_fy[fy]
            if not all(q in qmap for q in ('Q1', 'Q2', 'Q3', 'Q4')):
                continue
            q = {k: row.get(qmap[k]) for k in ('Q1', 'Q2', 'Q3', 'Q4')}
            if any(pd.isna(v) for v in q.values()):
                continue
            fy_sum = sum(abs(v) for v in q.values())
            if fy_sum <= 0:
                continue
            early = abs(q['Q1']) + abs(q['Q2']) + abs(q['Q3'])
            # Signature part 1: Q4 dominates and early quarters are a thin
            # slice of the year (the YTD remainder piled into Q4).
            if not (abs(q['Q4']) > 0.55 * fy_sum and early < 0.30 * fy_sum):
                continue
            # Signature part 2: neighbors prove these early quarters are
            # undercaptured, not merely seasonal. Compare each of Q1-Q3 to
            # the same quarter in the prior and next fiscal year.
            neigh_ratio = []
            for q3 in ('Q1', 'Q2', 'Q3'):
                neigh_vals = []
                for j in (i - 1, i + 1):
                    if 0 <= j < len(fys) and q3 in by_fy[fys[j]]:
                        nv = row.get(by_fy[fys[j]][q3])
                        if pd.notna(nv):
                            neigh_vals.append(abs(nv))
                if neigh_vals:
                    nmed = np.median(neigh_vals)
                    if nmed > 0:
                        neigh_ratio.append(abs(q[q3]) / nmed)
            # Undercapture confirmed when this year's early quarters run far
            # below neighbors (median ratio < 0.5). Real seasonal patterns
            # (IBM Q1-heavy every year) keep ratios near 1.0 and are spared.
            if neigh_ratio and np.median(neigh_ratio) < 0.5:
                for qn in ('Q1', 'Q2', 'Q3', 'Q4'):
                    df.loc[idx, qmap[qn]] = np.nan
                touched.append((idx[1], fy))
    if touched:
        for lbl, fy in touched:
            print(f"  [Scope Fix] Neutralized FY{fy} quarterly split for '{lbl}': "
                  f"Q1-Q3 undercaptured vs adjacent years, derived Q4 would be "
                  f"a YTD artifact (annual figure retained, quarters set NaN).")
    return df


def _repair_q4_from_annual_ytd(df: pd.DataFrame) -> pd.DataFrame:
    """
    Universal Q4 gap-fill: derive missing Q4 values from Annual - YTD9.

    This runs AFTER _repair_always_positive so values are sign-corrected,
    and BEFORE _remerge_false_positive_splits.

    Design
    ------
    For EVERY row in the pivot (income-statement, cash-flow, and segment),
    check if Q4 is missing but the other three quarters and a matching annual
    value exist. If so: Q4 = Annual - Q1 - Q2 - Q3.

    Safety rules (to prevent bad derivation):
    - Skip Balance Sheet rows (instantaneous, not summable).
    - Skip ratio/average rows (EPS, %, Shares, RPO, Useful Life).
    - The derived Q4 must have the same sign as the majority of Q1/Q2/Q3.
      (Prevents e.g. CapEx Q4 = Annual - big-YTD from going negative when
      the annual HTML parse captured a slightly wrong number.)
    - For CapEx and always-positive items: clamp to 0 if derived value is
      slightly negative (rounding artifact), NaN if materially negative.
    - Only derives when the annual column is present and labelled *-Q4
      (i.e. it ended in the Q4 period â€” the annual duration fact sits on
      the Q4 period boundary).

    This is intentionally kept separate from the main per-group loop in
    build_pivoted_data so it operates on the fully-deduped pivot and can
    use neighbouring columns for cross-validation.
    """
    # The derivation logic now lives in build_pivoted_data where duration and
    # source-row metadata are still available. This post-pivot hook is retained
    # for call-site compatibility, but it must be a no-op.
    return df

    SKIP_KW = {'EPS', 'Shares', 'Margin', 'Ratio', 'ROE', 'ROA', '%',
               'Useful Life', 'RPO', 'Metric:'}
    ALWAYS_POS = {'Capital Expenditures', 'Depreciation', 'Amortization',
                  'Stock-Based Compensation', 'Share Repurchases',
                  'Dividends Paid', 'Taxes Paid on Stock Awards',
                  'Total Debt Repaid', 'Purchases of Investments', 'Acquisitions'}

    df = df.copy()
    cols = df.columns.tolist()

    # Group columns by fiscal year
    from collections import defaultdict
    fy_map = defaultdict(dict)   # fy -> {q_label -> col_name}
    for col in cols:
        if not isinstance(col, str) or '-' not in col:
            continue
        parts = col.split('-')
        if len(parts) == 2:
            fy_map[parts[0]][parts[1]] = col

    for fy, q_cols in fy_map.items():
        q1c = q_cols.get('Q1')
        q2c = q_cols.get('Q2')
        q3c = q_cols.get('Q3')
        q4c = q_cols.get('Q4')
        if q4c is None:
            continue   # no Q4 column at all for this FY

        for idx in df.index:
            cat, label = idx

            # Skip balance-sheet and ratio rows
            if cat == '2_Balance_Sheet':
                continue
            if any(kw in label for kw in SKIP_KW):
                continue

            # Only fill if Q4 is actually missing
            q4_val = pd.to_numeric(df.at[idx, q4c], errors='coerce')
            if pd.notna(q4_val):
                continue

            # Gather Q1/Q2/Q3
            def _get(col):
                if col is None or col not in df.columns:
                    return None
                v = pd.to_numeric(df.at[idx, col], errors='coerce')
                return float(v) if pd.notna(v) else None

            v1, v2, v3 = _get(q1c), _get(q2c), _get(q3c)
            n_known = sum(v is not None for v in [v1, v2, v3])
            if n_known < 2:
                continue   # not enough baseline

            # Annual value: must be in the Q4 column with ~365d duration.
            # In the post-pivot frame we don't have Duration, but the HTML
            # recovery function injected it as a regular fact that already
            # won dedup (TagRank=998 < actual Q4 90d = None -> no conflict).
            # So if Q4 is NaN here, there was no 90d fact. The annual fact
            # sits in the same Q4 column IF it was the only value.
            # We treat q4_val = annual only when n_known == 3 (all quarters
            # present) -- at that point the only thing that can populate Q4
            # is an annual-duration value, which is what HTML recovery fills.
            # Re-query from the raw df to check.
            annual_val = _get(q4c)   # already confirmed NaN above -- skip

            # Instead: compute YTD3 and look for a plausible annual in the
            # col using the Investing/Operating anchor approach.
            # Since q4_val is NaN, there is no annual either -- rely on the
            # HTML recovery having already injected the annual as a separate
            # TagRank=998 row. After dedup that becomes our Q4 column value
            # ONLY when the 90d fact was absent. If it's still NaN here, we
            # have nothing to derive from, so skip.
            continue   # derivation already handled in build_pivoted_data

    return df


def _apply_accounting_engine(df, is_financial=False, is_insurance=False, is_oil_gas=False, is_reit=False):
    """
    Applies fundamental accounting rules to derive missing values from known anchors.
    Runs iteratively to allow derived values to feed into subsequent calculations.
    """
    df = df.copy()
    
    def get_v(cat, lbl, col):
        try:
            v = df.at[(cat, lbl), col]
            return float(v) if pd.notna(v) else None
        except KeyError: return None

    def set_v(cat, lbl, col, val):
        if pd.notna(val):
            try:
                df.at[(cat, lbl), col] = val
            except KeyError: pass
            
    # --- NEW: Add this drop_v helper function ---
    def drop_v(cat, lbl, col):
        try:
            df.at[(cat, lbl), col] = None
        except KeyError: pass

    biz_segs = [
        idx[1] for idx in df.index
        if idx[0] == '4a_Segments_Business' and idx[1].startswith('Revenue - ')
    ]

    for col in df.columns:
        
        # --- NEW: TAXONOMY COLLISION RESOLVER ---
        # If an industrial/bank company had its interest tag "stolen" by the insurance label
        if not is_insurance:
            stray_insurance_val = get_v('1_Income_Statement', 'Net Investment Income', col)
            if stray_insurance_val is not None:
                # Re-route the stolen value back to Interest Income
                if get_v('1_Income_Statement', 'Interest Income', col) is None:
                    set_v('1_Income_Statement', 'Interest Income', col, stray_insurance_val)
                # Erase the invalid insurance label from this corporate filing
                drop_v('1_Income_Statement', 'Net Investment Income', col)

        # Run 3 passes to allow cascading (e.g. GP -> OpInc -> Pretax)
        for _ in range(3):
            # --- PHASE 1: INCOME STATEMENT ---
            rev = get_v('1_Income_Statement', 'Revenue', col)
            
            if is_insurance:
                # --- INSURANCE & HYBRID LOGIC (e.g., UNH) ---
                prem = get_v('1_Income_Statement', 'Premiums Earned', col)
                inv_inc = get_v('1_Income_Statement', 'Net Investment Income', col)
                claims = get_v('1_Income_Statement', 'Policyholder Claims/Benefits', col)
                dac = get_v('1_Income_Statement', 'Amortization of DAC', col)
                cogs = get_v('1_Income_Statement', 'Cost of Revenue', col)

                # 1. Derive Total Revenue if standard Revenue tag is missing
                if rev is None and prem is not None:
                    inv_val = inv_inc if inv_inc is not None else 0
                    set_v('1_Income_Statement', 'Revenue', col, prem + inv_val)
                    rev = prem + inv_val
                    
                # 2. Derive Gross Profit for insurance / hybrid companies.
                # Claims/benefits are an insurance COGS equivalent, so a plain
                # Revenue - CostOfRevenue derivation materially overstates gross
                # profit for healthcare insurers.
                gp = get_v('1_Income_Statement', 'Gross Profit', col)
                if rev is not None:
                    claims_val = claims if claims is not None else 0
                    cogs_val = cogs if cogs is not None else 0
                    dac_val = dac if dac is not None else 0
                    insurance_gp = rev - claims_val - cogs_val - dac_val
                    if claims is not None and claims != 0:
                        if gp is None or abs(gp - (rev - cogs_val)) <= max(1.0, abs(rev) * 0.01):
                            set_v('1_Income_Statement', 'Gross Profit', col, insurance_gp)
                            gp = insurance_gp
                    elif gp is None and cogs is not None:
                        set_v('1_Income_Statement', 'Gross Profit', col, rev - cogs)
                        gp = rev - cogs

                # 3. Derive Operating Income and Total Operating Expenses
                opinc = get_v('1_Income_Statement', 'Operating Income', col)
                opex = get_v('1_Income_Statement', 'Total Operating Expenses', col)
                
                if opinc is None and gp is not None and opex is not None:
                    set_v('1_Income_Statement', 'Operating Income', col, gp - opex)
                elif opex is None and gp is not None and opinc is not None:
                    set_v('1_Income_Statement', 'Total Operating Expenses', col, gp - opinc)

                # 4. Aggregation Fallback (If Opex is missing)
                opex = get_v('1_Income_Statement', 'Total Operating Expenses', col)
                if opex is None:
                    sga = get_v('1_Income_Statement', 'Selling, General & Admin', col)
                    known_exps = [x for x in [dac, sga] if x is not None]
                    if known_exps:
                        set_v('1_Income_Statement', 'Total Operating Expenses', col, sum(known_exps))
                        if opinc is None and gp is not None:
                            set_v('1_Income_Statement', 'Operating Income', col, gp - sum(known_exps))
                            
            elif is_financial:
                # --- FINANCIAL INSTITUTION LOGIC ---
                int_inc = get_v('1_Income_Statement', 'Interest Income', col)
                int_exp = get_v('1_Income_Statement', 'Interest Expense', col)
                net_int = get_v('1_Income_Statement', 'Net Interest Income (Expense)', col)
                
                # If Interest Income is missing, intelligently sum its known components
                if False:
                    valid_comps = []
                    for l in df.index.get_level_values('Label').unique():
                        if l.startswith('Income - ') and 'Tax' not in l and 'Comprehensive' not in l:
                            val = get_v('1_Income_Statement', l, col)
                            if val is None or pd.isna(val):
                                val = get_v('4a_Segments_Business', l, col)
                            if pd.notna(val):
                                valid_comps.append(val)
                                
                    if valid_comps and sum(valid_comps) > 0:
                        int_inc = sum(valid_comps)
                        set_v('1_Income_Statement', 'Interest Income', col, int_inc)
                
                if net_int is None and int_inc is not None and int_exp is not None:
                    set_v('1_Income_Statement', 'Net Interest Income (Expense)', col, int_inc - int_exp)
                elif int_inc is None and net_int is not None and int_exp is not None:
                    set_v('1_Income_Statement', 'Interest Income', col, net_int + int_exp)
                    
            else:
                # --- STANDARD CORPORATE LOGIC ---
                cogs = get_v('1_Income_Statement', 'Cost of Revenue', col)
                gp = get_v('1_Income_Statement', 'Gross Profit', col)

                if gp is None and rev is not None and cogs is not None and not is_oil_gas and not is_reit:
                    set_v('1_Income_Statement', 'Gross Profit', col, rev - cogs)
                elif cogs is None and rev is not None and gp is not None:
                    set_v('1_Income_Statement', 'Cost of Revenue', col, rev - gp)
                elif rev is None and gp is not None and cogs is not None:
                    set_v('1_Income_Statement', 'Revenue', col, gp + cogs)

                # Operating Income logic
                opinc = get_v('1_Income_Statement', 'Operating Income', col)
                opex = get_v('1_Income_Statement', 'Total Operating Expenses', col)
                if opinc is None and gp is not None and opex is not None and not is_oil_gas and not is_reit:
                    set_v('1_Income_Statement', 'Operating Income', col, gp - opex)
                elif opex is None and gp is not None and opinc is not None and not is_oil_gas and not is_reit:
                    set_v('1_Income_Statement', 'Total Operating Expenses', col, gp - opinc)

                opex = get_v('1_Income_Statement', 'Total Operating Expenses', col)
                if opinc is not None and opex is not None and gp is not None and rev is not None:
                    incl_ok = abs((rev - opinc) - opex) <= max(1.0, abs(rev) * 0.01)
                    excl_ok = abs((gp - opinc) - opex) <= max(1.0, abs(gp) * 0.01)
                    if not incl_ok and not excl_ok:
                        set_v('1_Income_Statement', 'Total Operating Expenses', col, gp - opinc)

            # Pretax & Net Income
            pretax = get_v('1_Income_Statement', 'Pretax Income', col)
            tax = get_v('1_Income_Statement', 'Income Tax Expense', col)
            ni = get_v('1_Income_Statement', 'Net Income', col)
            
            if ni is None and pretax is not None and tax is not None:
                set_v('1_Income_Statement', 'Net Income', col, pretax - tax)
            elif pretax is None and ni is not None and tax is not None:
                set_v('1_Income_Statement', 'Pretax Income', col, ni + tax)
            elif tax is None and pretax is not None and ni is not None:
                set_v('1_Income_Statement', 'Income Tax Expense', col, pretax - ni)

            # --- PHASE 2: BALANCE SHEET ---
            assets = get_v('2_Balance_Sheet', 'Total Assets', col)
            liab = get_v('2_Balance_Sheet', 'Total Liabilities', col)
            equity = get_v('2_Balance_Sheet', 'Total Equity', col)
            
            if assets is None and liab is not None and equity is not None:
                set_v('2_Balance_Sheet', 'Total Assets', col, liab + equity)
            elif liab is None and assets is not None and equity is not None:
                set_v('2_Balance_Sheet', 'Total Liabilities', col, assets - equity)
            elif equity is None and assets is not None and liab is not None:
                set_v('2_Balance_Sheet', 'Total Equity', col, assets - liab)

            # --- PHASE 3: CASH FLOW ---
            ni_cf = get_v('3_Cash_Flow', 'Net Income (CF)', col)
            if ni_cf is None and ni is not None:
                set_v('3_Cash_Flow', 'Net Income (CF)', col, ni)
            
            # --- PHASE 4: SEGMENTS ---
            if rev is not None and biz_segs:
                known_seg_vals = []
                missing_segs = []
                for s in biz_segs:
                    v = get_v('4a_Segments_Business', s, col)
                    if v is not None: known_seg_vals.append(v)
                    else: missing_segs.append(s)
                
                if len(missing_segs) == 1:
                    missing_val = rev - sum(known_seg_vals)
                    if missing_val >= 0:
                        set_v('4a_Segments_Business', missing_segs[0], col, missing_val)

            if biz_segs:
                for s in biz_segs:
                    base = s.replace('Revenue - ', '')
                    seg_rev = get_v('4a_Segments_Business', s, col)
                    seg_opinc = get_v('4a_Segments_Business', f'Operating Income - {base}', col)
                    seg_opex = get_v('4a_Segments_Business', f'Operating Expenses - {base}', col)
                    
                    if seg_opex is None and seg_rev is not None and seg_opinc is not None:
                        set_v('4a_Segments_Business', f'Operating Expenses - {base}', col, seg_rev - seg_opinc)
                    elif seg_opinc is None and seg_rev is not None and seg_opex is not None:
                        set_v('4a_Segments_Business', f'Operating Income - {base}', col, seg_rev - seg_opex)
                    elif seg_rev is None and seg_opinc is not None and seg_opex is not None:
                        set_v('4a_Segments_Business', f'Revenue - {base}', col, seg_opinc + seg_opex)

    return df

def _apply_industry_kpis(df, is_financial=False, is_insurance=False):
    """
    Calculates dynamic KPI metrics based on the company's industry format.
    """
    df = df.copy()
    
    def get_v(cat, lbl, col):
        try:
            v = df.at[(cat, lbl), col]
            return float(v) if pd.notna(v) else None
        except KeyError: return None

    def set_v(cat, lbl, col, val):
        if pd.notna(val):
            try:
                # Round to 2 decimal places for clean percentage metrics
                df.at[(cat, lbl), col] = round(val, 2)
            except KeyError: pass

    for col in df.columns:
        # Baseline Anchors
        rev = get_v('1_Income_Statement', 'Revenue', col)
        ni = get_v('1_Income_Statement', 'Net Income', col)
        opex = get_v('1_Income_Statement', 'Total Operating Expenses', col)

        # 1. UNIVERSAL KPI
        if ni is not None and rev and rev != 0:
            set_v('5_KPI_Metrics', 'Net Margin (%)', col, (ni / rev) * 100)

        # 2. INDUSTRY SPECIFIC KPIs
        if is_insurance:
            prem = get_v('1_Income_Statement', 'Premiums Earned', col)
            claims = get_v('1_Income_Statement', 'Policyholder Claims/Benefits', col)
            
            if prem and prem != 0:
                if claims is not None:
                    loss_ratio = (claims / prem) * 100
                    set_v('5_KPI_Metrics', 'Loss Ratio (%)', col, loss_ratio)
                if opex is not None:
                    expense_ratio = (opex / prem) * 100
                    set_v('5_KPI_Metrics', 'Expense Ratio (%)', col, expense_ratio)
                    if claims is not None and opex is not None:
                        set_v('5_KPI_Metrics', 'Combined Ratio (%)', col, loss_ratio + expense_ratio)

        elif is_financial:
            # Efficiency Ratio is the golden KPI for banks/brokers
            if opex is not None and rev and rev != 0:
                set_v('5_KPI_Metrics', 'Efficiency Ratio (%)', col, (opex / rev) * 100)
                
        else:
            # Standard Industrial Margins
            gp = get_v('1_Income_Statement', 'Gross Profit', col)
            opinc = get_v('1_Income_Statement', 'Operating Income', col)
            
            if rev and rev != 0:
                if gp is not None:
                    set_v('5_KPI_Metrics', 'Gross Margin (%)', col, (gp / rev) * 100)
                if opinc is not None:
                    set_v('5_KPI_Metrics', 'Operating Margin (%)', col, (opinc / rev) * 100)

    return df

def _demote_non_face_nature_split(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only the income-statement FACE nature-of-revenue split on the face.

    A filer that discloses revenue/cost by ProductOrServiceAxis frequently tags
    BOTH the coarse face split (e.g. Microsoft: Product / Service-and-other;
    IBM: Services / Sales / Financing) AND a finer product-line breakdown
    (Server products, Office, Windows, Gaming, LinkedIn, Search, ...) on the
    same axis. Both partitions foot to the consolidated line, but only the
    coarse one is presented on the face of the income statement. We keep the
    SMALLEST member set that foots to the consolidated total and move every
    finer member back to the business-segment block (nothing is deleted).
    """
    import itertools as _it
    period_cols = list(df.columns)
    to_demote: list[str] = []
    for prefix in ('Revenue', 'Cost of Revenue'):
        tot_idx = ('1_Income_Statement', prefix)
        if tot_idx not in df.index:
            continue
        total = pd.to_numeric(df.loc[tot_idx], errors='coerce')
        all_members = [idx for idx in df.index
                       if idx[0] == '1_Income_Statement'
                       and idx[1].startswith(prefix + ' - ')]
        # A genuine face nature-of-revenue member carries exactly ONE axis, so
        # its label is "<prefix> - <member>" with no further ' - '. Rows that
        # contain additional ' - ' segments are multi-axis sub-disaggregations
        # (e.g. a cash-flow-hedge reclassification crossed with a segment) and
        # are never part of the income-statement face -> demote unconditionally
        # and exclude them from the footing search.
        members = [idx for idx in all_members
                   if ' - ' not in idx[1][len(prefix) + 3:]]
        for idx in all_members:
            if idx not in members:
                to_demote.append(idx[1])
                print(f"  [Face Split] Moving multi-axis sub-disaggregation to "
                      f"segments: '{idx[1]}'")
        if len(members) < 2:
            continue
        mv = {idx: pd.to_numeric(df.loc[idx], errors='coerce') for idx in members}
        ref_cols = [c for c in period_cols
                    if pd.notna(total[c]) and abs(total[c]) > 0
                    and sum(1 for idx in members if pd.notna(mv[idx][c])) >= 2]
        if not ref_cols:
            continue
        ref = ref_cols[0]
        present = [idx for idx in members if pd.notna(mv[idx][ref])]
        tref = float(total[ref])

        def _ref_err(sub):
            # footing error at the most recent reference quarter, where the
            # company's CURRENT revenue structure applies and the true face
            # split foots exactly. Averaging over all history instead would
            # penalise the real split for early-period restatement drift and
            # reward a spurious subset that only foots in a couple of quarters.
            return abs(sum(float(mv[idx][ref]) for idx in sub) - tref) / abs(tref)

        def _foot_count(sub):
            return sum(1 for c in ref_cols
                       if pd.notna(total[c]) and abs(total[c]) > 0
                       and all(pd.notna(mv[idx][c]) for idx in sub)
                       and abs(sum(float(mv[idx][c]) for idx in sub) - float(total[c]))
                       <= abs(float(total[c])) * 0.01)

        # Gather every member subset (up to 6) that foots at the reference
        # quarter within a loose tolerance, then rank: most exact first, then
        # fewest members, then most generic (shortest) labels. The exact-error
        # ranking keeps a complete split (IBM: Services+Sales+Financing) over a
        # near-miss subset that merely drops a small member.
        candidates = []
        for k in range(1, min(len(present), 6) + 1):
            for sub in _it.combinations(present, k):
                if abs(sum(float(mv[idx][ref]) for idx in sub) - tref) <= abs(tref) * 0.02:
                    candidates.append(sub)
        # require footing in >= 2 quarters when possible (robustness)
        robust = [c for c in candidates if _foot_count(c) >= 2]
        pool = robust or candidates
        if not pool:
            continue
        # most exact at the recent quarter, then coarsest, then broadest
        # footing history, then most generic labels.
        pool.sort(key=lambda s: (round(_ref_err(s), 5), len(s),
                                 -_foot_count(s), sum(len(idx[1]) for idx in s)))
        chosen = set(pool[0])
        for idx in members:
            if idx not in chosen:
                to_demote.append(idx[1])
                print(f"  [Face Split] Moving finer product/service member to "
                      f"segments: '{idx[1]}'")

    if not to_demote:
        return df
    flat = df.reset_index()
    m = (flat['Category'] == '1_Income_Statement') & (flat['Label'].isin(to_demote))
    flat.loc[m, 'Category'] = '4a_Segments_Business'
    if flat.duplicated(['Category', 'Label']).any():
        flat = (flat.groupby(['Category', 'Label'], sort=False)
                    .agg(lambda s: s.dropna().iloc[0] if s.dropna().size else np.nan)
                    .reset_index())
    return flat.set_index(['Category', 'Label'])


_ERA_STITCH_STOPWORDS = {
    'and', 'or', 'the', 'of', 'for', 'to', 'in', 'a', 'an', 'net', 'other',
    'services', 'service', 'products', 'product', 'revenue', 'sales', 'gross',
    'income', 'total', 'inc', 'corporation', 'corp', 'segment', 'segments',
    'change', 'pre', 'post',
}


def _stitch_era_renamed_members(df: pd.DataFrame) -> pd.DataFrame:
    """Merge a segment/product member that was RENAMED across eras into one
    continuous series (e.g. Microsoft 'Search advertising' -> 'Search and news
    advertising' -> 'Search advertising').

    Complements the value-continuity merge (which needs >=3 overlapping matching
    periods) by also handling clean successions. Strong false-positive guards:
    the two labels must share a *distinctive* token, any overlapping periods
    must agree within 15%, and the boundary values must be of similar magnitude.
    The surviving name is the one carrying the most recent data point (the
    latest era name); gaps are back-filled from the older series.
    """
    SEG = ('4a_Segments_Business', '4b_Segments_Geographic_Regions',
           '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated')
    period_cols = list(df.columns)
    pos = {c: k for k, c in enumerate(period_cols)}  # lower index = more recent
    from collections import defaultdict as _dd
    groups = _dd(list)
    for idx in df.index:
        cat, lbl = idx
        if cat not in SEG:
            continue
        # Recast bridge rows ('(Pre Change)'/'(Post Change)') deliberately keep
        # the two eras separate -- never stitch them.
        if '(Pre Change)' in lbl or '(Post Change)' in lbl:
            continue
        parts = lbl.split(' - ', 1)
        if len(parts) == 2 and ' - ' not in parts[1]:
            groups[(cat, parts[0])].append(idx)

    def a_old_max(periods, pos):
        return max(pos[c] for c in periods)

    def _toks(member):
        return {t for t in re.split(r'[\s\-]+', member.lower())
                if t and t not in _ERA_STITCH_STOPWORDS and len(t) >= 3}

    merges = []  # (loser_idx, winner_label)
    handled = set()
    for (cat, prefix), idxs in groups.items():
        if len(idxs) < 2:
            continue
        vals = {idx: pd.to_numeric(df.loc[idx], errors='coerce') for idx in idxs}
        pops = {idx: [c for c in period_cols if pd.notna(vals[idx][c])] for idx in idxs}
        for a, b in __import__('itertools').combinations(idxs, 2):
            if a in handled or b in handled:
                continue
            pa, pb = pops[a], pops[b]
            if len(pa) < 2 or len(pb) < 2:
                continue
            ma, mb = a[1].split(' - ', 1)[1], b[1].split(' - ', 1)[1]
            if not (_toks(ma) & _toks(mb)):
                continue  # no distinctive shared token -> unrelated
            overlap = set(pa) & set(pb)
            # overlapping periods must agree (same economic line, slight
            # definitional drift tolerated up to 15%).
            bad = False
            for c in overlap:
                va, vb = vals[a][c], vals[b][c]
                if pd.notna(va) and pd.notna(vb) and max(abs(va), abs(vb)) > 0:
                    if abs(va - vb) / max(abs(va), abs(vb)) > 0.15:
                        bad = True
                        break
            if bad:
                continue
            # winner = series carrying the most recent data point (latest era
            # name); loser back-fills the older gaps.
            a_recent, b_recent = min(pos[c] for c in pa), min(pos[c] for c in pb)
            if a_recent == b_recent:
                continue
            winner, loser = (a, b) if a_recent < b_recent else (b, a)
            if overlap:
                # overlapping periods already agree within 15% above -> same
                # line; that is sufficient evidence of continuity.
                pass
            else:
                # No overlap: require a clean temporal succession whose boundary
                # values are continuous. Compare the earlier series' newest
                # value against the later series' oldest value at the gap.
                earlier, later = (a, b) if a_old_max(pa, pos) > a_old_max(pb, pos) else (b, a)
                e_newest = period_cols[min(pos[c] for c in pops[earlier])]
                l_oldest = period_cols[max(pos[c] for c in pops[later])]
                ve, vl = vals[earlier][e_newest], vals[later][l_oldest]
                if pd.isna(ve) or pd.isna(vl) or ve == 0 or vl == 0:
                    continue
                # later must actually start after earlier ends (no time travel)
                if min(pos[c] for c in pops[later]) >= min(pos[c] for c in pops[earlier]):
                    pass  # later is more recent overall (expected)
                if not (0.5 <= abs(vl / ve) <= 2.0):
                    continue
            merges.append((loser, winner[1]))
            handled.add(loser)
    if not merges:
        return df
    flat = df.reset_index()
    for loser_idx, winner_label in merges:
        print(f"  [Era Rename Stitch] Merging '{loser_idx[1]}' -> "
              f"'{winner_label}' (continuous succession)")
        mask = (flat['Category'] == loser_idx[0]) & (flat['Label'] == loser_idx[1])
        flat.loc[mask, 'Label'] = winner_label
    if flat.duplicated(['Category', 'Label']).any():
        # winner (most recent) takes priority per period; loser back-fills gaps.
        flat = (flat.groupby(['Category', 'Label'], sort=False)
                    .agg(lambda s: s.dropna().iloc[0] if s.dropna().size else np.nan)
                    .reset_index())
    return flat.set_index(['Category', 'Label'])


def _stitch_or_drop_abandoned_face_disagg(df: pd.DataFrame) -> pd.DataFrame:
    """Reconcile income-statement disaggregation members ('Revenue - X') that
    carry no data in the most recent quarters.

    Such a member is not part of the filer's *current* face presentation, but it
    is frequently the older half of a series whose recent half now lives in the
    operating-segment block.  Uber is the canonical case: Mobility/Delivery/
    Freight quarterly revenue for the early years came from a since-retired
    revenue-disaggregation note carried on the face, while the recent quarters
    come from the segment note (category 4a).  The face member foots to total
    revenue at its newest available quarter, so it is kept on the face, yet it
    has nothing in the last four columns -- which previously caused it to be
    dropped outright, taking years of segment history with it.

    So before discarding the member, splice its history into a same-named
    segment row (gap-fill only -- a reported segment value is never overwritten).
    Members with no segment counterpart are dropped exactly as before, and
    genuine current face members (IBM's Services/Sales/Financing, which keep
    recent data) are never considered here at all.
    """
    if df.empty:
        return df
    _SEG = ('4a_Segments_Business', '4b_Segments_Geographic_Regions',
            '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated')
    abandoned = [
        (cat, lbl) for cat, lbl in df.index
        if cat == '1_Income_Statement' and ' - ' in lbl
        and not df.loc[(cat, lbl)].iloc[:4].notna().any()
    ]
    for idx in abandoned:
        cat, lbl = idx
        src = df.loc[idx]
        counterpart = next((sc for sc in _SEG if (sc, lbl) in df.index), None)
        if counterpart is not None:
            dst = df.loc[(counterpart, lbl)].copy()
            fill = dst.isna() & src.notna()
            n = int(fill.sum())
            if n:
                dst[fill] = src[fill]
                df.loc[(counterpart, lbl)] = dst
                print(f"  [Cleanup] Spliced {n} historical period(s) of face member "
                      f"'{lbl}' into {counterpart} before removing it from the income statement")
            else:
                print(f"  [Cleanup] Removing redundant face disaggregation member "
                      f"'{lbl}' (already present in {counterpart})")
        else:
            print(f"  [Cleanup] Dropping abandoned face disaggregation member "
                  f"'{lbl}' (no data in recent quarters)")
        df = df.drop(idx)
    return df


def _null_segment_total_leaks(df: pd.DataFrame) -> pd.DataFrame:
    """Null segment-revenue cells contaminated by the consolidated total -- a
    granular member that is tagged with, or leaks, the whole-company revenue.

    Both patterns are judged against the segment's OWN behaviour, never against
    a single total-revenue cell, so a correct segment value that merely looks
    oversized because *its own total is understated* is never touched:

      Type A (annual leak): a cell exceeds total revenue (>102%) AND is more
        than 2x the median of that same row's other quarters -- an annual figure
        leaked into one quarter. A correct quarter sitting above an understated
        total is ~1x its own median and is spared.

      Type B (total duplicate): a cell is ~100% of total (98-102%) AND the same
        row also reports a genuine small value (<=50% of total) somewhere else,
        proving it is really a sub-component, so the ~100% readings are the
        company total leaking in. A steadily dominant segment (e.g. Meta's
        Family of Apps, ~98% every quarter with no small period) is spared, as
        is a sole segment that legitimately equals the total.

    Runs on the final pivot right after _validate_and_repair_segment_data and
    before the footing audit, across every segment category, so the audit (and
    any HTML rescue) sees clean values.
    """
    SEG = ('4a_Segments_Business', '4b_Segments_Geographic_Regions',
           '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated')
    if ('1_Income_Statement', 'Revenue') not in df.index:
        return df
    total = pd.to_numeric(df.loc[('1_Income_Statement', 'Revenue')], errors='coerce')
    cols = list(df.columns)
    for idx in list(df.index):
        cat, lbl = idx
        if cat not in SEG or not str(lbl).startswith('Revenue - '):
            continue
        row = pd.to_numeric(df.loc[idx], errors='coerce')
        present = {c: float(row[c]) for c in cols if pd.notna(row.get(c))}
        if len(present) < 2:
            continue
        share = {c: present[c] / float(total[c]) for c in present
                 if pd.notna(total.get(c)) and float(total[c]) > 0}
        if not share:
            continue
        has_small = any(v <= 0.50 for v in share.values())
        nulled = []
        for c in list(present.keys()):
            sh = share.get(c)
            if sh is None:
                continue
            others_med = np.median([present[o] for o in present if o != c])
            kind = None
            if sh > 1.02 and others_med > 0 and present[c] > 2.0 * others_med:
                kind = 'A'
            elif 0.98 <= sh <= 1.02 and has_small:
                kind = 'B'
            if kind:
                df.at[idx, c] = np.nan
                nulled.append((c, kind))
        if nulled:
            kinds = '/'.join(sorted({k for _, k in nulled}))
            cells = ', '.join(c for c, _ in nulled[:6]) + ('...' if len(nulled) > 6 else '')
            print(f"  [Segment Leak Guard] {lbl}: nulled {len(nulled)} "
                  f"total-contaminated cell(s) [{kinds}]: {cells}")
    return df


def _move_noisy_business_segment_rows_to_disclosures(df: pd.DataFrame) -> pd.DataFrame:
    """Move obvious non-additive business-segment noise to disclosures.

    These rows are dimensional facts that landed in 4a_Segments_Business but
    are not reportable segment performance lines (acquisition/reclass/gain-loss
    rows, investment-agreement disclosures, tax/receivable leaks).  Moving them
    preserves every value while keeping segment tables and segment-footing checks
    focused on real business segments.
    """
    if df is None or df.empty:
        return df
    noisy = []
    for idx in list(df.index):
        cat, lbl = idx
        if cat != '4a_Segments_Business':
            continue
        text = str(lbl or '')
        if ' - ' in text:
            member = text.split(' - ', 1)[1]
        else:
            member = text
        m = re.sub(r"\s+", " ", member.lower()).strip()
        full = re.sub(r"\s+", " ", text.lower()).strip()
        metric = text.split(' - ', 1)[0].strip().lower()
        if metric == 'net income':
            noisy.append(idx)
            continue
        if any(k in m or k in full for k in (
            'gain loss', 'gain/loss', 'reclass', 'acquisition',
            'investment agreement', 'accounts receivable', 'receivable',
            'tax jurisdiction', 'deferred tax', 'deferred income tax',
            'income taxes', 'tax expense', 'asset impairment',
            'concentration risk', 'share based', 'stock-based',
            'common class', 'common stock', 'available to common stockholders',
            'contract with customer liability', 'deferred revenue', 'revenue deferred',
            'hedging', 'foreign exchange contract', 'foreign currency denominated debt',
            'net investment hedge', 'net investment hedging', 'interest income other',
            'not from contract with customer', 'nonredeemable noncontrolling interest',
            'chief executive officer', 'material reconciling items',
        )):
            noisy.append(idx)
            continue
        if metric.startswith('depreciation') and any(k in m for k in (
            'server', 'network asset', 'servers and network', 'property plant',
            'equipment', 'ppe', 'useful life'
        )):
            noisy.append(idx)
    if not noisy:
        return df
    df = df.copy()
    for idx in noisy:
        row = df.loc[idx].copy()
        new_idx = ('6_Disclosures', f"Business Segment Disclosure - {idx[1]}")
        if new_idx in df.index:
            df.loc[new_idx, :] = df.loc[new_idx].combine_first(row).values
        else:
            df.loc[new_idx, :] = row.values
        df = df.drop(idx)
    print(f"  [Segment Cleanup] Moved {len(noisy)} non-additive business-segment disclosure row(s) out of 4a_Segments_Business.")
    return df


def _clean_segment_member_name(lbl: str) -> str:
    text = str(lbl or '')
    return text.split(' - ', 1)[1].strip() if ' - ' in text else ''


def _is_clean_top_level_segment_revenue_label(lbl: str) -> bool:
    text = str(lbl or '')
    if not text.startswith('Revenue - ') or text.count(' - ') != 1:
        return False
    member = _clean_segment_member_name(text)
    m = re.sub(r"\s+", " ", member.lower()).strip()
    if not m:
        return False
    if any(k in m for k in (
        'external customers', 'intersegment', 'elimination', 'product', 'service',
        'financial service other', 'contract with customer', 'deferred', 'liability',
        'not from contract', 'revenue deferred', 'all other', 'other', 'total optum',
        'total', 'corporate', 'unallocated', 'acquisition', 'reclass', 'gain loss',
        'hedging', 'foreign exchange', 'accounts receivable', 'investment agreement',
    )):
        return False
    return True


def _repair_segment_revenue_residuals(df: pd.DataFrame) -> pd.DataFrame:
    """Repair one understated/missing clean segment revenue row using the
    consolidated revenue residual, when the same basis otherwise ties.

    This is intentionally conservative and generic.  It fixes cases such as
    META Reality Labs Q4 where Family of Apps + Reality Labs is the clean
    segment basis, one small segment is far below the residual, and replacing it
    makes the basis foot to consolidated revenue.  It does not run for partial
    revenue-category matrices such as HOOD or UNH.
    """
    if df is None or df.empty:
        return df
    total_idx = ('1_Income_Statement', 'Revenue')
    if total_idx not in df.index:
        return df
    labels = [lbl for cat, lbl in df.index
              if cat == '4a_Segments_Business' and _is_clean_top_level_segment_revenue_label(lbl)]
    if len(labels) < 2 or len(labels) > 8:
        return df

    total = pd.to_numeric(df.loc[total_idx], errors='coerce')
    seg = pd.DataFrame({lbl: pd.to_numeric(df.loc[('4a_Segments_Business', lbl)], errors='coerce')
                        for lbl in labels}, index=df.columns)
    count = seg.notna().sum(axis=1)
    comparable = total.notna() & (total.abs() > 0) & (count >= 2)
    if int(comparable.sum()) < 4:
        return df
    seg_sum = seg.fillna(0).sum(axis=1)
    gap = total - seg_sum
    gap_ratio = gap.abs() / total.abs().replace(0, np.nan)
    tie_mask = comparable & (gap_ratio <= 0.03)
    # Require this exact basis to usually foot; otherwise it may be a product /
    # customer / partial category disclosure and should not be repaired.
    if int(tie_mask.sum()) < max(3, int(0.50 * int(comparable.sum()))):
        return df

    df = df.copy()
    repaired = []
    for col in df.columns:
        if not comparable.get(col, False):
            continue
        rev_v = total.get(col, np.nan)
        if pd.isna(rev_v) or abs(float(rev_v)) <= 0:
            continue
        cur_vals = seg.loc[col]
        gap_abs = abs(float(rev_v) - float(cur_vals.fillna(0).sum()))
        # Even a 2%-of-sales segment error can be material for a small segment;
        # require both absolute and relative materiality to avoid no-op churn.
        if gap_abs <= max(100e6, 0.002 * abs(float(rev_v))):
            continue
        candidates = []
        for lbl in labels:
            cur_v = cur_vals.get(lbl, np.nan)
            other_sum = cur_vals.drop(labels=[lbl]).fillna(0).sum()
            repl = float(rev_v) - float(other_sum)
            if pd.isna(repl) or repl <= 0:
                continue
            if pd.notna(cur_v):
                cur_f = float(cur_v)
                if not (cur_f <= 0.60 * repl and abs(repl - cur_f) >= max(100e6, 0.40 * abs(repl))):
                    continue
            hist = seg[lbl].drop(labels=[col], errors='ignore').dropna()
            hist = hist[hist > 0]
            if len(hist) >= 4:
                med = float(hist.median())
                if med > 0 and not (0.10 * med <= repl <= 5.0 * med):
                    continue
            candidates.append((lbl, repl, cur_v))
        if len(candidates) == 1:
            lbl, repl, cur_v = candidates[0]
            df.at[('4a_Segments_Business', lbl), col] = repl
            seg.at[col, lbl] = repl
            repaired.append((lbl, col, cur_v, repl))

    if repaired:
        shown = ', '.join(f"{lbl} {col}" for lbl, col, _, _ in repaired[:6])
        extra = '' if len(repaired) <= 6 else f" (+{len(repaired)-6} more)"
        print(f"  [Segment Repair] Repaired clean business-segment revenue residual(s): {shown}{extra}")
    return df


def _move_parent_only_balance_sheet_rows_to_disclosures(df: pd.DataFrame) -> pd.DataFrame:
    """Move parent-company-only balance-sheet disclosures out of the main BS.

    Rows like UNH's 'Equity in Net Assets of Subsidiaries' are standalone /
    parent-company investment disclosures, not consolidated operating assets.
    They can dwarf the balance sheet and confuse downstream users, so preserve
    their values in disclosures instead of the main balance sheet.
    """
    if df is None or df.empty:
        return df
    move = []
    for idx in list(df.index):
        cat, lbl = idx
        if cat != '2_Balance_Sheet':
            continue
        l = re.sub(r"\s+", " ", str(lbl or '').lower()).strip()
        if any(k in l for k in (
            'equity in net assets of subsidiaries',
            'investment in subsidiaries',
            'net assets of subsidiaries',
            'parent company only',
            'parent-only',
        )):
            move.append(idx)
    if not move:
        return df
    df = df.copy()
    for idx in move:
        row = df.loc[idx].copy()
        new_idx = ('6_Disclosures', f"Balance Sheet Disclosure - {idx[1]}")
        if new_idx in df.index:
            df.loc[new_idx, :] = df.loc[new_idx].combine_first(row).values
        else:
            df.loc[new_idx, :] = row.values
        df = df.drop(idx)
    print(f"  [Balance Sheet Cleanup] Moved {len(move)} parent-only balance-sheet disclosure row(s) out of 2_Balance_Sheet.")
    return df


def _effective_capex_series(df: pd.DataFrame) -> pd.Series:
    """Return capex using detail rows when the generic row is missing or stale.

    Some filers/reporting paths expose detailed capex lines (software,
    intangibles, equipment/buildings) while the normalized generic
    ``Capital Expenditures`` row is blank.  A second edge case is a stale
    generic zero while detail rows are nonzero (AVGO 2019-Q1).  In both cases
    the detail sum is stronger evidence for FCF than the missing/stale generic
    row.  A true zero is preserved when no nonzero detail exists.
    """
    idx = df.columns
    capex = (pd.to_numeric(df.loc[('3_Cash_Flow', 'Capital Expenditures')], errors='coerce')
             if ('3_Cash_Flow', 'Capital Expenditures') in df.index
             else pd.Series(np.nan, index=idx))
    detail_sum = pd.Series(0.0, index=idx)
    detail_mask = pd.Series(False, index=idx)
    for lbl in ('Capital Expenditures (Software)', 'Capital Expenditures (Intangibles)', 'Capital Expenditures (Equipment & Buildings)'):
        key = ('3_Cash_Flow', lbl)
        if key in df.index:
            row = pd.to_numeric(df.loc[key], errors='coerce')
            detail_sum = detail_sum + row.fillna(0)
            detail_mask = detail_mask | row.notna()
    detail_nonzero = detail_mask & (detail_sum.abs() > 1e-9)
    capex_missing_or_stale_zero = capex.isna() | ((capex.abs() <= 1e-9) & detail_nonzero)
    return capex.where(~capex_missing_or_stale_zero, detail_sum.where(detail_mask))


def _effective_depreciation_amortization_series(df: pd.DataFrame) -> pd.Series:
    """Return the best supported D&A amount for EBITDA for every period.

    A filed combined cash-flow add-back is strongest because it captures D&A
    allocated across both cost of revenue and operating expenses.  When it is
    absent, derive the amount from separately surfaced depreciation and
    amortization components.  Income-statement D&A totals are the final fallback.
    Components are only added when a combined amount is unavailable (or is a
    stale zero), preventing a filed total from being counted again as detail.
    """
    if df is None or df.empty:
        return pd.Series(dtype='float64')
    cols = df.columns

    def row(category, label):
        key = (category, label)
        if key not in df.index:
            return pd.Series(np.nan, index=cols, dtype='float64')
        values = pd.to_numeric(df.loc[key], errors='coerce')
        # D&A is an expense/add-back.  Reject sign-inverted or reversal facts
        # instead of silently reducing EBITDA with an economically invalid input.
        return values.where(values >= 0)

    combined_cf = row('3_Cash_Flow', 'Depreciation & Amortization')

    depreciation = row('3_Cash_Flow', 'Depreciation')
    dep_and_depletion = row('3_Cash_Flow', 'Depreciation & Depletion')
    depreciation = depreciation.combine_first(dep_and_depletion)

    amortization = row('3_Cash_Flow', 'Amortization')
    intangible_amortization = row('3_Cash_Flow', 'Amortization of Intangibles (CF)')
    amortization = amortization.combine_first(intangible_amortization)

    component_mask = depreciation.notna() | amortization.notna()
    derived_components = (
        depreciation.fillna(0) + amortization.fillna(0)
    ).where(component_mask)

    # A real reported zero is retained unless nonzero component evidence proves
    # the aggregate is stale/incomplete for that period.
    stale_zero = (
        combined_cf.notna()
        & (combined_cf.abs() <= 1e-9)
        & derived_components.notna()
        & (derived_components.abs() > 1e-9)
    )
    effective = combined_cf.where(~stale_zero, derived_components)
    effective = effective.combine_first(derived_components)

    income_statement_total = row(
        '1_Income_Statement', 'Depreciation, Depletion & Amortization'
    ).combine_first(row('1_Income_Statement', 'Depreciation & Amortization Expense'))
    return effective.combine_first(income_statement_total)


def _refresh_ebitda(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute EBITDA and its margin from the latest EBIT and D&A."""
    if df is None or df.empty:
        return df
    ebit_key = ('1_Income_Statement', 'Operating Income')
    if ebit_key not in df.index:
        return df
    ebit = pd.to_numeric(df.loc[ebit_key], errors='coerce')
    da = _effective_depreciation_amortization_series(df).reindex(df.columns)
    have = ebit.notna() & da.notna()
    if not have.any():
        return df
    out = df.copy()
    ebitda = (ebit + da).where(have)
    out.loc[('5_KPI_Metrics', 'Metric: EBITDA'), :] = ebitda.values
    revenue_key = ('1_Income_Statement', 'Revenue')
    if revenue_key in out.index:
        revenue = pd.to_numeric(out.loc[revenue_key], errors='coerce')
        margin = ((ebitda / revenue) * 100).where(revenue.notna() & (revenue != 0))
        margin_key = ('5_KPI_Metrics', 'Metric: EBITDA Margin %')
        if margin.notna().any() or margin_key in out.index:
            out.loc[('5_KPI_Metrics', 'Metric: EBITDA Margin %'), :] = margin.values
    return out


def _refresh_fcf_from_effective_capex(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute FCF/UFCF from OCF and effective capex for every period.

    This final pass catches historical quarters whose KPI rows came from older
    cached/legacy logic before capex detail rows were promoted to effective
    capex.  It only touches FCF-related KPI rows and fills the generic capex row
    where it was blank but capex detail rows exist.
    """
    if df is None or df.empty or ('3_Cash_Flow', 'Operating Cash Flow') not in df.index:
        return df
    df = df.copy()
    idx = df.columns
    ocf = pd.to_numeric(df.loc[('3_Cash_Flow', 'Operating Cash Flow')], errors='coerce')
    eff_capex = _effective_capex_series(df)
    if ('3_Cash_Flow', 'Capital Expenditures') not in df.index and eff_capex.notna().any():
        df.loc[('3_Cash_Flow', 'Capital Expenditures'), :] = eff_capex.values
    elif ('3_Cash_Flow', 'Capital Expenditures') in df.index:
        capex = pd.to_numeric(df.loc[('3_Cash_Flow', 'Capital Expenditures')], errors='coerce')
        fill = (capex.isna() | ((capex.abs() <= 1e-9) & eff_capex.notna() & (eff_capex.abs() > 1e-9))) & eff_capex.notna()
        if fill.any():
            df.loc[('3_Cash_Flow', 'Capital Expenditures'), fill.index[fill]] = eff_capex[fill].values
    fcf = ocf - eff_capex.fillna(0)
    if fcf.notna().any():
        df.loc[('5_KPI_Metrics', 'Metric: Free Cash Flow'), :] = fcf.values
        rev = pd.to_numeric(df.loc[('1_Income_Statement', 'Revenue')], errors='coerce') if ('1_Income_Statement', 'Revenue') in df.index else pd.Series(np.nan, index=idx)
        df.loc[('5_KPI_Metrics', 'Metric: FCF Margin %'), :] = ((fcf / rev) * 100).values
    int_exp = pd.to_numeric(df.loc[('1_Income_Statement', 'Interest Expense')], errors='coerce').fillna(0) if ('1_Income_Statement', 'Interest Expense') in df.index else pd.Series(0.0, index=idx)
    tax = pd.to_numeric(df.loc[('1_Income_Statement', 'Income Tax Expense')], errors='coerce').fillna(0) if ('1_Income_Statement', 'Income Tax Expense') in df.index else pd.Series(0.0, index=idx)
    pretax = pd.to_numeric(df.loc[('1_Income_Statement', 'Pretax Income')], errors='coerce') if ('1_Income_Statement', 'Pretax Income') in df.index else pd.Series(np.nan, index=idx)
    tax_rate = (tax / pretax).where(pretax > 0, 0.0).clip(0, 1)
    ufcf = fcf + int_exp * (1 - tax_rate)
    if ufcf.notna().any():
        df.loc[('5_KPI_Metrics', 'Metric: Unlevered Free Cash Flow'), :] = ufcf.values
    return df


def _repair_total_net_debt_from_components(df: pd.DataFrame) -> pd.DataFrame:
    """Refresh Total Net Debt Issued (Repaid) with layered debt evidence.

    Priority by period:
      1) if both gross total issued and repaid are visible, net = issued - repaid;
      2) if short-term net-debt evidence is available, combine it with long-term
         net evidence instead of letting one-sided gross rows erase it;
      3) if only one gross total is visible and no short-term net evidence can
         explain the missing side, use the one-sided gross value (missing side=0);
      4) otherwise preserve a directly reported net-debt fact.

    This fixes AVGO-style periods where total repayments exist but the old net
    row stayed at zero, while preserving UNH-style periods where the correct net
    line is short-term net change plus long-term debt activity.
    """
    if df is None or df.empty:
        return df
    idx = df.columns
    def ser(label):
        key = ('3_Cash_Flow', label)
        return pd.to_numeric(df.loc[key], errors='coerce') if key in df.index else pd.Series(np.nan, index=idx)
    def first_nonmissing(*series):
        out = pd.Series(np.nan, index=idx)
        for s in series:
            out = out.where(out.notna(), s)
        return out

    issued = ser('Total Debt Issued')
    repaid = ser('Total Debt Repaid')
    net_existing = ser('Total Net Debt Issued (Repaid)')

    st_issued = ser('Short-term Debt Issued')
    st_repaid = ser('Short-term Debt Repaid')
    st_direct = first_nonmissing(ser('Net Short-Term Debt Issued (Repaid)'),
                                 ser('Net Change in Short-term Debt'))
    st_from_components = (st_issued.fillna(0) - st_repaid.fillna(0)).where(st_issued.notna() | st_repaid.notna())
    st_net = first_nonmissing(st_direct, st_from_components)

    lt_issued = ser('Long-term Debt Issued')
    lt_repaid = ser('Long-term Debt Repaid')
    lt_direct = ser('Net Long-Term Debt Issued (Repaid)')
    lt_from_components = (lt_issued.fillna(0) - lt_repaid.fillna(0)).where(lt_issued.notna() | lt_repaid.notna())
    lt_net = first_nonmissing(lt_from_components, lt_direct)

    component_mask = st_net.notna() | lt_net.notna()
    component_net = st_net.fillna(0) + lt_net.fillna(0)
    gross_mask = issued.notna() | repaid.notna()
    gross_pair = issued.notna() & repaid.notna()
    gross_net = issued.fillna(0) - repaid.fillna(0)

    # Start from direct reported net facts, then replace only where stronger
    # evidence is available.
    net = net_existing.copy()

    # Complete gross totals are the strongest display identity.
    net[gross_pair] = gross_net[gross_pair]

    # Component net is strongest for one-sided gross periods when short-term net
    # debt is explicitly reported (UNH-style: short-term net + long-term activity).
    explicit_short_net = st_direct.notna() & (st_direct.abs() > 1e-9)
    use_component = component_mask & ~gross_pair & explicit_short_net
    net[use_component] = component_net[use_component]

    # If no explicit short-term net evidence exists, one-sided gross debt rows
    # should still affect total net debt. This catches AVGO 2024-Q2 and similar
    # periods where repayment totals were visible but the previous net stayed 0.
    one_sided_gross = gross_mask & ~gross_pair
    use_gross_one_sided = one_sided_gross & ~explicit_short_net
    net[use_gross_one_sided] = gross_net[use_gross_one_sided]

    # If there is no gross evidence, but component net exists, use it.
    use_component_no_gross = component_mask & ~gross_mask
    net[use_component_no_gross] = component_net[use_component_no_gross]

    if net.notna().any():
        df = df.copy()
        df.loc[('3_Cash_Flow', 'Total Net Debt Issued (Repaid)'), :] = net.values
    return df


def _move_duplicate_short_term_debt_net_change_to_disclosures(df: pd.DataFrame) -> pd.DataFrame:
    """Preserve duplicate raw short-term-debt net-change rows outside main CF.

    ``Net Change in Short-term Debt`` is often the raw XBRL source used to build
    normalized ``Net Short-Term Debt Issued (Repaid)``.  When both rows carry the
    same payload, keeping both in the main cash-flow statement is noisy and can
    make the debt block look duplicated.  Move the raw duplicate to disclosures;
    if it differs materially, keep it in cash flow and let the sorter place it
    before the normalized net row.
    """
    raw_idx = ('3_Cash_Flow', 'Net Change in Short-term Debt')
    norm_idx = ('3_Cash_Flow', 'Net Short-Term Debt Issued (Repaid)')
    if df is None or df.empty or raw_idx not in df.index or norm_idx not in df.index:
        return df
    raw = pd.to_numeric(df.loc[raw_idx], errors='coerce')
    norm = pd.to_numeric(df.loc[norm_idx], errors='coerce')
    both = raw.notna() & norm.notna()
    raw_only = raw.notna() & norm.isna()
    if not both.any() or raw_only.any():
        return df
    scale = pd.concat([raw.abs(), norm.abs()], axis=1).max(axis=1).clip(lower=1.0)
    equal = ((raw - norm).abs() <= scale * 1e-9) | (raw.isna() & norm.isna())
    # Move only if every populated overlapping period is equal and the raw row
    # has no distinct value that the normalized row lacks.
    if not bool(equal[both].all()):
        return df
    df = df.copy()
    disc_idx = ('6_Disclosures', 'Cash Flow Disclosure - Net Change in Short-term Debt')
    if disc_idx in df.index:
        df.loc[disc_idx, :] = df.loc[disc_idx].combine_first(df.loc[raw_idx]).values
    else:
        df.loc[disc_idx, :] = df.loc[raw_idx].values
    df = df.drop(raw_idx)
    print("  [Cash Flow Cleanup] Moved duplicate raw 'Net Change in Short-term Debt' row to disclosures.")
    return df


def _refresh_insurance_statement_subtotals(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize insurance/health-benefit P&L subtotals.

    When policyholder/medical claims are present, they are cost-of-revenue
    equivalents.  A generic Revenue - Cost of Revenue gross-profit derivation
    overstates gross profit for insurers such as UNH.  This pass recomputes:
      Gross Profit = Revenue - Claims/Benefits - Cost of Revenue - DAC amort.
      Total Operating Expenses = Gross Profit - Operating Income
    and refreshes related margin / insurance KPI rows.
    """
    if df is None or df.empty:
        return df
    claims_key = ('1_Income_Statement', 'Policyholder Claims/Benefits')
    rev_key = ('1_Income_Statement', 'Revenue')
    if claims_key not in df.index or rev_key not in df.index:
        return df
    df = df.copy()
    cols = df.columns
    def row(cat, label):
        key = (cat, label)
        return pd.to_numeric(df.loc[key], errors='coerce') if key in df.index else pd.Series(np.nan, index=cols)
    claims = row('1_Income_Statement', 'Policyholder Claims/Benefits')
    rev = row('1_Income_Statement', 'Revenue')
    cogs = row('1_Income_Statement', 'Cost of Revenue').fillna(0)
    dac = row('1_Income_Statement', 'Amortization of DAC').fillna(0)
    have = rev.notna() & claims.notna() & (claims.abs() > 0)
    if not have.any():
        return df
    gp = rev - claims - cogs - dac
    df.loc[('1_Income_Statement', 'Gross Profit'), have.index[have]] = gp[have].values
    opinc = row('1_Income_Statement', 'Operating Income')
    opex = (gp - opinc).where(have & opinc.notna())
    if opex.notna().any():
        df.loc[('1_Income_Statement', 'Total Operating Expenses'), opex.index[opex.notna()]] = opex.dropna().values
    total_costs = (rev - opinc).where(have & opinc.notna())
    if total_costs.notna().any():
        df.loc[('1_Income_Statement', 'Total Costs and Expenses'), total_costs.index[total_costs.notna()]] = total_costs.dropna().values
    # KPI refreshes.
    if rev.notna().any():
        df.loc[('5_KPI_Metrics', 'Metric: Gross Margin %'), :] = ((gp / rev) * 100).values
        df.loc[('5_KPI_Metrics', 'Metric: EBIT Margin %'), :] = ((opinc / rev) * 100).values
    prem = row('1_Income_Statement', 'Premiums Earned')
    if prem.notna().any():
        loss_ratio = (claims / prem) * 100
        exp_ratio = (opex / prem) * 100
        df.loc[('5_KPI_Metrics', 'Loss Ratio (%)'), :] = loss_ratio.values
        df.loc[('5_KPI_Metrics', 'Expense Ratio (%)'), :] = exp_ratio.values
        df.loc[('5_KPI_Metrics', 'Combined Ratio (%)'), :] = (loss_ratio + exp_ratio).values
    return df


def _refresh_granular_footing_checks(df: pd.DataFrame) -> pd.DataFrame:
    """Publish granular, low-noise footing checks.

    The old version treated nonzero "Other Adjustments" rows as failures, even
    though they are explicit reconciliation plugs.  This version checks the
    visible accounting identities directly and suppresses segment footing when
    the segment basis is mixed/overlapping instead of emitting false failures.
    """
    if df is None or df.empty:
        return df
    df = df.copy()
    cols = df.columns
    def row(cat, label):
        key = (cat, label)
        return pd.to_numeric(df.loc[key], errors='coerce') if key in df.index else pd.Series(np.nan, index=cols)
    def flag(mask_have, mask_pass):
        out = pd.Series(np.nan, index=cols)
        out[mask_have] = 0.0
        out[mask_have & mask_pass] = 1.0
        return out
    def close(a, b, scale, pct=0.01, floor=50e6):
        return (a - b).abs() <= (scale.abs() * pct).clip(lower=floor)

    # Income statement: require the identities that are actually available.
    rev = row('1_Income_Statement', 'Revenue')
    cor = row('1_Income_Statement', 'Cost of Revenue')
    claims = row('1_Income_Statement', 'Policyholder Claims/Benefits').fillna(0)
    dac = row('1_Income_Statement', 'Amortization of DAC').fillna(0)
    gp = row('1_Income_Statement', 'Gross Profit')
    opx = row('1_Income_Statement', 'Total Operating Expenses')
    opinc = row('1_Income_Statement', 'Operating Income')
    tax = row('1_Income_Statement', 'Income Tax Expense')
    pretax = row('1_Income_Statement', 'Pretax Income')
    ni = row('1_Income_Statement', 'Net Income')
    nci = row('1_Income_Statement', 'Net Income to Noncontrolling Interest').fillna(0)

    is_have = pd.Series(False, index=cols)
    is_pass = pd.Series(True, index=cols)
    gp_have = rev.notna() & cor.notna() & gp.notna()
    gp_expected = rev - cor - claims - dac
    is_have |= gp_have
    is_pass &= (~gp_have | close(gp, gp_expected, rev, pct=0.015))

    op_have = gp.notna() & opx.notna() & opinc.notna()
    is_have |= op_have
    is_pass &= (~op_have | close(opinc, gp - opx, rev.where(rev.notna(), gp), pct=0.015))

    disc = row('1_Income_Statement', 'Income from Discontinued Operations').fillna(0)
    disc_alt = row('1_Income_Statement', 'Income (Loss) from Discontinued Operations').fillna(0)
    discontinued = disc.where(disc.abs() > 1e-9, disc_alt)
    ni_have = pretax.notna() & tax.notna() & ni.notna()
    is_have |= ni_have
    # Allow NCI presentation differences and discontinued-operation presentation.
    # Broadcom/AVGO-style quarters can reconcile only after discontinued ops are
    # added below continuing operations: NI = Pretax - Tax +/- NCI + DiscOps.
    base_cont = pretax - tax
    ni_parent_ok = close(ni, base_cont - nci, pretax.where(pretax.notna(), ni), pct=0.015)
    ni_consol_ok = close(ni, base_cont, pretax.where(pretax.notna(), ni), pct=0.015)
    ni_parent_disc_ok = close(ni, base_cont - nci + discontinued, pretax.where(pretax.notna(), ni), pct=0.015)
    ni_consol_disc_ok = close(ni, base_cont + discontinued, pretax.where(pretax.notna(), ni), pct=0.015)
    is_pass &= (~ni_have | ni_parent_ok | ni_consol_ok | ni_parent_disc_ok | ni_consol_disc_ok)
    is_flag = flag(is_have, is_pass)

    # Balance sheet closure.
    assets, liab, eq = row('2_Balance_Sheet', 'Total Assets'), row('2_Balance_Sheet', 'Total Liabilities'), row('2_Balance_Sheet', 'Total Equity')
    bs_have = assets.notna() & liab.notna() & eq.notna()
    bs_pass = (assets - liab - eq).abs() <= (assets.abs() * 0.005).clip(lower=50e6)
    bs_flag = flag(bs_have, bs_pass)

    # Cash flow bridge: Net Cash Flow is before FX in this script's convention.
    ocf, icf, fcf_fin, netcf = row('3_Cash_Flow', 'Operating Cash Flow'), row('3_Cash_Flow', 'Investing Cash Flow'), row('3_Cash_Flow', 'Financing Cash Flow'), row('3_Cash_Flow', 'Net Cash Flow')
    cf_have = ocf.notna() & icf.notna() & fcf_fin.notna() & netcf.notna()
    cf_pass = (ocf + icf + fcf_fin - netcf).abs() <= (netcf.abs() * 0.01).clip(lower=50e6)
    cf_flag = flag(cf_have, cf_pass)

    # Segment revenue footing: find a clean non-overlapping basis per period.
    all_segment_business_labels = [lbl for cat, lbl in df.index if cat == '4a_Segments_Business']
    seg_labels = [lbl for lbl in all_segment_business_labels if _is_clean_top_level_segment_revenue_label(lbl)]
    seg_flag = pd.Series(np.nan, index=cols)
    # Basis code: 1=clean additive/ties, -1=clean additive/mismatch,
    # 2=mixed or overlapping segment basis not forced, 0=no usable segment basis.
    seg_basis_code = pd.Series(np.nan, index=cols)
    seg_error_pct = pd.Series(np.nan, index=cols)
    if len(seg_labels) >= 2:
        total = row('1_Income_Statement', 'Revenue')
        seg_rows = {lbl: pd.to_numeric(df.loc[('4a_Segments_Business', lbl)], errors='coerce') for lbl in seg_labels}
        def basis_groups(labels):
            groups = []
            groups.append(labels)
            for marker in ('(Post Change)', '(Pre Change)'):
                g = [l for l in labels if marker.lower() in l.lower()]
                if len(g) >= 2:
                    groups.append(g)
            # Prefer clean named segment rows without nested product/customer rows.
            base = [l for l in labels if not any(k in l.lower() for k in (
                'advertising', 'subscription', 'platform', 'devices', 'search', 'network',
                'external customers', 'intersegment', 'elimination', 'product', 'service',
                'other products', 'other services', 'cloud services', 'office products',
                'server products', 'windows division', 'search advertising', 'contract'
            ))]
            if len(base) >= 2:
                groups.append(base)
            # Deduplicate preserving order.
            out=[]; seen=set()
            for g in groups:
                key=tuple(g)
                if key not in seen:
                    seen.add(key); out.append(g)
            return out
        for col in cols:
            if pd.isna(total.get(col)) or float(total.get(col)) == 0:
                continue
            present = [l for l in seg_labels if pd.notna(seg_rows[l].get(col))]
            best = None
            for g in basis_groups(present):
                if not (2 <= len(g) <= 8):
                    continue
                s = sum(float(seg_rows[l][col]) for l in g if pd.notna(seg_rows[l].get(col)))
                ratio = abs(float(total[col]) - s) / abs(float(total[col]))
                if best is None or ratio < best[0]:
                    best = (ratio, g)
            if best is not None:
                seg_error_pct[col] = best[0] * 100.0
                if best[0] <= 0.05:
                    seg_flag[col] = 1.0
                    seg_basis_code[col] = 1.0
                else:
                    # A clean-looking basis exists but does not tie closely.
                    # Preserve this as a basis tag without forcing a broad
                    # financials failure; historical segment tables can change
                    # basis or include reconciling items.
                    seg_basis_code[col] = -1.0
            elif present:
                seg_basis_code[col] = 2.0
            else:
                seg_basis_code[col] = 0.0
            # If no clean basis ties, leave the verification flag NaN rather
            # than false-failing a mixed dimensional disclosure (UNH/MSFT-style matrices).
    elif all_segment_business_labels:
        total = row('1_Income_Statement', 'Revenue')
        seg_basis_code[total.notna()] = 0.0

    checks = {
        'Metric: Income Statement Footing Verified': is_flag,
        'Metric: Balance Sheet Footing Verified': bs_flag,
        'Metric: Cash Flow Footing Verified': cf_flag,
        'Metric: Segment Revenue Footing Verified': seg_flag,
    }
    for label, series in checks.items():
        if series.notna().any():
            df.loc[('8_Integrity_Checks', label), :] = series.values

    # Basis rows are diagnostics, not pass/fail checks.  Keep them out of the
    # broad Financials Footing roll-up so mixed-basis segment disclosures do
    # not turn into false failures.
    diagnostic_rows = {
        'Metric: Segment Revenue Footing Basis Code': seg_basis_code,
        'Metric: Segment Revenue Footing Error %': seg_error_pct,
    }
    for label, series in diagnostic_rows.items():
        if series.notna().any():
            df.loc[('8_Integrity_Checks', label), :] = series.values

    valid = [s for s in checks.values() if s.notna().any()]
    broad = pd.DataFrame(valid).min(axis=0, skipna=True) if valid else pd.Series(np.nan, index=cols)
    if broad.notna().any():
        df.loc[('8_Integrity_Checks', 'Metric: Financials Footing Verified'), :] = broad.values
    return df


def _apply_quality_result_fixes(df: pd.DataFrame) -> pd.DataFrame:
    """Final quality repair bundle for issues found in the 2026-07 run."""
    if df is None or df.empty:
        return df
    df = _repair_total_net_debt_from_components(df)
    df = _move_duplicate_short_term_debt_net_change_to_disclosures(df)
    df = _refresh_insurance_statement_subtotals(df)
    df = _refresh_ebitda(df)
    df = _refresh_fcf_from_effective_capex(df)
    df = _repair_segment_revenue_residuals(df)
    df = _move_noisy_business_segment_rows_to_disclosures(df)
    df = _move_parent_only_balance_sheet_rows_to_disclosures(df)
    df = _refresh_granular_footing_checks(df)
    # Collapse any category moves/added rows safely without changing values for
    # duplicate row labels. Existing newest/rightmost value wins per pandas first.
    if df.index.duplicated().any():
        df = df.groupby(level=['Category', 'Label']).first()
    return df


def _merge_prefix_continuation_members(df: pd.DataFrame) -> pd.DataFrame:
    """Merge a segment member whose name is a mangled extension of another
    member of the SAME underlying series.

    Uber's pre-2020 mobility revenue is tagged with a member that humanises to
    'Mobilityharing' -- a corruption of the modern 'Mobility' member, of which
    'Mobility' is a literal string prefix. The era-stitch misses it because the
    two share no whole *token*. Here two segment 'Revenue - X' labels are paired
    when one member string is a strict prefix of the other (>=5 chars) and their
    values track within 8% across >=2 overlapping quarters -- strong evidence of
    one series. The label with the longer history (the clean modern name)
    survives; the fragment is folded in by gap-fill across every metric, then
    dropped. Gap-fill never transfers a value exceeding total revenue for its
    quarter, so a contaminated donor cell cannot reintroduce a leak.
    """
    SEG = ('1_Income_Statement', '4a_Segments_Business', '4b_Segments_Geographic_Regions',
           '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated')
    total = None
    if ('1_Income_Statement', 'Revenue') in df.index:
        total = pd.to_numeric(df.loc[('1_Income_Statement', 'Revenue')], errors='coerce')
    cols = list(df.columns)
    rename_map = {}   # donor member -> survivor member (applied across all metrics)
    for cat in SEG:
        labs = [lbl for (c, lbl) in df.index
                if c == cat and str(lbl).startswith('Revenue - ') and lbl.count(' - ') == 1]
        members = {lbl: lbl.split(' - ', 1)[1] for lbl in labs}
        vals = {lbl: pd.to_numeric(df.loc[(cat, lbl)], errors='coerce') for lbl in labs}
        for a in labs:
            ma = members[a].lower().replace(' ', '')
            if len(ma) < 5:
                continue
            for b in labs:
                if b == a:
                    continue
                mb = members[b].lower().replace(' ', '')
                if not (ma in mb and mb != ma):
                    continue   # require a (string) substring relation: a in b
                ov = [c for c in cols if pd.notna(vals[a].get(c)) and pd.notna(vals[b].get(c))]
                if len(ov) < 2:
                    continue
                diffs = []
                for c in ov:
                    va, vb = float(vals[a][c]), float(vals[b][c])
                    d = max(abs(va), abs(vb))
                    if d > 0:
                        diffs.append(abs(va - vb) / d)
                if not diffs or (sum(diffs) / len(diffs)) >= 0.08:
                    continue
                na = int(vals[a].notna().sum())
                nb = int(vals[b].notna().sum())
                survivor_lbl, donor_lbl = (a, b) if na >= nb else (b, a)
                rename_map[members[donor_lbl]] = members[survivor_lbl]
    if not rename_map:
        return df
    for donor_mem, surv_mem in rename_map.items():
        suffix = f" - {donor_mem}"
        for (cat, lbl) in list(df.index):
            if not str(lbl).endswith(suffix):
                continue
            surv_lbl = lbl[: -len(suffix)] + f" - {surv_mem}"
            donor_row = pd.to_numeric(df.loc[(cat, lbl)], errors='coerce')
            if (cat, surv_lbl) in df.index:
                surv_row = pd.to_numeric(df.loc[(cat, surv_lbl)], errors='coerce')
                for c in cols:
                    dv = donor_row.get(c)
                    if pd.isna(dv) or pd.notna(surv_row.get(c)):
                        continue   # gap-fill only; never overwrite survivor
                    if total is not None and pd.notna(total.get(c)) and float(total[c]) > 0                             and float(dv) > float(total[c]) * 1.5:
                        continue   # never transfer a >total (contaminated) cell
                    df.at[(cat, surv_lbl), c] = dv
                df = df.drop((cat, lbl))
                print(f"  [Member Variant] Merged mangled member '{lbl}' -> "
                      f"'{surv_lbl}' (prefix continuation of one series)")
            else:
                df = df.rename(index={(cat, lbl): (cat, surv_lbl)})
    return df


def _refresh_balance_sheet_closure(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute 'Metric: Balance Sheet Closure Verified' on the final pivot.

    calculate_kpis emits this flag early, from the PRE-reconciliation Total
    Assets / Liabilities / Equity. When the accounting engine then rebuilds a
    mis-contexted Total Assets, a period that now closes exactly is still
    flagged 0.0 from that stale pass (seen on Uber 2020-Q4 / 2021-Q2 /
    2024-Q3). Here we recompute the flag from the final reconciled rows using
    the identical tolerance (2% or $50M).

    Crucially we only refresh periods that ALREADY hold a closure value -- i.e.
    where Liabilities+Equity were independently reported at calc time. We never
    fabricate a flag for a period/company that had none, because by this point
    a derived Liabilities (= Assets - Equity) would make the check a tautology
    (this is exactly why the original metric stayed absent for such filers,
    e.g. Amazon). A period that genuinely fails (a real L or E capture gap)
    stays 0.0.
    """
    idx_clo = ('8_Integrity_Checks', 'Metric: Balance Sheet Closure Verified')
    BS = '2_Balance_Sheet'
    if idx_clo not in df.index or (BS, 'Total Assets') not in df.index:
        return df
    ta = pd.to_numeric(df.loc[(BS, 'Total Assets')], errors='coerce')
    tl = (pd.to_numeric(df.loc[(BS, 'Total Liabilities')], errors='coerce')
          if (BS, 'Total Liabilities') in df.index else pd.Series(np.nan, index=df.columns))
    te = (pd.to_numeric(df.loc[(BS, 'Total Equity')], errors='coerce')
          if (BS, 'Total Equity') in df.index else pd.Series(np.nan, index=df.columns))
    ident = tl + te
    # Redeemable / temporary (mezzanine) equity -- pre-IPO redeemable convertible
    # preferred, redeemable NCI -- is presented OUTSIDE permanent stockholders'
    # equity, between liabilities and equity, so for a filer that carries it the
    # true identity is A = L + TempEq + E. Treat a period as closing if it
    # reconciles under EITHER reading; taking the better residual can only rescue
    # a period, never turn an already-closing one into a failure (verified:
    # rescues PLTR 2019-Q4, zero regressions incl. TSLA/O redeemable-NCI periods).
    _temp_mask = (df.index.get_level_values(0) == BS) & df.index.get_level_values(1).str.contains(
        r'^Temporary Equity|Redeemable Noncontrolling Interest|^Redeemable Convertible Preferred',
        case=False, regex=True, na=False)
    if _temp_mask.any():
        _tempsum = df[_temp_mask].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=0)
        resid = pd.concat([(ta - ident).abs(), (ta - ident - _tempsum).abs()], axis=1).min(axis=1)
    else:
        resid = (ta - ident).abs()
    tol = (ta.abs() * 0.02).clip(lower=50e6)
    existing = pd.to_numeric(df.loc[idx_clo], errors='coerce')
    refreshed = 0
    for c in df.columns:
        if pd.isna(existing.get(c)):
            continue   # never had an independently-verifiable closure -> leave absent
        if pd.isna(ta.get(c)) or pd.isna(ident.get(c)):
            continue
        new_v = 1.0 if resid[c] <= tol[c] else 0.0
        if new_v != float(existing[c]):
            df.at[idx_clo, c] = new_v
            refreshed += 1
    if refreshed:
        print(f"  [Closure Refresh] Recomputed {refreshed} balance-sheet "
              f"closure flag(s) on reconciled totals")
    return df


def _repair_balance_sheet_identity(df: pd.DataFrame) -> pd.DataFrame:
    """Recover an under-captured balance-sheet total from the identity A = L + E.

    Two distinct, mutually-exclusive capture failures are repaired -- only on
    periods that currently FAIL to close (|A-(L+E)| over 2% of A or $50M); a
    sheet already within tolerance is never touched.

    1. Liabilities under-captured.  A whole liability line is never tagged, so the
       derived Total Liabilities (which equals the sum of its captured parts)
       sits a few percent below Assets-Equity.  Seen on AMD after the Xilinx
       acquisition -- the deferred-tax-liability booked against ~$27B of acquired
       intangibles (a $3.1B line, amortising down 2022-2023) is uncaptured -- and
       on AMZN 2018-Q4 and GOOGL 2015-Q3 (uncaptured long-term tax/other lines).
       Trigger: a positive gap, modest (<15% of Assets), that the equity side
       cannot explain.  Action: Total Liabilities = Assets - Equity.

    2. Equity total on the wrong restatement basis.  The captured Total Equity is
       a later restated figure while Liabilities is still original-basis, so L+E
       overshoots Assets.  Seen on MSFT FY2016-17: Total Equity carries the
       ASC-606 restated value (+$11B from deferred revenue pulled into equity)
       while Liabilities still holds the original deferred revenue.  Trigger: the
       sum of the captured equity components (common stock + APIC + retained
       earnings + AOCI) closes the identity while the captured total does not.
       Action: Total Equity = sum(components) -- the original-basis figure that is
       consistent with the original-basis Assets and Liabilities.

    Both are grounded: Assets and the equity components are independently-reported
    face values, so the recovered total matches the as-filed balance sheet
    (verified to the dollar on AMD, AMZN, GOOGL, MSFT).  This mirrors the existing
    Total-Assets reconstruction in calculate_kpis.  A '... Reconstructed' marker
    is emitted so the derived totals are distinguishable from captured ones.
    """
    BS = '2_Balance_Sheet'; INT = '8_Integrity_Checks'
    if (BS, 'Total Assets') not in df.index:
        return df
    df = df.copy()
    EQ_COMP = ['Common Stock', 'Additional Paid-In Capital', 'Retained Earnings',
               'Comprehensive Income', 'Preferred Stock, Value, Outstanding',
               'Convertible Preferred Stock, Nonredeemable or Redeemable, Issuer Option, Value',
               'Accumulated Distributions in Excess of Net Income']
    def _num(lbl):
        idx = (BS, lbl)
        return (pd.to_numeric(df.loc[idx], errors='coerce')
                if idx in df.index else pd.Series(np.nan, index=df.columns))
    ta, tl, te = _num('Total Assets'), _num('Total Liabilities'), _num('Total Equity')
    eq_rows = [l for l in EQ_COMP if (BS, l) in df.index]
    core_present = (BS, 'Common Stock') in df.index or (BS, 'Additional Paid-In Capital') in df.index
    has_re = ((BS, 'Retained Earnings') in df.index
              or (BS, 'Accumulated Distributions in Excess of Net Income') in df.index)
    eq_comp = (sum((_num(l).fillna(0) for l in eq_rows), start=pd.Series(0.0, index=df.columns))
               if eq_rows else pd.Series(np.nan, index=df.columns))
    n_eq = (sum((((BS, l) in df.index and df.loc[(BS, l)].notna()) for l in eq_rows))
            if eq_rows else pd.Series(0, index=df.columns))
    fixed_tl, fixed_te = {}, {}
    for c in df.columns:
        a, l, e = ta.get(c), tl.get(c), te.get(c)
        if pd.isna(a) or pd.isna(l) or pd.isna(e):
            continue
        tol = max(abs(a) * 0.02, 50e6)
        resid = a - (l + e)
        if abs(resid) <= tol:
            continue
        ec = eq_comp.get(c) if eq_rows else np.nan
        npres = int(n_eq.get(c, 0)) if eq_rows else 0
        # (2) equity total on the wrong basis: components close it, captured total does not
        if (pd.notna(ec) and core_present and has_re and npres >= 2
                and abs(a - l - ec) <= tol and abs(e - ec) > tol and abs(a - l - ec) < abs(resid)):
            df.at[(BS, 'Total Equity'), c] = ec
            fixed_te[c] = (e, ec)
            continue
        # (1) liabilities under-captured: modest positive gap the equity side cannot explain
        if resid > tol and resid < abs(a) * 0.15 and e > 0 and (not eq_rows or ec <= e + tol):
            df.at[(BS, 'Total Liabilities'), c] = a - e
            fixed_tl[c] = (l, a - e)
    for c, (old, new) in sorted(fixed_tl.items()):
        print(f"  [BS Identity] {c}: Total Liabilities {old/1e9:.2f}B -> {new/1e9:.2f}B "
              f"(recovered from Assets - Equity; a liability line is uncaptured)")
    if fixed_tl:
        df.loc[(INT, 'Metric: Total Liabilities Reconstructed'), list(fixed_tl)] = 1.0
    for c, (old, new) in sorted(fixed_te.items()):
        print(f"  [BS Identity] {c}: Total Equity {old/1e9:.2f}B -> {new/1e9:.2f}B "
              f"(captured total was on a different restatement basis; using components)")
    if fixed_te:
        df.loc[(INT, 'Metric: Total Equity Reconstructed'), list(fixed_te)] = 1.0
    return df


def _validate_and_repair_segment_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    SEG_CATS_LOCAL = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'}
    seg_rev_labels = [
        (cat, lbl)
        for (cat, lbl) in df.index
        if cat in SEG_CATS_LOCAL
        and lbl.startswith('Revenue - ')
    ]
    if not seg_rev_labels:
        return df

    # Consolidated revenue series for plausibility gate
    total_rev: pd.Series | None = None
    if ('1_Income_Statement', 'Revenue') in df.index:
        total_rev = pd.to_numeric(df.loc[('1_Income_Statement', 'Revenue')], errors='coerce')

    # Group columns by fiscal year
    fy_to_cols: dict[str, list[str]] = {}
    for col in df.columns:
        if isinstance(col, str) and '-' in col:
            fy = col.split('-')[0]
            fy_to_cols.setdefault(fy, []).append(col)

    for fy, cols in fy_to_cols.items():
        q_map = {str(c).split('-')[1]: c for c in cols if '-' in str(c)}
        q1c = q_map.get('Q1'); q2c = q_map.get('Q2')
        q3c = q_map.get('Q3'); q4c = q_map.get('Q4')

        for seg_idx in seg_rev_labels:
            if seg_idx not in df.index:
                continue

            def _get(col):
                if col is None or col not in df.columns:
                    return None
                v = pd.to_numeric(df.at[seg_idx, col], errors='coerce')
                return float(v) if pd.notna(v) else None

            v1, v2, v3, v4 = _get(q1c), _get(q2c), _get(q3c), _get(q4c)

            def _share(vv, col):
                if vv is None or col is None or total_rev is None or col not in total_rev.index:
                    return None
                t = pd.to_numeric(total_rev[col], errors='coerce')
                return (vv / float(t)) if (pd.notna(t) and float(t) > 0) else None

            # 0. Fix YTD9 contamination in Q3
            if v1 is not None and v2 is not None and v3 is not None:
                s1 = _share(v1, q1c)
                s2 = _share(v2, q2c)
                s3 = _share(v3, q3c)
                if s1 is not None and s2 is not None and s3 is not None:
                    s_avg = (s1 + s2) / 2.0
                    if s_avg > 0 and s3 > s_avg * 1.8:
                        discrete_q3 = v3 - (v1 + v2)
                        s3_discrete = _share(discrete_q3, q3c)
                        if discrete_q3 > 0 and s3_discrete is not None and abs(s3_discrete - s_avg) < s_avg * 0.6 + 0.08:
                            df.at[seg_idx, q3c] = discrete_q3
                            print(f"  [Segment Q3 Fix] YTD9 contamination corrected: {seg_idx[1]} {fy}-Q3: {v3:,.0f} -> {discrete_q3:,.0f}")
                            
                            # Re-derive Q4 if it was negative due to YTD9 overlap
                            if v4 is not None and v4 < 0 and q4c is not None:
                                annual_val = v4 + v1 + v2 + v3 
                                new_q4 = annual_val - (v1 + v2 + discrete_q3)
                                df.at[seg_idx, q4c] = new_q4
                                print(f"  [Segment Q4 Fix] Recovered Q4 after YTD9 fix: {seg_idx[1]} {fy}-Q4: -> {new_q4:,.0f}")
                                v4 = new_q4
                            v3 = discrete_q3

            # 0.5 Fix YTD6 contamination in Q2
            if v1 is not None and v2 is not None:
                s1 = _share(v1, q1c)
                s2 = _share(v2, q2c)
                if s1 is not None and s2 is not None and s1 > 0 and s2 > s1 * 1.8:
                    discrete_q2 = v2 - v1
                    s2_discrete = _share(discrete_q2, q2c)
                    if discrete_q2 > 0 and s2_discrete is not None and abs(s2_discrete - s1) < s1 * 0.6 + 0.08:
                        df.at[seg_idx, q2c] = discrete_q2
                        print(f"  [Segment Q2 Fix] YTD6 contamination corrected: {seg_idx[1]} {fy}-Q2: {v2:,.0f} -> {discrete_q2:,.0f}")
                        
                        if v4 is not None and v4 < 0 and v3 is not None and q4c is not None:
                            annual_val = v4 + v1 + v2 + v3 
                            new_q4 = annual_val - (v1 + discrete_q2 + v3)
                            df.at[seg_idx, q4c] = new_q4
                            print(f"  [Segment Q4 Fix] Recovered Q4 after YTD6 fix: {seg_idx[1]} {fy}-Q4: -> {new_q4:,.0f}")
                            v4 = new_q4
                        v2 = discrete_q2

            # -- Check 1: Annual contamination ------------------------------
            known_quarters = [v for v in [v1, v2, v3] if v is not None]
            if len(known_quarters) == 3 and v4 is not None and q4c is not None:
                ytd3 = v1 + v2 + v3
                if ytd3 > 0 and v4 > ytd3 * 0.8:
                    s4 = _share(v4, q4c)
                    s_avg = np.mean([s for s in [_share(v1, q1c), _share(v2, q2c), _share(v3, q3c)] if s is not None])
                    if s4 is not None and s_avg > 0 and s4 > s_avg * 1.8:
                        discrete_q4 = v4 - ytd3
                        max_q = max(v1, v2, v3)
                        if 0 < discrete_q4 < max_q * 6:
                            df.at[seg_idx, q4c] = discrete_q4
                            lbl = seg_idx[1]
                            print(f"  [Segment Q4 Fix] Annual-value contamination corrected: "
                                  f"{lbl} {fy}-Q4: {v4:,.0f} -> {discrete_q4:,.0f}")
                            v4 = discrete_q4

            # -- Check 2: Plausibility gate (segment > total revenue) --------
            if total_rev is not None and q4c is not None and q4c in total_rev.index:
                tot = pd.to_numeric(total_rev[q4c], errors='coerce')
                if pd.notna(tot) and tot > 0 and v4 is not None and v4 > tot * 1.05:
                    df.at[seg_idx, q4c] = np.nan
                    lbl = seg_idx[1]
                    print(f"  [Segment Q4 Fix] Value exceeds total revenue ({v4:,.0f} > {tot:,.0f}), "
                          f"nulled for HTML rescue: {lbl} {fy}-Q4")
                    v4 = None

            # -- Check 4: Annual-leak via implausible Q4 revenue share -------
            if v4 is not None and v4 > 0 and total_rev is not None and q4c is not None:
                _prior = [s for s in (_share(v1, q1c), _share(v2, q2c), _share(v3, q3c)) if s is not None]
                _q4s = _share(v4, q4c)
                if len(_prior) >= 2 and _q4s is not None:
                    _avg = sum(_prior) / len(_prior)
                    if _avg > 0 and _q4s > _avg * 2.5 and (_q4s - _avg) > 0.20:
                        df.at[seg_idx, q4c] = np.nan
                        lbl = seg_idx[1]
                        print(f"  [Segment Q4 Fix] Annual-leak nulled (Q4 share "
                              f"{_q4s:.0%} vs typical {_avg:.0%}): {lbl} {fy}-Q4 = {v4:,.0f}")
                        v4 = None

            # -- Check 3: Implausible negative segment revenue ---------------
            if v4 is not None and v4 < 0 and q4c is not None:
                max_q = max(v1 or 0, v2 or 0, v3 or 0)
                if abs(v4) > max_q * 0.5:
                    df.at[seg_idx, q4c] = np.nan
                    lbl = seg_idx[1]
                    print(f"  [Segment Q4 Fix] Implausible negative revenue nulled: {lbl} {fy}-Q4 = {v4:,.0f}")

    return df


def _reconcile_segment_partition_from_total(df: pd.DataFrame) -> pd.DataFrame:
    """
    Generic, data-driven repair of segment-revenue partitions.

    Many filers disaggregate consolidated Revenue into a set of mutually
    exclusive, collectively exhaustive members (e.g. Mastercard's
    North America + International Markets, or Broadcom's Americas + Asia +
    EMEA). When such a partition is *empirically validated* -- its leaf members
    actually sum to consolidated Revenue across most periods -- then any single
    period in which the partition fails to foot must be caused by exactly one
    bad-or-missing leaf, which is recoverable from the pure accounting identity

        bad_or_missing_leaf = Revenue - sum(other leaves)

    This recovers, for instance, Mastercard's Q4 geographic figures (disclosed
    only annually, so they otherwise arrive missing or as a mis-scaled residual)
    with no ticker-specific logic and no extra network round-trip.

    The repair is deliberately conservative:
      * Only categories proven to partition Revenue are eligible -- at least
        MIN_GOOD periods, and at least GOOD_FRAC of all fully-present periods,
        must foot within FOOT_TOL. Overlapping / partial / mis-scaled
        breakdowns (which never foot) are therefore left completely untouched.
      * A missing leaf is filled only when EXACTLY ONE leaf is absent and the
        implied residual is positive, material, and in line with that leaf's
        own neighbouring quarters.
      * A present-but-wrong leaf is corrected only when EXACTLY ONE leaf is a
        clear temporal outlier whose identity-implied value lands back in line
        with its neighbours. If two or more leaves look wrong (a wholesale
        capture failure) nothing is changed -- that is left for HTML rescue.
    """
    SEG_CATS = {'4a_Segments_Business', '4b_Segments_Geographic_Regions',
                '4c_Segments_Geographic_Countries'}
    PARENT = ('1_Income_Statement', 'Revenue')
    FOOT_TOL   = 0.015   # |sum(leaf) - total| <= 1.5% counts as footing
    MIN_GOOD   = 4       # need this many footing periods to trust a partition
    GOOD_FRAC  = 0.60    # ... and this share of all fully-present periods
    OUT_LO     = 0.55    # leaf < 0.55x neighbour-ref  => suspiciously low
    OUT_HI     = 1.80    # leaf > 1.80x neighbour-ref  => suspiciously high
    FILL_FLOOR = 0.01    # only fill a hole worth > 1% of total revenue

    if PARENT not in df.index:
        return df
    df = df.copy()
    total = pd.to_numeric(df.loc[PARENT], errors='coerce')
    import os as _os
    _dbg = bool(_os.environ.get('SEG_DEBUG'))
    if _dbg:
        print(f"[SEG_DBG] ENTER _reconcile_segment_partition_from_total: "
              f"{len(df.columns)} cols, parent_in_index={PARENT in df.index}")

    def _porder(col):
        try:
            fy, q = str(col).split('-')
            return (int(fy), int(q.lstrip('Qq')))
        except Exception:
            return (0, 0)

    ordered_cols = sorted(
        [c for c in df.columns
         if pd.notna(total.get(c)) and float(total.get(c)) != 0],
        key=_porder)

    for cat in SEG_CATS:
        leaves = [(c, l) for (c, l) in df.index
                  if c == cat and isinstance(l, str)
                  and l.startswith('Revenue - ') and l.count(' - ') == 1]
        if len(leaves) < 2:
            continue   # a single member cannot partition anything

        vals = {leaf: pd.to_numeric(df.loc[leaf], errors='coerce') for leaf in leaves}

        # --- 1. validate that these leaves really partition Revenue --------
        good_cols, foot_ratio, n_full = [], {}, 0
        for col in ordered_cols:
            t = float(total[col])
            present = [vals[leaf].get(col) for leaf in leaves]
            if any(pd.isna(p) for p in present):
                continue
            n_full += 1
            s = float(np.nansum(present))
            if t > 0 and abs(s - t) / t <= FOOT_TOL:
                good_cols.append(col)
                foot_ratio[col] = s / t
        if _dbg:
            _passes = (len(good_cols) >= MIN_GOOD and n_full > 0
                       and len(good_cols) / n_full >= GOOD_FRAC)
            print(f"[SEG_DBG] cat={cat} leaves={[l[1] for l in leaves]} "
                  f"n_full={n_full} good={len(good_cols)} eligible={_passes}")
            for _c in [c for c in ordered_cols if str(c).endswith('-Q4')]:
                _cells = '  '.join(
                    (l[1].split(' - ')[-1][:7] + '=' +
                     ('NA' if pd.isna(vals[l].get(_c)) else format(float(vals[l].get(_c)), ',.0f')))
                    for l in leaves)
                _t = total.get(_c)
                _ts = 'NA' if pd.isna(_t) else format(float(_t), ',.0f')
                print(f"[SEG_DBG]    {_c}: total={_ts}  |  {_cells}")
        if len(good_cols) < MIN_GOOD or n_full == 0 or len(good_cols) / n_full < GOOD_FRAC:
            continue   # not a trustworthy partition -> leave category untouched

        global_R = float(np.median(list(foot_ratio.values())))

        def _ref(leaf, col):
            # median of the (up to 4) temporally-nearest good-period values
            tk = _porder(col)[0] * 4 + _porder(col)[1]
            near = sorted(good_cols,
                          key=lambda g: abs((_porder(g)[0] * 4 + _porder(g)[1]) - tk))
            picks = []
            for g in near:
                x = vals[leaf].get(g)
                if pd.notna(x):
                    picks.append(float(x))
                if len(picks) >= 4:
                    break
            return float(np.median(picks)) if picks else None

        def _Rlocal(col):
            # prefer the footing ratio of good periods in the same fiscal year
            fy = _porder(col)[0]
            same = [foot_ratio[g] for g in good_cols if _porder(g)[0] == fy]
            return float(np.median(same)) if same else global_R

        # --- 2. repair only the periods that fail to foot ------------------
        for col in ordered_cols:
            if col in good_cols:
                continue
            t = float(total[col])
            present = {leaf: vals[leaf].get(col) for leaf in leaves}
            missing = [leaf for leaf, v in present.items() if pd.isna(v)]
            have    = {leaf: float(v) for leaf, v in present.items() if pd.notna(v)}
            s_present = sum(have.values())

            # CASE A: exactly one leaf missing -> identity fill
            if len(missing) == 1 and len(have) == len(leaves) - 1:
                residual = t * _Rlocal(col) - s_present
                if residual <= 0 or residual < t * FILL_FLOOR:
                    continue          # nothing meaningful missing (e.g. era design)
                leaf = missing[0]
                ref = _ref(leaf, col)
                if ref is not None and not (OUT_LO * ref <= residual <= OUT_HI * ref):
                    continue          # implied value wildly off this leaf's own trend
                df.at[leaf, col] = residual
                vals[leaf].at[col] = residual
                print(f"  [Segment Partition Fill] {leaf[1]} {col}: "
                      f"derived {residual:,.0f} = Revenue - sum(other members)")
                continue

            # CASE B: all leaves present but the partition does not foot ------
            if not missing and len(have) == len(leaves):
                if t > 0 and abs(s_present - t) / t <= FOOT_TOL:
                    continue          # actually fine
                culprits = []
                for leaf in leaves:
                    v = have[leaf]
                    ref = _ref(leaf, col)
                    if ref is None or ref <= 0:
                        continue
                    is_outlier = (v < OUT_LO * ref) or (v > OUT_HI * ref)
                    residual = t * _Rlocal(col) - (s_present - v)
                    fits = (residual > 0) and (OUT_LO * ref <= residual <= OUT_HI * ref)
                    if is_outlier and fits:
                        culprits.append((leaf, residual, v, ref))
                if len(culprits) == 1:   # unambiguous single bad member
                    leaf, residual, oldv, ref = culprits[0]
                    df.at[leaf, col] = residual
                    vals[leaf].at[col] = residual
                    print(f"  [Segment Partition Fix] {leaf[1]} {col}: "
                          f"{oldv:,.0f} -> {residual:,.0f} (outlier vs ~{ref:,.0f}; "
                          f"reset to Revenue - sum(others))")
                # 0 or >1 culprits -> ambiguous / wholesale failure -> untouched

    return df


def _has_subset_sum_near_target(items, target, tolerance=0.05):
    """Return whether any non-empty value subset is within tolerance of target.

    This is mathematically equivalent to the exhaustive combinations loop used
    by the segment-footing audit. For 13-18 values it uses meet-in-the-middle
    subset sums (O(2^(n/2)) instead of O(2^n)); small sets retain the original
    enumeration order and arithmetic.
    """
    values = [v for _label, v in items]
    n = len(values)
    if n == 0:
        return False

    denominator = max(target, 1)
    if n <= 12:
        for r in range(1, n + 1):
            for combo in itertools.combinations(values, r):
                if abs(target - sum(combo)) / denominator <= tolerance:
                    return True
        return False

    # The caller caps this path at 18 positive segment values. Convert the
    # original relative-error predicate into an inclusive sum interval.
    radius = tolerance * denominator
    lower, upper = target - radius, target + radius
    split = n // 2
    left_values, right_values = values[:split], values[split:]

    left_sums = [0.0]
    for value in left_values:
        left_sums += [subtotal + value for subtotal in left_sums]
    right_sums = [0.0]
    for value in right_values:
        right_sums += [subtotal + value for subtotal in right_sums]
    right_sums.sort()

    from bisect import bisect_left
    for left_sum in left_sums:
        minimum_right = lower - left_sum
        maximum_right = upper - left_sum
        pos = bisect_left(right_sums, minimum_right)
        if pos < len(right_sums) and right_sums[pos] <= maximum_right:
            # All inputs reaching this helper are positive, so a zero combined
            # sum can only be the forbidden empty subset.
            if left_sum != 0 or right_sums[pos] != 0:
                return True
    return False


def _audit_segment_footing(df, ticker, ye_month):
    """
    Performs an accounting-driven footing audit using subset sums to avoid 
    double-counting when multiple restatement eras coexist in the same quarter.
    """
    df = df.copy()
    if ('1_Income_Statement', 'Revenue') not in df.index:
        return df
        
    rev_n = pd.to_numeric(df.loc[('1_Income_Statement', 'Revenue')], errors='coerce')
    seg_cats = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'}
    all_failed_quarters = set()
    
    for cat in seg_cats:
        # Require exactly 1 ' - ' for 4a/4b/4c to prevent matching cross-tabulated sub-breakdowns that don't directly foot
        # For 4d, we allow 2 ' - 's.
        if cat == '4d_Segments_Cross_Tabulated':
            seg_labels = [(c, lbl) for (c, lbl) in df.index if c == cat and lbl.startswith('Revenue - ') and lbl.count(' - ') >= 1]
        else:
            seg_labels = [(c, lbl) for (c, lbl) in df.index if c == cat and lbl.startswith('Revenue - ') and lbl.count(' - ') == 1]

        if not seg_labels:
            continue

        for q_col in df.columns:
            if pd.isna(rev_n[q_col]) or rev_n[q_col] == 0: continue
            
            q_vals = {lbl: pd.to_numeric(df.loc[(c, lbl), q_col], errors='coerce') for (c, lbl) in seg_labels}
            q_vals = {k: v for k, v in q_vals.items() if pd.notna(v) and v > 0}
            
            if not q_vals: continue
            
            total_rev = rev_n[q_col]
            
            clean_vals = {k: v for k, v in q_vals.items() if not any(kw in k.lower() for kw in ['elimination', 'corporate', 'unallocated', 'reconcil'])}
            
            subtotal_labels = set()
            clean_items_list = list(clean_vals.items())
            for i, (lbl_i, val_i) in enumerate(clean_items_list):
                others = [(l, v) for j, (l, v) in enumerate(clean_items_list) if j != i]
                if len(others) >= 2:
                    for r in range(2, min(5, len(others) + 1)):
                        found_match = False
                        for combo in itertools.combinations(others, r):
                            combo_sum = sum(v for _, v in combo)
                            if combo_sum > 0 and abs(val_i - combo_sum) / max(abs(val_i), 1) < 0.02:
                                subtotal_labels.add(lbl_i)
                                found_match = True
                                break
                        if found_match:
                            break
            if subtotal_labels:
                clean_vals = {k: v for k, v in clean_vals.items() if k not in subtotal_labels}
            
            items = list(clean_vals.items())
            valid_combo_found = False
            
            if len(items) <= 18:
                valid_combo_found = _has_subset_sum_near_target(
                    items, total_rev, tolerance=0.05
                )
            else:
                sorted_v = sorted(items, key=lambda x: x[1], reverse=True)
                current_sum = 0
                for k, v in sorted_v:
                    if current_sum + v <= total_rev * 1.05:
                        current_sum += v
                if abs(total_rev - current_sum) / max(total_rev, 1) <= 0.05:
                    valid_combo_found = True
                    
            if not valid_combo_found:
                all_failed_quarters.add(q_col)

    if all_failed_quarters:
        failed_list = sorted(list(all_failed_quarters))
        print(f"  [Segment Footing Audit] Failed footing check in quarters: {failed_list}. Attempting HTML rescue...")
        df = _rescue_segments_from_html(df, failed_list, ticker, ye_month)
                        
    return df

def _rescue_segments_from_html(df, failed_quarters, ticker, ye_month):
    with _ProfileTimer("_rescue_segments_from_html"):
        return _rescue_segments_from_html_impl(df, failed_quarters, ticker, ye_month)


def _rescue_segments_from_html_impl(df, failed_quarters, ticker, ye_month):
    """
    Fallback: Fetch the actual segment revenue from the filing's HTML when
    XBRL-derived values fail the footing audit.
    """
    seg_cats = {'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'}
    try:
        company = fetch_company(ticker)
        filings = list(get_company_filings(company, ["10-K", "10-Q"]).head(20))
        filing_by_period = _build_filing_period_lookup(filings, ye_month, set(failed_quarters))
        revenue_row = pd.to_numeric(
            df.loc[('1_Income_Statement', 'Revenue')], errors='coerce'
        ) if ('1_Income_Statement', 'Revenue') in df.index else pd.Series(dtype='float64')
        _expected_segments_by_cat = {
            _cat: [
                idx[1].replace('Revenue - ', '')
                for idx in df.index
                if idx[0] == _cat and idx[1].startswith('Revenue - ')
            ]
            for _cat in ['4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries']
        }

        for q_col in failed_quarters:
            fy_str, q_str = q_col.split('-')

            # -- Find the matching filing ---------------------------------------
            target_filing = filing_by_period.get(q_col)
            if target_filing is None:
                print(f"    [Warning] No filing found for {q_col}.")
                continue

            is_10k_filing = '10-K' in target_filing.form
            
            try:
                rescue_tables = _get_filing_html_rescue_rows_cached(target_filing)
            except Exception as e:
                print(f"    [Warning] Failed to parse HTML for {q_col}: {e}")
                continue

            total_rev = revenue_row.get(q_col)
            if pd.isna(total_rev) or total_rev == 0:
                continue

            for cat in ['4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries']:
                rescued = False
                expected_segments = list(_expected_segments_by_cat.get(cat, ()))
                if not expected_segments:
                    continue
                    
                _seg_vals_q = {}
                for seg in expected_segments:
                    idx_key = (cat, f'Revenue - {seg}')
                    if idx_key in df.index:
                        v = pd.to_numeric(df.at[idx_key, q_col], errors='coerce')
                        if pd.notna(v) and v > 0:
                            _seg_vals_q[seg] = v
                _parent_segs: set[str] = set()
                _seg_items = list(_seg_vals_q.items())
                for i, (s_i, v_i) in enumerate(_seg_items):
                    others = [(s, v) for j, (s, v) in enumerate(_seg_items) if j != i]
                    for r in range(2, min(5, len(others) + 1)):
                        if any(
                            abs(v_i - sum(v for _, v in c)) / max(abs(v_i), 1) < 0.02
                            for c in itertools.combinations(others, r)
                        ):
                            _parent_segs.add(s_i)
                            break
                if _parent_segs:
                    expected_segments = [s for s in expected_segments if s not in _parent_segs]
                expected_lower = {s.split(' (')[0].lower(): s for s in expected_segments}
                if not expected_lower:
                    continue

                ytd9_by_seg: dict[str, float] = {}
                if is_10k_filing:
                    q1_col = f"{fy_str}-Q1"
                    q2_col = f"{fy_str}-Q2"
                    q3_col = f"{fy_str}-Q3"
                    for seg in expected_segments:
                        seg_idx = (cat, f'Revenue - {seg}')
                        if seg_idx not in df.index:
                            continue
                        vals = [
                            pd.to_numeric(df.at[seg_idx, c], errors='coerce')
                            for c in [q1_col, q2_col, q3_col]
                            if c in df.columns
                        ]
                        if all(pd.notna(v) for v in vals) and len(vals) == 3:
                            ytd9_by_seg[seg] = sum(vals)

                for table_rows in rescue_tables:
                    results: dict[str, list[float]] = {}
                    for row_vals in table_rows:
                        matched_seg = None
                        for nl in (row_vals.name0_lower, row_vals.name01_lower):
                            for lower_seg, orig_seg in expected_lower.items():
                                if (
                                    nl == lower_seg
                                    or nl.startswith(lower_seg + ' ')
                                    or nl.endswith(' ' + lower_seg)
                                ):
                                    matched_seg = orig_seg
                                    break
                            if matched_seg:
                                break
                        if matched_seg:
                            if row_vals.nums:
                                results[matched_seg] = row_vals.nums

                    if len(results) < max(2, len(expected_lower) // 2):
                        continue

                    for col_idx in range(min(4, max(len(v) for v in results.values()))):
                        col_vals_raw = {
                            seg: vals[col_idx]
                            for seg, vals in results.items()
                            if len(vals) > col_idx
                        }
                        if not col_vals_raw:
                            continue
                        col_sum_raw = sum(col_vals_raw.values())
                        if col_sum_raw <= 0:
                            continue

                        if is_10k_filing:
                            q1c = f"{fy_str}-Q1"; q2c = f"{fy_str}-Q2"; q3c = f"{fy_str}-Q3"
                            annual_rev = total_rev 
                            for qc in [q1c, q2c, q3c]:
                                if qc in df.columns:
                                    v = revenue_row.get(qc)
                                    if pd.notna(v):
                                        annual_rev += v
                            ref_val = annual_rev
                        else:
                            ref_val = total_rev

                        best_scale: int | None = None
                        for test_scale in [1, 1_000, 1_000_000, 1_000_000_000]:
                            col_sum_test = col_sum_raw * test_scale
                            if abs(ref_val - col_sum_test) / max(ref_val, 1) < 0.05:
                                best_scale = test_scale
                                break
                        if best_scale is None:
                            continue

                        if is_10k_filing:
                            discrete_vals: dict[str, float] = {}
                            all_ytd9_available = True
                            for seg, raw_annual in col_vals_raw.items():
                                annual_seg = raw_annual * best_scale
                                if seg in ytd9_by_seg:
                                    discrete_q4 = annual_seg - ytd9_by_seg[seg]
                                    if discrete_q4 < 0:
                                        all_ytd9_available = False
                                        break
                                    discrete_vals[seg] = discrete_q4
                                else:
                                    all_ytd9_available = False
                                    break

                            if not all_ytd9_available:
                                continue

                            q4_sum = sum(discrete_vals.values())
                            if abs(total_rev - q4_sum) / max(total_rev, 1) > 0.05:
                                continue

                            print(f"    [OK] Rescued {cat} {q_col} from 10-K HTML (annual âˆ’ YTD9 method).")
                            for seg, q4_val in discrete_vals.items():
                                df.at[(cat, f'Revenue - {seg}'), q_col] = q4_val
                            rescued = True
                            break

                        else:
                            print(f"    [OK] Rescued {cat} {q_col} from 10-Q HTML table.")
                            for seg, raw_val in col_vals_raw.items():
                                df.at[(cat, f'Revenue - {seg}'), q_col] = raw_val * best_scale
                            rescued = True
                            break

                    if rescued:
                        break

                if not rescued:
                    if cat == '4a_Segments_Business':
                        print(f"    [Warning] HTML rescue for {cat} {q_col} failed -- keeping XBRL-derived data.")

    except Exception as e:
        print(f"    [Error] HTML rescue failed: {e}")

    return df

def calculate_kpis(pivoted, is_reit=False):
    with _ProfileTimer("calculate_kpis"):
        return _calculate_kpis_impl(pivoted, is_reit=is_reit)


def _calculate_kpis_impl(pivoted, is_reit=False):
    _IBM_STYLE_STATE['active'] = False
    kpi_rows = []
    _row_cache: dict[tuple[str | None, str], pd.Series] = {}
    _num_cache: dict[tuple[str | None, str], pd.Series] = {}
    _nan_row = pd.Series(np.nan, index=pivoted.columns)

    def get_row(name, preferred_cat=None):
        key = (preferred_cat, name)
        cached = _row_cache.get(key)
        if cached is not None:
            return cached
        if preferred_cat and (preferred_cat, name) in pivoted.index:
            result = pivoted.loc[(preferred_cat, name)]
            _row_cache[key] = result
            return result
        for cat in ['1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow']:
            if (cat, name) in pivoted.index:
                result = pivoted.loc[(cat, name)]
                _row_cache[key] = result
                return result
        # Preserve the old behavior for absent rows: each get_row() call got a
        # fresh all-NaN Series.  Several downstream repair paths mutate numeric
        # copies of missing rows, and sharing one cached NaN row can leak state
        # between unrelated missing labels.
        return _nan_row.copy()

    def get_num(name, preferred_cat=None, fillna=None, mutable=False):
        key = (preferred_cat, name)
        base = _num_cache.get(key)
        if base is None:
            base = pd.to_numeric(get_row(name, preferred_cat), errors='coerce')
            _num_cache[key] = base
        result = base.copy() if mutable else base
        if fillna is not None:
            result = result.fillna(fillna)
        return result

    def add_val(cat, label, values):
        for p, v in values.items():
            if pd.notna(v): kpi_rows.append({'Category': cat, 'Label': label, 'Period': p, 'Value': v})
            
# --- NEW: GENERALIZED ANOMALY RESCUE (ULTRA-SAFE) ---
    def needs_anomaly_rescue(raw_series, derived_series, components_present_mask, is_exact_identity=False):
        """
        Safeguard against XBRL leaks with zero false-positives.
        - Exact Identities: Strict bidirectional enforcement (GAAP mathematically requires A = B + C).
        - Partial Sums: Strict unidirectional enforcement (only overwrites if parts > whole, 
          proving the 'whole' is actually just a subcomponent).
        """
        raw_filled = pd.to_numeric(raw_series, errors='coerce').fillna(0)
        derived_filled = pd.to_numeric(derived_series, errors='coerce').fillna(0)
        
        if is_exact_identity:
            # Bidirectional check: The reported number MUST equal the derived math.
            divergence = (derived_filled - raw_filled).abs() > derived_filled.abs().replace(0, 1) * 0.10
            return raw_series.isna() | (components_present_mask & divergence)
        else:
            # Unidirectional check: Only overwrite if we found MORE components than the total.
            understatement = derived_filled > raw_filled * 1.1
            return raw_series.isna() | (components_present_mask & understatement)
    
    rev, cor, gp, ni = get_row('Revenue'), get_row('Cost of Revenue'), get_row('Gross Profit'), get_row('Net Income')
    tax_exp, pretax, op_inc = get_row('Income Tax Expense'), get_row('Pretax Income'), get_row('Operating Income')
    rd, sga, assets, equity = get_row('Research & Development'), get_row('Selling, General & Admin'), get_row('Total Assets'), get_row('Total Equity')
    curr_assets = get_row('Total Current Assets')
    noncurr_assets = get_row('Total Non-Current Assets')

    rev_n, cor_n, gp_n, ni_n = pd.to_numeric(rev, errors='coerce'), pd.to_numeric(cor, errors='coerce'), pd.to_numeric(gp, errors='coerce'), pd.to_numeric(ni, errors='coerce')
    
    rev_parts = get_num('Product Revenue').fillna(0) + \
                get_num('Service Revenue').fillna(0) + \
                get_num('Subscription Revenue').fillna(0) + \
                get_num('Licenses Revenue').fillna(0) + \
                get_num('Lease & Other Revenue').fillna(0)
    rev_n = rev_n.where(rev_n.notna() & (rev_n != 0), rev_parts.replace(0, np.nan))
    
    cor_parts = get_num('Cost of Goods Sold').fillna(0) + \
                get_num('Cost of Services').fillna(0) + \
                get_num('Cost of Lease & Other Revenue').fillna(0)
    cor_n = cor_n.where(cor_n.notna() & (cor_n != 0), cor_parts.replace(0, np.nan))
    tax_n, pretax_n, op_inc_n = pd.to_numeric(tax_exp, errors='coerce'), pd.to_numeric(pretax, errors='coerce'), pd.to_numeric(op_inc, errors='coerce')
    # Snapshot BEFORE any gap-fill: did the filer natively report a separate
    # operating-income line? IBM-style filers (interest folded into a single
    # expense-and-other-income block, only a pretax subtotal) do not.
    _op_inc_was_filed = op_inc_n.notna()
    rd_n, assets_n, equity_n = pd.to_numeric(rd, errors='coerce'), pd.to_numeric(assets, errors='coerce'), pd.to_numeric(equity, errors='coerce')
    rd_parts = get_num('In-Process R&D').fillna(0) + \
               get_num('Related Party R&D').fillna(0)
    rd_n = rd_n.where(rd_n.notna() & (rd_n != 0), rd_parts.replace(0, np.nan))
    curr_assets_n = pd.to_numeric(curr_assets, errors='coerce')
    noncurr_assets_n = pd.to_numeric(noncurr_assets, errors='coerce')
    op_exp_n = get_num('Total Operating Expenses')
    total_costs_exp = get_row('Total Costs and Expenses')
    total_costs_exp_n = pd.to_numeric(total_costs_exp, errors='coerce')
    sga_n = pd.to_numeric(sga, errors='coerce')

    # -- Smart SGA: build effective_sga from components or combined ----
    # If the company files component tags (S&M + G&A), use them.
    # If the company files the combined tag, use that instead.
    # This prevents double-counting for companies that file both.
    sm  = get_row('Sales & Marketing')
    ga  = get_row('General & Administrative')
    sm_n = pd.to_numeric(sm, errors='coerce')
    ga_n = pd.to_numeric(ga, errors='coerce')
    has_sga_components = sm_n.notna() | ga_n.notna()
    effective_sga = pd.Series(0.0, index=pivoted.columns)
    effective_sga = effective_sga.where(~has_sga_components, sm_n.fillna(0) + ga_n.fillna(0))
    effective_sga = effective_sga.where(has_sga_components, sga_n.fillna(0))

    # -- Financing variables ------------------------------------------
    st_debt_issued  = get_row('Short-term Debt Issued')
    lt_debt_issued  = get_row('Long-term Debt Issued')
    debt_issued     = get_row('Total Debt Issued')
    
    st_debt_repaid  = get_row('Short-term Debt Repaid')
    lt_debt_repaid  = get_row('Long-term Debt Repaid')
    debt_repaid     = get_row('Total Debt Repaid')
    
    short_debt_net  = get_row('Net Change in Short-term Debt')

    st_issued_n     = pd.to_numeric(st_debt_issued, errors='coerce').fillna(0)
    lt_issued_n     = pd.to_numeric(lt_debt_issued, errors='coerce').fillna(0)
    st_repaid_n     = pd.to_numeric(st_debt_repaid, errors='coerce').fillna(0)
    lt_repaid_n     = pd.to_numeric(lt_debt_repaid, errors='coerce').fillna(0)

    # Derive missing Total Debt Issued
    derived_issued = st_issued_n + lt_issued_n
    issued_raw_n = pd.to_numeric(debt_issued, errors='coerce')
    issued_n = issued_raw_n.copy()
    components_issued = st_debt_issued.notna() | lt_debt_issued.notna()
    mask_issued = needs_anomaly_rescue(issued_n, derived_issued, components_issued, is_exact_identity=False)
    mask_issued = mask_issued & (derived_issued > 0)
    if mask_issued.any():
        issued_n[mask_issued] = derived_issued[mask_issued]
        add_val('3_Cash_Flow', 'Total Debt Issued', issued_n[mask_issued])
        debt_issued = issued_n # update series for missing_gross downstream
    issued_available = issued_n.notna()
    issued_n = issued_n.fillna(0)

    # Derive missing Total Debt Repaid
    derived_repaid = st_repaid_n + lt_repaid_n
    repaid_raw_n = pd.to_numeric(debt_repaid, errors='coerce')
    repaid_n = repaid_raw_n.copy()
    components_repaid = st_debt_repaid.notna() | lt_debt_repaid.notna()
    mask_repaid = needs_anomaly_rescue(repaid_n, derived_repaid, components_repaid, is_exact_identity=False)
    mask_repaid = mask_repaid & (derived_repaid > 0)
    if mask_repaid.any():
        repaid_n[mask_repaid] = derived_repaid[mask_repaid]
        add_val('3_Cash_Flow', 'Total Debt Repaid', repaid_n[mask_repaid])
        debt_repaid = repaid_n # update series for missing_gross downstream
    repaid_available = repaid_n.notna()
    repaid_n = repaid_n.fillna(0)

    short_net_raw   = pd.to_numeric(short_debt_net, errors='coerce')
    short_net_n     = short_net_raw.fillna(0)

    # -- Calculate Explicit Net Debt Lines ----------------------------
    net_lt_debt = lt_issued_n - lt_repaid_n
    mask_lt = lt_debt_issued.notna() | lt_debt_repaid.notna()
    if mask_lt.any():
        add_val('3_Cash_Flow', 'Net Long-Term Debt Issued (Repaid)', net_lt_debt[mask_lt])

    net_st_debt = st_issued_n - st_repaid_n
    # If the company explicitly files a net short-term change, prioritize it
    if short_net_raw.notna().any():
        net_st_debt = net_st_debt.where(short_net_raw.isna(), short_net_raw)
        
    mask_st = st_debt_issued.notna() | st_debt_repaid.notna() | short_net_raw.notna()
    if mask_st.any():
        add_val('3_Cash_Flow', 'Net Short-Term Debt Issued (Repaid)', net_st_debt[mask_st])
        
    # -- Calculate Total Net Debt -------------------------------------
    # If BOTH gross issued and gross repaid totals are available, the net line
    # must equal issued minus repaid.  If both are absent but the filer directly
    # reported a net-debt row, preserve that direct fact; only fall back to
    # ST/LT net components where no direct/gross evidence exists.
    total_net_reported = get_row('Total Net Debt Issued (Repaid)', preferred_cat='3_Cash_Flow')
    total_net_reported_n = pd.to_numeric(total_net_reported, errors='coerce')
    gross_pair_available = issued_available & repaid_available
    one_sided_gross = (issued_available | repaid_available) & ~gross_pair_available
    net_total_debt = total_net_reported_n.copy()
    net_total_debt[gross_pair_available] = issued_n[gross_pair_available] - repaid_n[gross_pair_available]

    derived_total_net = net_st_debt + net_lt_debt
    mask_use_derived = net_total_debt.isna() & (~gross_pair_available) & (mask_st | mask_lt) & derived_total_net.notna()
    net_total_debt[mask_use_derived] = derived_total_net[mask_use_derived]
    # Last resort for one-sided gross rows only when no direct net fact exists.
    net_total_debt[one_sided_gross & net_total_debt.isna()] = (issued_n - repaid_n)[one_sided_gross & net_total_debt.isna()]

    mask_total = gross_pair_available | one_sided_gross | total_net_reported_n.notna() | mask_st | mask_lt
    if mask_total.any():
        add_val('3_Cash_Flow', 'Total Net Debt Issued (Repaid)', net_total_debt[mask_total])
        
    # -- Calculate Net Equity / Shares --------------------------------
    shares_issued = get_row('Shares Issued')
    share_repur   = get_row('Share Repurchases')
    
    sh_issued_n   = pd.to_numeric(shares_issued, errors='coerce').fillna(0)
    sh_repur_n    = pd.to_numeric(share_repur, errors='coerce').fillna(0)
    
    net_shares    = sh_issued_n - sh_repur_n
    mask_shares   = shares_issued.notna() | share_repur.notna()
    if mask_shares.any():
        add_val('3_Cash_Flow', 'Net Shares Issued (Repurchased)', net_shares[mask_shares])

    # -- Financing Balance: infer missing Gross Debt Issued/Repaid ----
    # If a company only files "Net Change in Short-term Debt" (common
    # for Commercial Paper issuers), split it into Issued vs Repaid
    # so the financing bridge doesn't leave a phantom residual.
    missing_gross = debt_issued.isna() & debt_repaid.isna()
    if missing_gross.any() and short_debt_net.notna().any():
        for p in short_debt_net.index:
            if missing_gross.get(p) and pd.notna(short_debt_net[p]):
                val = short_debt_net[p]
                if val > 0:
                    add_val('3_Cash_Flow', 'Total Debt Issued', {p: val})
                    issued_n[p] = val
                elif val < 0:
                    add_val('3_Cash_Flow', 'Total Debt Repaid', {p: abs(val)})
                    repaid_n[p] = abs(val)

    # -- Fill missing Gross Profit ------------------------------------
    gp_calc = rev_n - cor_n
    gp_to_add = gp_calc[gp_n.isna() & gp_calc.notna()]
    if not gp_to_add.empty:
        add_val('1_Income_Statement', 'Gross Profit', gp_to_add)
        gp_n = gp_n.fillna(gp_to_add)

    # -- Fill missing Pretax Income -----------------------------------
    pretax_calc = ni_n + tax_n.fillna(0)
    pretax_to_add = pretax_calc[pretax_n.isna() & ni_n.notna()]
    if not pretax_to_add.empty:
        add_val('1_Income_Statement', 'Pretax Income', pretax_to_add)
        pretax_n = pretax_n.fillna(pretax_to_add)
    
    # -- Double-Entry Balance Sheet Derivations --
    # Derive missing Total Assets
    derived_assets = curr_assets_n.fillna(0) + noncurr_assets_n.fillna(0)
    assets_mask = assets_n.isna() & (derived_assets > 0)
    if assets_mask.any():
        assets_n[assets_mask] = derived_assets[assets_mask]
        add_val('2_Balance_Sheet', 'Total Assets', assets_n[assets_mask])
        
    # Derive missing Total Liabilities
    curr_liab = get_num('Total Current Liabilities').fillna(0)
    noncurr_liab = get_num('Other Non-Current Liabilities').fillna(0)
    lt_debt = get_num('Long-term Debt').fillna(0)
    op_lease = get_num('Operating Lease Liability (Non-current)').fillna(0)
    fin_lease_nc = get_num('Finance Lease Liability (Non-current)').fillna(0)
    
    derived_liab = curr_liab + noncurr_liab + lt_debt + op_lease + fin_lease_nc
    liab_n = get_num('Total Liabilities', mutable=True)
    liab_mask = liab_n.isna() & (derived_liab > 0)
    if liab_mask.any():
        liab_n[liab_mask] = derived_liab[liab_mask]
        add_val('2_Balance_Sheet', 'Total Liabilities', liab_n[liab_mask])
        
    # Derive missing Total Equity
    equity_mask = equity_n.isna() & assets_n.notna() & liab_n.notna()
    if equity_mask.any():
        equity_n[equity_mask] = assets_n[equity_mask] - liab_n[equity_mask]
        add_val('2_Balance_Sheet', 'Total Equity', equity_n[equity_mask])

    # Fallback: Total Liabilities = Assets âˆ’ Equity (when component sum failed)
    liab_from_ae = assets_n - equity_n
    liab_still_missing = liab_n.isna() & liab_from_ae.notna()
    if liab_still_missing.any():
        liab_n[liab_still_missing] = liab_from_ae[liab_still_missing]
        add_val('2_Balance_Sheet', 'Total Liabilities', liab_from_ae[liab_still_missing])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # TOTAL COSTS VS TRUE OPERATING EXPENSES
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Some filers report us-gaap:CostsAndExpenses / OperatingCostsAndExpenses,
    # which equals Revenue - Operating Income and therefore includes Cost of
    # Revenue.  Keep that filed subtotal as "Total Costs and Expenses" and use
    # GP - Operating Income (or Total Costs - Cost of Revenue) for true opex.
    # This fixes META-style statements without changing revenue/opinc math.
    if rev_n.notna().any() and op_inc_n.notna().any():
        _total_costs_expected = rev_n - op_inc_n
        _total_costs_tol = (_total_costs_expected.abs() * 0.01).clip(lower=50e6)
        _true_opex_from_gp = gp_n - op_inc_n
        _op_exp_as_total_costs = (
            op_exp_n.notna() & _total_costs_expected.notna()
            & ((_op_exp_as_tc_err := (op_exp_n - _total_costs_expected).abs()) <= _total_costs_tol)
            & cor_n.notna() & (cor_n.abs() > 0)
            & _true_opex_from_gp.notna() & (_true_opex_from_gp >= 0)
            & ((op_exp_n - _true_opex_from_gp).abs() > (_total_costs_tol * 0.5))
        )
        if _op_exp_as_total_costs.any():
            add_val('1_Income_Statement', 'Total Costs and Expenses', op_exp_n[_op_exp_as_total_costs])
            op_exp_n[_op_exp_as_total_costs] = _true_opex_from_gp[_op_exp_as_total_costs]
            add_val('1_Income_Statement', 'Total Operating Expenses', op_exp_n[_op_exp_as_total_costs])
            print("  [OpEx Correction] COGS-inclusive total-cost subtotal relabeled as "
                  "Total Costs and Expenses; true Total Operating Expenses derived from GP - Operating Income.")

    if total_costs_exp_n.notna().any():
        _opex_from_total_costs = total_costs_exp_n - cor_n.fillna(0)
        _from_tc_mask = (op_exp_n.isna() & total_costs_exp_n.notna()
                         & cor_n.notna() & _opex_from_total_costs.notna()
                         & (_opex_from_total_costs >= 0))
        if _from_tc_mask.any():
            op_exp_n[_from_tc_mask] = _opex_from_total_costs[_from_tc_mask]
            add_val('1_Income_Statement', 'Total Operating Expenses', op_exp_n[_from_tc_mask])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # TOTAL OPERATING EXPENSES: COGS-INCLUSIVE TAG CORRECTION
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # XBRL tags OperatingCostsAndExpenses (TagRank 1) and
    # CostsAndExpenses (TagRank 2) include Cost of Revenue for some
    # filers (Amazon, Costco, etc.).  When the preferred tag
    # OperatingExpenses (TagRank 0) is absent, the tag-locking engine
    # falls back to these COGS-inclusive tags, causing Total Operating
    # Expenses to double-count COGS relative to Gross Profit.
    #
    # Detection (3 independent signals -- any 1 is sufficient):
    #   A) TotalOpEx > GP in >50% of comparable periods
    #   B) TotalOpEx â‰ˆ Revenue (within 25%) -- it's clearly total costs
    #   C) OpInc cross-check: (TotalOpEx âˆ’ COGS) â‰ˆ (GP âˆ’ OpInc)
    #
    # Correction: TotalOpEx_corrected = TotalOpEx âˆ’ Cost of Revenue
    #
    # Safety: Normal filers (MSFT, GOOG) have TotalOpEx < GP -> no change
    # ------------------------------------------------------------------
    _opex_corrected = False
    if cor_n.notna().any() and op_exp_n.notna().any() and gp_n.notna().any():
        compare_mask = op_exp_n.notna() & gp_n.notna() & (gp_n.abs() > 0)
        if compare_mask.sum() >= 2:
            # Signal A: TotalOpEx exceeds Gross Profit
            exceeds_gp = op_exp_n[compare_mask] > gp_n[compare_mask] * 1.05
            signal_a = exceeds_gp.sum() > 0.5 * compare_mask.sum()

            # Signal B: TotalOpEx â‰ˆ Revenue (within 25%)
            rev_compare = op_exp_n.notna() & rev_n.notna() & (rev_n.abs() > 0)
            if rev_compare.any():
                opex_to_rev = op_exp_n[rev_compare] / rev_n[rev_compare]
                signal_b = (opex_to_rev > 0.75).sum() > 0.5 * rev_compare.sum()
            else:
                signal_b = False

            if signal_a or signal_b:
                corrected_opex = op_exp_n - cor_n.fillna(0)

                # Validate: corrected should be positive and < GP
                basic_valid = (corrected_opex > 0) & corrected_opex.notna()
                if basic_valid.sum() >= 2:
                    # Signal C: Cross-check against Operating Income if available
                    if op_inc_n.notna().any():
                        implied_opex = gp_n - op_inc_n
                        cross_mask = corrected_opex.notna() & implied_opex.notna() & (implied_opex.abs() > 0)
                        if cross_mask.sum() >= 2:
                            # Corrected should be much closer to implied than original
                            orig_err  = (op_exp_n[cross_mask] - implied_opex[cross_mask]).abs()
                            corr_err  = (corrected_opex[cross_mask] - implied_opex[cross_mask]).abs()
                            improvement = (corr_err < orig_err)
                            if improvement.sum() < 0.5 * cross_mask.sum():
                                # Correction made things worse -- abort
                                corrected_opex = None

                    if corrected_opex is not None:
                        op_exp_n = corrected_opex
                        add_val('1_Income_Statement', 'Total Operating Expenses',
                                op_exp_n[op_exp_n.notna()])
                        _opex_corrected = True
                        print("  [OpEx Correction] Total Operating Expenses included COGS "
                              "(CostsAndExpenses/OperatingCostsAndExpenses tag) "
                              "-- corrected by subtracting Cost of Revenue")

    # Granular Expense Summation (Institutional Aggregation)
    fulfillment = get_num('Fulfillment').fillna(0)
    marketing = get_num('Marketing').fillna(0)
    ga = get_num('General & Administrative').fillna(0)
    amort = get_num('Amortization of Intangibles').fillna(0)
    
    derived_opex = rd_n.fillna(0) + fulfillment + marketing + ga + amort
    components_opex = rd_n.notna() | get_row('Fulfillment').notna() | get_row('Marketing').notna() | get_row('General & Administrative').notna() | get_row('Amortization of Intangibles').notna()

    if is_reit:
        _reit_da   = get_num('Depreciation & Amortization').fillna(0)
        _reit_int  = get_num('Interest Expense').fillna(0)
        _reit_prop = get_num('Direct Costs of Leased and Rented Property or Equipment').fillna(0)
        _reit_imp  = get_num('Impairment Charges').fillna(0)
        _reit_acq  = get_num('Acquisition-Related Costs').fillna(0)
        _reit_rest = get_num('Restructuring & Related Charges').fillna(0)
        derived_opex = derived_opex + _reit_da + _reit_int + _reit_prop + _reit_imp + _reit_acq + _reit_rest
        components_opex = (components_opex | get_row('Depreciation & Amortization').notna()
                           | get_row('Interest Expense').notna()
                           | get_row('Direct Costs of Leased and Rented Property or Equipment').notna())

    # Priority 1: Reported Total. Priority 2: Sum of components.
    mask = needs_anomaly_rescue(op_exp_n, derived_opex, components_opex, is_exact_identity=False)
    mask = mask & (derived_opex > 0)
    if mask.any():
        # Only inject if derived is materially non-zero.
        # Forcing 0.0 breaks downstream accounting engine fallbacks.
        op_exp_n[mask] = derived_opex[mask]
        add_val('1_Income_Statement', 'Total Operating Expenses', op_exp_n[mask])

    # North Star Principle: Always prioritize reported Operating Income unless NaN.
    if gp_n.isna().all() and rev_n.notna().any(): gp_n = rev_n - cor_n.fillna(0)

    # -- Smart Operating Income gap fill ------------------------------
    # Safety net: if the COGS correction above didn't fire (e.g. too few
    # periods), this catches remaining cases where op_exp_n still includes
    # COGS and uses the component bridge instead.
    _n_opex_q = int(op_exp_n.notna().sum())
    opex_still_includes_cogs = (not _opex_corrected
                                and _n_opex_q > 0
                                and gp_n.notna().any()
                                and int((op_exp_n > gp_n * 1.05).sum()) > 0.5 * _n_opex_q)
    if opex_still_includes_cogs:
        opinc_calc = gp_n - (rd_n.fillna(0) + effective_sga)
    else:
        opinc_calc = gp_n - op_exp_n
    opinc_to_add = opinc_calc[op_inc_n.isna() & opinc_calc.notna()]
    if not opinc_to_add.empty:
        add_val('1_Income_Statement', 'Operating Income', opinc_to_add)
        op_inc_n = op_inc_n.fillna(opinc_to_add)

    # -- Total Operating Expenses Method B: GP âˆ’ OpInc ----------------
    opexp_calc_b = gp_n - op_inc_n
    opexp_to_add_b = opexp_calc_b[op_exp_n.isna() & opexp_calc_b.notna()]
    if not opexp_to_add_b.empty:
        add_val('1_Income_Statement', 'Total Operating Expenses', opexp_to_add_b)
        op_exp_n = op_exp_n.fillna(opexp_to_add_b)
    
    goodwill = get_row('Goodwill'); cash = get_row('Cash & Equivalents')
    goodwill_n, cash_n = pd.to_numeric(goodwill, errors='coerce'), pd.to_numeric(cash, errors='coerce')
    suspicious = ((assets_n == goodwill_n) | (assets_n == cash_n)) & (assets_n > 0)
    if suspicious.any():
        noncurr = get_row('Total Non-Current Assets')
        derived = pd.to_numeric(curr_assets, errors='coerce').fillna(0) + pd.to_numeric(noncurr, errors='coerce').fillna(0)
        assets_n[suspicious] = derived[suspicious]; add_val('2_Balance_Sheet', 'Total Assets', assets_n[suspicious])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ROBUST TOTAL ASSETS VALIDATION  (structural impossibility guard)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # A correctly-extracted Total Assets is the sum of the asset side, so
    # it can NEVER be smaller than any individual asset line -- this holds
    # for every company and every capital structure (negative equity
    # included).  A mis-mapped XBRL context (a consolidated-VIE, parent-
    # only, or segment `us-gaap:Assets` fact, common in complex post-
    # merger filings) can return a value an order of magnitude too small.
    # We detect that purely structurally, then rebuild the figure from the
    # accounting identity A = L + E using the *reported* Liabilities and
    # Equity (read from `pivoted`, never a value we derived, so the rebuild
    # cannot be circular).  Two safeguards make this regression-proof:
    #   1. The trigger is a logical impossibility, so a correct Total
    #      Assets is never flagged on its own merits.
    #   2. We only overwrite when the rebuilt figure *materially* differs
    #      from the reported one.  So when an asset SUBLINE is the thing
    #      that is overstated (annual real-estate-gross quirks, etc.), the
    #      rebuild equals the already-correct total and we leave it alone.
    # Any period we genuinely override is marked for transparency, and a
    # separate closure metric reports A == L + E for every period.
    _ta_reported = assets_n.copy()
    _ta_liab = get_num('Total Liabilities')
    _ta_eq   = get_num('Total Equity')
    _ta_ident = _ta_liab + _ta_eq                       # GAAP identity (any sign)
    _ta_cn = get_num('Total Current Assets').fillna(0) \
           + get_num('Total Non-Current Assets').fillna(0)
    # largest single asset line present each period (subtotals included);
    # Total Assets must be >= this for every filer.
    _ta_floor_labels = ('Total Current Assets', 'Total Non-Current Assets', 'Real Estate, Net',
        'Property, Plant & Equipment', 'Cash & Equivalents', 'Long-term Investments',
        'Short-term Investments', 'Goodwill', 'Intangible Assets (Net)',
        'Intangible Assets & Goodwill', 'Loans & Leases (Net)', 'Financial Instruments Owned',
        'Operating Lease ROU Asset', 'Accounts Receivable', 'Inventory',
        'Other Non-Current Assets', 'Securities Purchased under Agreements to Resell',
        'Securities Borrowed', 'Receivables from Brokers & Dealers', 'Customer Receivables',
        'Margin Loans')
    _ta_cols = [get_num(_l) for _l in _ta_floor_labels]
    _ta_floor = pd.concat(_ta_cols, axis=1).max(axis=1) if _ta_cols else pd.Series(np.nan, index=assets_n.index)
    # Trigger: reported Total Assets is below an individual asset line.
    _ta_bad = assets_n.notna() & _ta_floor.notna() & (assets_n < 0.999 * _ta_floor)
    if _ta_bad.any():
        # Rebuild from the reported identity whenever both L and E are
        # reported and their sum is a positive total (so negative-equity
        # filers are covered too); fall back to the current+non-current
        # sum (floor-checked) only where the identity is unavailable.
        _ta_recon = pd.Series(np.nan, index=assets_n.index)
        _use_id = _ta_liab.notna() & _ta_eq.notna() & (_ta_ident > 0)
        _ta_recon[_use_id] = _ta_ident[_use_id]
        _use_cn = (_ta_recon.isna() & (_ta_cn > 0)
                   & (_ta_floor.isna() | (_ta_cn >= 0.999 * _ta_floor)))
        _ta_recon[_use_cn] = _ta_cn[_use_cn]
        # Overwrite only on a material change vs the reported value, so a
        # merely-overstated subline never triggers a no-op "fix".
        _ta_otol = (_ta_reported.abs() * 0.02).clip(lower=50e6)
        _ta_fix = (_ta_bad & _ta_recon.notna()
                   & (_ta_reported.isna() | ((_ta_recon - _ta_reported).abs() > _ta_otol)))
        if _ta_fix.any():
            assets_n[_ta_fix] = _ta_recon[_ta_fix]
            add_val('2_Balance_Sheet', 'Total Assets', assets_n[_ta_fix])
            _ta_marker = pd.Series(np.nan, index=assets_n.index)
            _ta_marker[_ta_fix] = 1.0
            add_val('8_Integrity_Checks', 'Metric: Total Assets Reconstructed', _ta_marker[_ta_fix])
    # Balance-sheet closure transparency (additive; independent of the
    # footing metric).  1.0 = A == L + E within tolerance (2% or $50M),
    # 0.0 = does not reconcile -- e.g. a wrong Total Assets we could not
    # rebuild because reported L/E were unavailable.
    _ta_resid = (assets_n - _ta_ident).abs()
    _ta_ctol = (assets_n.abs() * 0.02).clip(lower=50e6)
    _ta_have = assets_n.notna() & _ta_ident.notna()
    _ta_closure = pd.Series(np.nan, index=assets_n.index)
    _ta_closure[_ta_have] = 1.0
    _ta_closure[_ta_have & (_ta_resid > _ta_ctol)] = 0.0
    add_val('8_Integrity_Checks', 'Metric: Balance Sheet Closure Verified', _ta_closure[_ta_have])

    add_val('5_KPI_Metrics', 'Metric: Gross Margin %', (gp_n / rev_n) * 100)
    add_val('5_KPI_Metrics', 'Metric: Net Margin %', (ni_n / rev_n) * 100)
    add_val('5_KPI_Metrics', 'Metric: EBIT Margin %', (op_inc_n / rev_n) * 100)
    da_n = _effective_depreciation_amortization_series(pivoted).reindex(pivoted.columns)
    ebitda = (op_inc_n + da_n).where(op_inc_n.notna() & da_n.notna())
    add_val('5_KPI_Metrics', 'Metric: EBITDA', ebitda)
    ebitda_margin = ((ebitda / rev_n) * 100).where(rev_n.notna() & (rev_n != 0))
    add_val('5_KPI_Metrics', 'Metric: EBITDA Margin %', ebitda_margin)
    
    ocf = get_row('Operating Cash Flow'); capex = get_row('Capital Expenditures')
    ocf_n, capex_n = pd.to_numeric(ocf, errors='coerce'), pd.to_numeric(capex, errors='coerce')
    # Some filers break capex into detail rows (software/intangibles/equipment)
    # without filing the generic Capital Expenditures row.  FCF should use the
    # effective capex total, and the main CF row should be filled only where the
    # aggregate is blank.  This is additive and never double-counts a filed total.
    capex_detail_labels = [
        'Capital Expenditures (Software)',
        'Capital Expenditures (Intangibles)',
        'Capital Expenditures (Equipment & Buildings)',
    ]
    capex_detail_sum = pd.Series(0.0, index=pivoted.columns)
    capex_detail_mask = pd.Series(False, index=pivoted.columns)
    for _capex_lbl in capex_detail_labels:
        _row = get_row(_capex_lbl)
        _num = pd.to_numeric(_row, errors='coerce')
        capex_detail_sum = capex_detail_sum + _num.fillna(0)
        capex_detail_mask = capex_detail_mask | _num.notna()
    capex_effective_n = capex_n.where(capex_n.notna(), capex_detail_sum.where(capex_detail_mask))
    _capex_fill_mask = capex_n.isna() & capex_effective_n.notna()
    if _capex_fill_mask.any():
        add_val('3_Cash_Flow', 'Capital Expenditures', capex_effective_n[_capex_fill_mask])
    fcf = ocf_n - capex_effective_n.fillna(0)
    add_val('5_KPI_Metrics', 'Metric: Free Cash Flow', fcf)
    add_val('5_KPI_Metrics', 'Metric: FCF Margin %', (fcf / rev_n) * 100)
    
    st_debt  = get_num('Short-term Debt').fillna(0)
    st_borr  = get_num('Short-term Borrowings').fillna(0)
    cp_ltd   = get_num('Current Portion of Long-term Debt').fillna(0)
    lt_debt_total = get_num('Long-term Debt').fillna(0)
    lt_debt_parts = get_num('Senior Notes').fillna(0) + \
                    get_num('Convertible Debt').fillna(0) + \
                    get_num('Other Long-term Borrowings').fillna(0)
    lt_debt = lt_debt_total.where(lt_debt_total != 0, lt_debt_parts)
    op_lc    = get_num('Operating Lease Liability (Current)').fillna(0)
    op_lnc   = get_num('Operating Lease Liability (Non-current)').fillna(0)
    fin_lc   = get_num('Finance Lease Liability (Current)').fillna(0)
    fin_lnc  = get_num('Finance Lease Liability (Non-current)').fillna(0)
    
    total_st_debt = st_debt.where(st_debt != 0, st_borr + cp_ltd)
    total_debt = total_st_debt + lt_debt + op_lc + op_lnc + fin_lc + fin_lnc
    add_val('5_KPI_Metrics', 'Metric: Total Debt', total_debt)

    # -- Lease Liabilities (ASC 842) ----------------------------------
    lease_curr = get_row('Operating Lease Liability (Current)')
    lease_noncurr = get_row('Operating Lease Liability (Non-current)')
    fin_lease_curr = get_row('Finance Lease Liability (Current)')
    fin_lease_noncurr = get_row('Finance Lease Liability (Non-current)')
    lc_n = pd.to_numeric(lease_curr, errors='coerce').fillna(0)
    lnc_n = pd.to_numeric(lease_noncurr, errors='coerce').fillna(0)
    flc_n = pd.to_numeric(fin_lease_curr, errors='coerce').fillna(0)
    flnc_n = pd.to_numeric(fin_lease_noncurr, errors='coerce').fillna(0)
    total_lease = lc_n + lnc_n + flc_n + flnc_n
    mask_lease = lease_curr.notna() | lease_noncurr.notna() | fin_lease_curr.notna() | fin_lease_noncurr.notna()
    if mask_lease.any():
        add_val('5_KPI_Metrics', 'Metric: Total Lease Liabilities', total_lease[mask_lease])

    # -- Net Cash (Debt) ----------------------------------------------
    cash_equiv = get_row('Cash & Equivalents')
    short_invest = get_row('Short-term Investments')
    ce_n = pd.to_numeric(cash_equiv, errors='coerce').fillna(0)
    si_n = pd.to_numeric(short_invest, errors='coerce').fillna(0)
    total_cash = ce_n + si_n
    total_cash[cash_equiv.isna() & short_invest.isna()] = np.nan
    if not total_cash.isna().all() or not total_debt.isna().all():
        tc_fill = total_cash.fillna(0)
        td_fill = total_debt.fillna(0)
        net_cash = tc_fill - td_fill
        net_cash[total_cash.isna() & total_debt.isna()] = np.nan
        add_val('5_KPI_Metrics', 'Metric: Net Cash (Debt)', net_cash)

    # -- ROE % (Annualised) -------------------------------------------
    if equity_n.notna().any():
        roe_raw = (ni_n / equity_n) * 100
        roe_ann = roe_raw.copy()
        for col in roe_raw.index:
            q_label = str(col).split('-')[-1] if '-' in str(col) else ''
            if q_label in ('Q1', 'Q2', 'Q3'):
                roe_ann[col] = roe_raw[col] * 4
        add_val('5_KPI_Metrics', 'Metric: ROE % (Annualised)', roe_ann)

    # -- Effective Tax Rate % -----------------------------------------
    if not tax_exp.isna().all() and not pretax.isna().all():
        add_val('5_KPI_Metrics', 'Metric: Effective Tax Rate %', (tax_n / pretax_n) * 100)

    # -- Unlevered Free Cash Flow (Quarterly) -------------------------
    # Formula: Levered FCF + Interest Expense Ã— (1 âˆ’ Effective Tax Rate)
    fcf_q      = ocf_n - capex_effective_n.fillna(0)   # quarterly levered FCF
    int_exp_q  = get_num('Interest Expense').fillna(0)
    tax_q      = pd.to_numeric(tax_exp, errors='coerce').fillna(0)
    pretax_q   = pd.to_numeric(pretax, errors='coerce')

    # Calculate discrete quarterly tax rate
    eff_tax_rate_q = tax_q / pretax_q
    
    # Safety Guards: Quarterly tax rates can be highly volatile. 
    # 1. If pretax income is <= 0 (a loss), assume a 0% tax shield.
    eff_tax_rate_q = eff_tax_rate_q.where(pretax_q > 0, 0.0)
    # 2. Cap the tax shield between 0% and 100% to prevent extreme outliers.
    eff_tax_rate_q = eff_tax_rate_q.clip(0, 1)

    # Vectorized UFCF calculation
    ufcf_series = fcf_q + (int_exp_q * (1 - eff_tax_rate_q))

    if ufcf_series.notna().any():
        add_val('5_KPI_Metrics', 'Metric: Unlevered Free Cash Flow', ufcf_series[ufcf_series.notna()])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # SHARES OUTSTANDING & EPS -- Thorough Derivation
    # Step 1: Derive missing shares from NI / EPS (with backfill)
    # Step 2: Derive missing EPS   from NI / shares
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    # -- Basic Shares -------------------------------------------------
    eps_b_orig = get_num('EPS Basic').replace(0, np.nan)
    shares_b   = get_row('Shares Outstanding Basic')
    shares_b_n = pd.to_numeric(shares_b, errors='coerce').replace(0, np.nan)

    # -- EPS numerator calibration --------------------------------------
    # For umbrella structures (e.g. IBKR, where the public company owns a
    # minority of the group), the consolidated 'Net Income' row includes
    # earnings attributable to noncontrolling interests, so NI / shares
    # massively overstates derived EPS.  Calibrate against quarters where
    # BOTH a filed EPS and a filed share count exist: whichever numerator
    # (consolidated vs ex-NCI) reproduces the filed EPS wins.
    ni_for_eps = ni_n
    _nci_kpi = get_num('Net Income to Noncontrolling Interest').fillna(0)
    if (_nci_kpi != 0).any():
        _ni_ex_nci = ni_n - _nci_kpi
        _calib = eps_b_orig.notna() & shares_b_n.notna() & ni_n.notna()
        # Calibrate on the 12 most recent qualifying quarters: old eras mix
        # stale share counts / pre-split figures and dilute the signal.
        _cq = sorted([str(c) for c in _calib.index[_calib]], reverse=True)[:12]
        _calib = _calib & _calib.index.isin(_cq)
        if _calib.any():
            _implied = (eps_b_orig * shares_b_n)[_calib]
            _denom = _implied.abs().clip(lower=1)
            _err_incl = ((ni_n[_calib] - _implied).abs() / _denom).median()
            _err_ex   = ((_ni_ex_nci[_calib] - _implied).abs() / _denom).median()
            if pd.notna(_err_ex) and pd.notna(_err_incl) and _err_ex + 0.02 < _err_incl:
                ni_for_eps = _ni_ex_nci
                print(f"  [EPS Basis] Filed EPS implies NI excluding noncontrolling "
                      f"interests (err {_err_ex:.1%} vs {_err_incl:.1%}); "
                      f"derived EPS/shares will use NI - NCI.")

    # Derive shares from NI / EPS where shares are missing
    shares_b_calc = ni_for_eps / eps_b_orig
    shares_b_n = shares_b_n.fillna(shares_b_calc)
    # Forward-fill (bfill on columns sorted newest->oldest) to cover missing Q4 shares
    shares_b_n = shares_b_n.bfill()

    shares_b_to_add = shares_b_n[shares_b.isna() | (pd.to_numeric(shares_b, errors='coerce') == 0)]
    if not shares_b_to_add.empty:
        add_val('1_Income_Statement', 'Shares Outstanding Basic', shares_b_to_add)

    # -- Diluted Shares -----------------------------------------------
    eps_d_orig = get_num('EPS Diluted').replace(0, np.nan)
    shares_d   = get_row('Shares Outstanding Diluted')
    shares_d_n = pd.to_numeric(shares_d, errors='coerce').replace(0, np.nan)

    shares_d_calc = ni_for_eps / eps_d_orig
    shares_d_n = shares_d_n.fillna(shares_d_calc)
    shares_d_n = shares_d_n.bfill()

    shares_d_to_add = shares_d_n[shares_d.isna() | (pd.to_numeric(shares_d, errors='coerce') == 0)]
    if not shares_d_to_add.empty:
        add_val('1_Income_Statement', 'Shares Outstanding Diluted', shares_d_to_add)

    # -- EPS from NI / Shares (fills missing EPS using derived shares) -
    eps_b_calc = ni_for_eps / shares_b_n
    eps_b_to_add = eps_b_calc[eps_b_orig.isna() & eps_b_calc.notna()]
    if not eps_b_to_add.empty:
        add_val('1_Income_Statement', 'EPS Basic', eps_b_to_add)

    eps_d_calc = ni_for_eps / shares_d_n
    eps_d_to_add = eps_d_calc[eps_d_orig.isna() & eps_d_calc.notna()]
    if not eps_d_to_add.empty:
        add_val('1_Income_Statement', 'EPS Diluted', eps_d_to_add)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ADDITIONAL VISIBILITY ROWS
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    # -- Cash Flow Reconciliation (ASU 2016-18) -----------------------
    # Ending Cash on the CF statement must include Restricted Cash.
    cash_recon = get_row('Cash Reconciliation: Total')
    if cash_recon.notna().any():
        add_val('3_Cash_Flow', 'Cash & Equivalents (incl. Restricted)', cash_recon)

    # -- Net Income to Noncontrolling Interest (Parent vs Entity) -----
    nci = get_row('Net Income to Noncontrolling Interest')
    if nci.notna().any():
        add_val('1_Income_Statement', 'Net Income to Noncontrolling Interest', nci)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # CASH FLOW FOOTING -- Granular Bridges
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    # -- Operating CF bridge: NI + non-cash + working capital changes -
    op_components = [
        'Net Income (CF)', 'Depreciation & Amortization', 'Amortization of Intangibles (CF)',
        'Stock-Based Compensation', 'Deferred Income Taxes',
        'Gain/Loss on Investments (CF)', 'Other Non-Cash Items',
        'Change in AR', 'Change in Inventory', 'Change in AP',
        'Change in Deferred Revenue', 'Change in Accrued Expenses',
        'Change in Income Taxes', 'Change in Prepaid & Other Assets',
        'Change in Other Liabilities'
    ]
    asset_keywords = ['AR', 'Inventory', 'Prepaid', 'Asset']
    ocf_total = get_row('Operating Cash Flow', preferred_cat='3_Cash_Flow')
    ocf_residual = pd.Series(0.0, index=pivoted.columns)
    if not ocf_total.isna().all():
        ocf_sum = pd.Series(0.0, index=pivoted.columns)
        _ocf_contribs = {}
        for lbl in op_components:
            val = get_num(lbl, preferred_cat='3_Cash_Flow').fillna(0)
            
            _w = None
            _info = CONCEPT_MAP.get(lbl)
            if isinstance(_info, dict) and _info.get('tags'):
                _kind, _calc_w = _classify_calc_lineage(
                    _info['tags'], set(),
                    targets=_CF_OPERATING_PARENTS, blockers=_CF_NONOPERATING_PARENTS)
                if _kind == 'target' and _calc_w:
                    _w = _calc_w
            
            if _w is None:
                is_asset = any(kw in lbl for kw in asset_keywords) and 'Change in' in lbl
                is_gain = 'gain' in lbl.lower()
                _w = -1.0 if (is_asset or is_gain) else 1.0
                
            ocf_sum += val * _w
            _ocf_contribs[lbl] = val * _w
        # Learned operating-CF lines (e.g. broker working-capital moves,
        # 'Amortization of client incentives') join the bridge with their
        # filed calculation weight, shrinking the residual to true noise.
        _cf_counted = set()
        for _l in op_components:
            _info = CONCEPT_MAP.get(_l)
            if isinstance(_info, dict):
                _cf_counted.update(_info.get('tags') or [])
        _cf_counted = frozenset(_cf_counted)
        _cf_claimed = set()
        _dyn_cf = []
        for _cat0, _lbl in pivoted.index:
            if _cat0 != '3_Cash_Flow' or _lbl in op_components or _lbl in _cf_claimed:
                continue
            if ':' in _lbl or ' - ' in _lbl or '(Net)' in _lbl:
                continue
            _info = CONCEPT_MAP.get(_lbl)
            if not (isinstance(_info, dict) and _info.get('auto')):
                continue
            _kind, _w = _classify_calc_lineage(
                _info.get('tags'), _cf_counted,
                targets=_CF_OPERATING_PARENTS, blockers=_CF_NONOPERATING_PARENTS)
            if _kind is None or not _w:
                _fw = _classify_cf_label_fallback(_lbl, 'op')
                if _fw is not None:
                    _kind, _w = 'target', _fw
            if _kind == 'target' and _w:
                _val = get_num(_lbl, preferred_cat='3_Cash_Flow')
                if _val.notna().any():
                    ocf_sum += _val.fillna(0) * _w
                    _ocf_contribs[_lbl] = _val.fillna(0) * _w
                    _cf_claimed.add(_lbl)
                    _dyn_cf.append(_lbl)
        if _dyn_cf:
            print(f"  [Bridge] Operating-CF components via calc linkbase: {', '.join(_dyn_cf)}")

        ocf_sum = _bridge_sign_self_check(
            pd.to_numeric(ocf_total, errors='coerce'), ocf_sum, _ocf_contribs, 'operating-CF')
        ocf_residual = pd.to_numeric(ocf_total, errors='coerce').fillna(0) - ocf_sum
        if ocf_total.notna().any() and (ocf_residual.abs() > 1e6).any():
            add_val('3_Cash_Flow', 'Other Operating Adjustments (Net)', ocf_residual[ocf_total.notna()])
            _ocf_ratio = (ocf_residual.abs()
                          / pd.to_numeric(ocf_total, errors='coerce').abs().clip(lower=1.0)).median()
            if pd.notna(_ocf_ratio) and _ocf_ratio > 0.10:
                # Supplemental cash-paid disclosures (cash paid for taxes /
                # interest) sit below the OCF subtotal and are NOT reconciling
                # operating components -- exclude them so the warning points at
                # genuine gaps only.
                def _is_supplemental_cashpaid(_l):
                    _i = CONCEPT_MAP.get(_l)
                    _tags = (_i.get('tags') or []) if isinstance(_i, dict) else []
                    return any(re.search(r'(incometaxespaid|interestpaid|'
                                         r'interestpaidnet|incometaxespaidnet)',
                                         str(_t).lower()) for _t in _tags)
                _unwired = [l for (c0, l) in pivoted.index
                            if c0 == '3_Cash_Flow' and l not in op_components
                            and l not in _dyn_cf
                            and isinstance(CONCEPT_MAP.get(l), dict)
                            and CONCEPT_MAP.get(l, {}).get('auto')
                            and not _is_supplemental_cashpaid(l)]
                print(f"  [Bridge Warning] 'Other Operating Adjustments (Net)' is "
                      f"{_ocf_ratio:.0%} of OCF (median) -- a face CF line is likely "
                      f"uncaptured or unclassified. Unwired learned CF rows: "
                      f"{', '.join(_unwired) if _unwired else 'none'}")

    # -- Investing CF bridge ------------------------------------------
    inv_total = get_row('Investing Cash Flow', preferred_cat='3_Cash_Flow')
    inv_residual = pd.Series(0.0, index=pivoted.columns)
    if inv_total.notna().any():
        inv_sum = pd.Series(0.0, index=pivoted.columns)
        _inv_contribs = {}
        
        inv_components = [
            'Capital Expenditures', 'Acquisitions', 'Purchases of Investments',
            'Proceeds from Investments', 'Proceeds from Asset Sales', 'Divestitures'
        ]
        
        for _lbl0, _w0 in (('Capital Expenditures', -1.0), ('Acquisitions', -1.0),
                           ('Purchases of Investments', -1.0), ('Proceeds from Investments', 1.0),
                           ('Proceeds from Asset Sales', 1.0), ('Divestitures', 1.0)):
            _v0 = get_num(_lbl0).fillna(0)
            inv_sum += _v0 * _w0
            _inv_contribs[_lbl0] = _v0 * _w0
        
        _inv_counted = set()
        for _l in inv_components:
            _info = CONCEPT_MAP.get(_l)
            if isinstance(_info, dict): _inv_counted.update(_info.get('tags') or [])
        _inv_counted = frozenset(_inv_counted)
        _dyn_inv = []
        for _cat0, _lbl in pivoted.index:
            if _cat0 != '3_Cash_Flow' or _lbl in inv_components or _lbl in _cf_claimed: continue
            if ':' in _lbl or ' - ' in _lbl or '(Net)' in _lbl: continue
            _info = CONCEPT_MAP.get(_lbl)
            if not (isinstance(_info, dict) and _info.get('auto')): continue
            _kind, _w = _classify_calc_lineage(_info.get('tags'), _inv_counted, targets=_CF_INVESTING_PARENTS, blockers=_CF_OPERATING_PARENTS | _CF_FINANCING_PARENTS)
            if _kind is None or not _w:
                _fw = _classify_cf_label_fallback(_lbl, 'inv')
                if _fw is not None:
                    _kind, _w = 'target', _fw
            if _kind == 'target' and _w:
                _val = get_num(_lbl, preferred_cat='3_Cash_Flow')
                if _val.notna().any():
                    inv_sum += _val.fillna(0) * _w
                    _inv_contribs[_lbl] = _val.fillna(0) * _w
                    _cf_claimed.add(_lbl)
                    _dyn_inv.append(_lbl)
        if _dyn_inv: print(f"  [Bridge] Investing-CF components via calc linkbase: {', '.join(_dyn_inv)}")

        inv_sum = _bridge_sign_self_check(
            pd.to_numeric(inv_total, errors='coerce'), inv_sum, _inv_contribs, 'investing-CF')
        inv_residual = pd.to_numeric(inv_total, errors='coerce').fillna(0) - inv_sum
        if (inv_residual.abs() > 1e6).any():
            add_val('3_Cash_Flow', 'Other Investing Adjustments (Net)', inv_residual[inv_total.notna()])

    # -- Financing CF bridge ------------------------------------------
    fin_total = get_row('Financing Cash Flow', preferred_cat='3_Cash_Flow')
    fin_residual = pd.Series(0.0, index=pivoted.columns)
    if fin_total.notna().any():
        fin_sum = pd.Series(0.0, index=pivoted.columns)
        _fin_contribs = {}
        
        fin_components = [
            'Total Debt Issued', 'Total Debt Repaid', 'Share Repurchases',
            'Dividends Paid', 'Stock Option Proceeds', 'Taxes Paid on Stock Awards'
        ]
        
        fin_sum += issued_n
        _fin_contribs['Total Debt Issued'] = issued_n.copy()
        fin_sum -= repaid_n
        _fin_contribs['Total Debt Repaid'] = -repaid_n
        for _lbl0, _w0 in (('Share Repurchases', -1.0), ('Dividends Paid', -1.0),
                           ('Stock Option Proceeds', 1.0), ('Taxes Paid on Stock Awards', -1.0)):
            _v0 = get_num(_lbl0, preferred_cat='3_Cash_Flow').fillna(0)
            fin_sum += _v0 * _w0
            _fin_contribs[_lbl0] = _v0 * _w0
        
        _fin_counted = set()
        for _l in fin_components:
            _info = CONCEPT_MAP.get(_l)
            if isinstance(_info, dict): _fin_counted.update(_info.get('tags') or [])
        _fin_counted = frozenset(_fin_counted)
        _dyn_fin = []
        for _cat0, _lbl in pivoted.index:
            if _cat0 != '3_Cash_Flow' or _lbl in fin_components or _lbl in _cf_claimed: continue
            if ':' in _lbl or ' - ' in _lbl or '(Net)' in _lbl: continue
            _info = CONCEPT_MAP.get(_lbl)
            if not (isinstance(_info, dict) and _info.get('auto')): continue
            _kind, _w = _classify_calc_lineage(_info.get('tags'), _fin_counted, targets=_CF_FINANCING_PARENTS, blockers=_CF_OPERATING_PARENTS | _CF_INVESTING_PARENTS)
            if _kind is None or not _w:
                _fw = _classify_cf_label_fallback(_lbl, 'fin')
                if _fw is not None:
                    _kind, _w = 'target', _fw
            if _kind == 'target' and _w:
                _val = get_num(_lbl, preferred_cat='3_Cash_Flow')
                if _val.notna().any():
                    fin_sum += _val.fillna(0) * _w
                    _fin_contribs[_lbl] = _val.fillna(0) * _w
                    _cf_claimed.add(_lbl)
                    _dyn_fin.append(_lbl)
        if _dyn_fin: print(f"  [Bridge] Financing-CF components via calc linkbase: {', '.join(_dyn_fin)}")

        fin_sum = _bridge_sign_self_check(
            pd.to_numeric(fin_total, errors='coerce'), fin_sum, _fin_contribs, 'financing-CF')
        fin_residual = pd.to_numeric(fin_total, errors='coerce').fillna(0) - fin_sum
        _CF_BRIDGE_SPEC.clear()
        for _sec, _ctr in (('op', _ocf_contribs), ('inv', _inv_contribs), ('fin', _fin_contribs)):
            _sp, _cm = _spec_from_contribs(_ctr, lambda l: get_row(l, preferred_cat='3_Cash_Flow'))
            _CF_BRIDGE_SPEC[_sec] = {'spec': _sp, 'complete': _cm}
        _BRIDGE_USED_LABELS.clear()
        for _ctr in (_ocf_contribs, _inv_contribs, _fin_contribs):
            _BRIDGE_USED_LABELS.update(_ctr.keys())
        _BRIDGE_USED_LABELS.update(op_components)
        for _l in list(CONCEPT_MAP):
            if isinstance(CONCEPT_MAP.get(_l), dict) and CONCEPT_MAP[_l].get('is_dyn_opex'):
                _BRIDGE_USED_LABELS.add(_l)
        if (fin_residual.abs() > 1e6).any():
            add_val('3_Cash_Flow', 'Other Financing Adjustments (Net)', fin_residual[fin_total.notna()])

    # -- Net Cash Flow (Operating + Investing + Financing CF) ---------
    ocf_ncf   = pd.to_numeric(ocf_total,  errors='coerce')
    icf_ncf   = pd.to_numeric(inv_total,  errors='coerce')
    fincf_ncf = pd.to_numeric(fin_total,  errors='coerce')
    net_cf    = ocf_ncf + icf_ncf + fincf_ncf
    mask_ncf  = ocf_ncf.notna() & icf_ncf.notna() & fincf_ncf.notna()
    if mask_ncf.any():
        add_val('3_Cash_Flow', 'Net Cash Flow', net_cf[mask_ncf])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # INCOME STATEMENT FOOTING -- Granular Bridges
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    # -- Gross Profit Level: GP = Rev âˆ’ COGS (+ adj) -----------------
    gp_res = gp_n - (rev_n - cor_n.fillna(0))
    if gp_res.notna().any() and gp_res.abs().max() > 1e6:
        add_val('1_Income_Statement', 'Gross Profit: Other Adjustments', gp_res[gp_n.notna()])

    # -- Operating Income Bridge --------------------------------------
    # Build from individual opex items (never reference TotalOpEx which
    # may include COGS via the CostsAndExpenses tag).
    opex_bridge = rd_n.fillna(0) + effective_sga

    # IBM-style detection: does a filed total-expense subtotal already
    # reconcile operating income (GP - TotalOpEx == OpInc)? If so, the
    # filed subtotal is authoritative and components must NOT be summed.
    _toe_filed = get_num('Total Operating Expenses', preferred_cat='1_Income_Statement')
    _pretax_id = get_num('Pretax Income', preferred_cat='1_Income_Statement')
    _scale = gp_n.abs().where(gp_n.abs() > 0, np.nan)
    # The expense-and-other-income block nets straight to PRETAX (there is no
    # separate operating subtotal). Anchor on pretax, which is reliably tagged.
    _rel_id = (gp_n - _toe_filed - _pretax_id).abs() / _scale
    _ibm_mask = (gp_n.notna() & _toe_filed.notna() & _pretax_id.notna()
                 & (_rel_id < 0.02))
    # Structural gate: a genuine IBM-style statement reports NO operating-income
    # line. Ordinary filers DO (even when OpInc ~= Pretax), so they are excluded
    # -- their interest stays below the operating line.
    _files_op_inc = bool(_op_inc_was_filed.sum() >= 4)
    _ibm_style = bool(_ibm_mask.sum() >= 4) and not _files_op_inc
    _IBM_STYLE_STATE['active'] = _ibm_style
    if _ibm_style:
        print(f"  [Bridge] IBM-style expense structure detected in "
              f"{int(_ibm_mask.sum())} quarter(s): the filed expense-and-other-"
              f"income block nets straight to pretax (GP - TotalOpEx = Pretax) "
              f"and no separate operating-income line is reported. Using the "
              f"filed subtotal as the opex bridge; operating income equals "
              f"pretax (interest sits inside the block, as filed).")

    # -- Dynamic opex components (calc-linkbase classified) ------------
    # Learned face lines (Amazon's Fulfillment / Technology & Infrastructure,
    # Visa's Network & Processing, ...) and financial-statement opex labels
    # are wired into the bridge so 'Operating Income: Other Adjustments'
    # stays a small true residual instead of absorbing the company's
    # largest cost lines. By dynamically processing all non-core lines via the SEC 
    # linkbase, we ensure exact math regardless of varying taxonomies.
    _already_in_bridge = {
        'Revenue', 'Cost of Revenue', 'Gross Profit', 'Research & Development',
        'Selling, General & Admin', 'Sales & Marketing', 'General & Administrative',
        'Marketing Expense', 'Operating Income', 'Total Operating Expenses',
    }
    _counted_concepts = set()
    for _l in _already_in_bridge:
        _info = CONCEPT_MAP.get(_l)
        if isinstance(_info, dict):
            _counted_concepts.update(_info.get('tags') or [])
    _counted_concepts = frozenset(_counted_concepts)

    _dyn_opex = []
    _fb_opex_sum = pd.Series(0.0, index=pivoted.columns)
    _fb_opex_labels = []
    for _cat0, _lbl in pivoted.index:
        if _cat0 != '1_Income_Statement' or _lbl in _already_in_bridge:
            continue
        if ':' in _lbl or ' - ' in _lbl:
            continue
        _ll = _lbl.lower()
        if any(_k in _ll for _k in ('total', 'eps', 'per share', 'shares outstanding',
                                    'net income', 'pretax', 'income tax', 'interest',
                                    'comprehensive', 'noncontrolling')):
            continue
        _info = CONCEPT_MAP.get(_lbl)
        if not isinstance(_info, dict):
            continue
        _kind, _w = _classify_calc_lineage(
            _info.get('tags'), _counted_concepts,
            targets=_OPEX_ROLLUP_PARENTS, blockers=_REVENUE_ROLLUP_PARENTS)
        if _kind == 'target' and _w:
            _r = get_num(_lbl, preferred_cat='1_Income_Statement')
            if _r.notna().any():
                opex_bridge += _r.fillna(0) * _w
                _dyn_opex.append(_lbl)
                if isinstance(CONCEPT_MAP.get(_lbl), dict):
                    CONCEPT_MAP[_lbl]['is_dyn_opex'] = True
        elif _kind is None:
            if _ibm_style and bool(_ibm_mask.all()):
                continue
            # Financial-statement expense labels (broker/bank noninterest
            # expenses) often carry custom concepts with no calc arcs.
            # Collect as a candidate group; adopted below only if it
            # decisively improves the filed operating-income fit.
            if any(_k in _ll for _k in ('compensation', 'employee benefit', 'occupancy',
                                        'communication', 'execution', 'clearing',
                                        'clearance', 'brokerage', 'regulatory',
                                        'data processing', 'professional fee',
                                        'advertising', 'bad debt', 'credit loss',
                                        'other operating expense')):
                _r = get_num(_lbl, preferred_cat='1_Income_Statement')
                if _r.notna().any():
                    _fb_opex_sum += _r.fillna(0)
                    _fb_opex_labels.append(_lbl)
    if _ibm_style:
        opex_bridge = opex_bridge.where(~_ibm_mask, _toe_filed)
    if _dyn_opex:
        print(f"  [Bridge] Opex components via calc linkbase: {', '.join(_dyn_opex)}")
    if _fb_opex_labels:
        # Reference for adoption: the filed OpInc row when it exists; the
        # filed pretax row otherwise (no-OpInc filers like HOOD would
        # otherwise have nothing to calibrate against and the derived
        # OpInc would inherit the incomplete bridge).
        _ref = op_inc_n if op_inc_n.notna().sum() >= 4 else pd.to_numeric(
            get_row('Pretax Income', preferred_cat='1_Income_Statement'), errors='coerce')
        _res0 = (_ref - (gp_n.fillna(0) - opex_bridge))[_ref.notna()].abs().median()
        _res1 = (_ref - (gp_n.fillna(0) - (opex_bridge + _fb_opex_sum)))[_ref.notna()].abs().median()
        if pd.notna(_res1) and pd.notna(_res0) and _res1 < 0.5 * _res0:
            opex_bridge = opex_bridge + _fb_opex_sum
            for _l in _fb_opex_labels:
                if isinstance(CONCEPT_MAP.get(_l), dict):
                    CONCEPT_MAP[_l]['is_dyn_opex'] = True
            print(f"  [Bridge] Opex components via expense-label fallback "
                  f"(financial-statement structure): {', '.join(_fb_opex_labels)}")

    if is_reit:
        opex_bridge = op_exp_n.where(op_exp_n.notna(), opex_bridge)

    opinc_res = op_inc_n - (gp_n.fillna(0) - opex_bridge)
    # Financial-statement structures: net interest income is OPERATING for
    # banks/brokers. If including it fits the filed operating income
    # decisively better, adopt it.
    _ii = get_num('Interest Income', preferred_cat='1_Income_Statement').fillna(0)
    _ie = get_num('Interest Expense', preferred_cat='1_Income_Statement').fillna(0)
    _nii = get_num('Net Interest Income (Expense)', preferred_cat='1_Income_Statement')
    _net_int = _nii.where(_nii.notna(), _ii - _ie).fillna(0)
    if (_net_int != 0).any():
        _res_fin = op_inc_n - (gp_n.fillna(0) + _net_int - opex_bridge)
        _m_base = opinc_res[op_inc_n.notna()].abs().median()
        _m_fin = _res_fin[op_inc_n.notna()].abs().median()
        if pd.notna(_m_fin) and pd.notna(_m_base) and _m_fin < 0.5 * _m_base:
            opinc_res = _res_fin
            print("  [Bridge] Operating bridge includes net interest income "
                  "(financial-statement structure).")
    if opinc_res.notna().any() and opinc_res.abs().max() > 1e6:
        add_val('1_Income_Statement', 'Operating Income: Other Adjustments',
                opinc_res[op_inc_n.notna()])
        _plug_ratio = (opinc_res.abs() / rev_n.abs().clip(lower=1.0)).median()
        if pd.notna(_plug_ratio) and _plug_ratio > 0.05:
            print(f"  [Bridge Warning] 'Operating Income: Other Adjustments' is "
                  f"{_plug_ratio:.0%} of revenue (median) -- face lines are likely "
                  f"missing from the income statement or unclassified.")

    # -- Pretax Income Bridge -----------------------------------------
    # Sum all individually captured non-operating items.
    # If "Total Non-operating Income" exists AND is more complete than
    # the component sum, prefer it to avoid residual inflation.
    
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # INTEREST INCOME: SUBTOTAL CORRECTION
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Filers like Robinhood sometimes file only partial interest components
    # (e.g., InterestIncomeOperating) without the total gross interest tag.
    # If Net Interest Income + Interest Expense > reported Interest Income,
    # the reported value is just a subcomponent. Overwrite it.
    net_int_inc_n = get_num('Net Interest Income (Expense)', preferred_cat='1_Income_Statement')
    int_exp_raw = get_num('Interest Expense', preferred_cat='1_Income_Statement')
    int_inc_raw = get_num('Interest Income', preferred_cat='1_Income_Statement')
    
    if net_int_inc_n.notna().any() and int_exp_raw.notna().any():
        derived_int_inc = net_int_inc_n + int_exp_raw.fillna(0)
        components_int = net_int_inc_n.notna()
        
        # Generic anomaly rescue: enforces exact mathematical identity bidirectionally
        int_inc_mask = needs_anomaly_rescue(int_inc_raw, derived_int_inc, components_int, is_exact_identity=True)
        if int_inc_mask.any():
            int_inc_raw = int_inc_raw.where(~int_inc_mask, derived_int_inc)
            add_val('1_Income_Statement', 'Interest Income', derived_int_inc[int_inc_mask])
            print("  [Subtotal Correction] Overwrote anomalous 'Interest Income' using exact identity (Net Interest + Interest Expense)")

    int_inc_n     = int_inc_raw.fillna(0)
    int_exp_kpi_n = get_num('Interest Expense').fillna(0)
    inv_gl_n      = get_num('Gain/Loss on Investments').fillna(0)
    eq_method_n   = get_num('Equity Method Income').fillna(0)
    other_inc_n   = get_num('Other Income / (Expense)').fillna(0)
    total_nonop_n = get_num('Total Non-operating Income')

    nonop_component_sum = int_inc_n - int_exp_kpi_n + inv_gl_n + eq_method_n + other_inc_n

    # Candidate non-operating bases: component sum, filed total, and ZERO
    # (financial-statement structures fold interest into operating income,
    # so adding it again double-counts). Choose per company by global fit,
    # then per quarter between the two best where applicable.
    _zero_nonop = pd.Series(0.0, index=pretax_n.index)
    _cands = [('components', nonop_component_sum)]
    if total_nonop_n.notna().any():
        _cands.append(('total', total_nonop_n.fillna(0)))
    _cands.append(('zero', _zero_nonop))
    _fits = {}
    for _nm, _c in _cands:
        _r = (pretax_n - (op_inc_n.fillna(0) + _c)).abs()
        _fits[_nm] = _r[pretax_n.notna()].median()
    _best_nm = min(_fits, key=lambda k: _fits[k] if pd.notna(_fits[k]) else np.inf)
    if _best_nm == 'zero' and pd.notna(_fits.get('components')):
        print("  [Bridge] Pretax bridge: operating income already contains "
              "non-operating items (financial-statement structure); "
              "interest/other income not re-added.")
    if _best_nm == 'total':
        res_with_components = (pretax_n - (op_inc_n.fillna(0) + nonop_component_sum)).abs()
        res_with_total      = (pretax_n - (op_inc_n.fillna(0) + total_nonop_n.fillna(0))).abs()
        use_total = res_with_total < res_with_components
        effective_nonop = nonop_component_sum.where(~use_total, total_nonop_n.fillna(0))
    else:
        effective_nonop = dict(_cands)[_best_nm]

    if _ibm_style:
        # Interest already inside the expense block in IBM-style quarters:
        # operating income == pretax there, so the non-operating bridge is zero.
        effective_nonop = effective_nonop.where(~_ibm_mask, 0.0)

    pretax_res = pretax_n - (op_inc_n.fillna(0) + effective_nonop)
    if pretax_res.notna().any() and pretax_res.abs().max() > 1e6:
        add_val('1_Income_Statement', 'Pretax Income: Other Adjustments', pretax_res[pretax_n.notna()])

    # -- Net Income Bridge: Pretax âˆ’ Tax + DiscOps âˆ’ NCI --------------
    disc_ops_n = get_num('Income from Discontinued Operations').fillna(0)
    nci_n_bridge = get_num('Net Income to Noncontrolling Interest').fillna(0)
    _res_with_nci = ni_n - (pretax_n.fillna(0) - tax_n.fillna(0) + disc_ops_n - nci_n_bridge)
    _res_no_nci   = ni_n - (pretax_n.fillna(0) - tax_n.fillna(0) + disc_ops_n)
    if (nci_n_bridge != 0).any():
        _m_with = _res_with_nci[ni_n.notna()].abs().median()
        _m_no   = _res_no_nci[ni_n.notna()].abs().median()
        if pd.notna(_m_no) and pd.notna(_m_with) and _m_no < 0.5 * _m_with:
            ni_res = _res_no_nci
            print("  [Bridge] Net-income bridge: 'Net Income' row is consolidated "
                  "(includes noncontrolling interests); NCI not re-subtracted.")
        else:
            ni_res = _res_with_nci
    else:
        ni_res = _res_with_nci
    if ni_res.notna().any() and ni_res.abs().max() > 1e6:
        add_val('1_Income_Statement', 'Net Income: Other Adjustments', ni_res[ni_n.notna()])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ENHANCED FOOTING VERIFICATION
    # 1.0 = Verified (all residuals < 50M), 0.0 = Not Verified
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    footing_verified = pd.Series(1.0, index=pivoted.columns)

    # Check Income Statement residuals
    for res in [gp_res, opinc_res, pretax_res, ni_res]:
        if res.notna().any():
            footing_verified[res.abs() > 50e6] = 0.0

    # Check Cash Flow residuals (only if totals exist)
    if inv_total.notna().any():
        footing_verified[inv_residual.abs() > 50e6] = 0.0
    if fin_total.notna().any():
        footing_verified[fin_residual.abs() > 50e6] = 0.0

    add_val('8_Integrity_Checks', 'Metric: Financials Footing Verified', footing_verified)
    
    # -- Basis-aware Segment Revenue Verification and Residual Repair --
    # Only verify business-segment revenue when we can identify a clean,
    # additive top-level segment basis that usually ties to consolidated
    # revenue. This prevents noisy partial cuts (HOOD revenue categories,
    # UNH external/intersegment matrices, acquisition/reclass rows) from
    # causing false failures while still catching true segment gaps (META Q4).
    def _seg_member_from_rev_label(lbl):
        text = str(lbl or '')
        if not text.startswith('Revenue - '):
            return ''
        return text[len('Revenue - '):].strip()

    def _is_top_level_revenue_segment(lbl):
        text = str(lbl or '')
        return text.startswith('Revenue - ') and text.count(' - ') == 1

    def _is_noise_segment_member(member):
        m = re.sub(r"\s+", " ", str(member or '').lower()).strip()
        if not m:
            return True
        # Non-additive disclosures or noisy one-off rows.
        if any(k in m for k in (
            'gain loss', 'gain/loss', 'reclass', 'acquisition',
            'investment agreement', 'accounts receivable', 'receivable',
            'tax jurisdiction', 'deferred tax', 'asset impairment',
            'concentration risk', 'share based', 'stock-based',
        )):
            return True
        # Matrix / channel cuts are not full reportable business segments.
        if any(k in m for k in (
            'external customers', 'intersegment', 'elimination',
            'product', 'service', 'financial service other',
        )):
            return True
        if m in {'other', 'total', 'total optum', 'all segments', 'corporate', 'unallocated'}:
            return True
        return False

    seg_rev_labels_all = [
        lbl for (cat, lbl) in pivoted.index
        if cat == '4a_Segments_Business' and _is_top_level_revenue_segment(lbl)
    ]
    seg_rev_labels = [
        lbl for lbl in seg_rev_labels_all
        if not _is_noise_segment_member(_seg_member_from_rev_label(lbl))
    ]

    if len(seg_rev_labels) >= 2 and rev_n.notna().any():
        seg_frame = pd.DataFrame({
            lbl: pd.to_numeric(pivoted.loc[('4a_Segments_Business', lbl)], errors='coerce')
            for lbl in seg_rev_labels
        }, index=pivoted.columns)
        seg_sum = seg_frame.fillna(0).sum(axis=1)
        seg_count = seg_frame.notna().sum(axis=1)
        comparable = rev_n.notna() & (rev_n.abs() > 0) & (seg_count >= 2)
        gap_ratio = (rev_n - seg_sum).abs() / rev_n.abs().replace(0, np.nan)
        tie_mask = comparable & (gap_ratio <= 0.05)

        # Accept the basis only if it ties often enough across the run.  This
        # keeps partial revenue-category tables out of the integrity metric.
        min_ties = max(3, int(0.50 * max(1, int(comparable.sum()))))
        basis_valid = bool(tie_mask.sum() >= min_ties)

        if basis_valid:
            # Repair one clearly bad/missing segment when the accepted basis is
            # otherwise additive: replacement = consolidated revenue - other
            # clean segment rows.  Extremely conservative: only one candidate
            # may qualify, and the current value must be missing or far below
            # the residual (META Reality Labs Q4 case).
            repaired = {}
            for p_col in pivoted.columns:
                if not comparable.get(p_col, False):
                    continue
                rev_v = rev_n.get(p_col, np.nan)
                if pd.isna(rev_v) or abs(rev_v) <= 0:
                    continue
                period_vals = seg_frame.loc[p_col]
                gap_abs = abs(float(rev_v) - float(period_vals.fillna(0).sum()))
                # Even a consolidated gap under 5% can be a huge segment error
                # when one small segment is understated (META Reality Labs Q4).
                if gap_abs <= max(50e6, 0.0025 * abs(float(rev_v))):
                    continue
                candidates = []
                for lbl in seg_rev_labels:
                    cur_v = period_vals.get(lbl, np.nan)
                    other_sum = period_vals.drop(labels=[lbl]).fillna(0).sum()
                    repl = rev_v - other_sum
                    if pd.isna(repl) or repl <= 0:
                        continue
                    if pd.notna(cur_v):
                        rel_change = abs(repl - cur_v) / max(abs(repl), 1.0)
                        if not (cur_v <= 0.55 * repl and rel_change >= 0.40):
                            continue
                    hist = seg_frame[lbl].drop(labels=[p_col], errors='ignore').dropna()
                    hist = hist[hist > 0]
                    if len(hist) >= 4:
                        med = float(hist.median())
                        # Q4 seasonality can be large, but a valid residual
                        # should still be within a broad historical envelope.
                        if med > 0 and not (0.15 * med <= repl <= 4.0 * med):
                            continue
                    candidates.append((lbl, repl, cur_v))
                if len(candidates) == 1:
                    lbl, repl, cur_v = candidates[0]
                    add_val('4a_Segments_Business', lbl, {p_col: repl})
                    repaired[(lbl, p_col)] = (cur_v, repl)
                    seg_frame.at[p_col, lbl] = repl

            if repaired:
                shown = ', '.join(f"{lbl} {period}" for (lbl, period) in list(repaired.keys())[:6])
                extra = '' if len(repaired) <= 6 else f" (+{len(repaired)-6} more)"
                print(f"  [Segment Repair] Repaired segment revenue residual(s): {shown}{extra}")

            seg_sum = seg_frame.fillna(0).sum(axis=1)
            gap_ratio = (rev_n - seg_sum).abs() / rev_n.abs().replace(0, np.nan)
            coverage = pd.Series(np.nan, index=pivoted.columns)
            comparable = rev_n.notna() & (rev_n.abs() > 0) & (seg_frame.notna().sum(axis=1) >= 2)
            coverage[comparable] = 1.0
            coverage[comparable & (gap_ratio > 0.05)] = 0.0
            add_val('8_Integrity_Checks', 'Metric: Segment Sum Verified (Revenue)', coverage[coverage.notna()])

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Q4 DE-CUMULATION FIX
    # If Q4 was derived as an annual total but now has quarterly
    # neighbors (derived in calculate_kpis), detect and fix it.
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    latest_vals = {}
    for r in kpi_rows:
        latest_vals[(r['Category'], r['Label'], r['Period'])] = r['Value']

    for label in ['Total Operating Expenses', 'Pretax Income', 'Gross Profit', 'Operating Income']:
        idx_base = ('1_Income_Statement', label)
        fy_list = set(col.split('-')[0] for col in pivoted.columns if '-' in str(col))
        for fy in fy_list:
            q1, q2, q3, q4 = f"{fy}-Q1", f"{fy}-Q2", f"{fy}-Q3", f"{fy}-Q4"

            def get_v(p):
                if (idx_base[0], label, p) in latest_vals:
                    return latest_vals[(idx_base[0], label, p)]
                if idx_base in pivoted.index and p in pivoted.columns:
                    return pivoted.loc[idx_base, p]
                return np.nan

            v1, v2, v3, v4 = get_v(q1), get_v(q2), get_v(q3), get_v(q4)
            if pd.notna(v4) and pd.notna(v1) and pd.notna(v2) and pd.notna(v3):
                q_sum = v1 + v2 + v3
                if v4 > 1.5 * q_sum and v4 > 0 and q_sum > 0:
                    discrete_q4 = v4 - q_sum
                    add_val('1_Income_Statement', label, {q4: discrete_q4})

    return pd.DataFrame(kpi_rows)

def _repair_always_positive(df):
    """
    Ensures that economically always-positive items (D&A, CapEx, etc.)
    don't have negative quarterly values due to tag mismatches.
    """
    ALWAYS_POSITIVE_CF_LABELS = {
        'Depreciation & Amortization', 'Depreciation', 'Amortization',
        'Stock-Based Compensation', 'Capital Expenditures', 'Capital Expenditure',
        'Capex', 'Dividends Paid',
        'Taxes Paid on Stock Awards', 'Total Debt Repaid', 'Short-term Debt Repaid', 
        'Long-term Debt Repaid', 'Share Repurchases',
        'Purchases of Investments',
    }
    df = df.copy()
    num_cols = df.columns.tolist()
    # Iterate physical rows without building a Series per row.  ``t[0]`` is the
    # (Category, Label) index tuple; ``t[j + 1]`` is the value for
    # ``num_cols[j]``.  The cell value only feeds ``pd.notna``/``float`` and the
    # write target/print are unchanged, so behavior is identical to iterrows.
    for _t in df.itertuples(index=True, name=None):
        idx = _t[0]
        label = idx[1]
        base_label = label.split(' - ')[0]
        
        is_strictly_positive = False
        label_lower = label.lower()
        if 'share based' in label_lower or 'stock based' in label_lower:
            is_strictly_positive = True
        elif any(base_label == kw or base_label.startswith(kw) for kw in ALWAYS_POSITIVE_CF_LABELS):
            is_strictly_positive = True
            
        if is_strictly_positive:
            for _j, col in enumerate(num_cols):
                val = _t[_j + 1]
                if pd.notna(val) and float(val) < 0:
                    print(f"    [!] Negative value mapped to NaN (Strictly Positive Fallback) -> {label} {col}")
                    df.at[idx, col] = np.nan
    return df

def _remerge_false_positive_splits(df):
    """
    Check for ' (Pre Change)' and ' (Post Change)' labels.
    If they match 100% on overlapping non-NaN periods, re-merge them.
    Used to fix false positive segment splits (e.g. in Adobe).
    """
    labels = df.index.get_level_values('Label')
    post_suffix = ' (Post Change)'
    pre_suffix = ' (Pre Change)'
    
    if not any(l.endswith(post_suffix) for l in labels):
        return df
        
    df = df.copy()
    to_drop = []
    
    # Iterate over unique combinations of (Category, Base Label)
    checked_bases = set()
    for cat, label in df.index:
        if label.endswith(post_suffix):
            base_label = label[:-len(post_suffix)]
            if (cat, base_label) in checked_bases: continue
            checked_bases.add((cat, base_label))
            
            pre_label = base_label + pre_suffix
            if (cat, pre_label) in df.index:
                row_pre = df.loc[(cat, pre_label)]
                row_post = df.loc[(cat, label)]
                
                # Check for overlap and 100% match
                common_mask = row_pre.notna() & row_post.notna()
                if common_mask.any():
                    v_pre = row_pre[common_mask].values.astype(float)
                    v_post = row_post[common_mask].values.astype(float)
                    
                    # Institutional-grade match check (allowing for tiny float noise but conceptually 100%)
                    if np.allclose(v_pre, v_post, rtol=0, atol=1e-2):
                        print(f"  [Segment Era Split Fix] Re-merging false positive split: {cat} - {base_label}")
                        merged_row = row_post.fillna(row_pre)
                        df.loc[(cat, base_label), :] = merged_row
                        to_drop.append((cat, pre_label))
                        to_drop.append((cat, label))
    
    if to_drop:
        df = df.drop(index=list(set(to_drop)))
    return df

def apply_stock_splits(df):
    """
    Detect and correct stock splits using level-clustering for shares and
    Net-Income consistency checks for EPS.
    """
    SHARES_BASIC   = ('1_Income_Statement', 'Shares Outstanding Basic')
    SHARES_DILUTED = ('1_Income_Statement', 'Shares Outstanding Diluted')
    EPS_BASIC      = ('1_Income_Statement', 'EPS Basic')
    EPS_DILUTED    = ('1_Income_Statement', 'EPS Diluted')
    NI_IDX         = ('1_Income_Statement', 'Net Income')

    shares_indices = [SHARES_BASIC, SHARES_DILUTED]
    eps_indices    = [EPS_BASIC, EPS_DILUTED]

    has_any = any(idx in df.index for idx in shares_indices + eps_indices)
    if not has_any: return df

    df = df.copy()
    for c in df.columns:
        if isinstance(df[c].dtype, pd.StringDtype): df[c] = df[c].astype(object)

    cols = df.columns.tolist()
    col_pos = {c: i for i, c in enumerate(cols)}
    _num = {idx: pd.to_numeric(df.loc[idx], errors='coerce') for idx in shares_indices + eps_indices + [NI_IDX] if idx in df.index}

    base_splits = [1.5, 2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 20, 25, 28, 50, 100]
    known_factors = sorted(set(base_splits) | {a * b for a in base_splits for b in base_splits if a * b <= 10000})
    MIN_SHARES = 1_000_000

    def _closest_factor(ratio):
        c = min(known_factors, key=lambda x: abs(x - ratio))
        return c, abs(ratio - c) / c

    shares_adj = {}
    for sh_idx in shares_indices:
        if sh_idx not in _num: continue
        series = _num[sh_idx]
        valid = [(col, float(series[col])) for col in cols if pd.notna(series.get(col)) and float(series[col]) > MIN_SHARES]
        if len(valid) < 3: continue
        by_val = sorted(valid, key=lambda x: x[1])
        clusters = [[by_val[0]]]
        for i in range(1, len(by_val)):
            if by_val[i][1] / by_val[i - 1][1] > 2.0: clusters.append([])
            clusters[-1].append(by_val[i])
        if len(clusters) <= 1: continue
        clusters.sort(key=lambda c: np.median([v for _, v in c]), reverse=True)
        cumulative = 1.0
        for k in range(len(clusters) - 1):
            higher, lower = {col: val for col, val in clusters[k]}, {col: val for col, val in clusters[k + 1]}
            ratios = []
            for ci, col in enumerate(cols):
                if col in higher:
                    for delta in [1, -1, 2, -2, 3, -3]:
                        ni = ci + delta
                        if 0 <= ni < len(cols) and cols[ni] in lower:
                            ratios.append(higher[col] / lower[cols[ni]]); break
            if not ratios: ratios = [np.median([v for _, v in clusters[k]]) / np.median([v for _, v in clusters[k + 1]])]
            closest, err = _closest_factor(float(np.median(ratios)))
            if err >= 0.10: continue
            h_pos, l_pos = sorted(col_pos[c] for c in higher), sorted(col_pos[c] for c in lower)
            all_pos = sorted(h_pos + l_pos)
            valid_split = False
            
            for p_idx in range(len(all_pos) - 1):
                cut = (all_pos[p_idx] + all_pos[p_idx+1]) / 2.0
                len_L = sum(1 for p in all_pos if p < cut)
                len_R = sum(1 for p in all_pos if p > cut)
                
                err_L_fwd = sum(1 for p in l_pos if p < cut)
                err_R_fwd = sum(1 for p in h_pos if p > cut)
                
                err_L_rev = sum(1 for p in h_pos if p < cut)
                err_R_rev = sum(1 for p in l_pos if p > cut)
                
                # Requires one clean side (0 errors) while allowing 10-K restatement noise (up to 30%) on the other
                if (err_L_fwd == 0 and err_R_fwd <= 0.30 * len_R) or (err_R_fwd == 0 and err_L_fwd <= 0.30 * len_L) or \
                   (err_L_rev == 0 and err_R_rev <= 0.30 * len_R) or (err_R_rev == 0 and err_L_rev <= 0.30 * len_L):
                    valid_split = True
                    break
            
            if not valid_split: continue
            cumulative *= closest
            for col, _ in clusters[k+1]: shares_adj[col] = max(shares_adj.get(col, 1.0), cumulative)
        if shares_adj: break

    eps_adj = {}
    valid_eps_factors = sorted(set(shares_adj.values())) if shares_adj else []
    if NI_IDX in _num and valid_eps_factors:
        ni_s = _num[NI_IDX]
        # EPS numerator basis: for umbrella structures (e.g. IBKR) the
        # consolidated 'Net Income' row includes large noncontrolling
        # interests, so NI/shares overstates expected EPS -- which can make a
        # PRE-SPLIT EPS look 'consistent' by coincidence (the two errors
        # nearly cancel) and escape adjustment.  Pick the basis (consolidated
        # vs ex-NCI) whose implied filed-EPS ratios cluster closest to 1x or
        # a known split factor.
        NCI_IDX = ('1_Income_Statement', 'Net Income to Noncontrolling Interest')
        if NCI_IDX in df.index:
            nci_s = pd.to_numeric(df.loc[NCI_IDX], errors='coerce').fillna(0)
            if (nci_s != 0).any():
                _targets = [1.0] + [float(f) for f in valid_eps_factors]
                _sh0 = next((_num[i] for i in shares_indices if i in _num), None)
                _eps0 = next((_num[i] for i in eps_indices if i in _num), None)

                def _basis_score(basis):
                    if _sh0 is None or _eps0 is None:
                        return np.inf
                    dists = []
                    for col in cols:
                        b, s, e = basis.get(col, np.nan), _sh0.get(col, np.nan), _eps0.get(col, np.nan)
                        if pd.isna(b) or pd.isna(s) or pd.isna(e) or s < MIN_SHARES or abs(e) < 1e-6:
                            continue
                        exp = b / (s * shares_adj.get(col, 1.0))
                        if abs(exp) < 1e-9:
                            continue
                        r = abs(e / exp)
                        dists.append(min(abs(r - t) / t for t in _targets))
                    return float(np.median(dists)) if dists else np.inf

                _sc_incl = _basis_score(ni_s)
                _sc_ex = _basis_score(ni_s - nci_s)
                if _sc_ex < _sc_incl:
                    ni_s = ni_s - nci_s
                    print(f"  [Stock Split] EPS consistency basis: NI excluding noncontrolling "
                          f"interests (fit {_sc_ex:.2f} vs {_sc_incl:.2f}).")
        sh_series = next((_num[idx] for idx in shares_indices if idx in _num), None)
        if sh_series is not None:
            for col in cols:
                ni_v, sh_v = ni_s.get(col, np.nan), sh_series.get(col, np.nan)
                if pd.isna(ni_v) or pd.isna(sh_v) or abs(ni_v) < 1e3 or sh_v < MIN_SHARES: continue
                expected_eps = ni_v / (sh_v * shares_adj.get(col, 1.0))
                for ep_idx in eps_indices:
                    if ep_idx not in _num: continue
                    eps_v = _num[ep_idx].get(col, np.nan)
                    if pd.isna(eps_v) or abs(expected_eps) < 1e-6: continue
                    ratio = abs(eps_v / expected_eps)
                    if ratio < 1.3: continue
                    closest = min(valid_eps_factors, key=lambda x: abs(x - ratio))
                    if abs(ratio - closest) / closest < 0.15: eps_adj[col] = max(eps_adj.get(col, 1.0), closest)

    from collections import defaultdict
    def _log_group(adj_dict, signal_name, verb):
        by_factor = defaultdict(list)
        for col, f in sorted(adj_dict.items(), key=lambda x: col_pos.get(x[0], 9999)): by_factor[f].append(col)
        for factor, qs in sorted(by_factor.items()):
            nice = int(factor) if factor == int(factor) else factor
            print(f"  [Stock Split] {signal_name} {verb}{nice} for {len(qs)} quarters: {qs[0]}..{qs[-1]}")
    if shares_adj: _log_group(shares_adj, "Shares", "x")
    if eps_adj: _log_group(eps_adj, "EPS", "/")
    for col, factor in shares_adj.items():
        for idx in shares_indices:
            if idx in df.index:
                v = _num[idx].get(col, np.nan)
                if pd.notna(v): df.at[idx, col] = float(v) * factor
    for col, factor in eps_adj.items():
        for idx in eps_indices:
            if idx in df.index:
                v = _num[idx].get(col, np.nan)
                if pd.notna(v): df.at[idx, col] = float(v) / factor
    if shares_adj or eps_adj: print("  [Stock Split] Adjusted to current split basis.")
    return df

# ---------------------------------------------------------------------------
# Institutional Cleanup Engine (post-pivot)
# ---------------------------------------------------------------------------
# Dynamic learning + multiple resolution tiers can surface the SAME economic
# line under two labels (e.g. curated 'Acquisitions' vs the filer's learned
# 'Acquisitions Net Of Cash Acquired And Purchases Of Intangible And Other
# Assets', or 'Sales & Marketing' vs 'Marketing Expense' which share a tag).
# Filers also rename lines across eras (Amazon's 'Technology and content' ->
# 'Technology and infrastructure'), and broker-dealers repeat face revenue
# lines in the dimensional segment data.  This pass reconciles all of that
# the way S&P Capital IQ does: one row per economic line, gaps filled from
# whichever source had the data, redundant echoes removed.
# ---------------------------------------------------------------------------

_MERGE_STOP_TOKENS = {'of', 'and', 'the', 'for', 'to', 'in', 'a', 'an', 'from', 'on'}

# Tokens that flag two similarly-named rows as DIFFERENT economic lines
# (mirror/contra entries, direction, tenor, share class) -- never merge
# across these even when values coincide.
_MERGE_CONTRAST_TOKENS = {
    'operating', 'investing', 'financing', 'basic', 'diluted',
    'current', 'noncurrent', 'short', 'long', 'term', 'total', 'gross',
    'issued', 'repaid', 'repayment', 'proceed', 'purchase', 'payment',
    'paid', 'received', 'borrowing', 'beginning', 'ending',
    'continuing', 'discontinued',
    # NOTE: this set is for the cleanup merge; the era-stitch set is separate.
    # asset/liability mirror pairs are often EQUAL by construction
    # (e.g. crypto safeguarding asset == safeguarding liability) but are
    # distinct lines -- never merge across these.
    'asset', 'liability', 'obligation', 'receivable', 'payable',
    'custody', 'safeguarding', 'held',
}


@lru_cache(maxsize=65536, typed=True)
def _merge_tokens(label: str) -> frozenset:
    l = re.sub(r'\(.*?\)', ' ', str(label).lower()).replace('&', ' and ')
    toks = set()
    for t in re.sub(r'[^\w\s]', ' ', l).split():
        if t in _MERGE_STOP_TOKENS or len(t) <= 1:
            continue
        if len(t) > 3 and t.endswith('s') and not t.endswith('ss'):
            t = t[:-1]
        toks.add(t)
    return frozenset(toks)


def _is_auto_label(label: str) -> bool:
    info = CONCEPT_MAP.get(label)
    return bool(isinstance(info, dict) and info.get('auto'))


def _is_protected_row(label: str) -> bool:
    """Derived bridge/plug rows must never merge with filed rows: their
    values frequently coincide with a real line by construction (e.g.
    'Net Income: Other Adjustments' == NCI, 'Total Net Debt Issued
    (Repaid)' == 'Total Debt Issued' whenever repayments are ~0)."""
    if ': Other Adjustments' in label or label.startswith('Metric:'):
        return True
    info = CONCEPT_MAP.get(label)
    if isinstance(info, dict) and not info.get('tags') and not info.get('auto'):
        return True   # pure-derived entry (empty tag list)
    return False


def _rows_equal(v1, v2, min_overlap=3):
    """True if the two numeric series agree on every overlapping period
    (>= min_overlap periods).  Returns (equal, overlap_count)."""
    o = v1.notna() & v2.notna()
    n = int(o.sum())
    if n < min_overlap:
        return False, n
    a, b = v1[o].astype(float).values, v2[o].astype(float).values
    if np.allclose(a, 0, atol=1e-9) and np.allclose(b, 0, atol=1e-9):
        return False, n   # all-zero rows are not evidence of identity
    return bool(np.allclose(a, b, rtol=1e-4, atol=0.02)), n


def _rows_mostly_equal(v1, v2, min_overlap=3, min_rate=0.7):
    """True if the two series agree near-exactly on at least `min_rate` of
    their overlapping periods (>= min_overlap of them).  Lets a stale curated
    line that was renamed/re-tagged merge into the current dimensional line
    when a single boundary quarter disagrees only because it is annual-derived
    (e.g. IBM 'Product Revenue' 2012-2017 vs 'Revenue - Product' 2017-2026:
    three of four 2017 quarters match to the dollar, Q4 is YTD-derived).
    Kept deliberately tight (0.1% per-period) so genuinely re-scoped lines
    (IBM 'Cost of Services' post-Kyndryl) do NOT merge."""
    o = v1.notna() & v2.notna()
    n = int(o.sum())
    if n < min_overlap:
        return False
    a, b = v1[o].astype(float).values, v2[o].astype(float).values
    if np.allclose(a, 0, atol=1e-9) and np.allclose(b, 0, atol=1e-9):
        return False
    match = np.isclose(a, b, rtol=1e-3, atol=0.02)
    return bool(match.mean() >= min_rate and match.sum() >= min_overlap)


def _mergeable_names(lbl1, lbl2):
    """Name-compatibility gate: token subset (one line name contained in the
    other) always passes; otherwise moderate overlap with no contrast token."""
    t1, t2 = _merge_tokens(lbl1), _merge_tokens(lbl2)
    if not t1 or not t2:
        return False
    if t1 <= t2 or t2 <= t1:
        return True
    jac = len(t1 & t2) / len(t1 | t2)
    if jac < 0.25:
        return False
    return not ((t1 ^ t2) & _MERGE_CONTRAST_TOKENS)


def _reconcile_equity_with_nci(df: pd.DataFrame) -> pd.DataFrame:
    """Fold noncontrolling interest into Total Equity when it closes the balance.

    Some consolidators (e.g. Interactive Brokers, Palantir) report "Total Equity"
    as the parent-only figure (us-gaap:StockholdersEquity) and present
    noncontrolling interest as a separate balance-sheet line. The accounting
    identity Assets = Liabilities + Total Equity then fails by exactly the NCI
    amount. When a balance-sheet NCI row exists and adding it to Total Equity
    demonstrably closes the balance across the majority of periods, fold NCI in so
    "Total Equity" means total equity including NCI (the standard definition).

    Guarded to fire ONLY when it provably closes the balance, so companies that
    already report an NCI-inclusive total (gap ~ 0) or have no NCI are untouched.
    """
    BS = '2_Balance_Sheet'
    idx = df.index
    def _row(label):
        key = (BS, label)
        return pd.to_numeric(df.loc[key], errors='coerce') if key in idx else None
    assets = _row('Total Assets')
    liab = _row('Total Liabilities')
    equity = _row('Total Equity')
    if assets is None or liab is None or equity is None:
        return df
    # Locate the balance-sheet NCI row by label (auto-learned; label varies).
    nci_key = None
    for cat, label in idx:
        if cat != BS:
            continue
        ll = label.lower()
        if ('noncontrolling' in ll or 'minority interest' in ll) and 'attributable to' not in ll.split('noncontrolling')[0]:
            # exclude IS-style "net income attributable to ..." that may share words; BS only here
            nci_key = (cat, label)
            break
    if nci_key is None:
        # fallback: any BS row mentioning noncontrolling/minority
        for cat, label in idx:
            if cat == BS and ('noncontrolling' in label.lower() or 'minority interest' in label.lower()):
                nci_key = (cat, label); break
    if nci_key is None:
        return df
    nci = pd.to_numeric(df.loc[nci_key], errors='coerce')

    gap = assets - liab - equity
    both = assets.notna() & liab.notna() & equity.notna() & nci.notna()
    if both.sum() == 0:
        return df
    # Materiality: gap is non-trivial relative to assets.
    material = (gap.abs() > assets.abs().clip(lower=1) * 0.001) & both
    if material.sum() == 0:
        return df
    # Does adding NCI close the gap? (residual after adding NCI is ~0)
    closes = (gap - nci).abs() <= (assets.abs().clip(lower=1) * 0.002)
    # Only proceed if NCI explains the gap for the majority of material periods.
    m = material & nci.notna()
    if m.sum() == 0 or (closes & m).sum() < (m.sum() * 0.6):
        return df
    # Fold NCI into Total Equity ONLY for periods where the gap is material and
    # adding NCI closes the balance. Periods that already report an NCI-inclusive
    # total (gap ~ 0) are left untouched.
    add_mask = material & nci.notna() & closes
    if add_mask.any():
        cur = df.loc[('2_Balance_Sheet', 'Total Equity')].copy()
        cur.loc[add_mask] = (equity[add_mask] + nci[add_mask])
        df.loc[('2_Balance_Sheet', 'Total Equity')] = cur.values
    return df


def _surface_income_statement_dda(df: pd.DataFrame, is_oil_gas: bool = False, is_reit: bool = False) -> pd.DataFrame:
    """Surface depreciation & (for oil & gas) depletion & amortization on the income
    statement for industries where it is a major income-statement cost line rather
    than merely a cash-flow add-back -- oil & gas (depletion of reserves) and REITs
    (depreciation of real estate, typically a REIT's single largest expense).

    The same DepreciationDepletionAndAmortization fact is referenced by both the
    income statement and the cash flow statement, so we surface the captured value as
    an income-statement line too. Additive and industry-guarded: it never runs for any
    other company, never overwrites a value already present on the income statement,
    and is a no-op if no depreciation was captured.
    """
    if not (is_oil_gas or is_reit):
        return df
    is_label = 'Depreciation, Depletion & Amortization' if is_oil_gas else 'Depreciation & Amortization Expense'
    is_key = ('1_Income_Statement', is_label)
    if is_key in df.index and pd.to_numeric(df.loc[is_key], errors='coerce').notna().any():
        return df
    cf_dda = None
    for lbl in ('Depreciation & Amortization', 'Depreciation & Depletion', 'Depreciation'):
        k = ('3_Cash_Flow', lbl)
        if k in df.index:
            vals = pd.to_numeric(df.loc[k], errors='coerce')
            if vals.notna().any():
                cf_dda = vals
                break
    if cf_dda is None:
        return df
    df.loc[is_key, :] = cf_dda.values
    return df


def _institutional_cleanup(df):
    with _ProfileTimer("_institutional_cleanup"):
        return _institutional_cleanup_impl(df)


def _institutional_cleanup_impl(df):
    # Labels currently wired into the CF/IS bridges are load-bearing:
    # dropping them desynchronizes the recomputed residuals (AMZN's
    # unearned-revenue pair and IBKR's payables/segregated rows were
    # wired, then dropped, leaving their ghost in the plug rows).
    _bridge_locked = set()
    try:
        for _sec_entry in _CF_BRIDGE_SPEC.values():
            _bridge_locked.update((_sec_entry.get('spec') or {}).keys())
        _bridge_locked |= _BRIDGE_USED_LABELS
    except Exception:
        pass
    """
    Post-pivot reconciliation:
      1. Drop rows with no data at all (artifacts of period filtering).
      2. Within each core statement, merge rows that are the same economic
         line (identical overlapping values + compatible names).  Curated
         labels win over auto-learned ones; gaps are filled from the loser.
      3. Stitch era-renamed auto rows (compatible names, disjoint periods)
         under the most recent label.
      4. Drop dimensional segment rows that merely echo a face-statement
         line (identical values, name containment, no extra data).
    """
    if df is None or df.empty:
        return df
    df = df.copy()
    num = df.apply(pd.to_numeric, errors='coerce')

    # Per-run string/metadata caches for the nested cleanup comparisons.
    # These call the same helpers as before; they only avoid recomputing them
    # for the same labels during one cleanup pass.
    _tokens_cache = {}
    _auto_cache = {}
    _protected_cache = {}

    def _tokens_for(label):
        key = str(label)
        cached = _tokens_cache.get(key)
        if cached is not None:
            return cached
        val = _merge_tokens(key)
        _tokens_cache[key] = val
        return val

    def _is_auto_cached(label):
        key = str(label)
        cached = _auto_cache.get(key)
        if cached is not None:
            return cached
        val = _is_auto_label(key)
        _auto_cache[key] = val
        return val

    def _is_protected_cached(label):
        key = str(label)
        cached = _protected_cache.get(key)
        if cached is not None:
            return cached
        val = _is_protected_row(key)
        _protected_cache[key] = val
        return val

    # ---- Pass 1: all-empty rows -----------------------------------------
    empty = ~num.notna().any(axis=1)
    if empty.any():
        for cat_l, lbl_l in df.index[empty]:
            print(f"  [Cleanup] Dropping empty row '{lbl_l}'")
        df, num = df[~empty], num[~empty]

    def _pref_key(idx):
        """Lower = preferred survivor: curated first, then the row whose
        data extends most recently (columns are newest-first) so era merges
        keep the company's CURRENT line name, then more data."""
        cat_k, lbl_k = idx
        row = num.loc[idx]
        nn = int(row.notna().sum())
        first_nn = next((p for p, c in enumerate(df.columns) if pd.notna(row[c])), 9999)
        return (1 if _is_auto_cached(lbl_k) else 0, first_nn, -nn)

    dead, fills = set(), []   # fills: (survivor_idx, source_idx)

    # ---- Pass 2 & 3: within-category merges ------------------------------
    for cat in ('1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow'):
        rows = [i for i in df.index if i[0] == cat]
        for a_pos in range(len(rows)):
            ia = rows[a_pos]
            if ia in dead:
                continue
            for b_pos in range(a_pos + 1, len(rows)):
                ib = rows[b_pos]
                if ib in dead or ia in dead:
                    continue
                la, lb = ia[1], ib[1]
                if _is_protected_cached(la) or _is_protected_cached(lb):
                    continue
                if not _mergeable_names(la, lb):
                    continue
                va, vb = num.loc[ia], num.loc[ib]
                equal, overlap = _rows_equal(va, vb)
                # Near-duplicate: same line, one annual-derived boundary quarter
                # disagrees. Require a clean token-subset name match for safety.
                ta, tb = _tokens_for(la), _tokens_for(lb)
                mostly = (not equal and (ta <= tb or tb <= ta)
                          and _rows_mostly_equal(va, vb))
                both_auto = _is_auto_cached(la) and _is_auto_cached(lb)
                era_stitch = (both_auto and overlap == 0
                              and va.notna().any() and vb.notna().any()
                              and (_tokens_for(la) <= _tokens_for(lb)
                                   or _tokens_for(lb) <= _tokens_for(la)
                                   or len(_tokens_for(la) & _tokens_for(lb))
                                   / max(len(_tokens_for(la) | _tokens_for(lb)), 1) >= 0.6))
                if not (equal or era_stitch or mostly):
                    continue
                surv, src = (ia, ib) if _pref_key(ia) <= _pref_key(ib) else (ib, ia)
                why = 'duplicate' if equal else ('near-duplicate' if mostly else 'era-renamed')
                if src[1] in _bridge_locked and not equal:
                    print(f"  [Cleanup] PROTECTED bridge component from merge: "
                          f"'{src[1]}' (would have merged into '{surv[1]}')")
                    continue
                # If the bridge component IS an exact duplicate, prefer the
                # curated/non-auto label as survivor so the bridge's recorded
                # label keeps pointing at a row that exists.
                if equal and src[1] in _bridge_locked and surv[1] not in _bridge_locked:
                    surv, src = src, surv
                print(f"  [Cleanup] Merging {why} row '{src[1]}' -> '{surv[1]}'")
                fills.append((surv, src))
                dead.add(src)

    for surv, src in fills:
        merged = df.loc[surv].combine_first(df.loc[src])
        df.loc[surv, :] = merged.values
        num.loc[surv, :] = pd.to_numeric(merged, errors='coerce').values

    # ---- Pass 4: segment rows echoing face-statement lines ---------------
    face_rows = [i for i in df.index
                 if i[0] in ('1_Income_Statement', '3_Cash_Flow') and i not in dead]
    for idx in [i for i in df.index if i[0] in SEG_CATS]:
        if idx in dead:
            continue
        parts = idx[1].split(' - ')
        if len(parts) < 2:
            continue
        mem_toks = _tokens_for(' '.join(parts[1:]))
        if not mem_toks:
            continue
        vseg = num.loc[idx]
        seg_nn = vseg.notna()
        for fidx in face_rows:
            ftoks = _tokens_for(fidx[1])
            if not (mem_toks <= ftoks or ftoks <= mem_toks):
                continue
            vface = num.loc[fidx]
            equal, overlap = _rows_equal(vseg, vface, min_overlap=4)
            if not equal:
                continue
            # face row must cover everything the segment row has
            if (seg_nn & ~vface.notna()).any():
                continue
            print(f"  [Cleanup] Dropping segment echo '{idx[1]}' "
                  f"(duplicates face line '{fidx[1]}')")
            dead.add(idx)
            break

    # ---- Pass 5: supplemental cash-paid continuity ------------------------
    # Companies move 'cash paid for interest / income taxes' between the CF
    # face and note disclosures across eras, splitting one series over
    # several labels (and categories). Stitch each topic into one canonical
    # supplemental row; conflicting concurrent components (e.g. interest on
    # debt vs on finance leases) are left separate.
    _SUPP_TOPICS = (
        ('Cash Taxes Paid', ('tax', 'taxe'), ('paid', 'payment')),
        ('Cash Interest Paid', ('interest',), ('paid', 'payment')),
    )
    _BRIDGE_COMPONENT_LABELS = {
        'Dividends Paid', 'Taxes Paid on Stock Awards', 'Share Repurchases',
        'Stock Option Proceeds', 'Total Debt Issued', 'Total Debt Repaid',
        'Other Non-Cash Items',   # OCF bridge component, not a supplement
    }
    # Reuse the numeric view maintained above.  Passes 1-4 only mark dead
    # rows after the initial conversion/fill updates, so a full-frame numeric
    # rebuild here is redundant; stitches below still rebuild after structural
    # changes to keep num aligned with df.
    for _canon, _need_any, _need_pay in _SUPP_TOPICS:
        group = []
        for idx in df.index:
            cat_t, lbl_t = idx
            if idx in dead or cat_t not in ('3_Cash_Flow', '6_Disclosures'):
                continue
            if lbl_t in _BRIDGE_COMPONENT_LABELS:
                continue
            if _is_protected_cached(lbl_t) and lbl_t != _canon:
                continue
            toks = _tokens_for(lbl_t.split(' - ')[-1])
            if any(t in toks for t in _need_any) and any(t in toks for t in _need_pay):
                group.append(idx)
        if len(group) < 2:
            continue
        # Disclosure-only series stay in 6_Disclosures: a stitch (and its
        # relocation to the CF supplemental block) requires at least one
        # CF-face member in the group.
        if not any(i[0] == '3_Cash_Flow' for i in group):
            continue
        # most-recent data first; newest series wins conflicting cells
        group.sort(key=lambda i: next(
            (p for p, c in enumerate(df.columns) if pd.notna(num.loc[i, c])), 9999))
        survivor = num.loc[group[0]].copy()
        absorbed = [group[0]]
        for idx in group[1:]:
            cand = num.loc[idx]
            ov = survivor.notna() & cand.notna()
            n_ov = int(ov.sum())
            if n_ov:
                conflicts = int((~np.isclose(survivor[ov].astype(float),
                                             cand[ov].astype(float),
                                             rtol=1e-4, atol=0.02)).sum())
                # tolerate isolated YTD/Q4-derivation artifacts only
                if conflicts > 0 and (n_ov < 3 or conflicts > max(1, int(0.34 * n_ov))):
                    continue
            survivor = survivor.combine_first(cand)
            absorbed.append(idx)
        if len(absorbed) < 2:
            continue
        print(f"  [Cleanup] Stitched supplemental series '{_canon}' from: "
              + ", ".join(f"'{i[1]}'" for i in absorbed))
        # place the canonical row where the first absorbed CF row sat
        new_idx = ('3_Cash_Flow', _canon)
        if new_idx in df.index and new_idx not in absorbed:
            continue   # canonical row exists but was value-incompatible
        anchor_idx = next((i for i in absorbed if i[0] == '3_Cash_Flow'), absorbed[0])
        dead.update(absorbed)
        new_row = pd.DataFrame([survivor.values], columns=df.columns,
                               index=pd.MultiIndex.from_tuples([new_idx], names=df.index.names))
        keep = [i for i in df.index if i not in dead and i != new_idx]
        pos = 0
        for k, i in enumerate(keep):
            if i == anchor_idx:
                pos = k
                break
        else:
            # anchor itself was absorbed; insert where it used to be
            before = [i for i in df.index if i == anchor_idx or (i not in dead and i != new_idx)]
            pos = before.index(anchor_idx)
        df = pd.concat([df.loc[keep[:pos]], new_row, df.loc[keep[pos:]]])
        dead = {i for i in dead if i in df.index}
        dead.discard(new_idx)   # the freshly stitched row is not a casualty
        if _canon not in CONCEPT_MAP:
            CONCEPT_MAP[_canon] = {'tags': [], 'cat': '3_Cash_Flow'}
            _auto_cache.clear()
            _protected_cache.clear()
        num = df.apply(pd.to_numeric, errors='coerce')

    # ---- Pass 6: abandoned supplemental disclosures -----------------------
    # Supplemental items (cash paid for interest/taxes, noncash exchanges,
    # period-increase echoes) have no calculation-linkbase parent, so the
    # sparse rollup cannot absorb them. Once abandoned -- no data in the four
    # most recent periods -- they are removed outright.
    _SUPP_PHRASES = ('obtained in exchange', 'obligations incurred', 'noncash',
                     'non-cash', 'period increase')
    _recent_cols = list(df.columns[:4])
    for idx in list(df.index):
        if idx in dead or idx[0] != '3_Cash_Flow':
            continue
        lbl_t = idx[1]
        if lbl_t in _BRIDGE_COMPONENT_LABELS or lbl_t in _bridge_locked:
            continue
        if _is_protected_cached(lbl_t) and lbl_t not in ('Cash Interest Paid', 'Cash Taxes Paid'):
            continue
        ll = lbl_t.lower()
        toks = _tokens_for(lbl_t)
        is_supp = (any(p in ll for p in _SUPP_PHRASES)
                   or ('paid' in toks or 'payment' in toks))
        if not is_supp:
            continue
        if num.loc[idx, _recent_cols].notna().any():
            continue
        print(f"  [Cleanup] Dropping abandoned supplemental row '{lbl_t}'")
        dead.add(idx)

    # ---- Pass 7: abandoned learned rows (era debris) ----------------------
    # After era-stitching and duplicate merges have preserved every renamed
    # line's history under its current name, auto-learned core-statement rows
    # with no data in the most recent four periods are era debris (2013
    # gross-netting disclosures, redeemed-NCI mechanics, superseded cash
    # reconciliation lines, ...). They have no calculation-linkbase parent,
    # so the sparse rollup cannot absorb them; per policy they are removed
    # outright. Stale derived bridge rows go with them. Curated rows are
    # never touched here.
    for idx in list(df.index):
        _is_seg_row = idx[0] in SEG_CATS
        if idx in dead or (idx[0] not in ('1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow')
                           and not _is_seg_row):
            continue
        lbl_t = idx[1]
        if lbl_t in _bridge_locked or (isinstance(CONCEPT_MAP.get(lbl_t), dict)
                                       and CONCEPT_MAP[lbl_t].get('is_dyn_opex')):
            continue  # bridge component: keep as granular history
        # Era-tagged segment rows ('(Pre Change)' / '(Post Change)') are
        # intentionally preserved historical segmentation -- stale by
        # design, never debris.
        if '(Pre Change)' in lbl_t or '(Post Change)' in lbl_t:
            continue
        info = CONCEPT_MAP.get(lbl_t)
        is_auto = isinstance(info, dict) and info.get('auto')
        # Segment/dimensional rows are inherently learned (never curated).
        if not (is_auto or _is_protected_cached(lbl_t) or _is_seg_row):
            continue
        row_vals = num.loc[idx]
        if not row_vals.notna().any():
            continue   # fully-empty rows were handled in Pass 1
        if row_vals[_recent_cols].notna().any():
            continue
        # Materiality guard: a stale row at institutional scale (>=5% of
        # peak Total Assets) is more likely a capture gap or an unstitched
        # era rename than abandoned debris -- keep it and flag instead of
        # silently destroying history (e.g. a $35B crypto safeguarding
        # asset whose recent quarters failed to extract).
        _base = None
        # Flow-metric segment rows scale with revenue, not assets.
        _b_order = ((('1_Income_Statement', 'Revenue'), ('2_Balance_Sheet', 'Total Assets'))
                    if _is_seg_row else
                    (('2_Balance_Sheet', 'Total Assets'), ('1_Income_Statement', 'Revenue')))
        for _b_idx in _b_order:
            if _b_idx in num.index:
                _bv = num.loc[_b_idx].abs().max()
                if pd.notna(_bv) and _bv > 0:
                    _base = float(_bv)
                    break
        _peak = float(row_vals.abs().max())
        if _base is None:
            continue   # no scale reference -- be conservative, keep
        if _peak >= 0.05 * _base:
            print(f"  [Cleanup Warning] Stale but material row kept: '{lbl_t}' "
                  f"(peak {_peak / _base:.0%} of company scale) -- possible "
                  f"capture gap or unstitched era rename.")
            continue
        print(f"  [Cleanup] Dropping abandoned learned row '{lbl_t}' "
              f"(peak {_peak / _base:.1%} of company scale)")
        dead.add(idx)

    for _d in dead:
        try:
            if _d[0] == '3_Cash_Flow':
                _info_d = CONCEPT_MAP.get(_d[1])
                if isinstance(_info_d, dict) and not _info_d.get('auto'):
                    print(f"  [Cleanup] NOTE: curated cash-flow row removed: '{_d[1]}'")
        except Exception:
            pass
    if dead:
        df = df.drop(index=[i for i in dead if i in df.index])
    return df


def _save_pivot_xlsx(final_pivot, out_path):
    """Write the final pivot to .xlsx, one sheet per statement category.

    Same values as the CSV (blanks stay blank), but each top-level Category --
    Income Statement, Balance Sheet, Cash Flow, ... -- lands on its own sheet
    with a frozen header, periods ordered oldest-to-newest, the period-ending
    dates repeated at the top, and a
    per-row number format (dollars vs per-share / percentage / flag, applied
    for display only -- the stored value is never rounded). Raises ImportError
    when openpyxl is missing so the caller can fall back to CSV.
    """
    import re as _re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR = '0_Period_Header'
    _dmask = ((final_pivot.index.get_level_values(0) == HDR)
              & (final_pivot.index.get_level_values(1) == 'Period Ending'))
    date_row = final_pivot.loc[_dmask].iloc[0] if _dmask.any() else None

    # XLSX is intentionally chronological (oldest -> newest), while the CSV
    # keeps the existing newest -> oldest order. Prefer the actual period-end
    # date when available, then fall back to labels such as 2025-Q3, FY2025,
    # 2025-FY, or 2025. Unrecognized columns remain stable at the end.
    _original_columns = list(final_pivot.columns)

    def _xlsx_period_sort_key(column):
        if date_row is not None:
            try:
                parsed = pd.to_datetime(date_row.get(column), errors='coerce')
                if pd.notna(parsed):
                    return (0, int(parsed.year), int(parsed.month), int(parsed.day), 0)
            except Exception:
                pass

        label = str(column).strip()
        quarter_match = _re.match(r'^(\d{4})\s*[-_/ ]?\s*Q([1-4])$', label, _re.I)
        if quarter_match:
            year, quarter = map(int, quarter_match.groups())
            return (0, year, quarter * 3, 31, 0)

        annual_match = _re.match(
            r'^(?:FY\s*[-_/ ]?)?(\d{4})(?:\s*[-_/ ]?\s*FY)?$',
            label,
            _re.I,
        )
        if annual_match:
            return (0, int(annual_match.group(1)), 12, 31, 0)

        return (1, 0, 0, 0, _original_columns.index(column))

    _xlsx_columns = sorted(_original_columns, key=_xlsx_period_sort_key)
    final_pivot = final_pivot.loc[:, _xlsx_columns]
    if date_row is not None:
        date_row = date_row.reindex(_xlsx_columns)

    def _friendly(cat):
        return _re.sub(r'^\d+[a-z]?_', '', str(cat)).replace('_', ' ').strip() or str(cat)

    def _sheet_name(name, used):
        nm = (_re.sub(r'[:\\/?*\[\]]', ' ', name).strip()[:31]) or 'Sheet'
        base, i = nm, 2
        while nm.lower() in used:
            suf = ' (%d)' % i
            nm = base[:31 - len(suf)] + suf
            i += 1
        used.add(nm.lower())
        return nm

    def _fmt_for(label):
        l = str(label).lower()
        if any(k in l for k in ('per share', 'eps', 'margin', 'ratio', '%',
                                'yield', 'tax rate', 'multiple', 'turnover')):
            return '#,##0.00;(#,##0.00)'
        return '#,##0;(#,##0)'

    _BOLD = ('Revenue', 'Gross Profit', 'Operating Income', 'Net Interest Income (Expense)',
             'Pretax Income', 'Net Income', 'EBITDA', 'Operating Cash Flow',
             'Investing Cash Flow', 'Financing Cash Flow', 'Net Cash Flow')

    def _is_bold(label):
        l = str(label)
        return l in _BOLD or l.startswith('Total ')

    cats = [c for c in dict.fromkeys(final_pivot.index.get_level_values(0)) if c != HDR]
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    date_font = Font(name='Arial', bold=True, italic=True, color='595959')
    item_font = Font(name='Arial', bold=True)
    val_font = Font(name='Arial')
    edge = Border(bottom=Side(style='thin', color='D9D9D9'))
    imp_border = Border(top=Side(style='thin', color='808080'),
                        bottom=Side(style='thin', color='808080'))

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        used = set()
        for cat in cats:
            sub = final_pivot.xs(cat, level=0, drop_level=True).copy()
            if date_row is not None:
                top = pd.DataFrame([list(date_row.values)],
                                   index=['Period Ending'], columns=sub.columns)
                sub = pd.concat([top, sub])
            sub.index.name = 'Line Item'
            sn = _sheet_name(_friendly(cat), used)
            sub.to_excel(writer, sheet_name=sn, startrow=1, startcol=1)
            ws = writer.sheets[sn]
            ws.sheet_view.showGridLines = False
            R0, C0 = 2, 2                       # table origin -> cell B2
            ncol_d = sub.shape[1]               # period (data) columns
            nrow_d = sub.shape[0]               # line-item (data) rows
            for j in range(C0, C0 + ncol_d + 1):
                c = ws.cell(row=R0, column=j)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal=('left' if j == C0 else 'center'))
                c.border = edge
            for i in range(R0 + 1, R0 + 1 + nrow_d):
                is_date = date_row is not None and i == R0 + 1
                lc = ws.cell(row=i, column=C0)
                lbl = lc.value
                bold = (not is_date) and _is_bold(lbl)
                numfmt = _fmt_for(lbl)
                lc.alignment = Alignment(horizontal='left')
                if is_date:
                    lc.font = item_font
                elif bold:
                    lc.font = item_font
                    lc.border = imp_border
                else:
                    lc.font = val_font
                    lc.value = '   ' + str(lbl)        # indent detail lines
                for j in range(C0 + 1, C0 + 1 + ncol_d):
                    c = ws.cell(row=i, column=j)
                    if is_date:
                        c.font = date_font
                        c.alignment = Alignment(horizontal='center')
                    else:
                        c.font = item_font if bold else val_font
                        c.alignment = Alignment(horizontal='right')
                        if isinstance(c.value, (int, float)) and c.value == c.value:
                            c.number_format = numfmt
                    if bold:
                        c.border = imp_border
            maxlen = max([len('Line Item')] + [len(str(x)) for x in sub.index])
            ws.column_dimensions[get_column_letter(C0)].width = min(max(maxlen + 2, 24), 60)
            for j in range(C0 + 1, C0 + 1 + ncol_d):
                ws.column_dimensions[get_column_letter(j)].width = 18
            ws.column_dimensions['A'].width = 3
            ws.freeze_panes = get_column_letter(C0 + 1) + str(R0 + 1) 


# ###########################################################################
# FOREIGN PRIVATE ISSUER (20-F / 40-F) BRANCH -- annual only
# ---------------------------------------------------------------------------
# Runs ONLY when _fx_is_foreign_20f() is true (20-F/40-F filer, no 10-K).
# No US-pipeline function mutates foreign data; no foreign code runs for US.
#
# Data layers (each degrades gracefully -- "blank over wrong"):
#   L1 annual backbone : SEC companyfacts JSON (face + disclosures, rep. ccy)
#   L2 annual XBRL     : per-filing 20-F/40-F XBRL via fetch_xbrl() -> face,
#                        dimensional segments, concentration facts, ADS-ratio
#                        evidence when it is a structured fact.
#   Output             : annual FY columns only.
# ###########################################################################

_FX_IC, _FX_BS, _FX_CF = "1_Income_Statement", "2_Balance_Sheet", "3_Cash_Flow"
_FX_SEG_PLATFORM = "4a_Segments_Platform"
_FX_SEG_PRODUCT = "4b_Segments_Product_Type"
_FX_SEG_GEO = "4c_Segments_Geographic_Regions"
_FX_SEG_BUS = _FX_SEG_PLATFORM  # backwards-compatible local alias
_FX_KPI, _FX_DIS, _FX_CON, _FX_CHK = ("5_KPI_Metrics", "6_Disclosures",
                                      "7_Concentration_Risk", "8_Integrity_Checks")
_FX_HYBRID_FALLBACK = True   # emit an FY column for years lacking an H1 scrape
_FX_DEBUG = _DEBUG_OUTPUT_ENABLED  # print discovery diagnostics only when SEC_DEBUG=1
_FX_TOL = 0.006              # 0.6% tolerance for sum/agreement checks
_FX_SCALE_TOL = 0.025        # looser: infer table units across restated comparatives


def _fx_env_int(name, default, minimum=1):
    try:
        return max(int(minimum), int(os.environ.get(name, default)))
    except Exception:
        return int(default)


def _fx_env_float(name, default, minimum=1.0):
    try:
        return max(float(minimum), float(os.environ.get(name, default)))
    except Exception:
        return float(default)


_FX_CACHE_VERSION = "2026-07-04.fx-cache.v1"
_FX_CACHE_DISABLED = {"0", "false", "no", "off", "disable", "disabled"}
_FX_PERSISTENT_CACHE_ENABLED = (
    os.environ.get("SEC_FX_CACHE", "1").strip().lower()
    not in _FX_CACHE_DISABLED
)
_FX_PERSISTENT_CACHE_ROOT = os.path.abspath(
    os.environ.get("SEC_FX_CACHE_DIR", _script_cache_dir("fx"))
)
_FX_CACHE_MISS = object()
_FX_PERSISTENT_CACHE_LOCK = threading.RLock()
_FX_VISIBLE_TEXT_CACHE = {}
_FX_VISIBLE_TEXT_CACHE_MAX = _fx_env_int("SEC_FX_TEXT_CACHE_SIZE", 512, 1)


def _fx_hash_blob(value):
    if isinstance(value, (bytes, bytearray)):
        data = bytes(value)
    else:
        data = str(value).encode("utf-8", "surrogatepass")
    return hashlib.sha256(data).hexdigest(), len(data)


@lru_cache(maxsize=1)
def _fx_code_fingerprint():
    try:
        with open(__file__, "rb") as fh:
            return hashlib.sha256(fh.read()).hexdigest()
    except Exception:
        return _FX_CACHE_VERSION


def _fx_freeze_for_cache(value):
    if isinstance(value, dict):
        return tuple(sorted(
            (str(k), _fx_freeze_for_cache(v)) for k, v in value.items()
        ))
    if isinstance(value, (list, tuple)):
        return tuple(_fx_freeze_for_cache(v) for v in value)
    if isinstance(value, (set, frozenset)):
        return tuple(sorted(_fx_freeze_for_cache(v) for v in value))
    if isinstance(value, (str, int, float, bool, type(None))):
        return value
    return str(value)


def _fx_cache_path(bucket, key):
    if not _FX_PERSISTENT_CACHE_ENABLED:
        return None
    try:
        key_blob = pickle.dumps((_FX_CACHE_VERSION, bucket, key), protocol=4)
    except Exception:
        key_blob = repr((_FX_CACHE_VERSION, bucket, key)).encode("utf-8", "replace")
    digest = hashlib.sha256(key_blob).hexdigest()
    return os.path.join(_FX_PERSISTENT_CACHE_ROOT, bucket, digest[:2], f"{digest}.pkl")


def _fx_cache_get(bucket, key):
    path = _fx_cache_path(bucket, key)
    if not path:
        return _FX_CACHE_MISS
    try:
        with _FX_PERSISTENT_CACHE_LOCK:
            if not os.path.exists(path):
                return _FX_CACHE_MISS
            with open(path, "rb") as fh:
                payload = pickle.load(fh)
        if not isinstance(payload, dict) or payload.get("version") != _FX_CACHE_VERSION:
            return _FX_CACHE_MISS
        return payload.get("value")
    except Exception:
        return _FX_CACHE_MISS


def _fx_cache_set(bucket, key, value):
    path = _fx_cache_path(bucket, key)
    if not path:
        return
    tmp = f"{path}.{os.getpid()}.{threading.get_ident()}.tmp"
    try:
        with _FX_PERSISTENT_CACHE_LOCK:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(tmp, "wb") as fh:
                pickle.dump(
                    {"version": _FX_CACHE_VERSION, "value": value},
                    fh,
                    protocol=pickle.HIGHEST_PROTOCOL,
                )
            os.replace(tmp, path)
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


def _fx_empty_diagnostics():
    return {"observations": [], "selected": [], "rejections": []}

# ---- face map (IFRS) ------------------------------------------------------
_FX_IFRS_MAP = {
    "Revenue":                       (_FX_IC, ["Revenue", "RevenueFromContractsWithCustomers"]),
    "Cost of Revenue":               (_FX_IC, ["CostOfSales"]),
    "Gross Profit":                  (_FX_IC, ["GrossProfit"]),
    "Research & Development":         (_FX_IC, ["ResearchAndDevelopmentExpense"]),
    "General & Administrative":       (_FX_IC, ["GeneralAndAdministrativeExpense"]),
    "Sales & Marketing":              (_FX_IC, ["SalesAndMarketingExpense", "DistributionCosts"]),
    "Other Operating Income/Expense": (_FX_IC, ["OtherOperatingIncomeExpense"]),
    "Operating Income":               (_FX_IC, ["ProfitLossFromOperatingActivities"]),
    "Equity Method Income":           (_FX_IC, ["ShareOfProfitLossOfAssociatesAndJointVenturesAccountedForUsingEquityMethod",
                                                "ShareOfProfitLossOfAssociatesAccountedForUsingEquityMethod"]),
    "Interest Income":                (_FX_IC, [
        "FinanceIncome", "RevenueFromInterest", "InterestIncome",
        "InterestRevenue", "InterestRevenueCalculatedUsingEffectiveInterestMethod",
        "InterestIncomeExpenseNoninsuranceActivities",
    ]),
    "Finance Costs":                  (_FX_IC, [
        "FinanceCosts", "InterestExpense", "InterestExpenseOnBorrowings",
        "InterestExpenseNoninsuranceActivities",
    ]),
    "Other Income/(Expense)":         (_FX_IC, [
        "OtherGainsLosses", "OtherIncomeExpense",
        "OtherNonOperatingIncomeExpense", "OtherNonOperatingIncomeExpenses",
        "OtherNonoperatingIncomeExpense", "OtherNonoperatingIncomeExpenses",
    ]),
    "Total Non-operating Income":      (_FX_IC, [
        "NonOperatingIncomeExpense", "NonoperatingIncomeExpense",
        "TotalNonOperatingIncomeExpense", "TotalNonoperatingIncomeExpense",
        "TotalNonOperatingIncomeExpenses", "TotalNonoperatingIncomeExpenses",
        "FinanceIncomeCostsAndOtherNonOperatingIncomeExpenses",
        "FinanceIncomeCostsAndOtherNonoperatingIncomeExpenses",
    ]),
    "Pretax Income":                  (_FX_IC, ["ProfitLossBeforeTax", "AccountingProfit"]),
    "Income Tax Expense":             (_FX_IC, ["IncomeTaxExpenseContinuingOperations"]),
    "Net Income":                     (_FX_IC, ["ProfitLoss"]),
    "Net Income to NCI":              (_FX_IC, ["ProfitLossAttributableToNoncontrollingInterests"]),
    "Net Income to Parent":           (_FX_IC, ["ProfitLossAttributableToOwnersOfParent"]),
    "EPS Basic":                      (_FX_IC, ["BasicEarningsLossPerShare",
                                                    "BasicEarningsLossPerShareFromContinuingOperations"]),
    "EPS Diluted":                    (_FX_IC, ["DilutedEarningsLossPerShare",
                                                    "DilutedEarningsLossPerShareFromContinuingOperations"]),
    "Shares Outstanding Basic":       (_FX_IC, ["WeightedAverageShares",
                                                    "WeightedAverageNumberOfOrdinarySharesOutstanding",
                                                    "WeightedAverageNumberOfSharesOutstanding"]),
    "Shares Outstanding Diluted":     (_FX_IC, ["AdjustedWeightedAverageShares",
                                                    "AdjustedWeightedAverageNumberOfShares",
                                                    "AdjustedWeightedAverageNumberOfOrdinarySharesOutstanding",
                                                    "WeightedAverageNumberOfDilutedSharesOutstanding",
                                                    "WeightedAverageShares"]),
    "Cash & Equivalents":             (_FX_BS, ["CashAndCashEquivalents"]),
    "Inventories":                    (_FX_BS, ["Inventories"]),
    "Trade Receivables":              (_FX_BS, ["CurrentTradeReceivables", "TradeAndOtherCurrentReceivables"]),
    "Total Current Assets":           (_FX_BS, ["CurrentAssets"]),
    "Property, Plant & Equipment":    (_FX_BS, ["PropertyPlantAndEquipment"]),
    "Right-of-Use Assets":            (_FX_BS, ["RightofuseAssets"]),
    "Intangibles & Goodwill":         (_FX_BS, ["IntangibleAssetsAndGoodwill"]),
    "Total Non-current Assets":       (_FX_BS, ["NoncurrentAssets"]),
    "Total Assets":                   (_FX_BS, ["Assets"]),
    "Accounts Payable":               (_FX_BS, ["TradeAndOtherCurrentPayables", "TradeAndOtherCurrentPayablesToTradeSuppliers"]),
    "Contract Liabilities":           (_FX_BS, ["ContractLiabilities"]),
    "Current Lease Liabilities":      (_FX_BS, ["CurrentLeaseLiabilities"]),
    "Total Current Liabilities":      (_FX_BS, ["CurrentLiabilities"]),
    "Noncurrent Lease Liabilities":   (_FX_BS, ["NoncurrentLeaseLiabilities"]),
    "Total Non-current Liabilities":  (_FX_BS, ["NoncurrentLiabilities"]),
    "Total Liabilities":              (_FX_BS, ["Liabilities"]),
    "Short-term Debt":                (_FX_BS, ["ShorttermBorrowings", "CurrentPortionOfLongtermBorrowings"]),
    "Long-term Debt":                 (_FX_BS, ["LongtermBorrowings", "NoncurrentPortionOfNoncurrentBondsIssued", "BondsIssued"]),
    "Retained Earnings":              (_FX_BS, ["RetainedEarnings"]),
    "Equity to Parent":               (_FX_BS, ["EquityAttributableToOwnersOfParent"]),
    "Noncontrolling Interest":        (_FX_BS, ["NoncontrollingInterests"]),
    "Total Equity":                   (_FX_BS, ["Equity"]),
    "Operating Cash Flow":            (_FX_CF, ["CashFlowsFromUsedInOperatingActivities"]),
    "Investing Cash Flow":            (_FX_CF, ["CashFlowsFromUsedInInvestingActivities"]),
    "Financing Cash Flow":            (_FX_CF, ["CashFlowsFromUsedInFinancingActivities"]),
    "Capital Expenditures":           (_FX_CF, ["PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities"]),
    "Depreciation":                   (_FX_CF, ["DepreciationExpense"]),
    "Amortization":                   (_FX_CF, ["AmortisationExpense"]),
    "Dividends Paid":                 (_FX_CF, ["DividendsPaidClassifiedAsFinancingActivities", "DividendsPaid"]),
    "Effect of FX on Cash":           (_FX_CF, ["EffectOfExchangeRateChangesOnCashAndCashEquivalents"]),
    "Net Cash Flow":                  (_FX_CF, ["IncreaseDecreaseInCashAndCashEquivalents"]),
}

# ---- disclosures map (IFRS).  kind: 'flow' (duration) or 'stock' (instant).
_FX_IFRS_DISCL = {
    "Stock-Based Compensation":            (["AdjustmentsForSharebasedPayments"], "flow"),
    "Depreciation - Right-of-Use Assets":  (["DepreciationRightofuseAssets"], "flow"),
    "Income Taxes Paid":                   (["IncomeTaxesPaidClassifiedAsOperatingActivities", "IncomeTaxesPaidRefundClassifiedAsOperatingActivities"], "flow"),
    "Interest Paid":                       (["InterestPaidClassifiedAsFinancingActivities", "InterestPaidClassifiedAsOperatingActivities"], "flow"),
    "Interest Received":                   (["InterestReceivedClassifiedAsInvestingActivities", "InterestReceivedClassifiedAsOperatingActivities"], "flow"),
    "Dividends Received":                  (["DividendsReceivedClassifiedAsInvestingActivities"], "flow"),
    "Employee Benefits Expense":           (["EmployeeBenefitsExpense"], "flow"),
    "Purchase of Intangible Assets":       (["PurchaseOfIntangibleAssetsClassifiedAsInvestingActivities"], "flow"),
    "Additions to Right-of-Use Assets":    (["AdditionsToRightofuseAssets"], "flow"),
    "Cash Outflow for Leases":             (["CashOutflowForLeases"], "flow"),
    "Impairment - Financial Assets":       (["ImpairmentLossOnFinancialAssets"], "flow"),
    "Inventory Write-down":                (["InventoryWritedown2011", "InventoryWritedown"], "flow"),
    "Government Grants Revenue":           (["RevenueFromGovernmentGrants"], "flow"),
    "Proceeds from Borrowings":            (["ProceedsFromBorrowingsClassifiedAsFinancingActivities"], "flow"),
    "Repayments of Borrowings":            (["RepaymentsOfBorrowingsClassifiedAsFinancingActivities"], "flow"),
    "Interest Expense on Borrowings":      (["InterestExpenseOnBorrowings"], "flow"),
}

_FX_PER_SHARE = {"EPS Basic", "EPS Diluted"}
_FX_SHARES = {"Shares Outstanding Basic", "Shares Outstanding Diluted"}
_FX_NON_ADDITIVE = _FX_PER_SHARE | _FX_SHARES
_FX_FLOW_CATS = {_FX_IC, _FX_CF}
_FX_NEGATIVE_FLOW_LABELS = {
    "Capital Expenditures",
    "Dividends Paid",
    "Purchase of Intangible Assets",
    "Repayments of Borrowings",
    "Cash Outflow for Leases",
}
_FX_CCY = {"TWD", "USD", "EUR", "JPY", "KRW", "GBP", "CNY", "HKD", "CHF", "SEK",
           "DKK", "NOK", "AUD", "CAD", "SGD", "ILS", "INR", "BRL", "MXN", "ZAR",
           "TRY", "PLN", "RUB", "IDR", "THB", "MYR", "PHP", "NZD", "CLP", "ARS"}

# canonical in-category row order for presentation
_FX_ORDER = {
    _FX_IC: ["Revenue", "Cost of Revenue", "Gross Profit", "Research & Development",
             "Sales & Marketing", "General & Administrative", "Total Operating Expenses",
             "Other Operating Income/Expense", "Operating Income", "Equity Method Income",
             "Interest Income", "Finance Costs", "Other Income/(Expense)",
             "Total Non-operating Income", "Pretax Income",
             "Income Tax Expense", "Net Income", "Net Income to Parent", "Net Income to NCI",
             "EPS Basic", "EPS Diluted", "Shares Outstanding Basic", "Shares Outstanding Diluted"],
    _FX_BS: ["Cash & Equivalents", "Trade Receivables", "Inventories", "Total Current Assets",
             "Property, Plant & Equipment", "Right-of-Use Assets", "Intangibles & Goodwill",
             "Total Non-current Assets", "Total Assets", "Accounts Payable", "Contract Liabilities",
             "Short-term Debt", "Current Lease Liabilities", "Total Current Liabilities",
             "Long-term Debt", "Noncurrent Lease Liabilities", "Total Non-current Liabilities",
             "Total Liabilities", "Retained Earnings", "Equity to Parent",
             "Noncontrolling Interest", "Total Equity"],
    _FX_CF: ["Pretax Income", "Net Income", "Depreciation & Amortization",
             "Share-Based Compensation", "Other Adjustments",
             "Change in Receivables", "Changes in Inventories",
             "Changes in Accounts Payable", "Changes in Accrued Expenses",
             "Changes in Income Taxes Payable",
             "Changes in Other Operating Activities", "Operating Cash Flow",
             "Capital Expenditures", "Sale of Property, Plant & Equipment",
             "Purchases of Intangible Assets",
             "Proceeds from Sale of Intangible Assets",
             "Purchases of Investments", "Proceeds from Sale of Investments",
             "Proceeds from Business Divestments", "Other Investing Activities",
             "Investing Cash Flow", "Short-Term Debt Issued",
             "Short-Term Debt Repaid", "Net Short-Term Debt Issued (Repaid)",
             "Long-Term Debt Issued", "Long-Term Debt Repaid",
             "Net Long-Term Debt Issued (Repaid)", "Issuance of Common Stock",
             "Repurchase of Common Stock", "Net Common Stock Issued (Repurchased)",
             "Common Dividends Paid", "Other Financing Activities",
             "Financing Cash Flow",
             "Effect of Exchange Rate Changes on Cash and Cash Equivalents",
             "Net Cash Flow", "Free Cash Flow"],
    _FX_KPI: KPI_ORDER,
}



# ---------------------------------------------------------------------------
# Annual-only foreign helper definitions retained from the original FPI path.
# These are required by the 20-F/40-F annual companyfacts/XBRL extractor.
# They do not discover or parse interim furnished reports.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class FilingProfile:
    pipeline: str
    has_10k: bool
    has_20f: bool
    has_40f: bool
    fiscal_year_end_month: int | None
    reason: str


@lru_cache(maxsize=65536, typed=True)
def _fx_norm(s):
    t = re.sub(r"\(Notes?[^)]*\)", "", str(s))
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t.strip(" :.-\u2013\u2014")


@lru_cache(maxsize=65536, typed=True)
def _fx_concept_words(concept):
    local = _fx_local(concept)
    s2 = re.sub(r"[_\-]+", " ", local)
    s2 = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", s2)
    s2 = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", s2)
    return re.sub(r"\s+", " ", s2).strip()


@lru_cache(maxsize=65536, typed=True)
def _fx_concept_compact(concept):
    return re.sub(r"[^a-z0-9]+", "", _fx_local(concept).lower())


def _fx_build_concept_alias():
    alias = {}
    try:
        for label, spec in (globals().get("CONCEPT_MAP") or {}).items():
            cat = spec.get("cat")
            for concept in spec.get("tags", []) or []:
                if concept:
                    alias.setdefault(re.sub(r"[^a-z0-9]+", "", str(concept).lower()), (cat, label))
    except Exception:
        pass
    try:
        for label, (cat, concepts) in (globals().get("_FX_IFRS_MAP") or {}).items():
            for concept in concepts or []:
                alias.setdefault(re.sub(r"[^a-z0-9]+", "", str(concept).lower()), (cat, label))
    except Exception:
        pass
    try:
        for label, (concepts, _kind) in (globals().get("_FX_IFRS_DISCL") or {}).items():
            for concept in concepts or []:
                alias.setdefault(re.sub(r"[^a-z0-9]+", "", str(concept).lower()), (_FX_DIS, label))
    except Exception:
        pass
    alias.update({
        "revenuefromcontractswithcustomers": (_FX_IC, "Revenue"),
        "revenues": (_FX_IC, "Revenue"),
        "salesrevenuenet": (_FX_IC, "Revenue"),
        "costofsales": (_FX_IC, "Cost of Revenue"),
        "costofrevenue": (_FX_IC, "Cost of Revenue"),
        "grossprofit": (_FX_IC, "Gross Profit"),
        "researchanddevelopmentexpense": (_FX_IC, "Research & Development"),
        "generalandadministrativeexpense": (_FX_IC, "General & Administrative"),
        "salesandmarketingexpense": (_FX_IC, "Sales & Marketing"),
        "sellinggeneraladministrativeexpense": (_FX_IC, "Selling, General & Admin"),
        "operatingincomeloss": (_FX_IC, "Operating Income"),
        "profitlossfromoperatingactivities": (_FX_IC, "Operating Income"),
        "financeincome": (_FX_IC, "Interest Income"),
        "interestincome": (_FX_IC, "Interest Income"),
        "financecosts": (_FX_IC, "Finance Costs"),
        "interestexpense": (_FX_IC, "Finance Costs"),
        "profitlossbeforetax": (_FX_IC, "Pretax Income"),
        "incomelossfromcontinuingoperationsbeforeincometaxes": (_FX_IC, "Pretax Income"),
        "incometaxexpensebenefit": (_FX_IC, "Income Tax Expense"),
        "profitloss": (_FX_IC, "Net Income"),
        "netincomeloss": (_FX_IC, "Net Income"),
        "profitlossattributabletoownersofparent": (_FX_IC, "Net Income to Parent"),
        "basicearningslosspershare": (_FX_IC, "EPS Basic"),
        "dilutedearningslosspershare": (_FX_IC, "EPS Diluted"),
        "weightedaveragenumberofsharesoutstandingbasic": (_FX_IC, "Shares Outstanding Basic"),
        "weightedaveragenumberofdilutedsharesoutstanding": (_FX_IC, "Shares Outstanding Diluted"),
        "cashandcashequivalents": (_FX_BS, "Cash & Equivalents"),
        "cashandcashequivalentsatcarryingvalue": (_FX_BS, "Cash & Equivalents"),
        "assets": (_FX_BS, "Total Assets"),
        "liabilities": (_FX_BS, "Total Liabilities"),
        "equity": (_FX_BS, "Total Equity"),
        "stockholdersequity": (_FX_BS, "Total Equity"),
        "netcashprovidedbyusedinoperatingactivities": (_FX_CF, "Operating Cash Flow"),
        "cashflowsfromusedinoperatingactivities": (_FX_CF, "Operating Cash Flow"),
        "netcashprovidedbyusedininvestingactivities": (_FX_CF, "Investing Cash Flow"),
        "cashflowsfromusedininvestingactivities": (_FX_CF, "Investing Cash Flow"),
        "netcashprovidedbyusedinfinancingactivities": (_FX_CF, "Financing Cash Flow"),
        "cashflowsfromusedinfinancingactivities": (_FX_CF, "Financing Cash Flow"),
    })
    return alias


_FX_CONCEPT_ALIAS = _fx_build_concept_alias()


def _fx_build_exact_label_map():
    exact = {}
    for cat, labels in (globals().get("_FX_ORDER") or {}).items():
        for label in labels:
            exact[_fx_norm(label)] = (cat, label)
    exact.update({
        "revenue": (_FX_IC, "Revenue"),
        "revenues": (_FX_IC, "Revenue"),
        "net sales": (_FX_IC, "Revenue"),
        "total revenue": (_FX_IC, "Revenue"),
        "cost of sales": (_FX_IC, "Cost of Revenue"),
        "gross profit": (_FX_IC, "Gross Profit"),
        "research and development": (_FX_IC, "Research & Development"),
        "selling general and administrative": (_FX_IC, "Selling, General & Admin"),
        "operating income": (_FX_IC, "Operating Income"),
        "income from operations": (_FX_IC, "Operating Income"),
        "income before income taxes": (_FX_IC, "Pretax Income"),
        "profit before tax": (_FX_IC, "Pretax Income"),
        "income tax expense": (_FX_IC, "Income Tax Expense"),
        "net income": (_FX_IC, "Net Income"),
        "profit loss": (_FX_IC, "Net Income"),
        "basic earnings per share": (_FX_IC, "EPS Basic"),
        "diluted earnings per share": (_FX_IC, "EPS Diluted"),
        "total assets": (_FX_BS, "Total Assets"),
        "total liabilities": (_FX_BS, "Total Liabilities"),
        "total equity": (_FX_BS, "Total Equity"),
        "net cash provided by operating activities": (_FX_CF, "Operating Cash Flow"),
        "net cash used in operating activities": (_FX_CF, "Operating Cash Flow"),
        "net cash provided by investing activities": (_FX_CF, "Investing Cash Flow"),
        "net cash used in investing activities": (_FX_CF, "Investing Cash Flow"),
        "net cash provided by financing activities": (_FX_CF, "Financing Cash Flow"),
        "net cash used in financing activities": (_FX_CF, "Financing Cash Flow"),
    })
    return exact


_FX_EXACT = _fx_build_exact_label_map()


_FX_REGEX = [
    (r"^basic(?: earnings| earnings loss| loss)?(?: per)? (?:ordinary )?share", (_FX_IC, "EPS Basic")),
    (r"^diluted(?: earnings| earnings loss| loss)?(?: per)? (?:ordinary )?share", (_FX_IC, "EPS Diluted")),
    (r"^weighted average (?:number of )?(?:ordinary |common )?shares.*basic", (_FX_IC, "Shares Outstanding Basic")),
    (r"^weighted average (?:number of )?(?:ordinary |common )?shares.*diluted", (_FX_IC, "Shares Outstanding Diluted")),
    (r"^net cash (generated by|provided by|used in) operating", (_FX_CF, "Operating Cash Flow")),
    (r"^net cash (generated by|provided by|used in) investing", (_FX_CF, "Investing Cash Flow")),
    (r"^net cash (generated by|provided by|used in) financing", (_FX_CF, "Financing Cash Flow")),
    (r"^net (increase|decrease).{0,40}in cash", (_FX_CF, "Net Cash Flow")),
    (r"^acquisitions? of property, plant and equipment", (_FX_CF, "Capital Expenditures")),
    (r"^(purchases? of|additions to) property, plant and equipment", (_FX_CF, "Capital Expenditures")),
    (r"^cash dividends", (_FX_CF, "Dividends Paid")),
    (r"^dividends paid", (_FX_CF, "Dividends Paid")),
    (r"^effects? of exchange rate", (_FX_CF, "Effect of FX on Cash")),
]
_FX_REGEX_COMPILED = [(re.compile(pat), tgt) for pat, tgt in _FX_REGEX]


_FX_FOREIGN_ANNUAL_EXTRA_CONCEPTS = {
    (_FX_IC, "Pretax Income"): [
        # Common consolidated US-GAAP concept used by some FPIs (ARM).
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxes",
    ],
    (_FX_IC, "Income Tax Expense"): [
        # Keep only the consolidated total tax concept for the annual face.
        "IncomeTaxExpenseBenefit",
        "IncomeTaxExpenseBenefitContinuingOperations",
    ],
}

_FX_FOREIGN_ANNUAL_COMPONENT_CONCEPT_BLOCKLIST = {
    (_FX_IC, "Pretax Income"): {
        # Geographic tax-note components are not consolidated pretax income.
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesForeign",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesDomestic",
    },
    (_FX_IC, "Income Tax Expense"): {
        # Current/deferred tax-note components are not the total tax line.
        "CurrentIncomeTaxExpenseBenefit",
        "CurrentIncomeTaxExpenseBenefitContinuingOperations",
        "DeferredIncomeTaxExpenseBenefit",
        "DeferredIncomeTaxExpenseBenefitContinuingOperations",
    },
}


def _fx_foreign_annual_component_blocked(concept, tgt, human_label=None):
    """Reject tax-note component concepts when building consolidated annual rows.

    This is intentionally scoped to the foreign annual path.  The global
    CONCEPT_MAP remains untouched so the native 10-K/10-Q route is not changed.
    """
    if not concept or not tgt:
        return False
    cat, label = tgt
    local = _fx_local(concept)
    blocked = _FX_FOREIGN_ANNUAL_COMPONENT_CONCEPT_BLOCKLIST.get((cat, label), set())
    if local in blocked:
        return True
    compact = _fx_concept_compact(local)
    norm_label = _fx_norm(human_label or "")
    if (cat, label) == (_FX_IC, "Pretax Income"):
        if (
            local.endswith("Foreign")
            or local.endswith("Domestic")
            or "BeforeIncomeTaxesForeign" in local
            or "BeforeIncomeTaxesDomestic" in local
        ):
            return True
    if (cat, label) == (_FX_IC, "Income Tax Expense"):
        if (
            compact.startswith("currentincometax")
            or compact.startswith("deferredincometax")
            or norm_label.startswith("current ")
            or norm_label.startswith("deferred ")
            or " current tax " in f" {norm_label} "
            or " deferred tax " in f" {norm_label} "
        ):
            return True
    return False



@lru_cache(maxsize=65536, typed=True)
def _fx_map_label(nl, section=None):
    if nl in _FX_EXACT:
        return _FX_EXACT[nl]
    for pat, tgt in _FX_REGEX_COMPILED:
        if pat.search(nl):
            return tgt
    if nl == "total" and section == "assets":
        return (_FX_BS, "Total Assets")
    return None


def _fx_label_from_concept_name(concept, label=None, allowed_cats=None):
    """Map custom IFRS/foreign concept names or human labels to a canonical row."""
    compact = _fx_concept_compact(concept)
    tgt = _FX_CONCEPT_ALIAS.get(compact)
    if tgt and (allowed_cats is None or tgt[0] in allowed_cats):
        if not _fx_foreign_annual_component_blocked(concept, tgt, label):
            return tgt

    texts = []
    if label:
        texts.append(label)
    words = _fx_concept_words(concept)
    if words:
        texts.append(words)
    for text in texts:
        tgt = _fx_map_label(_fx_norm(text), None)
        if tgt and (allowed_cats is None or tgt[0] in allowed_cats):
            return tgt
        compact_text = re.sub(r"[^a-z0-9]+", "", str(text).lower())
        tgt = _FX_CONCEPT_ALIAS.get(compact_text)
        if tgt and (allowed_cats is None or tgt[0] in allowed_cats):
            if not _fx_foreign_annual_component_blocked(concept, tgt, label):
                return tgt
    return None


def _fx_fiscal_quarter_end(fy, quarter, ye_month=12):
    """Return the calendar end date for fiscal quarter `quarter` of fiscal year `fy`."""
    fy, quarter, ye_month = int(fy), int(quarter), int(ye_month or 12)
    month = ((ye_month - (4 - quarter) * 3 - 1) % 12) + 1
    year = fy if month <= ye_month else fy - 1
    day = int(pd.Period(f"{year:04d}-{month:02d}").days_in_month)
    return f"{year:04d}-{month:02d}-{day:02d}"


def _fx_segment_member_cat(label, current_cat=None):
    """Conservative annual-XBRL segment member classifier."""
    n = _fx_norm_member(label, None)
    if not n:
        return current_cat
    geo_terms = {
        "united states", "united kingdom", "taiwan", "china", "japan", "korea",
        "south korea", "hong kong", "singapore", "germany", "france", "netherlands",
        "europe", "asia", "north america", "south america", "latin america",
        "middle east", "asia pacific", "greater china", "rest of world",
        "rest of the world", "domestic", "international",
    }
    if n in geo_terms or re.search(r"\b(?:country|countries|geograph\w*|region|regions)\b", n):
        return _FX_SEG_GEO
    if n in {"wafer", "wafers", "product", "products", "goods"} or re.search(r"\b(?:wafer|product type|other products?)\b", n):
        return _FX_SEG_PRODUCT
    return current_cat


# Display filter for final annual-only foreign output.
_FX_DISPLAY_KEEP = {
    _FX_IC: set(_FX_ORDER.get(_FX_IC, [])) | {
        "Selling, General & Admin", "Other Operating Expenses",
        "Other Income/(Expense)", "Finance Costs", "Interest Income",
    },
    _FX_BS: set(_FX_ORDER.get(_FX_BS, [])) | {
        "Cash & Equivalents", "Total Assets", "Total Liabilities", "Total Equity",
    },
    _FX_CF: set(_FX_ORDER.get(_FX_CF, [])) | {
        "Pretax Income", "Depreciation", "Amortization", "Dividends Paid",
        "Effect of FX on Cash", "Free Cash Flow",
    },
}


def _fx_normalize_value(cat, label, val):
    if val is None:
        return None
    try:
        v = float(val)
    except (TypeError, ValueError):
        return None
    if cat == _FX_IC and label in {"Finance Costs", "Interest Expense"}:
        return -abs(v)
    if cat == _FX_CF and label in _FX_NEGATIVE_FLOW_LABELS:
        return -abs(v)
    return v

def _fx_has_filings(company, form):
    try:
        return len(get_company_filings(company, form).head(1)) > 0
    except Exception:
        return False


def _fx_latest_filing_period_month(company, forms):
    for form in forms:
        try:
            filings = get_company_filings(company, form)
            if len(filings) <= 0:
                continue
            for f in list(filings.head(3)):
                por = getattr(f, "period_of_report", None)
                if por:
                    return int(pd.to_datetime(por).month)
        except Exception:
            continue
    return None


def _fx_profile(company, filings_10k):
    has_10k = bool(filings_10k is not None and len(filings_10k) > 0)
    has_20f = False if has_10k else _fx_has_filings(company, "20-F")
    has_40f = False if has_10k or has_20f else _fx_has_filings(company, "40-F")
    if has_10k:
        ye_month = None
        try:
            ye_month = int(pd.to_datetime(filings_10k[0].period_of_report).month)
        except Exception:
            pass
        return FilingProfile("US_NATIVE", True, has_20f, has_40f, ye_month,
                             "10-K filings present; native U.S. route has precedence")
    if has_20f:
        ye_month = _fx_latest_filing_period_month(company, ("20-F", "20-F/A"))
        return FilingProfile("FOREIGN_20F", False, True, False, ye_month,
                             "20-F filings present and no 10-K filings")
    if has_40f:
        ye_month = _fx_latest_filing_period_month(company, ("40-F", "40-F/A"))
        return FilingProfile("FOREIGN_40F", False, False, True, ye_month,
                             "40-F filings present and no 10-K/20-F filings")
    return FilingProfile("UNSUPPORTED", False, False, False, None,
                         "No supported annual filing profile found")


def _fx_is_foreign_20f(company, filings_10k):
    """Compatibility wrapper: foreign private issuer only when no 10-K exists."""
    return _fx_profile(company, filings_10k).pipeline in ("FOREIGN_20F", "FOREIGN_40F")


# ===========================================================================
# L1 -- companyfacts backbone (face + disclosures)
# ===========================================================================
def _fx_foreign_filing_context(company, limit):
    """Stable annual filing identity for safe foreign persistent-cache keys."""
    try:
        n_filings = max(1, min(int(limit or 1), 10))
    except Exception:
        n_filings = 1
    filings = []
    try:
        # Include amendments in the invalidation context so companyfacts cache
        # refreshes after a 20-F/A or 40-F/A even though the annual XBRL
        # enrichment loop itself preserves the existing 20-F/40-F selection.
        filings = list(get_company_filings(company, ["20-F", "20-F/A"]).head(n_filings + 4))
    except Exception:
        try:
            filings = list(get_company_filings(company, "20-F").head(n_filings))
        except Exception:
            filings = []
    if not filings:
        try:
            filings = list(get_company_filings(company, ["40-F", "40-F/A"]).head(n_filings + 4))
        except Exception:
            try:
                filings = list(get_company_filings(company, "40-F").head(n_filings))
            except Exception:
                filings = []
    out = []
    for filing in filings:
        ident = _stable_filing_identity_for_persistent_cache(filing)
        if ident is None:
            return None
        out.append(ident)
    return tuple(out) if out else None


def _fx_fetch_companyfacts(cik, cache_context=None):
    import json as _json
    cache_key = None
    if cache_context is not None:
        cache_key = _fx_freeze_for_cache({
            "source_fingerprint": _fx_code_fingerprint(),
            "cik": f"{int(cik):010d}",
            "context": cache_context,
        })
        cached = _fx_cache_get("companyfacts", cache_key)
        if cached is not _FX_CACHE_MISS:
            _profile_count("fx_companyfacts_cache_hits")
            return cached
    _profile_count("fx_companyfacts_cache_misses")
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{int(cik):010d}.json"
    hdrs = {"User-Agent": _sec_user_agent(), "Accept-Encoding": "gzip, deflate"}
    sec_limiter.wait()
    resp = _SHARED_HTTP_CLIENT.get(url, headers=hdrs, timeout=60.0, follow_redirects=True)
    resp.raise_for_status()
    facts = resp.json()
    if cache_key is not None:
        _fx_cache_set("companyfacts", cache_key, facts)
    return facts


def _fx_detect_currency(facts_json):
    root = facts_json.get("facts", {})
    counts = {}
    for taxo in ("ifrs-full", "us-gaap"):
        for cd in root.get(taxo, {}).values():
            for unit, arr in cd.get("units", {}).items():
                if unit in _FX_CCY:
                    counts[unit] = counts.get(unit, 0) + len(arr)
    return max(counts, key=counts.get) if counts else None


def _fx_months(start, end):
    if not start:
        return None
    sy, sm, sd = (int(x) for x in start.split("-"))
    ey, em, ed = (int(x) for x in end.split("-"))
    return round((ey - sy) * 12 + (em - sm) + (ed - sd) / 30.0)


def _fx_dedup_latest(records):
    best = {}
    for r in records:
        key = (r.get("start"), r.get("end"))
        cur = best.get(key)
        if cur is None or (r.get("filed", "") > cur.get("filed", "")):
            best[key] = r
    return list(best.values())


def _fx_facts_for(taxo, concepts, ccy):
    """(concept, deduped records, unit): reporting ccy -> ccy/shares -> any
    /shares -> shares -> anything."""
    candidates = []
    for c in concepts:
        cd = taxo.get(c)
        if not cd:
            continue
        units = cd.get("units", {})
        prefs = [ccy, f"{ccy}/shares"]
        prefs += [u for u in units if u.endswith("/shares")]
        if "shares" in units:
            prefs.append("shares")
        prefs += list(units)
        seen_units = []
        for unit in prefs:
            if unit in seen_units:
                continue
            seen_units.append(unit)
            if unit and unit in units and units[unit]:
                recs = _fx_dedup_latest(units[unit])
                unit_rank = 0 if unit == ccy else (1 if unit == f"{ccy}/shares" else
                            (2 if unit.endswith("/shares") else
                             (3 if unit == "shares" else 4)))
                latest = max((r.get("filed", "") for r in recs), default="")
                latest_rank = -int(re.sub(r"\D", "", latest) or "0")
                form_rank = 0 if any(r.get("form") in ("20-F", "40-F")
                                     for r in recs) else 1
                candidates.append((unit_rank, form_rank, -len(recs), latest_rank,
                                   c, recs, unit))
    if not candidates:
        return None, [], None
    candidates.sort(key=lambda x: (x[0], x[1], x[2], x[3]), reverse=False)
    _ur, _fr, _nr, _latest, c, recs, unit = candidates[0]
    return c, recs, unit


def _fx_annual_concept_iter(facts_root):
    """Yield (label, cat, concepts, taxo, kind) for the filer's taxonomy.
    IFRS filers: _FX_IFRS_MAP + _FX_IFRS_DISCL.  us-gaap FPIs: the pipeline's
    own rich CONCEPT_MAP (face + disclosures + concentration for free)."""
    ifrs = facts_root.get("ifrs-full", {})
    ugaap = facts_root.get("us-gaap", {})
    if len(ifrs) >= len(ugaap):
        mapped = set()
        for label, (cat, concepts) in _FX_IFRS_MAP.items():
            mapped.update(concepts)
            yield label, cat, concepts, ifrs, ("stock" if cat == _FX_BS else "flow")
        for label, (concepts, kind) in _FX_IFRS_DISCL.items():
            mapped.update(concepts)
            yield label, _FX_DIS, concepts, ifrs, kind
        for concept, meta in ifrs.items():
            if concept in mapped:
                continue
            human_label = meta.get("label") if isinstance(meta, dict) else None
            tgt = _fx_label_from_concept_name(
                concept, human_label,
                allowed_cats={_FX_IC, _FX_BS, _FX_CF},
            )
            if not tgt:
                continue
            cat, label = tgt
            yield label, cat, [concept], ifrs, ("stock" if cat == _FX_BS else "flow")
    else:
        cmap = globals().get("CONCEPT_MAP", {})
        mapped = set()
        for label, spec in cmap.items():
            cat = spec.get("cat", "")
            kind = "stock" if cat == _FX_BS else "flow"
            raw_concepts = list(_FX_FOREIGN_ANNUAL_EXTRA_CONCEPTS.get((cat, label), []))
            raw_concepts.extend(spec.get("tags", []) or [])
            concepts = []
            seen = set()
            for concept in raw_concepts:
                if not concept or concept in seen:
                    continue
                seen.add(concept)
                if _fx_foreign_annual_component_blocked(concept, (cat, label)):
                    continue
                concepts.append(concept)
            mapped.update(concepts)
            yield label, cat, concepts, ugaap, kind

        # Some US-GAAP FPIs use issuer-specific/custom face concepts in 20-F.
        # Let safe label/concept-name mapping supplement CONCEPT_MAP, but still
        # reject tax-note components via _fx_foreign_annual_component_blocked().
        for concept, meta in ugaap.items():
            if concept in mapped:
                continue
            human_label = meta.get("label") if isinstance(meta, dict) else None
            tgt = _fx_label_from_concept_name(
                concept, human_label,
                allowed_cats={_FX_IC, _FX_BS, _FX_CF},
            )
            if not tgt:
                continue
            cat, label = tgt
            if _fx_foreign_annual_component_blocked(concept, tgt, human_label):
                continue
            yield label, cat, [concept], ugaap, ("stock" if cat == _FX_BS else "flow")


def _fx_annual_unit_rank(label, unit, ccy):
    """Rank units for one canonical annual line.

    Companyfacts can retain several taxonomy concepts and units for the same
    line across a long history.  Ranking is performed per fiscal year rather
    than selecting one concept globally, which lets an issuer change taxonomy
    tags without deleting older annual anchors needed for Q4 recovery.
    """
    u = str(unit or "")
    u = re.sub(r"\b(?:iso4217|xbrli):", "", u, flags=re.I)
    if label in _FX_PER_SHARE:
        if ccy and u == f"{ccy}/shares":
            return 0
        if u.endswith("/shares"):
            return 1
        return 20
    if label in _FX_SHARES:
        if u == "shares":
            return 0
        if "share" in u.lower() and not u.endswith("/shares"):
            return 1
        return 20
    if ccy and u == ccy:
        return 0
    if u in _FX_CCY:
        return 3
    # Monetary concepts with non-currency units are usually not the face fact.
    return 10


def _fx_record_filed_rank(record):
    return int(re.sub(r"\D", "", str(record.get("filed", ""))) or "0")


def _fx_extract_annual(facts_json):
    facts_root = facts_json.get("facts", {})
    ccy = _fx_detect_currency(facts_json)
    rows, fy_end, non_add, kinds = {}, {}, set(), {}

    for label, cat, concepts, taxo, kind in _fx_annual_concept_iter(facts_root):
        if not concepts:
            continue
        is_stock = (kind == "stock")
        by_year = {}

        # Evaluate every mapped concept and unit per year.  The previous
        # implementation selected one concept for the issuer's entire history;
        # a taxonomy-tag change therefore erased otherwise valid older years.
        for concept_rank, concept in enumerate(concepts):
            cd = taxo.get(concept)
            if not cd:
                continue
            for unit, records in cd.get("units", {}).items():
                unit_rank = _fx_annual_unit_rank(label, unit, ccy)
                if unit_rank >= 20:
                    continue
                for r in _fx_dedup_latest(records):
                    if r.get("form") not in ("20-F", "40-F", "20-F/A", "40-F/A"):
                        continue
                    end_date = r.get("end")
                    if not end_date or "val" not in r:
                        continue
                    if is_stock:
                        yr = int(end_date[:4])
                        duration_penalty = 0.0
                    else:
                        months = _fx_months(r.get("start"), end_date)
                        if months is None or not (11 <= months <= 13):
                            continue
                        yr = int(end_date[:4])
                        duration_penalty = abs(float(months) - 12.0)
                    form_rank = 0 if r.get("form") in ("20-F", "40-F") else 1
                    # Lowest tuple wins, except the newest filing wins within
                    # an otherwise identical concept/unit choice.
                    score = (unit_rank, form_rank, duration_penalty,
                             -_fx_record_filed_rank(r), concept_rank)
                    prev = by_year.get(yr)
                    if prev is None or score < prev[0]:
                        by_year[yr] = (score, r, unit, concept)

        if not by_year:
            continue
        series = {}
        for yr, (_score, r, unit, _concept) in by_year.items():
            try:
                series[yr] = _fx_normalize_value(cat, label, float(r["val"]))
            except (TypeError, ValueError):
                continue
            end_date = r.get("end")
            if end_date and (yr not in fy_end or end_date > fy_end[yr]):
                fy_end[yr] = end_date
            if label in _FX_NON_ADDITIVE or unit == "shares" or str(unit).endswith("/shares"):
                non_add.add((cat, label))
        if series:
            rows[(cat, label)] = series
            kinds[(cat, label)] = kind

    return {"reporting_currency": ccy, "rows": rows, "fy_period_end": fy_end,
            "segments": {}, "non_additive": non_add, "kinds": kinds}


# ===========================================================================
# L2 -- per-filing 20-F XBRL: dimensional segments + face cross-check
# ===========================================================================
_FX_REV_CONCEPTS = {"Revenue", "Revenues", "RevenueFromContractsWithCustomers",
                    "RevenueFromContractWithCustomerExcludingAssessedTax",
                    "RevenueFromContractWithCustomerIncludingAssessedTax",
                    "SalesRevenueNet", "RevenueFromSaleOfGoods",
                    "RevenueFromRenderingOfServices"}
_FX_GEO_AXIS_KEYS = ("geograph",)
_FX_PLATFORM_AXIS_KEYS = ("platform", "business")
_FX_PRODUCT_AXIS_KEYS = ("productorservice", "product", "service")
_FX_BUS_AXIS_KEYS = ("segment", "reportable", "operating")
_FX_AXIS_SKIP = ("consolidat", "equitycomponent", "legalentity", "scenario",
                 "reclass", "retrospectiv", "range", "currency", "continuing",
                 "relatedparty", "majorcustomer", "counterparty")
_FX_MEM_SKIP = ("intersegment", "eliminat", "corporatenonsegment", "consolidated",
                "allother", "reconcil", "total", "operatingsegments", "material")
_FX_FACE_DIM_BLOCKERS = (
    "geograph", "country", "region", "segment", "reportable", "operatingsegment",
    "platform", "business", "product", "service", "customer", "majorcustomer",
    "counterparty", "relatedparty", "intersegment", "eliminat", "reconcil",
)
_FX_FACE_DIM_ALLOWED = (
    "consolidat", "total", "continuing", "ordinaryshare", "commonshare",
    "classesofshare", "sharecapital", "retrospectiv", "restatement",
)


@lru_cache(maxsize=65536, typed=True)
def _fx_local(qname):
    return str(qname).split(":")[-1].strip()


@lru_cache(maxsize=65536, typed=True)
def _fx_norm_unit(u):
    for tok in re.split(r"[^A-Za-z]+", str(u).upper()):
        if tok in _FX_CCY:
            return tok
    return None


@lru_cache(maxsize=65536, typed=True)
def _fx_member_label(qname):
    """Convert a taxonomy QName member into a readable label."""
    m = _fx_local(qname)
    m = re.sub(r"(Member|Domain)$", "", m)
    m = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", m)
    m = re.sub(r"(?<=[A-Za-z])(?=\d)", " ", m)
    return re.sub(r"\s+", " ", m).strip()


def _fx_face_dim_score(row, dim_cols):
    """Score dimensions for face-statement annual facts."""
    score = 0
    for dcol in dim_cols:
        raw = row.get(dcol)
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        raw_s = str(raw).strip()
        if not raw_s or raw_s.lower() in ("nan", "none"):
            continue
        axis = _fx_local(dcol).lower()
        member = _fx_local(raw_s).lower()
        combo = f"{axis} {member}"
        if any(k in combo for k in _FX_FACE_DIM_BLOCKERS):
            if not any(k in combo for k in _FX_FACE_DIM_ALLOWED):
                return None
        score += 1
    return score


def _fx_face_dim_score_values(values, dim_cols):
    """Tuple-backed equivalent of _fx_face_dim_score for hot dataframe loops."""
    score = 0
    for dcol, raw in zip(dim_cols, values):
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        raw_s = str(raw).strip()
        if not raw_s or raw_s.lower() in ("nan", "none"):
            continue
        axis = _fx_local(dcol).lower()
        member = _fx_local(raw_s).lower()
        combo = f"{axis} {member}"
        if any(k in combo for k in _FX_FACE_DIM_BLOCKERS):
            if not any(k in combo for k in _FX_FACE_DIM_ALLOWED):
                return None
        score += 1
    return score


_FX_MEM_ALIAS = {
    "us": "united states", "usa": "united states", "u s": "united states",
    "emea": "europe middle east africa", "uk": "united kingdom",
    "tw": "taiwan", "cn": "china", "jp": "japan", "kr": "korea",
    "de": "germany", "nl": "netherlands", "sg": "singapore",
    "hk": "hong kong", "apac": "asia pacific",
}


@lru_cache(maxsize=65536, typed=True)
def _fx_norm_member(label, axis_family=None):
    """Shared normalizer for XBRL members and HTML labels -> match key.
    Axis-aware aliases keep product/platform/geography "other" buckets from
    colliding while still matching the same member across sources and years."""
    t = re.sub(r"[^a-z0-9 ]", " ", str(label).lower())
    t = re.sub(r"\s+", " ", t).strip()
    t = _FX_MEM_ALIAS.get(t, t)
    filler = ["the", "and", "of"]
    if axis_family == _FX_SEG_GEO:
        filler += ["region", "regions", "country", "countries", "area", "areas"]
    elif axis_family == _FX_SEG_PLATFORM:
        filler += ["platform", "platforms", "segment", "segments", "business"]
    elif axis_family == _FX_SEG_PRODUCT:
        filler += ["product", "products", "type", "types", "service", "services"]
    else:
        filler += ["segment", "segments"]
    t = re.sub(r"\b(" + "|".join(filler) + r")\b", " ", t)
    toks = []
    for w in t.split():
        if axis_family and len(w) > 3 and w.endswith("s") and not w.endswith("ss"):
            w = w[:-1]
        toks.append(w)
    base = " ".join(toks)
    if base in ("other", "others") and axis_family:
        return f"{axis_family}:{base}"
    return base


@lru_cache(maxsize=65536, typed=True)
def _fx_axis_cat(dim_col):
    d = dim_col.lower()
    if any(k in d for k in _FX_AXIS_SKIP):
        return None
    if any(k in d for k in _FX_GEO_AXIS_KEYS):
        return _FX_SEG_GEO
    if any(k in d for k in _FX_PRODUCT_AXIS_KEYS):
        return _FX_SEG_PRODUCT
    if any(k in d for k in _FX_PLATFORM_AXIS_KEYS):
        return _FX_SEG_PLATFORM
    if any(k in d for k in _FX_BUS_AXIS_KEYS):
        return _FX_SEG_PLATFORM
    return None


def _fx_xbrl_extract(facts_df, ye_month):
    """One 20-F facts_df -> face FY values, dimensional segment FY values,
    concentration facts, currency.  Uses the same to_dataframe() schema as the
    US pipeline (concept, value, period_start/end/instant, unit_ref, dim_*)."""
    out = {"face": {}, "segments": {}, "concentration": {}, "ccy": None,
           "fy": None, "fy_end": None}
    if facts_df is None or len(facts_df) == 0:
        return out
    df = facts_df.copy()
    if "value" not in df.columns or "concept" not in df.columns:
        return out
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["value"])
    dim_cols = [c for c in df.columns if c.startswith("dim_")]
    df["_local"] = df["concept"].map(_fx_local)
    df["_raw_unit"] = df["unit_ref"].astype(str) if "unit_ref" in df.columns else ""
    df["_unit"] = df["_raw_unit"].map(_fx_norm_unit) if "unit_ref" in df.columns else None
    ccy = df["_unit"].value_counts().idxmax() if df["_unit"].notna().any() else None
    out["ccy"] = ccy

    end = pd.to_datetime(df.get("period_end"), errors="coerce")
    start = pd.to_datetime(df.get("period_start"), errors="coerce")
    inst = pd.to_datetime(df.get("period_instant"), errors="coerce") \
        if "period_instant" in df.columns else pd.Series(pd.NaT, index=df.index)
    dur_days = (end - start).dt.days
    is_fy = dur_days.between(330, 400)

    # fiscal-year end = latest FY-duration period_end in this filing
    if is_fy.any():
        fy_end_dt = end[is_fy].max()
    elif inst.notna().any():
        fy_end_dt = inst.max()
    else:
        return out
    fy = int(fy_end_dt.year)
    out["fy"] = fy
    out["fy_end"] = str(fy_end_dt.date())

    ndims = df[dim_cols].notna().sum(axis=1) if dim_cols else pd.Series(0, index=df.index)
    if dim_cols:
        df["_face_dim_score"] = [
            _fx_face_dim_score_values(values, dim_cols)
            for values in df[dim_cols].itertuples(index=False, name=None)
        ]
    else:
        df["_face_dim_score"] = 0

    # ---- face (undimensioned or benignly dimensioned) ----
    base = df[df["_face_dim_score"].notna()]
    ifrs_rev = {}
    for label, (cat, concepts) in _FX_IFRS_MAP.items():
        want_ccy = label not in _FX_NON_ADDITIVE
        for c in concepts:
            m = base[(base["_local"] == c)]
            if want_ccy and ccy:
                m = m[m["_unit"] == ccy]
            if cat == _FX_BS:
                m = m[inst.loc[m.index].dt.date == fy_end_dt.date()] if len(m) else m
            else:
                m = m[is_fy.loc[m.index] & (end.loc[m.index].dt.date == fy_end_dt.date())] if len(m) else m
            if len(m):
                m = m.sort_values(["_face_dim_score"], kind="stable")
                out["face"][(cat, label)] = _fx_normalize_value(cat, label, float(m.iloc[0]["value"]))
                if label == "Revenue":
                    ifrs_rev[fy] = out["face"][(cat, label)]
                break

    base_col_pos = {col: pos + 1 for pos, col in enumerate(base.columns)}  # +1: tuple[0] is the index

    def _base_tuple_get(row_values, col, default=None):
        pos = base_col_pos.get(col)
        return default if pos is None else row_values[pos]

    def _row_human_label_values(row_values):
        for col in ("label", "label_text", "standard_label", "concept_label", "name"):
            pos = base_col_pos.get(col)
            if pos is None:
                continue
            val = row_values[pos]
            if val is not None and str(val).strip() and str(val).lower() not in ("nan", "none"):
                return str(val)
        return None

    fallback = []
    for row_values in base.itertuples(index=True, name=None):
        idx = row_values[0]
        tgt = _fx_label_from_concept_name(
            _base_tuple_get(row_values, "_local"), _row_human_label_values(row_values),
            allowed_cats={_FX_IC, _FX_BS, _FX_CF},
        )
        if not tgt or tgt in out["face"]:
            continue
        cat, label = tgt
        if cat == _FX_BS:
            if pd.isna(inst.loc[idx]) or inst.loc[idx].date() != fy_end_dt.date():
                continue
        else:
            if (pd.isna(end.loc[idx]) or not bool(is_fy.loc[idx])
                    or end.loc[idx].date() != fy_end_dt.date()):
                continue
        rank_unit = (_base_tuple_get(row_values, "_raw_unit")
                     if label in _FX_NON_ADDITIVE else _base_tuple_get(row_values, "_unit"))
        unit_rank = _fx_annual_unit_rank(label, rank_unit, ccy)
        if unit_rank >= 20:
            continue
        fallback.append((unit_rank, float(_base_tuple_get(row_values, "_face_dim_score") or 0.0),
                         tgt, float(_base_tuple_get(row_values, "value"))))
    for _unit_rank, _dim_score, (cat, label), value in sorted(fallback, key=lambda x: (x[0], x[1])):
        out["face"].setdefault((cat, label), _fx_normalize_value(cat, label, value))

    # ---- dimensional: segments + concentration ----
    if dim_cols:
        one = df[(ndims == 1) & is_fy & (end.dt.date == fy_end_dt.date())]
        for dcol in dim_cols:
            sub = one[one[dcol].notna()]
            if not len(sub):
                continue
            dl = dcol.lower()
            if "majorcustomer" in dl or "concentration" in dl:
                conc_sub = sub[sub["_local"].isin(_FX_REV_CONCEPTS) |
                               sub["_local"].str.contains("Concentration", na=False)]
                d_pos = conc_sub.columns.get_loc(dcol)
                value_pos = conc_sub.columns.get_loc("value")
                for row_values in conc_sub.itertuples(index=False, name=None):
                    mem = _fx_member_label(row_values[d_pos])
                    if mem:
                        out["concentration"].setdefault(
                            ("7_Concentration_Risk", f"Revenue - Major Customer - {mem}"), {})[fy] = float(row_values[value_pos])
                continue
            cat = _fx_axis_cat(dcol)
            if not cat:
                continue
            sub = sub[sub["_local"].isin(_FX_REV_CONCEPTS)]
            if ccy:
                sub = sub[sub["_unit"] == ccy]
            d_pos = sub.columns.get_loc(dcol)
            value_pos = sub.columns.get_loc("value")
            for row_values in sub.itertuples(index=False, name=None):
                raw = row_values[d_pos]
                mem = _fx_member_label(raw)
                if not mem or any(k in _fx_local(raw).lower() for k in _FX_MEM_SKIP):
                    continue
                out["segments"].setdefault((cat, mem), {})[fy] = float(row_values[value_pos])
    return out



# ---------------------------------------------------------------------------
# Dynamic American Depositary Share (ADS) ratio detection
# ---------------------------------------------------------------------------
# Some foreign issuers report EPS per local ordinary/common share while the
# US-listed security is an ADS representing multiple local shares.  The old
# implementation used a ticker-specific ratio.  These helpers instead extract
# the ratio from the issuer's own XBRL concepts or filing text.
def _fx_valid_ads_ratio(value):
    try:
        value = float(value)
    except (TypeError, ValueError):
        return False
    return np.isfinite(value) and 0.01 <= value <= 1000.0


def _fx_ads_candidate(ratio, confidence, source, detail=None):
    if not _fx_valid_ads_ratio(ratio):
        return None
    return {
        "ratio": float(ratio),
        "confidence": float(confidence),
        "source": str(source),
        "detail": str(detail or ""),
    }


def _fx_ads_ratio_candidates_from_facts(facts_df):
    """Extract ADS ratios from custom numeric facts or text-block facts."""
    if facts_df is None or len(facts_df) == 0 or "concept" not in facts_df.columns:
        return []
    out = []
    col_pos = {col: pos for pos, col in enumerate(facts_df.columns)}
    concept_pos = col_pos.get("concept")
    value_pos = col_pos.get("value")
    for row_values in facts_df.itertuples(index=False, name=None):
        concept = _fx_local(row_values[concept_pos] if concept_pos is not None else "")
        compact = re.sub(r"[^a-z0-9]", "", concept.lower())
        raw_value = row_values[value_pos] if value_pos is not None else None

        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue
        if not _fx_valid_ads_ratio(value):
            continue

        has_ads = ("americandepositar" in compact or
                   "americandepositor" in compact or
                   re.search(r"(?:^|numberof|each|per)ads(?:share|shares)?", compact))
        if not has_ads or "share" not in compact:
            continue

        forward = bool(
            re.search(r"(?:common|ordinary|local|h)shares?.*(?:representedbyeach|per).*"
                      r"(?:americandepositar|americandepositor|ads)", compact)
            or re.search(r"(?:americandepositar|americandepositor|ads).*"
                         r"(?:represents?|representing).*(?:common|ordinary|local|h)shares?",
                         compact)
        )
        inverse = bool(
            re.search(r"(?:americandepositar|americandepositor|ads)(?:shares?)?per"
                      r"(?:common|ordinary|local|h)share", compact)
        )
        if forward:
            cand = _fx_ads_candidate(value, 0.995, f"xbrl_fact:{concept}", concept)
            if cand:
                out.append(cand)
        elif inverse and value != 0:
            cand = _fx_ads_candidate(1.0 / value, 0.99, f"xbrl_fact:{concept}", concept)
            if cand:
                out.append(cand)
    return out


def _fx_select_ads_ratio(candidates):
    """Select a ratio only when the evidence is strong and unambiguous."""
    groups = {}
    for cand in candidates or ():
        if not cand or not _fx_valid_ads_ratio(cand.get("ratio")):
            continue
        key = round(float(cand["ratio"]), 8)
        g = groups.setdefault(key, {"score": 0.0, "count": 0, "max_conf": 0.0,
                                    "sources": [], "details": []})
        conf = float(cand.get("confidence", 0.0))
        g["score"] += conf
        g["count"] += 1
        g["max_conf"] = max(g["max_conf"], conf)
        g["sources"].append(str(cand.get("source", "")))
        if cand.get("detail"):
            g["details"].append(str(cand["detail"]))
    if not groups:
        return None, None
    ranked = sorted(groups.items(), key=lambda kv: (kv[1]["score"], kv[1]["max_conf"],
                                                     kv[1]["count"]), reverse=True)
    ratio, evidence = ranked[0]
    if evidence["max_conf"] < 0.95 and evidence["count"] < 2:
        return None, None
    if len(ranked) > 1 and ranked[1][1]["score"] >= 0.85 * evidence["score"]:
        return None, None
    source = ", ".join(dict.fromkeys(evidence["sources"]))
    return float(ratio), source


def _fx_ratio_for_column(column, ratio_by_year=None, stable_ratio=None):
    """Return the applicable local-shares-per-ADS ratio for an output column."""
    try:
        year = int(str(column).split("-", 1)[0])
    except Exception:
        return stable_ratio
    if ratio_by_year and year in ratio_by_year:
        return ratio_by_year[year]
    return stable_ratio


def _fx_annual_from_xbrl(company, limit, progress=None, progress_start=18.0, progress_end=70.0,
                         annual_backbone=None, ye_month=12, use_arelle=False):
    """Loop recent 20-F/40-F filings -> merged annual XBRL layer only.

    Annual-only build note
    ----------------------
    This function reads structured XBRL facts from annual 20-F/40-F filings
    and uses companyfacts as the annual backbone.  ADS-ratio evidence is accepted only when it appears as a
    structured XBRL fact; prose-based ratio detection has been removed from
    this build to keep the foreign path annual/XBRL-only.
    """
    merged = {"face": {}, "segments": {}, "concentration": {}, "ccy": None,
              "fy_end": {}, "ads_ratio_by_year": {}, "ads_ratio": None,
              "ads_ratio_source": None, "ads_ratio_evidence": []}
    n_ok = 0
    ratio_candidates_all = []
    try:
        # `limit` now means annual filing count for the optional per-filing
        # XBRL enrichment layer.  Companyfacts still provides the broad annual
        # backbone across available history.
        n_filings = max(1, min(int(limit or 1), 10))
        filings = list(get_company_filings(company, "20-F").head(n_filings))
        if not filings:
            filings = list(get_company_filings(company, "40-F").head(n_filings))
    except Exception as e:
        if _FX_DEBUG:
            print(f"  [Foreign Annual][debug] annual filing listing failed: {type(e).__name__}")
        return merged
    filing_context = []
    for _filing in filings:
        _ident = _stable_filing_identity_for_persistent_cache(_filing)
        if _ident is None:
            filing_context = None
            break
        filing_context.append(_ident)
    xbrl_cache_key = None
    if filing_context:
        xbrl_cache_key = _fx_freeze_for_cache({
            "source_fingerprint": _fx_code_fingerprint(),
            "cik": str(getattr(company, "cik", "") or ""),
            "limit": int(limit or 0),
            "ye_month": int(ye_month or 12),
            "use_arelle": bool(use_arelle),
            "filings": tuple(filing_context),
        })
        cached_xbrl_layer = _fx_cache_get("annual_xbrl_layer", xbrl_cache_key)
        if cached_xbrl_layer is not _FX_CACHE_MISS:
            _profile_count("fx_annual_xbrl_layer_cache_hits")
            if progress is not None:
                progress.set(progress_end, "Loaded annual XBRL cache")
            return cached_xbrl_layer
    _profile_count("fx_annual_xbrl_layer_cache_misses")

    import io as _io
    import contextlib as _ctx
    total_filings = max(1, len(filings))
    stage_prefix = "20-F annual Arelle/XBRL" if use_arelle else "20-F annual XBRL"
    if progress is not None:
        progress.set(progress_start, f"{stage_prefix} 0/{len(filings)}")
    for filing_no, f in enumerate(filings, start=1):
        facts_df = None
        fy = None
        filing_candidates = []

        try:
            if use_arelle:
                # Run the same Arelle concept-map learner for foreign annual
                # filings before the normal edgartools XBRL parse. Arelle is
                # used only as metadata/enrichment: the actual values still come
                # from the existing 20-F/40-F XBRL/companyfacts pipeline.
                url = _get_filing_url_cached(f)
                if url:
                    arelle_update_concept_map(url, foreign=True)
            with _ctx.redirect_stdout(_io.StringIO()):
                xb = fetch_xbrl(f)
            if xb:
                facts_df = xb.facts.to_dataframe()
        except Exception as xbrl_err:
            if _FX_DEBUG:
                print(f"  [Foreign Annual][debug] XBRL fetch skipped "
                      f"({_fx_safe_accno(f)}): {type(xbrl_err).__name__}")

        if facts_df is not None:
            try:
                filing_candidates.extend(_fx_ads_ratio_candidates_from_facts(facts_df))
            except Exception as ratio_fact_err:
                if _FX_DEBUG:
                    print(f"  [Foreign Annual][debug] ADS ratio fact scan skipped "
                          f"({_fx_safe_accno(f)}): {type(ratio_fact_err).__name__}")
            try:
                res = _fx_xbrl_extract(facts_df, None)
                fy = res.get("fy")
                if fy is not None:
                    n_ok += 1
                    merged["ccy"] = merged["ccy"] or res["ccy"]
                    if res.get("fy_end"):
                        merged["fy_end"].setdefault(fy, res["fy_end"])
                    for key, val in res["face"].items():
                        merged["face"].setdefault(key, {}).setdefault(fy, val)
                    for key, series in res["segments"].items():
                        for y, v in series.items():
                            merged["segments"].setdefault(key, {}).setdefault(y, v)
                    for key, series in res["concentration"].items():
                        for y, v in series.items():
                            merged["concentration"].setdefault(key, {}).setdefault(y, v)
            except Exception as parse_err:
                if _FX_DEBUG:
                    print(f"  [Foreign Annual][debug] XBRL parse skipped "
                          f"({_fx_safe_accno(f)}): {type(parse_err).__name__}")

        # If XBRL did not provide a fiscal year, use filing metadata only for
        # associating structured ADS-ratio evidence; never for statement values.
        if fy is None:
            for attr in ("period_of_report", "filing_date"):
                try:
                    dt = pd.to_datetime(getattr(f, attr, None), errors="coerce")
                    if pd.notna(dt):
                        fy = int(dt.year)
                        break
                except Exception:
                    pass

        ratio, _source = _fx_select_ads_ratio(filing_candidates)
        if ratio is not None:
            if fy is not None:
                merged["ads_ratio_by_year"].setdefault(int(fy), ratio)
            for cand in filing_candidates:
                if abs(float(cand.get("ratio", 0.0)) - ratio) > 1e-8:
                    continue
                tagged = dict(cand)
                tagged["fy"] = fy
                ratio_candidates_all.append(tagged)
                merged["ads_ratio_evidence"].append(tagged)

        if progress is not None:
            pct = progress_start + (progress_end - progress_start) * filing_no / total_filings
            progress.set(pct, f"{stage_prefix} {filing_no}/{len(filings)}")

    unique_ratios = {round(float(v), 8)
                     for v in merged["ads_ratio_by_year"].values()
                     if _fx_valid_ads_ratio(v)}
    if len(unique_ratios) == 1:
        merged["ads_ratio"] = float(next(iter(unique_ratios)))
        _, merged["ads_ratio_source"] = _fx_select_ads_ratio(ratio_candidates_all)
    elif not unique_ratios:
        ratio, source = _fx_select_ads_ratio(ratio_candidates_all)
        merged["ads_ratio"] = ratio
        merged["ads_ratio_source"] = source

    if _FX_DEBUG:
        print(f"  [Foreign Annual][debug] annual XBRL: filings parsed={n_ok}  "
              f"seg members={len(merged['segments'])}  ccy={merged['ccy']}  "
              f"ADS ratio={merged.get('ads_ratio')}  "
              f"ratio years={sorted(merged.get('ads_ratio_by_year', {}))}")
    if xbrl_cache_key is not None:
        _fx_cache_set("annual_xbrl_layer", xbrl_cache_key, merged)
        _profile_count("fx_annual_xbrl_layer_cache_writes")
    return merged

def _fx_safe_accno(f):
    try:
        return getattr(f, "accession_no", "") or getattr(f, "accession_number", "")
    except Exception:
        return ""


# ===========================================================================
# L3 -- quarterly foreign-interim parser removed
# ===========================================================================
# The foreign branch uses annual 20-F/40-F XBRL/companyfacts only.
# Quarterly FPI extraction belongs in a separate script/engine.

# ===========================================================================
# L4 -- annual output assembly helpers
# ===========================================================================
_FX_Q_RANK = {"Q4": 0, "FY": 1, "Q3": 2, "Q2": 3, "Q1": 4}


# Quarterly bucketing/report folding removed for annual-only build.


def _fx_merge_annual_face(annual, xbrl_layer, qz):
    """Annual face precedence: 20-F/40-F XBRL face > companyfacts.

    The `qz` argument is retained only so the existing assembler can be reused;
    annual-only callers pass an empty container.
    """
    ann_rows = {k: dict(v) for k, v in annual["rows"].items()}
    kinds = dict(annual.get("kinds", {}))
    fy_end = dict(annual.get("fy_period_end", {}))
    agree = {}                                  # fy -> (checked, agreed)
    for key, series in xbrl_layer.get("face", {}).items():
        for fy, v in series.items():
            a = ann_rows.get(key, {}).get(fy)
            if a is not None and a != 0:
                c, g = agree.get(fy, (0, 0))
                agree[fy] = (c + 1, g + (1 if abs(a - v) <= _FX_TOL * abs(a) else 0))
            ann_rows.setdefault(key, {})[fy] = v
            kinds.setdefault(key, "stock" if key[0] == _FX_BS else "flow")
    for fy, d in xbrl_layer.get("fy_end", {}).items():
        fy_end.setdefault(fy, d)
    for okey, series in qz["cum"].items():
        if okey[0] != "face":
            continue
        key = okey[1:]
        for (fy, n), v in series.items():
            if n != 4 or (okey, fy) in qz["bad"]:
                continue
            ann_rows.setdefault(key, {}).setdefault(fy, v)
            kinds.setdefault(key, "stock" if key[0] == _FX_BS else "flow")
            fy_end.setdefault(fy, qz["cum_end"].get((fy, 4)))
    return ann_rows, kinds, fy_end, agree


def _fx_plausible_derived_q4(key, ann_v, got, q4_v):
    cat, lbl = key
    if q4_v is None:
        return False, "missing"
    if lbl in _FX_NON_ADDITIVE:
        return False, "non_additive"
    prior = [got.get(s) for s in ("Q1", "Q2", "Q3") if got.get(s) is not None]
    if not prior:
        return True, "no_prior_direct"
    max_prior = max(abs(v) for v in prior)
    if max_prior and abs(q4_v) > 3.5 * max_prior:
        return False, "implausible_vs_prior_quarters"
    if ann_v not in (None, 0) and abs(q4_v) > 1.25 * abs(ann_v):
        return False, "implausible_vs_annual"
    if cat == _FX_CF and lbl in _FX_NEGATIVE_FLOW_LABELS and q4_v > 0:
        return False, "wrong_outflow_sign"
    return True, "pass"


def _fx_align_annual_nonadditive_to_quarters(annual_value, quarter_values):
    """Align annual EPS/share units against observed interim values.

    This helper is dormant in the annual-only foreign route because no interim
    reports are parsed; it is retained only so the shared assembler remains
    import-compatible.
    """
    if annual_value is None:
        return None
    direct = [abs(float(v)) for v in quarter_values
              if v is not None and np.isfinite(float(v)) and float(v) != 0]
    if not direct:
        return float(annual_value)
    target = float(np.median(direct))
    if target <= 0:
        return None
    candidates = []
    for factor in (1e-6, 1e-3, 1.0, 1e3, 1e6):
        candidate = float(annual_value) * factor
        if candidate == 0 or not np.isfinite(candidate):
            continue
        score = abs(np.log(abs(candidate) / target))
        candidates.append((score, candidate))
    if not candidates:
        return None
    score, candidate = min(candidates)
    # More than a 20x scale mismatch is not a credible unit conversion.
    return candidate if score <= np.log(20.0) else None


def _fx_fiscal_quarter_day_weights(fy, qz, fy_end, ye_month):
    """Return exact day weights for Q1-Q4 where dates are available."""
    try:
        ends = []
        for qn in (1, 2, 3):
            d = qz.get("q_end", {}).get((fy, f"Q{qn}"))
            d = d or _fx_fiscal_quarter_end(fy, qn, ye_month)
            ends.append(pd.Timestamp(d))
        q4_end = fy_end.get(fy) or _fx_fiscal_quarter_end(fy, 4, ye_month)
        ends.append(pd.Timestamp(q4_end))
        prev_end = fy_end.get(fy - 1)
        if prev_end:
            prev = pd.Timestamp(prev_end)
        else:
            prev = ends[-1] - pd.DateOffset(years=1)
        points = [prev] + ends
        days = [(points[i + 1] - points[i]).days for i in range(4)]
        if any(d <= 0 or d > 120 for d in days):
            raise ValueError("implausible fiscal-quarter day count")
        return [float(d) for d in days]
    except Exception:
        return [1.0, 1.0, 1.0, 1.0]


def _fx_derive_q4_weighted_shares(annual_shares, got, fy, qz, fy_end, ye_month):
    """Solve the exact annual weighted-average-share identity for Q4.

    Annual weighted shares are a day-weighted average of the four discrete
    quarterly averages.  This is non-additive, but unlike EPS it has an exact
    recoverable Q4 when Q1-Q3 and the annual average are known.
    """
    direct = [got.get(s) for s in ("Q1", "Q2", "Q3")]
    if annual_shares is None or any(v is None for v in direct):
        return None
    aligned = _fx_align_annual_nonadditive_to_quarters(annual_shares, direct)
    if aligned is None:
        return None
    weights = _fx_fiscal_quarter_day_weights(fy, qz, fy_end, ye_month)
    numerator = aligned * sum(weights) - sum(float(v) * weights[i]
                                               for i, v in enumerate(direct))
    q4 = numerator / weights[3]
    if not np.isfinite(q4) or q4 <= 0:
        return None
    median_q = float(np.median([abs(float(v)) for v in direct]))
    if median_q > 0 and not (0.20 * median_q <= abs(q4) <= 5.0 * median_q):
        return None
    # Recheck the identity after solving to protect against date/unit errors.
    reconstructed = (sum(float(direct[i]) * weights[i] for i in range(3))
                     + q4 * weights[3]) / sum(weights)
    if abs(reconstructed - aligned) > 1e-9 * max(abs(aligned), 1.0):
        return None
    return q4


def _fx_q4_flow_derivation_allowed(key, fy, got, c2, c3, qz):
    """Permit derivation from the selected current observations.

    A historical duplicate disagreement should not permanently blank Q4 when
    the newest annual/cumulative facts form an exact identity.  Cumulative
    checks are used when present; otherwise all three direct quarters must come
    from snapshot-verified periods.
    """
    okey = ("face",) + key
    if (okey, fy) not in qz.get("bad", set()):
        return True
    tol = max(_FX_TOL * 2.0, 0.012)
    q1, q2, q3 = got.get("Q1"), got.get("Q2"), got.get("Q3")
    checks = []
    if c2 is not None and q1 is not None and q2 is not None:
        checks.append(abs((q1 + q2) - c2) <= tol * max(abs(c2), 1.0))
    if c3 is not None:
        if c2 is not None and q3 is not None:
            checks.append(abs((c2 + q3) - c3) <= tol * max(abs(c3), 1.0))
        elif None not in (q1, q2, q3):
            checks.append(abs((q1 + q2 + q3) - c3) <= tol * max(abs(c3), 1.0))
    if checks:
        return all(checks)
    if None not in (q1, q2, q3):
        verified = qz.get("verified_q", {})
        return all(bool(verified.get((fy, f"Q{i}"))) for i in (1, 2, 3))
    return False


def _fx_face_columns(ann_rows, kinds, fy_end, annual, qz, ye_month):
    """Face values per output column.

    Annual-only foreign runs pass no quarterly observations, so every year is
    emitted as a single FY column.
    """
    non_add_keys = set(annual.get("non_additive", set()))
    columns, col_end, sum_ok = {}, {}, {}
    q, cum, bad = qz["q"], qz["cum"], qz["bad"]
    q_years = sorted({fy for (fy, _s) in qz["q_end"]}, reverse=True)
    ann_years = sorted({y for s in ann_rows.values() for y in s}, reverse=True)
    years = sorted(set(q_years) | set(ann_years), reverse=True)

    face_keys = set(ann_rows)
    q_obs = set()
    for okey in list(q) + list(cum):
        if okey[0] == "face":
            face_keys.add(okey[1:])
            q_obs.add(okey[1:])

    def dv(key, fy, sub):
        s = q.get(("face",) + key, {})
        return s.get((fy, sub))

    def cv(key, fy, n):
        s = cum.get(("face",) + key, {})
        return s.get((fy, n))

    for fy in years:
        if fy in q_years:
            subs = sorted({s for (y, s) in qz["q_end"] if y == fy},
                          key=lambda s: _FX_Q_RANK.get(s, 9))
            have_ann = any(fy in ann_rows.get(k, {}) for k in face_keys)
            if have_ann and "Q4" not in subs:
                subs = ["Q4"] + subs
            for sub in subs:
                col = f"{fy}-{sub}"
                col_end[col] = (qz["q_end"].get((fy, sub))
                                or (fy_end.get(fy) if sub == "Q4" else None))
            for key in face_keys:
                cat, lbl = key
                kind = kinds.get(key, "stock" if cat == _FX_BS else "flow")
                unreliable = (("face",) + key, fy) in bad
                ann_v = ann_rows.get(key, {}).get(fy)
                source_ann_v = annual.get("rows", {}).get(key, {}).get(fy)
                per_share = (key in non_add_keys) or (lbl in _FX_NON_ADDITIVE)
                cum4_v = cv(key, fy, 4)
                q4_anchor = (cum4_v if (kind != "stock" and cum4_v is not None
                                        and source_ann_v is None)
                             else ann_v)
                got = {s: dv(key, fy, s) for s in ("Q1", "Q2", "Q3", "Q4")}
                q4_derived = False
                if key not in q_obs and kind != "stock" and not per_share:
                    got = {"Q1": None, "Q2": None, "Q3": None, "Q4": None}
                elif kind == "stock":
                    if got["Q4"] is None:
                        got["Q4"] = ann_v            # fiscal year-end snapshot
                    elif ann_v is not None:
                        got["Q4"] = ann_v            # annual source preferred
                elif per_share:
                    # EPS itself is not additive.  Weighted-average shares,
                    # however, obey an exact day-weighted annual identity.
                    # Recover Q4 shares from that identity, then the cleanup
                    # stage derives Q4 EPS from parent earnings / Q4 shares.
                    if lbl in _FX_SHARES and got["Q4"] is None:
                        eps_label = ("EPS Basic" if "Basic" in lbl else "EPS Diluted")
                        for _sub in ("Q1", "Q2", "Q3"):
                            if got.get(_sub) is not None:
                                continue
                            eps_q = dv((_FX_IC, eps_label), fy, _sub)
                            parent_q = dv((_FX_IC, "Net Income to Parent"), fy, _sub)
                            if parent_q is None:
                                parent_q = dv((_FX_IC, "Net Income"), fy, _sub)
                            try:
                                if eps_q not in (None, 0) and parent_q is not None:
                                    got[_sub] = float(parent_q) / float(eps_q)
                            except (TypeError, ValueError, ZeroDivisionError):
                                pass
                        annual_shares = ann_v
                        if annual_shares is None:
                            annual_eps = ann_rows.get((_FX_IC, eps_label), {}).get(fy)
                            annual_parent = ann_rows.get((_FX_IC, "Net Income to Parent"), {}).get(fy)
                            if annual_parent is None:
                                annual_parent = ann_rows.get((_FX_IC, "Net Income"), {}).get(fy)
                            if annual_eps not in (None, 0) and annual_parent is not None:
                                annual_shares = annual_parent / annual_eps
                        got["Q4"] = _fx_derive_q4_weighted_shares(
                            annual_shares, got, fy, qz, fy_end, ye_month)
                else:
                    c2, c3 = cv(key, fy, 2), cv(key, fy, 3)
                    can_derive = _fx_q4_flow_derivation_allowed(
                        key, fy, got, c2, c3, qz)
                    if not unreliable or can_derive:
                        if got["Q2"] is None and c2 is not None and got["Q1"] is not None:
                            got["Q2"] = c2 - got["Q1"]
                        if got["Q3"] is None and c3 is not None:
                            if c2 is not None:
                                got["Q3"] = c3 - c2
                            elif got["Q1"] is not None and got["Q2"] is not None:
                                got["Q3"] = c3 - got["Q1"] - got["Q2"]
                        if got["Q4"] is None and q4_anchor is not None:
                            if c3 is not None:
                                got["Q4"] = q4_anchor - c3
                                q4_derived = True
                            elif all(got[s] is not None for s in ("Q1", "Q2", "Q3")):
                                got["Q4"] = q4_anchor - got["Q1"] - got["Q2"] - got["Q3"]
                                q4_derived = True
                            if q4_derived:
                                ok_q4, _why = _fx_plausible_derived_q4(key, q4_anchor, got, got["Q4"])
                                if not ok_q4:
                                    got["Q4"] = None
                                    q4_derived = False
                if (kind != "stock" and not per_share and q4_anchor is not None
                        and all(got[s] is not None for s in ("Q1", "Q2", "Q3", "Q4"))):
                    ssum = got["Q1"] + got["Q2"] + got["Q3"] + got["Q4"]
                    ok = abs(ssum - q4_anchor) <= _FX_TOL * max(abs(q4_anchor), 1.0)
                    sum_ok[(key, fy)] = 1.0 if ok else 0.0
                    if not ok and q4_derived:
                        got["Q4"] = None         # blank over wrong
                for sub, v in got.items():
                    col = f"{fy}-{sub}"
                    if v is not None and col in col_end:
                        columns.setdefault(key, {})[col] = v
        else:
            col = f"{fy}-FY"
            wrote = False
            for key in face_keys:
                v = ann_rows.get(key, {}).get(fy)
                if v is not None:
                    columns.setdefault(key, {})[col] = v
                    wrote = True
            if wrote:
                col_end[col] = fy_end.get(fy, f"{fy}-12-31")
    return columns, col_end, q_years, sum_ok


def _fx_segments_quarterly(qz, ann_rows, xbrl_layer, q_years):
    """Segment rows.

    Annual-only foreign runs have no quarterly observations, so segment rows
    come from annual 20-F/40-F XBRL dimensions and are emitted as FY columns.
    """
    rows, checks = {}, {}
    q, cum, bad = qz["q"], qz["cum"], qz["bad"]
    q_pct, cum_pct = qz.get("q_pct", {}), qz.get("cum_pct", {})
    display, origs = {}, {}                         # (cat, norm) -> label / originals

    # Annual 20-F XBRL dimensions are authoritative axis evidence. When an
    # HTML member has a unique normalized match in XBRL, use that XBRL axis.
    xbrl_cats_by_member = {}
    for (xcat, xmem), _series in xbrl_layer.get("segments", {}).items():
        generic = _fx_norm_member(xmem, None)
        if generic:
            xbrl_cats_by_member.setdefault(generic, set()).add(xcat)

    def resolve_html_cat(cat, label):
        generic = _fx_norm_member(label, None)
        semantic = _fx_segment_member_cat(label, cat)
        # Strong label semantics outrank both the surrounding table header and
        # annual XBRL.  This is useful for explicit labels such as countries,
        # regions, or product/resolution names.
        if semantic == _FX_SEG_GEO:
            return _FX_SEG_GEO
        if semantic == _FX_SEG_PRODUCT and cat == _FX_SEG_GEO:
            return _FX_SEG_PRODUCT

        # Never use a generic residual member to infer an axis.  Labels such as
        # "Others" appear independently in platform, product, geography, and
        # technology tables; mapping them to the one annual XBRL axis where the
        # same word happens to exist causes cross-axis leakage and can turn a
        # valid 100% partition into 102%.
        compact = re.sub(r"[^a-z0-9]", "", str(generic or "").lower())
        if compact in {"other", "others", "allother", "allothers",
                       "miscellaneous", "unallocated"}:
            return cat

        xcats = xbrl_cats_by_member.get(generic, set())
        if len(xcats) == 1:
            return next(iter(xcats))
        return cat

    def note_label(cat, label):
        k = (cat, _fx_norm_member(label, cat))
        cur = display.get(k)
        if cur is None or len(str(label)) > len(str(cur)):
            display[k] = str(label)
        origs.setdefault(k, set()).add(str(label))
        return k

    seg_q = {}                                      # (cat, norm) -> {(fy,sub): v}
    for okey, series in q.items():
        if okey[0] != "seg":
            continue
        resolved_cat = resolve_html_cat(okey[1], okey[2])
        k = note_label(resolved_cat, okey[2])
        for (fy, sub), v in series.items():
            seg_q.setdefault(k, {})[(fy, sub)] = v
    seg_c = {}
    for okey, series in cum.items():
        if okey[0] != "seg":
            continue
        resolved_cat = resolve_html_cat(okey[1], okey[2])
        k = note_label(resolved_cat, okey[2])
        for (fy, n), v in series.items():
            seg_c.setdefault(k, {})[(fy, n)] = v
    seg_pq = {}
    for okey, series in q_pct.items():
        if okey[0] != "seg_pct":
            continue
        resolved_cat = resolve_html_cat(okey[1], okey[2])
        k = note_label(resolved_cat, okey[2])
        for (fy, sub), v in series.items():
            seg_pq.setdefault(k, {})[(fy, sub)] = float(v)

    seg_pc = {}
    for okey, series in cum_pct.items():
        if okey[0] != "seg_pct":
            continue
        resolved_cat = resolve_html_cat(okey[1], okey[2])
        k = note_label(resolved_cat, okey[2])
        for (fy, n), v in series.items():
            seg_pc.setdefault(k, {})[(fy, n)] = float(v)

    seg_x = {}
    for (cat, mem), series in xbrl_layer.get("segments", {}).items():
        k = note_label(cat, mem)
        for fy, v in series.items():
            seg_x.setdefault(k, {}).setdefault(fy, v)

    cats = {c for (c, _n) in list(seg_q) + list(seg_c) + list(seg_pq) + list(seg_pc) + list(seg_x)}
    fys = set(q_years) | {fy for s in seg_x.values() for fy in s}
    ann_rev = ann_rows.get((_FX_IC, "Revenue"), {})

    def face_q(fy, sub):
        return qz["q"].get(("face", _FX_IC, "Revenue"), {}).get((fy, sub))

    def face_c(fy, n):
        return qz["cum"].get(("face", _FX_IC, "Revenue"), {}).get((fy, n))

    def partition_ties(part, rev, tol_mult=1.0):
        if not part:
            return False
        if rev in (None, 0):
            return True
        return abs(sum(part.values()) - rev) <= _FX_TOL * tol_mult * abs(rev)

    def pct_to_amount(pct, rev):
        pct_sum = sum(pct.values()) if pct else 0.0
        if (rev not in (None, 0) and len(pct) >= 2
                and all(-1e-9 <= v <= 1.001 for v in pct.values())
                and 0.96 <= pct_sum <= 1.04):
            return {k: rev * (v / pct_sum) for k, v in pct.items()}
        return None

    def residual_member_key(k):
        norm = str(k[1]).lower()
        compact = re.sub(r"[^a-z0-9]", "", norm)
        return (
            compact.endswith("other")
            or compact.endswith("others")
            or compact in {
                "allother", "allothers", "restofworld", "restworld",
                "miscellaneous", "unallocated",
            }
        )

    def plausible_partition(part, rev, prior_parts=(),
                            allow_residual_negative=False):
        if not part:
            return False
        scale = max(abs(rev) if rev not in (None, 0) else
                    abs(sum(part.values())), 1.0)
        residual_negative = 0.0
        for k, v in part.items():
            if v > 1.10 * scale:
                return False
            if v >= -0.005 * scale:
                continue
            if (not allow_residual_negative
                    or not residual_member_key(k)
                    or v < -0.05 * scale):
                return False
            residual_negative += abs(v)
        if residual_negative > 0.05 * scale:
            return False
        for k, v in part.items():
            priors = []
            for p in prior_parts or ():
                if p and p.get(k) is not None:
                    priors.append(abs(p[k]))
            if priors and max(priors) > 0 and abs(v) > 4.0 * max(priors):
                return False
        return True

    for cat in sorted(cats):
        norms = [k for k in set(list(seg_q) + list(seg_c) + list(seg_pq) +
                               list(seg_pc) + list(seg_x)) if k[0] == cat]
        for fy in sorted(fys, reverse=True):
            if fy not in q_years:
                for k in norms:                     # hybrid FY rows from XBRL
                    v = seg_x.get(k, {}).get(fy)
                    if v is not None:
                        rows.setdefault((cat, f"Revenue - {display[k]}"), {})[f"{fy}-FY"] = v
                continue
            qmem = {s: {k: seg_q[k][(fy, s)] for k in norms
                        if (fy, s) in seg_q.get(k, {})}
                    for s in ("Q1", "Q2", "Q3", "Q4")}
            qpmem = {s: {k: seg_pq[k][(fy, s)] for k in norms
                         if (fy, s) in seg_pq.get(k, {})}
                     for s in ("Q1", "Q2", "Q3", "Q4")}
            unreliable = any((("seg", cat, o), fy) in bad
                             for k in norms for o in origs.get(k, ()))
            # Select a direct amount partition when it ties.  When an amount
            # partition is absent, partial, or pathological, a complete
            # disclosed percentage partition is converted to amounts using
            # consolidated revenue.  Rounded percentages are normalized only
            # when they already sum to approximately 100%.
            chosen_by_sub = {}
            for s in ("Q1", "Q2", "Q3", "Q4"):
                fr = face_q(fy, s)
                direct = qmem[s]
                direct_tie = False
                pathological = False
                if direct and fr not in (None, 0):
                    total = sum(direct.values())
                    direct_tie = abs(total - fr) <= _FX_TOL * abs(fr)
                    if cat in (_FX_SEG_PRODUCT, _FX_SEG_GEO):
                        scale = max(abs(fr), 1.0)
                        pathological = (
                            any(v < -0.005 * scale or v > 1.10 * scale
                                for v in direct.values())
                            or total > 1.20 * scale
                        )
                pct = qpmem[s]
                pct_sum = sum(pct.values()) if pct else 0.0
                pct_complete = (
                    fr not in (None, 0) and len(pct) >= 2
                    and all(-1e-9 <= v <= 1.001 for v in pct.values())
                    and 0.96 <= pct_sum <= 1.04
                )
                use_pct = pct_complete and (not direct_tie or pathological)
                chosen = None
                if use_pct:
                    chosen = {k: fr * (v / pct_sum) for k, v in pct.items()}
                    checks[(cat, fy, s)] = 1.0
                elif direct and not pathological:
                    chosen = direct
                    if fr not in (None, 0):
                        checks[(cat, fy, s)] = 1.0 if direct_tie else 0.0
                elif pathological:
                    checks[(cat, fy, s)] = 0.0
                if chosen:
                    chosen_by_sub[s] = chosen

            def cumulative_partition(n):
                rev = face_c(fy, n)
                direct = {k: seg_c[k][(fy, n)] for k in norms
                          if (fy, n) in seg_c.get(k, {})}
                if direct and partition_ties(direct, rev, tol_mult=3.0):
                    return direct, rev
                pct = {k: seg_pc[k][(fy, n)] for k in norms
                       if (fy, n) in seg_pc.get(k, {})}
                amount = pct_to_amount(pct, rev)
                if amount:
                    return amount, rev
                return None, rev

            def annual_partition():
                rev = face_c(fy, 4) or ann_rev.get(fy)
                direct = {k: seg_c[k][(fy, 4)] for k in norms
                          if (fy, 4) in seg_c.get(k, {})}
                if direct and partition_ties(direct, rev, tol_mult=3.0):
                    return direct, rev
                xbrl = {k: seg_x[k][fy] for k in norms if fy in seg_x.get(k, {})}
                if xbrl and partition_ties(xbrl, ann_rev.get(fy), tol_mult=3.0):
                    return xbrl, ann_rev.get(fy)
                pct = {k: seg_pc[k][(fy, 4)] for k in norms
                       if (fy, 4) in seg_pc.get(k, {})}
                amount = pct_to_amount(pct, rev)
                if amount:
                    return amount, rev
                return None, rev

            def install_derived(sub, derived, expected_rev, prior_parts):
                if chosen_by_sub.get(sub) or not derived:
                    return False
                if not partition_ties(derived, expected_rev, tol_mult=3.0):
                    return False
                if not plausible_partition(derived, expected_rev, prior_parts):
                    return False
                chosen_by_sub[sub] = derived
                checks[(cat, fy, sub)] = 1.0
                return True

            c2, c2_rev = cumulative_partition(2)
            c3, c3_rev = cumulative_partition(3)
            anchor, anchor_rev = annual_partition()

            if c2 and chosen_by_sub.get("Q1"):
                q1 = chosen_by_sub["Q1"]
                if set(c2) == set(q1):
                    q2_rev = face_q(fy, "Q2")
                    if q2_rev is None and None not in (c2_rev, face_q(fy, "Q1")):
                        q2_rev = c2_rev - face_q(fy, "Q1")
                    install_derived(
                        "Q2",
                        {k: c2[k] - q1[k] for k in c2},
                        q2_rev,
                        (q1,),
                    )

            if c3:
                if c2 and set(c3) == set(c2):
                    q3_rev = face_q(fy, "Q3")
                    if q3_rev is None and None not in (c3_rev, c2_rev):
                        q3_rev = c3_rev - c2_rev
                    install_derived(
                        "Q3",
                        {k: c3[k] - c2[k] for k in c3},
                        q3_rev,
                        (chosen_by_sub.get("Q1"), chosen_by_sub.get("Q2")),
                    )
                elif chosen_by_sub.get("Q1") and chosen_by_sub.get("Q2"):
                    q1, q2 = chosen_by_sub["Q1"], chosen_by_sub["Q2"]
                    if set(c3) == set(q1) == set(q2):
                        q3_rev = face_q(fy, "Q3")
                        prev_rev = None
                        if None not in (face_q(fy, "Q1"), face_q(fy, "Q2")):
                            prev_rev = face_q(fy, "Q1") + face_q(fy, "Q2")
                        if q3_rev is None and None not in (c3_rev, prev_rev):
                            q3_rev = c3_rev - prev_rev
                        install_derived(
                            "Q3",
                            {k: c3[k] - q1[k] - q2[k] for k in c3},
                            q3_rev,
                            (q1, q2),
                        )

            if (not chosen_by_sub.get("Q3") and anchor
                    and all(chosen_by_sub.get(s) for s in ("Q1", "Q2", "Q4"))):
                q1, q2, q4 = (chosen_by_sub["Q1"], chosen_by_sub["Q2"],
                              chosen_by_sub["Q4"])
                if set(anchor) == set(q1) == set(q2) == set(q4):
                    q3_rev = face_q(fy, "Q3")
                    prev_rev = None
                    if None not in (face_q(fy, "Q1"), face_q(fy, "Q2"), face_q(fy, "Q4")):
                        prev_rev = face_q(fy, "Q1") + face_q(fy, "Q2") + face_q(fy, "Q4")
                    if q3_rev is None and None not in (anchor_rev, prev_rev):
                        q3_rev = anchor_rev - prev_rev
                    install_derived(
                        "Q3",
                        {k: anchor[k] - q1[k] - q2[k] - q4[k] for k in anchor},
                        q3_rev,
                        (q1, q2, q4),
                    )

            if not chosen_by_sub.get("Q4") and anchor:
                if c3:
                    base, base_rev = c3, c3_rev
                elif all(chosen_by_sub.get(s) for s in ("Q1", "Q2", "Q3")):
                    q1, q2, q3 = (chosen_by_sub["Q1"], chosen_by_sub["Q2"],
                                  chosen_by_sub["Q3"])
                    if set(q1) == set(q2) == set(q3):
                        base = {k: q1[k] + q2[k] + q3[k] for k in q1}
                        base_rev = None
                        if None not in (face_q(fy, "Q1"), face_q(fy, "Q2"), face_q(fy, "Q3")):
                            base_rev = face_q(fy, "Q1") + face_q(fy, "Q2") + face_q(fy, "Q3")
                    else:
                        base, base_rev = None, None
                else:
                    base, base_rev = None, None

                if base:
                    if base_rev is None:
                        base_rev = sum(base.values())
                    bijective = set(base) == set(anchor)
                    derived = {k: anchor[k] - base[k] for k in set(anchor) & set(base)}
                    q4_rev = face_q(fy, "Q4")
                    if q4_rev is None and None not in (anchor_rev, base_rev):
                        q4_rev = anchor_rev - base_rev
                    a_ties = partition_ties(anchor, anchor_rev)
                    b_ties = partition_ties(base, base_rev)
                    q4_ties = partition_ties(derived, q4_rev, tol_mult=3.0)
                    evidence_ok = ((a_ties and b_ties) or q4_ties)
                    if unreliable:
                        evidence_ok = bool(a_ties and b_ties)
                    ok_derive = (
                        bijective and evidence_ok
                        and plausible_partition(
                            derived, q4_rev,
                            (chosen_by_sub.get("Q1"), chosen_by_sub.get("Q2"),
                             chosen_by_sub.get("Q3")),
                            allow_residual_negative=True,
                        )
                    )
                    checks[(cat, fy, "Q4")] = 1.0 if ok_derive else 0.0
                    if ok_derive:
                        chosen_by_sub["Q4"] = derived

            for sub, chosen in chosen_by_sub.items():
                for k, v in chosen.items():
                    rows.setdefault((cat, f"Revenue - {display[k]}"), {})[
                        f"{fy}-{sub}"] = v
    return rows, checks


def _fx_fmt_date(d):
    try:
        return pd.to_datetime(d).strftime("%m/%d/%y")
    except Exception:
        return ""


def _fx_safe_div(a, b):
    try:
        if a is None or b in (None, 0):
            return None
        return a / b
    except Exception:
        return None



def _fx_relation_is_valid(columns, relation, col_end, tolerance=_FX_TOL):
    """Return True when a canonical accounting identity is demonstrated by
    the issuer's already-observed periods.  This prevents a mapping mistake in
    one company from becoming a synthetic fill everywhere else."""
    residuals = []
    for col in col_end:
        vals = []
        complete = True
        for key, coeff in relation:
            v = columns.get(key, {}).get(col)
            if v is None or not np.isfinite(float(v)):
                complete = False
                break
            vals.append((float(v), float(coeff)))
        if not complete:
            continue
        resid = abs(sum(v * c for v, c in vals))
        scale = max([abs(v) for v, _c in vals] + [1.0])
        residuals.append(resid / scale)
    if not residuals:
        return False
    pass_rate = sum(r <= tolerance for r in residuals) / len(residuals)
    # One exceptionally clean observation is useful for young issuers; with
    # more history require broad consistency rather than one accidental tie.
    if len(residuals) == 1:
        return residuals[0] <= min(tolerance, 0.001)
    return pass_rate >= 0.80 and float(np.median(residuals)) <= tolerance


def _fx_plausible_relation_fill(key, value, known_values):
    if value is None or not np.isfinite(float(value)):
        return False
    v = float(value)
    scale = max([abs(float(x)) for x in known_values if x is not None] + [1.0])
    if abs(v) > 20.0 * scale:
        return False
    cat, label = key
    if cat == _FX_BS and label in {
            "Total Assets", "Total Current Assets", "Total Non-current Assets",
            "Total Liabilities", "Total Current Liabilities",
            "Total Non-current Liabilities", "Noncontrolling Interest"}:
        if v < -_FX_TOL * scale:
            return False
    return True


def _fx_complete_accounting_relations(columns, col_end):
    """Fill only one missing term in issuer-validated exact identities.

    The function never overwrites a reported value.  Each identity must first
    reconcile in the same issuer's observed periods, so it generalizes across
    IFRS and US-GAAP FPIs without assuming that similarly named custom rows are
    interchangeable.
    """
    relations = [
        # Income statement and attribution.
        [((_FX_IC, "Revenue"), 1.0),
         ((_FX_IC, "Cost of Revenue"), -1.0),
         ((_FX_IC, "Gross Profit"), -1.0)],
        [((_FX_IC, "Pretax Income"), 1.0),
         ((_FX_IC, "Income Tax Expense"), -1.0),
         ((_FX_IC, "Net Income"), -1.0)],
        [((_FX_IC, "Net Income"), 1.0),
         ((_FX_IC, "Net Income to Parent"), -1.0),
         ((_FX_IC, "Net Income to NCI"), -1.0)],
        [((_FX_IC, "Pretax Income"), 1.0),
         ((_FX_IC, "Operating Income"), -1.0),
         ((_FX_IC, "Total Non-operating Income"), -1.0)],
        [((_FX_IC, "Total Non-operating Income"), 1.0),
         ((_FX_IC, "Equity Method Income"), -1.0),
         ((_FX_IC, "Interest Income"), -1.0),
         ((_FX_IC, "Finance Costs"), -1.0),
         ((_FX_IC, "Other Income/(Expense)"), -1.0)],
        [((_FX_IC, "Pretax Income"), 1.0),
         ((_FX_IC, "Operating Income"), -1.0),
         ((_FX_IC, "Equity Method Income"), -1.0),
         ((_FX_IC, "Interest Income"), -1.0),
         ((_FX_IC, "Finance Costs"), -1.0),
         ((_FX_IC, "Other Income/(Expense)"), -1.0)],
        # Balance-sheet closure and current/non-current decompositions.
        [((_FX_BS, "Total Assets"), 1.0),
         ((_FX_BS, "Total Liabilities"), -1.0),
         ((_FX_BS, "Total Equity"), -1.0)],
        [((_FX_BS, "Total Assets"), 1.0),
         ((_FX_BS, "Total Current Assets"), -1.0),
         ((_FX_BS, "Total Non-current Assets"), -1.0)],
        [((_FX_BS, "Total Liabilities"), 1.0),
         ((_FX_BS, "Total Current Liabilities"), -1.0),
         ((_FX_BS, "Total Non-current Liabilities"), -1.0)],
        [((_FX_BS, "Total Equity"), 1.0),
         ((_FX_BS, "Equity to Parent"), -1.0),
         ((_FX_BS, "Noncontrolling Interest"), -1.0)],
        # Cash-flow bridge.  FX is an explicit term, never silently assumed 0
        # when it is the only missing observation.
        [((_FX_CF, "Net Cash Flow"), 1.0),
         ((_FX_CF, "Operating Cash Flow"), -1.0),
         ((_FX_CF, "Investing Cash Flow"), -1.0),
         ((_FX_CF, "Financing Cash Flow"), -1.0),
         ((_FX_CF, "Effect of FX on Cash"), -1.0)],
        [((_FX_CF, "Cash & Equivalents, End of Period"), 1.0),
         ((_FX_CF, "Cash & Equivalents, Beginning of Period"), -1.0),
         ((_FX_CF, "Net Cash Flow"), -1.0)],
    ]
    valid = [r for r in relations if _fx_relation_is_valid(columns, r, col_end)]
    if not valid:
        return columns
    # Several identities can unlock another one, so iterate to a fixed point.
    for _ in range(4):
        changed = False
        for relation in valid:
            for col in col_end:
                missing = []
                known_sum = 0.0
                known_values = []
                for key, coeff in relation:
                    v = columns.get(key, {}).get(col)
                    if v is None:
                        missing.append((key, float(coeff)))
                    else:
                        fv = float(v)
                        known_sum += fv * float(coeff)
                        known_values.append(fv)
                if len(missing) != 1:
                    continue
                key, coeff = missing[0]
                if coeff == 0:
                    continue
                value = -known_sum / coeff
                if not _fx_plausible_relation_fill(key, value, known_values):
                    continue
                columns.setdefault(key, {})[col] = value
                changed = True
        if not changed:
            break
    return columns


def _fx_reconcile_foreign_annual_tax_identity(columns, col_end):
    """Normalize the displayed tax row to expense sign using the face identity.

    Some FPI 20-F XBRL tags/labels present "income tax benefit (expense)" with
    benefits as positive values, while this CSV row is named "Income Tax
    Expense".  Also, companyfacts can expose current/deferred tax components.
    For the final annual face, the safest display value is the identity:
        Income Tax Expense = Pretax Income - Net Income
    when both face totals are present.
    """
    tax_key = (_FX_IC, "Income Tax Expense")
    ptx_key = (_FX_IC, "Pretax Income")
    ni_key = (_FX_IC, "Net Income")
    for col in list(col_end or []):
        ptx = columns.get(ptx_key, {}).get(col)
        ni = columns.get(ni_key, {}).get(col)
        if ptx is None or ni is None:
            continue
        try:
            expected = float(ptx) - float(ni)
        except Exception:
            continue
        current = columns.get(tax_key, {}).get(col)
        if current is None:
            columns.setdefault(tax_key, {})[col] = expected
            continue
        try:
            cur = float(current)
        except Exception:
            columns.setdefault(tax_key, {})[col] = expected
            continue
        tol = _FX_TOL * max(abs(float(ptx)), abs(float(ni)), abs(expected), 1.0)
        if abs(cur - expected) > tol:
            columns.setdefault(tax_key, {})[col] = expected
    return columns


def _fx_parent_income_series(row_getter):
    """Build a conservative common-parent earnings series for EPS recovery."""
    parent = pd.to_numeric(row_getter(_FX_IC, "Net Income to Parent"), errors="coerce")
    ni = pd.to_numeric(row_getter(_FX_IC, "Net Income"), errors="coerce")
    nci = pd.to_numeric(row_getter(_FX_IC, "Net Income to NCI"), errors="coerce")
    parent = parent.copy()
    direct_identity = ni - nci
    parent = parent.fillna(direct_identity.where(ni.notna() & nci.notna()))

    # When NCI is demonstrably immaterial in this issuer's observed history,
    # consolidated profit is a defensible numerator fallback for remaining
    # gaps.  Do not make this assumption for issuers with material minorities.
    rel_samples = []
    both = ni.notna() & parent.notna() & (ni.abs() > 0)
    if both.any():
        rel_samples.extend(((parent[both] - ni[both]).abs() / ni[both].abs()).tolist())
    nci_samples = ni.notna() & nci.notna() & (ni.abs() > 0)
    if nci_samples.any():
        rel_samples.extend((nci[nci_samples].abs() / ni[nci_samples].abs()).tolist())
    rel_samples = [float(x) for x in rel_samples if np.isfinite(x)]
    if len(rel_samples) >= 3 and float(np.quantile(rel_samples, 0.90)) <= 0.02:
        parent = parent.fillna(ni)
    return parent


def _fx_share_scale_fallback(raw):
    vals = pd.to_numeric(raw, errors="coerce").abs().dropna()
    if vals.empty:
        return 1.0
    med = float(vals.median())
    # XBRL shares are normally absolute; HTML tables often use thousands or
    # millions.  These thresholds convert all three representations to
    # millions of shares when no EPS overlap is available to infer the scale.
    if med >= 1_000_000:
        return 1_000_000.0
    if med >= 1_000:
        return 1_000.0
    return 1.0


def _fx_prepare_eps_shares(raw_eps, raw_shares, parent, ratio_by_col=None):
    """Normalize raw EPS and weighted-share units on the issuer basis.

    The EPS/net-income relationship is used only to infer whether weighted
    shares are absolute, thousands, or millions.  ADS conversion is deliberately
    deferred until after missing EPS/share observations have been completed, so
    issuer-reported EPS remains unchanged while the displayed share count can be
    expressed as listed ADS-equivalent units.
    """
    eps = pd.to_numeric(raw_eps, errors="coerce")
    shares = pd.to_numeric(raw_shares, errors="coerce")
    parent = pd.to_numeric(parent, errors="coerce")

    divisors = (1.0, 1_000.0, 1_000_000.0)
    candidates = []
    for div_rank, div in enumerate(divisors):
        share_out = shares / div
        mask = (parent.notna() & eps.notna() & share_out.notna()
                & (eps.abs() > 0) & (share_out.abs() > 0))
        if not mask.any():
            continue
        implied = (parent[mask].abs() / 1e6) / eps[mask].abs()
        actual = share_out[mask].abs()
        good = (implied > 0) & (actual > 0)
        if good.any():
            score = float(np.median(np.abs(np.log(actual[good] / implied[good]))))
            candidates.append((score, div_rank, div))

    if candidates:
        _score, _rank, divisor = min(candidates)
    else:
        divisor = _fx_share_scale_fallback(shares)

    eps_out = eps.copy()
    shares_out = shares / divisor
    return eps_out, shares_out


def _fx_assemble(annual, xbrl_layer, qz, ye_month):
    """Statement assembly for the foreign branch.

    In this annual-only build the caller passes an empty quarterly container,
    so only FY columns are emitted.  The historical quarterly code remains as
    dormant helper logic for compatibility with the surrounding assembly
    functions, but the foreign route does not discover or parse interim reports.
    """
    ann_rows, kinds, fy_end, agree = _fx_merge_annual_face(annual, xbrl_layer, qz)
    columns, col_end, q_years, sum_ok = _fx_face_columns(
        ann_rows, kinds, fy_end, annual, qz, ye_month)

    def put(key, col, val):
        if val is not None and col in col_end:
            columns.setdefault(key, {})[col] = val

    seg_rows, seg_checks = _fx_segments_quarterly(qz, ann_rows, xbrl_layer, q_years)
    for key, d in seg_rows.items():
        for col, val in d.items():
            put(key, col, val)

    for key, series in xbrl_layer.get("concentration", {}).items():
        for fy, val in series.items():
            put(key, f"{fy}-Q4" if fy in q_years else f"{fy}-FY", val)

    # Recover exact accounting terms only after direct/derived observations
    # have been selected.  The identities self-validate on the issuer's own
    # reported periods and never overwrite a reported value.
    columns = _fx_complete_accounting_relations(columns, col_end)
    columns = _fx_reconcile_foreign_annual_tax_identity(columns, col_end)

    if not col_end:
        return pd.DataFrame()

    def sort_key(c):
        y, s = c.split("-")
        return (-int(y), _FX_Q_RANK.get(s, 9))
    col_order = sorted(col_end, key=sort_key)

    def g(label, col, cat=_FX_IC):
        return columns.get((cat, label), {}).get(col)

    for col in col_order:
        rev = g("Revenue", col)
        gp, oi, ni = g("Gross Profit", col), g("Operating Income", col), g("Net Income", col)
        tax, ptx = g("Income Tax Expense", col), g("Pretax Income", col)
        cfo = g("Operating Cash Flow", col, _FX_CF)
        capex = g("Capital Expenditures", col, _FX_CF)
        da_total = g("Depreciation & Amortization", col, _FX_CF)
        if da_total is None:
            da_parts = [
                v for v in (
                    g("Depreciation", col, _FX_CF),
                    g("Amortization", col, _FX_CF),
                ) if v is not None and v >= 0
            ]
            da_total = sum(da_parts) if da_parts else None
        cash = g("Cash & Equivalents", col, _FX_BS)
        std = g("Short-term Debt", col, _FX_BS) or 0.0
        ltd = g("Long-term Debt", col, _FX_BS) or 0.0
        cll = g("Current Lease Liabilities", col, _FX_BS) or 0.0
        nll = g("Noncurrent Lease Liabilities", col, _FX_BS) or 0.0
        for lab, num in (("Gross Margin (%)", gp), ("Operating Margin (%)", oi),
                         ("Net Margin (%)", ni)):
            m = _fx_safe_div(num, rev)
            if m is not None:
                put((_FX_KPI, lab), col, round(m * 100, 2))
        if oi is not None and da_total is not None and da_total >= 0:
            ebitda = oi + da_total
            put((_FX_KPI, "Metric: EBITDA"), col, ebitda)
            ebitda_margin = _fx_safe_div(ebitda, rev)
            if ebitda_margin is not None:
                put((_FX_KPI, "Metric: EBITDA Margin %"), col,
                    round(ebitda_margin * 100, 2))
        if cfo is not None and capex is not None:
            fcf = cfo - abs(capex)
            put((_FX_KPI, "Metric: Free Cash Flow"), col, fcf)
            m = _fx_safe_div(fcf, rev)
            if m is not None:
                put((_FX_KPI, "Metric: FCF Margin %"), col, round(m * 100, 2))
        m = _fx_safe_div(tax, ptx)
        if m is not None:
            put((_FX_KPI, "Metric: Effective Tax Rate %"), col, round(m * 100, 2))
        if std or ltd:
            put((_FX_KPI, "Metric: Total Debt"), col, std + ltd)
            if cash is not None:
                put((_FX_KPI, "Metric: Net Cash (Debt)"), col, cash - (std + ltd))
        if cll or nll:
            put((_FX_KPI, "Metric: Total Lease Liabilities"), col, cll + nll)
    for col in col_order:
        y, s = col.split("-")
        prev = f"{int(y)-1}-{s}"
        a = g("Revenue", col)
        b = g("Revenue", prev) if prev in col_end else None
        if a is not None and b not in (None, 0):
            put((_FX_KPI, "Metric: Revenue Growth %"), col, round((a - b) / b * 100, 2))

    for col in col_order:
        A = g("Total Assets", col, _FX_BS)
        L = g("Total Liabilities", col, _FX_BS)
        E = g("Total Equity", col, _FX_BS)
        if None not in (A, L, E) and A:
            put((_FX_CHK, "Metric: Balance Sheet Closure Verified"), col,
                1.0 if abs(A - L - E) <= _FX_TOL * abs(A) else 0.0)
        cfo = g("Operating Cash Flow", col, _FX_CF)
        cfi = g("Investing Cash Flow", col, _FX_CF)
        cff = g("Financing Cash Flow", col, _FX_CF)
        fxe_raw = g("Effect of FX on Cash", col, _FX_CF)
        fxe = fxe_raw or 0.0
        dch = g("Net Cash Flow", col, _FX_CF)
        if None not in (cfo, cfi, cff, dch):
            # Issuers differ: some report "Net Cash Flow" before FX translation
            # effects (e.g. Spotify), while others report net increase/decrease
            # after FX (e.g. TSMC). Accept either presentation when the explicit
            # FX row is available.
            base_tie = cfo + cfi + cff - dch
            fx_tie = cfo + cfi + cff + fxe - dch
            best_tie = fx_tie if fxe_raw is not None and abs(fx_tie) < abs(base_tie) else base_tie
            put((_FX_CHK, "Metric: Cash Flow Ties Verified"), col,
                1.0 if abs(best_tie) <= _FX_TOL * max(abs(dch), abs(cfo), 1.0) else 0.0)
    for (cat, fy, sub), flag in seg_checks.items():
        col = f"{fy}-{sub}"
        if cat == _FX_SEG_PRODUCT:
            lab = "Metric: Segment Sum Verified (Product Type)"
        elif cat == _FX_SEG_GEO:
            lab = "Metric: Segment Sum Verified (Geographic)"
        else:
            lab = "Metric: Segment Sum Verified (Platform)"
        put((_FX_CHK, lab), col, flag)
    for fy, (checked, agreed) in agree.items():
        if checked:
            col = f"{fy}-Q4" if fy in q_years else f"{fy}-FY"
            put((_FX_CHK, "Metric: Annual Source Agreement Verified"), col,
                1.0 if agreed == checked else 0.0)
    for key in ((_FX_IC, "Revenue"), (_FX_IC, "Net Income")):
        for (k, fy), flag in sum_ok.items():
            if k == key:
                put((_FX_CHK, "Metric: Quarters Sum to FY Verified"),
                    f"{fy}-Q4", flag)
    for (fy, sub), ver in qz["verified_q"].items():
        put((_FX_CHK, "Metric: Interim Snapshot Verified"), f"{fy}-{sub}",
            1.0 if ver else 0.0)

    hdr = ("0_Period_Header", "Period Ending")
    for col in col_order:
        put(hdr, col, _fx_fmt_date(col_end.get(col)))

    def row_sort(key):
        cat, lbl = key
        if cat == _FX_CF:
            order = _FX_ORDER.get(cat, [])
            subtotal_order = {
                "Operating Cash Flow": 900,
                "Investing Cash Flow": 920,
                "Financing Cash Flow": 940,
                "Effect of FX on Cash": 960,
                "Net Cash Flow": 980,
            }
            if lbl in subtotal_order:
                return (cat, subtotal_order[lbl], lbl)
            if lbl in order:
                return (cat, order.index(lbl), lbl)
            return (cat, 500, lbl)
        order = _FX_ORDER.get(cat)
        pos = order.index(lbl) if (order and lbl in order) else 999
        return (cat, pos, lbl)

    idx_keys = sorted(columns, key=row_sort)
    idx = pd.MultiIndex.from_tuples(idx_keys, names=["Category", "Label"])
    df = pd.DataFrame(index=idx, columns=col_order, dtype="object")
    for key, d in columns.items():
        for col, val in d.items():
            df.at[key, col] = val
    return df.dropna(how="all")


def _fx_write_foreign_outputs(final_pivot, ticker, out_dir, diagnostics=None,
                              ads_ratio_by_year=None, stable_ads_ratio=None,
                              save_xlsx=False):
    final_pivot = _fx_display_cleanup(
        final_pivot,
        ads_ratio_by_year=ads_ratio_by_year,
        stable_ads_ratio=stable_ads_ratio,
    )
    if save_xlsx:
        xlsx_dir = f"{out_dir}/excel"
        os.makedirs(xlsx_dir, exist_ok=True)
        out_path = f"{xlsx_dir}/{ticker}_annual_financials.xlsx"
        try:
            _save_pivot_xlsx(final_pivot, out_path)
            return out_path
        except ImportError:
            print("  [xlsx] openpyxl not installed -- saved CSV instead.")
        except Exception as _xe:
            print(f"  [xlsx] export failed ({_xe}); saved CSV instead.")

    final_pivot = _normalize_output_margin_rows(final_pivot)
    out_path = f"{out_dir}/{ticker}_annual_financials.csv"
    _profile_call("write_csv", final_pivot.to_csv, out_path)
    return out_path


def _fx_display_cleanup(df, ads_ratio_by_year=None, stable_ads_ratio=None):
    if df is None or df.empty:
        return df
    out = df.copy().astype("object")
    qcols = [c for c in out.columns if not str(c).endswith("-FY")]
    if qcols:
        out = out.loc[:, qcols]

    def _row(cat, label):
        if (cat, label) in out.index:
            return pd.to_numeric(out.loc[(cat, label)], errors="coerce")
        return pd.Series(np.nan, index=out.columns)

    def _put(cat, label, series):
        if series is None:
            return
        s = pd.to_numeric(series, errors="coerce")
        if s.notna().any():
            out.loc[(cat, label), :] = s.reindex(out.columns).values

    def _sum_present(*series):
        valid = [pd.to_numeric(x, errors="coerce") for x in series if x is not None]
        if not valid:
            return pd.Series(np.nan, index=out.columns)
        return pd.concat(valid, axis=1).sum(axis=1, min_count=1)

    # Reference-style IFRS balance-sheet rollups from commonly split lines.
    _put(_FX_BS, "Trade Receivables",
         _row(_FX_BS, "Notes And Accounts Receivable Net"))
    rou_assets = _row(_FX_BS, "Right Of Use Assets")
    _put(_FX_BS, "Right-of-Use Assets", rou_assets)
    _put(_FX_BS, "Property, Plant & Equipment",
         _sum_present(_row(_FX_BS, "Property, Plant & Equipment"), rou_assets))
    _put(_FX_BS, "Intangibles & Goodwill", _row(_FX_BS, "Intangible Assets"))
    _put(_FX_BS, "Accounts Payable",
         _sum_present(_row(_FX_BS, "Accounts Payable"),
                      _row(_FX_BS, "Payables To Contractors And Equipment Suppliers"),
                      _row(_FX_BS, "Payables To Related Parties")))
    _put(_FX_BS, "Long-term Debt",
         _sum_present(_row(_FX_BS, "Bonds Payable"),
                      _row(_FX_BS, "Long Term Bank Loans")))
    _put(_FX_BS, "Noncurrent Lease Liabilities", _row(_FX_BS, "Lease Liabilities"))

    # Reference-style cash-flow rollups from the raw parsed IFRS rows.
    dep = _row(_FX_CF, "Depreciation")
    amo = _row(_FX_CF, "Amortization")
    da = _sum_present(dep, amo)
    _put(_FX_CF, "Depreciation & Amortization", da)
    operating_income = _row(_FX_IC, "Operating Income")
    ebitda = (operating_income + da).where(
        operating_income.notna() & da.notna() & (da >= 0)
    )
    _put(_FX_KPI, "Metric: EBITDA", ebitda)
    revenue = _row(_FX_IC, "Revenue")
    _put(_FX_KPI, "Metric: EBITDA Margin %",
         ((ebitda / revenue) * 100).where(revenue.notna() & (revenue != 0)).round(2))
    _put(_FX_CF, "Share-Based Compensation", _row(_FX_CF, "Share Based Compensation"))
    # IFRS cash-flow statements often start from income before income tax.
    # Do not relabel that starter as "Net Income"; keep it as a distinct
    # cash-flow bridge row so the IS Net Income row remains unambiguous.
    _put(_FX_CF, "Pretax Income", _row(_FX_IC, "Pretax Income"))
    _put(_FX_CF, "Change in Receivables", _row(_FX_CF, "Notes And Accounts Receivable Net"))
    _put(_FX_CF, "Changes in Inventories", _row(_FX_CF, "Inventories"))
    _put(_FX_CF, "Changes in Accounts Payable", _row(_FX_CF, "Accounts Payable"))
    accrued = _sum_present(
        _row(_FX_CF, "Accrued Expenses And Other Current Liabilities"),
        _row(_FX_CF, "Accrued Profit Sharing Bonus To Employees And Compensation To Directors"),
        _row(_FX_CF, "Salary And Bonus Payable"),
    )
    _put(_FX_CF, "Changes in Accrued Expenses", accrued)
    _put(_FX_CF, "Changes in Income Taxes Payable", _row(_FX_CF, "Income Taxes Paid"))
    _put(_FX_CF, "Purchases of Intangible Assets", _row(_FX_CF, "Intangible Assets"))
    _put(_FX_CF, "Purchases of Investments",
         -_sum_present(
             _row(_FX_CF, "Financial Assets At Fair Value Through Other Comprehensive Income").abs(),
             _row(_FX_CF, "Financial Assets At Amortized Cost").abs(),
         ))
    _put(_FX_CF, "Proceeds from Sale of Investments",
         _sum_present(
             _row(_FX_CF, "Financial Assets At Fair Value Through Other Comprehensive Income").clip(lower=0),
             _row(_FX_CF, "Financial Assets At Amortized Cost").clip(lower=0),
             _row(_FX_CF, "Proceeds From Return Of Capital Of Investments In Equity Instruments At Fair Value Through Other Comprehensive Income"),
         ))
    _put(_FX_CF, "Long-Term Debt Issued",
         _sum_present(_row(_FX_CF, "Proceeds From Issuance Of Bonds"),
                      _row(_FX_CF, "Proceeds From Long Term Bank Loans")))
    _put(_FX_CF, "Long-Term Debt Repaid",
         _sum_present(_row(_FX_CF, "Repayment Of Bonds"),
                      _row(_FX_CF, "Repayment Of Long Term Bank Loans")))
    _put(_FX_CF, "Common Dividends Paid", _row(_FX_CF, "Dividends Paid"))
    _put(_FX_CF, "Effect of Exchange Rate Changes on Cash and Cash Equivalents",
         _row(_FX_CF, "Effect of FX on Cash"))
    ocf = _row(_FX_CF, "Operating Cash Flow")
    capex = _row(_FX_CF, "Capital Expenditures")
    _put(_FX_CF, "Free Cash Flow", _sum_present(ocf, capex))

    pretax = _row(_FX_IC, "Pretax Income")
    operating = _row(_FX_IC, "Operating Income")
    total_nonop = (pretax - operating).where(pretax.notna() & operating.notna())
    _put(_FX_IC, "Total Non-operating Income", total_nonop)
    residual_other = (
        total_nonop
        - _row(_FX_IC, "Equity Method Income").fillna(0)
        - _row(_FX_IC, "Interest Income").fillna(0)
        - _row(_FX_IC, "Finance Costs").fillna(0)
    ).where(total_nonop.notna())
    existing_other = _row(_FX_IC, "Other Income/(Expense)")
    _put(_FX_IC, "Other Income/(Expense)", existing_other.fillna(residual_other))

    # Normalize weighted-share units to millions.  Missing EPS/share values are
    # completed on the issuer's ordinary-share basis first.  A validated ADS
    # ratio is applied only to the final displayed share counts, preserving the
    # issuer-reported EPS values while matching the listed-security share basis.
    ratio_by_year = {int(y): float(v) for y, v in (ads_ratio_by_year or {}).items()
                     if _fx_valid_ads_ratio(v)}
    ratio_by_col = pd.Series(
        {_c: _fx_ratio_for_column(_c, ratio_by_year, stable_ads_ratio)
         for _c in out.columns},
        index=out.columns,
        dtype="float64",
    )
    parent = _fx_parent_income_series(_row)
    _put(_FX_IC, "Net Income to Parent", parent)

    prepared = {}
    for eps_label, share_label in (
            ("EPS Basic", "Shares Outstanding Basic"),
            ("EPS Diluted", "Shares Outstanding Diluted")):
        eps_direct, shares_direct = _fx_prepare_eps_shares(
            _row(_FX_IC, eps_label),
            _row(_FX_IC, share_label),
            parent,
            ratio_by_col,
        )
        with np.errstate(divide="ignore", invalid="ignore"):
            eps_from_shares = (parent / 1e6) / shares_direct
            shares_from_eps = (parent / 1e6) / eps_direct
        eps_from_shares = eps_from_shares.where(
            shares_direct.notna() & (shares_direct > 0))
        shares_from_eps = shares_from_eps.where(
            eps_direct.notna() & (shares_from_eps > 0))
        eps_final = eps_direct.fillna(eps_from_shares)
        shares_final = shares_direct.fillna(shares_from_eps)
        # A first fill can unlock the reciprocal observation.
        with np.errstate(divide="ignore", invalid="ignore"):
            eps_final = eps_final.fillna(((parent / 1e6) / shares_final).where(shares_final > 0))
            shares_final = shares_final.fillna(((parent / 1e6) / eps_final).where(
                ((parent / 1e6) / eps_final) > 0))
        prepared[(eps_label, share_label)] = [eps_final, shares_final]

    # Basic and diluted weighted shares are often identical.  Use one class as
    # a fallback only in periods where the issuer's EPS classes agree within
    # 2%; otherwise preserve the distinction and leave the gap blank.
    b_eps, b_sh = prepared[("EPS Basic", "Shares Outstanding Basic")]
    d_eps, d_sh = prepared[("EPS Diluted", "Shares Outstanding Diluted")]
    eps_close = (b_eps.notna() & d_eps.notna() &
                 ((b_eps - d_eps).abs() <= 0.02 * np.maximum(b_eps.abs(), d_eps.abs()).clip(lower=1e-9)))
    b_sh = b_sh.fillna(d_sh.where(eps_close))
    d_sh = d_sh.fillna(b_sh.where(eps_close))
    shares_close = (b_sh.notna() & d_sh.notna() &
                    ((b_sh - d_sh).abs() <= 0.02 * np.maximum(b_sh.abs(), d_sh.abs()).clip(lower=1e-9)))
    b_eps = b_eps.fillna(d_eps.where(shares_close))
    d_eps = d_eps.fillna(b_eps.where(shares_close))

    # Convert only the final displayed weighted-share counts to ADS-equivalent
    # units.  Ratios are filing-derived and validated; periods without reliable
    # ratio evidence remain on the issuer basis rather than being guessed.
    valid_ratio = ratio_by_col.where(ratio_by_col.apply(_fx_valid_ads_ratio))
    b_sh_display = b_sh.where(valid_ratio.isna(), b_sh / valid_ratio)
    d_sh_display = d_sh.where(valid_ratio.isna(), d_sh / valid_ratio)

    # Internal foreign-filer share arithmetic is deliberately performed in
    # millions so NI / EPS identities remain numerically stable.  The public
    # CSV contract, however, uses raw share counts just like the native 10-K /
    # 10-Q pipeline.  Convert only at the final display boundary and round to
    # whole shares to avoid floating-point artifacts such as 5185505000.000001.
    b_sh_display = (b_sh_display * 1_000_000.0).round()
    d_sh_display = (d_sh_display * 1_000_000.0).round()

    _put(_FX_IC, "EPS Basic", b_eps)
    _put(_FX_IC, "EPS Diluted", d_eps)
    _put(_FX_IC, "Shares Outstanding Basic", b_sh_display)
    _put(_FX_IC, "Shares Outstanding Diluted", d_sh_display)

    keep_idx = []
    for cat, label in out.index:
        if cat in (_FX_IC, _FX_BS, _FX_CF):
            if label in _FX_DISPLAY_KEEP.get(cat, set()):
                keep_idx.append((cat, label))
            continue
        keep_idx.append((cat, label))
    out = out.loc[pd.MultiIndex.from_tuples(keep_idx, names=out.index.names)]

    # Drop display rows with no quarterly data after cleanup.
    if qcols:
        out = out[~out.loc[:, qcols].isna().all(axis=1)]

    def _sort_idx(idx):
        cat, label = idx
        order = _FX_ORDER.get(cat, [])
        return (cat, order.index(label) if label in order else 999, label)

    out = out.reindex(sorted(out.index, key=_sort_idx))
    return out


# ===========================================================================
# DRIVER
# ===========================================================================
def _fx_empty_quarterization():
    """Empty quarterly container for annual-only foreign output."""
    return {"q": {}, "cum": {}, "q_pct": {}, "cum_pct": {},
            "q_end": {}, "cum_end": {}, "bad": set(), "verified_q": {}}


def _fx_trim_annual_columns(df, limit):
    """Keep the most recent `limit` FY columns for annual-only foreign output."""
    if df is None or df.empty:
        return df
    try:
        n = int(limit)
    except Exception:
        n = 0
    if n <= 0:
        return df
    fy_cols = [c for c in df.columns if str(c).endswith("-FY")]
    if len(fy_cols) <= n:
        return df

    def _fy_sort(c):
        try:
            return -int(str(c).split("-")[0])
        except Exception:
            return 0

    keep = sorted(fy_cols, key=_fy_sort)[:n]
    return df.loc[:, keep]


def _fx_main_foreign_20f(ticker, company, ye_month, out_dir, limit=8, save_xlsx=False, progress=None, use_arelle=False):
    print(f"  [Foreign Annual] Detected foreign private issuer -- annual 20-F/40-F XBRL path.")
    diagnostics = _fx_empty_diagnostics()
    if progress is not None:
        progress.set(6.0, "Loading SEC company facts")

    try:
        foreign_cache_context = _fx_foreign_filing_context(company, limit)
        facts = _fx_fetch_companyfacts(company.cik, cache_context=foreign_cache_context)
        annual = _fx_extract_annual(facts)
        rev_years = sorted(annual["rows"].get((_FX_IC, "Revenue"), {}), reverse=True)
        print(f"  [Foreign Annual] Annual companyfacts rows: {len(annual['rows'])}  "
              f"currency: {annual['reporting_currency']}  "
              f"Revenue FYs: {rev_years[:3]}..{rev_years[-1:] if rev_years else []}")
        if progress is not None:
            progress.set(18.0, "Annual company facts parsed")
    except Exception as e:
        print(f"  [Foreign Annual][ERROR] companyfacts failed: {type(e).__name__}: {e}")
        annual = {"reporting_currency": None, "rows": {}, "fy_period_end": {},
                  "segments": {}, "non_additive": set(), "kinds": {}}
        if progress is not None:
            progress.set(18.0, "Company facts unavailable; continuing")

    try:
        xbrl_layer = _fx_annual_from_xbrl(
            company, limit, progress=progress, progress_start=20.0, progress_end=72.0,
            annual_backbone=annual, ye_month=ye_month, use_arelle=use_arelle,
        )
    except Exception as e:
        print(f"  [Foreign Annual][warn] annual XBRL layer skipped: {type(e).__name__}")
        xbrl_layer = {"face": {}, "segments": {}, "concentration": {},
                      "ccy": None, "fy_end": {}, "ads_ratio_by_year": {},
                      "ads_ratio": None, "ads_ratio_source": None}

    if progress is not None:
        progress.set(78.0, "Assembling annual statements")
    final_pivot = _fx_assemble(annual, xbrl_layer, _fx_empty_quarterization(), ye_month)
    final_pivot = _fx_trim_annual_columns(final_pivot, limit)
    if progress is not None:
        progress.set(94.0, "Annual statement assembly complete")
    if final_pivot.empty:
        print("  [Foreign Annual] No annual data extracted.")
        return

    if progress is not None:
        progress.set(96.0, "Writing output")
    ads_ratio = xbrl_layer.get("ads_ratio")
    ads_ratio_by_year = xbrl_layer.get("ads_ratio_by_year", {})
    if ads_ratio is not None:
        print(f"  [Foreign Annual] Filing-derived ADS ratio: {ads_ratio:g} "
              f"local shares per ADS ({xbrl_layer.get('ads_ratio_source') or 'structured filing facts'}).")
    elif ads_ratio_by_year:
        print(f"  [Foreign Annual] Filing-derived ADS ratios by year: {ads_ratio_by_year}")
    out_path = _fx_write_foreign_outputs(
        final_pivot, ticker, out_dir, diagnostics,
        ads_ratio_by_year=ads_ratio_by_year,
        stable_ads_ratio=ads_ratio,
        save_xlsx=save_xlsx,
    )
    if progress is not None:
        progress.set_result_footer(f"Saved to {out_path}")
    n_by_cat = final_pivot.groupby(level=0).size().to_dict()
    print(f"  [Foreign Annual] Rows by category: {n_by_cat}")
    print(f"  [Foreign Annual] Output version: annual-only 20-F/40-F XBRL/companyfacts + Arelle {'on' if use_arelle else 'off'}")
    print(f"Success! Data saved to {out_path}")
    return final_pivot


# ---------------------------------------------------------------------------
# NATIVE 10-K ANNUAL MODE (--annual)
# ---------------------------------------------------------------------------
_NATIVE_ANNUAL_FORMS = {"10-K", "10-K/A"}
_NATIVE_ANNUAL_INSTANT_PREFIXES = {
    "Assets", "Total Assets", "Total Current Assets", "Total Non-Current Assets",
    "Liabilities", "Total Liabilities", "Total Current Liabilities",
    "Total Non-Current Liabilities", "Total Equity", "Total Debt",
    "Cash & Equivalents", "Short-Term Investments", "Inventory",
    "Accounts Receivable", "Accounts Payable", "Property, Plant & Equipment",
    "Deferred Revenue", "Operating Lease ROU Asset", "Operating Lease Liability",
}
_NATIVE_ANNUAL_WEIGHTED_SHARE_LABELS = {
    "Shares Outstanding Basic", "Shares Outstanding Diluted",
}
_NATIVE_ANNUAL_REPORTED_ONLY_LABELS = {
    "EPS Basic", "EPS Diluted",
    "Shares Outstanding Basic", "Shares Outstanding Diluted",
}


def _native_annual_effective_month_year(end_dt):
    end_dt = pd.to_datetime(end_dt, errors="coerce")
    if pd.isna(end_dt):
        return None, None
    month, year = int(end_dt.month), int(end_dt.year)
    if int(end_dt.day) < 15:
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return month, year


def _native_annual_effective_ye_month_from_filings(filings_10k, fallback_month=12):
    """Return annual-mode fiscal year-end month using the same 52/53-week
    early-month normalization as fact period selection.

    The normal quarterly pipeline keeps its existing ``ye_month`` derivation.
    This helper is used only by native ``--annual`` mode so early-November
    fiscal-year ends (for example a Saturday nearest October 31) are not
    rejected as non-year-end facts after the selector shifts dates before the
    15th into the previous fiscal month.
    """
    try:
        fallback = int(fallback_month or 12)
    except Exception:
        fallback = 12
    months = []
    try:
        iterable = list(filings_10k.head(8)) if hasattr(filings_10k, 'head') else list(filings_10k)[:8]
    except Exception:
        iterable = []
    for filing in iterable:
        period = None
        for attr in ("period_of_report", "period_end", "report_date"):
            try:
                period = getattr(filing, attr, None)
            except Exception:
                period = None
            if period:
                break
        month, _year = _native_annual_effective_month_year(period)
        if month is not None:
            months.append(int(month))
    if not months:
        return fallback
    counts = pd.Series(months).value_counts()
    effective = int(counts.index[0])
    # Require at least two observations when available before overriding the
    # raw/profile month; with a single filing, the effective month is still the
    # safest annual selector because it matches get_period_info exactly.
    return effective


def _native_annual_fy_from_end(end_dt, ye_month):
    month, year = _native_annual_effective_month_year(end_dt)
    if month is None:
        return None
    return int(year + (month > int(ye_month)))


def _native_annual_is_fiscal_year_end(row, ye_month):
    end_dt = row.get("_End_dt") if hasattr(row, "get") else None
    if pd.isna(end_dt):
        return False
    month, _year = _native_annual_effective_month_year(end_dt)
    if month != int(ye_month):
        return False
    try:
        return int(row.get("FY")) == _native_annual_fy_from_end(end_dt, ye_month)
    except Exception:
        return False


def _native_annual_is_full_year_duration(duration):
    try:
        duration = float(duration)
    except Exception:
        return False
    return 330 <= duration <= 390


def _native_annual_q4_start_month(ye_month):
    try:
        ym = int(ye_month)
    except Exception:
        ym = 12
    return ((ym - 2 - 1) % 12) + 1


def _native_annual_is_ordinary_q4_stub(row, ye_month):
    """True when a short segment period looks like a normal Q4-only fact.

    A reported new segment may have a real fiscal-year-ending stub, but an
    ordinary Q4 segment fact commonly starts on day 1 of the fiscal Q4 start
    month and lasts about one quarter.  Rejecting only that boundary-shaped case
    preserves mid-year and late-year starts such as Oct 15-Dec 31 while avoiding
    the most likely false positive.
    """
    try:
        dur = float(row.get('_Duration', 0) or 0)
        start_dt = pd.to_datetime(row.get('_Start_dt'), errors='coerce')
        if pd.isna(start_dt):
            return False
        q4_start_month = _native_annual_q4_start_month(ye_month)
        return 75 <= dur <= 105 and int(start_dt.month) == q4_start_month and int(start_dt.day) <= 7
    except Exception:
        return False


def _native_annual_effective_month_year_series(end_series):
    """Vectorized equivalent of _native_annual_effective_month_year."""
    end = pd.to_datetime(end_series, errors='coerce')
    effective = end.where(end.dt.day >= 15, end - pd.DateOffset(months=1))
    return effective.dt.month, effective.dt.year


def _native_annual_fiscal_year_end_mask_vectorized(df, ye_month):
    """Vectorized equivalent of _native_annual_is_fiscal_year_end."""
    if df is None or len(df) == 0:
        return pd.Series(False, index=getattr(df, 'index', None), dtype=bool)
    end = df['_End_dt'] if '_End_dt' in df.columns else pd.to_datetime(df.get('End'), errors='coerce')
    effective_month, effective_year = _native_annual_effective_month_year_series(end)
    fy_num = pd.to_numeric(df['FY'], errors='coerce') if 'FY' in df.columns else pd.Series(np.nan, index=df.index)
    try:
        ye_month_i = int(ye_month)
    except Exception:
        ye_month_i = 12
    if ye_month_i == 12:
        computed_fy = effective_year
    else:
        computed_fy = effective_year + (effective_month > ye_month_i).astype('int64')
    return (
        pd.to_datetime(end, errors='coerce').notna()
        & effective_month.eq(ye_month_i)
        & computed_fy.eq(fy_num)
    ).fillna(False)


def _native_annual_ordinary_q4_stub_mask_vectorized(df, ye_month):
    """Vectorized equivalent of _native_annual_is_ordinary_q4_stub."""
    if df is None or len(df) == 0:
        return pd.Series(False, index=getattr(df, 'index', None), dtype=bool)
    start = df['_Start_dt'] if '_Start_dt' in df.columns else pd.to_datetime(df.get('Start'), errors='coerce')
    start = pd.to_datetime(start, errors='coerce')
    duration = pd.to_numeric(df['_Duration'] if '_Duration' in df.columns else df.get('Duration'), errors='coerce').fillna(0.0)
    q4_start_month = _native_annual_q4_start_month(ye_month)
    return (
        start.notna()
        & start.dt.month.eq(q4_start_month)
        & start.dt.day.le(7)
        & duration.between(75, 105, inclusive='both')
    ).fillna(False)


def _prepare_native_annual_df(all_facts):
    """Create the native annual working DataFrame and reusable typed columns."""
    df = pd.DataFrame(all_facts)
    if df.empty:
        return df

    df['Value'] = pd.to_numeric(df.get('Value'), errors='coerce')
    df = df.dropna(subset=['Value']).copy()
    if df.empty:
        return df

    for col, default in (
        ('Form', None), ('Accession', None), ('FilingUrl', None),
        ('Start', None), ('Concept', None), ('DimCount', 0),
        ('TagRank', 999), ('IsCalculated', False), ('StartEstimated', False),
    ):
        if col not in df.columns:
            df[col] = default

    df['_End_dt'] = pd.to_datetime(df['End'], errors='coerce')
    df['_Start_dt'] = pd.to_datetime(df['Start'], errors='coerce')
    df['_Filed_dt'] = pd.to_datetime(df['Filed'], errors='coerce')
    df['_Duration'] = pd.to_numeric(df['Duration'], errors='coerce').fillna(0.0)
    df['_StartWasReconstructed'] = df['StartEstimated'].fillna(False).astype(bool)
    df['_FY_num'] = pd.to_numeric(df['FY'], errors='coerce')
    df['_Form_norm'] = df['Form'].astype(str).str.upper()
    df['_FullYear'] = df['_Duration'].between(330, 390, inclusive='both')
    df['_Instant'] = (df['_Duration'].abs() <= 15) & df['_Start_dt'].isna()
    df['_SourceRank'] = df['Concept'].map(_native_annual_source_rank)
    df['_IsCalcSort'] = df['IsCalculated'].astype(bool).astype('int8')
    df['_DimSort'] = pd.to_numeric(df['DimCount'], errors='coerce').fillna(0)
    df['_TagRankNum'] = pd.to_numeric(df['TagRank'], errors='coerce').fillna(999)
    df['_LabelStr'] = df['Label'].astype(str)
    df['_Prefix'] = df['_LabelStr'].str.split(' - ').str[0]
    df['_IsSegment'] = df['Category'].isin(SEG_CATS)
    df['_IsAnnualForm'] = df['_Form_norm'].isin(_NATIVE_ANNUAL_FORMS)
    df['_IsAmendment'] = df['_Form_norm'].eq('10-K/A')
    return df


def _refresh_native_annual_label_columns(df):
    """Refresh derived label/category columns after segment label rewrites."""
    if df is None or df.empty:
        return df
    df['_LabelStr'] = df['Label'].astype(str)
    df['_Prefix'] = df['_LabelStr'].str.split(' - ').str[0]
    df['_IsSegment'] = df['Category'].isin(SEG_CATS)
    return df


def _native_annual_loose_candidate_mask(df, ye_month):
    """Cheap superset filter for rows native annual mode can possibly select."""
    if df is None or len(df) == 0:
        return pd.Series(False, index=getattr(df, 'index', None), dtype=bool)
    fy_end = _native_annual_fiscal_year_end_mask_vectorized(df, ye_month)
    duration = pd.to_numeric(df['_Duration'], errors='coerce').fillna(0.0)
    is_segment = df['Category'].isin(SEG_CATS)
    label_str = df['Label'].astype(str)
    full_year_possible = duration.between(330, 390, inclusive='both')
    instant_possible = (duration.abs() <= 15) & df['_Start_dt'].isna()
    segment_stub_possible = (
        is_segment
        & label_str.str.startswith('Revenue - ', na=False)
        & duration.ge(60)
        & duration.lt(330)
        & df['_Start_dt'].notna()
    )
    return (df['_Form_norm'].isin(_NATIVE_ANNUAL_FORMS) & fy_end & (
        full_year_possible | instant_possible | segment_stub_possible
    )).fillna(False)


def _native_annual_requested_fy_sets(selected, limit):
    """Return visible FYs and working FYs; working includes one prior dependency year."""
    years = sorted(
        pd.to_numeric(selected['FY'], errors='coerce').dropna().astype(int).unique(),
        reverse=True,
    )
    if not years:
        return set(), set()
    if limit is None:
        return set(years), set(years)
    try:
        n = max(1, int(limit))
    except Exception:
        n = 1
    visible = years[:n]
    working = years[:n + 1]
    return set(visible), set(working)


class _AnnualStageTimer:
    """Optional annual-build profiler, enabled only with SEC_PROFILE=1."""
    def __init__(self):
        self.enabled = os.environ.get('SEC_PROFILE', '').strip().lower() in ('1', 'true', 'yes', 'on')
        self.rows = []

    def stage(self, name):
        timer = self
        class _Stage:
            def __enter__(self_inner):
                self_inner.t0 = time.perf_counter() if timer.enabled else None
                return self_inner
            def __exit__(self_inner, exc_type, exc, tb):
                if timer.enabled and self_inner.t0 is not None:
                    timer.rows.append((name, time.perf_counter() - self_inner.t0))
                return False
        return _Stage()

    def emit(self):
        if not self.enabled or not self.rows:
            return
        total = sum(v for _n, v in self.rows)
        print('  [Annual Profile] ' + ' | '.join(f'{n}={v:.3f}s' for n, v in self.rows) + f' | total={total:.3f}s')


def _native_annual_source_rank(concept):
    c = str(concept or "").lower()
    if "synthetic" in c:
        return 3
    if "html" in c or "fallback" in c:
        return 2
    if not c or c == "nan":
        return 1
    return 0


def _native_annual_segment_is_disclosure(label, protected_labels=None):
    normalized_label = re.sub(r'\s+', ' ', str(label or '')).strip().casefold()
    if protected_labels and normalized_label in protected_labels:
        return False
    parts = [p.strip() for p in str(label or "").split(' - ')]
    if not parts:
        return False
    prefix = parts[0]
    if prefix not in GENUINE_SEGMENT_METRICS:
        return True
    if len(parts) >= 2:
        member_full = ' - '.join(parts[1:]).lower()
        if any(kw == member_full for kw in ('operating segments', 'reportable segments',
                                            'consolidated', 'total', 'all segments')):
            return True
        if any(pat in member_full for pat in DISCLOSURE_PATTERNS):
            return True
        ratio_kws = {'percent', 'percentage', 'ratio', 'weighted average'}
        if any(kw in member_full for kw in ratio_kws) and 'Concentration %' not in prefix:
            return True
        if '_' in member_full:
            return True
    return False


def _audit_income_statement_sort(final_pivot, is_item_order=None, is_financial=False,
                                 is_insurance=False, max_warnings=8):
    """Warn when an income-statement row lands in an implausible section.

    Audit only: it never edits data or row order.  It catches future custom
    labels that escape the sorter, especially rows below Net Income and rows
    whose computed structural bucket disagrees sharply with their displayed
    location.  The sorter remains deterministic and company-agnostic.
    """
    if final_pivot is None or final_pivot.empty:
        return final_pivot
    try:
        idx = list(final_pivot.index)
    except Exception:
        return final_pivot

    labels = [label for cat, label in idx if cat == '1_Income_Statement']
    if not labels:
        return final_pivot

    order = is_item_order or {}
    positions = {label: i for i, label in enumerate(labels)}
    net_positions = [i for i, label in enumerate(labels)
                     if str(label).split(' - ')[0] == 'Net Income']
    net_i = min(net_positions) if net_positions else None

    # Rows allowed below Net Income.
    below_net_allowed = (
        'eps', 'earnings per share', 'per share',
        'share outstanding', 'shares outstanding', 'weighted average share',
        'antidilutive', 'preferred stock dividend', 'preferred dividends',
        'redemption premium', 'noncontrolling', 'minority interest',
        'discontinued operation',
    )

    suspicious_below_net = []
    if net_i is not None:
        for i, label in enumerate(labels[net_i + 1:], start=net_i + 1):
            base = str(label).split(' - ')[0]
            l = base.lower()
            if any(k in l for k in below_net_allowed):
                continue
            pos = order.get(base)
            if pos is None:
                pos = _evidence_is_pos(base, is_financial, is_insurance)
            if pos < 31.3:
                suspicious_below_net.append(base)

    # Structural monotonicity audit.  Compare the row's displayed section to
    # the best generic evidence bucket.  This detects mistakes such as revenue
    # detail rows after cost rows, cost rows below gross profit, or operating
    # expenses below operating income before they silently ship.
    suspicious_sections = []
    for label in labels:
        base = str(label).split(' - ')[0]
        l = base.lower()
        if any(k in l for k in below_net_allowed):
            continue
        # Use generic evidence for the audit section.  The active display
        # order can be intentionally compacted for financial/insurance filers,
        # so its raw numeric slots should not be interpreted as semantic buckets.
        pos = _evidence_is_pos(base, is_financial, is_insurance)
        section = _is_sort_section(pos)
        displayed_section = _is_sort_section(float(positions[label]))
        # The displayed ordinal and structural bucket are not on the same scale
        # for compact filers, so only warn for clear subtotal-boundary violations
        # using nearby anchors.
        if section == 0:  # revenue/top-line
            blockers = ('Cost of Revenue', 'Gross Profit', 'Research & Development',
                        'Total Operating Expenses', 'Operating Income', 'Net Income')
            if any(b in positions and positions[label] > positions[b] for b in blockers):
                suspicious_sections.append(base)
        elif section == 1:  # cost/gross-profit block
            blockers = ('Research & Development', 'Total Operating Expenses',
                        'Operating Income', 'Net Income')
            if any(b in positions and positions[label] > positions[b] for b in blockers):
                suspicious_sections.append(base)
        elif section == 2:  # operating expense detail
            blockers = ('Total Operating Expenses', 'Operating Income', 'Net Income')
            if any(b in positions and positions[label] > positions[b] for b in blockers):
                suspicious_sections.append(base)
        elif section == 4 and not (_IBM_STYLE_STATE.get('active') and not is_financial and not is_insurance):
            # Non-operating rows should generally not appear above Operating Income.
            if 'Operating Income' in positions and positions[label] < positions['Operating Income']:
                suspicious_sections.append(base)

    if suspicious_below_net:
        shown = ', '.join(dict.fromkeys(suspicious_below_net[:max_warnings]))
        extra = '' if len(suspicious_below_net) <= max_warnings else f" (+{len(suspicious_below_net)-max_warnings} more)"
        print(f"  [Sort Audit] Possible income-statement row(s) below Net Income: {shown}{extra}")

    suspicious_sections = [x for x in dict.fromkeys(suspicious_sections)
                           if x not in dict.fromkeys(suspicious_below_net)]
    if suspicious_sections:
        shown = ', '.join(suspicious_sections[:max_warnings])
        extra = '' if len(suspicious_sections) <= max_warnings else f" (+{len(suspicious_sections)-max_warnings} more)"
        print(f"  [Sort Audit] Possible income-statement row(s) outside expected section: {shown}{extra}")

    return final_pivot


def _sort_value_signature(final_pivot):
    """Order-independent fingerprint of every row and value.

    Used as a guard around display-only sorting.  It treats NaN/NA values as a
    stable token so a sort pass cannot silently alter, drop, duplicate, or
    rewrite numeric cells.
    """
    if final_pivot is None:
        return None
    records = []
    for idx, row in final_pivot.iterrows():
        norm = []
        for value in row.tolist():
            try:
                if pd.isna(value):
                    norm.append(('NA', None))
                    continue
            except Exception:
                pass
            if isinstance(value, np.generic):
                value = value.item()
            norm.append((type(value).__name__, value))
        records.append((tuple(idx) if isinstance(idx, tuple) else (idx,), tuple(norm)))
    records.sort(key=lambda x: (repr(x[0]), repr(x[1])))
    return records


def _sort_preserving_values(final_pivot, sort_key, is_item_order=None,
                            is_financial=False, is_insurance=False,
                            context="final output"):
    """Apply a row-order-only sort and verify the data payload is unchanged."""
    if final_pivot is None or final_pivot.empty:
        return final_pivot
    before_sig = _sort_value_signature(final_pivot)
    sorted_pivot = final_pivot.reindex(sorted(final_pivot.index, key=sort_key))
    after_sig = _sort_value_signature(sorted_pivot)
    if before_sig != after_sig:
        print(f"  [Sort Guard] WARNING: {context} sort was skipped because it would change row/value payload.")
        return _audit_face_statement_sort(final_pivot, is_item_order, is_financial, is_insurance)
    return _audit_face_statement_sort(sorted_pivot, is_item_order, is_financial, is_insurance)


def _sort_final_output_pivot(final_pivot, is_financial=False, is_insurance=False,
                             context="final output"):
    """Authoritative final display-order pass for both fresh and cached outputs.

    This is deliberately row-order-only and is safe to run immediately before
    writing CSV/XLSX.  It prevents old final-pivot cache entries or late repair
    passes from preserving a stale/bad display order.
    """
    if final_pivot is None or final_pivot.empty:
        return final_pivot

    base_order = {n: i for i, n in enumerate(CONCEPT_MAP.keys())}
    item_order = dict(base_order)
    statement_order_maps = {
        '2_Balance_Sheet': _build_statement_order('2_Balance_Sheet'),
        '3_Cash_Flow': _build_statement_order('3_Cash_Flow'),
    }
    is_item_order = _build_item_order(is_insurance, is_financial)
    item_order = _apply_presentation_anchors(item_order, is_financial, is_insurance)
    is_item_order = _apply_presentation_anchors(is_item_order, is_financial, is_insurance)

    # Preserve the existing non-IS correction used by the normal final sorter:
    # Goodwill belongs with non-current assets, not in the liabilities block.
    if 'Goodwill' in item_order:
        if 'Intangible Assets (Net)' in item_order:
            item_order['Goodwill'] = item_order['Intangible Assets (Net)'] + 0.5
        elif 'Intangible Assets & Goodwill' in item_order:
            item_order['Goodwill'] = item_order['Intangible Assets & Goodwill'] + 0.5
        elif 'Total Assets' in item_order:
            item_order['Goodwill'] = item_order['Total Assets'] - 0.5

    # Preserve IBM-style statement ordering in the authoritative pre-write pass.
    if (
        _IBM_STYLE_STATE.get('active')
        and not is_financial
        and not is_insurance
        and 'Total Operating Expenses' in is_item_order
    ):
        _toe_pos = is_item_order['Total Operating Expenses']
        for _blk_lbl, _off in (('Other Income / (Expense)', 0.03),
                               ('Gain/Loss on Investments', 0.02),
                               ('Interest Expense', 0.01)):
            if _blk_lbl in is_item_order:
                is_item_order[_blk_lbl] = _toe_pos - _off

    for k in final_pivot.index.get_level_values('Label').unique():
        base = str(k).split(' - ')[0]
        if base not in item_order:
            item_order[base] = 999
        if base not in is_item_order:
            is_item_order[base] = _evidence_is_pos(base, is_financial, is_insurance)

    seg_met_order = {m: i for i, m in enumerate(SEGMENT_METRIC_ORDER)}
    if is_insurance:
        dynamic_kpis = ['Loss Ratio (%)', 'Expense Ratio (%)', 'Combined Ratio (%)', 'Net Margin (%)'] + KPI_ORDER
    elif is_financial:
        dynamic_kpis = ['Efficiency Ratio (%)', 'Net Margin (%)'] + KPI_ORDER
    else:
        dynamic_kpis = ['Gross Margin (%)', 'Operating Margin (%)', 'Net Margin (%)'] + KPI_ORDER
    kpi_order = {k: i for i, k in enumerate(dynamic_kpis)}
    for alias, canonical in KPI_ORDER_ALIASES.items():
        if canonical in kpi_order:
            kpi_order[alias] = kpi_order[canonical]
    for k in final_pivot.index.get_level_values('Label').unique():
        if k not in kpi_order:
            kpi_order[k] = 99

    business_segment_labels = [idx[1] for idx in final_pivot.index if idx[0] == '4a_Segments_Business']
    business_segment_context = _business_segment_context(business_segment_labels)

    adjustment_sort_map = {
        'Gross Profit: Other Adjustments': 'Gross Profit',
        'Operating Income: Other Adjustments': 'Operating Income',
        'Pretax Income: Other Adjustments': 'Pretax Income',
        'Net Income: Other Adjustments': 'Net Income',
        'Other Operating Adjustments (Net)': 'Operating Cash Flow',
        'Other Investing Adjustments (Net)': 'Investing Cash Flow',
        'Other Financing Adjustments (Net)': 'Financing Cash Flow',
    }

    def sort_key(idx):
        cat, label = idx
        label = str(label)
        if label in adjustment_sort_map:
            _anchor = adjustment_sort_map[label]
            if cat == '1_Income_Statement':
                _anchor_pos = is_item_order.get(_anchor, item_order.get(_anchor, 999))
            elif cat in ('2_Balance_Sheet', '3_Cash_Flow'):
                _anchor_pos = _statement_order_for_label(cat, _anchor, statement_order_maps)
            else:
                _anchor_pos = item_order.get(_anchor, 999)
            return (CAT_ORDER.get(cat, 99), _anchor_pos, -1, label)
        base = label.split(' - ')[0]
        if cat == '1_Income_Statement':
            if base not in is_item_order:
                is_item_order[base] = _evidence_is_pos(base, is_financial, is_insurance)
            return (CAT_ORDER.get(cat, 99), is_item_order.get(base, 999),
                    0 if ' - ' not in label else 1, label)
        if cat == '4a_Segments_Business':
            return (CAT_ORDER.get(cat, 99), *_segment_business_sort_key(label, business_segment_context))
        if cat == '4a_Segments_Business':
            return (CAT_ORDER.get(cat, 99), *_segment_business_sort_key(label, business_segment_context))
        if cat in SEG_CATS or cat == '7_Concentration_Risk':
            return (CAT_ORDER.get(cat, 99), seg_met_order.get(base, 999), label)
        if cat == '5_KPI_Metrics':
            return (CAT_ORDER.get(cat, 99), kpi_order.get(label, 999), label)
        if cat in ('2_Balance_Sheet', '3_Cash_Flow'):
            return (CAT_ORDER.get(cat, 99),
                    _statement_order_for_label(cat, label, statement_order_maps),
                    0 if ' - ' not in label else 1, label)
        return (CAT_ORDER.get(cat, 99), item_order.get(base, 999),
                0 if ' - ' not in label else 1, label)

    return _sort_preserving_values(final_pivot, sort_key, is_item_order,
                                   is_financial, is_insurance, context)


def _normalize_output_margin_rows(final_pivot):
    """Expose one public row for each margin in CSV/XLSX output.

    KPI construction still uses the older ``Metric: ...`` rows internally,
    while the industry KPI pass creates friendlier display labels.  Keeping
    both in the exported file made the same gross, operating, and net margins
    appear twice.  Consolidate those aliases only at the output boundary so
    downstream calculations remain unchanged and CSV consumers get one stable,
    human-readable label per metric.
    """
    if final_pivot is None or final_pivot.empty:
        return final_pivot

    aliases = {
        'Metric: Gross Margin %': 'Gross Margin (%)',
        'Metric: EBIT Margin %': 'Operating Margin (%)',
        'Metric: Net Margin %': 'Net Margin (%)',
        'Metric: FCF Margin %': 'FCF Margin (%)',
    }
    result = final_pivot.copy()
    for legacy_label, public_label in aliases.items():
        legacy_idx = ('5_KPI_Metrics', legacy_label)
        public_idx = ('5_KPI_Metrics', public_label)
        if legacy_idx not in result.index:
            continue
        if public_idx in result.index:
            # Prefer the display row and use the legacy calculation only to
            # fill periods that the display pass did not populate.
            result.loc[public_idx, :] = result.loc[public_idx, :].combine_first(
                result.loc[legacy_idx, :]
            )
            result = result.drop(index=legacy_idx)
        else:
            result = result.rename(index={legacy_label: public_label}, level='Label')
    return result

def _native_annual_sort_output(final_pivot, is_financial=False, is_insurance=False):
    return _sort_final_output_pivot(
        final_pivot,
        is_financial=is_financial,
        is_insurance=is_insurance,
        context="native annual final sort",
    )

def _snapshot_native_mutable_state():
    """Capture per-run mutable learning/bridge state before isolated annual mode."""
    return {
        'CONCEPT_MAP': copy.deepcopy(CONCEPT_MAP),
        'STANDARD_TAG_MAP': copy.deepcopy(STANDARD_TAG_MAP),
        '_CONCEPT_TAG_TO_LABEL': copy.deepcopy(_CONCEPT_TAG_TO_LABEL),
        '_NORMALIZED_LABEL_INDEX': copy.deepcopy(_NORMALIZED_LABEL_INDEX),
        '_AUTO_SEEN_CONCEPTS': set(_AUTO_SEEN_CONCEPTS),
        '_FACE_PRESENTED': copy.deepcopy(_FACE_PRESENTED),
        '_FACE_PRESENTATION_POS': copy.deepcopy(_FACE_PRESENTATION_POS),
        '_RESOLVE_OVERRIDDEN': set(_RESOLVE_OVERRIDDEN),
        'GLOBAL_CALC_PARENT': copy.deepcopy(GLOBAL_CALC_PARENT),
        '_CF_BRIDGE_SPEC': copy.deepcopy(_CF_BRIDGE_SPEC),
        '_BRIDGE_USED_LABELS': set(_BRIDGE_USED_LABELS),
        '_IBM_STYLE_STATE': dict(_IBM_STYLE_STATE),
        '_FUZZY_CACHE': copy.deepcopy(_FUZZY_CACHE),
        '_ANCHOR_SEQ': copy.deepcopy(_ANCHOR_SEQ),
    }


def _restore_native_mutable_state(snapshot):
    if not snapshot:
        return
    CONCEPT_MAP.clear(); CONCEPT_MAP.update(snapshot['CONCEPT_MAP'])
    STANDARD_TAG_MAP.clear(); STANDARD_TAG_MAP.update(snapshot['STANDARD_TAG_MAP'])
    _CONCEPT_TAG_TO_LABEL.clear(); _CONCEPT_TAG_TO_LABEL.update(snapshot['_CONCEPT_TAG_TO_LABEL'])
    _NORMALIZED_LABEL_INDEX.clear(); _NORMALIZED_LABEL_INDEX.update(snapshot['_NORMALIZED_LABEL_INDEX'])
    _AUTO_SEEN_CONCEPTS.clear(); _AUTO_SEEN_CONCEPTS.update(snapshot['_AUTO_SEEN_CONCEPTS'])
    _FACE_PRESENTED.clear(); _FACE_PRESENTED.update(snapshot['_FACE_PRESENTED'])
    _FACE_PRESENTATION_POS.clear(); _FACE_PRESENTATION_POS.update(snapshot.get('_FACE_PRESENTATION_POS', {}))
    _RESOLVE_OVERRIDDEN.clear(); _RESOLVE_OVERRIDDEN.update(snapshot['_RESOLVE_OVERRIDDEN'])
    GLOBAL_CALC_PARENT.clear(); GLOBAL_CALC_PARENT.update(snapshot['GLOBAL_CALC_PARENT'])
    _CF_BRIDGE_SPEC.clear(); _CF_BRIDGE_SPEC.update(snapshot['_CF_BRIDGE_SPEC'])
    _BRIDGE_USED_LABELS.clear(); _BRIDGE_USED_LABELS.update(snapshot['_BRIDGE_USED_LABELS'])
    _IBM_STYLE_STATE.clear(); _IBM_STYLE_STATE.update(snapshot['_IBM_STYLE_STATE'])
    if '_FUZZY_CACHE' in snapshot:
        _FUZZY_CACHE.clear(); _FUZZY_CACHE.update(snapshot['_FUZZY_CACHE'])
    if '_ANCHOR_SEQ' in snapshot:
        _ANCHOR_SEQ.clear(); _ANCHOR_SEQ.update(snapshot['_ANCHOR_SEQ'])


# ---------------------------------------------------------------------------
# Persistent native parsed-extraction cache
# ---------------------------------------------------------------------------
# EDGAR filings are immutable by accession.  The expensive part of a repeat
# native run is not only downloading filings; it is rebuilding XBRL/HTML facts
# and the learned accounting/tag state produced while extracting those facts.
# This cache stores the extraction checkpoint after all selected filings have
# been parsed, then restores that exact checkpoint on the next identical run.
_NATIVE_EXTRACTION_CACHE_VERSION = "2026-07-14.native-extraction.v8-annual-consolidation-persist"
_NATIVE_EXTRACTION_CACHE_DISABLED = {"0", "false", "no", "off", "disable", "disabled"}
_NATIVE_EXTRACTION_CACHE_ENABLED = (
    os.environ.get("SEC_NATIVE_EXTRACTION_CACHE", "1").strip().lower()
    not in _NATIVE_EXTRACTION_CACHE_DISABLED
)
_NATIVE_EXTRACTION_CACHE_ROOT = os.path.abspath(
    os.environ.get("SEC_NATIVE_EXTRACTION_CACHE_DIR",
                   _script_cache_dir("native_extraction"))
)
_NATIVE_EXTRACTION_CACHE_MISS = object()
_NATIVE_EXTRACTION_CACHE_LOCK = threading.RLock()


def _native_extraction_source_fingerprint():
    # Reuse the existing full-script fingerprint helper.  This makes cached
    # parsed facts conservative during development: any script edit invalidates
    # old extraction checkpoints instead of silently replaying stale logic.
    try:
        return _fx_code_fingerprint()
    except Exception:
        try:
            with open(__file__, "rb") as fh:
                return hashlib.sha256(fh.read()).hexdigest()
        except Exception:
            return _NATIVE_EXTRACTION_CACHE_VERSION


def _native_extraction_filing_identity(filing):
    meta = _get_filing_local_metadata(filing)
    acc = meta.get("accession_no") or meta.get("accession_number")
    if not acc:
        # Object-id fallback is intentionally rejected for persistent cache keys:
        # it is process-local and would not be stable across CLI re-runs.
        return None
    return (
        str(meta.get("form") or ""),
        str(acc),
        str(meta.get("filing_date") or ""),
    )


def _native_extraction_cache_key(ticker, company, mode, limit, ye_month,
                                 use_arelle, filings):
    if not _NATIVE_EXTRACTION_CACHE_ENABLED:
        return None
    filing_ids = []
    for filing in filings or []:
        ident = _native_extraction_filing_identity(filing)
        if ident is None:
            return None
        filing_ids.append(ident)
    if not filing_ids:
        return None
    try:
        cik = str(getattr(company, "cik", "") or "")
    except Exception:
        cik = ""
    key = {
        "cache_version": _NATIVE_EXTRACTION_CACHE_VERSION,
        "source_fingerprint": _native_extraction_source_fingerprint(),
        "ticker": str(ticker or "").upper(),
        "cik": cik,
        "mode": str(mode or ""),
        "limit": int(limit or 0),
        "ye_month": int(ye_month or 12),
        "use_arelle": bool(use_arelle),
        "filings": tuple(filing_ids),
    }
    return _fx_freeze_for_cache(key)


def _native_extraction_cache_path(key):
    if not _NATIVE_EXTRACTION_CACHE_ENABLED or key is None:
        return None
    try:
        key_blob = pickle.dumps((_NATIVE_EXTRACTION_CACHE_VERSION, key), protocol=4)
    except Exception:
        key_blob = repr((_NATIVE_EXTRACTION_CACHE_VERSION, key)).encode("utf-8", "replace")
    digest = hashlib.sha256(key_blob).hexdigest()
    return os.path.join(_NATIVE_EXTRACTION_CACHE_ROOT, digest[:2], f"{digest}.pkl")


def _native_extraction_cache_get(key):
    path = _native_extraction_cache_path(key)
    if not path:
        return _NATIVE_EXTRACTION_CACHE_MISS
    try:
        with _NATIVE_EXTRACTION_CACHE_LOCK:
            if not os.path.exists(path):
                return _NATIVE_EXTRACTION_CACHE_MISS
            with open(path, "rb") as fh:
                payload = pickle.load(fh)
        if not isinstance(payload, dict):
            return _NATIVE_EXTRACTION_CACHE_MISS
        if payload.get("version") != _NATIVE_EXTRACTION_CACHE_VERSION:
            return _NATIVE_EXTRACTION_CACHE_MISS
        value = payload.get("value")
        if not isinstance(value, dict):
            return _NATIVE_EXTRACTION_CACHE_MISS
        if not isinstance(value.get("all_facts"), list):
            return _NATIVE_EXTRACTION_CACHE_MISS
        if not isinstance(value.get("period_dates"), dict):
            return _NATIVE_EXTRACTION_CACHE_MISS
        if not isinstance(value.get("native_state"), dict):
            return _NATIVE_EXTRACTION_CACHE_MISS
        _profile_count("native_extraction_cache_hits")
        return value
    except Exception as exc:
        _profile_count("native_extraction_cache_read_failures")
        try:
            print(f"  [Cache] Native extraction cache ignored ({type(exc).__name__}).")
        except Exception:
            pass
        return _NATIVE_EXTRACTION_CACHE_MISS


def _native_extraction_cache_set(key, all_facts, period_dates, native_state,
                                 metadata=None):
    path = _native_extraction_cache_path(key)
    if not path or key is None or not all_facts or not native_state:
        return
    payload = {
        "version": _NATIVE_EXTRACTION_CACHE_VERSION,
        "created_at": time.time(),
        "value": {
            "all_facts": all_facts,
            "period_dates": dict(period_dates or {}),
            "native_state": native_state,
            "metadata": dict(metadata or {}),
        },
    }
    tmp = f"{path}.{os.getpid()}.{threading.get_ident()}.tmp"
    try:
        with _NATIVE_EXTRACTION_CACHE_LOCK:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(tmp, "wb") as fh:
                pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)
            os.replace(tmp, path)
        _profile_count("native_extraction_cache_writes")
    except Exception as exc:
        _profile_count("native_extraction_cache_write_failures")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        try:
            print(f"  [Cache] Native extraction cache write skipped ({type(exc).__name__}).")
        except Exception:
            pass


def _restore_cached_native_extraction(cache_value, all_facts, period_dates):
    # Mutate the caller-owned lists/dicts so the downstream build path receives
    # the same shapes as a live extraction run.  Restore learned state before
    # building; later accounting passes can then behave exactly as they do after
    # a normal extraction phase.
    _restore_native_mutable_state(cache_value.get("native_state"))
    all_facts[:] = list(cache_value.get("all_facts") or [])
    period_dates.clear()
    period_dates.update(dict(cache_value.get("period_dates") or {}))


# ---------------------------------------------------------------------------
# Persistent final-pivot cache
# ---------------------------------------------------------------------------
# This is deliberately a checkpoint cache, not a final-file cache.  The script
# still writes CSV/XLSX normally.  The cached object is the fully repaired
# DataFrame that would otherwise be recomputed from the same extracted facts.
_FINAL_PIVOT_CACHE_VERSION = "2026-07-14.final-pivot.v19-geography-alias-consolidation"
_FINAL_PIVOT_CACHE_DISABLED = {"0", "false", "no", "off", "disable", "disabled"}
_FINAL_PIVOT_CACHE_ENABLED = (
    os.environ.get("SEC_FINAL_PIVOT_CACHE", "1").strip().lower()
    not in _FINAL_PIVOT_CACHE_DISABLED
)
_FINAL_PIVOT_CACHE_ROOT = os.path.abspath(
    os.environ.get("SEC_FINAL_PIVOT_CACHE_DIR",
                   _script_cache_dir("final_pivot"))
)
_FINAL_PIVOT_CACHE_MISS = object()
_FINAL_PIVOT_CACHE_LOCK = threading.RLock()


def _final_pivot_cache_key(ticker, company, mode, limit, ye_month, use_arelle,
                           filings, company_name=None, is_financial=False,
                           is_insurance=False, is_oil_gas=False, is_reit=False):
    if not _FINAL_PIVOT_CACHE_ENABLED:
        return None
    filing_ids = []
    for filing in filings or []:
        ident = _native_extraction_filing_identity(filing)
        if ident is None:
            return None
        filing_ids.append(ident)
    if not filing_ids:
        return None
    try:
        cik = str(getattr(company, "cik", "") or "")
    except Exception:
        cik = ""
    key = {
        "cache_version": _FINAL_PIVOT_CACHE_VERSION,
        "source_fingerprint": _native_extraction_source_fingerprint(),
        "ticker": str(ticker or "").upper(),
        "cik": cik,
        "mode": str(mode or ""),
        "limit": int(limit or 0),
        "ye_month": int(ye_month or 12),
        "use_arelle": bool(use_arelle),
        "company_name": str(company_name or ""),
        "is_financial": bool(is_financial),
        "is_insurance": bool(is_insurance),
        "is_oil_gas": bool(is_oil_gas),
        "is_reit": bool(is_reit),
        "pandas": getattr(pd, "__version__", ""),
        "numpy": getattr(np, "__version__", ""),
        "filings": tuple(filing_ids),
    }
    return _fx_freeze_for_cache(key)


def _final_pivot_cache_path(key):
    if not _FINAL_PIVOT_CACHE_ENABLED or key is None:
        return None
    try:
        key_blob = pickle.dumps((_FINAL_PIVOT_CACHE_VERSION, key), protocol=4)
    except Exception:
        key_blob = repr((_FINAL_PIVOT_CACHE_VERSION, key)).encode("utf-8", "replace")
    digest = hashlib.sha256(key_blob).hexdigest()
    return os.path.join(_FINAL_PIVOT_CACHE_ROOT, digest[:2], f"{digest}.pkl")


def _final_pivot_cache_get(key):
    path = _final_pivot_cache_path(key)
    if not path:
        return _FINAL_PIVOT_CACHE_MISS
    try:
        with _FINAL_PIVOT_CACHE_LOCK:
            if not os.path.exists(path):
                return _FINAL_PIVOT_CACHE_MISS
            with open(path, "rb") as fh:
                payload = pickle.load(fh)
        if not isinstance(payload, dict):
            return _FINAL_PIVOT_CACHE_MISS
        if payload.get("version") != _FINAL_PIVOT_CACHE_VERSION:
            return _FINAL_PIVOT_CACHE_MISS
        df = payload.get("value")
        if not isinstance(df, pd.DataFrame):
            return _FINAL_PIVOT_CACHE_MISS
        _profile_count("final_pivot_cache_hits")
        return df.copy(deep=True)
    except Exception as exc:
        _profile_count("final_pivot_cache_read_failures")
        try:
            print(f"  [Cache] Final pivot cache ignored ({type(exc).__name__}).")
        except Exception:
            pass
        return _FINAL_PIVOT_CACHE_MISS


def _final_pivot_cache_set(key, final_pivot, metadata=None):
    path = _final_pivot_cache_path(key)
    if not path or key is None or final_pivot is None or final_pivot.empty:
        return
    payload = {
        "version": _FINAL_PIVOT_CACHE_VERSION,
        "created_at": time.time(),
        "metadata": dict(metadata or {}),
        "value": final_pivot.copy(deep=True),
    }
    tmp = f"{path}.{os.getpid()}.{threading.get_ident()}.tmp"
    try:
        with _FINAL_PIVOT_CACHE_LOCK:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(tmp, "wb") as fh:
                pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)
            os.replace(tmp, path)
        _profile_count("final_pivot_cache_writes")
    except Exception as exc:
        _profile_count("final_pivot_cache_write_failures")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        try:
            print(f"  [Cache] Final pivot cache write skipped ({type(exc).__name__}).")
        except Exception:
            pass



# ---------------------------------------------------------------------------
# Native --annual only: consolidated-entity roll-up rescue
# ---------------------------------------------------------------------------
# Some historical 10-Ks (redomiciliations, guarantor structures, predecessor/
# successor filings) tag the audited face-statement comparative values only as
# consolidating entity members.  The dimensionless consolidated fact can then
# be absent or zero, while the member rows already present in the annual pivot
# sum exactly to the filed consolidated subtotal.
#
# This repair is deliberately narrow:
#   * native --annual path only (called only from build_annual_pivoted_data)
#   * income statement only
#   * never changes a valid nonzero consolidated winner
#   * requires an explicit consolidation-elimination member and at least two
#     entity buckets
#   * uses only exact accounting identities for derived subtotals
#   * skipped for financial/insurance issuers

_ANNUAL_CONSOLIDATION_ENTITY_TERMS = (
    'parent company',
    'guarantor subsidiaries',
    'non guarantor subsidiaries',
    'non-guarantor subsidiaries',
    'subsidiaries',
    'consolidation eliminations',
    'consolidating eliminations',
    'eliminations',
)

_ANNUAL_CONSOLIDATION_EXCLUDED_SCOPE_TERMS = (
    'semiconductor solutions', 'infrastructure software', 'wireless',
    'wired infrastructure', 'enterprise storage', 'industrial and other',
    'geographic', 'country', 'region', 'product or service',
)

def _annual_consolidation_suffix_kind(label, base_label):
    text = re.sub(r'\s+', ' ', str(label or '')).strip()
    prefix = str(base_label) + ' - '
    if not text.lower().startswith(prefix.lower()):
        return None
    suffix = text[len(prefix):].strip().lower()
    if any(term in suffix for term in _ANNUAL_CONSOLIDATION_EXCLUDED_SCOPE_TERMS):
        # Explicit elimination labels sometimes include a trailing word such as
        # "Product".  Keep those only when the consolidation role is explicit.
        if 'elimination' not in suffix:
            return None
    if 'elimination' in suffix:
        return 'elimination'
    if any(term in suffix for term in _ANNUAL_CONSOLIDATION_ENTITY_TERMS):
        return 'entity'
    return None

def _annual_consolidation_rollup_series(pivoted, base_label):
    """Return (series, evidence) for an explicit consolidating roll-up.

    A period is eligible only when at least two entity-member rows and one
    elimination row are reported.  This prevents ordinary business/geographic
    segment rows from being summed into the consolidated face statement.
    """
    if pivoted is None or pivoted.empty or not isinstance(pivoted.index, pd.MultiIndex):
        return pd.Series(dtype='float64'), {}
    rows=[]
    for idx in pivoted.index:
        if len(idx) < 2:
            continue
        cat, label = idx[0], idx[1]
        if cat not in {'4a_Segments_Business', '6_Disclosures', '1_Income_Statement'}:
            continue
        kind = _annual_consolidation_suffix_kind(label, base_label)
        if kind:
            rows.append((idx, kind))
    if not rows:
        return pd.Series(np.nan, index=pivoted.columns, dtype='float64'), {}

    result = pd.Series(np.nan, index=pivoted.columns, dtype='float64')
    evidence = {}
    for col in pivoted.columns:
        entity_values=[]
        elimination_values=[]
        used=[]
        for idx, kind in rows:
            value = pd.to_numeric(pd.Series([pivoted.at[idx, col]]), errors='coerce').iloc[0]
            if pd.isna(value):
                continue
            used.append((idx[1], float(value), kind))
            if kind == 'elimination':
                elimination_values.append(float(value))
            else:
                entity_values.append(float(value))
        nonzero_entities=[v for v in entity_values if abs(v) > 1e-9]
        if len(nonzero_entities) < 2 or not elimination_values:
            continue
        total=float(sum(entity_values) + sum(elimination_values))
        gross_abs=float(sum(abs(v) for v in entity_values + elimination_values))
        if not np.isfinite(total) or abs(total) < 1e6:
            continue
        # Reject pathological cancellation where the purported consolidated
        # amount is tiny relative to the member activity.
        if gross_abs > max(abs(total), 1.0) * 50.0:
            continue
        result[col]=total
        evidence[col]=used
    return result, evidence

def _annual_get_numeric_row(pivoted, category, label):
    idx=(category, label)
    if idx in pivoted.index:
        return pd.to_numeric(pivoted.loc[idx], errors='coerce')
    return pd.Series(np.nan, index=pivoted.columns, dtype='float64')

def _annual_set_numeric_cell(pivoted, category, label, period, value):
    idx=(category, label)
    if idx not in pivoted.index:
        pivoted.loc[idx, :] = np.nan
    pivoted.at[idx, period] = float(value)

def _apply_native_annual_consolidation_rollup_rescue(
        pivoted, is_financial=False, is_insurance=False):
    """Repair missing/zero annual face subtotals from consolidating members.

    The function is intentionally not used by quarterly or foreign-filer paths.
    It never overwrites a valid nonzero face value.
    """
    if pivoted is None or pivoted.empty or is_financial or is_insurance:
        return pivoted
    out=pivoted.copy()

    gp_rollup, gp_evidence = _annual_consolidation_rollup_series(out, 'Gross Profit')
    rev = _annual_get_numeric_row(out, '1_Income_Statement', 'Revenue')
    cor = _annual_get_numeric_row(out, '1_Income_Statement', 'Cost of Revenue')
    gp = _annual_get_numeric_row(out, '1_Income_Statement', 'Gross Profit')
    opinc = _annual_get_numeric_row(out, '1_Income_Statement', 'Operating Income')
    opex = _annual_get_numeric_row(out, '1_Income_Statement', 'Total Operating Expenses')

    changed=[]
    for period in out.columns:
        gp_candidate = gp_rollup.get(period, np.nan)
        revenue = rev.get(period, np.nan)
        current_gp = gp.get(period, np.nan)
        if pd.isna(gp_candidate) or pd.isna(revenue):
            continue
        if revenue <= 0 or gp_candidate <= 0 or gp_candidate > revenue * 1.05:
            continue
        # Only an absent/zero gross-profit winner may be repaired.
        if pd.notna(current_gp) and abs(float(current_gp)) > 1e-9:
            continue

        cost_candidate = float(revenue - gp_candidate)
        if cost_candidate < 0:
            continue

        # Require exact mechanical agreement with an already valid nonzero cost
        # row when one exists.  Otherwise the exact Revenue-GP identity supplies
        # the missing cost subtotal.
        current_cor = cor.get(period, np.nan)
        tolerance=max(2_000_000.0, 0.01 * max(abs(revenue), abs(gp_candidate), abs(cost_candidate), 1.0))
        if pd.notna(current_cor) and abs(float(current_cor)) > 1e-9:
            if abs(float(current_cor) - cost_candidate) > tolerance:
                continue

        _annual_set_numeric_cell(out, '1_Income_Statement', 'Gross Profit', period, gp_candidate)
        if pd.isna(current_cor) or abs(float(current_cor)) <= 1e-9:
            _annual_set_numeric_cell(out, '1_Income_Statement', 'Cost of Revenue', period, cost_candidate)

        # Do not directly repair individual operating-expense components or
        # Operating Income from consolidating member rows.  Those disclosures
        # can use different eliminations/bases.  Once Gross Profit and Cost of
        # Revenue are restored, the existing accounting engine may derive only
        # exact subtotals through its normal guarded equations.
        changed.append((period, float(gp_candidate), float(cost_candidate), len(gp_evidence.get(period, []))))

    for period, gp_value, cost_value, n_evidence in changed:
        print(
            f"  [Annual Consolidation Rescue] {period}: Gross Profit={gp_value:,.0f}, "
            f"Cost of Revenue={cost_value:,.0f} from {n_evidence} audited "
            f"consolidating member row(s)."
        )
    return out


def _native_annual_rescue_overrides(before, after):
    """Return long-form rows changed by the annual-only rescue.

    The annual builder historically rebuilt ``final_long`` from the original
    selected facts after KPI calculation.  Without explicitly carrying these
    overrides forward, valid repairs made to the temporary pivot are silently
    discarded.  This helper persists only cells whose values actually changed.
    """
    rows = []
    if before is None or after is None or after.empty:
        return pd.DataFrame(columns=['Category', 'Label', 'Period', 'Value'])
    for idx in after.index:
        if not isinstance(idx, tuple) or len(idx) < 2:
            continue
        category, label = idx[0], idx[1]
        if category != '1_Income_Statement':
            continue
        for period in after.columns:
            new_value = pd.to_numeric(pd.Series([after.at[idx, period]]), errors='coerce').iloc[0]
            if pd.isna(new_value):
                continue
            if idx in before.index and period in before.columns:
                old_value = pd.to_numeric(pd.Series([before.at[idx, period]]), errors='coerce').iloc[0]
            else:
                old_value = np.nan
            unchanged = (pd.isna(old_value) and pd.isna(new_value)) or (
                pd.notna(old_value) and float(old_value) == float(new_value)
            )
            if unchanged:
                continue
            rows.append({
                'Category': category,
                'Label': label,
                'Period': period,
                'Value': float(new_value),
            })
    return pd.DataFrame(rows, columns=['Category', 'Label', 'Period', 'Value'])


def build_annual_pivoted_data(all_facts, ticker, ye_month, company_name=None,
                              is_financial=False, is_insurance=False,
                              is_oil_gas=False, is_reit=False, limit=None):
    with _ProfileTimer("build_annual_pivoted_data_total"):
        return _build_annual_pivoted_data_impl(
            all_facts, ticker, ye_month, company_name=company_name,
            is_financial=is_financial, is_insurance=is_insurance,
            is_oil_gas=is_oil_gas, is_reit=is_reit, limit=limit,
        )


def _build_annual_pivoted_data_impl(all_facts, ticker, ye_month, company_name=None,
                                    is_financial=False, is_insurance=False,
                                    is_oil_gas=False, is_reit=False, limit=None):
    """Build native 10-K annual FY columns without using quarterly/Q4 repair logic."""
    timer = _AnnualStageTimer()

    with timer.stage('annual_dataframe_create'):
        df = _prepare_native_annual_df(all_facts)
        if df.empty:
            return pd.DataFrame()
        df['_FiscalYearEnd'] = _native_annual_fiscal_year_end_mask_vectorized(df, ye_month)

    # Keep the existing segment-name normalization/reconciliation, but do not run
    # quarterly repairs, Q4 derivations, YTD neutralization, or sparse annual
    # segment deletion.
    with timer.stage('annual_label_normalize'):
        def normalize_segment_label(label):
            parts = str(label).split(' - ')
            if len(parts) > 1:
                prefix, mems = parts[0], [p.strip() for p in parts[1:]]
                unique_mems, seen = [], set()
                for m in mems:
                    if m.lower() not in seen:
                        unique_mems.append(m); seen.add(m.lower())
                return prefix + ' - ' + ' - '.join(sorted(unique_mems, key=str.lower))
            return label

        label_norm_map = {_lbl: normalize_segment_label(_lbl) for _lbl in df['Label'].dropna().unique()}
        df['Label'] = df['Label'].map(lambda _lbl: label_norm_map.get(_lbl, _lbl))
        df = _refresh_native_annual_label_columns(df)

    # Expensive segment reconciliation is run only on a loose annual superset.
    # The strict candidate rules below are unchanged and remain the final judge.
    with timer.stage('annual_early_filter'):
        loose_mask = _native_annual_loose_candidate_mask(df, ye_month)
        df = df.loc[loose_mask].copy()
        if df.empty:
            return pd.DataFrame()

    with timer.stage('annual_segment_reconcile'):
        df = _reconcile_segment_labels(df, ticker=ticker, company_name=company_name)
        df = _detect_and_merge_renamed_segments(df)
        df = _merge_concurrent_member_variants(df)
        df = _refresh_native_annual_label_columns(df)
        # Segment rescue/relabeling can change Category/Label, so refresh the
        # annual form/year-end flags too before strict selection.
        df['_FiscalYearEnd'] = _native_annual_fiscal_year_end_mask_vectorized(df, ye_month)
        df['_FullYear'] = df['_Duration'].between(330, 390, inclusive='both')
        df['_Instant'] = (df['_Duration'].abs() <= 15) & df['_Start_dt'].isna()
        df['_IsAnnualForm'] = df['_Form_norm'].isin(_NATIVE_ANNUAL_FORMS)
        df['_IsAmendment'] = df['_Form_norm'].eq('10-K/A')

    with timer.stage('annual_candidate_select'):
        # Reject the point-in-time EntityCommonStockSharesOutstanding fact from the
        # weighted-average share rows.  The existing label names are historical, but
        # annual mode must use the annual weighted-average duration only.
        shares_instant_leak = (
            df['_LabelStr'].isin(_NATIVE_ANNUAL_WEIGHTED_SHARE_LABELS)
            & df['Concept'].astype(str).str.contains('EntityCommonStockSharesOutstanding', case=False, na=False)
        )

        is_balance_instant = (
            (df['Category'] == '2_Balance_Sheet')
            | (df['_IsSegment'] & df['_Prefix'].isin(_NATIVE_ANNUAL_INSTANT_PREFIXES))
        )
        is_flow_annual = (
            df['Category'].isin(['1_Income_Statement', '3_Cash_Flow', '6_Disclosures', '7_Concentration_Risk'])
            | (df['_IsSegment'] & ~df['_Prefix'].isin(_NATIVE_ANNUAL_INSTANT_PREFIXES))
        )

        full_year_candidates = df[
            df['_IsAnnualForm'] & df['_FiscalYearEnd'] & df['_FullYear'] & is_flow_annual & ~shares_instant_leak
        ].copy()
        full_year_candidates['SelectionType'] = 'annual_full_year'

        instant_candidates = df[
            df['_IsAnnualForm'] & df['_FiscalYearEnd'] & df['_Instant'] & is_balance_instant
        ].copy()
        instant_candidates['SelectionType'] = 'annual_instant'

        # Segment revenue stubs support new segments that began mid-year or late-year,
        # but they are the highest false-positive risk in annual mode.  Only accept
        # true-source XBRL periods (not reconstructed Start dates) and never accept
        # an ordinary ~Q4 duration unless the filing reports no full-year fact for
        # the same segment/year.  This preserves legitimate late-year segment launches
        # while avoiding Q4-only segment facts masquerading as FY values.
        stub_candidates = df[
            df['_IsAnnualForm']
            & df['_FiscalYearEnd']
            & df['_IsSegment']
            & df['_LabelStr'].str.startswith('Revenue - ')
            & (df['_Duration'] >= 60)
            & (df['_Duration'] < 330)
            & df['_Start_dt'].notna()
            & ~df['_StartWasReconstructed']
            & (pd.to_numeric(df['DimCount'], errors='coerce').fillna(0) > 0)
            & (df['_SourceRank'] <= 1)
        ].copy()
        if not stub_candidates.empty:
            ordinary_q4_mask = _native_annual_ordinary_q4_stub_mask_vectorized(stub_candidates, ye_month)
            if ordinary_q4_mask.any():
                n_q4 = len(stub_candidates.loc[ordinary_q4_mask, ['Category', 'Label', 'FY']].drop_duplicates())
                print(f"  [Annual] Rejected {n_q4} Q4-shaped segment revenue stub(s) to avoid treating quarterly data as FY.")
                stub_candidates = stub_candidates.loc[~ordinary_q4_mask].copy()
        if not stub_candidates.empty and not full_year_candidates.empty:
            full_keys = pd.MultiIndex.from_frame(full_year_candidates[['Category', 'Label', 'FY']])
            stub_keys = pd.MultiIndex.from_frame(stub_candidates[['Category', 'Label', 'FY']])
            stub_candidates = stub_candidates.loc[~stub_keys.isin(full_keys)].copy()
        if not stub_candidates.empty:
            stub_candidates['SelectionType'] = 'annual_segment_stub'
            short_stub = stub_candidates[stub_candidates['_Duration'] < 120]
            if not short_stub.empty:
                n_short = len(short_stub[['Category', 'Label', 'FY']].drop_duplicates())
                print(f"  [Annual] Preserving {n_short} short fiscal-year-ending segment revenue stub(s) "
                      f"(<120 days) from true XBRL start/end dates; verify newly-created late-year segments manually.")

        candidates = pd.concat([full_year_candidates, instant_candidates, stub_candidates], ignore_index=True)
        if candidates.empty:
            return pd.DataFrame()

    with timer.stage('annual_best_fact_select'):
        candidates['_DurationPreference'] = np.where(
            candidates['SelectionType'].eq('annual_segment_stub'),
            -pd.to_numeric(candidates['_Duration'], errors='coerce').fillna(0),
            (pd.to_numeric(candidates['_Duration'], errors='coerce').fillna(0) - 365).abs(),
        )

        candidates = candidates.sort_values(
            ['Category', 'Label', 'FY', '_SourceRank', '_IsAmendment', '_Filed_dt',
             '_IsCalcSort', '_TagRankNum', '_DimSort', '_DurationPreference'],
            ascending=[True, True, True, True, False, False, True, True, True, True],
        )
        selected = candidates.drop_duplicates(subset=['Category', 'Label', 'FY'], keep='first').copy()
        selected['Period'] = selected['FY'].astype(int).astype(str) + '-FY'

        selected = selected.sort_values(['Category', 'Label', 'Period', '_Filed_dt'], ascending=[True, True, True, False])
        selected = selected.drop_duplicates(subset=['Category', 'Label', 'Period'], keep='first')

    with timer.stage('annual_pre_kpi_trim'):
        visible_fys, working_fys = _native_annual_requested_fy_sets(selected, limit)
        if working_fys:
            selected_working = selected[selected['FY'].astype(int).isin(working_fys)].copy()
        else:
            selected_working = selected
        if selected_working.empty:
            return pd.DataFrame()

    with timer.stage('annual_pivot'):
        pivoted_before_rescue = selected_working.pivot(
            index=['Category', 'Label'], columns='Period', values='Value')
        pivoted_temp = _apply_native_annual_consolidation_rollup_rescue(
            pivoted_before_rescue, is_financial=is_financial, is_insurance=is_insurance)
        annual_rescue_overrides = _native_annual_rescue_overrides(
            pivoted_before_rescue, pivoted_temp)

    with timer.stage('annual_kpis'):
        kpi_long = calculate_kpis(pivoted_temp, is_reit=is_reit)
        # Overrides must be last.  The normal annual builder reconstructs the
        # output from ``selected_working`` and would otherwise discard repairs
        # made only to ``pivoted_temp``.
        final_long = pd.concat([
            selected_working[['Category', 'Label', 'Period', 'Value']],
            kpi_long,
            annual_rescue_overrides,
        ], ignore_index=True)
        final_long = final_long.drop_duplicates(subset=['Category', 'Label', 'Period'], keep='last')
        final_pivot = final_long.pivot(index=['Category', 'Label'], columns='Period', values='Value')

    with timer.stage('annual_accounting_engine'):
        final_pivot = _apply_accounting_engine(final_pivot, is_financial=is_financial,
                                               is_insurance=is_insurance,
                                               is_oil_gas=is_oil_gas,
                                               is_reit=is_reit)
        final_pivot = _apply_industry_kpis(final_pivot, is_financial=is_financial, is_insurance=is_insurance)
        final_pivot = _recompute_cf_residuals(final_pivot)
        final_pivot = _move_noisy_business_segment_rows_to_disclosures(final_pivot)
        final_pivot = _repair_always_positive(final_pivot)
        final_pivot = _refresh_balance_sheet_closure(final_pivot)
        final_pivot = _apply_quality_result_fixes(final_pivot)

        # Annual mode must not turn Q4/instant/share-count substitutes into FY EPS
        # or FY weighted-average share facts.  Restore reported annual-only EPS and
        # weighted-average shares after shared KPI/accounting helpers have run.
        for _lbl in _NATIVE_ANNUAL_REPORTED_ONLY_LABELS:
            _idx = ('1_Income_Statement', _lbl)
            if _idx in final_pivot.index:
                if _idx in pivoted_temp.index:
                    _reported = pivoted_temp.loc[_idx].reindex(final_pivot.columns)
                    final_pivot.loc[_idx, :] = _reported.values
                else:
                    final_pivot.loc[_idx, :] = np.nan

        # Do the same for segment rows.  The shared accounting engine can fill one
        # missing segment as a residual of total revenue; that is useful for
        # quarterly repair but wrong for annual mode when a segment did not exist in
        # a prior year.  Keep only actually reported annual/stub segment facts.
        for _idx in list(final_pivot.index):
            if _idx[0] in SEG_CATS:
                if _idx in pivoted_temp.index:
                    _reported = pivoted_temp.loc[_idx].reindex(final_pivot.columns)
                    final_pivot.loc[_idx, :] = _reported.values
                else:
                    final_pivot = final_pivot.drop(index=_idx)

    with timer.stage('annual_sort_trim'):
        # Annual-safe display cleanup.  Do not drop sparse segment rows; a segment
        # that starts in the newest year is valid annual output.
        if not final_pivot.empty:
            _html_business_labels = {
                re.sub(r'\s+', ' ', str(f.get('Label') or '')).strip().casefold()
                for f in all_facts
                if f.get('Category') == '4a_Segments_Business'
                and f.get('Concept') == 'HTMLBusinessBreakdown'
            }
            new_idx = [
                ('6_Disclosures', l)
                if c in SEG_CATS and _native_annual_segment_is_disclosure(
                    l, protected_labels=_html_business_labels)
                else (c, l)
                for c, l in final_pivot.index
            ]
            final_pivot.index = pd.MultiIndex.from_tuples(new_idx, names=['Category', 'Label'])
            final_pivot = final_pivot.groupby(level=['Category', 'Label']).first()
            final_pivot = final_pivot.dropna(how='all')
            final_pivot = _native_annual_sort_output(final_pivot, is_financial=is_financial, is_insurance=is_insurance)

        fy_cols = sorted(
            [c for c in final_pivot.columns if isinstance(c, str) and c.endswith('-FY')],
            key=lambda c: int(str(c).split('-')[0]),
            reverse=True,
        )
        if visible_fys:
            visible_cols = {f'{fy}-FY' for fy in visible_fys}
            fy_cols = [c for c in fy_cols if c in visible_cols]
        elif limit is not None:
            fy_cols = fy_cols[:max(1, int(limit))]
        final_pivot = final_pivot[fy_cols]

        period_dates = {}
        if fy_cols:
            end_by_period = selected.groupby('Period')['_End_dt'].max()
            for col in fy_cols:
                dt = end_by_period.get(col)
                period_dates[col] = '' if pd.isna(dt) else pd.to_datetime(dt).strftime('%m/%d/%y')
        final_pivot.attrs['period_dates'] = period_dates

    timer.emit()
    return final_pivot


def _run_native_annual_mode(ticker, company, ye_month, limit, use_arelle=False,
                            log_output=False, save_xlsx=False, workers=None,
                            progress=None, company_name=None,
                            is_financial=False, is_insurance=False,
                            is_oil_gas=False, is_reit=False):
    progress = progress or PipelineProgress(enabled=False)
    _status_print = progress.write if not log_output else print

    progress.set(6.0, "Listing native annual filings")
    with _ExternalConsoleSilencer(enabled=not log_output):
        filings = list(fetch_native_annual_filings(company, limit, ye_month=ye_month))
    progress.set(8.0, f"Queued {len(filings)} annual SEC filings")
    progress.set_stats(0, len(filings), f"Queued {len(filings)} annual SEC filings")

    all_facts, period_dates, failed_filings = [], {}, []

    def process_filing(filing):
        filing_id = safe_filing_id(filing)
        print(f"  Processing {filing_id}...")
        try:
            with _ExternalConsoleSilencer(enabled=not log_output):
                facts, period_end = extract_from_filing(filing, ye_month, ticker, use_arelle)
            return filing, facts, period_end, None
        except Exception as e:
            return filing, [], None, e

    def absorb_result(filing, facts, period_end):
        if facts:
            all_facts.extend(facts)
        if period_end:
            try:
                fy = _native_annual_fy_from_end(pd.to_datetime(period_end), ye_month)
                if fy is not None:
                    period_dates[f"{fy}-FY"] = period_end
            except Exception:
                pass

    requested_workers = workers if workers is not None else _fx_env_int("SEC_MAX_WORKERS", 8, 1)
    max_workers = min(max(1, int(requested_workers)), max(1, len(filings)))
    worker_label = f"{max_workers} native annual"
    if int(requested_workers) != max_workers:
        worker_label += f" ({requested_workers} requested)"

    extraction_cache_key = _native_extraction_cache_key(
        ticker, company, "US_NATIVE_ANNUAL", limit, ye_month, use_arelle, filings
    )

    card_width = progress.sync_terminal_width()
    _status_print(_format_run_card(
        ticker, limit, "US_NATIVE_ANNUAL", worker_label, _cache_display_name(),
        stream=getattr(progress, "_stream", sys.stdout), width=card_width,
        attach_progress=not log_output, warn=progress.warn_text(), stats=progress.stats_text(),
        company_name=company_name or getattr(company, "name", None), cik=getattr(company, "cik", None),
        mode=_mode_display_name(route="US_NATIVE_ANNUAL", use_arelle=use_arelle, save_xlsx=save_xlsx),
    ))
    progress.attach_outline()

    cached_extraction = _native_extraction_cache_get(extraction_cache_key)
    if cached_extraction is not _NATIVE_EXTRACTION_CACHE_MISS:
        _restore_cached_native_extraction(cached_extraction, all_facts, period_dates)
        progress.set_stats(len(filings), len(filings), "Loaded native annual extraction cache")
        progress.set(70.0, "Loaded cached annual SEC extraction")
        print(f"  [Cache] Loaded native annual extraction cache for {ticker.upper()} "
              f"({len(filings)} filing(s), {len(all_facts)} fact rows).")
    else:
        if log_output:
            print(f"\n[Annual Phase 1] Processing {len(filings)} 10-K/10-K/A filings with {max_workers} workers...")

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_filing = {executor.submit(process_filing, f): f for f in filings}
            total_futures = max(1, len(future_to_filing))
            for completed, future in enumerate(concurrent.futures.as_completed(future_to_filing), start=1):
                filing = future_to_filing[future]
                filing_id = safe_filing_id(filing)
                try:
                    _filing, facts, period_end, exc = future.result()
                    if exc:
                        print(f"  [Warning] {filing_id} failed in annual phase 1 "
                              f"({type(exc).__name__}: {exc}) -- queued for retry.")
                        failed_filings.append(filing)
                    else:
                        absorb_result(filing, facts, period_end)
                except Exception as exc:
                    print(f"  [Warning] Future error for {filing_id} "
                          f"({type(exc).__name__}: {exc}) -- queued for retry.")
                    failed_filings.append(filing)
                finally:
                    pct = 8.0 + 56.0 * completed / total_futures
                    progress.set(pct, f"Annual SEC filings {completed}/{len(future_to_filing)}")

        if failed_filings:
            if log_output:
                print(f"\n[Annual Phase 2] Retrying {len(failed_filings)} failed annual filing(s) sequentially...")
            total_retries = max(1, len(failed_filings))
            still_failed = []
            for retry_no, filing in enumerate(failed_filings, start=1):
                filing_id = safe_filing_id(filing)
                success = False
                for attempt in range(1, 5):
                    wait = 10 * attempt
                    print(f"  [Retry {attempt}/4] {filing_id} -- waiting {wait}s...")
                    time.sleep(wait)
                    try:
                        with _ExternalConsoleSilencer(enabled=not log_output):
                            facts, period_end = extract_from_filing(filing, ye_month, ticker, use_arelle)
                        absorb_result(filing, facts, period_end)
                        print(f"  [OK] {filing_id} recovered on retry {attempt}.")
                        success = True
                        break
                    except Exception as e:
                        print(f"  [Retry {attempt}/4] Still failing: {type(e).__name__}: {e}")
                if not success:
                    print(f"  [Error] {filing_id} permanently failed after all annual retries.")
                    still_failed.append(filing)
                progress.set(64.0 + 6.0 * retry_no / total_retries,
                             f"Annual retries {retry_no}/{len(failed_filings)}")


        if all_facts and not failed_filings:
            _native_extraction_cache_set(
                extraction_cache_key,
                all_facts,
                period_dates,
                _snapshot_native_mutable_state(),
                metadata={
                    "ticker": str(ticker).upper(),
                    "mode": "US_NATIVE_ANNUAL",
                    "filings": len(filings),
                    "facts": len(all_facts),
                },
            )

    if all_facts:
        all_facts[:] = _reconcile_html_business_breakdown_facts(all_facts)

    progress.set(70.0, "Annual SEC retrieval complete")

    if not all_facts:
        return None

    final_cache_key = _final_pivot_cache_key(
        ticker, company, "US_NATIVE_ANNUAL_FINAL", limit, ye_month, use_arelle, filings,
        company_name=company_name, is_financial=is_financial, is_insurance=is_insurance,
        is_oil_gas=is_oil_gas, is_reit=is_reit,
    )
    cached_final = _final_pivot_cache_get(final_cache_key)
    if cached_final is not _FINAL_PIVOT_CACHE_MISS:
        final_pivot = cached_final
        progress.set(97.0, "Loaded annual final pivot cache")
        print(f"  [Cache] Loaded native annual final pivot cache for {ticker.upper()} "
              f"({len(final_pivot)} rows).")
    else:
        progress.start_pulse(88.0, "Building annual FY statements", expected_seconds=45.0)
        try:
            final_pivot = build_annual_pivoted_data(
                all_facts, ticker, ye_month, company_name=company_name,
                is_financial=is_financial, is_insurance=is_insurance,
                is_oil_gas=is_oil_gas, is_reit=is_reit, limit=limit,
            )
        finally:
            progress.stop_pulse()

        if final_pivot is None or final_pivot.empty:
            return None

        progress.set(90.0, "Annual FY statements built")
        annual_period_dates = dict(final_pivot.attrs.get('period_dates') or {})
        annual_period_dates.update({k: v for k, v in period_dates.items() if k in final_pivot.columns})
        if annual_period_dates:
            header_row = pd.DataFrame(
                {col: annual_period_dates.get(col, '') for col in final_pivot.columns},
                index=pd.MultiIndex.from_tuples([('0_Period_Header', 'Period Ending')],
                                                names=['Category', 'Label'])
            )
            final_pivot = pd.concat([header_row, final_pivot])
        _final_pivot_cache_set(
            final_cache_key, final_pivot,
            metadata={
                "ticker": str(ticker).upper(),
                "mode": "US_NATIVE_ANNUAL_FINAL",
                "filings": len(filings),
                "rows": len(final_pivot),
                "columns": len(final_pivot.columns),
            },
        )

    final_pivot = _apply_quality_result_fixes(final_pivot)
    final_pivot = _normalize_output_margin_rows(final_pivot)
    final_pivot = _sort_final_output_pivot(final_pivot, is_financial=is_financial, is_insurance=is_insurance, context="native annual pre-write sort")

    progress.set(98.0, "Writing annual output file")
    out_dir = "output/financials"
    os.makedirs(out_dir, exist_ok=True)
    if save_xlsx:
        xlsx_dir = f"{out_dir}/excel"
        os.makedirs(xlsx_dir, exist_ok=True)
        out_path = f"{xlsx_dir}/{ticker}_annual_financials.xlsx"
        try:
            _save_pivot_xlsx(final_pivot, out_path)
        except ImportError:
            out_path = f"{out_dir}/{ticker}_annual_financials.csv"
            _profile_call("write_csv", final_pivot.to_csv, out_path)
            print("  [xlsx] openpyxl not installed -- run 'pip install openpyxl'; saved annual CSV instead.")
        except Exception as xe:
            out_path = f"{out_dir}/{ticker}_annual_financials.csv"
            _profile_call("write_csv", final_pivot.to_csv, out_path)
            print(f"  [xlsx] annual export failed ({xe}); saved CSV instead.")
    else:
        out_path = f"{out_dir}/{ticker}_annual_financials.csv"
        _profile_call("write_csv", final_pivot.to_csv, out_path)

    print(f"Success! Annual data saved to {out_path}")
    return out_path

def main(ticker, limit, use_arelle=False, dqc_ruleset=None, log_output=False,
         save_xlsx=False, workers=None, annual=False):
    _initialize_sec_identity()
    import builtins
    _original_print = builtins.print

    progress = PipelineProgress(enabled=not log_output, description=f"{ticker}: starting")
    if not log_output:
        builtins.print = progress.note_log
    _status_print = progress.write if not log_output else _original_print
    progress.set(1.0, "Fetching company profile")
    with _ExternalConsoleSilencer(enabled=not log_output):
        company = fetch_company(ticker)
    progress.set(4.0, "Detecting filing route")
    # Extract the company's legal name for segment-label reconciliation.
    # Used by Layer 1c to detect when a company adds/removes its own name
    # from XBRL segment tags across filings (e.g. "Boeing Defense Space
    # Security" -> "Defense Space And Security").
    is_insurance = False
    is_financial = False
    is_oil_gas = False
    is_reit = False
    try:
        sic = str(company.sic)
        if sic.startswith('63') or sic.startswith('64'):
            is_insurance = True
            print(f"  [Industry] Detected Insurance Institution (SIC: {sic})")
        elif sic == '6798' or sic.startswith('65'):
            is_reit = True
            print(f"  [Industry] Detected REIT / Real Estate (SIC: {sic})")
        elif sic.startswith('6'):
            is_financial = True
            print(f"  [Industry] Detected Financial Institution (SIC: {sic})")
        elif sic.startswith('131') or sic.startswith('29'):
            is_oil_gas = True
            print(f"  [Industry] Detected Oil & Gas / Extractive (SIC: {sic})")
    except Exception:
        pass
        
    _company_name = company.name
    with _ExternalConsoleSilencer(enabled=not log_output):
        filings_10k = get_company_filings(company, "10-K")
        profile = _fx_profile(company, filings_10k)
        ye_month = (profile.fiscal_year_end_month if profile.fiscal_year_end_month
                    else (pd.to_datetime(filings_10k[0].period_of_report).month
                          if len(filings_10k) > 0 else 12))

    # --- Foreign private issuer (20-F/40-F): annual-only isolated branch ---
    if profile.pipeline in ("FOREIGN_20F", "FOREIGN_40F"):
        _fx_out_dir = "output/financials"
        os.makedirs(_fx_out_dir, exist_ok=True)
        progress.set_stats(0, limit, "Starting annual 20-F scan")
        card_width = progress.sync_terminal_width()
        _status_print(_format_run_card(
            ticker,
            limit,
            profile.pipeline,
            "annual only (foreign FY)",
            _cache_display_name(),
            stream=getattr(progress, "_stream", sys.stdout),
            width=card_width,
            attach_progress=not log_output,
            warn=progress.warn_text(),
            stats=progress.stats_text(),
            company_name=getattr(company, "name", None),
            cik=getattr(company, "cik", None),
            mode=_mode_display_name(route=profile.pipeline, use_arelle=use_arelle, save_xlsx=save_xlsx),
        ))
        progress.attach_outline()
        with _ExternalConsoleSilencer(enabled=not log_output):
            result = _fx_main_foreign_20f(
                ticker, company, ye_month, _fx_out_dir, limit=limit,
                save_xlsx=save_xlsx, progress=progress, use_arelle=use_arelle,
            )
        progress.finish("Complete" if result is not None else "Complete (no data)")
        if not log_output:
            builtins.print = _original_print
        return

    # --- Native U.S. 10-K annual mode: isolated 10-K / 10-K/A FY branch ---
    if annual:
        annual_ye_month = _native_annual_effective_ye_month_from_filings(filings_10k, ye_month)
        if annual_ye_month != ye_month:
            print(f"  [Annual] Normalized fiscal year-end month for 52/53-week annual mode: {ye_month} -> {annual_ye_month}")
        _annual_state = _snapshot_native_mutable_state()
        try:
            out_path = _run_native_annual_mode(
                ticker, company, annual_ye_month, limit, use_arelle=use_arelle,
                log_output=log_output, save_xlsx=save_xlsx, workers=workers,
                progress=progress, company_name=_company_name,
                is_financial=is_financial, is_insurance=is_insurance,
                is_oil_gas=is_oil_gas, is_reit=is_reit,
            )
        finally:
            _restore_native_mutable_state(_annual_state)
        if not log_output:
            progress.finish("Complete" if out_path else "Complete (no data)",
                            f"Saved to {out_path}" if out_path else "No data extracted.")
            builtins.print = _original_print
        else:
            progress.finish("Complete" if out_path else "Complete (no data)")
            _original_print(f"Success! Data saved to {out_path}" if out_path else "No data extracted.")
        return

    progress.set(6.0, "Listing SEC filings")
    with _ExternalConsoleSilencer(enabled=not log_output):
        filings = list(fetch_filings(company, limit))
    progress.set(8.0, f"Queued {len(filings)} SEC filings")
    progress.set_stats(0, len(filings), f"Queued {len(filings)} SEC filings")

    all_facts, period_dates = [], {}
    failed_filings = []

    def process_filing(filing):
        # safe_filing_id never makes a network call -- safe inside the thread pool
        filing_id = safe_filing_id(filing)
        print(f"  Processing {filing_id}...")
        try:
            with _ExternalConsoleSilencer(enabled=not log_output):
                facts, period_end = extract_from_filing(filing, ye_month, ticker, use_arelle)
            return filing, facts, period_end, None
        except Exception as e:
            return filing, [], None, e

    def _absorb_result(filing, facts, period_end):
        """Record facts and period date from a successful fetch."""
        if facts:
            all_facts.extend(facts)
        if period_end:
            fy, q = get_period_info(pd.to_datetime(period_end), ye_month)
            period_dates[f"{fy}-{q}"] = period_end

    # -- Phase 1: Parallel fetch ---------------------------------------------
    requested_workers = (
        workers if workers is not None
        else _fx_env_int("SEC_MAX_WORKERS", 8, 1)
    )
    MAX_WORKERS = min(max(1, int(requested_workers)), max(1, len(filings)))
    # Worker count controls local overlap only; sec_limiter still caps SEC
    # request starts at the process level.
    worker_label = f"{MAX_WORKERS} native"
    if int(requested_workers) != MAX_WORKERS:
        worker_label += f" ({requested_workers} requested)"

    extraction_cache_key = _native_extraction_cache_key(
        ticker, company, profile.pipeline, limit, ye_month, use_arelle, filings
    )

    card_width = progress.sync_terminal_width()
    _status_print(_format_run_card(
        ticker,
        limit,
        profile.pipeline,
        worker_label,
        _cache_display_name(),
        stream=getattr(progress, "_stream", sys.stdout),
        width=card_width,
        attach_progress=not log_output,
        warn=progress.warn_text(),
        stats=progress.stats_text(),
        company_name=getattr(company, "name", None),
        cik=getattr(company, "cik", None),
        mode=_mode_display_name(route=profile.pipeline, use_arelle=use_arelle, save_xlsx=save_xlsx),
    ))
    progress.attach_outline()
    cached_extraction = _native_extraction_cache_get(extraction_cache_key)
    if cached_extraction is not _NATIVE_EXTRACTION_CACHE_MISS:
        _restore_cached_native_extraction(cached_extraction, all_facts, period_dates)
        progress.set_stats(len(filings), len(filings), "Loaded native extraction cache")
        progress.set(70.0, "Loaded cached SEC extraction")
        print(f"  [Cache] Loaded native extraction cache for {ticker.upper()} "
              f"({len(filings)} filing(s), {len(all_facts)} fact rows).")
    else:
        if log_output:
            print(f"\n[Phase 1] Processing {len(filings)} filings with {MAX_WORKERS} workers...")

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_filing = {executor.submit(process_filing, f): f for f in filings}

            iterator = concurrent.futures.as_completed(future_to_filing)

            total_futures = max(1, len(future_to_filing))
            for completed_futures, future in enumerate(iterator, start=1):
                filing = future_to_filing[future]
                # safe_filing_id: no network call -- safe to call even on a timed-out filing
                filing_id = safe_filing_id(filing)
                try:
                    _, facts, period_end, exc = future.result()
                    if exc:
                        print(f"  [Warning] {filing_id} failed in phase 1 "
                              f"({type(exc).__name__}: {exc}) -- queued for retry.")
                        failed_filings.append(filing)
                    else:
                        _absorb_result(filing, facts, period_end)
                except Exception as exc:
                    # Future itself raised (thread crash, pickling failure, etc.)
                    print(f"  [Warning] Future error for {filing_id} "
                          f"({type(exc).__name__}: {exc}) -- queued for retry.")
                    failed_filings.append(filing)
                finally:
                    pct = 8.0 + 50.0 * completed_futures / total_futures
                    progress.set(pct, f"SEC filings {completed_futures}/{len(future_to_filing)}")

        # -- Phase 2: Sequential retry for failed filings ------------------------
        if failed_filings:
            if log_output:
                print(f"\n[Phase 2] Retrying {len(failed_filings)} failed filing(s) sequentially "
                      f"(up to 4 attempts each, with exponential back-off)...")
            retry_iterator = failed_filings
            progress.set(59.0, f"Retrying {len(failed_filings)} failed filings")

            still_failed = []
            total_retries = max(1, len(retry_iterator))

            for retry_no, filing in enumerate(retry_iterator, start=1):
                filing_id = safe_filing_id(filing)
                success = False

                for attempt in range(1, 5):          # 4 sequential retries per filing
                    wait = 10 * attempt              # 10 s, 20 s, 30 s, 40 s
                    print(f"  [Retry {attempt}/4] {filing_id} -- waiting {wait}s...")
                    time.sleep(wait)
                    try:
                        with _ExternalConsoleSilencer(enabled=not log_output):
                            facts, period_end = extract_from_filing(filing, ye_month, ticker, use_arelle)
                        _absorb_result(filing, facts, period_end)
                        print(f"  [OK] {filing_id} recovered on retry {attempt}.")
                        success = True
                        break
                    except Exception as e:
                        print(f"  [Retry {attempt}/4] Still failing: "
                              f"{type(e).__name__}: {e}")

                if not success:
                    print(f"  [Error] {filing_id} permanently failed after all retries.")
                    still_failed.append(filing)
                progress.set(59.0 + 7.0 * retry_no / total_retries,
                             f"Retries {retry_no}/{len(retry_iterator)}")

            # -- Phase 3: Minimal period-date recovery ---------------------------
            # Even when we cannot retrieve facts, we try to discover the quarter
            # so the output CSV includes an (empty) column for it.  An empty cell
            # is far more informative than a silently missing period.
            if still_failed:
                if log_output:
                    print(f"\n[Phase 3] Recovering period metadata for "
                          f"{len(still_failed)} permanently failed filing(s)...")
                phase3_iterator = still_failed
                progress.set(67.0, f"Recovering metadata for {len(still_failed)} filings")
                total_phase3 = max(1, len(phase3_iterator))

                for phase3_no, filing in enumerate(phase3_iterator, start=1):
                    filing_id = safe_filing_id(filing)
                    progress.set(67.0 + 3.0 * (phase3_no - 1) / total_phase3,
                                 f"Metadata {phase3_no}/{len(phase3_iterator)}")

                    # Attempt 1: fetch just the SGML header (small file) with retries
                    with _ExternalConsoleSilencer(enabled=not log_output):
                        period_str = fetch_period_of_report(filing)

                    if period_str:
                        try:
                            period_dt = pd.to_datetime(period_str)
                            fy, q = get_period_info(period_dt, ye_month)
                            key = f"{fy}-{q}"
                            if key not in period_dates:
                                period_dates[key] = period_dt.strftime('%m/%d/%y')
                                print(f"  [Period Recovered] {filing_id} -> {key} "
                                      f"(column will be present, data unavailable).")
                            else:
                                print(f"  [Period Already Known] {filing_id} -> {key}.")
                            continue
                        except Exception:
                            pass

                    # Attempt 2: estimate from filing date (always available locally)
                    est_period, is_est = estimate_period_from_filing_date(filing)
                    if est_period:
                        try:
                            period_dt = pd.to_datetime(est_period)
                            fy, q = get_period_info(period_dt, ye_month)
                            key = f"{fy}-{q}"
                            if key not in period_dates:
                                period_dates[key] = period_dt.strftime('%m/%d/%y')
                                print(f"  [Period Estimated] {filing_id} -> {key} "
                                      f"(estimated from filing date; data unavailable).")
                            else:
                                print(f"  [Period Already Known] {filing_id} -> {key}.")
                        except Exception:
                            print(f"  [Period Unknown] Could not recover any period "
                                  f"metadata for {filing_id}.")
                    else:
                        print(f"  [Period Unknown] Could not recover any period "
                              f"metadata for {filing_id}.")


        if all_facts and not failed_filings:
            _native_extraction_cache_set(
                extraction_cache_key,
                all_facts,
                period_dates,
                _snapshot_native_mutable_state(),
                metadata={
                    "ticker": str(ticker).upper(),
                    "mode": str(profile.pipeline or "US_NATIVE"),
                    "filings": len(filings),
                    "facts": len(all_facts),
                },
            )

    if all_facts:
        all_facts[:] = _reconcile_html_business_breakdown_facts(all_facts)

    progress.set(70.0, "SEC retrieval complete")

    # -- Final processing & output --------------------------------------------
    if all_facts:
        final_cache_key = _final_pivot_cache_key(
            ticker, company, f"{profile.pipeline}_FINAL", limit, ye_month, use_arelle, filings,
            company_name=_company_name, is_financial=is_financial, is_insurance=is_insurance,
            is_oil_gas=is_oil_gas, is_reit=is_reit,
        )
        cached_final = _final_pivot_cache_get(final_cache_key)
        if cached_final is not _FINAL_PIVOT_CACHE_MISS:
            final_pivot = cached_final
            progress.set(97.0, "Loaded final pivot cache")
            print(f"  [Cache] Loaded native final pivot cache for {ticker.upper()} "
                  f"({len(final_pivot)} rows).")
        else:
            progress.start_pulse(82.0, "Building quarterly statements", expected_seconds=75.0)
            try:
                final_pivot = build_pivoted_data(
                    all_facts, ticker, ye_month,
                    company_name=_company_name,
                    is_financial=is_financial,
                    is_insurance=is_insurance,
                    is_oil_gas=is_oil_gas,
                    is_reit=is_reit,
                )
            finally:
                progress.stop_pulse()
            progress.set(82.0, "Quarterly statements built")
            progress.start_pulse(88.0, "Organizing output rows", expected_seconds=20.0)
        
            # Filter out 4a Business Segments that have no data in the most recent 8 quarters
            if len(final_pivot.columns) > 0:
                is_4a = final_pivot.index.get_level_values('Category') == '4a_Segments_Business'
                check_limit = min(8, len(final_pivot.columns))
                stale_4a_mask = is_4a & final_pivot.iloc[:, :check_limit].isna().all(axis=1)
                final_pivot = final_pivot[~stale_4a_mask]

                is_seg = final_pivot.index.get_level_values('Category').isin({'4a_Segments_Business', '4b_Segments_Geographic_Regions', '4c_Segments_Geographic_Countries', '4d_Segments_Cross_Tabulated'})
                sparse_seg_mask = is_seg & (final_pivot.notna().sum(axis=1) <= 2)
                final_pivot = final_pivot[~sparse_seg_mask]

            final_pivot = final_pivot.dropna(how='all')
        
            _html_business_labels = {
                re.sub(r'\s+', ' ', str(f.get('Label') or '')).strip().casefold()
                for f in all_facts
                if f.get('Category') == '4a_Segments_Business'
                and f.get('Concept') == 'HTMLBusinessBreakdown'
            }

            def is_disc(cat, lbl):
                normalized_label = re.sub(r'\s+', ' ', str(lbl or '')).strip().casefold()
                if cat == '4a_Segments_Business' and normalized_label in _html_business_labels:
                    return False
                # A segment label is typically "Metric - Member" or "Metric - SubMember - Member"
                parts = [p.strip() for p in lbl.split(' - ')]
                if not parts:
                    return False
                
                prefix = parts[0]
            
                # 1. Metric-based check: if it doesn't start with a "genuine" business segment metric, move to disclosures.
                # This handles things like "Useful Life", "RPO Timing", etc.
                if prefix not in GENUINE_SEGMENT_METRICS:
                    return True
                
                # 2. Member-based check: even if the metric is genuine, the member name might be noise.
                if len(parts) >= 2:
                    member_full = ' - '.join(parts[1:]).lower()
                
                    # Check for "Total" or "Consolidated" summaries that should be disclosures
                    if any(kw == member_full for kw in ('operating segments', 'reportable segments', 'consolidated', 'total', 'all segments')):
                        return True
                
                    # Check against DISCLOSURE_PATTERNS (noise tags like "state and local jurisdiction")
                    if any(pat in member_full for pat in DISCLOSURE_PATTERNS):
                        return True
                
                    # Ratio guard: if member name contains "percent", "percentage", or "ratio" 
                    # but the metric prefix itself isn't a known ratio metric, it's a disclosure.
                    # This catches items like "Revenue - Percentage of Net Sales" or 
                    # "Operating Income - Defined Contribution Plan Percent".
                    RATIO_KWS = {'percent', 'percentage', 'ratio', 'weighted average'}
                    if any(kw in member_full for kw in RATIO_KWS):
                        # Exception: if prefix is "Customer Concentration %", keep it in segments (though typically cat 7)
                        if 'Concentration %' not in prefix:
                            return True
                
                    # Catch internal/technical tags
                    if '_' in member_full:
                        return True
                    
                return False

            new_idx = [
                ('6_Disclosures', l) if c in SEG_CATS and is_disc(c, l) else (c, l)
                for c, l in final_pivot.index
            ]
            final_pivot.index = pd.MultiIndex.from_tuples(new_idx, names=['Category', 'Label'])
        
            final_pivot = final_pivot.groupby(level=['Category', 'Label']).first()
            progress.set(85.0, "Preparing display order")
        
            # --- NEW: Dynamic Sort Order ---
            base_order = {n: i for i, n in enumerate(CONCEPT_MAP.keys())}

            if is_insurance:
                ins_order = [
                    # --- Top Line / Revenues ---
                    'Premiums Earned', 
                    'Net Investment Income', 
                    'Revenue',
                
                    # --- Claims & Cost of Revenue ---
                    'Policyholder Claims/Benefits', 
                    'Amortization of DAC', 
                    'Cost of Revenue', 
                    'Gross Profit', 
                
                    # --- Operating Expenses ---
                    'Research & Development',
                    'Selling, General & Admin',
                    'Sales & Marketing',
                    'General & Administrative',
                    'Marketing Expense',
                    'Customer Bad Debt',
                    'Amortization of Intangibles',
                    'Restructuring & Related Charges',
                    'Impairment Charges',
                    'Acquisition-Related Costs',
                    'Litigation & Settlement Charges',
                
                    # --- Subtotals & Non-Operating ---
                    'Total Operating Expenses', 
                    'Operating Income', 
                    'Interest Income', 
                    'Interest Expense',
                    'Net Interest Income (Expense)',
                    'Gain/Loss on Investments',
                    'Equity Method Income',
                    'Other Income / (Expense)',
                    'Total Non-operating Income',
                
                    # --- Bottom Line ---
                    'Pretax Income', 
                    'Income Tax Expense', 
                    'Income from Discontinued Operations',
                    'Net Income',
                    'Net Income to Noncontrolling Interest',
                
                    # --- Per Share Data ---
                    'EPS Basic',
                    'EPS Diluted',
                    'Shares Outstanding Basic',
                    'Shares Outstanding Diluted'
                ]
                _ins_pos = {name: pos for pos, name in enumerate(ins_order)}
                item_order = {
                    n: _ins_pos.get(n, base_order[n] + 100)
                    for n in CONCEPT_MAP.keys()
                }
            elif is_financial:
                fin_order = [
                    'Interest Income', 
                    'Interest Expense', 
                    'Net Interest Income (Expense)',
                    'Revenue', 
                
                    # --- Operating Expenses ---
                    'Research & Development',
                    'Sales & Marketing', 
                    'Marketing Expense',
                    'Selling, General & Admin', 
                    'General & Administrative',
                    'Salaries & Employee Benefits',
                    'Customer Bad Debt',
                    'Other Operating Expenses',
                    'Amortization of Intangibles',
                    'Restructuring & Related Charges',
                    'Impairment Charges',
                    'Acquisition-Related Costs',
                    'Litigation & Settlement Charges',
                
                    # --- Subtotals & Below the Line ---
                    'Total Operating Expenses', 
                    'Operating Income',
                    'Pretax Income', 
                    'Income Tax Expense', 
                    'Net Income', 
                    'Net Income to Noncontrolling Interest'
                ]
                _fin_pos = {name: pos for pos, name in enumerate(fin_order)}
                item_order = {
                    n: _fin_pos.get(n, base_order[n] + 100)
                    for n in CONCEPT_MAP.keys()
                }
                # Keep the financial-company Gross Profit anchor near the top of
                # the statement even when the filer does not report a Gross
                # Profit row. Dynamic operating-expense lines use this anchor as
                # a floor; leaving it in the generic fallback bucket pushed
                # brokerage/clearance fees below Net Income for HOOD.
                if 'Gross Profit' in item_order and 'Revenue' in _fin_pos:
                    item_order['Gross Profit'] = _fin_pos['Revenue'] + 0.5

                # Any omitted financial-statement operating-expense line
                # discovered from the face statement should remain above Total
                # Operating Expenses, not fall into the post-Net-Income catch-all.
                _fin_opex_start = _fin_pos.get('Research & Development',
                                               _fin_pos.get('Sales & Marketing', 4))
                _fin_opex_end = _fin_pos.get('Total Operating Expenses', 99)
                for _name, _info in list(CONCEPT_MAP.items()):
                    if not (isinstance(_info, dict)
                            and _info.get('cat') == '1_Income_Statement'):
                        continue
                    if _name in _fin_pos:
                        continue
                    _kw_pos = _evidence_is_pos(_name, is_financial, is_insurance)
                    if 9 <= _kw_pos < 20:
                        # Preserve keyword-bucket relative order while fitting
                        # inside the compact financial operating-expense block.
                        _slot = _fin_opex_start + ((_kw_pos - 9) / 11.0) * max(1, (_fin_opex_end - _fin_opex_start - 1))
                        item_order[_name] = min(item_order.get(_name, 999), _slot)
            else:
                item_order = base_order

            # Separate generic income-statement sorter for every company type.
            # Non-IS sections keep their original CONCEPT_MAP order; only P&L
            # rows use accounting-bucket fallback positions.
            is_item_order = _build_item_order(is_insurance, is_financial)

            # Goodwill is defined in CONCEPT_MAP between Total Current Liabilities and
            # Long-term Debt, so the raw key order sorts it into the liabilities block.
            # It is a non-current asset: re-anchor it beside the intangibles (or just
            # above Total Assets) so it sorts within the asset section, for every filer.
            if 'Goodwill' in item_order:
                if 'Intangible Assets (Net)' in item_order:
                    item_order['Goodwill'] = item_order['Intangible Assets (Net)'] + 0.5
                elif 'Intangible Assets & Goodwill' in item_order:
                    item_order['Goodwill'] = item_order['Intangible Assets & Goodwill'] + 0.5
                elif 'Total Assets' in item_order:
                    item_order['Goodwill'] = item_order['Total Assets'] - 0.5

            # Slot auto-learned face-statement lines into the curated flow at the
            # position the company itself presents them (presentation order).
            item_order = _apply_presentation_anchors(item_order, is_financial, is_insurance)
            is_item_order = _apply_presentation_anchors(is_item_order, is_financial, is_insurance)

            # IBM-style filers fold interest (and other non-operating income) into a
            # single expense block with no separate operating-income subtotal. Mirror
            # the filing's reading order by moving Interest Expense -- and the block's
            # 'Other Income / (Expense)' -- to just ABOVE the Total Operating Expenses
            # line instead of below it. Gated on the detected structure.
            if (
                _IBM_STYLE_STATE.get('active')
                and not is_financial
                and not is_insurance
                and 'Total Operating Expenses' in is_item_order
            ):
                _toe_pos = is_item_order['Total Operating Expenses']
                for _blk_lbl, _off in (('Other Income / (Expense)', 0.03),
                                       ('Gain/Loss on Investments', 0.02),
                                       ('Interest Expense', 0.01)):
                    if _blk_lbl in is_item_order:
                        is_item_order[_blk_lbl] = _toe_pos - _off

            seg_met_order = {m: i for i, m in enumerate(SEGMENT_METRIC_ORDER)}
            business_segment_labels = [idx[1] for idx in final_pivot.index if idx[0] == '4a_Segments_Business']
            business_segment_context = _business_segment_context(business_segment_labels)
        
            # --- NEW: Dynamic KPI Sort Order ---
            if is_insurance:
                dynamic_kpis = ['Loss Ratio (%)', 'Expense Ratio (%)', 'Combined Ratio (%)', 'Net Margin (%)'] + KPI_ORDER
            elif is_financial:
                dynamic_kpis = ['Efficiency Ratio (%)', 'Net Margin (%)'] + KPI_ORDER
            else:
                dynamic_kpis = ['Gross Margin (%)', 'Operating Margin (%)', 'Net Margin (%)'] + KPI_ORDER
            
            kpi_order = {k: i for i, k in enumerate(dynamic_kpis)}
            for alias, canonical in KPI_ORDER_ALIASES.items():
                if canonical in kpi_order:
                    kpi_order[alias] = kpi_order[canonical]
            for k in final_pivot.index.get_level_values('Label').unique():
                if k not in kpi_order:
                    kpi_order[k] = 99

            _adjustment_sort_map = {
                'Gross Profit: Other Adjustments': 'Gross Profit',
                'Operating Income: Other Adjustments': 'Operating Income',
                'Pretax Income: Other Adjustments': 'Pretax Income',
                'Net Income: Other Adjustments': 'Net Income',
                'Other Operating Adjustments (Net)': 'Operating Cash Flow',
                'Other Investing Adjustments (Net)': 'Investing Cash Flow',
                'Other Financing Adjustments (Net)': 'Financing Cash Flow',
            }

            def sort_key(idx):
                cat, label = idx
                if label in _adjustment_sort_map:
                    _anchor = _adjustment_sort_map[label]
                    _anchor_order = is_item_order if cat == '1_Income_Statement' else item_order
                    return (CAT_ORDER.get(cat, 99),
                            _anchor_order.get(_anchor, item_order.get(_anchor, 999)),
                            -1, label)
                base = label.split(' - ')[0]
                if cat == '1_Income_Statement':
                    if base not in is_item_order:
                        is_item_order[base] = _evidence_is_pos(base, is_financial, is_insurance)
                    return (CAT_ORDER.get(cat, 99), is_item_order.get(base, 999),
                            0 if ' - ' not in label else 1, label)
                if cat in SEG_CATS or cat == '7_Concentration_Risk':
                    return (CAT_ORDER.get(cat, 99), seg_met_order.get(base, 999), label)
                if cat == '5_KPI_Metrics': return (CAT_ORDER.get(cat, 99), kpi_order.get(label, 999), label)
                return (CAT_ORDER.get(cat, 99), item_order.get(base, 999), 0 if ' - ' not in label else 1, label)

            # Keep only the coarse income-statement nature split on the face; move
            # finer product/service members back to the segment block. Then stitch
            # any cross-era renamed segment members into one continuous series.
            # (Both run before the sort so re-categorised rows land in the right
            # section.)
            progress.stop_pulse()
            progress.set(88.0, "Output rows organized")
            progress.start_pulse(97.0, "Final accounting repairs", expected_seconds=55.0)
            final_pivot = _demote_non_face_nature_split(final_pivot)
            final_pivot = _stitch_era_renamed_members(final_pivot)
            final_pivot = _surface_income_statement_dda(final_pivot, is_oil_gas, is_reit)

            final_pivot = _sort_preserving_values(final_pivot, sort_key, is_item_order, is_financial, is_insurance, context="native quarterly final sort")
            final_pivot = apply_stock_splits(final_pivot)
            final_pivot = _institutional_cleanup(final_pivot)
            final_pivot = _reconcile_equity_with_nci(final_pivot)
            final_pivot = _neutralize_ytd_undercapture(final_pivot)
            final_pivot = _recompute_cf_residuals(final_pivot)
        
            # Reconcile abandoned face-statement disaggregation members. A nature-
            # of-revenue (product/service) line carrying no data in the recent
            # window is not part of the filer's current face presentation, but it is
            # often the older half of a series whose recent half now lives in the
            # segment block -- so its history is spliced there before the row is
            # removed, rather than discarded. (Genuine current members -- IBM's
            # Services/Sales/Financing -- keep recent data and are retained.)
            final_pivot = _stitch_or_drop_abandoned_face_disagg(final_pivot)
            final_pivot = _merge_prefix_continuation_members(final_pivot)
            final_pivot = _repair_q4_from_annual_ytd(final_pivot)
            final_pivot = _repair_balance_sheet_identity(final_pivot)
            final_pivot = _refresh_balance_sheet_closure(final_pivot)

            # Final, authoritative segment-partition repair. MUST run last: it
            # corrects any mis-parsed / unfilled segment Q4 cells left behind by
            # the HTML rescue and the second repair phase (e.g. Mastercard
            # geographic) via the pure accounting identity, with no later pass
            # able to overwrite it before the CSV is written.
            final_pivot = _reconcile_segment_partition_from_total(final_pivot)
            final_pivot = _apply_quality_result_fixes(final_pivot)

            # --- Clean up sparse/abandoned rows via calculation linkbase mathematically ---
            label_to_concept = {}
            for f in all_facts:
                label_to_concept[(f['Category'], f['Label'])] = f['Concept']
            _present_output_labels = set(final_pivot.index.get_level_values('Label'))

            for cat, label in final_pivot.index:
                # The calculation-linkbase trace is only meaningful for face
                # statement rows: a segment row's lineage walks straight to
                # Gross Profit / Operating Income (Revenue's calc parents) and
                # would delete genuine segment history under a nonsense parent.
                if cat not in ('1_Income_Statement', '2_Balance_Sheet', '3_Cash_Flow'):
                    continue
                # Protect core curated tags: they are explicitly mapped granular data and should never be dropped.
                if label in CONCEPT_MAP and not CONCEPT_MAP[label].get('auto', False):
                    continue
                
                # Recency Guard: if the row has data in the most recent 4 periods, it is an active, 
                # newly-reported genuine metric (not abandoned garbage). Protect it!
                if final_pivot.loc[(cat, label)].iloc[:4].notna().any():
                    continue

                cpt = label_to_concept.get((cat, label))
                if cpt:
                    cpt_clean = cpt.split(':')[-1] if ':' in cpt else cpt.split('_')[-1]
                
                    current_cpts = {cpt_clean}
                    rolled_up_target = None
                    for _ in range(4):
                        next_cpts = set()
                        for c in current_cpts:
                            for p_clean, w in GLOBAL_CALC_PARENT.get(c, []):
                                next_cpts.add(p_clean)
                                target = STANDARD_TAG_MAP.get(p_clean) or _CONCEPT_TAG_TO_LABEL.get(p_clean)
                                if target and target in CONCEPT_MAP and not CONCEPT_MAP[target].get('auto'):
                                    if target in _present_output_labels:
                                        rolled_up_target = target
                                        break
                            if rolled_up_target: break
                        if rolled_up_target: break
                        current_cpts = next_cpts
                    
                    if rolled_up_target:
                        print(f"  [Sparse Rollup] Accurately rolling up abandoned sparse '{label}' into parent '{rolled_up_target}'")
                        final_pivot = final_pivot.drop((cat, label))

            final_pivot = _apply_quality_result_fixes(final_pivot)
            final_pivot = _sort_preserving_values(final_pivot, sort_key, is_item_order, is_financial, is_insurance, context="native quarterly final sort")

            if period_dates:
                header_row = pd.DataFrame({col: period_dates.get(col, '') for col in final_pivot.columns}, index=pd.MultiIndex.from_tuples([('0_Period_Header', 'Period Ending')], names=['Category', 'Label']))
                final_pivot = pd.concat([header_row, final_pivot])

            progress.stop_pulse()
            progress.set(97.0, "Final repairs complete")
            _final_pivot_cache_set(
                final_cache_key, final_pivot,
                metadata={
                    "ticker": str(ticker).upper(),
                    "mode": f"{profile.pipeline}_FINAL",
                    "filings": len(filings),
                    "rows": len(final_pivot),
                    "columns": len(final_pivot.columns),
                },
            )

        final_pivot = _apply_quality_result_fixes(final_pivot)
        final_pivot = _normalize_output_margin_rows(final_pivot)
        final_pivot = _sort_final_output_pivot(final_pivot, is_financial=is_financial, is_insurance=is_insurance, context="native quarterly pre-write sort")

        progress.set(98.0, "Writing output file")
        out_dir = "output/financials"
        os.makedirs(out_dir, exist_ok=True)
        
        if save_xlsx:
            # Create the excel sub-folder inside output/financials/
            xlsx_dir = f"{out_dir}/excel"
            os.makedirs(xlsx_dir, exist_ok=True)
            out_path = f"{xlsx_dir}/{ticker}_financials.xlsx"
            
            try:
                _save_pivot_xlsx(final_pivot, out_path)
            except ImportError:
                # Fallback to CSV in output/financials/
                out_path = f"{out_dir}/{ticker}_financials.csv"
                _profile_call("write_csv", final_pivot.to_csv, out_path)
                _status_print("  [xlsx] openpyxl not installed -- run 'pip install openpyxl'; saved CSV instead.")
            except Exception as _xe:
                # Fallback to CSV in output/financials/
                out_path = f"{out_dir}/{ticker}_financials.csv"
                _profile_call("write_csv", final_pivot.to_csv, out_path)
                _status_print(f"  [xlsx] export failed ({_xe}); saved CSV instead.")
        else:
            # Save standard CSV in output/financials/
            out_path = f"{out_dir}/{ticker}_financials.csv"
            _profile_call("write_csv", final_pivot.to_csv, out_path)

        if not log_output:
            progress.finish("Complete", f"Saved to {out_path}")
            builtins.print = _original_print
        else:
            progress.finish("Complete")
            _original_print(f"Success! Data saved to {out_path}")
    else:
        if not log_output:
            progress.finish("Complete (no data)", "No data extracted.")
            builtins.print = _original_print
        else:
            progress.finish("Complete (no data)")
            _original_print("No data extracted.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--limit", type=int, default=50,
                        help="Number of filings to pull (fewer = faster; less history).")
    parser.add_argument("--no-arelle", action="store_true",
                        help="Skip the slow Arelle custom-tag pre-pass "
                             "(10-K annual filings and 20-F/40-F annual enrichment).")
    parser.add_argument("--log", action="store_true", help="Print detailed logs")
    parser.add_argument("--xlsx", action="store_true",
                        help="Save as .xlsx (one sheet per statement) instead of CSV.")
    parser.add_argument("--workers", type=int, default=None,
                        help="Native 10-K/10-Q worker count. Defaults to "
                             "SEC_MAX_WORKERS or 8; SEC request starts remain "
                             "globally rate-limited.")
    parser.add_argument("--annual", action="store_true",
                        help="For native 10-K filers, fetch 10-K/10-K/A annual filings only and output FY columns.")
    parser.add_argument("--reset-identity", action="store_true",
                        help="Delete the saved SEC contact identity and show first-run setup again.")
    args = parser.parse_args()
    if args.reset_identity:
        _reset_cached_sec_identity()
    main(args.ticker.upper(), args.limit, use_arelle=not args.no_arelle,
         log_output=args.log, save_xlsx=args.xlsx, workers=args.workers,
         annual=args.annual)
